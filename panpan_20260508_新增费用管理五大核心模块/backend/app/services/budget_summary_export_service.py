from __future__ import annotations

from datetime import datetime
from io import BytesIO
from pathlib import Path
import re
from typing import Any, Awaitable, Callable

import aiosqlite
from fastapi import HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.db_paths import common_db_path
from app.formula_refs import extract_formula_codes
from app.schemas import BudgetSummaryExportPivotRequest


class BudgetSummaryExportService:
    def __init__(
        self,
        *,
        editable_context_provider: Callable[[], Awaitable[tuple[Path, int, int]]],
        fetch_budget_summary_all_rows: Callable[[Path], Awaitable[list[tuple[Any, ...]]]],
        normalize_formula: Callable[[str | None], str],
        normalize_summary_value: Callable[[str, Any], str],
        extract_data_acct_code_from_name: Callable[[str], str | None],
        extract_product_code_from_summary_name: Callable[[str | None], str | None],
        formula_ref_display_value: Callable[[float, str], str],
        budget_summary_field_meta: Callable[[], dict[str, tuple[str, str]]],
        write_template_pivot_data_area: Callable[..., None],
        build_export_year_datetime_text: Callable[[list[Any]], str],
        month_idx_from_label: Callable[[str | None], int | None],
        prepare_formula_expression: Callable[[str | None], str],
        autosize_worksheet_columns: Callable[[Any], None],
    ) -> None:
        self._editable_context_provider = editable_context_provider
        self._fetch_budget_summary_all_rows = fetch_budget_summary_all_rows
        self._normalize_formula = normalize_formula
        self._normalize_summary_value = normalize_summary_value
        self._extract_data_acct_code_from_name = extract_data_acct_code_from_name
        self._extract_product_code_from_summary_name = extract_product_code_from_summary_name
        self._formula_ref_display_value = formula_ref_display_value
        self._budget_summary_field_meta = budget_summary_field_meta
        self._write_template_pivot_data_area = write_template_pivot_data_area
        self._build_export_year_datetime_text = build_export_year_datetime_text
        self._month_idx_from_label = month_idx_from_label
        self._prepare_formula_expression = prepare_formula_expression
        self._autosize_worksheet_columns = autosize_worksheet_columns

    async def export_budget_summary_from_template(
        self,
        body: BudgetSummaryExportPivotRequest,
        output_filename: str,
    ) -> StreamingResponse:
        editable_budget_path, _editable_year, _editable_vid = await self._editable_context_provider()
        raw_rows = await self._fetch_budget_summary_all_rows(editable_budget_path)
        field_order = [
            "report_level1",
            "report_level2",
            "report_level3",
            "report_level4",
            "report_level5",
            "dept_level1",
            "dept_level2",
            "dept_level3",
            "data_code_name",
            "product_code_name",
            "year",
            "month",
            "quarter",
            "budget_actual",
            "version_id",
            "version_name",
            "value",
            "value_type",
            "update_time",
        ]
        records: list[dict[str, Any]] = []
        formula_tasks: list[dict[str, Any]] = []
        needed_ref_codes: set[str] = set()
        needed_period_ids: set[int] = set()
        needed_version_ids: set[int] = set()
        needed_budget_actuals: set[int] = set()
        needed_product_codes: set[str] = set()

        common_path = common_db_path()
        period_id_by_year_month: dict[tuple[str, str], int] = {}
        formula_map: dict[str, tuple[str, str]] = {}
        value_type_by_code: dict[str, str] = {}
        async with aiosqlite.connect(common_path) as cdb:
            await cdb.execute("PRAGMA foreign_keys = ON")
            cur_period = await cdb.execute("SELECT period_id, year, month FROM period")
            for row in await cur_period.fetchall():
                period_id_by_year_month[(str(row[1]), str(row[2]))] = int(row[0])
            cur_acct = await cdb.execute(
                """
                SELECT data_acct_code, budget_formula, actual_formula, value_type
                FROM data_account
                """
            )
            for row in await cur_acct.fetchall():
                code = str(row[0])
                formula_map[code] = (
                    self._normalize_formula(row[1]),
                    self._normalize_formula(row[2]),
                )
                value_type_by_code[code] = str(row[3] or "")

        for raw in raw_rows:
            rec = {field_order[i]: raw[i] for i in range(len(field_order))}
            rec["budget_actual"] = "预算" if int(raw[13] or 0) == 0 else "实际"
            rec["version_display"] = self._normalize_summary_value(
                "version_display",
                {"version_id": rec.get("version_id"), "version_name": rec.get("version_name")},
            )
            rec["calc_formula_raw"] = ""
            rec["calc_formula_values"] = ""
            records.append(rec)

            data_code = self._extract_data_acct_code_from_name(str(raw[8] or ""))
            if not data_code:
                continue
            budget_actual = int(raw[13] or 0)
            formula_pair = formula_map.get(data_code, ("", ""))
            formula = formula_pair[0] if budget_actual == 0 else formula_pair[1]
            if not formula:
                continue
            rec["calc_formula_raw"] = formula
            period_id = period_id_by_year_month.get((str(raw[10]), str(raw[11])))
            if period_id is None:
                continue
            refs = sorted(extract_formula_codes(formula))
            if not refs:
                continue
            row_pc = self._extract_product_code_from_summary_name(str(raw[9] or "")) or ""
            formula_tasks.append(
                {
                    "record": rec,
                    "formula": formula,
                    "refs": refs,
                    "period_id": period_id,
                    "version_id": int(raw[14] or 0),
                    "budget_actual": budget_actual,
                    "product_code": row_pc,
                }
            )
            needed_ref_codes.update(refs)
            needed_period_ids.add(period_id)
            needed_version_ids.add(int(raw[14] or 0))
            needed_budget_actuals.add(budget_actual)
            if row_pc:
                needed_product_codes.add(row_pc)

        ref_value_map: dict[tuple[str, str, int, int, int], float] = {}
        if (
            needed_ref_codes
            and needed_period_ids
            and needed_version_ids
            and needed_budget_actuals
            and needed_product_codes
        ):
            code_ph = ",".join(["?"] * len(needed_ref_codes))
            period_ph = ",".join(["?"] * len(needed_period_ids))
            version_ph = ",".join(["?"] * len(needed_version_ids))
            ba_ph = ",".join(["?"] * len(needed_budget_actuals))
            prod_ph = ",".join(["?"] * len(needed_product_codes))
            async with aiosqlite.connect(editable_budget_path) as bdb:
                await bdb.execute("PRAGMA foreign_keys = ON")
                cur_vals = await bdb.execute(
                    f"""
                    SELECT data_acct_code, product_code, period_id, version_id, budget_actual, value
                    FROM budget_data
                    WHERE data_acct_code IN ({code_ph})
                      AND period_id IN ({period_ph})
                      AND version_id IN ({version_ph})
                      AND budget_actual IN ({ba_ph})
                      AND product_code IN ({prod_ph})
                    """,
                    (
                        *sorted(needed_ref_codes),
                        *sorted(needed_period_ids),
                        *sorted(needed_version_ids),
                        *sorted(needed_budget_actuals),
                        *sorted(needed_product_codes),
                    ),
                )
                for row in await cur_vals.fetchall():
                    ref_value_map[
                        (str(row[0]), str(row[1]), int(row[2]), int(row[3]), int(row[4]))
                    ] = float(row[5] or 0.0)

        ref_pattern = re.compile(r"<\s*([A-Z]\d{4})[^>]*>")
        for task in formula_tasks:
            formula = task["formula"]
            refs = task["refs"]
            period_id = task["period_id"]
            version_id = task["version_id"]
            budget_actual = task["budget_actual"]

            task_pc = str(task.get("product_code") or "")

            def _to_display(code: str) -> str:
                raw_value = ref_value_map.get(
                    (code, task_pc, period_id, version_id, budget_actual),
                    0.0,
                )
                return self._formula_ref_display_value(raw_value, value_type_by_code.get(code, ""))

            if "<" in formula and ">" in formula:
                value_formula = ref_pattern.sub(lambda m: _to_display(str(m.group(1))), formula)
            else:
                value_formula = formula
                for code in sorted(refs, key=len, reverse=True):
                    value_formula = re.sub(rf"\b{re.escape(code)}\b", _to_display(code), value_formula)
            task["record"]["calc_formula_values"] = value_formula

        template_path = Path(__file__).resolve().parents[3] / "download_template" / "pivot_export_temp.xlsx"
        if not template_path.exists():
            raise HTTPException(status_code=404, detail="缺少导出模板 pivot_export_temp.xlsx")
        wb = load_workbook(template_path)
        if "Pivot数据区" not in wb.sheetnames:
            raise HTTPException(status_code=400, detail="模板缺少工作表：Pivot数据区")
        if "数据透视表" not in wb.sheetnames:
            raise HTTPException(status_code=400, detail="模板缺少工作表：数据透视表")
        pivot_count = sum(len(getattr(ws, "_pivots", [])) for ws in wb.worksheets)
        if pivot_count == 0:
            raise HTTPException(
                status_code=400,
                detail="模板未检测到Excel原生数据透视表实例，请在“数据透视表”中先创建透视表并保存",
            )
        ws_data = wb["Pivot数据区"]

        versions_info_text = self._build_export_year_datetime_text([rec.get("year") for rec in records])
        field_meta = self._budget_summary_field_meta()
        chinese_value_by_header = {
            field_meta["report_level1"][0]: lambda r: r["report_level1"] or "",
            field_meta["report_level2"][0]: lambda r: r["report_level2"] or "",
            field_meta["report_level3"][0]: lambda r: r["report_level3"] or "",
            field_meta["report_level4"][0]: lambda r: r["report_level4"] or "",
            field_meta["report_level5"][0]: lambda r: r["report_level5"] or "",
            field_meta["dept_level1"][0]: lambda r: r["dept_level1"] or "",
            field_meta["dept_level2"][0]: lambda r: r["dept_level2"] or "",
            field_meta["dept_level3"][0]: lambda r: r["dept_level3"] or "",
            field_meta["data_code_name"][0]: lambda r: r["data_code_name"] or "",
            field_meta["product_code_name"][0]: lambda r: r["product_code_name"] or "",
            field_meta["month"][0]: lambda r: r["month"] or "",
            field_meta["quarter"][0]: lambda r: r["quarter"] or "",
            field_meta["budget_actual"][0]: lambda r: r["budget_actual"] or "",
            field_meta["version_display"][0]: lambda r: r["version_display"] or "",
            "年度": lambda r: r.get("year") or "",
            "版本号": lambda r: r["version_display"] or "",
            "版本名称": lambda r: r["version_display"] or "",
            field_meta["value"][0]: lambda r: float(r["value"] or 0.0),
            field_meta["value_type"][0]: lambda r: r["value_type"] or "",
            field_meta["update_time"][0]: lambda r: r["update_time"] or "",
            "计算公式原文": lambda r: r["calc_formula_raw"] or "",
            "计算公式数值": lambda r: r["calc_formula_values"] or "",
        }
        self._write_template_pivot_data_area(
            wb=wb,
            ws_data=ws_data,
            records=records,
            chinese_value_by_header=chinese_value_by_header,
            versions_info_text=versions_info_text,
        )

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return StreamingResponse(
            out,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{output_filename}"'},
        )

    async def export_budget_summary_formula_tree_workbook(
        self,
        _body: BudgetSummaryExportPivotRequest,
        version_id: int,
        budget_path: Path,
        budget_year: int,
    ) -> StreamingResponse:
        common_path = common_db_path()
        async with aiosqlite.connect(common_path) as cdb:
            await cdb.execute("PRAGMA foreign_keys = ON")
            cur_reports = await cdb.execute(
                """
                SELECT report_acct_code, report_acct_name, parent_code, is_summary, is_minus
                FROM report_account
                """
            )
            report_rows = await cur_reports.fetchall()
            cur_mappings = await cdb.execute(
                """
                SELECT m.report_acct_code, m.data_acct_code, d.data_acct_name
                FROM report_data_mapping m
                LEFT JOIN data_account d ON d.data_acct_code = m.data_acct_code
                """
            )
            mapping_rows = await cur_mappings.fetchall()
            cur_data = await cdb.execute(
                """
                SELECT data_acct_code, data_acct_name, budget_formula, actual_formula, value_type, product_code
                FROM data_account
                """
            )
            data_rows = await cur_data.fetchall()
            cur_depts = await cdb.execute(
                """
                SELECT dept_code, dept_name, parent_code
                FROM dept_account
                """
            )
            dept_rows = await cur_depts.fetchall()
            cur_dept_map = await cdb.execute(
                """
                SELECT dept_code, product_code
                FROM dept_product_mapping
                """
            )
            dept_product_rows = await cur_dept_map.fetchall()
            cur_products = await cdb.execute(
                """
                SELECT product_code, product_name
                FROM product_type
                """
            )
            product_rows = await cur_products.fetchall()

        data_meta_map: dict[str, dict[str, Any]] = {}
        dept_info_map = {
            str(dept_code): {
                "name": str(dept_name or ""),
                "parent": str(parent_code) if parent_code is not None else None,
            }
            for dept_code, dept_name, parent_code in dept_rows
        }
        dept_code_by_product = {
            str(product_code): str(dept_code)
            for dept_code, product_code in dept_product_rows
            if product_code
        }
        product_name_map = {
            str(product_code): str(product_name or "")
            for product_code, product_name in product_rows
        }

        for code, name, budget_formula, actual_formula, value_type, product_code in data_rows:
            product_code_text = str(product_code or "")
            fallback_product = ""
            fallback_dept_levels = ["", "", ""]
            if product_code_text:
                product_name = product_name_map.get(product_code_text, "")
                fallback_product = f"{product_code_text} {product_name}".strip()
                mapped_dept_code = dept_code_by_product.get(product_code_text)
                if mapped_dept_code:
                    path_codes: list[str] = []
                    cur_code = mapped_dept_code
                    seen: set[str] = set()
                    while cur_code and cur_code not in seen:
                        seen.add(cur_code)
                        path_codes.append(cur_code)
                        cur_code = str(dept_info_map.get(cur_code, {}).get("parent") or "")
                    path_codes.reverse()
                    for idx, dept_code in enumerate(path_codes[:3]):
                        dept_name = str(dept_info_map.get(dept_code, {}).get("name") or "")
                        fallback_dept_levels[idx] = f"{dept_code} {dept_name}".strip()
            data_meta_map[str(code)] = {
                "name": str(name or ""),
                "budget_formula": self._normalize_formula(budget_formula),
                "actual_formula": self._normalize_formula(actual_formula),
                "value_type": str(value_type or ""),
                "fallback_dept_levels": fallback_dept_levels,
                "fallback_product": fallback_product,
            }

        report_nodes: dict[str, dict[str, Any]] = {}
        for report_code, report_name, parent_code, is_summary, is_minus in report_rows:
            code = str(report_code)
            report_nodes[code] = {
                "type": "report",
                "code": code,
                "name": str(report_name or ""),
                "parent_code": str(parent_code) if parent_code is not None else None,
                "is_summary": int(is_summary or 0),
                "is_minus": int(is_minus or 0),
                "children": [],
            }

        roots: list[dict[str, Any]] = []
        for node in report_nodes.values():
            parent_code = node.get("parent_code")
            if parent_code and parent_code in report_nodes:
                report_nodes[parent_code]["children"].append(node)
            else:
                roots.append(node)

        data_values: dict[str, dict[str, Any]] = {}
        current_month = 1
        version_name = ""
        async with aiosqlite.connect(budget_path) as bdb:
            await bdb.execute("PRAGMA foreign_keys = ON")
            cur_ver = await bdb.execute(
                "SELECT version_name, current_month FROM version WHERE version_id = ?",
                (version_id,),
            )
            ver_row = await cur_ver.fetchone()
            if not ver_row:
                raise HTTPException(status_code=400, detail=f"版本 {version_id} 不存在")
            version_name = str(ver_row[0] or "")
            current_month = int(ver_row[1] or 1)
            if current_month < 1 or current_month > 13:
                current_month = 1
            cur_summary = await bdb.execute(
                """
                SELECT data_code_name, dept_level1, dept_level2, dept_level3, product_code_name,
                       month, budget_actual, value
                FROM budget_summary
                WHERE version_id = ?
                """,
                (version_id,),
            )
            summary_rows = await cur_summary.fetchall()

        for (
            data_code_name,
            dept_level1,
            dept_level2,
            dept_level3,
            product_code_name,
            month_label,
            budget_actual,
            value,
        ) in summary_rows:
            code = self._extract_data_acct_code_from_name(str(data_code_name or ""))
            if not code:
                continue
            rec = data_values.setdefault(
                code,
                {
                    "data_display": str(data_code_name or code),
                    "dept_levels": ["", "", ""],
                    "product": "",
                    "budget_months": [0.0] * 12,
                    "actual_months": [0.0] * 12,
                },
            )
            if not rec["data_display"] and data_code_name:
                rec["data_display"] = str(data_code_name)
            for idx, raw in enumerate((dept_level1, dept_level2, dept_level3)):
                if not rec["dept_levels"][idx] and raw:
                    rec["dept_levels"][idx] = str(raw)
            if not rec["product"] and product_code_name:
                rec["product"] = str(product_code_name)
            month_idx = self._month_idx_from_label(str(month_label or ""))
            if month_idx is None:
                continue
            if int(budget_actual or 0) == 0:
                if abs(rec["budget_months"][month_idx]) < 1e-12:
                    rec["budget_months"][month_idx] = float(value or 0.0)
            else:
                if abs(rec["actual_months"][month_idx]) < 1e-12:
                    rec["actual_months"][month_idx] = float(value or 0.0)

        for idx, (report_code_raw, data_code_raw, data_name_raw) in enumerate(mapping_rows):
            report_code = str(report_code_raw or "")
            data_code = str(data_code_raw or "")
            parent = report_nodes.get(report_code)
            if not parent or not data_code:
                continue
            base = data_values.get(data_code, {})
            meta = data_meta_map.get(data_code, {})
            display_name = str(data_name_raw or meta.get("name", "")).strip()
            data_display = f"{data_code} {display_name}".strip() if display_name else data_code
            if base.get("data_display"):
                data_display = str(base["data_display"])
            parent["children"].append(
                {
                    "type": "data",
                    "id": f"{report_code}:{data_code}:{idx}",
                    "code": data_code,
                    "name": data_display,
                    "dept_levels": list(base.get("dept_levels", meta.get("fallback_dept_levels", ["", "", ""]))),
                    "product": str(base.get("product", meta.get("fallback_product", ""))),
                    "budget_months": list(base.get("budget_months", [0.0] * 12)),
                    "actual_months": list(base.get("actual_months", [0.0] * 12)),
                    "budget_formula": str(meta.get("budget_formula", "")),
                    "actual_formula": str(meta.get("actual_formula", "")),
                    "value_type": str(meta.get("value_type", "")),
                }
            )

        def _sort_tree(nodes: list[dict[str, Any]]) -> None:
            nodes.sort(key=lambda n: str(n.get("code", "")))
            for node in nodes:
                children = node.get("children") or []
                if children:
                    _sort_tree(children)

        _sort_tree(roots)

        row_entries: list[dict[str, Any]] = []

        def _walk(node: dict[str, Any], report_level: int) -> None:
            if node.get("type") == "report":
                entry = {
                    "type": "report",
                    "node": node,
                    "report_level": report_level,
                    "depth": report_level,
                }
                node["_entry"] = entry
                row_entries.append(entry)
                for child in node.get("children") or []:
                    _walk(child, report_level + 1)
            else:
                entry = {
                    "type": "data",
                    "node": node,
                    "report_level": report_level - 1,
                    "depth": report_level,
                }
                node["_entry"] = entry
                row_entries.append(entry)

        for root in roots:
            _walk(root, 1)

        x = max(1, min(13, current_month))

        def _month_uses_actual(month_num: int) -> bool:
            if x == 13:
                return True
            if x == 1:
                return False
            return month_num < x

        month_headers = [
            (f"实际M{m:02d}" if _month_uses_actual(m) else f"预算M{m:02d}")
            for m in range(1, 13)
        ]

        wb = Workbook()
        ws = wb.active
        ws.title = "导出带公式树状表"
        ws.cell(
            row=1,
            column=1,
            value=f"当前年度：{budget_year}__当前版本号：{version_id}__版本名称：{version_name}",
        )
        headers = [
            "报告科目1级",
            "报告科目2级",
            "报告科目3级",
            "报告科目4级",
            "报告科目5级",
            "数据科目",
            "数值类型",
            "部门科目1级",
            "部门科目2级",
            "部门科目3级",
            "产品科目",
            *month_headers,
            "季度Q1",
            "季度Q2",
            "季度Q3",
            "季度Q4",
            "年度合计",
        ]
        header_fill = PatternFill(fill_type="solid", fgColor="FF000000")
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)

        for idx, entry in enumerate(row_entries):
            entry["row_num"] = idx + 3

        data_row_by_code: dict[str, int] = {}
        for entry in row_entries:
            if entry["type"] == "data":
                code = str(entry["node"].get("code", ""))
                if code and code not in data_row_by_code:
                    data_row_by_code[code] = int(entry["row_num"])

        def _weighted_sum_formula_for_children(col_idx: int, children: list[dict[str, Any]]) -> str:
            terms: list[tuple[int, str]] = []
            for ch in children:
                if not isinstance(ch, dict) or "_entry" not in ch:
                    continue
                row_ref = f"{get_column_letter(col_idx)}{int(ch['_entry']['row_num'])}"
                if ch.get("type") == "report":
                    if int(ch.get("is_summary", 0)) != 1:
                        continue
                    sign = -1 if int(ch.get("is_minus", 0)) == 1 else 1
                    terms.append((sign, row_ref))
                else:
                    terms.append((1, row_ref))
            if not terms:
                return "0"
            expr = ""
            for idx, (sign, ref) in enumerate(terms):
                if idx == 0:
                    expr += f"-{ref}" if sign < 0 else ref
                else:
                    expr += f"-{ref}" if sign < 0 else f"+{ref}"
            return f"={expr}"

        gray_fill = PatternFill(fill_type="solid", fgColor="EDEDED")
        top_level_fill = PatternFill(fill_type="solid", fgColor="FFC0C0C0")
        formula_fill = PatternFill(fill_type="solid", fgColor="FFFFF2E6")
        muted_font = Font(color="FFB3B3B3")

        max_report_level_in_rows = 1
        for entry in row_entries:
            if entry["type"] == "report":
                max_report_level_in_rows = max(max_report_level_in_rows, int(entry["report_level"]))

        data_outline_level = min(max_report_level_in_rows, 8)

        for entry in row_entries:
            row_num = int(entry["row_num"])
            if entry["type"] == "report":
                node = entry["node"]
                level = max(1, min(int(entry["report_level"]), 5))
                report_display = f"{node.get('code', '')} {node.get('name', '')}".strip()
                if int(node.get("is_minus", 0)) == 1:
                    report_display = f"减： {report_display}"
                ws.cell(row=row_num, column=level, value=report_display)
                if level == 1:
                    ws.row_dimensions[row_num].height = 17
                    for col_idx in range(1, 29):
                        ws.cell(row=row_num, column=col_idx).fill = top_level_fill
                    for col_idx in range(1, 6):
                        ws.cell(row=row_num, column=col_idx).font = Font(
                            color="FF000000",
                            bold=True,
                            size=12,
                        )
                children = node.get("children") or []

                def _has_descendant_data(n: dict[str, Any]) -> bool:
                    if n.get("type") == "data":
                        return True
                    return any(_has_descendant_data(ch) for ch in (n.get("children") or []))

                has_data_descendant = _has_descendant_data(node)
                if int(node.get("is_summary", 0)) != 1:
                    for col_idx in range(12, 29):
                        c = ws.cell(row=row_num, column=col_idx, value="非汇总")
                        c.font = muted_font
                elif not has_data_descendant:
                    for col_idx in range(12, 29):
                        c = ws.cell(row=row_num, column=col_idx, value="无数据")
                        c.font = muted_font
                else:
                    for col_idx in range(12, 28):
                        ws.cell(
                            row=row_num,
                            column=col_idx,
                            value=_weighted_sum_formula_for_children(col_idx, children),
                        )
                    ws.cell(
                        row=row_num,
                        column=28,
                        value=f"=SUM({get_column_letter(12)}{row_num}:{get_column_letter(23)}{row_num})",
                    )

                for col_idx in range(12, 29):
                    ws.cell(row=row_num, column=col_idx).number_format = "#,##0.00;[Red]-#,##0.00"
            else:
                node = entry["node"]
                value_type = str(node.get("value_type", ""))
                ws.cell(row=row_num, column=6, value=str(node.get("name", "")))
                ws.cell(row=row_num, column=6).fill = gray_fill
                value_type_display = value_type
                if value_type == "金额":
                    value_type_display = "金额_亿元"
                elif value_type == "户数":
                    value_type_display = "户数_万户"
                ws.cell(row=row_num, column=7, value=value_type_display)
                ws.cell(row=row_num, column=7).fill = gray_fill
                dept_levels = list(node.get("dept_levels", ["", "", ""]))
                ws.cell(row=row_num, column=8, value=dept_levels[0] if len(dept_levels) > 0 else "")
                ws.cell(row=row_num, column=9, value=dept_levels[1] if len(dept_levels) > 1 else "")
                ws.cell(row=row_num, column=10, value=dept_levels[2] if len(dept_levels) > 2 else "")
                ws.cell(row=row_num, column=11, value=str(node.get("product", "")))
                ws.cell(row=row_num, column=11).fill = gray_fill

                budget_formula = self._normalize_formula(str(node.get("budget_formula", "")))
                actual_formula = self._normalize_formula(str(node.get("actual_formula", "")))
                budget_months = list(node.get("budget_months", [0.0] * 12))
                actual_months = list(node.get("actual_months", [0.0] * 12))

                def _normalize_direct_value(raw_val: float) -> float:
                    if value_type == "百分比":
                        return float(raw_val / 100.0) if abs(raw_val) > 1 else float(raw_val)
                    return float(raw_val)

                def _formula_to_excel(formula: str, col_idx: int) -> str:
                    expr = self._prepare_formula_expression(formula)
                    if not expr:
                        return "0"

                    def _replace_code(m: re.Match[str]) -> str:
                        code = str(m.group(1))
                        target_row = data_row_by_code.get(code)
                        if target_row is None:
                            return "0"
                        return f"{get_column_letter(col_idx)}{target_row}"

                    expr = re.sub(r"\b([A-Z]\d{4})\b", _replace_code, expr)
                    return f"={expr}"

                for month_idx in range(12):
                    month_num = month_idx + 1
                    month_col = 12 + month_idx
                    mcell = ws.cell(row=row_num, column=month_col)
                    use_actual = _month_uses_actual(month_num)
                    formula_to_use = actual_formula if use_actual else budget_formula
                    if formula_to_use:
                        mcell.value = _formula_to_excel(formula_to_use, month_col)
                        mcell.comment = Comment(
                            f"原始{'实际' if use_actual else '预算'}公式：{formula_to_use}",
                            "System",
                        )
                        mcell.comment.width = 300
                        mcell.fill = formula_fill
                    else:
                        source_vals = actual_months if use_actual else budget_months
                        mcell.value = _normalize_direct_value(float(source_vals[month_idx] or 0.0))
                        mcell.fill = gray_fill

                for q_idx in range(4):
                    q_start = 12 + q_idx * 3
                    q_end = q_start + 2
                    ws.cell(
                        row=row_num,
                        column=24 + q_idx,
                        value=f"=SUM({get_column_letter(q_start)}{row_num}:{get_column_letter(q_end)}{row_num})",
                    )
                ws.cell(
                    row=row_num,
                    column=28,
                    value=f"=SUM({get_column_letter(12)}{row_num}:{get_column_letter(23)}{row_num})",
                )

                num_fmt = "0.00%;[Red]-0.00%" if value_type == "百分比" else "#,##0.00;[Red]-#,##0.00"
                for col_idx in range(12, 29):
                    ws.cell(row=row_num, column=col_idx).number_format = num_fmt

            if entry["type"] == "report":
                outline_level = max(int(entry["report_level"]) - 1, 0)
            else:
                outline_level = data_outline_level
            ws.row_dimensions[row_num].outlineLevel = min(outline_level, 8)
            ws.row_dimensions[row_num].hidden = outline_level > 0

        ws.freeze_panes = "L3"
        ws.auto_filter.ref = f"A2:{get_column_letter(28)}2"
        ws.sheet_view.showOutlineSymbols = True
        if ws.sheet_properties.outlinePr is not None:
            ws.sheet_properties.outlinePr.summaryBelow = True

        ws.column_dimensions.group(get_column_letter(8), get_column_letter(10), outline_level=1, hidden=False)
        for col_idx in range(8, 11):
            ws.column_dimensions[get_column_letter(col_idx)].outlineLevel = 1
        ws.column_dimensions.group(get_column_letter(12), get_column_letter(23), outline_level=1, hidden=False)
        ws.column_dimensions.group(get_column_letter(24), get_column_letter(27), outline_level=1, hidden=False)
        for start_col, end_col in ((12, 23), (24, 27)):
            for col_idx in range(start_col, end_col + 1):
                ws.column_dimensions[get_column_letter(col_idx)].outlineLevel = 1

        self._autosize_worksheet_columns(ws)
        for col_idx in range(1, 6):
            ws.column_dimensions[get_column_letter(col_idx)].width = 8
        for col_idx in range(12, 29):
            letter = get_column_letter(col_idx)
            ws.column_dimensions[letter].width = min(max(ws.column_dimensions[letter].width or 8, 10), 16)
        for row_idx in range(3, ws.max_row + 1):
            for col_idx in range(12, 29):
                ws.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal="right")
            ws.cell(row=row_idx, column=28).fill = top_level_fill

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return StreamingResponse(
            out,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="budget_summary_formula_workbook.xlsx"'},
        )
