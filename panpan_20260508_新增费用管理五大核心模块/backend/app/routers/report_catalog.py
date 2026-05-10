from __future__ import annotations

from copy import copy
from io import BytesIO
import json
from pathlib import Path
import re
from typing import Any, Awaitable, Callable
import unicodedata

import aiosqlite
from fastapi import APIRouter, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from app.db_paths import common_db_path
from app.schemas import (
    ProductTypeCreate,
    ProductTypeRow,
    ProductTypeUpdate,
    ReportAccountCreate,
    ReportAccountRow,
    ReportAccountUpdate,
    ReportDataMappingCreate,
    ReportDataMappingRow,
)


def build_report_catalog_router(
    *,
    normalize_cell: Callable[[Any], str],
    color_row: Callable[[Any, int, int, str], None],
    validate_report_code_with_parent: Callable[[str, int, str | None], str | None],
    parse_bool_like: Callable[[str], int | None],
    write_operation_log: Callable[..., Awaitable[None]],
) -> APIRouter:
    router = APIRouter()

    @router.post("/api/report-accounts/import-preview")
    async def preview_report_account_import(file: UploadFile = File(...)):
        content = await file.read()
        try:
            wb = load_workbook(filename=BytesIO(content), data_only=True)
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"无法读取Excel文件：{exc}") from exc

        if "数据模版" not in wb.sheetnames:
            raise HTTPException(status_code=400, detail='上传文件缺失“数据模版”工作表，不能上传数据。')

        ws = wb["数据模版"]
        headers = [normalize_cell(c.value) for c in ws[1]]
        if not any(headers):
            raise HTTPException(status_code=400, detail="数据模版工作表第一行字段头为空")

        preview_rows: list[dict[str, str]] = []
        total_rows = 0
        for ridx in range(2, ws.max_row + 1):
            row_values = [normalize_cell(ws.cell(ridx, c).value) for c in range(1, ws.max_column + 1)]
            if not any(row_values):
                continue
            total_rows += 1
            if len(preview_rows) < 20:
                preview_rows.append(
                    {headers[i]: row_values[i] for i in range(min(len(headers), len(row_values))) if headers[i]}
                )

        return {
            "columns": [h for h in headers if h],
            "preview_rows": preview_rows,
            "row_count": total_rows,
        }

    @router.post("/api/report-accounts/import-apply")
    async def apply_report_account_import(
        file: UploadFile = File(...),
        mappings_json: str = Form(...),
    ):
        try:
            mappings = json.loads(mappings_json)
        except json.JSONDecodeError as exc:
            raise HTTPException(status_code=400, detail="字段映射格式不合法") from exc
        if not isinstance(mappings, dict):
            raise HTTPException(status_code=400, detail="字段映射格式不合法")

        required_fields = [
            "level1Code", "level1Name",
            "level2Code", "level2Name",
            "level3Code", "level3Name",
            "level4Code", "level4Name",
            "level5Code", "level5Name",
            "dataCode", "dataName",
            "isSummary", "isMinus", "remark",
        ]
        for f in required_fields:
            if not mappings.get(f):
                raise HTTPException(status_code=400, detail=f"字段映射缺失：{f}")

        content = await file.read()
        try:
            wb = load_workbook(filename=BytesIO(content))
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"无法读取Excel文件：{exc}") from exc
        if "数据模版" not in wb.sheetnames:
            raise HTTPException(status_code=400, detail='上传文件缺失“数据模版”工作表，不能上传数据。')

        ws = wb["数据模版"]
        headers = [normalize_cell(c.value) for c in ws[1]]
        header_col = {h: i + 1 for i, h in enumerate(headers) if h}
        for key, col_name in mappings.items():
            if col_name and col_name not in header_col:
                raise HTTPException(status_code=400, detail=f"映射列不存在：{col_name}")
        fail_reason_col = ws.max_column + 1
        ws.cell(row=1, column=fail_reason_col, value="失败原因")

        rows: list[dict[str, Any]] = []
        for ridx in range(2, ws.max_row + 1):
            raw_by_header = {h: normalize_cell(ws.cell(ridx, cidx).value) for h, cidx in header_col.items()}
            if not any(raw_by_header.values()):
                continue
            rows.append(
                {
                    "row_idx": ridx,
                    "lvl1_code": raw_by_header.get(mappings.get("level1Code", ""), "").upper(),
                    "lvl1_name": raw_by_header.get(mappings.get("level1Name", ""), ""),
                    "lvl2_code": raw_by_header.get(mappings.get("level2Code", ""), "").upper(),
                    "lvl2_name": raw_by_header.get(mappings.get("level2Name", ""), ""),
                    "lvl3_code": raw_by_header.get(mappings.get("level3Code", ""), "").upper(),
                    "lvl3_name": raw_by_header.get(mappings.get("level3Name", ""), ""),
                    "lvl4_code": raw_by_header.get(mappings.get("level4Code", ""), "").upper(),
                    "lvl4_name": raw_by_header.get(mappings.get("level4Name", ""), ""),
                    "lvl5_code": raw_by_header.get(mappings.get("level5Code", ""), "").upper(),
                    "lvl5_name": raw_by_header.get(mappings.get("level5Name", ""), ""),
                    "data_code": raw_by_header.get(mappings.get("dataCode", ""), "").upper(),
                    "data_name": raw_by_header.get(mappings.get("dataName", ""), ""),
                    "is_summary_raw": raw_by_header.get(mappings.get("isSummary", ""), ""),
                    "is_minus_raw": raw_by_header.get(mappings.get("isMinus", ""), ""),
                    "remark": raw_by_header.get(mappings.get("remark", ""), ""),
                }
            )

        report_code_re = re.compile(r"^[A-Z]\d+$")
        data_code_re = re.compile(r"^[A-Z]\d{4}$")
        success_count = 0
        overwrite_count = 0
        failed_count = 0
        current_path: dict[int, str] = {}

        path = common_db_path()
        async with aiosqlite.connect(path) as db:
            await db.execute("PRAGMA foreign_keys = ON")

            def _mark_failed(row_idx: int, reasons: list[str] | str) -> None:
                reason_text = "；".join(reasons) if isinstance(reasons, list) else str(reasons)
                ws.cell(row=row_idx, column=fail_reason_col, value=reason_text)
                color_row(ws, row_idx, ws.max_column, "FFFF0000")

            cur = await db.execute(
                "SELECT report_acct_code, report_acct_name, parent_code, is_summary, is_minus, level, remark FROM report_account"
            )
            report_rows = await cur.fetchall()
            existing_report: dict[str, dict[str, Any]] = {
                str(r[0]): {
                    "name": str(r[1] or ""),
                    "parent_code": str(r[2]) if r[2] is not None else None,
                    "is_summary": int(r[3] or 0),
                    "is_minus": int(r[4] or 0),
                    "level": int(r[5]),
                    "remark": str(r[6] or ""),
                }
                for r in report_rows
            }
            parent_to_children: dict[str | None, set[str]] = {}
            for code, info in existing_report.items():
                parent = info.get("parent_code")
                parent_to_children.setdefault(parent, set()).add(code)

            cur = await db.execute("SELECT data_acct_code FROM data_account")
            existing_data_codes = {str(r[0]) for r in await cur.fetchall()}
            cur = await db.execute("SELECT report_acct_code, data_acct_code FROM report_data_mapping")
            existing_mapping = {(str(r[0]), str(r[1])) for r in await cur.fetchall()}

            for row in rows:
                errors: list[str] = []
                level_entries = []
                for level, code_key, name_key in (
                    (1, "lvl1_code", "lvl1_name"),
                    (2, "lvl2_code", "lvl2_name"),
                    (3, "lvl3_code", "lvl3_name"),
                    (4, "lvl4_code", "lvl4_name"),
                    (5, "lvl5_code", "lvl5_name"),
                ):
                    c = row[code_key].strip()
                    n = row[name_key].strip()
                    if c or n:
                        level_entries.append((level, c, n))

                data_code = row["data_code"].strip()
                data_name = row["data_name"].strip()
                is_report_row = len(level_entries) > 0
                is_data_row = bool(data_code or data_name)

                if is_report_row and is_data_row:
                    errors.append("同一行不能同时填写报告科目和数据科目")
                elif not is_report_row and not is_data_row:
                    errors.append("空行无有效内容")

                if is_report_row:
                    if len(level_entries) != 1:
                        errors.append("每行只允许填写一个层级的报告科目代码和名称")
                    else:
                        level, code, name = level_entries[0]
                        if not code:
                            errors.append("报告科目代码不能为空")
                        elif not report_code_re.match(code):
                            errors.append("报告科目代码格式错误（示例：A01、A0101）")
                        if not name:
                            errors.append("报告科目名称不能为空")
                        parent_code = current_path.get(level - 1) if level > 1 else None
                        code_err = validate_report_code_with_parent(code, level, parent_code)
                        if code_err:
                            errors.append(code_err)

                        is_summary = parse_bool_like(row["is_summary_raw"])
                        is_minus = parse_bool_like(row["is_minus_raw"])
                        if is_summary is None:
                            errors.append("是否汇总字段必须是 0/1（或 是/否）")
                        if is_minus is None:
                            errors.append("是否减项字段必须是 0/1（或 是/否）")

                if is_data_row:
                    if level_entries:
                        errors.append("数据映射行必须仅填写数据科目列")
                    if not data_code:
                        errors.append("数据科目代码不能为空")
                    elif not data_code_re.match(data_code):
                        errors.append("数据科目代码格式错误（示例：A1001）")
                    elif data_code not in existing_data_codes:
                        errors.append("数据科目代码在系统中不存在")

                    target_report = None
                    for lv in (5, 4, 3, 2, 1):
                        if current_path.get(lv):
                            target_report = current_path[lv]
                            break
                    if not target_report:
                        errors.append("数据映射行之前必须先出现所属报告科目行")
                    elif target_report not in existing_report:
                        errors.append("映射目标报告科目在系统中不存在")
                    elif parent_to_children.get(target_report):
                        errors.append("该报告科目存在下级报告科目，不能直接挂接数据科目")

                if errors:
                    failed_count += 1
                    _mark_failed(row["row_idx"], errors)
                    continue

                if is_report_row:
                    level, code, name = level_entries[0]
                    parent_code = current_path.get(level - 1) if level > 1 else None
                    is_summary = parse_bool_like(row["is_summary_raw"])
                    is_minus = parse_bool_like(row["is_minus_raw"])
                    remark = row["remark"].strip() or None
                    assert is_summary is not None and is_minus is not None
                    existing = existing_report.get(code)

                    if existing:
                        if (existing.get("parent_code") or None) != (parent_code or None):
                            failed_count += 1
                            _mark_failed(row["row_idx"], "报告科目代码已存在，但上级科目不匹配")
                            continue
                        if int(existing.get("level") or 0) != level:
                            failed_count += 1
                            _mark_failed(row["row_idx"], "报告科目代码已存在，但层级不匹配")
                            continue
                        await db.execute(
                            """
                            UPDATE report_account
                            SET report_acct_name = ?, is_summary = ?, is_minus = ?, remark = ?
                            WHERE report_acct_code = ?
                            """,
                            (name, is_summary, is_minus, remark, code),
                        )
                        overwrite_count += 1
                        success_count += 1
                        color_row(ws, row["row_idx"], ws.max_column, "FF0000FF")
                    else:
                        await db.execute(
                            """
                            INSERT INTO report_account (
                              report_acct_code, report_acct_name, parent_code, is_summary, is_minus, level, is_leaf, remark
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                            """,
                            (code, name, parent_code, is_summary, is_minus, level, 0, remark),
                        )
                        existing_report[code] = {
                            "name": name,
                            "parent_code": parent_code,
                            "is_summary": is_summary,
                            "is_minus": is_minus,
                            "level": level,
                            "remark": remark or "",
                        }
                        parent_to_children.setdefault(parent_code, set()).add(code)
                        success_count += 1
                        color_row(ws, row["row_idx"], ws.max_column, "FF008000")

                    current_path[level] = code
                    for lv in range(level + 1, 6):
                        current_path.pop(lv, None)
                else:
                    target_report = None
                    for lv in (5, 4, 3, 2, 1):
                        if current_path.get(lv):
                            target_report = current_path[lv]
                            break
                    assert target_report is not None
                    if (target_report, data_code) in existing_mapping:
                        overwrite_count += 1
                        success_count += 1
                        color_row(ws, row["row_idx"], ws.max_column, "FF0000FF")
                    else:
                        await db.execute(
                            """
                            INSERT INTO report_data_mapping (report_acct_code, data_acct_code)
                            VALUES (?, ?)
                            """,
                            (target_report, data_code),
                        )
                        existing_mapping.add((target_report, data_code))
                        success_count += 1
                        color_row(ws, row["row_idx"], ws.max_column, "FF008000")

            await db.commit()

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        total = success_count + failed_count
        return StreamingResponse(
            out,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": 'attachment; filename="report_account_import_result.xlsx"',
                "X-Import-Total": str(total),
                "X-Import-Success": str(success_count),
                "X-Import-Overwrite": str(overwrite_count),
                "X-Import-Failed": str(failed_count),
                "Access-Control-Expose-Headers": (
                    "Content-Disposition,"
                    "X-Import-Total,X-Import-Success,X-Import-Overwrite,X-Import-Failed"
                ),
            },
        )

    @router.get("/api/product-types", response_model=list[ProductTypeRow])
    async def list_products():
        path = common_db_path()
        async with aiosqlite.connect(path) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            cur = await db.execute(
                "SELECT product_code, product_name, remark FROM product_type ORDER BY product_code"
            )
            rows = await cur.fetchall()
        return [ProductTypeRow(product_code=r[0], product_name=r[1], remark=r[2]) for r in rows]

    @router.post("/api/product-types", response_model=ProductTypeRow)
    async def create_product(body: ProductTypeCreate):
        path = common_db_path()
        async with aiosqlite.connect(path) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            cur = await db.execute(
                "SELECT 1 FROM product_type WHERE product_code = ?",
                (body.product_code,),
            )
            if await cur.fetchone():
                raise HTTPException(status_code=409, detail="产品代码已存在")
            await db.execute(
                "INSERT INTO product_type (product_code, product_name, remark) VALUES (?, ?, ?)",
                (body.product_code, body.product_name, body.remark),
            )
            await db.commit()
        await write_operation_log(
            action_type="INSERT",
            action_desc=f"新增产品科目 {body.product_code}",
            target_table="product_type",
            affected_rows=1,
            after_data=body.model_dump(),
        )
        return ProductTypeRow(
            product_code=body.product_code,
            product_name=body.product_name,
            remark=body.remark,
        )

    @router.patch("/api/product-types/{code}", response_model=ProductTypeRow)
    async def update_product(code: str, body: ProductTypeUpdate):
        path = common_db_path()
        async with aiosqlite.connect(path) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            cur = await db.execute(
                "SELECT product_code, product_name, remark FROM product_type WHERE product_code = ?",
                (code,),
            )
            before_row = await cur.fetchone()
            if not before_row:
                raise HTTPException(status_code=404, detail="产品不存在")
            updates: list[str] = []
            vals: list[Any] = []
            if body.product_name is not None:
                if not body.product_name.strip():
                    raise HTTPException(status_code=400, detail="产品名称不能为空")
                updates.append("product_name = ?")
                vals.append(body.product_name.strip())
            if body.remark is not None:
                updates.append("remark = ?")
                vals.append(body.remark)
            if updates:
                vals.append(code)
                await db.execute(
                    f"UPDATE product_type SET {', '.join(updates)} WHERE product_code = ?",
                    vals,
                )
                await db.commit()
            cur = await db.execute(
                "SELECT product_code, product_name, remark FROM product_type WHERE product_code = ?",
                (code,),
            )
            row = await cur.fetchone()
        await write_operation_log(
            action_type="UPDATE",
            action_desc=f"更新产品科目 {code}",
            target_table="product_type",
            affected_rows=1,
            before_data={
                "product_code": before_row[0],
                "product_name": before_row[1],
                "remark": before_row[2],
            },
            after_data={
                "product_code": row[0],
                "product_name": row[1],
                "remark": row[2],
            },
        )
        return ProductTypeRow(product_code=row[0], product_name=row[1], remark=row[2])

    @router.delete("/api/product-types/{code}")
    async def delete_product(code: str):
        path = common_db_path()
        async with aiosqlite.connect(path) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            cur = await db.execute(
                "SELECT product_code, product_name, remark FROM product_type WHERE product_code = ?",
                (code,),
            )
            before = await cur.fetchone()
            if not before:
                raise HTTPException(status_code=404, detail="产品不存在")
            cur = await db.execute(
                "SELECT COUNT(*) FROM data_account WHERE product_code = ?",
                (code,),
            )
            if (await cur.fetchone())[0] > 0:
                raise HTTPException(status_code=409, detail="仍有数据科目引用该产品，无法删除")
            cur = await db.execute(
                "SELECT COUNT(*) FROM dept_product_mapping WHERE product_code = ?",
                (code,),
            )
            if (await cur.fetchone())[0] > 0:
                raise HTTPException(status_code=409, detail="仍有部门映射引用该产品，无法删除")
            await db.execute("DELETE FROM product_type WHERE product_code = ?", (code,))
            await db.commit()
        await write_operation_log(
            action_type="DELETE",
            action_desc=f"删除产品科目 {code}",
            target_table="product_type",
            affected_rows=1,
            before_data={
                "product_code": before[0],
                "product_name": before[1],
                "remark": before[2],
            },
            after_data=None,
        )
        return {"ok": True}

    @router.get("/api/report-accounts", response_model=list[ReportAccountRow])
    async def list_report_accounts():
        path = common_db_path()
        async with aiosqlite.connect(path) as db:
            cur = await db.execute(
                """
                SELECT report_acct_code, report_acct_name, parent_code, is_summary,
                       is_minus, level, is_leaf, remark
                FROM report_account
                ORDER BY report_acct_code
                """
            )
            rows = await cur.fetchall()
        return [
            ReportAccountRow(
                report_acct_code=r[0],
                report_acct_name=r[1],
                parent_code=r[2],
                is_summary=bool(r[3]),
                is_minus=bool(r[4]),
                level=int(r[5]),
                is_leaf=bool(r[6]),
                remark=r[7],
            )
            for r in rows
        ]

    @router.get("/api/report-tree/export")
    async def export_report_tree():
        common_path = common_db_path()
        async with aiosqlite.connect(common_path) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            cur = await db.execute(
                """
                SELECT report_acct_code, report_acct_name, parent_code, is_summary, is_minus, remark
                FROM report_account
                """
            )
            report_rows = await cur.fetchall()
            cur = await db.execute(
                """
                SELECT m.report_acct_code, m.data_acct_code, d.data_acct_name
                FROM report_data_mapping m
                LEFT JOIN data_account d ON d.data_acct_code = m.data_acct_code
                """
            )
            mapping_rows = await cur.fetchall()

        report_nodes: dict[str, dict[str, Any]] = {}
        for r in report_rows:
            code = str(r[0])
            report_nodes[code] = {
                "type": "report",
                "code": code,
                "name": str(r[1] or ""),
                "parent_code": str(r[2]) if r[2] is not None else None,
                "is_summary": int(r[3] or 0),
                "is_minus": int(r[4] or 0),
                "remark": str(r[5] or ""),
                "children": [],
            }

        roots: list[dict[str, Any]] = []
        for node in report_nodes.values():
            parent_code = node["parent_code"]
            if parent_code and parent_code in report_nodes:
                report_nodes[parent_code]["children"].append(node)
            else:
                roots.append(node)

        for report_code_raw, data_code_raw, data_name_raw in mapping_rows:
            report_code = str(report_code_raw)
            parent = report_nodes.get(report_code)
            if not parent:
                continue
            parent["children"].append(
                {
                    "type": "data",
                    "code": str(data_code_raw),
                    "name": str(data_name_raw or ""),
                    "children": [],
                }
            )

        def _sort_tree(nodes: list[dict[str, Any]]) -> None:
            nodes.sort(key=lambda n: str(n.get("code", "")))
            for n in nodes:
                children = n.get("children") or []
                if children:
                    _sort_tree(children)

        _sort_tree(roots)

        report_levels = 5
        headers: list[str] = []
        for i in range(1, report_levels + 1):
            headers.extend([f"第{i}级报告科目代码", f"第{i}级报告科目名称"])
        headers.extend(["数据科目代码", "数据科目名称", "是否汇总", "是否减项", "备注"])

        template_path = Path(__file__).resolve().parents[2] / "download_template" / "data_acct_temp.xlsx"
        if template_path.exists():
            template_wb = load_workbook(template_path)
            style_ws = template_wb["数据模版"] if "数据模版" in template_wb.sheetnames else template_wb.active
            style_cell = style_ws.cell(row=1, column=1)
        else:
            style_cell = None

        wb = Workbook()
        ws = wb.active
        ws.title = "报告数据映射树"
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            if style_cell is not None:
                cell.font = copy(style_cell.font)
                cell.fill = copy(style_cell.fill)
                cell.border = copy(style_cell.border)
                cell.alignment = copy(style_cell.alignment)
                cell.number_format = style_cell.number_format
                cell.protection = copy(style_cell.protection)

        export_rows: list[list[Any]] = []

        def _walk(nodes: list[dict[str, Any]], depth: int) -> None:
            for node in nodes:
                row = [""] * len(headers)
                if node.get("type") == "report":
                    level = min(max(depth, 1), report_levels)
                    col_idx = (level - 1) * 2
                    row[col_idx] = str(node.get("code", ""))
                    row[col_idx + 1] = str(node.get("name", ""))
                    row[12] = int(node.get("is_summary", 0))
                    row[13] = int(node.get("is_minus", 0))
                    row[14] = str(node.get("remark", ""))
                else:
                    row[10] = str(node.get("code", ""))
                    row[11] = str(node.get("name", ""))
                export_rows.append(row)
                children = node.get("children") or []
                if children:
                    _walk(children, depth + 1)

        _walk(roots, 1)
        for i, row in enumerate(export_rows, start=2):
            for j, value in enumerate(row, start=1):
                ws.cell(row=i, column=j, value=value)

        ws.freeze_panes = "A2"
        for col_idx in (2, 4, 6, 8, 10):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions.group(col_letter, col_letter, outline_level=1, hidden=False)

        def _display_width(text: str) -> int:
            width = 0
            for ch in text:
                width += 2 if unicodedata.east_asian_width(ch) in ("F", "W") else 1
            return width

        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row_idx in range(1, ws.max_row + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is None:
                    continue
                text = str(val)
                part_max = max((_display_width(seg) for seg in text.splitlines()), default=0)
                max_len = max(max_len, part_max)
            ws.column_dimensions[col_letter].width = max(6, min(80, max_len + 2))
            ws.column_dimensions[col_letter].bestFit = True

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="report_tree_export.xlsx"'},
        )

    @router.post("/api/report-accounts", response_model=ReportAccountRow)
    async def create_report_account(body: ReportAccountCreate):
        path = common_db_path()
        async with aiosqlite.connect(path) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            cur = await db.execute(
                "SELECT 1 FROM report_account WHERE report_acct_code = ?",
                (body.report_acct_code,),
            )
            if await cur.fetchone():
                raise HTTPException(status_code=409, detail="报告科目代码已存在")
            if body.parent_code:
                cur = await db.execute(
                    "SELECT 1 FROM report_account WHERE report_acct_code = ?",
                    (body.parent_code,),
                )
                if not await cur.fetchone():
                    raise HTTPException(status_code=400, detail="上级报告科目不存在")
            await db.execute(
                """
                INSERT INTO report_account (
                  report_acct_code, report_acct_name, parent_code,
                  is_summary, is_minus, level, is_leaf, remark
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    body.report_acct_code,
                    body.report_acct_name,
                    body.parent_code,
                    int(body.is_summary),
                    int(body.is_minus),
                    body.level,
                    int(body.is_leaf),
                    body.remark,
                ),
            )
            await db.commit()
        await write_operation_log(
            action_type="INSERT",
            action_desc=f"新增报告科目 {body.report_acct_code}",
            target_table="report_account",
            affected_rows=1,
            after_data=body.model_dump(),
        )
        return ReportAccountRow(**body.model_dump())

    @router.patch("/api/report-accounts/{code}", response_model=ReportAccountRow)
    async def update_report_account(code: str, body: ReportAccountUpdate):
        path = common_db_path()
        async with aiosqlite.connect(path) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            cur = await db.execute(
                """
                SELECT report_acct_code, report_acct_name, parent_code, is_summary,
                       is_minus, level, is_leaf, remark
                FROM report_account
                WHERE report_acct_code = ?
                """,
                (code,),
            )
            before = await cur.fetchone()
            if not before:
                raise HTTPException(status_code=404, detail="报告科目不存在")
            updates: list[str] = []
            vals: list[Any] = []
            if body.report_acct_name is not None:
                if not body.report_acct_name.strip():
                    raise HTTPException(status_code=400, detail="报告科目名称不能为空")
                updates.append("report_acct_name = ?")
                vals.append(body.report_acct_name.strip())
            if body.parent_code is not None:
                parent_code = body.parent_code.strip() or None
                if parent_code:
                    if parent_code == code:
                        raise HTTPException(status_code=400, detail="上级科目不能为自身")
                    cur = await db.execute(
                        "SELECT 1 FROM report_account WHERE report_acct_code = ?",
                        (parent_code,),
                    )
                    if not await cur.fetchone():
                        raise HTTPException(status_code=400, detail="上级报告科目不存在")
                updates.append("parent_code = ?")
                vals.append(parent_code)
            if body.is_summary is not None:
                updates.append("is_summary = ?")
                vals.append(int(body.is_summary))
            if body.is_minus is not None:
                updates.append("is_minus = ?")
                vals.append(int(body.is_minus))
            if body.level is not None:
                updates.append("level = ?")
                vals.append(body.level)
            if body.is_leaf is not None:
                updates.append("is_leaf = ?")
                vals.append(int(body.is_leaf))
            if body.remark is not None:
                updates.append("remark = ?")
                vals.append(body.remark)
            if updates:
                vals.append(code)
                await db.execute(
                    f"UPDATE report_account SET {', '.join(updates)} WHERE report_acct_code = ?",
                    vals,
                )
                await db.commit()
            cur = await db.execute(
                """
                SELECT report_acct_code, report_acct_name, parent_code, is_summary,
                       is_minus, level, is_leaf, remark
                FROM report_account
                WHERE report_acct_code = ?
                """,
                (code,),
            )
            row = await cur.fetchone()
        await write_operation_log(
            action_type="UPDATE",
            action_desc=f"更新报告科目 {code}",
            target_table="report_account",
            affected_rows=1,
            before_data={
                "report_acct_code": before[0],
                "report_acct_name": before[1],
                "parent_code": before[2],
                "is_summary": bool(before[3]),
                "is_minus": bool(before[4]),
                "level": before[5],
                "is_leaf": bool(before[6]),
                "remark": before[7],
            },
            after_data={
                "report_acct_code": row[0],
                "report_acct_name": row[1],
                "parent_code": row[2],
                "is_summary": bool(row[3]),
                "is_minus": bool(row[4]),
                "level": row[5],
                "is_leaf": bool(row[6]),
                "remark": row[7],
            },
        )
        return ReportAccountRow(
            report_acct_code=row[0],
            report_acct_name=row[1],
            parent_code=row[2],
            is_summary=bool(row[3]),
            is_minus=bool(row[4]),
            level=int(row[5]),
            is_leaf=bool(row[6]),
            remark=row[7],
        )

    @router.delete("/api/report-accounts/{code}")
    async def delete_report_account(code: str):
        path = common_db_path()
        async with aiosqlite.connect(path) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            cur = await db.execute(
                """
                SELECT report_acct_code, report_acct_name, parent_code, is_summary,
                       is_minus, level, is_leaf, remark
                FROM report_account
                WHERE report_acct_code = ?
                """,
                (code,),
            )
            before = await cur.fetchone()
            if not before:
                raise HTTPException(status_code=404, detail="报告科目不存在")
            cur = await db.execute(
                "SELECT COUNT(*) FROM report_account WHERE parent_code = ?",
                (code,),
            )
            if (await cur.fetchone())[0] > 0:
                raise HTTPException(status_code=409, detail="该报告科目下仍有子节点，无法删除")
            cur = await db.execute(
                "SELECT COUNT(*) FROM report_data_mapping WHERE report_acct_code = ?",
                (code,),
            )
            if (await cur.fetchone())[0] > 0:
                raise HTTPException(status_code=409, detail="该报告科目仍有关联数据科目映射，无法删除")
            await db.execute("DELETE FROM report_account WHERE report_acct_code = ?", (code,))
            await db.commit()
        await write_operation_log(
            action_type="DELETE",
            action_desc=f"删除报告科目 {code}",
            target_table="report_account",
            affected_rows=1,
            before_data={
                "report_acct_code": before[0],
                "report_acct_name": before[1],
                "parent_code": before[2],
                "is_summary": bool(before[3]),
                "is_minus": bool(before[4]),
                "level": before[5],
                "is_leaf": bool(before[6]),
                "remark": before[7],
            },
            after_data=None,
        )
        return {"ok": True}

    @router.get("/api/report-data-mappings", response_model=list[ReportDataMappingRow])
    async def list_report_data_mappings():
        path = common_db_path()
        async with aiosqlite.connect(path) as db:
            cur = await db.execute(
                "SELECT report_acct_code, data_acct_code FROM report_data_mapping ORDER BY report_acct_code, data_acct_code"
            )
            rows = await cur.fetchall()
        return [ReportDataMappingRow(report_acct_code=r[0], data_acct_code=r[1]) for r in rows]

    @router.post("/api/report-data-mappings", response_model=ReportDataMappingRow)
    async def create_report_data_mapping(body: ReportDataMappingCreate):
        path = common_db_path()
        async with aiosqlite.connect(path) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            cur = await db.execute(
                "SELECT 1 FROM report_account WHERE report_acct_code = ?",
                (body.report_acct_code,),
            )
            if not await cur.fetchone():
                raise HTTPException(status_code=400, detail="报告科目不存在")
            cur = await db.execute(
                "SELECT 1 FROM data_account WHERE data_acct_code = ?",
                (body.data_acct_code,),
            )
            if not await cur.fetchone():
                raise HTTPException(status_code=400, detail="数据科目不存在")
            cur = await db.execute(
                """
                SELECT 1 FROM report_data_mapping
                WHERE report_acct_code = ? AND data_acct_code = ?
                """,
                (body.report_acct_code, body.data_acct_code),
            )
            if await cur.fetchone():
                raise HTTPException(status_code=409, detail="映射已存在")
            await db.execute(
                "INSERT INTO report_data_mapping (report_acct_code, data_acct_code) VALUES (?, ?)",
                (body.report_acct_code, body.data_acct_code),
            )
            await db.commit()
        await write_operation_log(
            action_type="INSERT",
            action_desc=f"新增报告-数据映射 {body.report_acct_code} -> {body.data_acct_code}",
            target_table="report_data_mapping",
            affected_rows=1,
            after_data=body.model_dump(),
        )
        return ReportDataMappingRow(**body.model_dump())

    @router.delete("/api/report-data-mappings/{report_code}/{data_code}")
    async def delete_report_data_mapping(report_code: str, data_code: str):
        path = common_db_path()
        async with aiosqlite.connect(path) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            cur = await db.execute(
                """
                SELECT 1 FROM report_data_mapping
                WHERE report_acct_code = ? AND data_acct_code = ?
                """,
                (report_code, data_code),
            )
            if not await cur.fetchone():
                raise HTTPException(status_code=404, detail="映射不存在")
            await db.execute(
                "DELETE FROM report_data_mapping WHERE report_acct_code = ? AND data_acct_code = ?",
                (report_code, data_code),
            )
            await db.commit()
        await write_operation_log(
            action_type="DELETE",
            action_desc=f"删除报告-数据映射 {report_code} -> {data_code}",
            target_table="report_data_mapping",
            affected_rows=1,
            before_data={"report_acct_code": report_code, "data_acct_code": data_code},
            after_data=None,
        )
        return {"ok": True}

    return router
