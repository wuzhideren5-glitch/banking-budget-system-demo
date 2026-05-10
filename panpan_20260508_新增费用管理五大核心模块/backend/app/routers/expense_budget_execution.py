from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
import re
import shutil
import sqlite3
from typing import Any, Awaitable, Callable

import aiosqlite
from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from pydantic import BaseModel
import xlrd

from app.audit import write_operation_log
from app.db_paths import common_db_path


@dataclass(frozen=True)
class FrameworkBudgetDepartmentRow:
    entity_name: str
    group_name: str
    owner_name: str
    budget_department: str


@dataclass(frozen=True)
class FrameworkProductDepartmentRow:
    entity_name: str
    group_name: str
    owner_name: str
    product_department: str


@dataclass(frozen=True)
class FrameworkSubjectRow:
    level_label: str
    budget_subject: str
    manage_department: str
    formula_text: str
    sort_order: int


@dataclass(frozen=True)
class ParsedFramework:
    source_file: Path
    budget_departments: list[FrameworkBudgetDepartmentRow]
    product_departments: list[FrameworkProductDepartmentRow]
    subjects: list[FrameworkSubjectRow]


@dataclass
class MasterPlan:
    dept_rows: list[tuple[str, str, str | None, int, int]]
    dept_mapping_rows: list[tuple[str, str]]
    data_account_upserts: list[tuple[str, str, None, int, str | None, str | None, int, str, str | None]]
    matched_subjects: int
    new_subjects: list[str]
    legacy_subjects: list[str]
    unmatched_products: list[str]


class ExpenseBudgetExecutionExportRequest(BaseModel):
    mode: str = "query"
    perspective: str = "group"
    keyword: str = ""
    include_zero_rows: bool = False
    entity_name: str = ""
    report_month: int | None = None
    include_monthly_actuals: bool = False
    include_last_year_monthly_actuals: bool = False


class ExpenseFrameworkSyncRequest(BaseModel):
    apply_to_master_data: bool = True


def build_expense_budget_execution_router(
    *,
    editable_context_provider: Callable[[], Awaitable[tuple[Path, int, int]]],
) -> APIRouter:
    router = APIRouter()

    class FrameworkContext:
        def __init__(self) -> None:
            self.owner_alias_map: dict[str, str] = {}
            self.owner_to_entity: dict[str, str] = {}
            self.owner_to_group: dict[str, str] = {}
            self.group_to_entity: dict[str, str] = {}
            self.product_department_to_entity: dict[str, str] = {}
            self.product_department_to_owner: dict[str, str] = {}
            self.product_department_to_group: dict[str, str] = {}
            self.subject_alias_map: dict[str, str] = {}

    GOVERNANCE_OWNER_ALIASES = {
        "董事会办公室": "公司治理部",
        "监事会办公室": "公司治理部",
        "董监事会办公室": "公司治理部",
    }

    def _desktop_budget_dir() -> Path:
        return Path.home() / "Desktop" / "预算系统"

    def _source_file(name: str) -> Path:
        path = _desktop_budget_dir() / name
        if not path.exists():
            raise HTTPException(status_code=400, detail=f"缺少源文件：{path}")
        return path

    def _text(value: Any) -> str:
        if value is None:
            return ""
        text = str(value).strip()
        return re.sub(r"\s+", " ", text)

    def _norm_key(value: str) -> str:
        return re.sub(r"\s+", "", _text(value)).lower()

    def _iso_now() -> str:
        return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    def _mtime_text(path: Path) -> str | None:
        try:
            return datetime.fromtimestamp(path.stat().st_mtime, timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
        except Exception:
            return None

    def _canonical_subject(raw_name: str, ctx: FrameworkContext) -> str:
        text = _text(raw_name)
        if not text:
            return ""
        return ctx.subject_alias_map.get(_norm_key(text), text)

    def _extract_data_account_name(data_code_name: str) -> str:
        text = _text(data_code_name)
        if not text:
            return ""
        match = re.match(r"^[A-Za-z]\d+\s+(.+)$", text)
        if match:
            return _text(match.group(1))
        parts = text.split(" ", 1)
        if len(parts) == 2 and re.match(r"^[A-Za-z]\d+$", parts[0]):
            return _text(parts[1])
        return text

    def _extract_product_department(product_code_name: str | None) -> str:
        text = _text(product_code_name)
        if not text:
            return ""
        if "-" in text:
            return _text(text.split("-", 1)[0])
        return text

    def _parse_month(value: Any) -> int | None:
        text = _text(value)
        if not text:
            return None
        tokens = re.findall(r"\d{1,2}", text)
        for token in reversed(tokens):
            month = int(token)
            if 1 <= month <= 12:
                return month
        try:
            month = int(float(text))
            return month if 1 <= month <= 12 else None
        except (TypeError, ValueError):
            return None

    def _to_float(value: Any) -> float:
        if value is None or value == "":
            return 0.0
        try:
            return float(value)
        except Exception:
            return 0.0

    def _strip_leading_code(value: str) -> str:
        text = _text(value)
        stripped = re.sub(r"^[A-Za-z]+\d+\s+", "", text)
        return _text(stripped or text)

    def _canonical_owner_name(owner_name: str, ctx: FrameworkContext) -> str:
        text = _text(owner_name)
        if not text:
            return ""
        return ctx.owner_alias_map.get(_norm_key(text), ctx.owner_alias_map.get(_norm_key(_strip_leading_code(text)), text))

    def _new_month_values() -> list[float]:
        return [0.0] * 12

    def _default_entity_name() -> str:
        return "科技子"

    ENTITY_ORDER = ["微众银行", "科技子", "科技孙"]

    def _entity_sort_key(entity_name: str) -> tuple[int, str]:
        text = _text(entity_name)
        try:
            return (ENTITY_ORDER.index(text), text)
        except ValueError:
            return (len(ENTITY_ORDER), text)

    def _list_available_entities(ctx: FrameworkContext) -> list[str]:
        values = {
            _text(value)
            for value in (
                list(ctx.owner_to_entity.values())
                + list(ctx.group_to_entity.values())
                + list(ctx.product_department_to_entity.values())
            )
            if _text(value)
        }
        return sorted(values, key=_entity_sort_key)

    def _entity_for_group(group_name: str, ctx: FrameworkContext) -> str:
        return ctx.group_to_entity.get(group_name, _default_entity_name())

    def _entity_for_owner(owner_name: str, ctx: FrameworkContext) -> str:
        return ctx.owner_to_entity.get(owner_name, _default_entity_name())

    def _default_group_name(entity_name: str) -> str:
        return entity_name or _default_entity_name()

    def _normalize_leaf_department_name(
        entity_name: str,
        group_name: str,
        owner_name: str,
        department_name: str,
    ) -> str:
        text = _text(department_name)
        if text:
            return text
        # New framework rows like "科技子/科技孙" leave the leaf department blank
        # but should still materialize as a self-mapping node in master data.
        if owner_name and (owner_name == group_name or owner_name == entity_name):
            return owner_name
        return ""

    def _compose_subject_remark(subject: FrameworkSubjectRow, existing_remark: str | None = None) -> str | None:
        parts: list[str] = []
        base = _text(existing_remark)
        if base:
            parts.append(base)
        if subject.manage_department:
            token = f"框架归口管理部门:{subject.manage_department}"
            if token not in parts:
                parts.append(token)
        if subject.formula_text and subject.formula_text != "0":
            token = f"框架公式:{subject.formula_text}"
            if token not in parts:
                parts.append(token)
        return "；".join(parts) or None

    def _parse_framework_source() -> ParsedFramework:
        path = _source_file("费用整体框架.xlsx")
        wb = load_workbook(path, data_only=True)
        required = {"预算部门", "部门预算科目"}
        if not required.issubset(set(wb.sheetnames)):
            raise HTTPException(status_code=400, detail="费用整体框架.xlsx 缺少必要工作表")

        budget_rows: list[FrameworkBudgetDepartmentRow] = []
        seen_budget: set[tuple[str, str, str, str]] = set()
        ws_budget = wb["预算部门"]
        for row_idx in range(2, ws_budget.max_row + 1):
            entity_name = _text(ws_budget.cell(row_idx, 1).value)
            group_name = _text(ws_budget.cell(row_idx, 2).value)
            owner_name = _text(ws_budget.cell(row_idx, 3).value)
            row = FrameworkBudgetDepartmentRow(
                entity_name=entity_name,
                group_name=group_name,
                owner_name=owner_name,
                budget_department=_normalize_leaf_department_name(
                    entity_name,
                    group_name,
                    owner_name,
                    _text(ws_budget.cell(row_idx, 4).value),
                ),
            )
            key = (row.entity_name, row.group_name, row.owner_name, row.budget_department)
            if not all(key) or key in seen_budget:
                continue
            seen_budget.add(key)
            budget_rows.append(row)

        product_rows: list[FrameworkProductDepartmentRow] = []
        if "产品部门" in wb.sheetnames:
            seen_product: set[tuple[str, str, str, str]] = set()
            ws_product = wb["产品部门"]
            for row_idx in range(2, ws_product.max_row + 1):
                entity_name = _text(ws_product.cell(row_idx, 1).value)
                group_name = _text(ws_product.cell(row_idx, 2).value)
                owner_name = _text(ws_product.cell(row_idx, 3).value)
                row = FrameworkProductDepartmentRow(
                    entity_name=entity_name,
                    group_name=group_name,
                    owner_name=owner_name,
                    product_department=_normalize_leaf_department_name(
                        entity_name,
                        group_name,
                        owner_name,
                        _text(ws_product.cell(row_idx, 4).value),
                    ),
                )
                key = (row.entity_name, row.group_name, row.owner_name, row.product_department)
                if not all(key) or key in seen_product:
                    continue
                seen_product.add(key)
                product_rows.append(row)

        subjects: list[FrameworkSubjectRow] = []
        seen_subject: set[str] = set()
        ws_subject = wb["部门预算科目"]
        for row_idx in range(2, ws_subject.max_row + 1):
            subject_name = _text(ws_subject.cell(row_idx, 2).value)
            if not subject_name or subject_name in seen_subject:
                continue
            seen_subject.add(subject_name)
            subjects.append(
                FrameworkSubjectRow(
                    level_label=_text(ws_subject.cell(row_idx, 1).value),
                    budget_subject=subject_name,
                    manage_department=_text(ws_subject.cell(row_idx, 3).value),
                    formula_text=_text(ws_subject.cell(row_idx, 4).value),
                    sort_order=len(subjects) + 1,
                )
            )

        return ParsedFramework(
            source_file=path,
            budget_departments=budget_rows,
            product_departments=product_rows,
            subjects=subjects,
        )

    async def _read_sync_meta() -> dict[str, dict[str, Any]]:
        async with aiosqlite.connect(common_db_path()) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            cur = await db.execute(
                "SELECT sync_key, source_file, source_mtime, synced_at, row_count, note FROM expense_sync_meta"
            )
            rows = await cur.fetchall()
        return {
            str(row[0]): {
                "source_file": str(row[1]),
                "source_mtime": str(row[2]) if row[2] is not None else None,
                "synced_at": str(row[3]),
                "row_count": int(row[4] or 0),
                "note": str(row[5]) if row[5] is not None else None,
            }
            for row in rows
        }

    async def _upsert_sync_meta(sync_key: str, source_file: str, row_count: int, note: str | None = None) -> None:
        async with aiosqlite.connect(common_db_path()) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            await db.execute(
                """
                INSERT INTO expense_sync_meta(sync_key, source_file, source_mtime, synced_at, row_count, note)
                VALUES (?, ?, ?, ?, ?, ?)
                ON CONFLICT(sync_key) DO UPDATE SET
                  source_file = excluded.source_file,
                  source_mtime = excluded.source_mtime,
                  synced_at = excluded.synced_at,
                  row_count = excluded.row_count,
                  note = excluded.note
                """,
                (
                    sync_key,
                    source_file,
                    _mtime_text(Path(source_file)) if source_file else None,
                    _iso_now(),
                    int(row_count),
                    note,
                ),
            )
            await db.commit()

    async def _persist_framework_snapshot(parsed: ParsedFramework) -> None:
        async with aiosqlite.connect(common_db_path()) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            await db.execute("DELETE FROM expense_framework_budget_department")
            await db.execute("DELETE FROM expense_framework_product_department")
            await db.execute("DELETE FROM expense_framework_subject")
            if parsed.budget_departments:
                await db.executemany(
                    """
                    INSERT INTO expense_framework_budget_department(entity_name, group_name, owner_name, budget_department)
                    VALUES (?, ?, ?, ?)
                    """,
                    [(r.entity_name, r.group_name, r.owner_name, r.budget_department) for r in parsed.budget_departments],
                )
            if parsed.product_departments:
                await db.executemany(
                    """
                    INSERT INTO expense_framework_product_department(entity_name, group_name, owner_name, product_department)
                    VALUES (?, ?, ?, ?)
                    """,
                    [(r.entity_name, r.group_name, r.owner_name, r.product_department) for r in parsed.product_departments],
                )
            if parsed.subjects:
                await db.executemany(
                    """
                    INSERT INTO expense_framework_subject(budget_subject, level_label, manage_department, formula_text, sort_order)
                    VALUES (?, ?, ?, ?, ?)
                    """,
                    [(r.budget_subject, r.level_label, r.manage_department, r.formula_text, r.sort_order) for r in parsed.subjects],
                )
            await db.commit()
        total_rows = len(parsed.budget_departments) + len(parsed.product_departments) + len(parsed.subjects)
        await _upsert_sync_meta(
            "framework_import",
            str(parsed.source_file),
            total_rows,
            note=(
                f"预算部门{len(parsed.budget_departments)}行；"
                f"产品部门{len(parsed.product_departments)}行；"
                f"部门预算科目{len(parsed.subjects)}行"
            ),
        )

    async def _load_framework_from_db() -> ParsedFramework | None:
        async with aiosqlite.connect(common_db_path()) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            cur = await db.execute(
                "SELECT entity_name, group_name, owner_name, budget_department FROM expense_framework_budget_department ORDER BY id"
            )
            budget_rows = [
                FrameworkBudgetDepartmentRow(
                    entity_name=_text(row[0]),
                    group_name=_text(row[1]),
                    owner_name=_text(row[2]),
                    budget_department=_text(row[3]),
                )
                for row in await cur.fetchall()
            ]
            cur = await db.execute(
                "SELECT entity_name, group_name, owner_name, product_department FROM expense_framework_product_department ORDER BY id"
            )
            product_rows = [
                FrameworkProductDepartmentRow(
                    entity_name=_text(row[0]),
                    group_name=_text(row[1]),
                    owner_name=_text(row[2]),
                    product_department=_text(row[3]),
                )
                for row in await cur.fetchall()
            ]
            cur = await db.execute(
                "SELECT budget_subject, level_label, manage_department, formula_text, sort_order "
                "FROM expense_framework_subject ORDER BY sort_order, budget_subject"
            )
            subject_rows = [
                FrameworkSubjectRow(
                    level_label=_text(row[1]),
                    budget_subject=_text(row[0]),
                    manage_department=_text(row[2]),
                    formula_text=_text(row[3]),
                    sort_order=int(row[4] or 0),
                )
                for row in await cur.fetchall()
            ]
        if not budget_rows and not product_rows and not subject_rows:
            return None
        meta = await _read_sync_meta()
        source_file = meta.get("framework_import", {}).get("source_file") or str(_source_file("费用整体框架.xlsx"))
        return ParsedFramework(
            source_file=Path(source_file),
            budget_departments=budget_rows,
            product_departments=product_rows,
            subjects=subject_rows,
        )

    def _derive_product_departments_from_budget_departments(
        budget_rows: list[FrameworkBudgetDepartmentRow],
    ) -> list[FrameworkProductDepartmentRow]:
        seen: set[tuple[str, str, str, str]] = set()
        product_rows: list[FrameworkProductDepartmentRow] = []
        for row in budget_rows:
            key = (row.entity_name, row.group_name, row.owner_name, row.budget_department)
            if not all(key) or key in seen:
                continue
            seen.add(key)
            product_rows.append(
                FrameworkProductDepartmentRow(
                    entity_name=row.entity_name,
                    group_name=row.group_name,
                    owner_name=row.owner_name,
                    product_department=row.budget_department,
                )
            )
        return product_rows

    async def _merge_framework_with_existing(parsed: ParsedFramework) -> ParsedFramework:
        existing = await _load_framework_from_db()
        return ParsedFramework(
            source_file=parsed.source_file,
            budget_departments=parsed.budget_departments or (existing.budget_departments if existing else []),
            product_departments=_derive_product_departments_from_budget_departments(
                parsed.budget_departments or (existing.budget_departments if existing else [])
            ),
            subjects=parsed.subjects or (existing.subjects if existing else []),
        )

    def _build_framework_context(parsed: ParsedFramework) -> FrameworkContext:
        ctx = FrameworkContext()
        for row in parsed.budget_departments:
            if row.owner_name:
                ctx.owner_alias_map[_norm_key(row.owner_name)] = row.owner_name
                ctx.owner_alias_map[_norm_key(_strip_leading_code(row.owner_name))] = row.owner_name
            if row.owner_name and row.entity_name and row.owner_name not in ctx.owner_to_entity:
                ctx.owner_to_entity[row.owner_name] = row.entity_name
            if row.group_name and row.entity_name and row.group_name not in ctx.group_to_entity:
                ctx.group_to_entity[row.group_name] = row.entity_name
            if row.owner_name and row.group_name and row.owner_name not in ctx.owner_to_group:
                ctx.owner_to_group[row.owner_name] = row.group_name

        for row in parsed.product_departments:
            if row.owner_name:
                ctx.owner_alias_map[_norm_key(row.owner_name)] = row.owner_name
                ctx.owner_alias_map[_norm_key(_strip_leading_code(row.owner_name))] = row.owner_name
            if row.product_department:
                ctx.product_department_to_entity[row.product_department] = row.entity_name
                ctx.product_department_to_owner[row.product_department] = row.owner_name
                ctx.product_department_to_group[row.product_department] = row.group_name
            if row.group_name and row.entity_name and row.group_name not in ctx.group_to_entity:
                ctx.group_to_entity[row.group_name] = row.entity_name
            if row.owner_name and row.entity_name and row.owner_name not in ctx.owner_to_entity:
                ctx.owner_to_entity[row.owner_name] = row.entity_name
            if row.owner_name and row.group_name and row.owner_name not in ctx.owner_to_group:
                ctx.owner_to_group[row.owner_name] = row.group_name

        for subject in parsed.subjects:
            if subject.budget_subject:
                ctx.subject_alias_map[_norm_key(subject.budget_subject)] = subject.budget_subject

        for alias_name, canonical_name in GOVERNANCE_OWNER_ALIASES.items():
            ctx.owner_alias_map[_norm_key(alias_name)] = canonical_name
            ctx.owner_alias_map[_norm_key(_strip_leading_code(alias_name))] = canonical_name

        return ctx

    async def _load_framework_context() -> tuple[FrameworkContext, str, str]:
        parsed = await _load_framework_from_db()
        meta = await _read_sync_meta()
        has_entity_info = parsed is not None and any(row.entity_name for row in parsed.product_departments)
        if parsed is not None and has_entity_info:
            import_meta = meta.get("framework_import", {})
            source_desc = (
                f"系统内部表（最近同步自 {import_meta.get('source_file', str(parsed.source_file))}"
                f"，同步时间 {import_meta.get('synced_at', '-')})"
            )
            return _build_framework_context(parsed), "internal", source_desc

        parsed = _parse_framework_source()
        return _build_framework_context(parsed), "source", f"{parsed.source_file}（当前直接读取源文件）"

    def _load_actual_rows_from_source(
        ctx: FrameworkContext,
    ) -> tuple[
        dict[tuple[str, str], list[float]],
        dict[tuple[str, str], list[float]],
        dict[tuple[str, str], list[float]],
        int,
        int,
        str,
    ]:
        path = _source_file("部门费用执行.xls")
        wb = xlrd.open_workbook(str(path))
        if not wb.sheet_names():
            raise HTTPException(status_code=400, detail="部门费用执行.xls 缺少工作表")
        ws = wb.sheet_by_name("费用执行表") if "费用执行表" in wb.sheet_names() else wb.sheet_by_index(0)
        if ws.nrows < 2:
            raise HTTPException(status_code=400, detail="部门费用执行.xls 缺少表头")

        headers = [_text(v) for v in ws.row_values(1)]
        header_idx = {name: idx for idx, name in enumerate(headers) if name}
        required_headers = ["费用归属部门", "期间", "金额", "科目描述"]
        missing = [h for h in required_headers if h not in header_idx]
        if missing:
            raise HTTPException(status_code=400, detail=f"部门费用执行.xls 缺少字段：{'、'.join(missing)}")

        actual_by_entity: dict[tuple[str, str], list[float]] = defaultdict(_new_month_values)
        actual_by_group: dict[tuple[str, str], list[float]] = defaultdict(_new_month_values)
        actual_by_owner: dict[tuple[str, str], list[float]] = defaultdict(_new_month_values)
        detail_rows = 0

        for row_idx in range(2, ws.nrows):
            owner_name = _canonical_owner_name(_text(ws.cell_value(row_idx, header_idx["费用归属部门"])), ctx)
            month_idx = _parse_month(ws.cell_value(row_idx, header_idx["期间"]))
            amount = _to_float(ws.cell_value(row_idx, header_idx["金额"]))
            budget_subject = _canonical_subject(
                _text(ws.cell_value(row_idx, header_idx["科目描述"])),
                ctx,
            )
            if not owner_name or month_idx is None or not budget_subject:
                continue
            detail_rows += 1
            entity_name = ctx.owner_to_entity.get(owner_name, _default_entity_name())
            group_name = ctx.owner_to_group.get(owner_name, _default_group_name(entity_name))
            actual_by_entity[(entity_name, budget_subject)][month_idx - 1] += amount
            actual_by_owner[(owner_name, budget_subject)][month_idx - 1] += amount
            actual_by_group[(group_name, budget_subject)][month_idx - 1] += amount

        monthly_rows = sum(1 for values in actual_by_owner.values() for value in values if value != 0)
        return actual_by_entity, actual_by_group, actual_by_owner, detail_rows, monthly_rows, str(path)

    async def _load_actual_rows(
        ctx: FrameworkContext,
    ) -> tuple[
        dict[tuple[str, str], list[float]],
        dict[tuple[str, str], list[float]],
        dict[tuple[str, str], list[float]],
        str,
        str,
    ]:
        async with aiosqlite.connect(common_db_path()) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            cur = await db.execute(
                """
                SELECT owner_name_mapped, budget_subject_mapped, period_ym, amount
                FROM expense_actual_detail_raw
                WHERE owner_matched = 1 AND subject_matched = 1
                ORDER BY owner_name_mapped, budget_subject_mapped, period_ym
                """
            )
            raw_detail_rows = await cur.fetchall()
        if raw_detail_rows:
            actual_by_entity: dict[tuple[str, str], list[float]] = defaultdict(_new_month_values)
            actual_by_group: dict[tuple[str, str], list[float]] = defaultdict(_new_month_values)
            actual_by_owner: dict[tuple[str, str], list[float]] = defaultdict(_new_month_values)
            for owner_name, budget_subject, period_ym, amount in raw_detail_rows:
                owner = _canonical_owner_name(_text(owner_name), ctx)
                subject = _canonical_subject(_text(budget_subject), ctx)
                month_idx = _parse_month(period_ym)
                if not owner or not subject or month_idx is None or month_idx < 1 or month_idx > 12:
                    continue
                numeric = round(float(amount or 0.0), 2)
                entity_name = ctx.owner_to_entity.get(owner, _default_entity_name())
                actual_by_owner[(owner, subject)][month_idx - 1] += numeric
                group_name = ctx.owner_to_group.get(owner, _default_group_name(entity_name))
                actual_by_entity[(entity_name, subject)][month_idx - 1] += numeric
                actual_by_group[(group_name, subject)][month_idx - 1] += numeric
            meta = await _read_sync_meta()
            import_meta = meta.get("actual_import", {})
            source_desc = (
                f"费用执行明细导入（最近同步时间 {import_meta.get('synced_at', '-')}"
                f"，来源 {import_meta.get('source_file', '费用执行明细导入')})"
            )
            return actual_by_entity, actual_by_group, actual_by_owner, "internal", source_desc

        async with aiosqlite.connect(common_db_path()) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            cur = await db.execute(
                "SELECT owner_name, budget_subject, month, amount FROM expense_execution_monthly ORDER BY owner_name, budget_subject, month"
            )
            db_rows = await cur.fetchall()
        if db_rows:
            actual_by_entity: dict[tuple[str, str], list[float]] = defaultdict(_new_month_values)
            actual_by_group: dict[tuple[str, str], list[float]] = defaultdict(_new_month_values)
            actual_by_owner: dict[tuple[str, str], list[float]] = defaultdict(_new_month_values)
            for owner_name, budget_subject, month, amount in db_rows:
                owner = _canonical_owner_name(_text(owner_name), ctx)
                subject = _canonical_subject(_text(budget_subject), ctx)
                month_idx = int(month or 0)
                if not owner or not subject or month_idx < 1 or month_idx > 12:
                    continue
                numeric = round(float(amount or 0.0), 2)
                entity_name = ctx.owner_to_entity.get(owner, _default_entity_name())
                actual_by_owner[(owner, subject)][month_idx - 1] += numeric
                group_name = ctx.owner_to_group.get(owner, _default_group_name(entity_name))
                actual_by_entity[(entity_name, subject)][month_idx - 1] += numeric
                actual_by_group[(group_name, subject)][month_idx - 1] += numeric
            meta = await _read_sync_meta()
            import_meta = meta.get("actual_import", {})
            source_desc = (
                f"系统内部表（最近同步自 {import_meta.get('source_file', str(_source_file('部门费用执行.xls')))}"
                f"，同步时间 {import_meta.get('synced_at', '-')})"
            )
            return actual_by_entity, actual_by_group, actual_by_owner, "internal", source_desc

        actual_by_entity, actual_by_group, actual_by_owner, _detail_rows, _monthly_rows, source_file = _load_actual_rows_from_source(ctx)
        return actual_by_entity, actual_by_group, actual_by_owner, "source", f"{source_file}（当前直接读取源文件）"

    async def _load_budget_rows(
        ctx: FrameworkContext,
        budget_db: Path,
        version_id: int,
    ) -> tuple[
        str,
        int,
        dict[tuple[str, str], float],
        dict[tuple[str, str], float],
        dict[tuple[str, str], float],
    ]:
        async with aiosqlite.connect(budget_db) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            cur = await db.execute(
                "SELECT version_name, current_month FROM version WHERE version_id = ?",
                (version_id,),
            )
            version_row = await cur.fetchone()
            if not version_row:
                raise HTTPException(status_code=400, detail=f"版本 {version_id} 不存在")
            version_name = _text(version_row[0]) or f"V{version_id}"
            current_month = int(version_row[1] or 1)

            cur = await db.execute(
                """
                SELECT product_code_name, data_code_name, value
                FROM budget_summary
                WHERE version_id = ? AND budget_actual = 0
                """,
                (version_id,),
            )
            rows = await cur.fetchall()

        budget_by_entity: dict[tuple[str, str], float] = defaultdict(float)
        budget_by_group: dict[tuple[str, str], float] = defaultdict(float)
        budget_by_owner: dict[tuple[str, str], float] = defaultdict(float)

        for product_code_name, data_code_name, value in rows:
            product_department = _extract_product_department(product_code_name)
            owner_name = ctx.product_department_to_owner.get(product_department, product_department or "未映射产品部门")
            group_name = ctx.product_department_to_group.get(
                product_department,
                ctx.owner_to_group.get(owner_name, _default_group_name(ctx.owner_to_entity.get(owner_name, _default_entity_name()))),
            )
            entity_name = ctx.product_department_to_entity.get(
                product_department,
                ctx.owner_to_entity.get(owner_name, ctx.group_to_entity.get(group_name, _default_entity_name())),
            )
            budget_subject = _canonical_subject(_extract_data_account_name(_text(data_code_name)), ctx)
            if not budget_subject:
                continue
            amount = _to_float(value)
            budget_by_entity[(entity_name, budget_subject)] += amount
            budget_by_owner[(owner_name, budget_subject)] += amount
            budget_by_group[(group_name, budget_subject)] += amount

        return version_name, current_month, budget_by_entity, budget_by_group, budget_by_owner

    async def _load_previous_year_actual_subject_monthly(
        ctx: FrameworkContext,
        budget_db: Path,
        budget_year: int,
        current_month: int,
        entity_name: str = "",
    ) -> tuple[dict[str, list[float]], dict[str, float], str]:
        previous_year_db = budget_db.parent / f"budget_{budget_year - 1}.db"
        if not previous_year_db.exists():
            return {}, {}, f"{previous_year_db}（未找到上一年度库）"

        async with aiosqlite.connect(previous_year_db) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            cur = await db.execute("SELECT version_id, version_name FROM version ORDER BY version_id DESC LIMIT 1")
            version_row = await cur.fetchone()
            if not version_row:
                return {}, {}, f"{previous_year_db}（缺少 version 配置）"
            version_id = int(version_row[0])
            version_name = _text(version_row[1]) or f"V{version_id}"
            cur = await db.execute(
                """
                SELECT product_code_name, data_code_name, month, value
                FROM budget_summary
                WHERE version_id = ? AND budget_actual = 1
                """,
                (version_id,),
            )
            rows = await cur.fetchall()

        monthly_map: dict[str, list[float]] = defaultdict(_new_month_values)
        selected_entity = _text(entity_name)
        for product_code_name, data_code_name, month_text, value in rows:
            month_idx = _parse_month(month_text)
            if month_idx is None:
                continue
            product_department = _extract_product_department(product_code_name)
            owner_name = ctx.product_department_to_owner.get(product_department, product_department or "未映射产品部门")
            group_name = ctx.product_department_to_group.get(
                product_department,
                ctx.owner_to_group.get(owner_name, _default_group_name(ctx.owner_to_entity.get(owner_name, _default_entity_name()))),
            )
            resolved_entity_name = ctx.product_department_to_entity.get(
                product_department,
                ctx.owner_to_entity.get(owner_name, ctx.group_to_entity.get(group_name, _default_entity_name())),
            )
            if selected_entity and resolved_entity_name != selected_entity:
                continue
            budget_subject = _canonical_subject(_extract_data_account_name(_text(data_code_name)), ctx)
            if not budget_subject:
                continue
            monthly_map[budget_subject][month_idx - 1] += _to_float(value)
        totals = {
            subject: round(sum(values[:current_month]), 2)
            for subject, values in monthly_map.items()
        }
        return (
            {subject: [round(amount, 2) for amount in values] for subject, values in monthly_map.items()},
            totals,
            f"{previous_year_db} / {version_name}",
        )

    async def _load_budget_subject_catalog_rows() -> list[dict[str, Any]]:
        async with aiosqlite.connect(common_db_path()) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            cur = await db.execute(
                """
                SELECT id, parent_id, level_number, subject_name, formula_text, sort_order
                FROM budget_subject_catalog
                ORDER BY sort_order, id
                """
            )
            rows = await cur.fetchall()
        if rows:
            return [
                {
                    "id": int(row[0]),
                    "parent_id": int(row[1]) if row[1] is not None else None,
                    "level_number": int(row[2] or 1),
                    "level_label": f"{int(row[2] or 1)}级",
                    "subject_name": _text(row[3]),
                    "formula_text": _text(row[4]) or None,
                    "sort_order": int(row[5] or 0),
                }
                for row in rows
            ]

        parsed = await _load_framework_from_db()
        if parsed is None:
            parsed = _parse_framework_source()
        return [
            {
                "id": idx,
                "parent_id": None,
                "level_number": 1,
                "level_label": _text(subject.level_label) or "1级",
                "subject_name": _text(subject.budget_subject),
                "formula_text": _text(subject.formula_text) or None,
                "sort_order": int(subject.sort_order or idx),
            }
            for idx, subject in enumerate(parsed.subjects, start=1)
            if _text(subject.budget_subject)
        ]

    def _metric_payload(
        monthly_actuals: list[float],
        annual_budget: float,
        last_year_actual: float,
    ) -> dict[str, Any]:
        normalized_monthly_actuals = [round(float(value or 0.0), 2) for value in monthly_actuals]
        current_amount = round(sum(normalized_monthly_actuals), 2)
        budget_amount = round(float(annual_budget), 2)
        last_year_amount = round(float(last_year_actual), 2)
        yoy_change = round(current_amount - last_year_amount, 2)
        return {
            "monthly_actuals": normalized_monthly_actuals,
            "current_actual": current_amount,
            "annual_budget": budget_amount,
            "budget_progress": round(current_amount / budget_amount, 6) if budget_amount else None,
            "yoy_change": yoy_change,
            "yoy_rate": round(yoy_change / last_year_amount, 6) if last_year_amount else None,
            "last_year_actual": last_year_amount,
        }

    def _filter_template_subject_tree(
        nodes: list[dict[str, Any]],
        keyword: str,
    ) -> list[dict[str, Any]]:
        keyword_text = _norm_key(keyword)
        if not keyword_text:
            return nodes
        filtered: list[dict[str, Any]] = []
        for node in nodes:
            child_nodes = _filter_template_subject_tree(list(node.get("children", [])), keyword)
            searchable = " ".join(
                [
                    _text(node.get("subject_name")),
                    _text(node.get("level_label")),
                    _text(node.get("formula_text")),
                ]
            )
            if keyword_text in _norm_key(searchable) or child_nodes:
                next_node = dict(node)
                next_node["children"] = child_nodes
                filtered.append(next_node)
        return filtered

    def _build_template_subject_tree(
        subject_rows: list[dict[str, Any]],
        current_subject_monthly_totals: dict[str, list[float]],
        budget_subject_totals: dict[str, float],
        previous_year_subject_monthly_totals: dict[str, list[float]],
        previous_year_subject_totals: dict[str, float],
    ) -> list[dict[str, Any]]:
        node_map: dict[int, dict[str, Any]] = {}
        roots: list[dict[str, Any]] = []
        for row in subject_rows:
            node_map[int(row["id"])] = {
                "id": int(row["id"]),
                "parent_id": row["parent_id"],
                "level_number": int(row["level_number"]),
                "level_label": _text(row["level_label"]) or f'{int(row["level_number"])}级',
                "subject_name": _text(row["subject_name"]),
                "formula_text": _text(row.get("formula_text")) or None,
                "sort_order": int(row.get("sort_order") or 0),
                "children": [],
            }
        for row in subject_rows:
            node = node_map[int(row["id"])]
            parent_id = row["parent_id"]
            if parent_id is not None and int(parent_id) in node_map:
                node_map[int(parent_id)]["children"].append(node)
            else:
                roots.append(node)

        def sort_nodes(nodes: list[dict[str, Any]]) -> None:
            nodes.sort(key=lambda item: (int(item["sort_order"]), int(item["id"])))
            for child in nodes:
                sort_nodes(list(child["children"]))

        def fill_metrics(node: dict[str, Any]) -> dict[str, Any]:
            monthly_actuals = list(current_subject_monthly_totals.get(node["subject_name"], _new_month_values()))
            previous_year_monthly_actuals = list(
                previous_year_subject_monthly_totals.get(node["subject_name"], _new_month_values())
            )
            annual_budget = float(budget_subject_totals.get(node["subject_name"], 0.0))
            last_year_actual = float(previous_year_subject_totals.get(node["subject_name"], 0.0))
            filled_children: list[dict[str, Any]] = []
            for child in list(node["children"]):
                child_filled = fill_metrics(child)
                filled_children.append(child_filled)
                child_monthly_actuals = list(child_filled.get("monthly_actuals", _new_month_values()))
                child_previous_year_monthly_actuals = list(
                    child_filled.get("previous_year_monthly_actuals", _new_month_values())
                )
                monthly_actuals = [
                    round(monthly_actuals[idx] + child_monthly_actuals[idx], 2)
                    for idx in range(12)
                ]
                previous_year_monthly_actuals = [
                    round(previous_year_monthly_actuals[idx] + child_previous_year_monthly_actuals[idx], 2)
                    for idx in range(12)
                ]
                annual_budget += float(child_filled["annual_budget"])
                last_year_actual += float(child_filled["last_year_actual"])
            next_node = dict(node)
            next_node.update(_metric_payload(monthly_actuals, annual_budget, last_year_actual))
            next_node["previous_year_monthly_actuals"] = [
                round(value, 2) for value in previous_year_monthly_actuals
            ]
            next_node["is_leaf"] = len(filled_children) == 0
            next_node["children"] = filled_children
            return next_node

        sort_nodes(roots)
        return [fill_metrics(node) for node in roots]

    def _build_report_rows(
        *,
        perspective: str,
        ctx: FrameworkContext,
        actual_map: dict[tuple[str, str], list[float]],
        budget_map: dict[tuple[str, str], float],
        keyword: str,
        include_zero_rows: bool,
    ) -> list[dict[str, Any]]:
        keyword_text = _norm_key(keyword)
        all_keys = sorted(set(actual_map.keys()) | set(budget_map.keys()), key=lambda item: (item[0], item[1]))
        rows: list[dict[str, Any]] = []
        for dimension_value, budget_subject in all_keys:
            monthly_actuals = [round(v, 2) for v in list(actual_map.get((dimension_value, budget_subject), _new_month_values()))]
            annual_budget = round(float(budget_map.get((dimension_value, budget_subject), 0.0)), 2)
            cumulative_actual = round(sum(monthly_actuals), 2)
            if not include_zero_rows and cumulative_actual == 0 and annual_budget == 0:
                continue

            if perspective == "entity":
                entity_name = dimension_value
                group_name = ""
                owner_name = ""
            elif perspective == "group":
                entity_name = _entity_for_group(dimension_value, ctx)
                group_name = dimension_value
                owner_name = ""
            else:
                entity_name = _entity_for_owner(dimension_value, ctx)
                owner_name = dimension_value
                group_name = ctx.owner_to_group.get(owner_name, _default_group_name(entity_name))

            searchable = " ".join(
                [
                    perspective,
                    dimension_value,
                    entity_name,
                    group_name,
                    owner_name,
                    budget_subject,
                ]
            )
            if keyword_text and keyword_text not in _norm_key(searchable):
                continue

            rows.append(
                {
                    "perspective": perspective,
                    "dimension_value": dimension_value,
                    "entity_name": entity_name,
                    "group_name": group_name,
                    "owner_dept": owner_name,
                    "budget_subject": budget_subject,
                    "monthly_actuals": monthly_actuals,
                    "cumulative_actual": cumulative_actual,
                    "annual_budget": annual_budget,
                    "execution_rate": round(cumulative_actual / annual_budget, 6) if annual_budget else None,
                }
            )
        return rows

    async def _build_import_status() -> dict[str, Any]:
        meta = await _read_sync_meta()
        async with aiosqlite.connect(common_db_path()) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            counts = {}
            for table_name in (
                "expense_framework_budget_department",
                "expense_framework_product_department",
                "expense_framework_subject",
                "expense_execution_monthly",
            ):
                cur = await db.execute(f"SELECT COUNT(*) FROM {table_name}")
                counts[table_name] = int((await cur.fetchone())[0] or 0)
        return {
            "framework_import": meta.get("framework_import"),
            "master_apply": meta.get("master_apply"),
            "actual_import": meta.get("actual_import"),
            "counts": counts,
        }

    def _match_product_department(product_name: str, product_departments: list[str]) -> str | None:
        candidates = [
            product_department
            for product_department in product_departments
            if product_name == product_department or product_name.startswith(f"{product_department}-")
        ]
        if not candidates:
            return None
        return max(candidates, key=len)

    async def _build_master_plan(parsed: ParsedFramework) -> MasterPlan:
        product_departments = [row.product_department for row in parsed.product_departments]
        product_tree: dict[tuple[str, str], list[str]] = defaultdict(list)
        group_order: list[str] = []
        owners_by_group: dict[str, list[str]] = defaultdict(list)
        for row in parsed.product_departments:
            if row.group_name not in group_order:
                group_order.append(row.group_name)
            if row.owner_name not in owners_by_group[row.group_name]:
                owners_by_group[row.group_name].append(row.owner_name)
            if row.product_department not in product_tree[(row.group_name, row.owner_name)]:
                product_tree[(row.group_name, row.owner_name)].append(row.product_department)

        dept_rows: list[tuple[str, str, str | None, int, int]] = []
        dept_code_by_product_department: dict[str, str] = {}
        for group_idx, group_name in enumerate(group_order, start=1):
            group_code = f"Y{group_idx}"
            dept_rows.append((group_code, group_name, None, 1, 0))
            owner_names = owners_by_group[group_name]
            for owner_idx, owner_name in enumerate(owner_names, start=1):
                owner_code = f"{group_code}{owner_idx:02d}"
                product_list = product_tree[(group_name, owner_name)]
                collapse_to_owner = len(product_list) == 1 and product_list[0] == owner_name
                dept_rows.append((owner_code, owner_name, group_code, 2, 1 if collapse_to_owner else 0))
                if collapse_to_owner:
                    dept_code_by_product_department[owner_name] = owner_code
                    continue
                for product_idx, product_department in enumerate(product_list, start=1):
                    product_code = f"{owner_code}{product_idx:02d}"
                    dept_rows.append((product_code, product_department, owner_code, 3, 1))
                    dept_code_by_product_department[product_department] = product_code

        async with aiosqlite.connect(common_db_path()) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            cur = await db.execute("SELECT product_code, product_name FROM product_type ORDER BY product_code")
            product_type_rows = [(str(row[0]), _text(row[1])) for row in await cur.fetchall()]
            cur = await db.execute(
                "SELECT dept_code, dept_name, parent_code, level, is_leaf FROM dept_account ORDER BY dept_code"
            )
            existing_dept_rows = [
                (
                    str(row[0]),
                    _text(row[1]),
                    str(row[2]) if row[2] is not None else None,
                    int(row[3] or 0),
                    int(row[4] or 0),
                )
                for row in await cur.fetchall()
            ]
            cur = await db.execute(
                """
                SELECT dpm.product_code, dpm.dept_code
                FROM dept_product_mapping dpm
                ORDER BY dpm.product_code
                """
            )
            existing_mapping_rows = [(str(row[0]), str(row[1])) for row in await cur.fetchall()]
            cur = await db.execute(
                """
                SELECT data_acct_code, data_acct_name, budget_formula, actual_formula, need_calc, value_type, remark
                FROM data_account
                ORDER BY data_acct_code
                """
            )
            data_account_rows = await cur.fetchall()

        dept_mapping_rows: list[tuple[str, str]] = []
        unmatched_products: list[str] = []
        existing_mapping_by_product = {product_code: dept_code for product_code, dept_code in existing_mapping_rows}
        existing_dept_by_code = {row[0]: row for row in existing_dept_rows}
        for product_code, product_name in product_type_rows:
            product_department = _match_product_department(product_name, product_departments)
            if product_department:
                dept_code = dept_code_by_product_department.get(product_department)
                if dept_code:
                    dept_mapping_rows.append((dept_code, product_code))
                    continue
            legacy_dept_code = existing_mapping_by_product.get(product_code)
            if legacy_dept_code:
                dept_mapping_rows.append((legacy_dept_code, product_code))
                continue
            unmatched_products.append(product_name)

        needed_legacy_dept_codes: set[str] = {dept_code for dept_code, _product_code in dept_mapping_rows if dept_code not in {row[0] for row in dept_rows}}
        closure_dept_codes: set[str] = set()
        for dept_code in needed_legacy_dept_codes:
            current = dept_code
            visited: set[str] = set()
            while current and current not in visited:
                visited.add(current)
                closure_dept_codes.add(current)
                parent = existing_dept_by_code.get(current, (None, None, None, None, None))[2]
                current = parent or ""
        existing_dept_code_set = {row[0] for row in dept_rows}
        legacy_rows = [
            existing_dept_by_code[dept_code]
            for dept_code in sorted(
                closure_dept_codes,
                key=lambda code: (
                    int(existing_dept_by_code.get(code, ("", "", None, 999, 0))[3] or 999),
                    code,
                ),
            )
            if dept_code in existing_dept_by_code and dept_code not in existing_dept_code_set
        ]
        dept_rows.extend(legacy_rows)

        existing_by_name = {
            _text(row[1]): {
                "data_acct_code": str(row[0]),
                "budget_formula": row[2],
                "actual_formula": row[3],
                "need_calc": int(row[4] or 0),
                "value_type": _text(row[5]) or "金额",
                "remark": row[6],
            }
            for row in data_account_rows
            if _text(row[1])
        }
        used_codes = {str(row[0]) for row in data_account_rows if _text(row[0])}
        max_code_num = 1000
        for code in used_codes:
            match = re.match(r"^A(\d{4})$", code)
            if match:
                max_code_num = max(max_code_num, int(match.group(1)))

        def next_data_code() -> str:
            nonlocal max_code_num
            while True:
                max_code_num += 1
                code = f"A{max_code_num:04d}"
                if code not in used_codes:
                    used_codes.add(code)
                    return code

        matched_subjects = 0
        new_subjects: list[str] = []
        data_account_upserts: list[tuple[str, str, None, int, str | None, str | None, int, str, str | None]] = []
        framework_subject_names = {_text(subject.budget_subject) for subject in parsed.subjects}
        for subject in parsed.subjects:
            existing = existing_by_name.get(subject.budget_subject)
            if existing is not None:
                matched_subjects += 1
                code = str(existing["data_acct_code"])
                budget_formula = existing["budget_formula"]
                actual_formula = existing["actual_formula"]
                need_calc = int(existing["need_calc"] or 0)
                value_type = existing["value_type"] or "金额"
                remark = _compose_subject_remark(subject, existing.get("remark"))
            else:
                code = next_data_code()
                budget_formula = None
                actual_formula = None
                need_calc = 0
                value_type = "金额"
                remark = _compose_subject_remark(subject)
                new_subjects.append(subject.budget_subject)
            data_account_upserts.append(
                (code, subject.budget_subject, None, 1, budget_formula, actual_formula, need_calc, value_type, remark)
            )

        legacy_subjects = sorted(name for name in existing_by_name.keys() if name not in framework_subject_names)
        return MasterPlan(
            dept_rows=dept_rows,
            dept_mapping_rows=dept_mapping_rows,
            data_account_upserts=data_account_upserts,
            matched_subjects=matched_subjects,
            new_subjects=new_subjects,
            legacy_subjects=legacy_subjects,
            unmatched_products=sorted(unmatched_products),
        )

    def _backup_common_db() -> Path:
        common_path = common_db_path()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = common_path.parent / f"common.db.backup_before_expense_framework_{timestamp}"
        shutil.copy2(common_path, backup_path)
        return backup_path

    def _apply_master_plan(plan: MasterPlan) -> Path:
        backup_path = _backup_common_db()
        conn = sqlite3.connect(common_db_path())
        try:
            conn.execute("PRAGMA foreign_keys = ON")
            conn.execute("BEGIN")
            conn.execute("DELETE FROM dept_product_mapping")
            conn.execute("DELETE FROM dept_account")
            conn.executemany(
                """
                INSERT INTO dept_account(dept_code, dept_name, parent_code, level, is_leaf)
                VALUES (?, ?, ?, ?, ?)
                """,
                plan.dept_rows,
            )
            if plan.dept_mapping_rows:
                conn.executemany(
                    "INSERT INTO dept_product_mapping(dept_code, product_code) VALUES (?, ?)",
                    plan.dept_mapping_rows,
                )
            for row in plan.data_account_upserts:
                conn.execute(
                    """
                    INSERT INTO data_account(
                      data_acct_code, data_acct_name, product_code, applies_to_all_products,
                      budget_formula, actual_formula, need_calc, value_type, remark
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(data_acct_code) DO UPDATE SET
                      data_acct_name = excluded.data_acct_name,
                      product_code = excluded.product_code,
                      applies_to_all_products = excluded.applies_to_all_products,
                      budget_formula = excluded.budget_formula,
                      actual_formula = excluded.actual_formula,
                      need_calc = excluded.need_calc,
                      value_type = excluded.value_type,
                      remark = excluded.remark
                    """,
                    row,
                )
            conn.commit()
            return backup_path
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()

    async def _import_actual_snapshot(ctx: FrameworkContext) -> dict[str, Any]:
        _actual_by_entity, _actual_by_group, actual_by_owner, detail_rows, monthly_rows, source_file = _load_actual_rows_from_source(ctx)
        upsert_rows: list[tuple[str, str, int, float]] = []
        for (owner_name, budget_subject), month_values in sorted(actual_by_owner.items()):
            for month_idx, amount in enumerate(month_values, start=1):
                if amount == 0:
                    continue
                upsert_rows.append((owner_name, budget_subject, month_idx, round(amount, 2)))

        async with aiosqlite.connect(common_db_path()) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            await db.execute("DELETE FROM expense_execution_monthly")
            if upsert_rows:
                await db.executemany(
                    """
                    INSERT INTO expense_execution_monthly(owner_name, budget_subject, month, amount)
                    VALUES (?, ?, ?, ?)
                    """,
                    upsert_rows,
                )
            await db.commit()

        await _upsert_sync_meta(
            "actual_import",
            source_file,
            len(upsert_rows),
            note=f"源明细{detail_rows}行；月度汇总{len(actual_by_owner)}组；非零月度单元格{monthly_rows}个",
        )
        await write_operation_log(
            action_type="IMPORT",
            action_desc=f"同步部门费用执行，写入 {len(upsert_rows)} 个月度汇总单元格",
            target_table="expense_execution_monthly",
            affected_rows=len(upsert_rows),
            after_data={
                "source_file": source_file,
                "detail_rows": detail_rows,
                "monthly_subject_rows": len(actual_by_owner),
                "saved_cells": len(upsert_rows),
            },
        )
        return {
            "source_file": source_file,
            "detail_rows": detail_rows,
            "monthly_subject_rows": len(actual_by_owner),
            "saved_cells": len(upsert_rows),
        }

    async def _framework_preview() -> dict[str, Any]:
        parsed = await _merge_framework_with_existing(_parse_framework_source())
        plan = await _build_master_plan(parsed)
        return {
            "source_file": str(parsed.source_file),
            "framework": {
                "group_count": len({row.group_name for row in parsed.product_departments}),
                "owner_count": len({row.owner_name for row in parsed.product_departments}),
                "budget_department_count": len(parsed.budget_departments),
                "product_department_count": len(parsed.product_departments),
                "subject_count": len(parsed.subjects),
            },
            "master_preview": {
                "dept_rows": len(plan.dept_rows),
                "dept_product_mapping_rows": len(plan.dept_mapping_rows),
                "matched_subjects": plan.matched_subjects,
                "new_subjects": len(plan.new_subjects),
                "legacy_subjects": len(plan.legacy_subjects),
                "unmatched_products": len(plan.unmatched_products),
                "sample_new_subjects": plan.new_subjects[:10],
                "sample_legacy_subjects": plan.legacy_subjects[:10],
                "sample_unmatched_products": plan.unmatched_products[:10],
            },
        }

    async def _sync_framework(body: ExpenseFrameworkSyncRequest) -> dict[str, Any]:
        parsed = await _merge_framework_with_existing(_parse_framework_source())
        await _persist_framework_snapshot(parsed)
        result: dict[str, Any] = {
            "source_file": str(parsed.source_file),
            "framework_rows": {
                "budget_departments": len(parsed.budget_departments),
                "product_departments": len(parsed.product_departments),
                "subjects": len(parsed.subjects),
            },
            "master_applied": False,
        }
        await write_operation_log(
            action_type="IMPORT",
            action_desc="同步费用整体框架到内部表",
            target_table="expense_framework_*",
            affected_rows=len(parsed.budget_departments) + len(parsed.product_departments) + len(parsed.subjects),
            after_data=result,
        )
        if body.apply_to_master_data:
            plan = await _build_master_plan(parsed)
            backup_path = _apply_master_plan(plan)
            await _upsert_sync_meta(
                "master_apply",
                str(parsed.source_file),
                len(plan.dept_rows) + len(plan.dept_mapping_rows) + len(plan.data_account_upserts),
                note=(
                    f"部门{len(plan.dept_rows)}行；部门产品映射{len(plan.dept_mapping_rows)}行；"
                    f"数据科目同步{len(plan.data_account_upserts)}行；新增科目{len(plan.new_subjects)}个"
                ),
            )
            await write_operation_log(
                action_type="UPDATE",
                action_desc="应用费用整体框架到部门/科目主数据",
                target_table="dept_account,data_account,dept_product_mapping",
                affected_rows=len(plan.dept_rows) + len(plan.dept_mapping_rows) + len(plan.data_account_upserts),
                after_data={
                    "backup_file": str(backup_path),
                    "dept_rows": len(plan.dept_rows),
                    "dept_product_mapping_rows": len(plan.dept_mapping_rows),
                    "data_account_upserts": len(plan.data_account_upserts),
                    "new_subjects": len(plan.new_subjects),
                    "legacy_subjects": len(plan.legacy_subjects),
                    "unmatched_products": len(plan.unmatched_products),
                },
            )
            result["master_applied"] = True
            result["master_apply"] = {
                "backup_file": str(backup_path),
                "dept_rows": len(plan.dept_rows),
                "dept_product_mapping_rows": len(plan.dept_mapping_rows),
                "data_account_upserts": len(plan.data_account_upserts),
                "matched_subjects": plan.matched_subjects,
                "new_subjects": len(plan.new_subjects),
                "legacy_subjects": len(plan.legacy_subjects),
                "unmatched_products": len(plan.unmatched_products),
                "sample_new_subjects": plan.new_subjects[:10],
                "sample_legacy_subjects": plan.legacy_subjects[:10],
                "sample_unmatched_products": plan.unmatched_products[:10],
            }
        return result

    async def _resolve_report(
        *,
        perspective: str,
        keyword: str,
        include_zero_rows: bool,
    ) -> dict[str, Any]:
        if perspective not in {"entity", "group", "owner_dept"}:
            raise HTTPException(status_code=400, detail="perspective 仅支持 entity、group 或 owner_dept")

        budget_db, budget_year, version_id = await editable_context_provider()
        ctx, framework_source_mode, framework_source_file = await _load_framework_context()
        actual_by_entity, actual_by_group, actual_by_owner, actual_source_mode, actual_source_file = await _load_actual_rows(ctx)
        version_name, current_month, budget_by_entity, budget_by_group, budget_by_owner = await _load_budget_rows(
            ctx,
            budget_db,
            version_id,
        )
        if perspective == "entity":
            actual_map = actual_by_entity
            budget_map = budget_by_entity
        elif perspective == "group":
            actual_map = actual_by_group
            budget_map = budget_by_group
        else:
            actual_map = actual_by_owner
            budget_map = budget_by_owner
        rows = _build_report_rows(
            perspective=perspective,
            ctx=ctx,
            actual_map=actual_map,
            budget_map=budget_map,
            keyword=keyword,
            include_zero_rows=include_zero_rows,
        )
        note_parts = [
            "当前版本支持按“主体”“事业群”“费用归属部门”三种维度查询；预算部门维度已从报表中移除。"
        ]
        if framework_source_mode == "source":
            note_parts.append("费用整体框架当前仍在直读桌面文件，建议先执行“同步框架”。")
        if actual_source_mode == "source":
            note_parts.append("当前未检测到已导入的费用执行明细，系统将直读桌面“部门费用执行.xls”。")
        return {
            "perspective": perspective,
            "budget_year": budget_year,
            "version_id": version_id,
            "version_name": version_name,
            "current_month": current_month,
            "framework_source_mode": framework_source_mode,
            "actual_source_mode": actual_source_mode,
            "framework_source_file": framework_source_file,
            "actual_source_file": actual_source_file,
            "rows": rows,
            "note": " ".join(note_parts),
        }

    async def _resolve_template_report(
        *,
        keyword: str,
        entity_name: str = "",
        report_month: int | None = None,
    ) -> dict[str, Any]:
        budget_db, budget_year, version_id = await editable_context_provider()
        ctx, framework_source_mode, framework_source_file = await _load_framework_context()
        available_entities = _list_available_entities(ctx)
        selected_entity = _text(entity_name)
        (
            _actual_by_entity,
            _actual_by_group,
            actual_by_owner,
            actual_source_mode,
            actual_source_file,
        ) = await _load_actual_rows(ctx)
        version_name, current_month, _budget_by_entity, _budget_by_group, budget_by_owner = await _load_budget_rows(
            ctx,
            budget_db,
            version_id,
        )
        selected_month = int(report_month or current_month)
        if selected_month < 1 or selected_month > 12:
            raise HTTPException(status_code=400, detail="report_month 仅支持 1-12")
        (
            previous_year_subject_monthly_totals,
            previous_year_subject_totals,
            previous_actual_source_file,
        ) = await _load_previous_year_actual_subject_monthly(
            ctx,
            budget_db,
            budget_year,
            selected_month,
            selected_entity,
        )
        current_subject_monthly_totals: dict[str, list[float]] = defaultdict(_new_month_values)
        for (owner_name, budget_subject), month_values in actual_by_owner.items():
            if selected_entity and _entity_for_owner(owner_name, ctx) != selected_entity:
                continue
            for idx in range(selected_month):
                current_subject_monthly_totals[budget_subject][idx] += round(float(month_values[idx] or 0.0), 2)
        budget_subject_totals: dict[str, float] = defaultdict(float)
        for (owner_name, budget_subject), amount in budget_by_owner.items():
            if selected_entity and _entity_for_owner(owner_name, ctx) != selected_entity:
                continue
            budget_subject_totals[budget_subject] += round(float(amount or 0.0), 2)

        subject_rows = await _load_budget_subject_catalog_rows()
        subject_tree = _build_template_subject_tree(
            subject_rows,
            {k: [round(item, 2) for item in values] for k, values in current_subject_monthly_totals.items()},
            {k: round(v, 2) for k, v in budget_subject_totals.items()},
            previous_year_subject_monthly_totals,
            previous_year_subject_totals,
        )
        filtered_tree = _filter_template_subject_tree(subject_tree, keyword)
        note_parts = [
            "月报模式按“部门预算科目”层级展示费用类型，支持逐层展开、收起和右键操作。",
            "本年实际取1月至当前月累计实际，本年预算取年度预算总额，去年同期取上一年度同月累计实际。",
        ]
        if selected_entity:
            note_parts.append(f"当前主体筛选：{selected_entity}。")
        if framework_source_mode == "source":
            note_parts.append("费用整体框架当前仍在直读桌面文件，建议先执行“同步框架”。")
        if actual_source_mode == "source":
            note_parts.append("当前未检测到已导入的费用执行明细，系统将直读桌面“部门费用执行.xls”。")
        return {
            "mode": "template",
            "budget_year": budget_year,
            "version_id": version_id,
            "version_name": version_name,
            "current_month": selected_month,
            "framework_source_mode": framework_source_mode,
            "actual_source_mode": actual_source_mode,
            "framework_source_file": framework_source_file,
            "actual_source_file": actual_source_file,
            "previous_actual_source_file": previous_actual_source_file,
            "available_entities": available_entities,
            "selected_entity_name": selected_entity,
            "template_title": f"{budget_year}年{selected_month}月费用统计表",
            "subject_tree": filtered_tree,
            "note": " ".join(note_parts),
        }

    @router.get("/api/expense-budget-execution")
    async def get_expense_budget_execution(
        mode: str = Query("query"),
        perspective: str = Query("group"),
        keyword: str = Query(""),
        include_zero_rows: bool = Query(False),
        entity_name: str = Query(""),
        report_month: int | None = Query(None),
    ):
        if mode == "template":
            return await _resolve_template_report(keyword=keyword, entity_name=entity_name, report_month=report_month)
        return await _resolve_report(
            perspective=perspective,
            keyword=keyword,
            include_zero_rows=include_zero_rows,
        )

    @router.get("/api/expense-budget-execution/status")
    async def get_expense_budget_execution_status():
        return await _build_import_status()

    @router.get("/api/expense-budget-execution/admin/framework-preview")
    async def preview_expense_framework_sync():
        return await _framework_preview()

    @router.post("/api/expense-budget-execution/admin/framework-sync")
    async def sync_expense_framework(body: ExpenseFrameworkSyncRequest):
        return await _sync_framework(body)

    @router.post("/api/expense-budget-execution/admin/actual-sync")
    async def sync_expense_actual():
        ctx, _framework_source_mode, _framework_source_file = await _load_framework_context()
        return await _import_actual_snapshot(ctx)

    @router.post("/api/expense-budget-execution/export")
    async def export_expense_budget_execution(body: ExpenseBudgetExecutionExportRequest):
        if body.mode == "template":
            report = await _resolve_template_report(
                keyword=body.keyword,
                entity_name=body.entity_name,
                report_month=body.report_month,
            )
            wb = Workbook()
            ws = wb.active
            ws.title = "月报模式"
            headers = [
                "层级",
                "费用类型",
                "本年实际",
            ]
            if body.include_monthly_actuals:
                headers.extend([f"{idx}月实际" for idx in range(1, int(report["current_month"]) + 1)])
            headers.extend(
                [
                    "本年预算",
                    "预算进度%",
                    "本年同比增减额",
                    "本年同比%",
                    "去年同期",
                ]
            )
            if body.include_last_year_monthly_actuals:
                previous_year_short = str(int(report["budget_year"]) - 1)[-2:]
                headers.extend([f"{previous_year_short}年{idx}月实际" for idx in range(1, 13)])
            ws.cell(row=1, column=1, value=report["template_title"])
            ws.cell(row=2, column=1, value=report["note"])
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=4, column=col_idx, value=header)

            def write_nodes(nodes: list[dict[str, Any]], row_idx: int) -> int:
                current_row = row_idx
                for node in nodes:
                    ws.cell(row=current_row, column=1, value=node["level_label"])
                    indent = "  " * max(int(node["level_number"]) - 1, 0)
                    ws.cell(row=current_row, column=2, value=f"{indent}{node['subject_name']}")
                    ws.cell(row=current_row, column=3, value=node["current_actual"])
                    col_idx = 4
                    if body.include_monthly_actuals:
                        for month_value in list(node.get("monthly_actuals", []))[: int(report["current_month"])]:
                            ws.cell(row=current_row, column=col_idx, value=month_value)
                            col_idx += 1
                    ws.cell(row=current_row, column=col_idx, value=node["annual_budget"])
                    progress_cell = ws.cell(row=current_row, column=col_idx + 1, value=node["budget_progress"])
                    progress_cell.number_format = "0.00%"
                    yoy_change_cell = ws.cell(row=current_row, column=col_idx + 2, value=node["yoy_change"])
                    yoy_change_cell.number_format = "0.00"
                    yoy_rate_cell = ws.cell(row=current_row, column=col_idx + 3, value=node["yoy_rate"])
                    yoy_rate_cell.number_format = "0.00%"
                    ws.cell(row=current_row, column=col_idx + 4, value=node["last_year_actual"])
                    if body.include_last_year_monthly_actuals:
                        last_year_col_idx = col_idx + 5
                        for month_value in list(node.get("previous_year_monthly_actuals", []))[:12]:
                            ws.cell(row=current_row, column=last_year_col_idx, value=month_value)
                            last_year_col_idx += 1
                    current_row += 1
                    current_row = write_nodes(list(node.get("children", [])), current_row)
                return current_row

            end_row = write_nodes(list(report.get("subject_tree", [])), 5)
            meta_row = end_row + 1
            ws.cell(row=meta_row, column=1, value="分月实际来源")
            ws.cell(row=meta_row, column=2, value=report["actual_source_file"])
            ws.cell(row=meta_row + 1, column=1, value="去年同期来源")
            ws.cell(row=meta_row + 1, column=2, value=report["previous_actual_source_file"])
            for col_idx, header in enumerate(headers, start=1):
                width = max(len(header) + 4, 14)
                ws.column_dimensions[get_column_letter(col_idx)].width = width

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            filename = "expense_budget_execution_monthly_report.xlsx"
            return StreamingResponse(
                buffer,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="{filename}"'},
            )

        report = await _resolve_report(
            perspective=body.perspective,
            keyword=body.keyword,
            include_zero_rows=body.include_zero_rows,
        )
        wb = Workbook()
        ws = wb.active
        ws.title = "费用预算执行报表"
        headers = [
            "视角",
            "查询维度值",
            "主体",
            "事业群",
            "费用归属部门",
            "部门预算科目",
            "1月实际",
            "2月实际",
            "3月实际",
            "4月实际",
            "5月实际",
            "6月实际",
            "7月实际",
            "8月实际",
            "9月实际",
            "10月实际",
            "11月实际",
            "12月实际",
            "累计实际",
            "年度预算",
            "年度预算执行率",
        ]
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        row_idx = 2
        for row in report["rows"]:
            view_label = {
                "entity": "主体",
                "group": "事业群",
                "owner_dept": "费用归属部门",
            }.get(row["perspective"], row["perspective"])
            ws.cell(row=row_idx, column=1, value=view_label)
            ws.cell(row=row_idx, column=2, value=row["dimension_value"])
            ws.cell(row=row_idx, column=3, value=row["entity_name"])
            ws.cell(row=row_idx, column=4, value=row["group_name"])
            ws.cell(row=row_idx, column=5, value=row["owner_dept"])
            ws.cell(row=row_idx, column=6, value=row["budget_subject"])
            for offset, month_value in enumerate(row["monthly_actuals"], start=7):
                ws.cell(row=row_idx, column=offset, value=month_value)
            ws.cell(row=row_idx, column=19, value=row["cumulative_actual"])
            ws.cell(row=row_idx, column=20, value=row["annual_budget"])
            rate_cell = ws.cell(row=row_idx, column=21, value=row["execution_rate"])
            rate_cell.number_format = "0.00%"
            row_idx += 1

        meta_row = row_idx + 1
        ws.cell(row=meta_row, column=1, value="说明")
        ws.cell(row=meta_row, column=2, value=report["note"])

        for col_idx, header in enumerate(headers, start=1):
            width = max(len(header) + 2, 12)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        filename = f"expense_budget_execution_{body.perspective}.xlsx"
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    return router
