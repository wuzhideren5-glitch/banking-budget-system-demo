from __future__ import annotations

from copy import copy
from io import BytesIO
import json
from pathlib import Path
import re
from typing import Any, Awaitable, Callable
import unicodedata

import aiosqlite
from fastapi import APIRouter, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from app.db_paths import common_db_path
from app.schemas import (
    DeptAccountCreate,
    DeptAccountRow,
    DeptAccountUpdate,
    DeptProductMappingCreate,
    DeptProductMappingRow,
)


def build_dept_catalog_router(
    *,
    normalize_cell: Callable[[Any], str],
    color_row: Callable[[Any, int, int, str], None],
    validate_dept_code_with_parent: Callable[[str, int, str | None], str | None],
    write_operation_log: Callable[..., Awaitable[None]],
) -> APIRouter:
    router = APIRouter()
    hidden_root_names = {"历史架构", "虚拟部门"}
    tech_entities = {"科技子", "科技孙"}
    entity_order = {"微众银行": 0, "科技子": 1, "科技孙": 2}

    def _build_dept_entity_map(
        dept_rows: list[tuple[str, str, str | None, int, int]],
    ) -> dict[str, str]:
        dept_by_code = {
            str(row[0]): {
                "dept_name": str(row[1] or ""),
                "parent_code": str(row[2]) if row[2] is not None else None,
            }
            for row in dept_rows
        }
        cache: dict[str, str] = {}

        def resolve(code: str) -> str:
            if code in cache:
                return cache[code]
            row = dept_by_code.get(code)
            if not row:
                return "微众银行"
            parent_code = row["parent_code"]
            if not parent_code or parent_code not in dept_by_code:
                entity_name = row["dept_name"] if row["dept_name"] in tech_entities else "微众银行"
            else:
                entity_name = resolve(parent_code)
            cache[code] = entity_name
            return entity_name

        return {code: resolve(code) for code in dept_by_code}

    @router.post("/api/dept-accounts/import-preview")
    async def preview_dept_account_import(file: UploadFile = File(...)):
        content = await file.read()
        try:
            wb = load_workbook(filename=BytesIO(content), data_only=True)
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"无法读取Excel文件：{exc}") from exc

        if "数据模版" not in wb.sheetnames:
            raise HTTPException(status_code=400, detail='上传文件缺失“数据模版”工作表，不能上传数据。')

        ws = wb["数据模版"]
        headers = [normalize_cell(c.value) for c in ws[1]]
        if not any(headers):
            raise HTTPException(status_code=400, detail="数据模版工作表第一行字段头为空")

        preview_rows: list[dict[str, str]] = []
        total_rows = 0
        for ridx in range(2, ws.max_row + 1):
            row_values = [normalize_cell(ws.cell(ridx, c).value) for c in range(1, ws.max_column + 1)]
            if not any(row_values):
                continue
            total_rows += 1
            if len(preview_rows) < 20:
                preview_rows.append(
                    {headers[i]: row_values[i] for i in range(min(len(headers), len(row_values))) if headers[i]}
                )

        return {
            "columns": [h for h in headers if h],
            "preview_rows": preview_rows,
            "row_count": total_rows,
        }

    @router.post("/api/dept-accounts/import-apply")
    async def apply_dept_account_import(
        file: UploadFile = File(...),
        mappings_json: str = Form(...),
    ):
        try:
            mappings = json.loads(mappings_json)
        except json.JSONDecodeError as exc:
            raise HTTPException(status_code=400, detail="字段映射格式不合法") from exc
        if not isinstance(mappings, dict):
            raise HTTPException(status_code=400, detail="字段映射格式不合法")

        required_fields = [
            "level1Code", "level1Name",
            "level2Code", "level2Name",
            "level3Code", "level3Name",
            "productCode", "productName",
        ]
        for f in required_fields:
            if not mappings.get(f):
                raise HTTPException(status_code=400, detail=f"字段映射缺失：{f}")

        content = await file.read()
        try:
            wb = load_workbook(filename=BytesIO(content))
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"无法读取Excel文件：{exc}") from exc
        if "数据模版" not in wb.sheetnames:
            raise HTTPException(status_code=400, detail='上传文件缺失“数据模版”工作表，不能上传数据。')

        ws = wb["数据模版"]
        headers = [normalize_cell(c.value) for c in ws[1]]
        header_col = {h: i + 1 for i, h in enumerate(headers) if h}
        for key, col_name in mappings.items():
            if col_name and col_name not in header_col:
                raise HTTPException(status_code=400, detail=f"映射列不存在：{col_name}")
        fail_reason_col = ws.max_column + 1
        ws.cell(row=1, column=fail_reason_col, value="失败原因")

        rows: list[dict[str, Any]] = []
        for ridx in range(2, ws.max_row + 1):
            raw_by_header = {h: normalize_cell(ws.cell(ridx, cidx).value) for h, cidx in header_col.items()}
            if not any(raw_by_header.values()):
                continue
            rows.append(
                {
                    "row_idx": ridx,
                    "lvl1_code": raw_by_header.get(mappings.get("level1Code", ""), "").upper(),
                    "lvl1_name": raw_by_header.get(mappings.get("level1Name", ""), ""),
                    "lvl2_code": raw_by_header.get(mappings.get("level2Code", ""), "").upper(),
                    "lvl2_name": raw_by_header.get(mappings.get("level2Name", ""), ""),
                    "lvl3_code": raw_by_header.get(mappings.get("level3Code", ""), "").upper(),
                    "lvl3_name": raw_by_header.get(mappings.get("level3Name", ""), ""),
                    "product_code": raw_by_header.get(mappings.get("productCode", ""), "").upper(),
                    "product_name": raw_by_header.get(mappings.get("productName", ""), ""),
                }
            )

        dept_code_re = re.compile(r"^Y\d+$")
        product_code_re = re.compile(r"^Z\d{4}$")
        success_count = 0
        overwrite_count = 0
        failed_count = 0
        current_path: dict[int, str] = {}

        path = common_db_path()
        async with aiosqlite.connect(path) as db:
            await db.execute("PRAGMA foreign_keys = ON")

            def _mark_failed(row_idx: int, reasons: list[str] | str) -> None:
                reason_text = "；".join(reasons) if isinstance(reasons, list) else str(reasons)
                ws.cell(row=row_idx, column=fail_reason_col, value=reason_text)
                color_row(ws, row_idx, ws.max_column, "FFFF0000")

            cur = await db.execute("SELECT dept_code, dept_name, parent_code, level FROM dept_account")
            dept_rows = await cur.fetchall()
            existing_dept: dict[str, dict[str, Any]] = {
                str(r[0]): {
                    "name": str(r[1] or ""),
                    "parent_code": str(r[2]) if r[2] is not None else None,
                    "level": int(r[3]),
                }
                for r in dept_rows
            }
            cur = await db.execute("SELECT product_code FROM product_type")
            existing_products = {str(r[0]) for r in await cur.fetchall()}
            cur = await db.execute("SELECT dept_code, product_code FROM dept_product_mapping")
            existing_mapping = {(str(r[0]), str(r[1])) for r in await cur.fetchall()}
            product_to_dept = {prod: dept for dept, prod in existing_mapping}

            for row in rows:
                errors: list[str] = []
                level_entries = []
                for level, code_key, name_key in (
                    (1, "lvl1_code", "lvl1_name"),
                    (2, "lvl2_code", "lvl2_name"),
                    (3, "lvl3_code", "lvl3_name"),
                ):
                    c = row[code_key].strip()
                    n = row[name_key].strip()
                    if c or n:
                        level_entries.append((level, c, n))
                product_code = row["product_code"].strip()
                product_name = row["product_name"].strip()

                is_dept_row = len(level_entries) > 0
                is_product_row = bool(product_code or product_name)
                if is_dept_row and is_product_row:
                    errors.append("同一行不能同时填写部门科目和产品科目")
                elif not is_dept_row and not is_product_row:
                    errors.append("空行无有效内容")

                if is_dept_row:
                    if len(level_entries) != 1:
                        errors.append("每行只允许填写一个层级的部门科目代码和名称")
                    else:
                        level, code, name = level_entries[0]
                        if not code:
                            errors.append("部门科目代码不能为空")
                        elif not dept_code_re.match(code):
                            errors.append("部门科目代码格式错误（示例：Y1、Y11、Y111）")
                        if not name:
                            errors.append("部门科目名称不能为空")
                        parent_code = current_path.get(level - 1) if level > 1 else None
                        code_err = validate_dept_code_with_parent(code, level, parent_code)
                        if code_err:
                            errors.append(code_err)
                        if level < 3 and any(l > level for l in current_path):
                            for l in list(current_path.keys()):
                                if l > level:
                                    current_path.pop(l, None)

                if is_product_row:
                    if level_entries:
                        errors.append("产品映射行必须仅填写产品科目列")
                    if not product_code:
                        errors.append("产品科目代码不能为空")
                    elif not product_code_re.match(product_code):
                        errors.append("产品科目代码格式错误（示例：Z0001）")
                    elif product_code not in existing_products:
                        errors.append("产品科目代码在系统中不存在")
                    target_dept = current_path.get(3) or current_path.get(2) or current_path.get(1)
                    if not target_dept:
                        errors.append("产品映射行之前必须先出现所属部门科目行")
                    elif target_dept not in existing_dept:
                        errors.append("映射目标部门科目在系统中不存在")
                    else:
                        already = product_to_dept.get(product_code)
                        if already and already != target_dept:
                            errors.append(f"产品科目 {product_code} 已映射到部门 {already}，不能重复映射")

                if errors:
                    failed_count += 1
                    _mark_failed(row["row_idx"], errors)
                    continue

                if is_dept_row:
                    level, code, name = level_entries[0]
                    parent_code = current_path.get(level - 1) if level > 1 else None
                    existing = existing_dept.get(code)
                    if existing:
                        if (existing.get("parent_code") or None) != (parent_code or None):
                            failed_count += 1
                            _mark_failed(row["row_idx"], "部门科目代码已存在，但上级科目不匹配")
                            continue
                        if int(existing.get("level") or 0) != level:
                            failed_count += 1
                            _mark_failed(row["row_idx"], "部门科目代码已存在，但层级不匹配")
                            continue
                        await db.execute(
                            "UPDATE dept_account SET dept_name = ? WHERE dept_code = ?",
                            (name, code),
                        )
                        overwrite_count += 1
                        success_count += 1
                        color_row(ws, row["row_idx"], ws.max_column, "FF0000FF")
                    else:
                        await db.execute(
                            """
                            INSERT INTO dept_account (dept_code, dept_name, parent_code, level, is_leaf)
                            VALUES (?, ?, ?, ?, 0)
                            """,
                            (code, name, parent_code, level),
                        )
                        existing_dept[code] = {"name": name, "parent_code": parent_code, "level": level}
                        success_count += 1
                        color_row(ws, row["row_idx"], ws.max_column, "FF008000")
                    current_path[level] = code
                    for l in (level + 1, level + 2):
                        current_path.pop(l, None)
                else:
                    target_dept = current_path.get(3) or current_path.get(2) or current_path.get(1)
                    assert target_dept is not None
                    if (target_dept, product_code) in existing_mapping:
                        overwrite_count += 1
                        success_count += 1
                        color_row(ws, row["row_idx"], ws.max_column, "FF0000FF")
                    else:
                        await db.execute(
                            """
                            INSERT INTO dept_product_mapping (dept_code, product_code)
                            VALUES (?, ?)
                            """,
                            (target_dept, product_code),
                        )
                        existing_mapping.add((target_dept, product_code))
                        product_to_dept[product_code] = target_dept
                        success_count += 1
                        color_row(ws, row["row_idx"], ws.max_column, "FF008000")

            await db.commit()

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        total = success_count + failed_count
        return StreamingResponse(
            out,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": 'attachment; filename="dept_account_import_result.xlsx"',
                "X-Import-Total": str(total),
                "X-Import-Success": str(success_count),
                "X-Import-Overwrite": str(overwrite_count),
                "X-Import-Failed": str(failed_count),
                "Access-Control-Expose-Headers": (
                    "Content-Disposition,"
                    "X-Import-Total,X-Import-Success,X-Import-Overwrite,X-Import-Failed"
                ),
            },
        )

    @router.get("/api/dept-accounts", response_model=list[DeptAccountRow])
    async def list_dept_accounts():
        path = common_db_path()
        async with aiosqlite.connect(path) as db:
            cur = await db.execute(
                "SELECT dept_code, dept_name, parent_code, level, is_leaf FROM dept_account ORDER BY dept_code"
            )
            rows = await cur.fetchall()
        entity_by_code = _build_dept_entity_map(
            [
                (
                    str(r[0]),
                    str(r[1] or ""),
                    str(r[2]) if r[2] is not None else None,
                    int(r[3] or 0),
                    int(r[4] or 0),
                )
                for r in rows
            ]
        )
        return [
            DeptAccountRow(
                dept_code=r[0],
                dept_name=r[1],
                parent_code=r[2],
                level=int(r[3]),
                is_leaf=bool(r[4]),
                entity_name=entity_by_code.get(str(r[0]), "微众银行"),
            )
            for r in rows
        ]

    @router.get("/api/dept-tree/export")
    async def export_dept_tree():
        common_path = common_db_path()
        async with aiosqlite.connect(common_path) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            cur = await db.execute(
                """
                SELECT dept_code, dept_name, parent_code
                FROM dept_account
                """
            )
            dept_rows = await cur.fetchall()
            cur = await db.execute(
                """
                SELECT m.dept_code, m.product_code, p.product_name
                FROM dept_product_mapping m
                LEFT JOIN product_type p ON p.product_code = m.product_code
                """
            )
            mapping_rows = await cur.fetchall()

        dept_nodes: dict[str, dict[str, Any]] = {}
        entity_by_code = _build_dept_entity_map(
            [
                (
                    str(r[0]),
                    str(r[1] or ""),
                    str(r[2]) if r[2] is not None else None,
                    0,
                    0,
                )
                for r in dept_rows
            ]
        )
        for r in dept_rows:
            code = str(r[0])
            dept_nodes[code] = {
                "type": "dept",
                "code": code,
                "name": str(r[1] or ""),
                "parent_code": str(r[2]) if r[2] is not None else None,
                "entity_name": entity_by_code.get(code, "微众银行"),
                "children": [],
            }

        roots: list[dict[str, Any]] = []
        for node in dept_nodes.values():
            parent_code = node["parent_code"]
            if parent_code and parent_code in dept_nodes:
                dept_nodes[parent_code]["children"].append(node)
            else:
                roots.append(node)
        roots = [node for node in roots if str(node.get("name", "")) not in hidden_root_names]
        entity_roots: list[dict[str, Any]] = []
        roots_by_entity: dict[str, list[dict[str, Any]]] = {}
        for node in roots:
            entity_name = str(node.get("entity_name", "") or "微众银行")
            roots_by_entity.setdefault(entity_name, []).append(node)
        for entity_name in sorted(
            roots_by_entity.keys(),
            key=lambda name: (entity_order.get(name, 999), name),
        ):
            entity_roots.append(
                {
                    "type": "entity",
                    "code": "",
                    "name": entity_name,
                    "entity_name": entity_name,
                    "children": roots_by_entity[entity_name],
                }
            )

        for dept_code_raw, product_code_raw, product_name_raw in mapping_rows:
            dept_code = str(dept_code_raw)
            parent = dept_nodes.get(dept_code)
            if not parent:
                continue
            parent["children"].append(
                {
                    "type": "product",
                    "code": str(product_code_raw),
                    "name": str(product_name_raw or ""),
                    "children": [],
                }
            )

        def _sort_tree(nodes: list[dict[str, Any]]) -> None:
            nodes.sort(key=lambda n: str(n.get("code", "")))
            for n in nodes:
                children = n.get("children") or []
                if children:
                    _sort_tree(children)

        _sort_tree(entity_roots)

        headers = [
            "主体",
            "第1级部门科目代码", "第1级部门科目名称",
            "第2级部门科目代码", "第2级部门科目名称",
            "第3级部门科目代码", "第3级部门科目名称",
            "产品科目代码", "产品科目名称",
        ]

        template_path = Path(__file__).resolve().parents[2] / "download_template" / "data_acct_temp.xlsx"
        if template_path.exists():
            template_wb = load_workbook(template_path)
            style_ws = template_wb["数据模版"] if "数据模版" in template_wb.sheetnames else template_wb.active
            style_cell = style_ws.cell(row=1, column=1)
        else:
            style_cell = None

        wb = Workbook()
        ws = wb.active
        ws.title = "部门产品映射树"
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            if style_cell is not None:
                cell.font = copy(style_cell.font)
                cell.fill = copy(style_cell.fill)
                cell.border = copy(style_cell.border)
                cell.alignment = copy(style_cell.alignment)
                cell.number_format = style_cell.number_format
                cell.protection = copy(style_cell.protection)

        export_rows: list[list[Any]] = []

        def _walk(nodes: list[dict[str, Any]], depth: int, current_entity: str = "") -> None:
            for node in nodes:
                row = [""] * len(headers)
                node_type = str(node.get("type", ""))
                if node_type == "entity":
                    entity_name = str(node.get("entity_name", "") or node.get("name", "") or "微众银行")
                    row[0] = entity_name
                elif node_type == "dept":
                    entity_name = str(node.get("entity_name", "") or current_entity or "微众银行")
                    level = min(max(depth, 1), 3)
                    row[0] = entity_name
                    col_idx = 1 + (level - 1) * 2
                    row[col_idx] = str(node.get("code", ""))
                    row[col_idx + 1] = str(node.get("name", ""))
                else:
                    row[0] = current_entity or "微众银行"
                    row[7] = str(node.get("code", ""))
                    row[8] = str(node.get("name", ""))
                export_rows.append(row)
                children = node.get("children") or []
                if children:
                    next_depth = depth if node_type == "entity" else depth + 1
                    next_entity = str(node.get("entity_name", current_entity) or current_entity)
                    _walk(children, next_depth, next_entity)

        _walk(entity_roots, 1)
        for i, row in enumerate(export_rows, start=2):
            for j, value in enumerate(row, start=1):
                ws.cell(row=i, column=j, value=value)

        ws.freeze_panes = "A2"
        for col_idx in (3, 5, 7):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions.group(col_letter, col_letter, outline_level=1, hidden=False)

        def _display_width(text: str) -> int:
            width = 0
            for ch in text:
                width += 2 if unicodedata.east_asian_width(ch) in ("F", "W") else 1
            return width

        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row_idx in range(1, ws.max_row + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is None:
                    continue
                text = str(val)
                part_max = max((_display_width(seg) for seg in text.splitlines()), default=0)
                max_len = max(max_len, part_max)
            ws.column_dimensions[col_letter].width = max(6, min(80, max_len + 2))
            ws.column_dimensions[col_letter].bestFit = True

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="dept_tree_export.xlsx"'},
        )

    @router.post("/api/dept-accounts", response_model=DeptAccountRow)
    async def create_dept_account(body: DeptAccountCreate):
        path = common_db_path()
        async with aiosqlite.connect(path) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            cur = await db.execute("SELECT 1 FROM dept_account WHERE dept_code = ?", (body.dept_code,))
            if await cur.fetchone():
                raise HTTPException(status_code=409, detail="部门代码已存在")
            if body.parent_code:
                cur = await db.execute("SELECT 1 FROM dept_account WHERE dept_code = ?", (body.parent_code,))
                if not await cur.fetchone():
                    raise HTTPException(status_code=400, detail="上级部门不存在")
            await db.execute(
                "INSERT INTO dept_account (dept_code, dept_name, parent_code, level, is_leaf) VALUES (?, ?, ?, ?, ?)",
                (body.dept_code, body.dept_name, body.parent_code, body.level, int(body.is_leaf)),
            )
            await db.commit()
        await write_operation_log(
            action_type="INSERT",
            action_desc=f"新增部门科目 {body.dept_code}",
            target_table="dept_account",
            affected_rows=1,
            after_data=body.model_dump(),
        )
        async with aiosqlite.connect(path) as db:
            cur = await db.execute(
                "SELECT dept_code, dept_name, parent_code, level, is_leaf FROM dept_account ORDER BY dept_code"
            )
            rows = await cur.fetchall()
        entity_by_code = _build_dept_entity_map(
            [
                (
                    str(r[0]),
                    str(r[1] or ""),
                    str(r[2]) if r[2] is not None else None,
                    int(r[3] or 0),
                    int(r[4] or 0),
                )
                for r in rows
            ]
        )
        return DeptAccountRow(
            dept_code=body.dept_code,
            dept_name=body.dept_name,
            parent_code=body.parent_code,
            level=body.level,
            is_leaf=body.is_leaf,
            entity_name=entity_by_code.get(body.dept_code, "微众银行"),
        )

    @router.patch("/api/dept-accounts/{code}", response_model=DeptAccountRow)
    async def update_dept_account(code: str, body: DeptAccountUpdate):
        path = common_db_path()
        async with aiosqlite.connect(path) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            cur = await db.execute(
                "SELECT dept_code, dept_name, parent_code, level, is_leaf FROM dept_account WHERE dept_code = ?",
                (code,),
            )
            before = await cur.fetchone()
            if not before:
                raise HTTPException(status_code=404, detail="部门不存在")
            updates: list[str] = []
            vals: list[Any] = []
            if body.dept_name is not None:
                if not body.dept_name.strip():
                    raise HTTPException(status_code=400, detail="部门名称不能为空")
                updates.append("dept_name = ?")
                vals.append(body.dept_name.strip())
            if body.parent_code is not None:
                parent_code = body.parent_code.strip() or None
                if parent_code:
                    if parent_code == code:
                        raise HTTPException(status_code=400, detail="上级部门不能为自身")
                    cur = await db.execute("SELECT 1 FROM dept_account WHERE dept_code = ?", (parent_code,))
                    if not await cur.fetchone():
                        raise HTTPException(status_code=400, detail="上级部门不存在")
                updates.append("parent_code = ?")
                vals.append(parent_code)
            if body.level is not None:
                updates.append("level = ?")
                vals.append(body.level)
            if body.is_leaf is not None:
                updates.append("is_leaf = ?")
                vals.append(int(body.is_leaf))
            if updates:
                vals.append(code)
                await db.execute(
                    f"UPDATE dept_account SET {', '.join(updates)} WHERE dept_code = ?",
                    vals,
                )
                await db.commit()
            cur = await db.execute(
                "SELECT dept_code, dept_name, parent_code, level, is_leaf FROM dept_account WHERE dept_code = ?",
                (code,),
            )
            row = await cur.fetchone()
        await write_operation_log(
            action_type="UPDATE",
            action_desc=f"更新部门科目 {code}",
            target_table="dept_account",
            affected_rows=1,
            before_data={
                "dept_code": before[0],
                "dept_name": before[1],
                "parent_code": before[2],
                "level": before[3],
                "is_leaf": bool(before[4]),
            },
            after_data={
                "dept_code": row[0],
                "dept_name": row[1],
                "parent_code": row[2],
                "level": row[3],
                "is_leaf": bool(row[4]),
            },
        )
        async with aiosqlite.connect(path) as db:
            cur = await db.execute(
                "SELECT dept_code, dept_name, parent_code, level, is_leaf FROM dept_account ORDER BY dept_code"
            )
            all_rows = await cur.fetchall()
        entity_by_code = _build_dept_entity_map(
            [
                (
                    str(r[0]),
                    str(r[1] or ""),
                    str(r[2]) if r[2] is not None else None,
                    int(r[3] or 0),
                    int(r[4] or 0),
                )
                for r in all_rows
            ]
        )
        return DeptAccountRow(
            dept_code=row[0],
            dept_name=row[1],
            parent_code=row[2],
            level=int(row[3]),
            is_leaf=bool(row[4]),
            entity_name=entity_by_code.get(str(row[0]), "微众银行"),
        )

    @router.delete("/api/dept-accounts/{code}")
    async def delete_dept_account(code: str):
        path = common_db_path()
        async with aiosqlite.connect(path) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            cur = await db.execute(
                "SELECT dept_code, dept_name, parent_code, level, is_leaf FROM dept_account WHERE dept_code = ?",
                (code,),
            )
            before = await cur.fetchone()
            if not before:
                raise HTTPException(status_code=404, detail="部门不存在")
            cur = await db.execute("SELECT COUNT(*) FROM dept_account WHERE parent_code = ?", (code,))
            if (await cur.fetchone())[0] > 0:
                raise HTTPException(status_code=409, detail="该部门下仍有子节点，无法删除")
            cur = await db.execute("SELECT COUNT(*) FROM dept_product_mapping WHERE dept_code = ?", (code,))
            if (await cur.fetchone())[0] > 0:
                raise HTTPException(status_code=409, detail="该部门仍有关联产品映射，无法删除")
            await db.execute("DELETE FROM dept_account WHERE dept_code = ?", (code,))
            await db.commit()
        await write_operation_log(
            action_type="DELETE",
            action_desc=f"删除部门科目 {code}",
            target_table="dept_account",
            affected_rows=1,
            before_data={
                "dept_code": before[0],
                "dept_name": before[1],
                "parent_code": before[2],
                "level": before[3],
                "is_leaf": bool(before[4]),
            },
            after_data=None,
        )
        return {"ok": True}

    @router.get("/api/dept-product-mappings", response_model=list[DeptProductMappingRow])
    async def list_dept_product_mappings():
        path = common_db_path()
        async with aiosqlite.connect(path) as db:
            cur = await db.execute(
                "SELECT dept_code, product_code FROM dept_product_mapping ORDER BY dept_code, product_code"
            )
            rows = await cur.fetchall()
        return [DeptProductMappingRow(dept_code=r[0], product_code=r[1]) for r in rows]

    @router.post("/api/dept-product-mappings", response_model=DeptProductMappingRow)
    async def create_dept_product_mapping(body: DeptProductMappingCreate):
        path = common_db_path()
        async with aiosqlite.connect(path) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            cur = await db.execute("SELECT 1 FROM dept_account WHERE dept_code = ?", (body.dept_code,))
            if not await cur.fetchone():
                raise HTTPException(status_code=400, detail="部门不存在")
            cur = await db.execute("SELECT COUNT(*) FROM dept_account WHERE parent_code = ?", (body.dept_code,))
            if (await cur.fetchone())[0] > 0:
                raise HTTPException(status_code=409, detail="仅允许部门叶子节点挂接产品科目")
            cur = await db.execute("SELECT 1 FROM product_type WHERE product_code = ?", (body.product_code,))
            if not await cur.fetchone():
                raise HTTPException(status_code=400, detail="产品不存在")
            cur = await db.execute(
                "SELECT 1 FROM dept_product_mapping WHERE dept_code = ? AND product_code = ?",
                (body.dept_code, body.product_code),
            )
            if await cur.fetchone():
                raise HTTPException(status_code=409, detail="映射已存在")
            cur = await db.execute(
                "SELECT dept_code FROM dept_product_mapping WHERE product_code = ?",
                (body.product_code,),
            )
            bound = await cur.fetchone()
            if bound:
                raise HTTPException(
                    status_code=409,
                    detail=f"产品 {body.product_code} 已映射到部门 {bound[0]}，不可重复映射",
                )
            await db.execute(
                "INSERT INTO dept_product_mapping (dept_code, product_code) VALUES (?, ?)",
                (body.dept_code, body.product_code),
            )
            await db.commit()
        await write_operation_log(
            action_type="INSERT",
            action_desc=f"新增部门-产品映射 {body.dept_code} -> {body.product_code}",
            target_table="dept_product_mapping",
            affected_rows=1,
            after_data=body.model_dump(),
        )
        return DeptProductMappingRow(**body.model_dump())

    @router.delete("/api/dept-product-mappings/{dept_code}/{product_code}")
    async def delete_dept_product_mapping(dept_code: str, product_code: str):
        path = common_db_path()
        async with aiosqlite.connect(path) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            cur = await db.execute(
                "SELECT 1 FROM dept_product_mapping WHERE dept_code = ? AND product_code = ?",
                (dept_code, product_code),
            )
            if not await cur.fetchone():
                raise HTTPException(status_code=404, detail="映射不存在")
            await db.execute(
                "DELETE FROM dept_product_mapping WHERE dept_code = ? AND product_code = ?",
                (dept_code, product_code),
            )
            await db.commit()
        await write_operation_log(
            action_type="DELETE",
            action_desc=f"删除部门-产品映射 {dept_code} -> {product_code}",
            target_table="dept_product_mapping",
            affected_rows=1,
            before_data={"dept_code": dept_code, "product_code": product_code},
            after_data=None,
        )
        return {"ok": True}

    return router
