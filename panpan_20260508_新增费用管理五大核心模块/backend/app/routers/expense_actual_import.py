from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
import re
from typing import Any, Awaitable, Callable

import aiosqlite
from fastapi import APIRouter, File, HTTPException, Query, UploadFile
import xlrd

from app.db_paths import common_db_path
from app.schemas import (
    ExpenseActualImportApplyResponse,
    ExpenseActualImportBatchRow,
    ExpenseActualImportPreviewResponse,
    ExpenseActualImportPreviewRow,
)


@dataclass(frozen=True)
class FrameworkBudgetDepartmentRow:
    entity_name: str
    group_name: str
    owner_name: str
    budget_department: str


@dataclass(frozen=True)
class FrameworkProductDepartmentRow:
    entity_name: str
    group_name: str
    owner_name: str
    product_department: str


@dataclass(frozen=True)
class FrameworkSubjectRow:
    level_label: str
    budget_subject: str
    manage_department: str
    formula_text: str
    sort_order: int


@dataclass(frozen=True)
class ParsedFramework:
    source_file: Path
    budget_departments: list[FrameworkBudgetDepartmentRow]
    product_departments: list[FrameworkProductDepartmentRow]
    subjects: list[FrameworkSubjectRow]


@dataclass
class ParsedActualDetailRow:
    period_ym: str
    period_text: str
    org_code: str
    org_name: str
    dep_code: str
    dep_name: str
    subject_code: str
    subject_name: str
    amount: float
    fee_type_code: str
    fee_type_name: str
    control_item_code: str
    control_item_name: str
    control_dept_code: str
    owner_name_raw: str
    owner_name_mapped: str | None
    monthly_caliber: str
    budget_subject_raw: str
    budget_subject_mapped: str | None
    owner_matched: bool
    subject_matched: bool
    match_note: str | None


def build_expense_actual_import_router(
    *,
    write_operation_log: Callable[..., Awaitable[None]],
) -> APIRouter:
    router = APIRouter()

    class FrameworkContext:
        def __init__(self) -> None:
            self.owner_alias_map: dict[str, str] = {}
            self.owner_names: set[str] = set()
            self.subject_alias_map: dict[str, str] = {}
            self.subject_names: set[str] = set()

    GOVERNANCE_OWNER_ALIASES = {
        "董事会办公室": "公司治理部",
        "监事会办公室": "公司治理部",
        "董监事会办公室": "公司治理部",
    }

    def _desktop_budget_dir() -> Path:
        return Path.home() / "Desktop" / "预算系统"

    def _source_file(name: str) -> Path:
        path = _desktop_budget_dir() / name
        if not path.exists():
            raise HTTPException(status_code=400, detail=f"缺少源文件：{path}")
        return path

    def _text(value: Any) -> str:
        if value is None:
            return ""
        text = str(value).strip()
        return re.sub(r"\s+", " ", text)

    def _norm_key(value: str) -> str:
        return re.sub(r"\s+", "", _text(value)).lower()

    def _iso_now() -> str:
        return datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")

    def _strip_leading_code(value: str) -> str:
        text = _text(value)
        stripped = re.sub(r"^[A-Za-z]+\d+\s+", "", text)
        return _text(stripped or text)

    def _canonical_owner_name(owner_name: str, ctx: FrameworkContext) -> str:
        text = _text(owner_name)
        if not text:
            return ""
        return ctx.owner_alias_map.get(_norm_key(text), ctx.owner_alias_map.get(_norm_key(_strip_leading_code(text)), text))

    def _canonical_subject(subject_name: str, ctx: FrameworkContext) -> str:
        text = _text(subject_name)
        if not text:
            return ""
        return ctx.subject_alias_map.get(_norm_key(text), text)

    def _normalize_leaf_department_name(entity_name: str, group_name: str, owner_name: str, department_name: str) -> str:
        text = _text(department_name)
        if text:
            return text
        if owner_name and (owner_name == group_name or owner_name == entity_name):
            return owner_name
        return ""

    def _parse_framework_source() -> ParsedFramework:
        path = _source_file("费用整体框架.xlsx")
        from openpyxl import load_workbook

        wb = load_workbook(path, data_only=True)
        required = {"预算部门", "部门预算科目"}
        if not required.issubset(set(wb.sheetnames)):
            raise HTTPException(status_code=400, detail="费用整体框架.xlsx 缺少必要工作表")

        budget_rows: list[FrameworkBudgetDepartmentRow] = []
        seen_budget: set[tuple[str, str, str, str]] = set()
        ws_budget = wb["预算部门"]
        for row_idx in range(2, ws_budget.max_row + 1):
            entity_name = _text(ws_budget.cell(row_idx, 1).value)
            group_name = _text(ws_budget.cell(row_idx, 2).value)
            owner_name = _text(ws_budget.cell(row_idx, 3).value)
            row = FrameworkBudgetDepartmentRow(
                entity_name=entity_name,
                group_name=group_name,
                owner_name=owner_name,
                budget_department=_normalize_leaf_department_name(
                    entity_name,
                    group_name,
                    owner_name,
                    _text(ws_budget.cell(row_idx, 4).value),
                ),
            )
            key = (row.entity_name, row.group_name, row.owner_name, row.budget_department)
            if not all(key) or key in seen_budget:
                continue
            seen_budget.add(key)
            budget_rows.append(row)

        product_rows: list[FrameworkProductDepartmentRow] = []
        if "产品部门" in wb.sheetnames:
            seen_product: set[tuple[str, str, str, str]] = set()
            ws_product = wb["产品部门"]
            for row_idx in range(2, ws_product.max_row + 1):
                entity_name = _text(ws_product.cell(row_idx, 1).value)
                group_name = _text(ws_product.cell(row_idx, 2).value)
                owner_name = _text(ws_product.cell(row_idx, 3).value)
                row = FrameworkProductDepartmentRow(
                    entity_name=entity_name,
                    group_name=group_name,
                    owner_name=owner_name,
                    product_department=_normalize_leaf_department_name(
                        entity_name,
                        group_name,
                        owner_name,
                        _text(ws_product.cell(row_idx, 4).value),
                    ),
                )
                key = (row.entity_name, row.group_name, row.owner_name, row.product_department)
                if not all(key) or key in seen_product:
                    continue
                seen_product.add(key)
                product_rows.append(row)

        subject_rows: list[FrameworkSubjectRow] = []
        seen_subject: set[str] = set()
        ws_subject = wb["部门预算科目"]
        for row_idx in range(2, ws_subject.max_row + 1):
            budget_subject = _text(ws_subject.cell(row_idx, 2).value)
            if not budget_subject or budget_subject in seen_subject:
                continue
            seen_subject.add(budget_subject)
            subject_rows.append(
                FrameworkSubjectRow(
                    level_label=_text(ws_subject.cell(row_idx, 1).value),
                    budget_subject=budget_subject,
                    manage_department=_text(ws_subject.cell(row_idx, 3).value),
                    formula_text=_text(ws_subject.cell(row_idx, 4).value),
                    sort_order=len(subject_rows) + 1,
                )
            )

        return ParsedFramework(path, budget_rows, product_rows, subject_rows)

    async def _load_framework_from_db() -> ParsedFramework | None:
        async with aiosqlite.connect(common_db_path()) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            cur = await db.execute(
                "SELECT entity_name, group_name, owner_name, budget_department FROM expense_framework_budget_department ORDER BY id"
            )
            budget_rows = [
                FrameworkBudgetDepartmentRow(
                    entity_name=_text(row[0]),
                    group_name=_text(row[1]),
                    owner_name=_text(row[2]),
                    budget_department=_text(row[3]),
                )
                for row in await cur.fetchall()
            ]
            cur = await db.execute(
                "SELECT entity_name, group_name, owner_name, product_department FROM expense_framework_product_department ORDER BY id"
            )
            product_rows = [
                FrameworkProductDepartmentRow(
                    entity_name=_text(row[0]),
                    group_name=_text(row[1]),
                    owner_name=_text(row[2]),
                    product_department=_text(row[3]),
                )
                for row in await cur.fetchall()
            ]
            cur = await db.execute(
                "SELECT budget_subject, level_label, manage_department, formula_text, sort_order FROM expense_framework_subject ORDER BY sort_order, budget_subject"
            )
            subject_rows = [
                FrameworkSubjectRow(
                    level_label=_text(row[1]),
                    budget_subject=_text(row[0]),
                    manage_department=_text(row[2]),
                    formula_text=_text(row[3]),
                    sort_order=int(row[4] or 0),
                )
                for row in await cur.fetchall()
            ]
        if not budget_rows and not product_rows and not subject_rows:
            return None
        return ParsedFramework(_source_file("费用整体框架.xlsx"), budget_rows, product_rows, subject_rows)

    async def _load_framework_context() -> FrameworkContext:
        parsed = await _load_framework_from_db()
        if parsed is None:
            parsed = _parse_framework_source()
        ctx = FrameworkContext()
        for row in parsed.budget_departments:
            if row.owner_name:
                ctx.owner_alias_map[_norm_key(row.owner_name)] = row.owner_name
                ctx.owner_alias_map[_norm_key(_strip_leading_code(row.owner_name))] = row.owner_name
                ctx.owner_names.add(row.owner_name)
            if row.budget_department:
                ctx.owner_alias_map[_norm_key(row.budget_department)] = row.owner_name or row.budget_department
        for row in parsed.product_departments:
            if row.owner_name:
                ctx.owner_alias_map[_norm_key(row.owner_name)] = row.owner_name
                ctx.owner_alias_map[_norm_key(_strip_leading_code(row.owner_name))] = row.owner_name
                ctx.owner_names.add(row.owner_name)
            if row.product_department and row.owner_name:
                ctx.owner_alias_map[_norm_key(row.product_department)] = row.owner_name
        for subject in parsed.subjects:
            if subject.budget_subject:
                ctx.subject_alias_map[_norm_key(subject.budget_subject)] = subject.budget_subject
                ctx.subject_names.add(subject.budget_subject)
        for alias_name, canonical_name in GOVERNANCE_OWNER_ALIASES.items():
            ctx.owner_alias_map[_norm_key(alias_name)] = canonical_name
            ctx.owner_alias_map[_norm_key(_strip_leading_code(alias_name))] = canonical_name
            ctx.owner_alias_map[_norm_key(canonical_name)] = canonical_name
            ctx.owner_names.add(canonical_name)
        return ctx

    def _parse_period_ym(value: Any, datemode: int) -> tuple[str, str]:
        if value is None or value == "":
            return "", ""
        if isinstance(value, (int, float)):
            dt = xlrd.xldate.xldate_as_datetime(float(value), datemode)
            return dt.strftime("%Y-%m"), dt.strftime("%Y-%m-%d")
        text = _text(value)
        match = re.search(r"(\d{4})[-/年]?(\d{1,2})", text)
        if match:
            year = int(match.group(1))
            month = int(match.group(2))
            return f"{year:04d}-{month:02d}", text
        return text, text

    def _build_match_note(owner_matched: bool, subject_matched: bool) -> str | None:
        problems: list[str] = []
        if not owner_matched:
            problems.append("费用归属部门未匹配")
        if not subject_matched:
            problems.append("预算科目未匹配")
        return "；".join(problems) or None

    def _parse_actual_file(file_name: str, raw: bytes, ctx: FrameworkContext) -> list[ParsedActualDetailRow]:
        try:
            wb = xlrd.open_workbook(file_contents=raw)
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"无法读取 Excel 文件：{exc}") from exc
        if not wb.sheet_names():
            raise HTTPException(status_code=400, detail="导入文件缺少工作表")
        ws = wb.sheet_by_name("费用执行表") if "费用执行表" in wb.sheet_names() else wb.sheet_by_index(0)
        if ws.nrows < 2:
            raise HTTPException(status_code=400, detail="导入文件缺少表头")

        headers = [_text(v) for v in ws.row_values(1)]
        header_idx = {name: idx for idx, name in enumerate(headers) if name}
        required = [
            "期间",
            "费用发生部门编码",
            "费用发生部门",
            "责任中心编码",
            "责任中心",
            "科目编码",
            "科目描述",
            "金额",
            "费用类别编码",
            "费用类别",
            "管控口径编码",
            "管控口径名称",
            "归口管理部门编码",
            "费用归属部门",
            "费用月报口径",
            "预算科目",
        ]
        missing = [name for name in required if name not in header_idx]
        if missing:
            raise HTTPException(status_code=400, detail=f"导入文件缺少字段：{'、'.join(missing)}")

        rows: list[ParsedActualDetailRow] = []
        for row_idx in range(2, ws.nrows):
            values = ws.row_values(row_idx)
            if not any(_text(v) for v in values):
                continue
            period_ym, period_text = _parse_period_ym(values[header_idx["期间"]], wb.datemode)
            owner_name_raw = _text(values[header_idx["费用归属部门"]])
            owner_candidate = _canonical_owner_name(owner_name_raw, ctx)
            owner_name_mapped = owner_candidate if owner_candidate in ctx.owner_names else None
            budget_subject_raw = _text(values[header_idx["预算科目"]])
            subject_candidate = _canonical_subject(budget_subject_raw, ctx)
            budget_subject_mapped = subject_candidate if subject_candidate in ctx.subject_names else None
            owner_matched = bool(owner_name_mapped)
            subject_matched = bool(budget_subject_mapped)
            rows.append(
                ParsedActualDetailRow(
                    period_ym=period_ym,
                    period_text=period_text,
                    org_code=_text(values[header_idx["费用发生部门编码"]]),
                    org_name=_text(values[header_idx["费用发生部门"]]),
                    dep_code=_text(values[header_idx["责任中心编码"]]),
                    dep_name=_text(values[header_idx["责任中心"]]),
                    subject_code=_text(values[header_idx["科目编码"]]),
                    subject_name=_text(values[header_idx["科目描述"]]),
                    amount=float(values[header_idx["金额"]] or 0.0),
                    fee_type_code=_text(values[header_idx["费用类别编码"]]),
                    fee_type_name=_text(values[header_idx["费用类别"]]),
                    control_item_code=_text(values[header_idx["管控口径编码"]]),
                    control_item_name=_text(values[header_idx["管控口径名称"]]),
                    control_dept_code=_text(values[header_idx["归口管理部门编码"]]),
                    owner_name_raw=owner_name_raw,
                    owner_name_mapped=owner_name_mapped,
                    monthly_caliber=_text(values[header_idx["费用月报口径"]]),
                    budget_subject_raw=budget_subject_raw,
                    budget_subject_mapped=budget_subject_mapped,
                    owner_matched=owner_matched,
                    subject_matched=subject_matched,
                    match_note=_build_match_note(owner_matched, subject_matched),
                )
            )
        if not rows:
            raise HTTPException(status_code=400, detail="导入文件没有可用明细数据")
        return rows

    def _build_preview_response(file_name: str, rows: list[ParsedActualDetailRow]) -> ExpenseActualImportPreviewResponse:
        periods = sorted({row.period_ym for row in rows if row.period_ym})
        matched_owner_rows = sum(1 for row in rows if row.owner_matched)
        matched_subject_rows = sum(1 for row in rows if row.subject_matched)
        unmatched_rows = sum(1 for row in rows if not (row.owner_matched and row.subject_matched))

        def to_preview_row(row: ParsedActualDetailRow) -> ExpenseActualImportPreviewRow:
            if row.owner_matched and row.subject_matched:
                match_status = "已匹配"
            elif row.owner_matched or row.subject_matched:
                match_status = "部分匹配"
            else:
                match_status = "未匹配"
            return ExpenseActualImportPreviewRow(
                period_ym=row.period_ym,
                owner_name_raw=row.owner_name_raw,
                owner_name_mapped=row.owner_name_mapped,
                budget_subject_raw=row.budget_subject_raw,
                budget_subject_mapped=row.budget_subject_mapped,
                amount=round(float(row.amount), 2),
                match_status=match_status,
                match_note=row.match_note,
            )

        preview_rows = [to_preview_row(row) for row in rows[:20]]
        unmatched_preview_rows = [to_preview_row(row) for row in rows if row.match_note][:20]
        return ExpenseActualImportPreviewResponse(
            file_name=file_name,
            row_count=len(rows),
            periods=periods,
            matched_owner_rows=matched_owner_rows,
            matched_subject_rows=matched_subject_rows,
            unmatched_rows=unmatched_rows,
            preview_rows=preview_rows,
            unmatched_preview_rows=unmatched_preview_rows,
        )

    async def _ensure_tables() -> None:
        async with aiosqlite.connect(common_db_path()) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            await db.execute(
                """
                CREATE TABLE IF NOT EXISTS expense_actual_import_batch (
                  id INTEGER PRIMARY KEY AUTOINCREMENT,
                  file_name TEXT NOT NULL,
                  import_mode TEXT NOT NULL,
                  periods_text TEXT,
                  total_rows INTEGER NOT NULL DEFAULT 0,
                  matched_owner_rows INTEGER NOT NULL DEFAULT 0,
                  matched_subject_rows INTEGER NOT NULL DEFAULT 0,
                  unmatched_rows INTEGER NOT NULL DEFAULT 0,
                  created_at TEXT NOT NULL,
                  note TEXT
                )
                """
            )
            await db.execute(
                """
                CREATE TABLE IF NOT EXISTS expense_actual_detail_raw (
                  id INTEGER PRIMARY KEY AUTOINCREMENT,
                  batch_id INTEGER REFERENCES expense_actual_import_batch(id) ON DELETE SET NULL,
                  period_ym TEXT NOT NULL,
                  period_text TEXT,
                  org_code TEXT,
                  org_name TEXT,
                  dep_code TEXT,
                  dep_name TEXT,
                  subject_code TEXT,
                  subject_name TEXT,
                  amount REAL NOT NULL DEFAULT 0,
                  fee_type_code TEXT,
                  fee_type_name TEXT,
                  control_item_code TEXT,
                  control_item_name TEXT,
                  control_dept_code TEXT,
                  owner_name_raw TEXT,
                  owner_name_mapped TEXT,
                  monthly_caliber TEXT,
                  budget_subject_raw TEXT,
                  budget_subject_mapped TEXT,
                  owner_matched INTEGER NOT NULL DEFAULT 0 CHECK (owner_matched IN (0, 1)),
                  subject_matched INTEGER NOT NULL DEFAULT 0 CHECK (subject_matched IN (0, 1)),
                  match_note TEXT
                )
                """
            )
            await db.commit()

    @router.get("/api/expense-actual-import/batches", response_model=list[ExpenseActualImportBatchRow])
    async def list_expense_actual_import_batches():
        await _ensure_tables()
        async with aiosqlite.connect(common_db_path()) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            cur = await db.execute(
                """
                SELECT id, file_name, import_mode, periods_text, total_rows,
                       matched_owner_rows, matched_subject_rows, unmatched_rows, created_at, note
                FROM expense_actual_import_batch
                ORDER BY id DESC
                LIMIT 20
                """
            )
            rows = await cur.fetchall()
        return [
            ExpenseActualImportBatchRow(
                id=int(row[0]),
                file_name=_text(row[1]),
                import_mode=_text(row[2]),
                periods=[item for item in _text(row[3]).split(",") if item],
                total_rows=int(row[4] or 0),
                matched_owner_rows=int(row[5] or 0),
                matched_subject_rows=int(row[6] or 0),
                unmatched_rows=int(row[7] or 0),
                created_at=_text(row[8]),
                note=_text(row[9]) or None,
            )
            for row in rows
        ]

    @router.post("/api/expense-actual-import/import-preview", response_model=ExpenseActualImportPreviewResponse)
    async def preview_expense_actual_import(file: UploadFile = File(...)):
        await _ensure_tables()
        raw = await file.read()
        if not raw:
            raise HTTPException(status_code=400, detail="上传文件为空")
        ctx = await _load_framework_context()
        rows = _parse_actual_file(file.filename or "部门费用执行.xls", raw, ctx)
        return _build_preview_response(file.filename or "部门费用执行.xls", rows)

    @router.post("/api/expense-actual-import/import-apply", response_model=ExpenseActualImportApplyResponse)
    async def apply_expense_actual_import(
        mode: str = Query("append"),
        file: UploadFile = File(...),
    ):
        await _ensure_tables()
        import_mode = _text(mode).lower()
        if import_mode not in {"append", "overwrite"}:
            raise HTTPException(status_code=400, detail="导入模式仅支持 append 或 overwrite")
        raw = await file.read()
        if not raw:
            raise HTTPException(status_code=400, detail="上传文件为空")
        ctx = await _load_framework_context()
        rows = _parse_actual_file(file.filename or "部门费用执行.xls", raw, ctx)
        preview = _build_preview_response(file.filename or "部门费用执行.xls", rows)
        periods = preview.periods
        note = "允许未匹配明细入库并预警"

        async with aiosqlite.connect(common_db_path()) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            if import_mode == "overwrite" and periods:
                placeholders = ",".join("?" for _ in periods)
                await db.execute(
                    f"DELETE FROM expense_actual_detail_raw WHERE period_ym IN ({placeholders})",
                    periods,
                )
            cur = await db.execute(
                """
                INSERT INTO expense_actual_import_batch(
                  file_name, import_mode, periods_text, total_rows,
                  matched_owner_rows, matched_subject_rows, unmatched_rows, created_at, note
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    file.filename or "部门费用执行.xls",
                    import_mode,
                    ",".join(periods),
                    preview.row_count,
                    preview.matched_owner_rows,
                    preview.matched_subject_rows,
                    preview.unmatched_rows,
                    _iso_now(),
                    note,
                ),
            )
            batch_id = int(cur.lastrowid)
            await db.executemany(
                """
                INSERT INTO expense_actual_detail_raw(
                  batch_id, period_ym, period_text, org_code, org_name, dep_code, dep_name,
                  subject_code, subject_name, amount, fee_type_code, fee_type_name,
                  control_item_code, control_item_name, control_dept_code,
                  owner_name_raw, owner_name_mapped, monthly_caliber,
                  budget_subject_raw, budget_subject_mapped, owner_matched, subject_matched, match_note
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                [
                    (
                        batch_id,
                        row.period_ym,
                        row.period_text,
                        row.org_code,
                        row.org_name,
                        row.dep_code,
                        row.dep_name,
                        row.subject_code,
                        row.subject_name,
                        row.amount,
                        row.fee_type_code,
                        row.fee_type_name,
                        row.control_item_code,
                        row.control_item_name,
                        row.control_dept_code,
                        row.owner_name_raw,
                        row.owner_name_mapped,
                        row.monthly_caliber,
                        row.budget_subject_raw,
                        row.budget_subject_mapped,
                        1 if row.owner_matched else 0,
                        1 if row.subject_matched else 0,
                        row.match_note,
                    )
                    for row in rows
                ],
            )
            await db.commit()

        await write_operation_log(
            action_type="IMPORT",
            action_desc=f"导入费用执行明细 {preview.row_count} 行（{import_mode}）",
            target_table="expense_actual_detail_raw",
            affected_rows=preview.row_count,
            after_data={
                "batch_id": batch_id,
                "file_name": file.filename or "部门费用执行.xls",
                "import_mode": import_mode,
                "periods": periods,
                "matched_owner_rows": preview.matched_owner_rows,
                "matched_subject_rows": preview.matched_subject_rows,
                "unmatched_rows": preview.unmatched_rows,
            },
        )
        return ExpenseActualImportApplyResponse(
            batch_id=batch_id,
            file_name=file.filename or "部门费用执行.xls",
            import_mode=import_mode,
            row_count=preview.row_count,
            periods=periods,
            matched_owner_rows=preview.matched_owner_rows,
            matched_subject_rows=preview.matched_subject_rows,
            unmatched_rows=preview.unmatched_rows,
            note=note,
        )

    return router
