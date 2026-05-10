from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from io import BytesIO
from typing import Any, Awaitable, Callable, Literal
from urllib.parse import quote

import aiosqlite
from fastapi import APIRouter, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, Field

from app.db_paths import common_db_path

ScopeType = Literal["entity", "group", "owner"]
ImportMode = Literal["append", "overwrite"]
MonthValueSource = Literal["actual", "forecast"]


class ExpenseForecastScopeOption(BaseModel):
    value: str
    label: str


class ExpenseForecastOwnerGroupOption(BaseModel):
    group_value: str
    group_label: str
    owner_options: list[ExpenseForecastScopeOption] = Field(default_factory=list)


class ExpenseForecastMetaResponse(BaseModel):
    default_year: int
    default_version: str
    version_suggestions: list[str] = Field(default_factory=list)
    entity_options: list[ExpenseForecastScopeOption] = Field(default_factory=list)
    group_options: list[ExpenseForecastScopeOption] = Field(default_factory=list)
    owner_options: list[ExpenseForecastScopeOption] = Field(default_factory=list)
    owner_group_options: list[ExpenseForecastOwnerGroupOption] = Field(default_factory=list)


class ExpenseForecastMonthCell(BaseModel):
    month: int
    value: float
    source: MonthValueSource
    editable: bool = False


class ExpenseForecastRow(BaseModel):
    id: int
    parent_id: int | None = None
    level_number: int
    subject_name: str
    formula_text: str | None = None
    sort_order: int = 0
    is_leaf: bool = False
    months: list[ExpenseForecastMonthCell] = Field(default_factory=list)
    total_value: float = 0


class ExpenseForecastViewResponse(BaseModel):
    year: int
    forecast_version: str
    scope_type: ScopeType
    scope_value: str
    actual_cutoff_month: int = 0
    rows: list[ExpenseForecastRow] = Field(default_factory=list)


class ExpenseForecastCellUpsertRequest(BaseModel):
    year: int
    forecast_version: str
    scope_type: ScopeType
    scope_value: str
    subject_id: int
    month: int
    value: float


class ExpenseForecastCellUpsertResponse(BaseModel):
    updated: bool
    actual_cutoff_month: int


class ExpenseForecastImportPreviewItem(BaseModel):
    row_number: int
    budget_subject: str
    month: int
    value: float
    action: str
    message: str | None = None


class ExpenseForecastImportPreviewResponse(BaseModel):
    file_name: str
    import_mode: ImportMode
    actual_cutoff_month: int
    preview_count: int
    insertable_cells: int
    updatable_cells: int
    skipped_cells: int
    error_cells: int
    items: list[ExpenseForecastImportPreviewItem] = Field(default_factory=list)


class ExpenseForecastImportApplyResponse(BaseModel):
    file_name: str
    import_mode: ImportMode
    actual_cutoff_month: int
    inserted_cells: int
    updated_cells: int
    skipped_cells: int
    error_cells: int


class ExpenseForecastExportRequest(BaseModel):
    year: int
    forecast_version: str
    scope_type: ScopeType
    scope_value: str


def build_expense_forecast_router(
    *,
    default_year: int,
    write_operation_log: Callable[..., Awaitable[None]],
) -> APIRouter:
    router = APIRouter()
    manage_department_alias = {
        "科技管理部": "科技业务",
        "董事会办公室": "公司治理部",
        "监事会办公室": "公司治理部",
    }

    def _text(value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip()

    def _default_version() -> str:
        return datetime.now().strftime("%y%m%d") + "v1"

    async def _ensure_tables() -> None:
        async with aiosqlite.connect(common_db_path()) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            await db.execute(
                """
                CREATE TABLE IF NOT EXISTS expense_forecast_entry (
                  id INTEGER PRIMARY KEY AUTOINCREMENT,
                  forecast_year INTEGER NOT NULL,
                  forecast_version TEXT NOT NULL,
                  scope_type TEXT NOT NULL CHECK (scope_type IN ('entity', 'group', 'owner')),
                  scope_value TEXT NOT NULL,
                  subject_id INTEGER NOT NULL REFERENCES budget_subject_catalog(id) ON DELETE CASCADE,
                  month INTEGER NOT NULL CHECK (month BETWEEN 1 AND 12),
                  forecast_value REAL NOT NULL DEFAULT 0,
                  create_time TEXT NOT NULL,
                  update_time TEXT NOT NULL,
                  UNIQUE (forecast_year, forecast_version, scope_type, scope_value, subject_id, month)
                )
                """
            )
            await db.execute(
                """
                CREATE INDEX IF NOT EXISTS idx_expense_forecast_lookup
                ON expense_forecast_entry(forecast_year, forecast_version, scope_type, scope_value)
                """
            )
            await db.commit()

    async def _load_scope_rows() -> list[tuple[str, str, str]]:
        async with aiosqlite.connect(common_db_path()) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            cur = await db.execute(
                """
                SELECT DISTINCT entity_name, group_name, owner_name
                FROM expense_framework_budget_department
                ORDER BY entity_name, group_name, owner_name
                """
            )
            rows = await cur.fetchall()
        return [(_text(row[0]), _text(row[1]), _text(row[2])) for row in rows]

    async def _load_scope_options() -> tuple[
        list[ExpenseForecastScopeOption],
        list[ExpenseForecastScopeOption],
        list[ExpenseForecastScopeOption],
        list[ExpenseForecastOwnerGroupOption],
    ]:
        rows = await _load_scope_rows()
        entity_values: list[str] = []
        group_values: list[str] = []
        owner_values: list[str] = []
        owner_values_by_group: dict[str, list[str]] = defaultdict(list)
        for entity_name, group_name, owner_name in rows:
            if entity_name and entity_name not in entity_values:
                entity_values.append(entity_name)
            if group_name and group_name not in group_values:
                group_values.append(group_name)
            if owner_name and owner_name not in owner_values:
                owner_values.append(owner_name)
            if group_name and owner_name and owner_name not in owner_values_by_group[group_name]:
                owner_values_by_group[group_name].append(owner_name)
        return (
            [ExpenseForecastScopeOption(value=value, label=value) for value in entity_values],
            [ExpenseForecastScopeOption(value=value, label=value) for value in group_values],
            [ExpenseForecastScopeOption(value=value, label=value) for value in owner_values],
            [
                ExpenseForecastOwnerGroupOption(
                    group_value=group_name,
                    group_label=group_name,
                    owner_options=[
                        ExpenseForecastScopeOption(value=owner_name, label=owner_name)
                        for owner_name in owner_values_by_group[group_name]
                    ],
                )
                for group_name in group_values
            ],
        )

    async def _resolve_scope_owners(scope_type: ScopeType, scope_value: str) -> list[str]:
        rows = await _load_scope_rows()
        owners: list[str] = []
        value = _text(scope_value)
        for entity_name, group_name, owner_name in rows:
            if scope_type == "entity" and entity_name == value and owner_name and owner_name not in owners:
                owners.append(owner_name)
            elif scope_type == "group" and group_name == value and owner_name and owner_name not in owners:
                owners.append(owner_name)
            elif scope_type == "owner" and owner_name == value and owner_name not in owners:
                owners.append(owner_name)
        if not owners:
            raise HTTPException(status_code=400, detail="当前编制口径下没有可用的费用归属部门")
        return owners

    async def _load_budget_subject_rows() -> list[dict[str, Any]]:
        async with aiosqlite.connect(common_db_path()) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            cur = await db.execute(
                """
                SELECT c.id, c.parent_id, c.level_number, c.subject_name, c.formula_text, c.sort_order,
                       EXISTS(SELECT 1 FROM budget_subject_catalog child WHERE child.parent_id = c.id) AS has_children
                FROM budget_subject_catalog c
                ORDER BY COALESCE(c.parent_id, 0), c.sort_order, c.id
                """
            )
            rows = await cur.fetchall()
        return [
            {
                "id": int(row[0]),
                "parent_id": int(row[1]) if row[1] is not None else None,
                "level_number": int(row[2]),
                "subject_name": _text(row[3]),
                "formula_text": _text(row[4]) or None,
                "sort_order": int(row[5] or 0),
                "is_leaf": not bool(row[6]),
            }
            for row in rows
        ]

    async def _load_manage_department_map() -> dict[str, str]:
        async with aiosqlite.connect(common_db_path()) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            cur = await db.execute(
                """
                SELECT budget_subject, manage_department
                FROM expense_framework_subject
                ORDER BY sort_order, budget_subject
                """
            )
            rows = await cur.fetchall()
        return {_text(row[0]): _text(row[1]) for row in rows}

    def _normalized_manage_department(raw_manage_department: str) -> str:
        value = _text(raw_manage_department)
        if not value or value == "使用部门":
            return ""
        return manage_department_alias.get(value, value)

    def _effective_manage_departments(
        subject_rows: list[dict[str, Any]],
        manage_department_map: dict[str, str],
    ) -> tuple[dict[int, str], dict[str, list[str]]]:
        row_by_id = {row["id"]: row for row in subject_rows}
        children_by_parent: dict[int | None, list[int]] = defaultdict(list)
        for row in subject_rows:
            children_by_parent[row["parent_id"]].append(int(row["id"]))

        effective_by_id: dict[int, str] = {}
        effective_by_name: dict[str, list[str]] = defaultdict(list)

        def _walk(node_id: int, inherited_department: str) -> None:
            row = row_by_id[node_id]
            current_department = _normalized_manage_department(
                manage_department_map.get(row["subject_name"], "")
            ) or inherited_department
            effective_by_id[node_id] = current_department
            if current_department and current_department not in effective_by_name[row["subject_name"]]:
                effective_by_name[row["subject_name"]].append(current_department)
            for child_id in children_by_parent.get(node_id, []):
                _walk(child_id, current_department)

        for root_id in children_by_parent.get(None, []):
            _walk(root_id, "")
        return effective_by_id, effective_by_name

    async def _forecast_versions(year: int) -> list[str]:
        await _ensure_tables()
        async with aiosqlite.connect(common_db_path()) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            cur = await db.execute(
                """
                SELECT forecast_version, MAX(update_time) AS latest_time
                FROM expense_forecast_entry
                WHERE forecast_year = ?
                GROUP BY forecast_version
                ORDER BY latest_time DESC, forecast_version DESC
                LIMIT 20
                """,
                (year,),
            )
            rows = await cur.fetchall()
        versions = [_text(row[0]) for row in rows if _text(row[0])]
        default_version = _default_version()
        if default_version not in versions:
            versions.insert(0, default_version)
        return versions

    async def _actual_cutoff_month(year: int) -> int:
        async with aiosqlite.connect(common_db_path()) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            cur = await db.execute(
                """
                SELECT MAX(CAST(substr(period_ym, 6, 2) AS INTEGER))
                FROM expense_actual_detail_raw
                WHERE substr(period_ym, 1, 4) = ?
                """,
                (str(year),),
            )
            row = await cur.fetchone()
            value = int(row[0] or 0) if row and row[0] is not None else 0
            if value > 0:
                return max(0, min(12, value))
            cur = await db.execute("SELECT MAX(month) FROM expense_execution_monthly")
            row = await cur.fetchone()
            value = int(row[0] or 0) if row and row[0] is not None else 0
            return max(0, min(12, value))

    async def _load_actual_map(year: int, owner_names: list[str]) -> dict[tuple[str, str, int], float]:
        if not owner_names:
            return {}
        placeholders = ",".join("?" for _ in owner_names)
        async with aiosqlite.connect(common_db_path()) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            cur = await db.execute(
                f"""
                SELECT owner_name_mapped, budget_subject_mapped, CAST(substr(period_ym, 6, 2) AS INTEGER) AS month, SUM(amount)
                FROM expense_actual_detail_raw
                WHERE owner_matched = 1
                  AND subject_matched = 1
                  AND substr(period_ym, 1, 4) = ?
                  AND owner_name_mapped IN ({placeholders})
                GROUP BY owner_name_mapped, budget_subject_mapped, CAST(substr(period_ym, 6, 2) AS INTEGER)
                """,
                (str(year), *owner_names),
            )
            rows = await cur.fetchall()
            if rows:
                return {
                    (_text(row[0]), _text(row[1]), int(row[2])): float(row[3] or 0)
                    for row in rows
                    if int(row[2] or 0) >= 1 and int(row[2] or 0) <= 12
                }

            cur = await db.execute(
                """
                SELECT 1
                FROM expense_actual_detail_raw
                WHERE owner_matched = 1
                  AND subject_matched = 1
                  AND substr(period_ym, 1, 4) = ?
                LIMIT 1
                """,
                (str(year),),
            )
            raw_exists = await cur.fetchone()
            if raw_exists:
                return {}

            cur = await db.execute(
                f"""
                SELECT owner_name, budget_subject, month, SUM(amount)
                FROM expense_execution_monthly
                WHERE owner_name IN ({placeholders})
                GROUP BY owner_name, budget_subject, month
                """,
                owner_names,
            )
            rows = await cur.fetchall()
            return {
                (_text(row[0]), _text(row[1]), int(row[2])): float(row[3] or 0)
                for row in rows
            }

    async def _load_forecast_map(
        *,
        year: int,
        forecast_version: str,
        owner_names: list[str],
    ) -> dict[tuple[str, int, int], float]:
        await _ensure_tables()
        if not owner_names:
            return {}
        placeholders = ",".join("?" for _ in owner_names)
        async with aiosqlite.connect(common_db_path()) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            cur = await db.execute(
                f"""
                SELECT scope_value, subject_id, month, forecast_value
                FROM expense_forecast_entry
                WHERE forecast_year = ? AND forecast_version = ? AND scope_type = 'owner'
                  AND scope_value IN ({placeholders})
                """,
                (year, forecast_version, *owner_names),
            )
            rows = await cur.fetchall()
        return {
            (_text(row[0]), int(row[1]), int(row[2])): float(row[3] or 0)
            for row in rows
        }

    async def _build_view(
        *,
        year: int,
        forecast_version: str,
        scope_type: ScopeType,
        scope_value: str,
    ) -> ExpenseForecastViewResponse:
        subject_rows = await _load_budget_subject_rows()
        if not subject_rows:
            raise HTTPException(status_code=400, detail="当前没有可用的预算科目树")
        owners = await _resolve_scope_owners(scope_type, scope_value)
        owner_set = set(owners)
        actual_cutoff_month = await _actual_cutoff_month(year)
        manage_department_map = await _load_manage_department_map()
        effective_manage_by_id, _effective_manage_by_name = _effective_manage_departments(
            subject_rows, manage_department_map
        )
        actual_map = await _load_actual_map(year, owners)
        forecast_map = await _load_forecast_map(
            year=year,
            forecast_version=forecast_version,
            owner_names=owners,
        )

        row_by_id = {row["id"]: row for row in subject_rows}
        children_by_parent: dict[int | None, list[int]] = defaultdict(list)
        for row in subject_rows:
            children_by_parent[row["parent_id"]].append(row["id"])

        aggregated_cache: dict[int, tuple[list[ExpenseForecastMonthCell], bool]] = {}

        def _zero_cells(editable: bool = False) -> list[ExpenseForecastMonthCell]:
            return [
                ExpenseForecastMonthCell(month=month, value=0.0, source="forecast", editable=editable)
                for month in range(1, 13)
            ]

        def _permitted_owners(subject_name: str) -> list[str]:
            matched_ids = [row["id"] for row in subject_rows if row["subject_name"] == subject_name]
            manage_department = ""
            if matched_ids:
                manage_department = effective_manage_by_id.get(int(matched_ids[0]), "")
            if not manage_department:
                return owners
            if manage_department in owner_set:
                return [manage_department]
            return []

        def _aggregate(node_id: int) -> tuple[list[ExpenseForecastMonthCell], bool]:
            cached = aggregated_cache.get(node_id)
            if cached is not None:
                return cached
            row = row_by_id[node_id]
            children = children_by_parent.get(node_id, [])
            result = _zero_cells(editable=False)
            permitted_owners = _permitted_owners(row["subject_name"])
            has_visible_child = False

            for month in range(1, 13):
                actual_value = sum(
                    float(actual_map.get((owner_name, row["subject_name"], month), 0.0))
                    for owner_name in permitted_owners
                )
                forecast_value = sum(
                    float(forecast_map.get((owner_name, node_id, month), 0.0))
                    for owner_name in permitted_owners
                )
                editable = (
                    scope_type == "owner"
                    and bool(row["is_leaf"])
                    and not row["formula_text"]
                    and bool(permitted_owners)
                    and month > actual_cutoff_month
                )
                source: MonthValueSource = "actual" if month <= actual_cutoff_month else "forecast"
                result[month - 1] = ExpenseForecastMonthCell(
                    month=month,
                    value=actual_value if source == "actual" else forecast_value,
                    source=source,
                    editable=editable,
                )

            for child_id in children:
                child_cells, child_visible = _aggregate(child_id)
                has_visible_child = has_visible_child or child_visible
                for idx in range(12):
                    result[idx] = ExpenseForecastMonthCell(
                        month=result[idx].month,
                        value=result[idx].value + child_cells[idx].value,
                        source=result[idx].source,
                        editable=False,
                    )
            self_visible = bool(permitted_owners) and (
                bool(row["is_leaf"]) or any(abs(cell.value) > 1e-9 for cell in result)
            )
            visible = has_visible_child or self_visible
            aggregated_cache[node_id] = (result, visible)
            return result, visible

        ordered_rows: list[ExpenseForecastRow] = []

        def _walk(parent_id: int | None) -> None:
            for node_id in children_by_parent.get(parent_id, []):
                row = row_by_id[node_id]
                cells, visible = _aggregate(node_id)
                if not visible:
                    continue
                ordered_rows.append(
                    ExpenseForecastRow(
                        id=node_id,
                        parent_id=row["parent_id"],
                        level_number=row["level_number"],
                        subject_name=row["subject_name"],
                        formula_text=row["formula_text"],
                        sort_order=row["sort_order"],
                        is_leaf=bool(row["is_leaf"]),
                        months=cells,
                        total_value=sum(cell.value for cell in cells),
                    )
                )
                _walk(node_id)

        _walk(None)
        return ExpenseForecastViewResponse(
            year=year,
            forecast_version=forecast_version,
            scope_type=scope_type,
            scope_value=scope_value,
            actual_cutoff_month=actual_cutoff_month,
            rows=ordered_rows,
        )

    async def _subject_lookup() -> tuple[dict[int, dict[str, Any]], dict[str, list[dict[str, Any]]]]:
        rows = await _load_budget_subject_rows()
        by_id = {row["id"]: row for row in rows}
        by_name: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in rows:
            by_name[row["subject_name"]].append(row)
        return by_id, by_name

    def _normalize_scope_type(raw: str) -> ScopeType:
        value = _text(raw).lower()
        if value not in {"entity", "group", "owner"}:
            raise HTTPException(status_code=400, detail="编制口径仅支持 entity、group、owner")
        return value  # type: ignore[return-value]

    def _normalize_import_mode(raw: str) -> ImportMode:
        value = _text(raw).lower()
        if value not in {"append", "overwrite"}:
            raise HTTPException(status_code=400, detail="导入模式仅支持 append 或 overwrite")
        return value  # type: ignore[return-value]

    def _parse_import_file(raw: bytes) -> list[dict[str, Any]]:
        if not raw:
            raise HTTPException(status_code=400, detail="上传文件为空")
        wb = load_workbook(BytesIO(raw), data_only=True)
        ws = wb[wb.sheetnames[0]]
        header_map: dict[str, int] = {}
        for col in range(1, ws.max_column + 1):
            key = _text(ws.cell(1, col).value)
            if key:
                header_map[key] = col
        subject_col = header_map.get("预算科目")
        if subject_col is None:
            raise HTTPException(status_code=400, detail="导入模板缺少“预算科目”列")
        month_cols: dict[int, int] = {}
        for key, col in header_map.items():
            normalized = key.upper().replace("月", "").replace("M", "").strip()
            if normalized.isdigit():
                month = int(normalized)
                if 1 <= month <= 12:
                    month_cols[month] = col
        if not month_cols:
            raise HTTPException(status_code=400, detail="导入模板缺少月份列，请使用 M1~M12 或 1月~12月")

        rows: list[dict[str, Any]] = []
        for row_idx in range(2, ws.max_row + 1):
            budget_subject = _text(ws.cell(row_idx, subject_col).value)
            if not budget_subject:
                continue
            for month, col in month_cols.items():
                raw_value = ws.cell(row_idx, col).value
                if raw_value in (None, ""):
                    continue
                try:
                    value = float(raw_value)
                except (TypeError, ValueError):
                    rows.append(
                        {
                            "row_number": row_idx,
                            "budget_subject": budget_subject,
                            "month": month,
                            "value": 0.0,
                            "error": f"月份 M{month} 不是有效数字",
                        }
                    )
                    continue
                rows.append(
                    {
                        "row_number": row_idx,
                        "budget_subject": budget_subject,
                        "month": month,
                        "value": value,
                        "error": None,
                    }
                )
        return rows

    async def _preview_import(
        *,
        file_name: str,
        raw: bytes,
        year: int,
        forecast_version: str,
        scope_type: ScopeType,
        scope_value: str,
        import_mode: ImportMode,
    ) -> tuple[ExpenseForecastImportPreviewResponse, list[dict[str, Any]]]:
        if scope_type != "owner":
            raise HTTPException(status_code=400, detail="费用预估导入仅支持在费用归属部门口径下执行")
        parsed_rows = _parse_import_file(raw)
        actual_cutoff_month = await _actual_cutoff_month(year)
        by_id, by_name = await _subject_lookup()
        manage_department_map = await _load_manage_department_map()
        _effective_manage_by_id, effective_manage_by_name = _effective_manage_departments(
            list(by_id.values()), manage_department_map
        )
        forecast_map = await _load_forecast_map(
            year=year,
            forecast_version=forecast_version,
            owner_names=[scope_value],
        )
        preview_items: list[ExpenseForecastImportPreviewItem] = []
        normalized_rows: list[dict[str, Any]] = []
        inserted = 0
        updated = 0
        skipped = 0
        errors = 0

        for item in parsed_rows:
            row_no = int(item["row_number"])
            subject_name = _text(item["budget_subject"])
            month = int(item["month"])
            value = float(item["value"])
            row_error = item.get("error")
            if row_error:
                errors += 1
                preview_items.append(
                    ExpenseForecastImportPreviewItem(
                        row_number=row_no,
                        budget_subject=subject_name,
                        month=month,
                        value=value,
                        action="error",
                        message=_text(row_error),
                    )
                )
                continue
            if month <= actual_cutoff_month:
                skipped += 1
                preview_items.append(
                    ExpenseForecastImportPreviewItem(
                        row_number=row_no,
                        budget_subject=subject_name,
                        month=month,
                        value=value,
                        action="skipped",
                        message="该月份已有实际数，不能导入预估",
                    )
                )
                continue
            matched_rows = by_name.get(subject_name, [])
            if not matched_rows:
                errors += 1
                preview_items.append(
                    ExpenseForecastImportPreviewItem(
                        row_number=row_no,
                        budget_subject=subject_name,
                        month=month,
                        value=value,
                        action="error",
                        message="预算科目不存在",
                    )
                )
                continue
            if len(matched_rows) > 1:
                errors += 1
                preview_items.append(
                    ExpenseForecastImportPreviewItem(
                        row_number=row_no,
                        budget_subject=subject_name,
                        month=month,
                        value=value,
                        action="error",
                        message="预算科目名称不唯一，请改为页面手工录入",
                    )
                )
                continue
            matched = matched_rows[0]
            if not bool(matched["is_leaf"]) or matched["formula_text"]:
                errors += 1
                preview_items.append(
                    ExpenseForecastImportPreviewItem(
                        row_number=row_no,
                        budget_subject=subject_name,
                        month=month,
                        value=value,
                        action="error",
                        message="当前预算科目不可录入预估",
                    )
                )
                continue
            normalized_manage_departments = effective_manage_by_name.get(subject_name, [])
            if normalized_manage_departments and scope_value not in normalized_manage_departments:
                errors += 1
                preview_items.append(
                    ExpenseForecastImportPreviewItem(
                        row_number=row_no,
                        budget_subject=subject_name,
                        month=month,
                        value=value,
                        action="error",
                        message=f"该预算科目仅归口管理部门“{normalized_manage_departments[0]}”可录入",
                    )
                )
                continue

            existing = forecast_map.get((scope_value, int(matched["id"]), month))
            action = "inserted"
            message = None
            if existing is not None:
                if import_mode == "append":
                    action = "skipped"
                    message = "追加模式下保留已有预估值"
                    skipped += 1
                else:
                    action = "updated"
                    updated += 1
            else:
                inserted += 1

            preview_items.append(
                ExpenseForecastImportPreviewItem(
                    row_number=row_no,
                    budget_subject=subject_name,
                    month=month,
                    value=value,
                    action=action,
                    message=message,
                )
            )
            normalized_rows.append(
                {
                    "subject_id": int(matched["id"]),
                    "budget_subject": subject_name,
                    "month": month,
                    "value": value,
                    "action": action,
                }
            )

        response = ExpenseForecastImportPreviewResponse(
            file_name=file_name,
            import_mode=import_mode,
            actual_cutoff_month=actual_cutoff_month,
            preview_count=min(len(preview_items), 200),
            insertable_cells=inserted,
            updatable_cells=updated,
            skipped_cells=skipped,
            error_cells=errors,
            items=preview_items[:200],
        )
        return response, normalized_rows

    @router.get("/api/expense-forecast/meta", response_model=ExpenseForecastMetaResponse)
    async def get_expense_forecast_meta(year: int = Query(default_year)):
        await _ensure_tables()
        entity_options, group_options, owner_options, owner_group_options = await _load_scope_options()
        versions = await _forecast_versions(year)
        return ExpenseForecastMetaResponse(
            default_year=year,
            default_version=versions[0] if versions else _default_version(),
            version_suggestions=versions,
            entity_options=entity_options,
            group_options=group_options,
            owner_options=owner_options,
            owner_group_options=owner_group_options,
        )

    @router.get("/api/expense-forecast/view", response_model=ExpenseForecastViewResponse)
    async def get_expense_forecast_view(
        year: int = Query(default_year),
        forecast_version: str = Query(...),
        scope_type: str = Query(...),
        scope_value: str = Query(...),
    ):
        return await _build_view(
            year=year,
            forecast_version=_text(forecast_version) or _default_version(),
            scope_type=_normalize_scope_type(scope_type),
            scope_value=_text(scope_value),
        )

    @router.post("/api/expense-forecast/cell", response_model=ExpenseForecastCellUpsertResponse)
    async def upsert_expense_forecast_cell(body: ExpenseForecastCellUpsertRequest):
        await _ensure_tables()
        if body.scope_type != "owner":
            raise HTTPException(status_code=400, detail="费用预估仅支持在费用归属部门口径下录入")
        by_id, _ = await _subject_lookup()
        manage_department_map = await _load_manage_department_map()
        effective_manage_by_id, _effective_manage_by_name = _effective_manage_departments(
            list(by_id.values()), manage_department_map
        )
        subject = by_id.get(int(body.subject_id))
        if not subject:
            raise HTTPException(status_code=404, detail="预算科目不存在")
        if not bool(subject["is_leaf"]) or subject["formula_text"]:
            raise HTTPException(status_code=400, detail="当前预算科目不可录入预估")
        normalized_manage_department = effective_manage_by_id.get(int(body.subject_id), "")
        if normalized_manage_department and normalized_manage_department != _text(body.scope_value):
            raise HTTPException(
                status_code=400,
                detail=f"该预算科目仅归口管理部门“{normalized_manage_department}”可录入",
            )
        if int(body.month) < 1 or int(body.month) > 12:
            raise HTTPException(status_code=400, detail="月份必须在 1 到 12 之间")
        actual_cutoff_month = await _actual_cutoff_month(int(body.year))
        if int(body.month) <= actual_cutoff_month:
            raise HTTPException(status_code=400, detail="该月份已有实际数，不允许修改预估")

        now = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
        async with aiosqlite.connect(common_db_path()) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            await db.execute(
                """
                INSERT INTO expense_forecast_entry(
                  forecast_year, forecast_version, scope_type, scope_value, subject_id, month,
                  forecast_value, create_time, update_time
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(forecast_year, forecast_version, scope_type, scope_value, subject_id, month)
                DO UPDATE SET forecast_value = excluded.forecast_value, update_time = excluded.update_time
                """,
                (
                    int(body.year),
                    _text(body.forecast_version) or _default_version(),
                    body.scope_type,
                    _text(body.scope_value),
                    int(body.subject_id),
                    int(body.month),
                    float(body.value),
                    now,
                    now,
                ),
            )
            await db.commit()

        await write_operation_log(
            action_type="UPSERT",
            action_desc=f"写入费用预测 {subject['subject_name']} M{int(body.month)}",
            target_table="expense_forecast_entry",
            affected_rows=1,
            after_data=body.model_dump(),
        )
        return ExpenseForecastCellUpsertResponse(updated=True, actual_cutoff_month=actual_cutoff_month)

    @router.post("/api/expense-forecast/import-preview", response_model=ExpenseForecastImportPreviewResponse)
    async def preview_expense_forecast_import(
        year: int = Query(default_year),
        forecast_version: str = Query(...),
        scope_type: str = Query(...),
        scope_value: str = Query(...),
        mode: str = Query("append"),
        file: UploadFile = File(...),
    ):
        await _ensure_tables()
        raw = await file.read()
        response, _rows = await _preview_import(
            file_name=file.filename or "费用预测导入.xlsx",
            raw=raw,
            year=year,
            forecast_version=_text(forecast_version) or _default_version(),
            scope_type=_normalize_scope_type(scope_type),
            scope_value=_text(scope_value),
            import_mode=_normalize_import_mode(mode),
        )
        return response

    @router.post("/api/expense-forecast/import-apply", response_model=ExpenseForecastImportApplyResponse)
    async def apply_expense_forecast_import(
        year: int = Query(default_year),
        forecast_version: str = Query(...),
        scope_type: str = Query(...),
        scope_value: str = Query(...),
        mode: str = Query("append"),
        file: UploadFile = File(...),
    ):
        await _ensure_tables()
        normalized_scope_type = _normalize_scope_type(scope_type)
        normalized_mode = _normalize_import_mode(mode)
        normalized_scope_value = _text(scope_value)
        normalized_version = _text(forecast_version) or _default_version()
        raw = await file.read()
        preview, rows = await _preview_import(
            file_name=file.filename or "费用预测导入.xlsx",
            raw=raw,
            year=year,
            forecast_version=normalized_version,
            scope_type=normalized_scope_type,
            scope_value=normalized_scope_value,
            import_mode=normalized_mode,
        )
        now = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
        inserted = 0
        updated = 0
        skipped = 0
        errors = preview.error_cells

        async with aiosqlite.connect(common_db_path()) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            for row in rows:
                action = _text(row["action"])
                if action == "skipped":
                    skipped += 1
                    continue
                if action not in {"inserted", "updated"}:
                    continue
                await db.execute(
                    """
                    INSERT INTO expense_forecast_entry(
                      forecast_year, forecast_version, scope_type, scope_value, subject_id, month,
                      forecast_value, create_time, update_time
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(forecast_year, forecast_version, scope_type, scope_value, subject_id, month)
                    DO UPDATE SET forecast_value = excluded.forecast_value, update_time = excluded.update_time
                    """,
                    (
                        int(year),
                        normalized_version,
                        normalized_scope_type,
                        normalized_scope_value,
                        int(row["subject_id"]),
                        int(row["month"]),
                        float(row["value"]),
                        now,
                        now,
                    ),
                )
                if action == "inserted":
                    inserted += 1
                else:
                    updated += 1
            await db.commit()

        await write_operation_log(
            action_type="IMPORT",
            action_desc=f"导入费用预测 {inserted + updated} 个单元格（{normalized_mode}）",
            target_table="expense_forecast_entry",
            affected_rows=inserted + updated,
            after_data={
                "year": year,
                "forecast_version": normalized_version,
                "scope_type": normalized_scope_type,
                "scope_value": normalized_scope_value,
                "mode": normalized_mode,
                "inserted_cells": inserted,
                "updated_cells": updated,
                "skipped_cells": skipped,
                "error_cells": errors,
            },
        )
        return ExpenseForecastImportApplyResponse(
            file_name=file.filename or "费用预测导入.xlsx",
            import_mode=normalized_mode,
            actual_cutoff_month=preview.actual_cutoff_month,
            inserted_cells=inserted,
            updated_cells=updated,
            skipped_cells=skipped,
            error_cells=errors,
        )

    @router.post("/api/expense-forecast/export")
    async def export_expense_forecast(body: ExpenseForecastExportRequest):
        view = await _build_view(
            year=int(body.year),
            forecast_version=_text(body.forecast_version) or _default_version(),
            scope_type=body.scope_type,
            scope_value=_text(body.scope_value),
        )
        wb = Workbook()
        ws = wb.active
        ws.title = "费用预测表"
        ws.append(["费用预测表"])
        scope_label = {"entity": "主体", "group": "事业群", "owner": "费用归属部门"}[body.scope_type]
        ws.append(
            [
                f"年份：{body.year}",
                f"版本：{body.forecast_version}",
                f"编制口径：{scope_label}",
                f"口径值：{body.scope_value}",
                f"实际截至月：{view.actual_cutoff_month}月",
            ]
        )
        header = ["预算科目"] + [f"{month}月" for month in range(1, 13)] + ["全年预测"]
        ws.append(header)
        for row in view.rows:
            indent = "  " * max(0, int(row.level_number) - 1)
            ws.append(
                [f"{indent}{row.subject_name}"]
                + [round(cell.value, 2) for cell in row.months]
                + [round(row.total_value, 2)]
            )
        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        display_file_name = f"费用预测表_{body.year}_{body.forecast_version}.xlsx"
        encoded_file_name = quote(display_file_name)
        return StreamingResponse(
            stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=expense-forecast.xlsx; filename*=UTF-8''{encoded_file_name}"
            },
        )

    return router
