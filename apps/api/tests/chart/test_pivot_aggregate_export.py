from __future__ import annotations

from io import BytesIO
import unittest

from openpyxl import Workbook, load_workbook

from app.schemas import BudgetSummaryExportPivotRequest, BudgetSummaryRowDto
from app.services.pivot_aggregate_export import aggregate_workbook_response, build_pivot_aggregate_workbook


async def _read_response_body(response) -> bytes:
    chunks: list[bytes] = []
    async for chunk in response.body_iterator:
        chunks.append(chunk if isinstance(chunk, bytes) else chunk.encode())
    return b"".join(chunks)


class PivotAggregateExportTests(unittest.IsolatedAsyncioTestCase):
    async def test_aggregate_workbook_response_uses_common_excel_download_contract(self) -> None:
        wb = Workbook()
        wb.active["A1"] = "透视聚合"

        response = aggregate_workbook_response(wb, "预算透视聚合.xlsx")

        self.assertEqual(
            response.media_type,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        disposition = response.headers["content-disposition"]
        self.assertIn("filename=workbook.xlsx", disposition)
        self.assertIn("filename*=UTF-8''%E9%A2%84%E7%AE%97", disposition)

        body = await _read_response_body(response)
        loaded = load_workbook(BytesIO(body))
        self.assertEqual(loaded.active["A1"].value, "透视聚合")

    def test_build_workbook_adds_org_product_refs_when_data_account_is_row_field(self) -> None:
        wb = build_pivot_aggregate_workbook(
            rows=[
                BudgetSummaryRowDto(
                    data_code_name="A01.03.01.01.01.074 贷款利息收入",
                    product_code_name="A01 泛微粒贷",
                    year="2026",
                    month="M01",
                    quarter="Q1",
                    budget_actual=0,
                    version_id=1,
                    value=123.45,
                    value_type="金额",
                )
            ],
            body=BudgetSummaryExportPivotRequest(
                row_field_ids=["data_code_name"],
                column_field_ids=[],
                show_row_total=False,
                show_column_total=False,
            ),
            title="测试透视",
            source_label="测试来源",
            org_product_refs_by_runtime_ref_code={
                "A01.03.01.01.01.074": [
                    "A01:业务状况表:A0114 贷款利息收入",
                    "A01:业务状况表:A01140103 利息收入",
                ]
            },
        )

        ws = wb.active

        self.assertEqual(ws.cell(row=8, column=3).value, "机构产品引用数量")
        self.assertEqual(ws.cell(row=8, column=4).value, "机构产品来源")
        self.assertEqual(ws.cell(row=9, column=3).value, 2)
        self.assertIn("A01:业务状况表:A0114", ws.cell(row=9, column=4).value)

    def test_build_workbook_keeps_legacy_headers_without_org_product_ref_index(self) -> None:
        wb = build_pivot_aggregate_workbook(
            rows=[
                BudgetSummaryRowDto(
                    data_code_name="A01.03.01.01.01.074 贷款利息收入",
                    product_code_name="A01 泛微粒贷",
                    year="2026",
                    month="M01",
                    quarter="Q1",
                    budget_actual=0,
                    version_id=1,
                    value=123.45,
                    value_type="金额",
                )
            ],
            body=BudgetSummaryExportPivotRequest(
                row_field_ids=["data_code_name"],
                column_field_ids=[],
                show_row_total=False,
                show_column_total=False,
            ),
            title="测试透视",
            source_label="测试来源",
        )

        ws = wb.active

        self.assertEqual(ws.cell(row=8, column=1).value, "机构及产品指标编码")
        self.assertEqual(ws.cell(row=8, column=2).value, "预算数值")
        self.assertIsNone(ws.cell(row=8, column=3).value)
        self.assertIsNone(ws.cell(row=8, column=4).value)


if __name__ == "__main__":
    unittest.main()
