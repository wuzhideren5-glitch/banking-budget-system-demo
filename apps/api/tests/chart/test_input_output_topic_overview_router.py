from __future__ import annotations

import unittest
from pathlib import Path
from types import SimpleNamespace

from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.routers import input_output_topic_overview as input_output_topic_overview_module
from app.routers.business_cost_income_ratio import build_business_cost_income_ratio_router
from app.services import input_output_topic_overview as input_output_topic_overview_service
from app.services.input_output_topic_overview import (
    _org_product_refs_for_data_acct_code,
    build_input_output_topic_workbook,
)


async def noop_operation_log(**_kwargs) -> None:
    return None


def route_paths(router) -> set[str]:
    return {getattr(route, "path", "") for route in router.routes}


class InputOutputTopicOverviewRouterTests(unittest.TestCase):
    def test_topic_overview_routes_have_their_own_router(self) -> None:
        router = input_output_topic_overview_module.build_input_output_topic_overview_router()

        self.assertEqual(
            route_paths(router),
            {
                "/api/input-output-topic-overview/meta",
                "/api/input-output-topic-overview/report",
                "/api/input-output-topic-overview/export",
            },
        )

    def test_business_cost_income_router_no_longer_registers_topic_overview_paths(self) -> None:
        router = build_business_cost_income_ratio_router(write_operation_log=noop_operation_log)

        self.assertFalse(
            any(path.startswith("/api/input-output-topic-overview") for path in route_paths(router))
        )

    def test_export_uses_common_excel_download_contract(self) -> None:
        previous_report = input_output_topic_overview_module.build_input_output_topic_report
        previous_workbook = input_output_topic_overview_module.build_input_output_topic_workbook

        async def fake_report(**kwargs):
            self.assertEqual(kwargs["entity_name"], "微众银行")
            self.assertEqual(kwargs["report_month"], "2026-05")
            self.assertEqual(kwargs["amount_unit"], "ten_thousand")
            return SimpleNamespace(rows=[])

        def fake_workbook(report, *, view_mode):
            self.assertEqual(report.rows, [])
            self.assertEqual(view_mode, "detail")
            wb = Workbook()
            wb.active.title = "投入产出专题"
            wb.active["A1"] = "export"
            return wb

        try:
            input_output_topic_overview_module.build_input_output_topic_report = fake_report
            input_output_topic_overview_module.build_input_output_topic_workbook = fake_workbook
            app = FastAPI()
            app.include_router(input_output_topic_overview_module.build_input_output_topic_overview_router())

            response = TestClient(app).get(
                "/api/input-output-topic-overview/export",
                params={
                    "entity_name": "微众银行",
                    "report_month": "2026-05",
                    "amount_unit": "ten_thousand",
                    "view_mode": "detail",
                },
            )
        finally:
            input_output_topic_overview_module.build_input_output_topic_report = previous_report
            input_output_topic_overview_module.build_input_output_topic_workbook = previous_workbook

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.headers["content-type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        disposition = response.headers["content-disposition"]
        self.assertIn("filename=input-output-topic-overview.xlsx", disposition)
        self.assertIn("filename*=UTF-8''%E6%8A%95%E5%85%A5%E4%BA%A7%E5%87%BA", disposition)

    def test_router_does_not_hand_roll_export_download_response(self) -> None:
        router_source = Path(input_output_topic_overview_module.__file__).read_text(encoding="utf-8")

        self.assertNotIn("StreamingResponse", router_source)
        self.assertNotIn("BytesIO", router_source)
        self.assertNotIn("quote(", router_source)
        self.assertNotIn("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", router_source)
        self.assertNotIn("Content-Disposition", router_source)
        self.assertIn("workbook_streaming_response", router_source)

    def test_workbook_exports_org_product_topic_metric_code(self) -> None:
        wb = build_input_output_topic_workbook(
            {
                "report_month": "2026-05",
                "group_name": None,
                "selected_product_codes": ["A01"],
                "amount_unit_label": "万元",
                "total_rows": [
                    {
                        "section": "indicator",
                        "id": 1,
                        "name": "投入产出比",
                        "parent_id": None,
                        "topic_metric_node_code": "A01:业务状况表:A0101",
                        "sort_order": 1,
                        "metrics": {},
                    }
                ],
                "product_blocks": [],
            },
            view_mode="total",
        )

        ws = wb.active

        self.assertIn("机构产品主题指标", [cell.value for cell in ws[4]])
        self.assertEqual(ws["G5"].value, "A01:业务状况表:A0101")

    def test_workbook_exports_org_product_sources_for_data_account_rows(self) -> None:
        wb = build_input_output_topic_workbook(
            {
                "report_month": "2026-05",
                "group_name": None,
                "selected_product_codes": ["A01"],
                "amount_unit_label": "万元",
                "total_rows": [
                    {
                        "section": "input",
                        "id": 1,
                        "name": "渠道成本",
                        "parent_id": None,
                        "data_acct_code": "A01.02.04.02.01.020",
                        "org_product_refs": [
                            "A01:业务状况表:A0114010702 渠道费率",
                            "A01:业务状况表:A01220601 渠道费率",
                        ],
                        "sort_order": 1,
                        "metrics": {},
                    }
                ],
                "product_blocks": [],
            },
            view_mode="total",
        )

        ws = wb.active
        headers = [cell.value for cell in ws[4]]

        metric_col = headers.index("机构及产品指标编码") + 1
        self.assertEqual(ws.cell(row=5, column=metric_col).value, "A0114010702")
        self.assertIn("机构产品来源", headers)
        source_col = headers.index("机构产品来源") + 1
        self.assertEqual(
            ws.cell(row=5, column=source_col).value,
            "A01:业务状况表:A0114010702 渠道费率\nA01:业务状况表:A01220601 渠道费率",
        )

    def test_normalizes_org_product_source_lookup_for_topic_rows(self) -> None:
        refs = {"A01.02.04.02.01.020": ["A01:业务状况表:A0114010702 渠道费率"]}

        self.assertEqual(
            _org_product_refs_for_data_acct_code(" a01.02.04.02.01.020 ", refs),
            ["A01:业务状况表:A0114010702 渠道费率"],
        )
        self.assertEqual(_org_product_refs_for_data_acct_code("", refs), [])

    def test_topic_overview_service_uses_mysql_gateway_path(self) -> None:
        source = Path(input_output_topic_overview_service.__file__).read_text(encoding="utf-8")

        self.assertNotIn("aiosqlite_compat", source)
        self.assertNotIn("import aiosqlite", source)
        self.assertIn("get_pool().fetch_all", source)
        self.assertIn("WHERE budget_year = ?", source)


if __name__ == "__main__":
    unittest.main()
