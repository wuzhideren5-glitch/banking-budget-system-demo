from __future__ import annotations

import asyncio
from io import BytesIO
import json
import sqlite3
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

import app.services.budget_summary_export_service as budget_summary_export_service_module
from app.schemas import BudgetSummaryExportPivotRequest
from app.services.budget_summary_export_service import (
    FORMULA_TREE_ORG_PRODUCT_REF_COUNT_COL,
    FORMULA_TREE_ORG_PRODUCT_REF_SOURCE_COL,
    BudgetSummaryExportService,
    month_idx_from_label,
    org_product_refs_for_formula_tree_data_code,
)


class BudgetSummaryExportServiceTests(unittest.TestCase):
    def test_parses_month_index_from_display_label(self) -> None:
        self.assertEqual(month_idx_from_label("M01"), 0)
        self.assertEqual(month_idx_from_label("实际M12"), 11)
        self.assertIsNone(month_idx_from_label("M13"))
        self.assertIsNone(month_idx_from_label("全年"))

    def test_normalizes_formula_tree_org_product_refs_lookup(self) -> None:
        refs = {"A01.01.01.001": ["A01:业务状况表:A0121 手续费收入"]}

        self.assertEqual(
            org_product_refs_for_formula_tree_data_code(" a01.01.01.001 ", refs),
            ["A01:业务状况表:A0121 手续费收入"],
        )
        self.assertEqual(org_product_refs_for_formula_tree_data_code("", refs), [])

    def test_formula_tree_workbook_appends_org_product_trace_columns(self) -> None:
        async def run_case() -> tuple[list[str], int, str]:
            with tempfile.TemporaryDirectory() as tmp:
                common_path = Path(tmp) / "common.db"
                budget_path = Path(tmp) / "budget_2026.db"
                self._seed_common_db(common_path)
                self._seed_budget_db(budget_path)
                previous_common_db_path = budget_summary_export_service_module.common_db_path
                budget_summary_export_service_module.common_db_path = lambda: common_path
                try:
                    service = BudgetSummaryExportService(
                        editable_context_provider=lambda: asyncio.sleep(
                            0,
                            result=(budget_path, 2026, 1),
                        )
                    )
                    response = await service.export_budget_summary_formula_tree_workbook(
                        BudgetSummaryExportPivotRequest(),
                        version_id=1,
                        budget_path=budget_path,
                        budget_year=2026,
                    )
                    payload = b""
                    async for chunk in response.body_iterator:
                        payload += chunk
                finally:
                    budget_summary_export_service_module.common_db_path = previous_common_db_path

                wb = load_workbook(BytesIO(payload), data_only=False)
                ws = wb["导出带公式树状表"]
                headers = [
                    ws.cell(row=2, column=FORMULA_TREE_ORG_PRODUCT_REF_COUNT_COL).value,
                    ws.cell(row=2, column=FORMULA_TREE_ORG_PRODUCT_REF_SOURCE_COL).value,
                ]
                data_row = 4
                return (
                    headers,
                    ws.cell(row=data_row, column=FORMULA_TREE_ORG_PRODUCT_REF_COUNT_COL).value,
                    ws.cell(row=data_row, column=FORMULA_TREE_ORG_PRODUCT_REF_SOURCE_COL).value,
                )

        headers, ref_count, source = asyncio.run(run_case())

        self.assertEqual(headers, ["机构产品引用数量", "机构产品来源"])
        self.assertEqual(ref_count, 1)
        self.assertEqual(source, "A01:业务状况表:A0121 手续费收入")

    def _seed_common_db(self, db_path: Path) -> None:
        payload = {
            "metrics": [
                {
                    "code": "A0121",
                    "name": "手续费收入",
                    "mapping_status": "MANUAL_CONFIRMED",
                    "data_acct_code": "A01.01.01.001",
                    "metric_node_code": "A01.01",
                },
                {
                    "code": "0503",
                    "name": "05未确认行",
                    "mapping_status": "ORG_PRODUCT_ONLY_OR_CREATE_LATER",
                    "data_acct_code": "A01.01.01.001",
                    "metric_node_code": "A01.01",
                },
            ]
        }
        with sqlite3.connect(db_path) as conn:
            conn.executescript(
                """
                CREATE TABLE data_account_metric_node (
                  node_code TEXT PRIMARY KEY,
                  node_name TEXT NOT NULL,
                  parent_code TEXT,
                  level INTEGER NOT NULL,
                  node_type TEXT,
                  sort_order INTEGER DEFAULT 0,
                  is_active INTEGER DEFAULT 1
                );
                CREATE TABLE data_account (
                  data_acct_code TEXT PRIMARY KEY,
                  data_acct_name TEXT NOT NULL,
                  budget_formula TEXT,
                  actual_formula TEXT,
                  value_type TEXT
                );
                CREATE TABLE data_account_metric_binding (
                  metric_node_code TEXT NOT NULL,
                  data_acct_code TEXT NOT NULL,
                  sort_order INTEGER DEFAULT 0,
                  scope_code TEXT,
                  is_active INTEGER DEFAULT 1
                );
                CREATE TABLE org_product_tree_snapshot (
                  id INTEGER PRIMARY KEY CHECK (id = 1),
                  payload_json TEXT NOT NULL,
                  updated_at TEXT NOT NULL
                );
                CREATE TABLE org_product_metric_table (
                  entity_code TEXT,
                  table_name TEXT,
                  payload_json TEXT
                );
                INSERT INTO data_account_metric_node
                  (node_code, node_name, parent_code, level, node_type, sort_order, is_active)
                VALUES ('A01.01', '手续费收入', NULL, 1, 'metric', 1, 1);
                INSERT INTO data_account
                  (data_acct_code, data_acct_name, budget_formula, actual_formula, value_type)
                VALUES ('A01.01.01.001', '手续费收入', '', '', '金额');
                INSERT INTO data_account_metric_binding
                  (metric_node_code, data_acct_code, sort_order, scope_code, is_active)
                VALUES ('A01.01', 'A01.01.01.001', 1, 'A01', 1);
                INSERT INTO org_product_tree_snapshot(id, payload_json, updated_at)
                VALUES(1, '{"code":"AA","name":"微众银行","children":[{"code":"A","name":"个金群","children":[{"code":"A01","name":"公司贷款","children":[]}]}]}', 'now');
                """
            )
            conn.execute(
                """
                INSERT INTO org_product_metric_table(entity_code, table_name, payload_json)
                VALUES (?, ?, ?)
                """,
                ("A01", "业务状况表", json.dumps(payload, ensure_ascii=False)),
            )

    def _seed_budget_db(self, db_path: Path) -> None:
        with sqlite3.connect(db_path) as conn:
            conn.executescript(
                """
                CREATE TABLE version (
                  version_id INTEGER PRIMARY KEY,
                  version_date_time TEXT NOT NULL,
                  version_name TEXT NOT NULL,
                  current_month INTEGER NOT NULL
                );
                CREATE TABLE budget_summary (
                  data_code_name TEXT,
                  dept_level1 TEXT,
                  dept_level2 TEXT,
                  dept_level3 TEXT,
                  product_code_name TEXT,
                  month TEXT,
                  budget_actual INTEGER,
                  value REAL,
                  version_id INTEGER
                );
                INSERT INTO version(version_id, version_date_time, version_name, current_month)
                VALUES (1, '2026-06-05T00:00:00', '测试版本', 1);
                INSERT INTO budget_summary
                  (data_code_name, dept_level1, dept_level2, dept_level3, product_code_name,
                   month, budget_actual, value, version_id)
                VALUES
                  ('A01.01.01.001 手续费收入', '公司部', '', '', 'A01 公司贷款', 'M01', 0, 12.5, 1);
                """
            )

    def test_main_no_longer_supplies_budget_summary_export_formatting_helpers(self) -> None:
        main_source = (Path(__file__).parent / "app" / "main.py").read_text(encoding="utf-8")

        self.assertNotIn("def _extract_data_acct_code_from_name", main_source)
        self.assertNotIn("def _extract_product_code_from_summary_name", main_source)
        self.assertNotIn("def _formula_ref_display_value", main_source)
        self.assertNotIn("def _month_idx_from_label", main_source)
        constructor_start = main_source.index("BudgetSummaryExportService(")
        constructor_end = main_source.index(")", constructor_start)
        constructor_source = main_source[constructor_start:constructor_end]
        self.assertNotIn("extract_data_acct_code_from_name=", constructor_source)
        self.assertNotIn("extract_product_code_from_summary_name=", constructor_source)
        self.assertNotIn("formula_ref_display_value=", constructor_source)
        self.assertNotIn("month_idx_from_label=", constructor_source)
        self.assertNotIn("prepare_formula_expression=", constructor_source)
        self.assertNotIn("autosize_worksheet_columns=", constructor_source)


if __name__ == "__main__":
    unittest.main()
