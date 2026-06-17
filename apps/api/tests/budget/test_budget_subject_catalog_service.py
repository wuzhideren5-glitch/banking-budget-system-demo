from __future__ import annotations

import asyncio
import sqlite3
import tempfile
import unittest
from pathlib import Path

from fastapi import HTTPException
from openpyxl import load_workbook

from app.schemas import BudgetSubjectCatalogCreate, BudgetSubjectCatalogUpdate
from app.services.budget_subject_catalog import (
    build_budget_subject_catalog_workbook,
    create_budget_subject_catalog,
    delete_budget_subject_catalog,
    list_budget_subject_catalog,
    update_budget_subject_catalog,
)


SCHEMA = """
CREATE TABLE budget_subject_catalog (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  parent_id INTEGER,
  level_number INTEGER NOT NULL DEFAULT 1,
  subject_name TEXT NOT NULL,
  manage_department TEXT,
  formula_text TEXT,
  sort_order INTEGER NOT NULL DEFAULT 0
);
CREATE TABLE expense_framework_subject (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  level_label TEXT,
  budget_subject TEXT,
  manage_department TEXT,
  formula_text TEXT,
  sort_order INTEGER NOT NULL DEFAULT 0
);
"""


async def _skip_schema_check(_) -> None:
    return None


def _create_db(path: Path, *, seed_catalog: bool = False, seed_framework: bool = False) -> None:
    with sqlite3.connect(path) as conn:
        conn.executescript(SCHEMA)
        if seed_catalog:
            conn.executescript(
                """
                INSERT INTO budget_subject_catalog(
                  id, parent_id, level_number, subject_name, manage_department, formula_text, sort_order
                ) VALUES
                  (10, NULL, 1, '业务费用', '归口A', '0', 20),
                  (11, 10, 2, 'IT费用', NULL, NULL, 30);
                """
            )
        if seed_framework:
            conn.executescript(
                """
                INSERT INTO expense_framework_subject(
                  level_label, budget_subject, manage_department, formula_text, sort_order
                ) VALUES
                  ('一级', '业务费用', '归口A', '0', 10),
                  ('二级', 'IT费用', '归口B', '业务费用*0.2', 20),
                  ('三级', '', '归口C', '0', 30),
                  ('一级', '人力费用', NULL, NULL, 40);
                """
            )
        conn.commit()


class BudgetSubjectCatalogServiceTests(unittest.TestCase):
    def test_bootstraps_from_current_expense_framework_subjects(self) -> None:
        async def run() -> None:
            with tempfile.TemporaryDirectory() as tmp:
                db_path = Path(tmp) / "common.db"
                _create_db(db_path, seed_framework=True)

                rows = await list_budget_subject_catalog(db_path, ensure_schema=_skip_schema_check)

            rows_by_name = {row.subject_name: row for row in rows}
            self.assertEqual(set(rows_by_name), {"业务费用", "IT费用", "人力费用"})
            self.assertEqual(rows_by_name["IT费用"].parent_id, rows_by_name["业务费用"].id)
            self.assertFalse(rows_by_name["业务费用"].is_leaf)
            self.assertTrue(rows_by_name["IT费用"].is_leaf)

        asyncio.run(run())

    def test_creates_updates_and_deletes_leaf_rows(self) -> None:
        async def run() -> None:
            with tempfile.TemporaryDirectory() as tmp:
                db_path = Path(tmp) / "common.db"
                _create_db(db_path, seed_catalog=True)

                child = await create_budget_subject_catalog(
                    db_path,
                    BudgetSubjectCatalogCreate(parent_id=11, subject_name="云资源", manage_department="归口C"),
                    ensure_schema=_skip_schema_check,
                )
                updated = await update_budget_subject_catalog(
                    db_path,
                    child.id,
                    BudgetSubjectCatalogUpdate(formula_text="IT费用*0.1"),
                    ensure_schema=_skip_schema_check,
                )
                with self.assertRaises(HTTPException) as blocked:
                    await delete_budget_subject_catalog(db_path, 11, ensure_schema=_skip_schema_check)
                deleted = await delete_budget_subject_catalog(db_path, child.id, ensure_schema=_skip_schema_check)
                rows = await list_budget_subject_catalog(db_path, ensure_schema=_skip_schema_check)

            self.assertEqual(child.level_number, 3)
            self.assertEqual(updated.formula_text, "IT费用*0.1")
            self.assertEqual(blocked.exception.status_code, 409)
            self.assertEqual(deleted.subject_name, "云资源")
            self.assertEqual([row.id for row in rows], [10, 11])

        asyncio.run(run())

    def test_builds_export_workbook_in_tree_order(self) -> None:
        async def run() -> tuple[str, str, str]:
            with tempfile.TemporaryDirectory() as tmp:
                db_path = Path(tmp) / "common.db"
                _create_db(db_path, seed_catalog=True)
                rows = await list_budget_subject_catalog(db_path, ensure_schema=_skip_schema_check)
                workbook = build_budget_subject_catalog_workbook(rows)
                out = Path(tmp) / workbook.filename
                out.write_bytes(workbook.content)
                wb = load_workbook(out)
                ws = wb.active
                return ws["A1"].value, ws["B2"].value, ws["B3"].value

        self.assertEqual(asyncio.run(run()), ("层级", "业务费用", "IT费用"))


if __name__ == "__main__":
    unittest.main()
