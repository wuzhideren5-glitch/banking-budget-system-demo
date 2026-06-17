from __future__ import annotations

from io import BytesIO

from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.routers.intelligent_budget_simulation import build_intelligent_budget_simulation_router


def _client() -> TestClient:
    app = FastAPI()
    app.include_router(build_intelligent_budget_simulation_router())
    return TestClient(app)


def test_parse_target_endpoint_requires_user_confirmation_before_solve() -> None:
    client = _client()

    response = client.post(
        "/api/intelligent-budget-simulation/parse-target",
        json={"target_text": "净利润增长10%，不良率控制在1.2%，规模不要太冒进"},
    )

    assert response.status_code == 200
    body = response.json()
    assert body["requires_confirmation"] is True
    assert body["min_net_profit_growth"] == 0.1
    assert body["max_npl_ratio"] == 0.012
    assert body["adjustable_factors"] == ["规模", "收益率", "费用", "风险"]


def test_create_task_rejects_unconfirmed_target() -> None:
    client = _client()

    response = client.post(
        "/api/intelligent-budget-simulation/tasks",
        json={"target_text": "净利润增长10%，不良率控制在1.2%", "confirmed": False},
    )

    assert response.status_code == 400
    assert "确认" in response.json()["detail"]


def test_create_and_read_completed_task_with_top_ten_solutions() -> None:
    client = _client()

    create_response = client.post(
        "/api/intelligent-budget-simulation/tasks",
        json={"target_text": "净利润增长6%，不良率控制在1.5%", "confirmed": True},
    )

    assert create_response.status_code == 200
    created = create_response.json()
    assert created["status"] == "completed"
    assert created["baseline_solution"]["display_role"] == "baseline"
    assert created["baseline_solution"]["budget_snapshot"]["loan_balance"] > 0
    assert len(created["solutions"]) == 10
    task_id = created["task_id"]

    read_response = client.get(f"/api/intelligent-budget-simulation/tasks/{task_id}")

    assert read_response.status_code == 200
    read = read_response.json()
    assert read["task_id"] == task_id
    assert read["solutions"][0]["rank"] == 1
    assert read["solutions"][0]["explanation"] == ""


def test_export_endpoint_returns_workbook_with_required_sheets() -> None:
    client = _client()
    created = client.post(
        "/api/intelligent-budget-simulation/tasks",
        json={"target_text": "净利润增长6%，不良率控制在1.5%", "confirmed": True},
    ).json()

    response = client.post(
        "/api/intelligent-budget-simulation/export",
        json={"task_id": created["task_id"]},
    )

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "intelligent_budget_simulation" in response.headers["content-disposition"]

    workbook = load_workbook(BytesIO(response.content), read_only=True)
    assert workbook.sheetnames == ["目标与约束", "步长摘要", "预算结果快照", "Top10方案", "二层因子", "产品拆解", "风险传导", "协商记录"]
    top10 = workbook["Top10方案"]
    assert top10.max_row == 11
    assert top10["A1"].value == "排名"
