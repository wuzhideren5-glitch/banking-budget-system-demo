from __future__ import annotations

import asyncio
import sqlite3
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from app.services.budget_simulation_export import build_budget_simulation_export_buffer
from app.services.budget_simulation_metrics import (
    SIMULATION_FACTOR_METRIC_HINTS,
    build_budget_simulation_baseline_rows,
    resolve_metric_runtime_ref_codes,
)
from app.services.budget_simulation_results import (
    SIM_INTEREST_INCOME,
    SIM_REVENUE,
    build_budget_simulation_result_rows,
)
from app.schemas import SimulationBaselineRequestItem, SimulationBaselineRow, SimulationInputItem, SimulationResultRow


class BudgetSimulationMetricResolutionTests(unittest.TestCase):
    def test_resolves_current_product_prefixed_binding_by_local_metric_code(self) -> None:
        codes = resolve_metric_runtime_ref_codes(
            [
                {
                    "metric_node_code": "A01.01.01.01.01.017",
                    "local_metric_code": "01.01.017",
                    "functional_group_code": "MGMT_LOAN_DAILY",
                    "node_name": "管理贷款日均",
                    "data_acct_code": "A01.01.01.01.01.017",
                    "data_acct_name": "管理贷款日均",
                }
            ],
            SIMULATION_FACTOR_METRIC_HINTS["MGMT_LOAN_DAILY"],
        )

        self.assertEqual(codes, ["A01.01.01.01.01.017"])

    def test_resolves_metric_codes_with_product_scope_when_requested(self) -> None:
        codes = resolve_metric_runtime_ref_codes(
            [
                {
                    "metric_node_code": "A01.01.01.01.01.017",
                    "local_metric_code": "01.01.017",
                    "functional_group_code": "MGMT_LOAN_DAILY",
                    "node_name": "管理贷款日均",
                    "scope_code": "A01",
                    "data_acct_code": "A01.01.01.01.01.017",
                    "data_acct_name": "管理贷款日均",
                },
                {
                    "metric_node_code": "B01.01.01.01.01.017",
                    "local_metric_code": "01.01.017",
                    "functional_group_code": "MGMT_LOAN_DAILY",
                    "node_name": "管理贷款日均",
                    "scope_code": "B01",
                    "data_acct_code": "B01.01.01.01.01.017",
                    "data_acct_name": "管理贷款日均",
                },
            ],
            SIMULATION_FACTOR_METRIC_HINTS["MGMT_LOAN_DAILY"],
            product_code="A01",
        )

        self.assertEqual(codes, ["A01.01.01.01.01.017"])

    def test_does_not_append_retired_corp_suffix_fallback_codes(self) -> None:
        codes = resolve_metric_runtime_ref_codes(
            [],
            SIMULATION_FACTOR_METRIC_HINTS["MGMT_LOAN_DAILY"],
        )

        self.assertEqual(codes, [])
        self.assertNotIn("data_codes", SIMULATION_FACTOR_METRIC_HINTS["MGMT_LOAN_DAILY"])

    def test_export_builder_keeps_workbook_shape_outside_router(self) -> None:
        buffer, filename = build_budget_simulation_export_buffer(
            params=[
                SimulationInputItem(
                    indicator_code="mgmt_loan_daily",
                    product_code="a01",
                    simulate_value=123.45,
                )
            ],
            result_rows=[
                SimulationResultRow(
                    metric_group="盈利性指标",
                    indicator_code="PROFIT_NET",
                    indicator_name="净利润",
                    value_type="金额",
                    baseline_2025=1.0,
                    baseline_2026=2.0,
                    simulation_2026=3.0,
                )
            ],
        )

        wb = load_workbook(buffer)

        self.assertRegex(filename, r"^budget_simulation_\d{14}\.xlsx$")
        self.assertEqual(wb.sheetnames, ["模拟参数", "测算结果"])
        self.assertEqual(wb["模拟参数"]["A2"].value, "MGMT_LOAN_DAILY")
        self.assertEqual(wb["模拟参数"]["B2"].value, "A01")
        self.assertEqual(wb["测算结果"]["C2"].value, "净利润")
        self.assertEqual(wb["测算结果"]["G2"].value, 3.0)

    def test_export_builder_writes_org_product_sources_for_simulation_params(self) -> None:
        buffer, _filename = build_budget_simulation_export_buffer(
            params=[
                SimulationInputItem(
                    indicator_code="mgmt_loan_daily",
                    product_code="a01",
                    simulate_value=123.45,
                )
            ],
            result_rows=[],
            baseline_rows=[
                SimulationBaselineRow(
                    indicator_code="MGMT_LOAN_DAILY",
                    indicator_name="管理贷款日均规模",
                    product_code="A01",
                    product_name="泛微粒贷",
                    value_type="金额",
                    baseline_value=42.5,
                    source_data_acct_codes=["A01.01.01.01.01.017"],
                    source_org_product_refs=["A01:业务状况表:A0111"],
                )
            ],
        )

        wb = load_workbook(buffer)
        ws = wb["模拟参数"]

        self.assertEqual(ws["D1"].value, "机构及产品指标编码")
        self.assertEqual(ws["E1"].value, "机构产品来源")
        self.assertEqual(ws["D2"].value, "A0111")
        self.assertEqual(ws["E2"].value, "A01:业务状况表:A0111")

    def test_result_builder_reads_current_metric_bindings_and_budget_facts(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            common_path = f"{tmp_dir}/common.db"
            with sqlite3.connect(common_path) as conn:
                conn.executescript(
                    """
                    CREATE TABLE org_product_tree_snapshot (
                      id INTEGER PRIMARY KEY CHECK (id = 1),
                      payload_json TEXT NOT NULL,
                      updated_at TEXT NOT NULL
                    );
                    INSERT INTO org_product_tree_snapshot(id, payload_json, updated_at)
                    VALUES(1, '{"code":"AA","name":"全行","children":[{"code":"A","name":"个金群","children":[{"code":"A01","name":"贷款产品","children":[]}]}]}', 'now');

                    CREATE TABLE data_account_metric_node (
                      node_code TEXT PRIMARY KEY,
                      node_name TEXT,
                      local_metric_code TEXT,
                      functional_group_code TEXT,
                      metric_table_name TEXT NOT NULL DEFAULT '',
                      is_active INTEGER
                    );
                    INSERT INTO data_account_metric_node(
                      node_code, node_name, local_metric_code, functional_group_code, metric_table_name, is_active
                    ) VALUES (
                      'A01.03.01.03.001', '利息收入', '03.01.03', 'INTEREST_INCOME', '', 1
                    );

                    CREATE TABLE data_account (
                      data_acct_code TEXT PRIMARY KEY,
                      data_acct_name TEXT,
                      value_type TEXT
                    );
                    INSERT INTO data_account(data_acct_code, data_acct_name, value_type)
                    VALUES ('A01.03.01.03.001', '利息收入', '金额');

                    CREATE TABLE data_account_metric_binding (
                      data_acct_code TEXT,
                      metric_node_code TEXT,
                      scope_type TEXT,
                      scope_code TEXT,
                      is_active INTEGER
                    );
                    INSERT INTO data_account_metric_binding(
                      data_acct_code, metric_node_code, scope_type, scope_code, is_active
                    ) VALUES (
                      'A01.03.01.03.001', 'A01.03.01.03.001', 'PRODUCT', 'A01', 1
                    );
                    """
                )

            for year, value in ((2025, 100.0), (2026, 150.0)):
                with sqlite3.connect(f"{tmp_dir}/budget_{year}.db") as conn:
                    conn.executescript(
                        """
                        CREATE TABLE version (version_id INTEGER PRIMARY KEY);
                        INSERT INTO version(version_id) VALUES (1);
                        CREATE TABLE budget_data (
                          data_acct_code TEXT,
                          product_code TEXT,
                          period_id INTEGER,
                          budget_actual INTEGER,
                          version_id INTEGER,
                          value REAL
                        );
                        """
                    )
                    conn.execute(
                        """
                        INSERT INTO budget_data(
                          data_acct_code, product_code, period_id, budget_actual, version_id, value
                        ) VALUES ('A01.03.01.03.001', 'A01', 1, 0, 1, ?)
                        """,
                        (value,),
                    )

            rows = asyncio.run(build_budget_simulation_result_rows(Path(common_path), []))

        rows_by_code = {row.indicator_code: row for row in rows}
        self.assertEqual(rows_by_code[SIM_INTEREST_INCOME].baseline_2025, 100.0)
        self.assertEqual(rows_by_code[SIM_INTEREST_INCOME].baseline_2026, 150.0)
        self.assertEqual(rows_by_code[SIM_INTEREST_INCOME].simulation_2026, 150.0)
        self.assertEqual(rows_by_code[SIM_REVENUE].baseline_2025, 100.0)
        self.assertEqual(rows_by_code[SIM_REVENUE].baseline_2026, 150.0)

    def test_baseline_builder_keeps_product_names_and_metric_reads_outside_router(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            common_path = Path(tmp_dir) / "common.db"
            budget_path = Path(tmp_dir) / "budget_2026.db"
            with sqlite3.connect(common_path) as conn:
                conn.executescript(
                    """
                    CREATE TABLE org_product_tree_snapshot (
                      id INTEGER PRIMARY KEY CHECK (id = 1),
                      payload_json TEXT NOT NULL,
                      updated_at TEXT NOT NULL
                    );
                    INSERT INTO org_product_tree_snapshot(id, payload_json, updated_at)
                    VALUES(1, '{"code":"AA","name":"微众银行","children":[{"code":"A","name":"个金群","children":[{"code":"A01","name":"贷款产品","children":[]}]}]}', 'now');

                    CREATE TABLE data_account_metric_node (
                      node_code TEXT PRIMARY KEY,
                      node_name TEXT,
                      local_metric_code TEXT,
                      functional_group_code TEXT,
                      metric_table_name TEXT NOT NULL DEFAULT '',
                      is_active INTEGER
                    );
                    INSERT INTO data_account_metric_node(
                      node_code, node_name, local_metric_code, functional_group_code, metric_table_name, is_active
                    ) VALUES (
                      'A01.01.01.01.017', '管理贷款日均', '01.01.017', 'MGMT_LOAN_DAILY', '', 1
                    );

                    CREATE TABLE data_account (
                      data_acct_code TEXT PRIMARY KEY,
                      data_acct_name TEXT,
                      value_type TEXT
                    );
                    INSERT INTO data_account(data_acct_code, data_acct_name, value_type)
                    VALUES ('A01.01.01.01.017', '管理贷款日均', '金额');

                    CREATE TABLE data_account_metric_binding (
                      data_acct_code TEXT,
                      metric_node_code TEXT,
                      scope_type TEXT,
                      scope_code TEXT,
                      is_active INTEGER
                    );
                    INSERT INTO data_account_metric_binding(
                      data_acct_code, metric_node_code, scope_type, scope_code, is_active
                    ) VALUES (
                      'A01.01.01.01.017', 'A01.01.01.01.017', 'PRODUCT', 'A01', 1
                    );
                    """
                )
            with sqlite3.connect(budget_path) as conn:
                conn.executescript(
                    """
                    CREATE TABLE budget_data (
                      data_acct_code TEXT,
                      product_code TEXT,
                      period_id INTEGER,
                      budget_actual INTEGER,
                      version_id INTEGER,
                      value REAL
                    );
                    INSERT INTO budget_data(
                      data_acct_code, product_code, period_id, budget_actual, version_id, value
                    ) VALUES ('A01.01.01.01.017', 'A01', 1, 0, 7, 42.5);
                    """
                )

            rows = asyncio.run(
                build_budget_simulation_baseline_rows(
                    common_path=common_path,
                    budget_path=budget_path,
                    version_id=7,
                    period_month_map={1: 1},
                    body=[
                        SimulationBaselineRequestItem(
                            indicator_code="mgmt_loan_daily",
                            product_code="a01",
                        )
                    ],
                )
            )

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].indicator_code, "MGMT_LOAN_DAILY")
        self.assertEqual(rows[0].indicator_name, "管理贷款日均规模")
        self.assertEqual(rows[0].product_code, "A01")
        self.assertEqual(rows[0].product_name, "贷款产品")
        self.assertEqual(rows[0].baseline_value, 42.5)


if __name__ == "__main__":
    unittest.main()
