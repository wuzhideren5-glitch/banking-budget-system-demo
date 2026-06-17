from __future__ import annotations

import sqlite3
import tempfile
import unittest
from pathlib import Path

import aiosqlite
from openpyxl import load_workbook

from app.services.runtime_ref_export import (
    RUNTIME_REF_EXPORT_HEADERS,
    build_runtime_ref_export_workbook,
    export_runtime_refs_workbook,
)


class RuntimeRefExportTests(unittest.IsolatedAsyncioTestCase):
    async def test_export_workbook_uses_current_product_metric_identity(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            db_path = Path(tmp) / "common.db"
            with sqlite3.connect(db_path) as conn:
                conn.executescript(
                    """
                    CREATE TABLE data_account (
                      data_acct_code TEXT PRIMARY KEY,
                      data_acct_name TEXT NOT NULL,
                      budget_formula TEXT,
                      actual_formula TEXT,
                      need_calc INTEGER NOT NULL DEFAULT 0,
                      formula_calc_mode INTEGER NOT NULL DEFAULT 0,
                      allow_manual_entry INTEGER NOT NULL DEFAULT 1,
                      value_type TEXT NOT NULL,
                      remark TEXT
                    );
                    CREATE TABLE data_account_metric_node (
                      node_code TEXT PRIMARY KEY,
                      node_name TEXT NOT NULL,
                      parent_code TEXT
                    );
                    CREATE TABLE data_account_metric_binding (
                      data_acct_code TEXT NOT NULL,
                      metric_node_code TEXT NOT NULL,
                      scope_type TEXT NOT NULL,
                      scope_code TEXT NOT NULL,
                      is_active INTEGER NOT NULL DEFAULT 1
                    );
                    CREATE TABLE org_product_tree_snapshot (
                      id INTEGER PRIMARY KEY CHECK (id = 1),
                      payload_json TEXT NOT NULL,
                      updated_at TEXT NOT NULL
                    );
                    CREATE TABLE org_product_metric_table (
                      entity_code TEXT,
                      table_name TEXT,
                      payload_json TEXT
                    );
                    INSERT INTO org_product_tree_snapshot(id, payload_json, updated_at)
                    VALUES(1, '{"code":"AA","name":"微众银行","children":[{"code":"A","name":"个金群","children":[{"code":"A01","name":"泛微粒贷","children":[]}]}]}', 'now');
                    INSERT INTO data_account_metric_node(node_code, node_name, parent_code)
                    VALUES
                      ('A01', '泛微粒贷', NULL),
                      ('A01.01.01.001', '产品利息收入', 'A01'),
                      ('CORP', '全行', NULL),
                      ('CORP.00', '全行指标', 'CORP');
                    INSERT INTO data_account(
                      data_acct_code, data_acct_name, budget_formula, actual_formula,
                      formula_calc_mode, allow_manual_entry, value_type, remark
                    )
                    VALUES
                      ('A01.01.01.001', '产品利息收入', NULL, NULL, 0, 1, '金额', '产品前缀科目'),
                      ('CORP.00', '全行指标', NULL, NULL, 0, 1, '金额', '全行口径');
                    INSERT INTO data_account_metric_binding(data_acct_code, metric_node_code, scope_type, scope_code)
                    VALUES
                      ('A01.01.01.001', 'A01.01.01.001', 'PRODUCT', 'A01'),
                      ('CORP.00', 'CORP.00', 'CORP', 'CORP');
                    INSERT INTO org_product_metric_table VALUES (
                      'A01',
                      '业务状况表',
                      '{"metrics":[{"code":"A010101001","name":"产品利息收入","mapping_status":"ORG_PRODUCT_ONLY_OR_CREATE_LATER","metric_node_code":"SHOULD_BE_IGNORED","data_acct_code":"SHOULD_BE_IGNORED"},{"code":"A010501","name":"05编码指标","mapping_status":"MANUAL_CONFIRMED","metric_node_code":"A01.05.01","data_acct_code":"A01.01.01.001"},{"code":"Z990101001","name":"孤立旧字段指标","mapping_status":"MANUAL_CONFIRMED","metric_node_code":"A01.01.01.001","data_acct_code":"A01.01.01.001"}]}'
                    );
                    """
                )

            async with aiosqlite.connect(db_path) as db:
                buffer = await build_runtime_ref_export_workbook(db)

        wb = load_workbook(buffer, data_only=True)
        self.assertEqual(wb.sheetnames[:2], ["运行说明", "机构及产品指标编码清单"])
        intro_ws = wb["运行说明"]
        self.assertEqual(intro_ws.cell(row=1, column=1).value, "项目")
        self.assertEqual(intro_ws.cell(row=2, column=1).value, "定位")
        self.assertIn("机构及产品指标体系主键", str(intro_ws.cell(row=2, column=2).value))
        self.assertEqual(intro_ws.cell(row=3, column=2).value, "机构及产品指标")
        self.assertIn("不作为独立配置导入模板", str(intro_ws.cell(row=5, column=2).value))
        ws = wb["机构及产品指标编码清单"]
        headers = [ws.cell(row=1, column=idx).value for idx in range(1, len(RUNTIME_REF_EXPORT_HEADERS) + 1)]
        self.assertEqual(headers, list(RUNTIME_REF_EXPORT_HEADERS))
        self.assertTrue(ws.cell(row=1, column=1).font.bold)

        product_row = [ws.cell(row=2, column=idx).value for idx in range(1, len(RUNTIME_REF_EXPORT_HEADERS) + 1)]
        corp_row = [ws.cell(row=3, column=idx).value for idx in range(1, len(RUNTIME_REF_EXPORT_HEADERS) + 1)]

        self.assertEqual(product_row[0], "泛微粒贷 / 产品利息收入")
        self.assertEqual(product_row[1], "A01.01.01.001")
        self.assertEqual(product_row[2], "A01")
        self.assertEqual(product_row[3], "A01.01.01.001")
        self.assertEqual(product_row[8], "泛微粒贷")
        self.assertEqual(product_row[13], 1)
        self.assertEqual(
            product_row[14],
            "A01:业务状况表:A010101001 产品利息收入",
        )

        self.assertEqual(corp_row[0], "全行 / 全行指标")
        self.assertEqual(corp_row[1], "CORP.00")
        self.assertEqual(corp_row[2], "CORP")
        self.assertEqual(corp_row[3], "CORP.00")
        self.assertEqual(corp_row[8], "全行")
        self.assertEqual(corp_row[13], 0)
        self.assertIsNone(corp_row[14])

    async def test_export_workbook_wrapper_opens_explicit_common_db(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            db_path = Path(tmp) / "common.db"
            with sqlite3.connect(db_path) as conn:
                conn.executescript(
                    """
                    CREATE TABLE data_account (
                      data_acct_code TEXT PRIMARY KEY,
                      data_acct_name TEXT NOT NULL,
                      budget_formula TEXT,
                      actual_formula TEXT,
                      need_calc INTEGER NOT NULL DEFAULT 0,
                      formula_calc_mode INTEGER NOT NULL DEFAULT 0,
                      allow_manual_entry INTEGER NOT NULL DEFAULT 1,
                      value_type TEXT NOT NULL,
                      remark TEXT
                    );
                    CREATE TABLE data_account_metric_node (
                      node_code TEXT PRIMARY KEY,
                      node_name TEXT NOT NULL,
                      parent_code TEXT
                    );
                    CREATE TABLE data_account_metric_binding (
                      data_acct_code TEXT NOT NULL,
                      metric_node_code TEXT NOT NULL,
                      scope_type TEXT NOT NULL,
                      scope_code TEXT NOT NULL,
                      is_active INTEGER NOT NULL DEFAULT 1
                    );
                    CREATE TABLE org_product_tree_snapshot (
                      id INTEGER PRIMARY KEY CHECK (id = 1),
                      payload_json TEXT NOT NULL,
                      updated_at TEXT NOT NULL
                    );
                    INSERT INTO data_account(data_acct_code, data_acct_name, value_type)
                    VALUES ('A01.01.01.001', '产品利息收入', '金额');
                    INSERT INTO data_account_metric_node(node_code, node_name, parent_code)
                    VALUES ('A01.01.01.001', '产品利息收入', NULL);
                    INSERT INTO org_product_tree_snapshot(id, payload_json, updated_at)
                    VALUES(1, '{"code":"AA","name":"微众银行","children":[{"code":"A","name":"个金群","children":[{"code":"A01","name":"泛微粒贷","children":[]}]}]}', 'now');
                    INSERT INTO data_account_metric_binding(data_acct_code, metric_node_code, scope_type, scope_code)
                    VALUES ('A01.01.01.001', 'A01.01.01.001', 'PRODUCT', 'A01');
                    """
                )

            buffer = await export_runtime_refs_workbook(db_path)

        wb = load_workbook(buffer, data_only=True)
        self.assertEqual(wb.sheetnames[:2], ["运行说明", "机构及产品指标编码清单"])
        ws = wb["机构及产品指标编码清单"]
        self.assertEqual(ws.cell(row=2, column=4).value, "A01.01.01.001")
        self.assertEqual(ws.cell(row=2, column=9).value, "泛微粒贷")


if __name__ == "__main__":
    unittest.main()
