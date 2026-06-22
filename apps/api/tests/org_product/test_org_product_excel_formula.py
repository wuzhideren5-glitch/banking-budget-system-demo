from __future__ import annotations

import unittest

from openpyxl import Workbook

from app.org_product_excel_formula import (
    SheetFormulaContext,
    convert_excel_formula_to_system,
    format_metric_code_for_display,
    index_sheet_contexts,
    normalize_sheet_lookup_key,
)


def _ctx(sheet: str, entity: str, table: str, rows: dict[int, str]) -> SheetFormulaContext:
    display = {r: format_metric_code_for_display(entity, c) for r, c in rows.items()}
    return SheetFormulaContext(
        sheet_name=sheet,
        sheet_key=normalize_sheet_lookup_key(sheet),
        entity_code=entity,
        table_name=table,
        row_display_codes=display,
    )


def test_same_sheet_addition():
    current = _ctx("A01业务状况表", "A01", "业务状况表", {2: "A0101", 3: "A010101", 10: "A010102", 13: "A010103"})
    index = index_sheet_contexts([current])
    out = convert_excel_formula_to_system("=E3+E10+E13", current=current, all_contexts=index)
    assert out == "A01.01.01 + A01.01.02 + A01.01.03", out


def test_cross_sheet_sum():
    a01 = _ctx("A01业务状况表", "A01", "业务状况表", {2: "A0101"})
    a02 = _ctx("A02业务状况表", "A02", "业务状况表", {2: "A0201"})
    aa = _ctx("AA损益表", "AA", "损益表", {2: "AA46"})
    index = index_sheet_contexts([a01, a02, aa])
    out = convert_excel_formula_to_system(
        "=A01业务状况表!E2+A02业务状况表!E2",
        current=aa,
        all_contexts=index,
    )
    assert out == "A01/业务状况表/A01.01 + A02/业务状况表/A02.01", out


def test_sum_range():
    current = _ctx("A01业务状况表", "A01", "业务状况表", {2: "A0101", 3: "A010101", 4: "A010102"})
    index = index_sheet_contexts([current])
    out = convert_excel_formula_to_system("=SUM(E3:E4)", current=current, all_contexts=index)
    assert out == "SUM(A01.01.01,A01.01.02)", out


class MetricImportExportTemplateTests(unittest.TestCase):
    def test_export_headers_include_rule_and_display_only_markers(self) -> None:
        from app.routers.org_product_helpers import metric_export_headers_v04

        headers = metric_export_headers_v04()
        self.assertIn("规则", headers)
        self.assertIn("科目层级※仅展示", headers)
        self.assertIn("逻辑码※仅展示", headers)

    def test_display_only_header_aliases(self) -> None:
        from app.routers.org_product_helpers import _canon_header_label

        self.assertEqual(_canon_header_label("科目层级※仅展示"), "科目层级")
        self.assertEqual(_canon_header_label("逻辑码（仅展示）"), "逻辑码")
        self.assertEqual(_canon_header_label("规则"), "规则")

    def test_import_ignores_display_only_logic_code_column(self) -> None:
        from app.routers.org_product_helpers import _parse_metric_worksheet_basic

        wb = Workbook()
        ws = wb.active
        ws.title = "A01业务状况表"
        ws.append(
            [
                "科目层级※仅展示",
                "科目性质",
                "科目代码",
                "科目名称",
                "实际月公式",
                "逻辑码※仅展示",
                "规则",
            ]
        )
        ws.append(["一级", "收入", "A01.01", "营业收入", "", "MANUAL", "SUM"])
        metrics, row_count, parse_error, _ = _parse_metric_worksheet_basic(
            ws,
            "test",
            entity_code="A01",
            strict=True,
        )
        self.assertIsNone(parse_error)
        self.assertEqual(row_count, 1)
        self.assertEqual(metrics[0]["logic_code"], "01")
        self.assertEqual(metrics[0]["annual_agg_rule"], "SUM")


    def test_import_template_workbook_has_v04_headers_and_guide(self) -> None:
        from openpyxl import load_workbook

        from app.routers.org_product_helpers import (
            METRIC_EXPORT_HEADER_LEVEL,
            METRIC_EXPORT_HEADER_LOGIC,
            build_org_product_metric_import_template_workbook,
        )

        wb = load_workbook(build_org_product_metric_import_template_workbook(), data_only=False)
        self.assertIn("AA业务状况表", wb.sheetnames)
        self.assertIn("填写说明", wb.sheetnames)
        ws = wb["AA业务状况表"]
        headers = [cell.value for cell in ws[1]]
        self.assertEqual(headers[0], METRIC_EXPORT_HEADER_LEVEL)
        self.assertEqual(headers[-1], "规则")
        self.assertIn(METRIC_EXPORT_HEADER_LOGIC, headers)
        self.assertEqual(ws["H2"].value, "=H3+H6+H9")

    def test_import_template_route_returns_xlsx_bytes(self) -> None:
        from fastapi import FastAPI
        from fastapi.testclient import TestClient

        from app.routers.org_product_metric_config import router

        app = FastAPI()
        app.include_router(router)
        client = TestClient(app)
        response = client.get("/api/org-product-metrics/import-template")
        self.assertEqual(response.status_code, 200)
        self.assertGreater(len(response.content), 1000)
        self.assertEqual(response.content[:2], b"PK")
        self.assertIn(
            "org_product_metric_import_template.xlsx",
            response.headers.get("content-disposition", ""),
        )


if __name__ == "__main__":
    test_same_sheet_addition()
    test_cross_sheet_sum()
    test_sum_range()
    print("ok")
