from __future__ import annotations

from copy import copy
from io import BytesIO
from pathlib import Path
import unittest

from openpyxl.styles import Font
from openpyxl import Workbook, load_workbook

from app.services.export_common import (
    binary_streaming_response,
    color_worksheet_row_font,
    excel_streaming_response,
    normalize_excel_cell,
    workbook_streaming_response,
)


async def _read_response_body(response) -> bytes:
    chunks: list[bytes] = []
    async for chunk in response.body_iterator:
        chunks.append(chunk if isinstance(chunk, bytes) else chunk.encode())
    return b"".join(chunks)


class ExportCommonTests(unittest.IsolatedAsyncioTestCase):
    async def test_workbook_streaming_response_preserves_workbook_and_utf8_filename(self) -> None:
        wb = Workbook()
        wb.active["A1"] = "预算展示"

        response = workbook_streaming_response(
            wb,
            filename="预算展示配置导入模板.xlsx",
            fallback_filename="budget-display-config.xlsx",
        )

        self.assertEqual(
            response.media_type,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        disposition = response.headers["content-disposition"]
        self.assertIn("filename=budget-display-config.xlsx", disposition)
        self.assertIn("filename*=UTF-8''%E9%A2%84%E7%AE%97", disposition)

        body = await _read_response_body(response)
        loaded = load_workbook(BytesIO(body))
        self.assertEqual(loaded.active["A1"].value, "预算展示")

    async def test_workbook_streaming_response_defaults_to_ascii_fallback_for_utf8_filename(self) -> None:
        wb = Workbook()

        response = workbook_streaming_response(wb, filename="预算透视聚合.xlsx")

        disposition = response.headers["content-disposition"]
        self.assertIn("filename=workbook.xlsx", disposition)
        self.assertIn("filename*=UTF-8''%E9%A2%84%E7%AE%97", disposition)

    async def test_excel_streaming_response_preserves_binary_body_extra_headers_and_utf8_filename(self) -> None:
        response = excel_streaming_response(
            b"fake workbook",
            filename="数据科目导入结果.xlsx",
            extra_headers={"X-Import-Total": "3"},
        )

        self.assertEqual(response.headers["x-import-total"], "3")
        disposition = response.headers["content-disposition"]
        self.assertIn("filename=workbook.xlsx", disposition)
        self.assertIn("filename*=UTF-8''%E6%95%B0%E6%8D%AE%E7%A7%91%E7%9B%AE", disposition)
        self.assertEqual(await _read_response_body(response), b"fake workbook")

    async def test_binary_streaming_response_preserves_body_media_type_and_ascii_filename(self) -> None:
        response = binary_streaming_response(
            b"fake pptx",
            media_type="application/vnd.openxmlformats-officedocument.presentationml.presentation",
            filename="chart_export_20260604.pptx",
            extra_headers={"X-Export-Source": "chart"},
        )

        self.assertEqual(
            response.media_type,
            "application/vnd.openxmlformats-officedocument.presentationml.presentation",
        )
        self.assertEqual(response.headers["x-export-source"], "chart")
        disposition = response.headers["content-disposition"]
        self.assertIn("filename=chart_export_20260604.pptx", disposition)
        self.assertIn("filename*=UTF-8''chart_export_20260604.pptx", disposition)
        self.assertEqual(await _read_response_body(response), b"fake pptx")

    async def test_normalizes_excel_cells_for_import_parsers(self) -> None:
        self.assertEqual(normalize_excel_cell(None), "")
        self.assertEqual(normalize_excel_cell("  数据科目  "), "数据科目")
        self.assertEqual(normalize_excel_cell(123), "123")

    async def test_colors_worksheet_row_font_without_losing_existing_font_shape(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="ok")
        existing_font = copy(ws.cell(row=1, column=1).font)
        ws.cell(row=1, column=1).font = Font(name=existing_font.name, bold=True, italic=True)

        color_worksheet_row_font(ws, row_idx=1, max_col=2, color="FF0000")

        self.assertTrue(ws.cell(row=1, column=1).font.bold)
        self.assertTrue(ws.cell(row=1, column=1).font.italic)
        self.assertEqual(ws.cell(row=1, column=1).font.color.rgb, "00FF0000")
        self.assertEqual(ws.cell(row=1, column=2).font.color.rgb, "00FF0000")

    async def test_main_no_longer_keeps_excel_import_helper_implementations(self) -> None:
        main_source = (Path(__file__).parent / "app" / "main.py").read_text(encoding="utf-8")

        self.assertNotIn("def _normalize_cell", main_source)
        self.assertNotIn("def _color_row", main_source)
        self.assertNotIn("from openpyxl.styles import Font", main_source)
        self.assertIn("normalize_excel_cell", main_source)
        self.assertIn("color_worksheet_row_font", main_source)


if __name__ == "__main__":
    unittest.main()
