from __future__ import annotations

import json
import sqlite3
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from app.routers.org_product_helpers import (
    DataEntryMetricRowPayload,
    DataEntryMetricValuesPayload,
    MetricNodePayload,
    _append_org_product_output_export_sheet,
    _build_data_entry_export_workbook,
    _build_metric_rows,
    _parse_metric_worksheet_basic,
    _sanitize_data_entry_payload_mapping_refs,
    _sanitize_data_entry_payload_for_response,
    _sanitize_metric_node_dicts_for_response,
    _sanitize_org_product_output_entity_for_snapshot,
    _sanitize_output_payload_for_response,
    _sanitize_metric_nodes_for_save,
)
from app.db_bootstrap.runtime_metric_tree import ensure_runtime_metric_identity_tables
from app.db_bootstrap.report_display import ensure_budget_output_display_item_schema_sync
from app.services.org_product_metric_runtime_sync import (
    assert_all_runtime_metric_refs_are_confirmed_org_product_metrics,
    merge_canonical_expense_metric_trees_into_org_product_metrics,
    normalize_org_product_metric_mapping_statuses,
    OrgProductMetricRuntimeSyncError,
    purge_unreferenced_legacy_aa05_metric_master,
    sync_org_product_metric_runtime_refs,
)
from app.services.org_product_metric_runtime_snapshot import (
    load_org_product_metric_payload_from_runtime_tree,
)
from app.services.runtime_metric_refs import (
    compact_org_product_metric_code,
    normalize_org_product_metric_code,
)
from scripts.verify_current_database_inventory import (
    org_product_metric_guard_violations,
    org_product_metric_runtime_ref_violations,
)


def _flatten_metric_payload(metrics: list[dict]) -> list[dict]:
    rows: list[dict] = []
    stack = [item for item in metrics if isinstance(item, dict)]
    while stack:
        node = stack.pop(0)
        rows.append(node)
        children = node.get("children")
        if isinstance(children, list):
            stack[0:0] = [child for child in children if isinstance(child, dict)]
    return rows


def _assert_no_legacy_runtime_identity_fields(testcase: unittest.TestCase, row: dict) -> None:
    testcase.assertNotIn("mapping_status", row)
    testcase.assertNotIn("metric_node_code", row)
    testcase.assertNotIn("data_acct_code", row)


class OrgProductMetricRuntimeRefsTests(unittest.TestCase):
    def test_shared_metric_code_normalizer_converts_compact_inputs_to_canonical_dotted_codes(self) -> None:
        self.assertEqual(
            normalize_org_product_metric_code("B01", "B0105010101001"),
            "B01.05.01.01.01.001",
        )
        self.assertEqual(
            normalize_org_product_metric_code("B01", "05010101001"),
            "B01.05.01.01.01.001",
        )
        self.assertEqual(
            normalize_org_product_metric_code("B01", "B01.05.01.01.01.001"),
            "B01.05.01.01.01.001",
        )
        self.assertEqual(
            compact_org_product_metric_code("B01.05.01.01.01.001"),
            "B0105010101001",
        )

    def test_runtime_sync_uses_derived_metric_binding_view_without_physical_binding_table(self) -> None:
        with sqlite3.connect(":memory:") as conn:
            conn.execute(
                """
                CREATE TABLE data_account (
                  data_acct_code TEXT PRIMARY KEY NOT NULL,
                  data_acct_name TEXT NOT NULL,
                  budget_formula TEXT,
                  actual_formula TEXT,
                  budget_rule_code TEXT,
                  budget_rule_config_json TEXT,
                  need_calc INTEGER NOT NULL DEFAULT 0 CHECK (need_calc IN (0, 1)),
                  formula_calc_mode INTEGER NOT NULL DEFAULT 0 CHECK (formula_calc_mode BETWEEN 0 AND 3),
                  allow_manual_entry INTEGER NOT NULL DEFAULT 1 CHECK (allow_manual_entry IN (0, 1)),
                  value_type TEXT NOT NULL,
                  remark TEXT
                )
                """
            )

            ensure_runtime_metric_identity_tables(conn)

            self.assertEqual(
                conn.execute(
                    "SELECT type FROM sqlite_master WHERE name='data_account_metric_binding'"
                ).fetchone(),
                ("view",),
            )

            result = sync_org_product_metric_runtime_refs(
                conn,
                entity_code="A01",
                table_name="业务状况表",
                metrics=[{"code": "A01050101", "name": "常规人力", "nature": "支出"}],
            )

            self.assertEqual(result.normalized_refs, 1)
            self.assertEqual(
                conn.execute(
                    """
                    SELECT data_acct_code, metric_node_code, scope_type, scope_code
                    FROM data_account_metric_binding
                    WHERE data_acct_code='A01.05.01.01'
                    """
                ).fetchone(),
                ("A01.05.01.01", "A01.05.01.01", "PRODUCT", "A01"),
            )

    def test_budget_output_display_item_no_longer_references_retired_data_account_table(self) -> None:
        with sqlite3.connect(":memory:") as conn:
            ensure_budget_output_display_item_schema_sync(conn)

            fk_targets = {
                str(row[2])
                for row in conn.execute("PRAGMA foreign_key_list(budget_output_display_item)")
            }

        self.assertNotIn("data_account", fk_targets)

    def test_runtime_identity_bootstrap_migrates_data_account_table_to_view(self) -> None:
        with sqlite3.connect(":memory:") as conn:
            conn.executescript(
                """
                CREATE TABLE data_account (
                  data_acct_code TEXT PRIMARY KEY NOT NULL,
                  data_acct_name TEXT NOT NULL,
                  budget_formula TEXT,
                  actual_formula TEXT,
                  budget_rule_code TEXT,
                  budget_rule_config_json TEXT,
                  need_calc INTEGER NOT NULL DEFAULT 0 CHECK (need_calc IN (0, 1)),
                  formula_calc_mode INTEGER NOT NULL DEFAULT 0 CHECK (formula_calc_mode BETWEEN 0 AND 3),
                  allow_manual_entry INTEGER NOT NULL DEFAULT 1 CHECK (allow_manual_entry IN (0, 1)),
                  value_type TEXT NOT NULL,
                  remark TEXT
                );
                INSERT INTO data_account(
                  data_acct_code, data_acct_name, budget_formula, actual_formula,
                  need_calc, formula_calc_mode, allow_manual_entry, value_type, remark
                )
                VALUES(
                  'A01.05.01.01', '常规人力', 'BUDGET()', 'ACTUAL()',
                  1, 3, 0, '金额', 'legacy account remark'
                );
                CREATE TABLE data_account_metric_node (
                  node_code TEXT PRIMARY KEY NOT NULL,
                  node_name TEXT NOT NULL,
                  parent_code TEXT REFERENCES data_account_metric_node(node_code),
                  product_code TEXT,
                  local_metric_code TEXT,
                  logic_code TEXT,
                  functional_group_code TEXT,
                  metric_table_name TEXT NOT NULL DEFAULT '',
                  level INTEGER NOT NULL CHECK (level BETWEEN 1 AND 8),
                  node_type TEXT NOT NULL CHECK (node_type IN ('CATEGORY', 'GROUP', 'METRIC')),
                  horizontal_rollup INTEGER NOT NULL DEFAULT 0 CHECK (horizontal_rollup IN (0, 1)),
                  vertical_rollup INTEGER NOT NULL DEFAULT 0 CHECK (vertical_rollup IN (0, 1)),
                  sort_order INTEGER NOT NULL DEFAULT 0,
                  is_active INTEGER NOT NULL DEFAULT 1 CHECK (is_active IN (0, 1)),
                  remark TEXT,
                  created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                  updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
                );
                INSERT INTO data_account_metric_node(
                  node_code, node_name, parent_code, product_code, local_metric_code, logic_code,
                  functional_group_code, metric_table_name, level, node_type, horizontal_rollup, vertical_rollup,
                  sort_order, is_active
                ) VALUES(
                  'A01.05.01.01', '待迁移', NULL, 'A01', '05.01.01', '05.01.01',
                  '业务状况表', '业务状况表', 4, 'METRIC', 0, 0, 1, 1
                );
                """
            )

            ensure_runtime_metric_identity_tables(conn)

            self.assertEqual(
                conn.execute("SELECT type FROM sqlite_master WHERE name='data_account'").fetchone(),
                ("view",),
            )
            self.assertEqual(
                conn.execute(
                    """
                    SELECT data_acct_name, budget_formula, actual_formula,
                           need_calc, formula_calc_mode, allow_manual_entry, value_type, remark
                    FROM data_account
                    WHERE data_acct_code='A01.05.01.01'
                    """
                ).fetchone(),
                ("常规人力", "BUDGET()", "ACTUAL()", 1, 3, 0, "金额", "legacy account remark"),
            )

    def _create_runtime_metric_tables(self, conn: sqlite3.Connection) -> None:
        conn.executescript(
            """
            CREATE TABLE data_account (
              data_acct_code TEXT PRIMARY KEY NOT NULL,
              data_acct_name TEXT NOT NULL,
              budget_formula TEXT,
              actual_formula TEXT,
              budget_rule_code TEXT,
              budget_rule_config_json TEXT,
              need_calc INTEGER NOT NULL DEFAULT 0 CHECK (need_calc IN (0, 1)),
              formula_calc_mode INTEGER NOT NULL DEFAULT 0 CHECK (formula_calc_mode BETWEEN 0 AND 3),
              allow_manual_entry INTEGER NOT NULL DEFAULT 1 CHECK (allow_manual_entry IN (0, 1)),
              value_type TEXT NOT NULL,
              remark TEXT
            );
            """
        )
        ensure_runtime_metric_identity_tables(conn)

    def test_runtime_sync_materializes_org_product_metric_refs(self) -> None:
        with sqlite3.connect(":memory:") as conn:
            self._create_runtime_metric_tables(conn)
            metrics = [
                {
                    "code": "A010101",
                    "name": "利息收入",
                    "nature": "收入",
                    "metric_node_code": "A01.01.01",
                    "data_acct_code": "",
                    "allow_manual_entry": 0,
                    "children": [],
                }
            ]

            result = sync_org_product_metric_runtime_refs(
                conn,
                entity_code="A01",
                table_name="业务状况表",
                metrics=metrics,
            )

            self.assertEqual(result.normalized_refs, 1)
            self.assertNotIn("metric_node_code", metrics[0])
            self.assertNotIn("data_acct_code", metrics[0])
            self.assertNotIn("mapping_status", metrics[0])
            self.assertEqual(
                conn.execute(
                    "SELECT data_acct_name, allow_manual_entry FROM data_account WHERE data_acct_code='A01.01.01'"
                ).fetchone(),
                ("利息收入", 0),
            )
            self.assertEqual(
                conn.execute(
                    "SELECT metric_node_code, scope_type, scope_code FROM data_account_metric_binding WHERE data_acct_code='A01.01.01'"
                ).fetchone(),
                ("A01.01.01", "PRODUCT", "A01"),
            )
            parent_codes = {
                row[0]
                for row in conn.execute(
                    "SELECT node_code FROM data_account_metric_node ORDER BY node_code"
                )
            }
            self.assertTrue({"A01", "A01.01", "A01.01.01"}.issubset(parent_codes))

    def test_runtime_sync_persists_annual_agg_rule_and_formulas(self) -> None:
        with sqlite3.connect(":memory:") as conn:
            self._create_runtime_metric_tables(conn)
            metrics = [
                {
                    "code": "AA.01.001",
                    "name": "分子",
                    "nature": "收入",
                    "annual_agg_rule": "SUM",
                    "formula_budget_annual": "",
                    "children": [],
                },
                {
                    "code": "AA.01.002",
                    "name": "分母",
                    "nature": "收入",
                    "annual_agg_rule": "SUM",
                    "children": [],
                },
                {
                    "code": "AA.01.003",
                    "name": "不良率",
                    "nature": "比例",
                    "annual_agg_rule": "CALC",
                    "formula_budget_annual": "AA.01.001/AA.01.002",
                    "children": [],
                },
            ]

            sync_org_product_metric_runtime_refs(
                conn,
                entity_code="AA",
                table_name="业务状况表",
                metrics=metrics,
                overwrite_existing_metadata=True,
            )

            self.assertEqual(
                conn.execute(
                    """
                    SELECT annual_agg_rule, budget_formula
                    FROM data_account_metric_node
                    WHERE node_code='AA.01.001'
                    """
                ).fetchone(),
                ("SUM", ""),
            )
            self.assertEqual(
                conn.execute(
                    """
                    SELECT annual_agg_rule, budget_formula
                    FROM data_account_metric_node
                    WHERE node_code='AA.01.003'
                    """
                ).fetchone(),
                ("CALC", "AA.01.001/AA.01.002"),
            )

    def test_runtime_sync_derives_confirmed_leaf_ref_from_org_product_metric_code(self) -> None:
        with sqlite3.connect(":memory:") as conn:
            self._create_runtime_metric_tables(conn)
            metrics = [
                {
                    "code": "A01050101",
                    "name": "常规人力",
                    "nature": "支出",
                    "metric_node_code": "",
                    "data_acct_code": "",
                    "children": [],
                }
            ]

            result = sync_org_product_metric_runtime_refs(
                conn,
                entity_code="A01",
                table_name="业务状况表",
                metrics=metrics,
            )

            self.assertEqual(result.normalized_refs, 1)
            self.assertNotIn("metric_node_code", metrics[0])
            self.assertNotIn("data_acct_code", metrics[0])
            self.assertNotIn("mapping_status", metrics[0])
            self.assertEqual(
                conn.execute(
                    "SELECT data_acct_name FROM data_account WHERE data_acct_code='A01.05.01.01'"
                ).fetchone(),
                ("常规人力",),
            )
            self.assertEqual(
                conn.execute(
                    "SELECT metric_node_code, scope_type, scope_code FROM data_account_metric_binding WHERE data_acct_code='A01.05.01.01'"
                ).fetchone(),
                ("A01.05.01.01", "PRODUCT", "A01"),
            )

    def test_runtime_sync_preserves_explicit_metric_tree_parent_relationships(self) -> None:
        with sqlite3.connect(":memory:") as conn:
            self._create_runtime_metric_tables(conn)
            metrics = [
                {
                    "code": "AA.05.01.02.03",
                    "name": "IT",
                    "children": [
                        {
                            "code": "AA.05.01.02.03.03.001",
                            "name": "IT职场",
                            "children": [],
                        }
                    ],
                }
            ]

            sync_org_product_metric_runtime_refs(
                conn,
                entity_code="AA",
                table_name="业务状况表",
                metrics=metrics,
            )

            self.assertEqual(
                conn.execute(
                    """
                    SELECT parent_code
                    FROM data_account_metric_node
                    WHERE node_code='AA.05.01.02.03.03.001'
                    """
                ).fetchone(),
                ("AA.05.01.02.03",),
            )
            self.assertIsNone(
                conn.execute(
                    """
                    SELECT node_code
                    FROM data_account_metric_node
                    WHERE node_code='AA.05.01.02.03.03'
                    """
                ).fetchone()
            )

    def test_purge_unreferenced_legacy_aa05_metric_master_removes_old_duplicate_branch(self) -> None:
        with sqlite3.connect(":memory:") as conn:
            self._create_runtime_metric_tables(conn)
            conn.executescript(
                """
                INSERT INTO data_account_metric_node(
                  node_code, node_name, parent_code, product_code, local_metric_code,
                  functional_group_code, metric_table_name, level, node_type, runtime_account_enabled,
                  allow_manual_entry, value_type
                )
                VALUES
                  ('AA', '微众银行', NULL, 'AA', '', '', '', 1, 'CATEGORY', 0, 1, '金额'),
                  ('AA.05', '减:业务及管理费', 'AA', 'AA', '05', '业务状况表', '业务状况表', 2, 'GROUP', 1, 1, '金额'),
                  ('AA.05.01', '人力费用', 'AA.05', 'AA', '05.01', '业务状况表', '业务状况表', 3, 'METRIC', 1, 1, '金额');
                """
            )

            removed = purge_unreferenced_legacy_aa05_metric_master(
                conn,
                budget_paths=(),
                read_model_paths=(),
            )

            self.assertEqual(removed, 2)
            self.assertEqual(
                conn.execute(
                    """
                    SELECT COUNT(*)
                    FROM data_account_metric_node
                    WHERE node_code='AA.05' OR node_code LIKE 'AA.05.%'
                    """
                ).fetchone(),
                (0,),
            )

    def test_canonical_expense_merge_renames_existing_runtime_group_accounts(self) -> None:
        with sqlite3.connect(":memory:") as conn:
            self._create_runtime_metric_tables(conn)
            conn.execute(
                """
                INSERT INTO data_account_metric_node(
                  node_code, node_name, product_code, local_metric_code, logic_code, level,
                  node_type, runtime_account_enabled, value_type
                )
                VALUES('A01.90.01', '旧名称', 'A01', '90.01', '90.01', 4,
                       'GROUP', 1, '金额')
                """
            )

            merge_canonical_expense_metric_trees_into_org_product_metrics(conn)

            self.assertEqual(
                conn.execute(
                    "SELECT data_acct_name FROM data_account WHERE data_acct_code='A01.90.01'"
                ).fetchone(),
                ("直接费用",),
            )

    def test_canonical_expense_merge_deduplicates_existing_payload_nodes(self) -> None:
        with sqlite3.connect(":memory:") as conn:
            self._create_runtime_metric_tables(conn)
            duplicate_child = {
                "id": "canonical-B010501010303001",
                "levelLabel": "六级",
                "nature": "其他",
                "code": "B010501010303001",
                "name": "IT职场",
                "metric_node_code": "B01.05.01.01.03.03.001",
                "data_acct_code": "B01.05.01.01.03.03.001",
                "mapping_status": "MANUAL_CONFIRMED",
                "children": [],
            }
            payload = {
                "id": "table-业务状况表",
                "name": "业务状况表",
                "metrics": [
                    {
                        "id": "canonical-B01050101",
                        "levelLabel": "三级",
                        "nature": "其他",
                        "code": "B01050101",
                        "name": "直接费用",
                        "children": [
                            {
                                "id": "canonical-B0105010103",
                                "levelLabel": "四级",
                                "nature": "其他",
                                "code": "B0105010103",
                                "name": "IT",
                                "children": [dict(duplicate_child), dict(duplicate_child)],
                            }
                        ],
                    }
                ],
            }
            # The old org_product_metric_table has been retired; seed the
            # pre-existing payload data directly into data_account_metric_node
            # so that the merge function can load it via the runtime tree.
            conn.execute(
                """
                INSERT INTO data_account_metric_node(
                  node_code, node_name, parent_code, product_code, local_metric_code,
                  logic_code, functional_group_code, metric_table_name, level, node_type,
                  horizontal_rollup, vertical_rollup, runtime_account_enabled,
                  sort_order, is_active, remark, value_type, allow_manual_entry
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                ("B01", "企业金融", None, "B01", "", "", "业务状况表", "业务状况表",
                 1, "CATEGORY", 0, 0, 0, 0, 1, "", "金额", 0),
            )
            conn.execute(
                """
                INSERT INTO data_account_metric_node(
                  node_code, node_name, parent_code, product_code, local_metric_code,
                  logic_code, functional_group_code, metric_table_name, level, node_type,
                  horizontal_rollup, vertical_rollup, runtime_account_enabled,
                  sort_order, is_active, remark, value_type, allow_manual_entry
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                ("B01.05.01.01", "直接费用", "B01.05.01", "B01", "05.01.01", "05.01.01",
                 "业务状况表", "业务状况表", 4, "GROUP", 0, 0, 0, 0, 1, "", "金额", 0),
            )
            conn.execute(
                """
                INSERT INTO data_account_metric_node(
                  node_code, node_name, parent_code, product_code, local_metric_code,
                  logic_code, functional_group_code, metric_table_name, level, node_type,
                  horizontal_rollup, vertical_rollup, runtime_account_enabled,
                  sort_order, is_active, remark, value_type, allow_manual_entry
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                ("B01.05.01.01.03.03.001", "IT职场", "B01.05.01.01.03.03", "B01",
                 "05.01.01.03.03.001", "05.01.01.03.03.001",
                 "业务状况表", "业务状况表", 7, "METRIC", 0, 0, 1, 0, 1, "", "金额", 1),
            )

            merge_canonical_expense_metric_trees_into_org_product_metrics(conn)

            saved = load_org_product_metric_payload_from_runtime_tree(
                conn,
                entity_code="B01",
                table_name="业务状况表",
            )
            self.assertIsNotNone(saved)
            matches = [
                node
                for node in _flatten_metric_payload(saved["metrics"])  # type: ignore[index]
                if node.get("code") == "B01.05.01.01.03.03.001"
            ]
            self.assertEqual(len(matches), 1)
            _assert_no_legacy_runtime_identity_fields(self, matches[0])

    def test_canonical_expense_merge_prunes_legacy_dotted_duplicate_payload_nodes(self) -> None:
        with sqlite3.connect(":memory:") as conn:
            self._create_runtime_metric_tables(conn)
            legacy_payload = {
                "id": "table-业务状况表",
                "name": "业务状况表",
                "metrics": [
                    {
                        "id": "import-v02-B01-legacy",
                        "levelLabel": "四级",
                        "nature": "其他",
                        "code": "B01.90.01",
                        "name": "旧名称",
                        "children": [],
                    }
                ],
            }
            conn.execute(
                """
                CREATE TABLE org_product_metric_table (
                  entity_code TEXT NOT NULL,
                  entity_name TEXT NOT NULL,
                  table_id TEXT NOT NULL,
                  table_name TEXT NOT NULL,
                  payload_json TEXT NOT NULL,
                  updated_at TEXT NOT NULL,
                  PRIMARY KEY(entity_code, table_name)
                )
                """
            )
            conn.execute(
                """
                INSERT INTO org_product_metric_table
                  (entity_code, entity_name, table_id, table_name, payload_json, updated_at)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                ("B01", "企业金融", "table-业务状况表", "业务状况表", json.dumps(legacy_payload, ensure_ascii=False), "2026-06-13T00:00:00Z"),
            )

            merge_canonical_expense_metric_trees_into_org_product_metrics(conn)

            saved = load_org_product_metric_payload_from_runtime_tree(
                conn,
                entity_code="B01",
                table_name="业务状况表",
            )
            self.assertIsNotNone(saved)
            matches = [
                node
                for node in _flatten_metric_payload(saved["metrics"])  # type: ignore[index]
                if node.get("code") == "B01.90.01"
            ]
            self.assertEqual(len(matches), 1)
            self.assertEqual(matches[0]["id"], "canonical-B019001")
            self.assertEqual(matches[0]["name"], "直接费用")

    def test_blank_org_product_mapping_statuses_are_removed_before_runtime_sync(self) -> None:
        with sqlite3.connect(":memory:") as conn:
            conn.executescript(
                """
                CREATE TABLE org_product_metric_table (
                  entity_code TEXT NOT NULL,
                  entity_name TEXT NOT NULL,
                  table_id TEXT NOT NULL,
                  table_name TEXT NOT NULL,
                  payload_json TEXT NOT NULL,
                  updated_at TEXT NOT NULL,
                  PRIMARY KEY(entity_code, table_name)
                );
                INSERT INTO org_product_metric_table
                  (entity_code, entity_name, table_id, table_name, payload_json, updated_at)
                VALUES
                  ('AA', '微众银行', 'table-main', '业务状况表',
                   '{"metrics":[{"code":"AA01N04","name":"新二级指标","mapping_status":"","metric_node_code":"","data_acct_code":"","children":[]}]}',
                   '2026-06-10T00:00:00Z');
                """
            )

            updated = normalize_org_product_metric_mapping_statuses(conn)

            self.assertEqual(updated, 1)
            payload = json.loads(
                conn.execute("SELECT payload_json FROM org_product_metric_table").fetchone()[0]
            )
            self.assertNotIn("mapping_status", payload["metrics"][0])
            self.assertNotIn("metric_node_code", payload["metrics"][0])
            self.assertNotIn("data_acct_code", payload["metrics"][0])

    def test_legacy_protected_mapping_statuses_are_removed_before_runtime_sync(self) -> None:
        with sqlite3.connect(":memory:") as conn:
            conn.executescript(
                """
                CREATE TABLE org_product_metric_table (
                  entity_code TEXT NOT NULL,
                  entity_name TEXT NOT NULL,
                  table_id TEXT NOT NULL,
                  table_name TEXT NOT NULL,
                  payload_json TEXT NOT NULL,
                  updated_at TEXT NOT NULL,
                  PRIMARY KEY(entity_code, table_name)
                );
                INSERT INTO org_product_metric_table
                  (entity_code, entity_name, table_id, table_name, payload_json, updated_at)
                VALUES
                  ('A01', '开鑫贷', 'table-main', '业务状况表',
                   '{"metrics":[{"code":"A0105","name":"旧保护费用类","mapping_status":"PROTECTED_05_REVIEW_ONLY","metric_node_code":"A01.05","data_acct_code":"A01.05","children":[]}]}',
                   '2026-06-10T00:00:00Z');
                """
            )

            updated = normalize_org_product_metric_mapping_statuses(conn)

            self.assertEqual(updated, 1)
            payload = json.loads(
                conn.execute("SELECT payload_json FROM org_product_metric_table").fetchone()[0]
            )
            metric = payload["metrics"][0]
            self.assertNotIn("mapping_status", metric)
            self.assertNotIn("metric_node_code", metric)
            self.assertNotIn("data_acct_code", metric)

    def test_runtime_tree_account_is_confirmed_without_json_metric_master_backfill(self) -> None:
        with sqlite3.connect(":memory:") as conn:
            self._create_runtime_metric_tables(conn)
            conn.executescript(
                """
                CREATE TABLE org_product_metric_table (
                  entity_code TEXT NOT NULL,
                  entity_name TEXT NOT NULL,
                  table_id TEXT NOT NULL,
                  table_name TEXT NOT NULL,
                  payload_json TEXT NOT NULL,
                  updated_at TEXT NOT NULL,
                  PRIMARY KEY(entity_code, table_name)
                );
                INSERT INTO data_account_metric_node
                  (node_code, node_name, parent_code, product_code, local_metric_code,
                   functional_group_code, metric_table_name, level, node_type, runtime_account_enabled,
                   allow_manual_entry, value_type)
                VALUES
                  ('A01', '产品A', NULL, 'A01', '', '', '', 1, 'CATEGORY', 0, 1, '金额'),
                  ('A01.14', '贷款利息收入', 'A01', 'A01', '14', '业务状况表', '业务状况表', 2, 'CATEGORY', 0, 1, '金额'),
                  ('A01.14.01', '自营贷款', 'A01.14', 'A01', '14.01', '业务状况表', '业务状况表', 3, 'CATEGORY', 0, 1, '金额'),
                  ('A01.14.01.03', '利息收入', 'A01.14.01', 'A01', '14.01.03', '业务状况表', '业务状况表', 4, 'METRIC', 1, 1, '金额');
                """
            )

            assert_all_runtime_metric_refs_are_confirmed_org_product_metrics(
                conn,
                budget_paths=(),
                read_model_paths=(),
            )
            self.assertIsNone(
                conn.execute("SELECT payload_json FROM org_product_metric_table").fetchone()
            )

    def test_existing_node_only_ref_does_not_hide_orphan_runtime_account(self) -> None:
        with sqlite3.connect(":memory:") as conn:
            self._create_runtime_metric_tables(conn)
            conn.executescript(
                """
                CREATE TABLE org_product_metric_table (
                  entity_code TEXT NOT NULL,
                  entity_name TEXT NOT NULL,
                  table_id TEXT NOT NULL,
                  table_name TEXT NOT NULL,
                  payload_json TEXT NOT NULL,
                  updated_at TEXT NOT NULL,
                  PRIMARY KEY(entity_code, table_name)
                );
                INSERT INTO data_account_metric_node
                  (node_code, node_name, parent_code, product_code, local_metric_code,
                   functional_group_code, metric_table_name, level, node_type, runtime_account_enabled,
                   allow_manual_entry, value_type)
                VALUES
                  ('A01', '产品A', NULL, 'A01', '', '', '', 1, 'CATEGORY', 0, 1, '金额'),
                  ('A01.02', '减:风险成本', 'A01', 'A01', '02', '业务状况表', '业务状况表', 2, 'CATEGORY', 0, 1, '金额'),
                  ('A01.02.03', '付息率', 'A01.02', 'A01', '02.03', '业务状况表', '业务状况表', 3, 'METRIC', 1, 1, '百分比');
                """
            )
            existing_payload = {
                "id": "table-业务状况表",
                "name": "业务状况表",
                "metrics": [
                    {
                        "id": "risk",
                        "levelLabel": "一级",
                        "nature": "其他",
                        "code": "A0102",
                        "name": "减:风险成本",
                        "metric_node_code": "A01.02",
                        "data_acct_code": "A01.02",
                        "mapping_status": "MANUAL_CONFIRMED",
                        "children": [
                            {
                                "id": "risk-rate",
                                "levelLabel": "二级",
                                "nature": "其他",
                                "code": "A010203",
                                "name": "付息率",
                                "metric_node_code": "A01.02.03",
                                "data_acct_code": "A01.02.03",
                                "mapping_status": "REVIEW_NODE_ONLY_NO_DATA_ACCOUNT",
                                "children": [],
                            }
                        ],
                    }
                ],
            }
            conn.execute(
                """
                INSERT INTO org_product_metric_table
                (entity_code, entity_name, table_id, table_name, payload_json, updated_at)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                ("A01", "产品A", "table-业务状况表", "业务状况表", json.dumps(existing_payload, ensure_ascii=False), "2026-06-10T00:00:00Z"),
            )

            assert_all_runtime_metric_refs_are_confirmed_org_product_metrics(
                conn,
                budget_paths=(),
                read_model_paths=(),
            )
            payload = json.loads(conn.execute("SELECT payload_json FROM org_product_metric_table WHERE entity_code='A01'").fetchone()[0])
            unchanged = payload["metrics"][0]["children"][0]
            self.assertEqual(unchanged["mapping_status"], "REVIEW_NODE_ONLY_NO_DATA_ACCOUNT")

    def test_budget_fact_refs_must_already_exist_in_org_product_metric_master(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            budget_path = Path(tmp) / "budget_2026.db"
            with sqlite3.connect(budget_path) as bconn:
                bconn.execute(
                    """
                    CREATE TABLE budget_data (
                      data_acct_code TEXT NOT NULL,
                      product_code TEXT NOT NULL,
                      period_id INTEGER NOT NULL,
                      budget_actual INTEGER NOT NULL,
                      version_id INTEGER NOT NULL,
                      value REAL NOT NULL
                    )
                    """
                )
                bconn.execute(
                    """
                    INSERT INTO budget_data
                    (data_acct_code, product_code, period_id, budget_actual, version_id, value)
                    VALUES ('A01.01.02.05.01.028', 'A01', 4, 0, 1, 100)
                    """
                )
            with sqlite3.connect(":memory:") as conn:
                self._create_runtime_metric_tables(conn)
                conn.execute(
                    """
                    CREATE TABLE org_product_metric_table (
                      entity_code TEXT NOT NULL,
                      entity_name TEXT NOT NULL,
                      table_id TEXT NOT NULL,
                      table_name TEXT NOT NULL,
                      payload_json TEXT NOT NULL,
                      updated_at TEXT NOT NULL,
                      PRIMARY KEY(entity_code, table_name)
                    )
                    """
                )

                with self.assertRaisesRegex(
                    OrgProductMetricRuntimeSyncError,
                    "A01.01.02.05.01.028",
                ):
                    assert_all_runtime_metric_refs_are_confirmed_org_product_metrics(
                        conn,
                        budget_paths=(budget_path,),
                        read_model_paths=(),
                    )

    def test_runtime_sync_ignores_mismatched_legacy_metric_refs(self) -> None:
        with sqlite3.connect(":memory:") as conn:
            self._create_runtime_metric_tables(conn)
            metrics = [
                {
                    "code": "A010101",
                    "name": "利息收入",
                    "metric_node_code": "A01.01.99",
                    "data_acct_code": "A01.01.02",
                }
            ]

            sync_org_product_metric_runtime_refs(
                conn,
                entity_code="A01",
                table_name="业务状况表",
                metrics=metrics,
            )

            self.assertNotIn("metric_node_code", metrics[0])
            self.assertNotIn("data_acct_code", metrics[0])
            self.assertEqual(
                conn.execute(
                    "SELECT data_acct_code FROM data_account WHERE data_acct_code='A01.01.01'"
                ).fetchone(),
                ("A01.01.01",),
            )

    def test_inventory_guard_rejects_derivable_leaf_without_runtime_ref(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            db_path = Path(tmp) / "common.db"
            payload = {
                "id": "table-业务状况表",
                "name": "业务状况表",
                "metrics": [
                    {
                        "id": "fee05",
                        "levelLabel": "三级",
                        "nature": "支出",
                        "code": "A01050101",
                        "name": "常规人力",
                        "children": [],
                    }
                ],
            }
            with sqlite3.connect(db_path) as conn:
                self._create_runtime_metric_tables(conn)
                conn.execute(
                    """
                    CREATE TABLE org_product_metric_table (
                      entity_code TEXT NOT NULL,
                      entity_name TEXT NOT NULL,
                      table_id TEXT NOT NULL,
                      table_name TEXT NOT NULL,
                      payload_json TEXT NOT NULL,
                      updated_at TEXT NOT NULL,
                      PRIMARY KEY (entity_code, table_name)
                    )
                    """
                )
                conn.execute(
                    """
                    INSERT INTO org_product_metric_table
                    (entity_code, entity_name, table_id, table_name, payload_json, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (
                        "A01",
                        "泛微粒贷",
                        "table-业务状况表",
                        "业务状况表",
                        json.dumps(payload, ensure_ascii=False),
                        "2026-06-10T00:00:00Z",
                    ),
                )

            violations = org_product_metric_runtime_ref_violations(db_path)

        self.assertEqual(
            violations,
            (
                "org_product_ref_missing_metric_node:A01/业务状况表/A01050101/A01.05.01.01/A01.05.01.01",
                "org_product_ref_missing_data_account:A01/业务状况表/A01050101/A01.05.01.01/A01.05.01.01",
                "org_product_ref_missing_or_invalid_binding:A01/业务状况表/A01050101/A01.05.01.01/A01.05.01.01",
            ),
        )

    def test_inventory_guard_rejects_data_account_without_org_product_confirmed_ref(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            db_path = Path(tmp) / "common.db"
            with sqlite3.connect(db_path) as conn:
                self._create_runtime_metric_tables(conn)
                conn.executescript(
                    """
                    CREATE TABLE org_product_metric_table (
                      entity_code TEXT NOT NULL,
                      entity_name TEXT NOT NULL,
                      table_id TEXT NOT NULL,
                      table_name TEXT NOT NULL,
                      payload_json TEXT NOT NULL,
                      updated_at TEXT NOT NULL,
                      PRIMARY KEY (entity_code, table_name)
                    );
                    INSERT INTO data_account_metric_node
                      (node_code, node_name, parent_code, product_code, local_metric_code,
                       logic_code, level, node_type, runtime_account_enabled, value_type)
                    VALUES ('A01.01.01', '利息收入', NULL, 'A01', '01.01',
                            '01.01', 3, 'METRIC', 1, '金额');
                    INSERT INTO org_product_metric_table
                    (entity_code, entity_name, table_id, table_name, payload_json, updated_at)
                    VALUES ('A01', '泛微粒贷', 'table-业务状况表', '业务状况表', '{"metrics":[]}', '2026-06-10T00:00:00Z');
                    """
                )

            violations = org_product_metric_runtime_ref_violations(db_path)

        self.assertEqual(
            violations,
            ("data_account_missing_org_product_confirmed_ref:A01.01.01",),
        )

    def test_save_sanitizer_drops_fee_05_legacy_references(self) -> None:
        rows = _sanitize_metric_nodes_for_save(
            "A01",
            [
                MetricNodePayload(
                    id="fee05",
                    levelLabel="二级",
                    nature="支出",
                    code="A010501",
                    name="05费用测试",
                    metric_node_code="A01.05.01",
                    data_acct_code="A01.05.01",
                    mapping_status="MANUAL_CONFIRMED",
                    note="",
                )
            ],
        )

        _assert_no_legacy_runtime_identity_fields(self, rows[0])

    def test_save_sanitizer_drops_non_05_legacy_references(self) -> None:
        rows = _sanitize_metric_nodes_for_save(
            "A01",
            [
                MetricNodePayload(
                    id="normal",
                    levelLabel="二级",
                    nature="收入",
                    code="A010101",
                    name="普通测试",
                    metric_node_code="A01.01.01",
                    data_acct_code="A01.01.01",
                    mapping_status="MANUAL_CONFIRMED",
                    value_type="百分比",
                    allow_manual_entry=0,
                    note="",
                )
            ],
        )

        _assert_no_legacy_runtime_identity_fields(self, rows[0])
        self.assertEqual(rows[0]["value_type"], "百分比")
        self.assertEqual(rows[0]["allow_manual_entry"], 0)

    def test_metric_response_sanitizer_drops_legacy_references(self) -> None:
        rows = _sanitize_metric_node_dicts_for_response(
            "A01",
            [
                {
                    "id": "normal",
                    "levelLabel": "一级",
                    "nature": "收入",
                    "code": "A0101",
                    "name": "营业收入",
                    "mapping_status": "MANUAL_CONFIRMED",
                    "metric_node_code": "A01.01.01",
                    "data_acct_code": "A01.01.01",
                    "value_type": "百分比",
                    "allow_manual_entry": 0,
                    "children": [
                        {
                            "id": "fee05-child",
                            "levelLabel": "二级",
                            "nature": "支出",
                            "code": "A010501",
                            "name": "直接费用",
                            "mapping_status": "MANUAL_CONFIRMED",
                            "metric_node_code": "A01.05.01",
                            "data_acct_code": "A01.05.01",
                        }
                    ],
                }
            ],
        )

        _assert_no_legacy_runtime_identity_fields(self, rows[0])
        _assert_no_legacy_runtime_identity_fields(self, rows[0]["children"][0])

    def test_metric_response_sanitizer_deduplicates_sibling_runtime_refs(self) -> None:
        rows = _sanitize_metric_node_dicts_for_response(
            "B01",
            [
                {
                    "id": "parent",
                    "levelLabel": "五级",
                    "nature": "支出",
                    "code": "B0105010103",
                    "name": "IT",
                    "children": [
                        {
                            "id": "first",
                            "levelLabel": "六级",
                            "nature": "支出",
                            "code": "B010501010303001",
                            "name": "IT职场",
                            "metric_node_code": "B01.05.01.01.03.03.001",
                            "data_acct_code": "B01.05.01.01.03.03.001",
                            "mapping_status": "MANUAL_CONFIRMED",
                        },
                        {
                            "id": "second",
                            "levelLabel": "六级",
                            "nature": "支出",
                            "code": "B010501010303001",
                            "name": "IT职场",
                            "metric_node_code": "B01.05.01.01.03.03.001",
                            "data_acct_code": "B01.05.01.01.03.03.001",
                            "mapping_status": "MANUAL_CONFIRMED",
                        },
                    ],
                }
            ],
        )

        self.assertEqual(len(rows[0]["children"]), 1)

    def test_data_entry_snapshot_sanitizer_deduplicates_legacy_and_canonical_metric_rows(self) -> None:
        with sqlite3.connect(":memory:") as conn:
            conn.execute(
                """
                CREATE TABLE org_product_metric_table (
                  entity_code TEXT NOT NULL,
                  entity_name TEXT NOT NULL,
                  table_id TEXT NOT NULL,
                  table_name TEXT NOT NULL,
                  payload_json TEXT NOT NULL,
                  updated_at TEXT NOT NULL,
                  PRIMARY KEY(entity_code, table_name)
                )
                """
            )
            metric_payload = {
                "metrics": [
                    {
                        "id": "canonical-B0105010101001",
                        "levelLabel": "五级",
                        "nature": "其他",
                        "code": "B01.05.01.01.01.001",
                        "name": "业务人力",
                        "children": [],
                    }
                ]
            }
            conn.execute(
                """
                INSERT INTO org_product_metric_table
                  (entity_code, entity_name, table_id, table_name, payload_json, updated_at)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                ("B01", "企业金融", "table-业务状况表", "业务状况表", json.dumps(metric_payload, ensure_ascii=False), "2026-06-13T00:00:00Z"),
            )
            payload = {
                "entity_code": "B01",
                "entity_name": "企业金融",
                "table_name": "业务状况表",
                "metrics": [
                    {
                        "metric_id": "import-v02-B01-legacy",
                        "metric_code": "B01.05.01.01.01.001",
                        "metric_name": "旧业务人力",
                        "values": {"months": {"1": "10"}},
                    },
                    {
                        "metric_id": "canonical-B0105010101001",
                        "metric_code": "B0105010101001",
                        "metric_name": "业务人力",
                        "values": {"months": {"2": "20"}},
                    },
                ],
            }

            sanitized = _sanitize_data_entry_payload_mapping_refs(conn, payload)

        self.assertEqual(len(sanitized["metrics"]), 1)
        row = sanitized["metrics"][0]
        self.assertEqual(row["metric_id"], "canonical-B0105010101001")
        self.assertEqual(row["metric_code"], "B01.05.01.01.01.001")
        self.assertEqual(row["metric_name"], "业务人力")
        self.assertEqual(row["values"]["months"], {"1": "10", "2": "20"})

    def test_output_snapshot_sanitizer_preserves_value_type_for_display_formatting(self) -> None:
        sanitized = _sanitize_org_product_output_entity_for_snapshot(
            {
                "rows": [
                    {
                        "id": "canonical-B01050101",
                        "levelLabel": "三级",
                        "nature": "其他",
                        "code": "B01.05.01.01",
                        "name": "直接费用",
                        "value_type": "金额",
                        "months": [123.4],
                        "annual": 123.4,
                    }
                ]
            }
        )

        self.assertEqual(sanitized["rows"][0]["value_type"], "金额")

    def test_save_sanitizer_deduplicates_sibling_runtime_refs(self) -> None:
        duplicate = MetricNodePayload(
            id="dup",
            levelLabel="六级",
            nature="支出",
            code="B010501010303001",
            name="IT职场",
            metric_node_code="B01.05.01.01.03.03.001",
            data_acct_code="B01.05.01.01.03.03.001",
            mapping_status="MANUAL_CONFIRMED",
        )
        rows = _sanitize_metric_nodes_for_save(
            "B01",
            [
                MetricNodePayload(
                    id="parent",
                    levelLabel="五级",
                    nature="支出",
                    code="B0105010103",
                    name="IT",
                    children=[duplicate, duplicate],
                )
            ],
        )

        self.assertEqual(len(rows[0]["children"]), 1)

    def test_metric_export_rows_drop_legacy_mapping_refs(self) -> None:
        rows: list[dict[str, str]] = []
        _build_metric_rows(
            [
                {
                    "id": "normal",
                    "levelLabel": "一级",
                    "nature": "收入",
                    "code": "A0101",
                    "name": "营业收入",
                    "mapping_status": "MANUAL_CONFIRMED",
                    "metric_node_code": "A01.01.01",
                    "data_acct_code": "A01.01.01",
                    "value_type": "百分比",
                    "allow_manual_entry": 0,
                    "children": [
                        {
                            "id": "fee05",
                            "levelLabel": "二级",
                            "nature": "支出",
                            "code": "A010501",
                            "name": "直接费用",
                            "mapping_status": "MANUAL_CONFIRMED",
                            "metric_node_code": "A01.05.01",
                            "data_acct_code": "A01.05.01",
                        }
                    ],
                }
            ],
            rows,
            "A01",
        )

        _assert_no_legacy_runtime_identity_fields(self, rows[0])
        self.assertEqual(rows[0]["value_type"], "百分比")
        self.assertEqual(rows[0]["allow_manual_entry"], 0)
        _assert_no_legacy_runtime_identity_fields(self, rows[1])

    def test_metric_import_accepts_current_org_product_metric_fields_only(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "A01业务状况表"
        ws.append(["科目层级", "科目性质", "科目代码", "科目名称", "机构及产品指标编码", "数值类型", "允许手工录入"])
        ws.append(["一级", "收入", "A0101", "营业收入", "A0101", "百分比", "不允许"])
        ws.append(["二级", "支出", "A010501", "直接费用", "A010501", "金额", "允许"])

        rows, row_count, parse_error, _header_map = _parse_metric_worksheet_basic(ws, "test", entity_code="A01", strict=True)

        self.assertIsNone(parse_error)
        self.assertEqual(row_count, 2)
        _assert_no_legacy_runtime_identity_fields(self, rows[0])
        self.assertEqual(rows[0]["value_type"], "百分比")
        self.assertEqual(rows[0]["allow_manual_entry"], 0)
        _assert_no_legacy_runtime_identity_fields(self, rows[0]["children"][0])

    def test_metric_import_no_longer_maps_legacy_runtime_ref_headers(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "A01业务状况表"
        ws.append(["科目层级", "科目性质", "科目代码", "科目名称", "运行引用", "数据科目", "指标节点", "映射状态"])
        ws.append(["一级", "收入", "A0101", "营业收入", "A01.01.01", "A01.01.01", "A01.01.01", "MANUAL_CONFIRMED"])

        _rows, _row_count, parse_error, header_map = _parse_metric_worksheet_basic(
            ws,
            "test",
            entity_code="A01",
            strict=True,
        )

        self.assertIsNone(parse_error)
        self.assertNotIn("机构及产品指标编码", header_map)
        self.assertNotIn("指标节点", header_map)
        self.assertNotIn("映射状态", header_map)

    def test_inventory_guard_rejects_protected_05_with_refs(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            db_path = Path(tmp) / "common.db"
            payload = {
                "id": "table-业务状况表",
                "name": "业务状况表",
                "metrics": [
                    {
                        "id": "fee05",
                        "levelLabel": "一级",
                        "nature": "支出",
                        "code": "A0105",
                        "name": "直接费用",
                        "mapping_status": "PROTECTED_05_REVIEW_ONLY",
                        "metric_node_code": "A01.05",
                        "data_acct_code": "A01.05",
                        "children": [],
                    }
                ],
            }
            with sqlite3.connect(db_path) as conn:
                conn.execute(
                    """
                    CREATE TABLE org_product_metric_table (
                      entity_code TEXT NOT NULL,
                      entity_name TEXT NOT NULL,
                      table_id TEXT NOT NULL,
                      table_name TEXT NOT NULL,
                      payload_json TEXT NOT NULL,
                      updated_at TEXT NOT NULL,
                      PRIMARY KEY (entity_code, table_name)
                    )
                    """
                )
                conn.execute(
                    """
                    INSERT INTO org_product_metric_table
                    (entity_code, entity_name, table_id, table_name, payload_json, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    ("A01", "泛微粒贷", "table-业务状况表", "业务状况表", json.dumps(payload, ensure_ascii=False), "2026-06-05T00:00:00Z"),
                )

            violations = org_product_metric_guard_violations(db_path)

        self.assertEqual(violations, ("legacy_protected_status:A01/业务状况表/A0105/A01.05/A01.05",))

    def test_inventory_guard_rejects_snapshot_05_with_refs(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            db_path = Path(tmp) / "common.db"
            data_entry_payload = {
                "metrics": [
                    {
                        "metric_code": "A0105",
                        "metric_name": "直接费用",
                        "mapping_status": "PROTECTED_05_REVIEW_ONLY",
                        "metric_node_code": "A01.05",
                        "data_acct_code": "A01.05",
                    }
                ]
            }
            output_payload = {
                "rows": [
                    {
                        "code": "A0105",
                        "name": "直接费用",
                        "mapping_status": "PROTECTED_05_REVIEW_ONLY",
                        "metric_node_code": "A01.05",
                        "data_acct_code": "A01.05",
                    }
                ]
            }
            with sqlite3.connect(db_path) as conn:
                conn.execute(
                    """
                    CREATE TABLE org_product_data_entry_snapshot (
                      entity_code TEXT NOT NULL,
                      entity_name TEXT NOT NULL,
                      year INTEGER NOT NULL,
                      month_index INTEGER,
                      table_id TEXT,
                      table_name TEXT,
                      payload_json TEXT NOT NULL,
                      updated_at TEXT NOT NULL,
                      PRIMARY KEY (entity_code, year)
                    )
                    """
                )
                conn.execute(
                    """
                    CREATE TABLE org_product_data_entry_snapshot_v2 (
                      entity_code TEXT NOT NULL,
                      entity_name TEXT NOT NULL,
                      year INTEGER NOT NULL,
                      version_id INTEGER NOT NULL,
                      version_name TEXT NOT NULL,
                      table_name TEXT NOT NULL,
                      month_index INTEGER,
                      table_id TEXT,
                      payload_json TEXT NOT NULL,
                      updated_at TEXT NOT NULL,
                      PRIMARY KEY (entity_code, year, version_id, table_name)
                    )
                    """
                )
                conn.execute(
                    """
                    CREATE TABLE org_product_data_entry_draft (
                      user_id INTEGER NOT NULL,
                      user_display_name TEXT NOT NULL,
                      entity_code TEXT NOT NULL,
                      entity_name TEXT NOT NULL,
                      year INTEGER NOT NULL,
                      table_name TEXT NOT NULL,
                      payload_json TEXT NOT NULL,
                      updated_at TEXT NOT NULL,
                      PRIMARY KEY (user_id, entity_code, year, table_name)
                    )
                    """
                )
                conn.execute(
                    """
                    CREATE TABLE org_product_output_snapshot_v1 (
                      entity_code TEXT NOT NULL,
                      entity_name TEXT NOT NULL,
                      year INTEGER NOT NULL,
                      input_version_id INTEGER NOT NULL,
                      output_version_id INTEGER NOT NULL,
                      output_version_name TEXT NOT NULL,
                      table_name TEXT NOT NULL,
                      payload_json TEXT NOT NULL,
                      updated_at TEXT NOT NULL,
                      PRIMARY KEY (entity_code, year, input_version_id, output_version_id, table_name)
                    )
                    """
                )
                conn.execute(
                    """
                    INSERT INTO org_product_data_entry_snapshot
                    (entity_code, entity_name, year, month_index, table_id, table_name, payload_json, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        "A01",
                        "泛微粒贷",
                        2026,
                        3,
                        "",
                        "业务状况表",
                        json.dumps(data_entry_payload, ensure_ascii=False),
                        "2026-06-05T00:00:00Z",
                    ),
                )
                conn.execute(
                    """
                    INSERT INTO org_product_data_entry_snapshot_v2
                    (entity_code, entity_name, year, version_id, version_name, table_name, month_index, table_id, payload_json, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        "A01",
                        "泛微粒贷",
                        2026,
                        1,
                        "202603v1",
                        "业务状况表",
                        3,
                        "",
                        json.dumps(data_entry_payload, ensure_ascii=False),
                        "2026-06-05T00:00:00Z",
                    ),
                )
                conn.execute(
                    """
                    INSERT INTO org_product_data_entry_draft
                    (user_id, user_display_name, entity_code, entity_name, year, table_name, payload_json, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        1,
                        "测试用户",
                        "A01",
                        "泛微粒贷",
                        2026,
                        "业务状况表",
                        json.dumps(data_entry_payload, ensure_ascii=False),
                        "2026-06-05T00:00:00Z",
                    ),
                )
                conn.execute(
                    """
                    INSERT INTO org_product_output_snapshot_v1
                    (entity_code, entity_name, year, input_version_id, output_version_id, output_version_name, table_name, payload_json, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        "A01",
                        "泛微粒贷",
                        2026,
                        1,
                        1,
                        "输出v1",
                        "业务状况表",
                        json.dumps(output_payload, ensure_ascii=False),
                        "2026-06-05T00:00:00Z",
                    ),
                )

            violations = org_product_metric_guard_violations(db_path)

        self.assertEqual(
            violations,
            (
                "data_entry_legacy_legacy_protected_status:A01/业务状况表/A0105/A01.05/A01.05",
                "data_entry_legacy_protected_status:A01/业务状况表/A0105/A01.05/A01.05",
                "data_entry_draft_legacy_protected_status:A01/业务状况表/A0105/A01.05/A01.05",
                "output_legacy_protected_status:A01/业务状况表/A0105/A01.05/A01.05",
            ),
        )

    def test_inventory_guard_rejects_business_cost_income_05_topic_refs(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            db_path = Path(tmp) / "budget_2026.db"
            with sqlite3.connect(db_path) as conn:
                conn.execute(
                    """
                    CREATE TABLE business_cost_income_indicator (
                      id INTEGER PRIMARY KEY,
                      product_code TEXT NOT NULL,
                      name TEXT NOT NULL,
                      topic_metric_node_code TEXT
                    )
                    """
                )
                conn.executemany(
                    """
                    INSERT INTO business_cost_income_indicator
                    (id, product_code, name, topic_metric_node_code)
                    VALUES (?, ?, ?, ?)
                    """,
                    [
                        (1, "A01", "正常主题", "A01:业务状况表:A0101"),
                        (2, "A01", "05主题", "A01:业务状况表:A010501"),
                    ],
                )

            violations = org_product_metric_guard_violations(db_path)

        self.assertEqual(violations, ())

    def test_data_entry_metric_row_ignores_legacy_mapping_refs(self) -> None:
        row = DataEntryMetricRowPayload(
            metric_id="metric-A01-1",
            metric_code="A0101",
            metric_name="营业收入",
            metric_node_code="A01.01.01",
            data_acct_code="A01.01.01",
            mapping_status="MANUAL_CONFIRMED",
            levelLabel="一级",
            nature="收入",
            values=DataEntryMetricValuesPayload(),
        )

        dumped = row.model_dump()

        _assert_no_legacy_runtime_identity_fields(self, dumped)

    def test_data_entry_payload_drops_legacy_mapping_refs_from_metric_table(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            db_path = Path(tmp) / "common.db"
            with sqlite3.connect(db_path) as conn:
                conn.execute(
                    """
                    CREATE TABLE org_product_metric_table (
                      entity_code TEXT NOT NULL,
                      entity_name TEXT NOT NULL,
                      table_id TEXT NOT NULL,
                      table_name TEXT NOT NULL,
                      payload_json TEXT NOT NULL,
                      updated_at TEXT NOT NULL,
                      PRIMARY KEY (entity_code, table_name)
                    )
                    """
                )
                conn.execute(
                    """
                    INSERT INTO org_product_metric_table
                    (entity_code, entity_name, table_id, table_name, payload_json, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (
                        "A01",
                        "泛微粒贷",
                        "table-业务状况表",
                        "业务状况表",
                        json.dumps(
                            {
                                "metrics": [
                                    {
                                        "id": "m-revenue",
                                        "code": "A0101",
                                        "name": "营业收入",
                                        "mapping_status": "MANUAL_CONFIRMED",
                                        "metric_node_code": "A01.01.01.NEW",
                                        "data_acct_code": "A01.01.01.NEW",
                                    },
                                    {
                                        "id": "m-fee05",
                                        "code": "A0105",
                                        "name": "直接费用",
                                        "mapping_status": "MANUAL_CONFIRMED",
                                        "metric_node_code": "A01.05",
                                        "data_acct_code": "A01.05",
                                    },
                                ]
                            },
                            ensure_ascii=False,
                        ),
                        "2026-06-05T00:00:00Z",
                    ),
                )
                payload = {
                    "entity_code": "A01",
                    "entity_name": "泛微粒贷",
                    "year": 2026,
                    "table_name": "业务状况表",
                    "metrics": [
                        {
                            "metric_id": "m-revenue",
                            "metric_code": "A0101",
                            "metric_name": "营业收入",
                            "mapping_status": "OLD",
                            "metric_node_code": "OLD_NODE",
                            "data_acct_code": "OLD_ACCT",
                            "values": {"months": {"a1": "1"}},
                        },
                        {
                            "metric_id": "m-fee05",
                            "metric_code": "A0105",
                            "metric_name": "直接费用",
                            "mapping_status": "OLD",
                            "metric_node_code": "OLD_05_NODE",
                            "data_acct_code": "OLD_05_ACCT",
                            "values": {"months": {"a1": "2"}},
                        },
                    ],
                }

                cleaned = _sanitize_data_entry_payload_mapping_refs(conn, payload)

        rows = cleaned["metrics"]

        _assert_no_legacy_runtime_identity_fields(self, rows[0])
        self.assertEqual(rows[0]["values"]["months"]["a1"], "1")
        _assert_no_legacy_runtime_identity_fields(self, rows[1])
        self.assertEqual(rows[1]["values"]["months"]["a1"], "2")

    def test_data_entry_response_sanitizer_drops_refs_and_preserves_values(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            db_path = Path(tmp) / "common.db"
            with sqlite3.connect(db_path) as conn:
                conn.execute(
                    """
                    CREATE TABLE org_product_metric_table (
                      entity_code TEXT NOT NULL,
                      entity_name TEXT NOT NULL,
                      table_id TEXT NOT NULL,
                      table_name TEXT NOT NULL,
                      payload_json TEXT NOT NULL,
                      updated_at TEXT NOT NULL,
                      PRIMARY KEY (entity_code, table_name)
                    )
                    """
                )
                conn.execute(
                    """
                    INSERT INTO org_product_metric_table
                    (entity_code, entity_name, table_id, table_name, payload_json, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (
                        "A01",
                        "泛微粒贷",
                        "table-业务状况表",
                        "业务状况表",
                        json.dumps(
                            {
                                "metrics": [
                                    {
                                        "id": "m-revenue",
                                        "code": "A0101",
                                        "name": "营业收入",
                                        "mapping_status": "MANUAL_CONFIRMED",
                                        "metric_node_code": "A01.01.01.LATEST",
                                        "data_acct_code": "A01.01.01.LATEST",
                                    }
                                ]
                            },
                            ensure_ascii=False,
                        ),
                        "2026-06-05T00:00:00Z",
                    ),
                )
                payload = {
                    "metrics": [
                        {
                            "metric_id": "m-revenue",
                            "metric_code": "A0101",
                            "metric_name": "营业收入",
                            "mapping_status": "OLD",
                            "metric_node_code": "OLD_NODE",
                            "data_acct_code": "OLD_ACCT",
                            "values": {"months": {"a1": "9"}},
                        },
                        {
                            "metric_id": "m-fee05",
                            "metric_code": "A0105",
                            "metric_name": "直接费用",
                            "mapping_status": "MANUAL_CONFIRMED",
                            "metric_node_code": "A01.05",
                            "data_acct_code": "A01.05",
                            "values": {"months": {"a1": "5"}},
                        },
                    ]
                }

                cleaned = _sanitize_data_entry_payload_for_response(
                    conn,
                    payload,
                    entity_code="A01",
                    table_name="业务状况表",
                )

        rows = cleaned["metrics"]

        self.assertEqual(cleaned["entity_code"], "A01")
        self.assertEqual(cleaned["table_name"], "业务状况表")
        _assert_no_legacy_runtime_identity_fields(self, rows[0])
        self.assertEqual(rows[0]["values"]["months"]["a1"], "9")
        _assert_no_legacy_runtime_identity_fields(self, rows[1])
        self.assertEqual(rows[1]["values"]["months"]["a1"], "5")

    def test_data_entry_export_workbook_excludes_legacy_mapping_refs(self) -> None:
        workbook = _build_data_entry_export_workbook(
            {
                "entity_code": "A01",
                "entity_name": "泛微粒贷",
                "year": 2026,
                "month_index": 3,
                "table_name": "业务状况表",
                "metrics": [
                    {
                        "metric_code": "A0101",
                        "metric_name": "营业收入",
                        "levelLabel": "一级",
                        "nature": "收入",
                        "mapping_status": "MANUAL_CONFIRMED",
                        "metric_node_code": "A01.01.01",
                        "data_acct_code": "A01.01.01",
                        "values": {
                            "prev_actual": "10",
                            "prev_budget": "11",
                            "prev_forecast": "12",
                            "months": {"a1": "1", "a2": "2", "a3": "3", "f4": "4"},
                        },
                    },
                    {
                        "metric_code": "A0105",
                        "metric_name": "直接费用",
                        "levelLabel": "一级",
                        "nature": "支出",
                        "mapping_status": "MANUAL_CONFIRMED",
                        "metric_node_code": "A01.05",
                        "data_acct_code": "A01.05",
                        "values": {"months": {"a1": "5"}},
                    },
                ],
            }
        )
        ws = workbook.active
        headers = [cell.value for cell in ws[1]]

        self.assertEqual(headers[:7], ["科目层级", "科目性质", "科目代码", "科目名称", "25年实际", "26年预算", "26年预测"])
        self.assertEqual(ws["E2"].value, "10")
        self.assertEqual(ws["F2"].value, "11")
        self.assertEqual(ws["G2"].value, "12")
        self.assertEqual(ws["H2"].value, "1")
        self.assertEqual(ws["K2"].value, "4")
        self.assertEqual(ws["E3"].value, "")
        self.assertEqual(ws["H3"].value, "5")

    def test_output_export_sheet_excludes_legacy_mapping_refs(self) -> None:
        from openpyxl import Workbook

        workbook = Workbook()
        ws = workbook.active

        _append_org_product_output_export_sheet(
            ws,
            [
                {
                    "levelLabel": "一级",
                    "nature": "收入",
                    "code": "A0101",
                    "name": "营业收入",
                    "mapping_status": "MANUAL_CONFIRMED",
                    "metric_node_code": "A01.01.01",
                    "data_acct_code": "A01.01.01",
                    "months": [1, 2],
                    "annual": 3,
                },
                {
                    "levelLabel": "一级",
                    "nature": "支出",
                    "code": "A0105",
                    "name": "直接费用",
                    "mapping_status": "MANUAL_CONFIRMED",
                    "metric_node_code": "A01.05",
                    "data_acct_code": "A01.05",
                    "months": [],
                    "annual": None,
                },
            ],
            "A01",
        )

        headers = [cell.value for cell in ws[1]]

        self.assertEqual(headers[:7], ["科目层级", "科目性质", "科目代码", "科目名称", "1月", "2月", "3月"])
        self.assertEqual(ws["E2"].value, 1)
        self.assertEqual(ws["F2"].value, 2)
        self.assertEqual(ws["Q2"].value, 3)
        self.assertEqual(ws["E3"].value, "")
        self.assertEqual(ws["Q3"].value, "")

    def test_output_snapshot_sanitizer_drops_legacy_refs(self) -> None:
        sanitized = _sanitize_org_product_output_entity_for_snapshot(
            {
                "entity_code": "A01",
                "entity_name": "泛微粒贷",
                "table_name": "业务状况表",
                "rows": [
                    {
                        "code": "A0101",
                        "name": "营业收入",
                        "mapping_status": "MANUAL_CONFIRMED",
                        "metric_node_code": "A01.01.01",
                        "data_acct_code": "A01.01.01",
                    },
                    {
                        "code": "A0105",
                        "name": "直接费用",
                        "mapping_status": "MANUAL_CONFIRMED",
                        "metric_node_code": "A01.05",
                        "data_acct_code": "A01.05",
                    },
                ],
            }
        )

        rows = sanitized["rows"]

        _assert_no_legacy_runtime_identity_fields(self, rows[0])
        _assert_no_legacy_runtime_identity_fields(self, rows[1])

    def test_output_response_sanitizer_drops_legacy_refs(self) -> None:
        sanitized = _sanitize_output_payload_for_response(
            {
                "rows": [
                    {
                        "code": "A0105",
                        "name": "直接费用",
                        "mapping_status": "MANUAL_CONFIRMED",
                        "metric_node_code": "A01.05",
                        "data_acct_code": "A01.05",
                    }
                ]
            },
            entity_code="A01",
            table_name="业务状况表",
        )

        self.assertEqual(sanitized["entity_code"], "A01")
        self.assertEqual(sanitized["table_name"], "业务状况表")
        _assert_no_legacy_runtime_identity_fields(self, sanitized["rows"][0])


if __name__ == "__main__":
    unittest.main()
