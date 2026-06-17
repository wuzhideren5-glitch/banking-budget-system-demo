from __future__ import annotations

import asyncio
import sqlite3
import tempfile
import unittest
from pathlib import Path
from typing import Any

from fastapi import HTTPException
from openpyxl import Workbook, load_workbook

from app.schemas import DeptAccountCreate, DeptAccountUpdate
from app.services.dept_catalog import (
    apply_dept_account_import,
    build_dept_tree_export_workbook,
    create_dept_account,
    delete_dept_account,
    list_dept_accounts,
    preview_dept_account_import,
    update_dept_account,
)


SCHEMA = """
CREATE TABLE dept_account (
  dept_code TEXT PRIMARY KEY,
  dept_name TEXT NOT NULL,
  entity_name TEXT NOT NULL DEFAULT '微众银行',
  parent_code TEXT,
  level INTEGER NOT NULL,
  is_leaf INTEGER NOT NULL DEFAULT 0
);
"""


def _create_db(path: Path) -> None:
    with sqlite3.connect(path) as conn:
        conn.executescript(SCHEMA)
        conn.executescript(
            """
            INSERT INTO dept_account(dept_code, dept_name, entity_name, parent_code, level, is_leaf)
            VALUES
              ('Y1', '个人金融事业群', '微众银行', NULL, 1, 0),
              ('Y101', '零售费用归属部门', '微众银行', 'Y1', 2, 1);
            """
        )
        conn.commit()


def _normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _color_row(_ws, _row_idx: int, _max_col: int, _color: str) -> None:
    return None


def _validate_dept_code_with_parent(code: str, level: int, parent_code: str | None) -> str | None:
    if level == 2 and parent_code and not code.startswith(parent_code):
        return "费用归属部门代码必须以前置事业群代码开头"
    return None


def _dept_import_workbook_bytes(rows: list[list[str]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "数据模版"
    ws.append(["主体", "事业群代码", "事业群名称", "费用归属部门代码", "费用归属部门名称"])
    for row in rows:
        ws.append(row)
    with tempfile.NamedTemporaryFile(suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        return Path(tmp.name).read_bytes()


class DeptCatalogServiceTests(unittest.TestCase):
    def test_previews_department_import_workbook(self) -> None:
        async def run() -> tuple[list[str], int, dict[str, str]]:
            content = _dept_import_workbook_bytes(
                [
                    ["微众银行", "Y3", "测试事业群", "", ""],
                    ["", "", "", "Y301", "测试费用部"],
                    ["", "", "", "", ""],
                ]
            )
            preview = await preview_dept_account_import(content, normalize_cell=_normalize_cell)
            return preview.columns, preview.row_count, preview.preview_rows[1]

        columns, row_count, second_row = asyncio.run(run())
        self.assertEqual(columns, ["主体", "事业群代码", "事业群名称", "费用归属部门代码", "费用归属部门名称"])
        self.assertEqual(row_count, 2)
        self.assertEqual(second_row["费用归属部门代码"], "Y301")

    def test_lists_and_creates_department_accounts(self) -> None:
        async def run() -> None:
            with tempfile.TemporaryDirectory() as tmp:
                db_path = Path(tmp) / "common.db"
                _create_db(db_path)

                created = await create_dept_account(
                    db_path,
                    DeptAccountCreate(
                        dept_code="Y2",
                        dept_name="企业金融事业群",
                        entity_name="微众银行",
                        level=1,
                        is_leaf=False,
                    ),
                )
                rows = await list_dept_accounts(db_path)

            self.assertEqual(created.dept_code, "Y2")
            self.assertEqual([row.dept_code for row in rows], ["Y1", "Y101", "Y2"])

        asyncio.run(run())

    def test_update_triggers_expense_rename_sync(self) -> None:
        sync_calls: list[dict[str, object]] = []

        async def fake_sync(_db, *, dept_level: int, old_name: str, new_name: str) -> dict[str, int]:
            sync_calls.append({"dept_level": dept_level, "old_name": old_name, "new_name": new_name})
            return {"expense_forecast_entry.scope_value[owner]": 2}

        async def run() -> None:
            with tempfile.TemporaryDirectory() as tmp:
                db_path = Path(tmp) / "common.db"
                _create_db(db_path)

                result = await update_dept_account(
                    db_path,
                    "Y101",
                    DeptAccountUpdate(dept_name="零售费用部"),
                    rename_sync=fake_sync,
                )
                rows = await list_dept_accounts(db_path)

            self.assertEqual(result.row.dept_name, "零售费用部")
            self.assertEqual(result.after_data["synced_related_updates"], {"expense_forecast_entry.scope_value[owner]": 2})
            self.assertEqual(sync_calls, [{"dept_level": 2, "old_name": "零售费用归属部门", "new_name": "零售费用部"}])
            self.assertEqual({row.dept_code: row.dept_name for row in rows}["Y101"], "零售费用部")

        asyncio.run(run())

    def test_delete_blocks_parent_and_removes_leaf(self) -> None:
        async def run() -> None:
            with tempfile.TemporaryDirectory() as tmp:
                db_path = Path(tmp) / "common.db"
                _create_db(db_path)

                with self.assertRaises(HTTPException) as blocked:
                    await delete_dept_account(db_path, "Y1")
                deleted = await delete_dept_account(db_path, "Y101")
                rows = await list_dept_accounts(db_path)

            self.assertEqual(blocked.exception.status_code, 409)
            self.assertEqual(deleted.before_data["dept_code"], "Y101")
            self.assertEqual([row.dept_code for row in rows], ["Y1"])

        asyncio.run(run())

    def test_builds_dept_tree_export_workbook(self) -> None:
        async def run() -> tuple[str, str, str, str, str]:
            with tempfile.TemporaryDirectory() as tmp:
                db_path = Path(tmp) / "common.db"
                _create_db(db_path)

                workbook = await build_dept_tree_export_workbook(db_path)
                out = Path(tmp) / workbook.filename
                out.write_bytes(workbook.content)
                wb = load_workbook(out)
                ws = wb.active
                return ws.title, ws["A1"].value, ws["A2"].value, ws["B3"].value, ws["D4"].value

        self.assertEqual(asyncio.run(run()), ("数据模版", "主体", "微众银行", "Y1", "Y101"))

    def test_applies_department_import_with_sparse_rows_and_failure_reason(self) -> None:
        async def run() -> tuple[int, int, int, dict[str, str], str | None]:
            with tempfile.TemporaryDirectory() as tmp:
                db_path = Path(tmp) / "common.db"
                _create_db(db_path)
                content = _dept_import_workbook_bytes(
                    [
                        ["微众银行", "Y3", "测试事业群", "", ""],
                        ["", "", "", "Y301", "测试费用部"],
                        ["", "", "", "Z999", "错误部门"],
                    ]
                )
                result = await apply_dept_account_import(
                    db_path,
                    content,
                    {
                        "entityName": "主体",
                        "level1Code": "事业群代码",
                        "level1Name": "事业群名称",
                        "level2Code": "费用归属部门代码",
                        "level2Name": "费用归属部门名称",
                    },
                    normalize_cell=_normalize_cell,
                    color_row=_color_row,
                    validate_dept_code_with_parent=_validate_dept_code_with_parent,
                )
                rows = await list_dept_accounts(db_path)
                out = Path(tmp) / result.filename
                out.write_bytes(result.content)
                wb = load_workbook(out)
                ws = wb["数据模版"]
                failure_reason = ws.cell(row=4, column=6).value

            return (
                result.success,
                result.failed,
                result.overwrite,
                {row.dept_code: row.dept_name for row in rows},
                failure_reason,
            )

        success, failed, overwrite, rows_by_code, failure_reason = asyncio.run(run())
        self.assertEqual((success, failed, overwrite), (2, 1, 0))
        self.assertEqual(rows_by_code["Y3"], "测试事业群")
        self.assertEqual(rows_by_code["Y301"], "测试费用部")
        self.assertIn("费用归属部门代码格式错误", str(failure_reason))


if __name__ == "__main__":
    unittest.main()
