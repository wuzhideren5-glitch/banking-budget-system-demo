from __future__ import annotations

import asyncio
import importlib
from pathlib import Path
from types import SimpleNamespace
import unittest

from openpyxl import load_workbook

from app.services.expense_forecast_export import (
    build_expense_forecast_export_workbook,
    build_expense_forecast_group_export_workbook,
)


def month_cell(value: float, *, source: str = "forecast", editable: bool = False) -> SimpleNamespace:
    return SimpleNamespace(value=value, source=source, editable=editable)


def month_cells(values: list[float], *, actual_cutoff_month: int = 0) -> list[SimpleNamespace]:
    return [
        month_cell(value, source="actual" if month <= actual_cutoff_month else "forecast")
        for month, value in enumerate(values, start=1)
    ]


class FakeExportWorkflowSource:
    def __init__(self) -> None:
        self.scope_view_requests: list[dict[str, object]] = []
        self.subject_view_requests: list[dict[str, object]] = []
        self.group_options_requests = 0

    async def build_scope_view(self, *, year: int, forecast_version: str, scope_type: str, scope_value: str) -> SimpleNamespace:
        self.scope_view_requests.append(
            {
                "year": year,
                "forecast_version": forecast_version,
                "scope_type": scope_type,
                "scope_value": scope_value,
            }
        )
        return SimpleNamespace(actual_cutoff_month=1, rows=[])

    async def build_subject_view(
        self,
        *,
        year: int,
        forecast_version: str,
        scope_type: str,
        scope_value: str,
        subject_id: int,
    ) -> SimpleNamespace:
        self.subject_view_requests.append(
            {
                "year": year,
                "forecast_version": forecast_version,
                "scope_type": scope_type,
                "scope_value": scope_value,
                "subject_id": subject_id,
            }
        )
        return SimpleNamespace(actual_cutoff_month=2, subject_name="差旅费", rows=[])

    async def load_owner_group_options(self) -> list[SimpleNamespace]:
        self.group_options_requests += 1
        return [
            SimpleNamespace(
                group_value="事业群A",
                owner_options=[SimpleNamespace(value="部门B"), SimpleNamespace(value="部门A")],
            )
        ]


class ExpenseForecastExportWorkflowTests(unittest.TestCase):
    def export_workflow_module(self):
        try:
            return importlib.import_module("app.services.expense_forecast_export_workflow")
        except ImportError as exc:
            self.fail(f"expense forecast export workflow module is missing: {exc}")

    def test_regular_export_workflow_plans_scope_request_and_builds_workbook(self) -> None:
        source = FakeExportWorkflowSource()
        export_workflow_module = self.export_workflow_module()

        result = asyncio.run(
            export_workflow_module.build_expense_forecast_export_from_source(
                year=2026,
                forecast_version="",
                default_version="260519v1",
                scope_type="owner",
                scope_value=" 部门A ",
                compile_mode="unknown",
                subject_id=None,
                amount_unit="yuan",
                exclude_fields=[],
                source=source,
            )
        )

        self.assertEqual(result.display_file_name, "费用预测表_2026_260519v1.xlsx")
        self.assertGreater(len(result.stream.getvalue()), 0)
        self.assertEqual(
            source.scope_view_requests,
            [
                {
                    "year": 2026,
                    "forecast_version": "260519v1",
                    "scope_type": "owner",
                    "scope_value": "部门A",
                }
            ],
        )
        self.assertEqual(source.subject_view_requests, [])

    def test_regular_export_workflow_plans_subject_request_and_builds_workbook(self) -> None:
        source = FakeExportWorkflowSource()
        export_workflow_module = self.export_workflow_module()

        result = asyncio.run(
            export_workflow_module.build_expense_forecast_export_from_source(
                year=2026,
                forecast_version="V1",
                default_version="260519v1",
                scope_type="owner",
                scope_value="部门A",
                compile_mode="subject",
                subject_id=11,
                amount_unit="ten_thousand",
                exclude_fields=[],
                source=source,
            )
        )

        self.assertEqual(result.display_file_name, "费用预测表_按预算科目_2026_V1_差旅费.xlsx")
        self.assertGreater(len(result.stream.getvalue()), 0)
        self.assertEqual(source.scope_view_requests, [])
        self.assertEqual(
            source.subject_view_requests,
            [
                {
                    "year": 2026,
                    "forecast_version": "V1",
                    "scope_type": "owner",
                    "scope_value": "部门A",
                    "subject_id": 11,
                }
            ],
        )

    def test_group_export_workflow_loads_group_options_and_owner_views(self) -> None:
        source = FakeExportWorkflowSource()
        export_workflow_module = self.export_workflow_module()

        result = asyncio.run(
            export_workflow_module.build_expense_forecast_group_export_from_source(
                year=2026,
                forecast_version="",
                default_version="260519v1",
                group_name="事业群A",
                amount_unit="yuan",
                exclude_fields=[],
                source=source,
            )
        )

        self.assertEqual(result.display_file_name, "费用预测表_2026_事业群A.xlsx")
        self.assertGreater(len(result.stream.getvalue()), 0)
        self.assertEqual(source.group_options_requests, 1)
        self.assertEqual(
            source.scope_view_requests,
            [
                {
                    "year": 2026,
                    "forecast_version": "260519v1",
                    "scope_type": "owner",
                    "scope_value": "部门A",
                },
                {
                    "year": 2026,
                    "forecast_version": "260519v1",
                    "scope_type": "owner",
                    "scope_value": "部门B",
                },
            ],
        )

    def test_regular_and_group_export_workflows_have_separate_source_interfaces(self) -> None:
        export_workflow_module = self.export_workflow_module()
        workflow_source = Path(export_workflow_module.__file__).read_text(encoding="utf-8")

        self.assertIn("class ExpenseForecastRegularExportSource(Protocol):", workflow_source)
        self.assertIn("class ExpenseForecastGroupExportSource(Protocol):", workflow_source)
        regular_source = workflow_source.split("class ExpenseForecastRegularExportSource(Protocol):", 1)[1].split(
            "class ExpenseForecastGroupExportSource(Protocol):",
            1,
        )[0]
        group_source = workflow_source.split("class ExpenseForecastGroupExportSource(Protocol):", 1)[1].split(
            "async def build_expense_forecast_export_from_source(",
            1,
        )[0]

        self.assertIn("build_scope_view", regular_source)
        self.assertIn("build_subject_view", regular_source)
        self.assertNotIn("load_owner_group_options", regular_source)
        self.assertIn("build_scope_view", group_source)
        self.assertIn("load_owner_group_options", group_source)
        self.assertNotIn("build_subject_view", group_source)


class ExpenseForecastExportTests(unittest.TestCase):
    def test_scope_export_writes_tree_formulas_and_unit_metadata(self) -> None:
        parent = SimpleNamespace(
            id=1,
            parent_id=None,
            level_number=1,
            subject_name="费用合计",
            formula_text=None,
            is_leaf=False,
            months=month_cells([0.0] * 12, actual_cutoff_month=1),
            total_value=0.0,
            annual_budget=0.0,
            business_submission=0.0,
            capital_advice=0.0,
            business_submission_editable=False,
            capital_advice_editable=False,
        )
        child_months = month_cells(
            [100.0, 200.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
            actual_cutoff_month=1,
        )
        child_months[1].editable = True
        child = SimpleNamespace(
            id=2,
            parent_id=1,
            level_number=2,
            subject_name="差旅费",
            formula_text=None,
            is_leaf=True,
            months=child_months,
            total_value=300.0,
            annual_budget=1_000.0,
            business_submission=320.0,
            capital_advice=330.0,
            business_submission_editable=True,
            capital_advice_editable=True,
        )
        view = SimpleNamespace(actual_cutoff_month=1, rows=[parent, child])

        stream, filename = build_expense_forecast_export_workbook(
            year=2026,
            forecast_version="V1",
            scope_type="owner",
            scope_value="部门A",
            compile_mode="scope",
            amount_unit="yuan",
            exclude_fields=[],
            view=view,
        )

        wb = load_workbook(stream, data_only=False)
        ws = wb["费用预测表"]

        self.assertEqual(filename, "费用预测表_2026_V1.xlsx")
        self.assertIn("单位：元", ws[2][6].value)
        self.assertEqual(ws["A3"].value, "预算科目")
        self.assertEqual(ws["B4"].value, "=B5")
        self.assertEqual(ws["N4"].value, "=N5")
        self.assertEqual(ws["O4"].value, "=O5")
        self.assertEqual(ws["B5"].value, 100.0)
        self.assertIsNone(ws["C5"].value)
        self.assertEqual(ws["N5"].value, "=B5+C5+D5+E5+F5+G5+H5+I5+J5+K5+L5+M5")
        self.assertEqual(ws["O5"].value, 1000.0)
        self.assertFalse(ws["C5"].protection.locked)
        self.assertTrue(ws["B5"].protection.locked)

    def test_subject_export_uses_owner_rows_and_subject_filename(self) -> None:
        owner_months = month_cells(
            [10_000.0, 20_000.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
            actual_cutoff_month=2,
        )
        row = SimpleNamespace(
            owner_name="部门A",
            months=owner_months,
            annual_budget=100_000.0,
            business_submission=30_000.0,
            capital_advice=28_000.0,
            business_submission_editable=False,
            capital_advice_editable=True,
        )
        subject_view = SimpleNamespace(actual_cutoff_month=2, subject_name="差旅费", rows=[row])

        stream, filename = build_expense_forecast_export_workbook(
            year=2026,
            forecast_version="V1",
            scope_type="owner",
            scope_value="部门A",
            compile_mode="subject",
            amount_unit="ten_thousand",
            exclude_fields=[],
            subject_view=subject_view,
        )

        wb = load_workbook(stream, data_only=False)
        ws = wb["费用预测表"]

        self.assertEqual(filename, "费用预测表_按预算科目_2026_V1_差旅费.xlsx")
        self.assertIn("预算科目：差旅费", ws[2][5].value)
        self.assertIn("单位：万元", ws[2][7].value)
        self.assertEqual(ws["A3"].value, "费用归属部门")
        self.assertEqual(ws["A4"].value, "部门A")
        self.assertEqual(ws["B4"].value, 1.0)
        self.assertEqual(ws["C4"].value, 2.0)
        self.assertEqual(ws["N4"].value, "=B4+C4+D4+E4+F4+G4+H4+I4+J4+K4+L4+M4")
        self.assertEqual(ws["O4"].value, 10.0)
        self.assertTrue(ws["R4"].protection.locked)
        self.assertFalse(ws["S4"].protection.locked)

    def test_group_export_aggregates_owner_sections_with_formulas(self) -> None:
        def owner_view(owner_value: float) -> SimpleNamespace:
            parent = SimpleNamespace(
                id=1,
                parent_id=None,
                level_number=1,
                subject_name="费用合计",
                formula_text=None,
                is_leaf=False,
                months=month_cells([0.0] * 12, actual_cutoff_month=1),
                total_value=0.0,
                annual_budget=0.0,
                business_submission=0.0,
                capital_advice=0.0,
                business_submission_editable=False,
                capital_advice_editable=False,
            )
            child = SimpleNamespace(
                id=2,
                parent_id=1,
                level_number=2,
                subject_name="差旅费",
                formula_text=None,
                is_leaf=True,
                months=month_cells([owner_value, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0], actual_cutoff_month=1),
                total_value=owner_value,
                annual_budget=owner_value * 10,
                business_submission=owner_value * 2,
                capital_advice=owner_value * 3,
                business_submission_editable=False,
                capital_advice_editable=True,
            )
            return SimpleNamespace(actual_cutoff_month=1, rows=[parent, child])

        stream, filename = build_expense_forecast_group_export_workbook(
            year=2026,
            forecast_version="V1",
            group_name="事业群A",
            amount_unit="yuan",
            exclude_fields=[],
            actual_cutoff=1,
            owner_names=["部门A", "部门B"],
            owner_views={"部门A": owner_view(100.0), "部门B": owner_view(200.0)},
        )

        wb = load_workbook(stream, data_only=False)
        ws = wb["费用预测表-事业群A"]

        self.assertEqual(filename, "费用预测表_2026_事业群A.xlsx")
        self.assertEqual(ws["A1"].value, "费用预测表（按事业群导出）")
        self.assertEqual(ws["B2"].value, "版本：V1")
        self.assertEqual(ws["A4"].value, "事业群A")
        self.assertEqual(ws["A7"].value, "  部门A")
        self.assertEqual(ws["A10"].value, "  部门B")
        self.assertEqual(ws["B4"].value, "=B5")
        self.assertEqual(ws["B5"].value, "=B8+B11")
        self.assertEqual(ws["B6"].value, "=B9+B12")
        self.assertEqual(ws["B8"].value, "=B9")
        self.assertEqual(ws["B11"].value, "=B12")
        self.assertEqual(ws["B9"].value, 100.0)
        self.assertEqual(ws["B12"].value, 200.0)
        self.assertEqual(ws["N4"].value, "=N5")
        self.assertEqual(ws["N5"].value, "=N8+N11")
        self.assertEqual(ws["N9"].value, "=B9+C9+D9+E9+F9+G9+H9+I9+J9+K9+L9+M9")
        self.assertTrue(ws["B4"].protection.locked)


if __name__ == "__main__":
    unittest.main()
