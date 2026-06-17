from __future__ import annotations

import unittest
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
from typing import Any

from fastapi import APIRouter, FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.routers import expense_forecast_rules as expense_forecast_rules_module
from app.routers.expense_forecast_rules import ExpenseForecastRuleSaveRequest
from app.services.export_common import excel_streaming_response
from app.services.expense_forecast_rule_import import build_expense_forecast_rule_template_workbook


async def _empty_rows(*_args, **_kwargs) -> list[dict[str, Any]]:
    return []


async def _list_rules(*_args, **_kwargs) -> list[dict[str, Any]]:
    return []


async def _none(*_args, **_kwargs):
    return None


async def _simulate_rule(*_args, **_kwargs) -> dict[str, Any]:
    return {"scheme_code": "MANUAL", "months": []}


async def _copy_rules(*_args, **_kwargs) -> int:
    return 0


async def _load_rule_detail(*_args, **_kwargs) -> dict[str, Any] | None:
    return None


async def _preview_rule_import(*_args, **_kwargs):
    return SimpleNamespace(
        preview_count=0,
        insertable_rules=0,
        updatable_rules=0,
        skipped_rules=0,
        error_rules=0,
        items=[],
    )


async def _apply_rule_import(*_args, **_kwargs):
    return SimpleNamespace(
        inserted_rules=0,
        updated_rules=0,
        skipped_rules=0,
        error_rules=0,
    )


def _download_rule_template():
    return excel_streaming_response(
        build_expense_forecast_rule_template_workbook(
            default_year=2026,
            default_version="baseline",
        ),
        filename="费用预测逻辑配置模板.xlsx",
        fallback_filename="expense-forecast-rule-template.xlsx",
    )


class ExpenseForecastRuleRouterTests(unittest.TestCase):
    def test_template_uses_common_excel_download_contract(self) -> None:
        router = APIRouter()
        expense_forecast_rules_module.register_expense_forecast_rule_routes(
            router,
            default_year=2026,
            list_rules=_list_rules,
            load_rule_detail=_load_rule_detail,
            save_rule=_none,
            delete_rule=_none,
            recalculate_rules=_none,
            simulate_rule=_simulate_rule,
            copy_rules=_copy_rules,
            download_rule_template=_download_rule_template,
            preview_rule_import=_preview_rule_import,
            apply_rule_import=_apply_rule_import,
        )
        app = FastAPI()
        app.include_router(router)

        response = TestClient(app).get("/api/expense-forecast/rules/template")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.headers["content-type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        disposition = response.headers["content-disposition"]
        self.assertIn("filename=expense-forecast-rule-template.xlsx", disposition)
        self.assertIn("filename*=UTF-8''%E8%B4%B9%E7%94%A8%E9%A2%84%E6%B5%8B", disposition)
        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(workbook["预测规则模板"].cell(2, 1).value, 2026)
        self.assertEqual(workbook["预测规则模板"].cell(2, 2).value, "baseline")

    def test_rule_payload_accepts_org_product_metric_variable_for_parent_resolution(self) -> None:
        body = ExpenseForecastRuleSaveRequest.model_validate(
            {
                "forecast_year": 2026,
                "forecast_version": "V1",
                "owner_name": "部门A",
                "subject_id": 11,
                "scheme_code": "METRIC_EXPR",
                "params": [
                    {
                        "param_group": "metric_expr",
                        "param_key": "expression",
                        "param_value": "base_amount",
                        "value_type": "string",
                    }
                ],
                "variables": [
                    {
                        "variable_code": "base_amount",
                        "source_type": "org_product_metric",
                        "source_key": "A0111",
                        "source_subkey": "A01",
                        "org_product_ref": "A01:业务状况表:A0111",
                        "sort_order": 1,
                    }
                ],
            }
        )

        variable = body.variables[0]
        self.assertEqual(variable.source_type, "org_product_metric")
        self.assertEqual(variable.org_product_ref, "A01:业务状况表:A0111")

    def test_router_does_not_hand_roll_rule_template_download_response(self) -> None:
        router_source = Path(expense_forecast_rules_module.__file__).read_text(encoding="utf-8")

        self.assertNotIn("build_expense_forecast_rule_template_workbook", router_source)
        self.assertNotIn("excel_streaming_response", router_source)
        self.assertNotIn("default_version", router_source)
        self.assertNotIn("StreamingResponse", router_source)
        self.assertNotIn("filename*=", router_source)
        self.assertNotIn("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", router_source)
        self.assertIn("download_rule_template", router_source)

    def test_router_delegates_rule_import_workflows(self) -> None:
        router_source = Path(expense_forecast_rules_module.__file__).read_text(encoding="utf-8")

        self.assertNotIn("parse_expense_forecast_rule_import_workbook(", router_source)
        self.assertNotIn("preview_expense_forecast_rule_import_rows(", router_source)
        self.assertNotIn("apply_expense_forecast_rule_import_rows(", router_source)
        self.assertNotIn("preview_expense_forecast_rule_import_workbook", router_source)
        self.assertNotIn("apply_expense_forecast_rule_import_workbook", router_source)
        self.assertNotIn("_ExpenseForecastRuleImportSource", router_source)
        self.assertNotIn("subject_lookup", router_source)
        self.assertIn("_run_rule_import_upload_workflow(file, preview_rule_import)", router_source)
        self.assertIn("_run_rule_import_upload_workflow(file, apply_rule_import)", router_source)

    def test_router_uses_shared_rule_import_upload_workflow(self) -> None:
        router_source = Path(expense_forecast_rules_module.__file__).read_text(encoding="utf-8")

        self.assertEqual(router_source.count("raw = await file.read()"), 1)
        self.assertEqual(router_source.count("except ValueError as exc"), 1)
        self.assertIn("_run_rule_import_upload_workflow", router_source)

    def test_router_delegates_rule_delete_workflow(self) -> None:
        router_source = Path(expense_forecast_rules_module.__file__).read_text(encoding="utf-8")

        self.assertNotIn("delete_expense_forecast_rule_definition(", router_source)
        self.assertNotIn("if not deleted", router_source)
        self.assertIn("await delete_rule(rule_id=rule_id)", router_source)

    def test_router_delegates_rule_save_audit_workflow(self) -> None:
        router_source = Path(expense_forecast_rules_module.__file__).read_text(encoding="utf-8")

        self.assertNotIn("write_operation_log", router_source)
        self.assertNotIn("新增费用预测规则", router_source)
        self.assertNotIn("更新费用预测规则", router_source)

    def test_router_registration_does_not_keep_unused_numeric_adapter(self) -> None:
        router_source = Path(expense_forecast_rules_module.__file__).read_text(encoding="utf-8")

        self.assertNotIn("safe_int", router_source)

    def test_router_uses_runtime_adapters_for_current_database_access(self) -> None:
        router_source = Path(expense_forecast_rules_module.__file__).read_text(encoding="utf-8")

        self.assertNotIn("common_db_path", router_source)
        self.assertNotIn("load_expense_forecast_rule_identity", router_source)
        self.assertNotIn("delete_expense_forecast_rule_definition_or_raise", router_source)
        self.assertNotIn("load_rule_rows", router_source)
        self.assertNotIn("load_rule_identity", router_source)

    def test_router_delegates_rule_list_runtime_context(self) -> None:
        router_source = Path(expense_forecast_rules_module.__file__).read_text(encoding="utf-8")

        self.assertNotIn("owner_names", router_source)
        self.assertIn("await list_rules(", router_source)

    def test_router_delegates_text_normalization_runtime_context(self) -> None:
        router_source = Path(expense_forecast_rules_module.__file__).read_text(encoding="utf-8")

        self.assertNotIn("text:", router_source)
        self.assertNotIn("text(", router_source)

    def test_router_delegates_rule_detail_workflow_runtime_context(self) -> None:
        router_source = Path(expense_forecast_rules_module.__file__).read_text(encoding="utf-8")

        self.assertNotIn("load_expense_forecast_rule_detail", router_source)
        self.assertNotIn("_ExpenseForecastRuleDetailSource", router_source)
        self.assertIn("await load_rule_detail(", router_source)

    def test_router_delegates_rule_simulation_runtime_context(self) -> None:
        router_source = Path(expense_forecast_rules_module.__file__).read_text(encoding="utf-8")

        self.assertNotIn("actual_cutoff_month", router_source)
        self.assertNotIn("load_actual_map", router_source)
        self.assertNotIn("load_annual_input_map", router_source)
        self.assertNotIn("load_forecast_map", router_source)
        self.assertNotIn("calculate_rule_months", router_source)
        self.assertNotIn("simulate_expense_forecast_rule_definition", router_source)
        self.assertIn("await simulate_rule(", router_source)

    def test_router_delegates_rule_copy_workflow_runtime_context(self) -> None:
        router_source = Path(expense_forecast_rules_module.__file__).read_text(encoding="utf-8")

        self.assertNotIn("copy_expense_forecast_rule_definitions", router_source)
        self.assertNotIn("_ExpenseForecastRuleCopySource", router_source)
        self.assertIn("await copy_rules(", router_source)


if __name__ == "__main__":
    unittest.main()
