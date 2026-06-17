from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

from app.services.expense_forecast_rule_import import (
    build_expense_forecast_rule_import_payload,
    build_expense_forecast_rule_template_workbook,
    parse_expense_forecast_rule_import_workbook,
)


def _workbook_bytes(headers: list[str], rows: list[list[object]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "预测规则模板"
    ws.append(headers)
    for row in rows:
        ws.append(row)
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def test_build_rule_template_contains_current_metric_expression_guidance() -> None:
    stream = build_expense_forecast_rule_template_workbook(
        default_year=2026,
        default_version="baseline",
        org_product_refs_by_runtime_ref_code={
            "A01.01.01.01.01.017": ("A01:业务状况表:A0111 管理贷款日均",),
        },
    )

    wb = load_workbook(stream)

    assert "预测规则模板" in wb.sheetnames
    assert "填写说明" in wb.sheetnames
    assert "机构产品指标候选" in wb.sheetnames
    assert wb["预测规则模板"].cell(2, 1).value == 2026
    assert wb["预测规则模板"].cell(2, 2).value == "baseline"
    guidance = "\n".join(str(cell.value or "") for cell in wb["填写说明"]["B"])
    assert "source_type=metric_tree" in guidance
    assert "A01.01.01.001" in guidance
    assert "机构及产品指标表" in guidance
    candidate_ws = wb["机构产品指标候选"]
    assert candidate_ws.cell(1, 2).value == "运行取数编码"
    assert candidate_ws.cell(1, 5).value == "机构及产品指标编码"
    assert candidate_ws.cell(2, 2).value == "A01.01.01.01.01.017"
    assert candidate_ws.cell(2, 3).value == "A01"
    assert candidate_ws.cell(2, 4).value == "业务状况表"
    assert candidate_ws.cell(2, 5).value == "A0111"
    assert candidate_ws.cell(2, 6).value == "管理贷款日均"
    assert candidate_ws.cell(2, 7).value == "A01:业务状况表:A0111"
    assert '"source_type": "org_product_metric"' in candidate_ws.cell(2, 8).value
    assert '"source_key": "A0111"' in candidate_ws.cell(2, 8).value
    assert '"source_subkey": "A01"' in candidate_ws.cell(2, 8).value
    assert '"org_product_ref": "A01:业务状况表:A0111"' in candidate_ws.cell(2, 8).value


def test_parse_rule_import_workbook_normalizes_labels_and_payloads() -> None:
    raw = _workbook_bytes(
        [
            "年度",
            "版本",
            "费用归属部门",
            "预算科目",
            "预测逻辑",
            "启用",
            "允许人工覆盖",
            "自动刷新",
            "手动重算",
            "生效开始月",
            "生效结束月",
            "规则优先级",
            "数据源优先级",
            "分摊模式",
            "自动反推方式",
            "是否允许负数",
            "自定义系数JSON",
            "表达式",
            "变量映射JSON",
            "备注",
        ],
        [
            [2026, "baseline", "日常费用部", "云资源费", "余额分摊", "是", "否", "是", "是", 1, 12, 90, "机构及产品指标编码优先", "自定义系数", "等差金额", "否", '{"7": 1}', "", "", "custom"],
            [2026, "baseline", "日常费用部", "短信费", "指标表达式", "是", "是", "是", "是", 3, 10, 80, "表内维护优先", "", "", "否", "", "base_amount * factor", '[{"variable_code":"base_amount","source_type":"metric_tree","source_key":"A01.01.01.001"}]', "expr"],
        ],
    )

    rows = parse_expense_forecast_rule_import_workbook(raw, default_year=2026, default_version="fallback")
    residual_params, residual_variables = build_expense_forecast_rule_import_payload(rows[0])
    expr_params, expr_variables = build_expense_forecast_rule_import_payload(rows[1])

    assert rows[0]["scheme_code"] == "RESIDUAL_ALLOC"
    assert rows[0]["allocation_mode"] == "custom"
    assert rows[1]["scheme_code"] == "METRIC_EXPR"
    assert rows[1]["metric_source_priority"] == "inline_first"
    assert residual_variables == []
    assert {item["param_key"]: item["param_value"] for item in residual_params}["weight_json"] == '{"7": 1}'
    assert {item["param_key"]: item["param_value"] for item in expr_params}["expression"] == "base_amount * factor"
    assert expr_variables == [
        {
            "variable_code": "base_amount",
            "source_type": "metric_tree",
            "source_key": "A01.01.01.001",
        }
    ]


def test_parse_rule_import_workbook_rejects_missing_required_header() -> None:
    raw = _workbook_bytes(["年度", "版本", "费用归属部门", "预测逻辑"], [[2026, "v1", "日常费用部", "余额分摊"]])

    with pytest.raises(ValueError, match="预算科目"):
        parse_expense_forecast_rule_import_workbook(raw, default_year=2026, default_version="baseline")


def test_metric_expression_payload_requires_expression_and_variable_codes() -> None:
    with pytest.raises(ValueError, match="必须填写表达式"):
        build_expense_forecast_rule_import_payload(
            {
                "scheme_code": "METRIC_EXPR",
                "owner_name": "日常费用部",
                "subject_name": "短信费",
                "effective_from_month": 1,
                "effective_to_month": 12,
                "expression": "",
                "variables_json": "[]",
            }
        )

    with pytest.raises(ValueError, match="variable_code"):
        build_expense_forecast_rule_import_payload(
            {
                "scheme_code": "METRIC_EXPR",
                "owner_name": "日常费用部",
                "subject_name": "短信费",
                "effective_from_month": 1,
                "effective_to_month": 12,
                "expression": "base_amount",
                "variables_json": '[{"source_type":"metric_tree"}]',
            }
        )
