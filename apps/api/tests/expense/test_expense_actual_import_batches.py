from __future__ import annotations

import asyncio
import sqlite3
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from app.db_bootstrap.expense import ensure_expense_actual_import_schema_sync
from app.services import expense_actual_import_batches as batch_module
from app.services.expense_actual_import_batches import (
    ExpenseActualImportBatchMissingError,
    ExpenseActualImportExportMissingError,
    delete_expense_actual_import_batch,
    export_expense_actual_import_batch,
    list_expense_actual_import_batches,
    normalize_import_kind,
)


def _seed_batch(db_path: Path) -> int:
    conn = sqlite3.connect(db_path)
    try:
        ensure_expense_actual_import_schema_sync(conn)
        cur = conn.execute(
            """
            INSERT INTO expense_actual_import_batch(
              import_kind, file_name, import_mode, periods_text, total_rows,
              matched_owner_rows, matched_subject_rows, unmatched_rows, created_at, note
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "current_year_actual",
                "actual.xlsx",
                "append",
                "2026-04",
                1,
                1,
                1,
                0,
                "2026-06-02T00:00:00Z",
                "note",
            ),
        )
        batch_id = int(cur.lastrowid)
        conn.execute(
            """
            INSERT INTO expense_actual_detail_raw(
              batch_id, import_kind, data_date, period_ym, period_text, org_code, org_name,
              dep_code, dep_name, subject_code, subject_name, journal_name, serial_no, line_desc,
              amount, fee_type_code, fee_type_name, bi_ai_source_code, bi_ai_source_name,
              manage_department_code, owner_name_raw, owner_name_mapped, monthly_caliber,
              budget_subject_raw, budget_subject_mapped, fee_major_mapped, fee_category_mapped,
              budget_release_caliber_mapped, manage_department2, special_control_tag,
              owner_matched, subject_matched, match_note
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                batch_id,
                "current_year_actual",
                "2026-04-30",
                "2026-04",
                "2026年4月",
                "ORG",
                "费用部门",
                "DEP",
                "责任中心",
                "SUB",
                "科目",
                "日记帐",
                "SN001",
                "说明",
                123.45,
                "F01",
                "费用类别",
                "C01",
                "管控口径",
                "CD01",
                "原始归口",
                "映射归口",
                "月报口径",
                "原始预算科目",
                "映射预算科目",
                "费用大类",
                "费用类别一级",
                "预算发布",
                "归口2",
                "专项",
                1,
                1,
                "",
            ),
        )
        conn.commit()
        return batch_id
    finally:
        conn.close()


class _FakeDeleteCursor:
    def __init__(self, state: dict[str, object]) -> None:
        self._state = state
        self.rowcount = 0
        self._row: tuple[object, ...] | None = None

    async def __aenter__(self) -> "_FakeDeleteCursor":
        return self

    async def __aexit__(self, exc_type: object, exc: object, tb: object) -> None:
        return None

    async def execute(self, sql: str, params: tuple[object, ...]) -> None:
        normalized_sql = " ".join(sql.lower().split())
        self._state.setdefault("executed_sql", []).append((normalized_sql, params))  # type: ignore[union-attr]
        if normalized_sql.startswith("select file_name"):
            self._row = ("mysql_actual.xlsx", 1, "current_year_actual")
            self.rowcount = 1
        elif normalized_sql.startswith("delete from expense_actual_detail_raw"):
            self._row = None
            self.rowcount = 1
        elif normalized_sql.startswith("delete from expense_actual_import_batch"):
            self._row = None
            self.rowcount = 1
        else:
            raise AssertionError(f"Unexpected SQL: {sql}")

    async def fetchone(self) -> tuple[object, ...] | None:
        return self._row


class _FakeDeleteConnection:
    def __init__(self, state: dict[str, object]) -> None:
        self._state = state

    async def begin(self) -> None:
        self._state["began"] = True

    async def commit(self) -> None:
        self._state["committed"] = True

    async def rollback(self) -> None:
        self._state["rolled_back"] = True

    def cursor(self) -> _FakeDeleteCursor:
        return _FakeDeleteCursor(self._state)


class _FakeAcquire:
    def __init__(self, state: dict[str, object]) -> None:
        self._state = state

    async def __aenter__(self) -> _FakeDeleteConnection:
        return _FakeDeleteConnection(self._state)

    async def __aexit__(self, exc_type: object, exc: object, tb: object) -> None:
        return None


class _FakeMysqlPool:
    def __init__(self) -> None:
        self.state: dict[str, object] = {"executed_sql": []}

    async def fetch_all(self, sql: str, params: tuple[object, ...]) -> list[dict[str, object]]:
        normalized_sql = " ".join(sql.lower().split())
        self.state.setdefault("fetch_all", []).append((normalized_sql, params))  # type: ignore[union-attr]
        if "from expense_actual_import_batch" in normalized_sql:
            return [
                {
                    "id": 7,
                    "import_kind": "current_year_actual",
                    "file_name": "mysql_actual.xlsx",
                    "import_mode": "append",
                    "periods_text": "2026-04",
                    "total_rows": 1,
                    "matched_owner_rows": 1,
                    "matched_subject_rows": 1,
                    "unmatched_rows": 0,
                    "created_at": "2026-06-02T00:00:00Z",
                    "note": "mysql",
                }
            ]
        if "from expense_actual_detail_raw" in normalized_sql:
            return [
                {
                    "period_ym": "2026-04",
                    "org_code": "ORG",
                    "org_name": "费用部门",
                    "dep_code": "DEP",
                    "dep_name": "责任中心",
                    "subject_code": "SUB",
                    "subject_name": "科目",
                    "amount": 123.45,
                    "fee_type_code": "F01",
                    "fee_type_name": "费用类别",
                    "bi_ai_source_code": "C01",
                    "bi_ai_source_name": "管控口径",
                    "manage_department_code": "CD01",
                    "owner_name_raw": "原始归口",
                    "owner_name_mapped": "映射归口",
                    "monthly_caliber": "月报口径",
                    "budget_subject_raw": "原始预算科目",
                    "budget_subject_mapped": "映射预算科目",
                    "fee_major_mapped": "费用大类",
                    "fee_category_mapped": "费用类别一级",
                    "budget_release_caliber_mapped": "预算发布",
                    "manage_department2": "归口2",
                    "special_control_tag": "专项",
                    "period_text": "2026年4月",
                    "match_note": "",
                    "journal_name": "日记帐",
                    "serial_no": "SN001",
                    "line_desc": "说明",
                    "data_date": "2026-04-30",
                }
            ]
        raise AssertionError(f"Unexpected SQL: {sql}")

    async def fetch_one(self, sql: str, params: tuple[object, ...]) -> dict[str, object] | None:
        normalized_sql = " ".join(sql.lower().split())
        self.state.setdefault("fetch_one", []).append((normalized_sql, params))  # type: ignore[union-attr]
        if normalized_sql.startswith("select max(id)"):
            return {"max_id": 7}
        if "from expense_actual_import_batch" in normalized_sql:
            return {"file_name": "mysql_actual.xlsx", "import_kind": "current_year_actual"}
        raise AssertionError(f"Unexpected SQL: {sql}")

    def acquire(self) -> _FakeAcquire:
        return _FakeAcquire(self.state)


class ExpenseActualImportBatchServiceTests(unittest.TestCase):
    def test_normalize_import_kind_rejects_unknown_value(self) -> None:
        with self.assertRaises(ValueError):
            normalize_import_kind("legacy_actual")

    def test_list_export_and_delete_batch(self) -> None:
        async def run() -> None:
            with tempfile.TemporaryDirectory() as tmp:
                db_path = Path(tmp) / "common.db"
                batch_id = _seed_batch(db_path)

                batches = await list_expense_actual_import_batches(db_path)
                exported = await export_expense_actual_import_batch(db_path, batch_id=batch_id)
                out_path = Path(tmp) / "export.xlsx"
                out_path.write_bytes(exported.content)
                wb = load_workbook(out_path, data_only=True)
                ws = wb.active
                deleted = await delete_expense_actual_import_batch(db_path, batch_id=batch_id)

                self.assertEqual(len(batches), 1)
                self.assertEqual(batches[0].id, batch_id)
                self.assertEqual(batches[0].periods, ["2026-04"])
                self.assertIn(f"批次{batch_id}", exported.filename)
                self.assertEqual(ws["A1"].value, "数据日期")
                self.assertEqual(ws["O1"].value, "BI-AI源编码")
                self.assertEqual(ws["P1"].value, "BI-AI源名称")
                self.assertEqual(ws["X2"].value, "已匹配")
                self.assertEqual(deleted.deleted_rows, 1)
                with sqlite3.connect(db_path) as conn:
                    detail_count = conn.execute("SELECT COUNT(*) FROM expense_actual_detail_raw").fetchone()[0]
                    batch_count = conn.execute("SELECT COUNT(*) FROM expense_actual_import_batch").fetchone()[0]
                self.assertEqual(detail_count, 0)
                self.assertEqual(batch_count, 0)

        asyncio.run(run())

    def test_export_missing_batch_raises(self) -> None:
        async def run() -> None:
            with tempfile.TemporaryDirectory() as tmp:
                db_path = Path(tmp) / "common.db"
                conn = sqlite3.connect(db_path)
                try:
                    ensure_expense_actual_import_schema_sync(conn)
                    conn.commit()
                finally:
                    conn.close()

                with self.assertRaises(ExpenseActualImportExportMissingError):
                    await export_expense_actual_import_batch(db_path)
                with self.assertRaises(ExpenseActualImportBatchMissingError):
                    await export_expense_actual_import_batch(db_path, batch_id=999)
                with self.assertRaises(ExpenseActualImportBatchMissingError):
                    await delete_expense_actual_import_batch(db_path, batch_id=999)

        asyncio.run(run())

    def test_runtime_common_db_uses_mysql_pool_for_list_export_and_delete(self) -> None:
        async def run() -> None:
            fake_pool = _FakeMysqlPool()
            db_path = Path(batch_module.settings.data_dir) / "common.db"
            with patch.object(batch_module, "get_pool", return_value=fake_pool):
                batches = await list_expense_actual_import_batches(db_path, limit=5)
                exported = await export_expense_actual_import_batch(db_path, batch_id=7)
                deleted = await delete_expense_actual_import_batch(db_path, batch_id=7)

            with tempfile.TemporaryDirectory() as tmp:
                out_path = Path(tmp) / "export.xlsx"
                out_path.write_bytes(exported.content)
                wb = load_workbook(out_path, data_only=True)
                ws = wb.active

                self.assertEqual([batch.id for batch in batches], [7])
                self.assertEqual(batches[0].note, "mysql")
                self.assertIn("批次7", exported.filename)
                self.assertEqual(ws["X2"].value, "已匹配")
                self.assertEqual(deleted.file_name, "mysql_actual.xlsx")
                self.assertEqual(deleted.deleted_rows, 1)
                self.assertTrue(fake_pool.state.get("began"))
                self.assertTrue(fake_pool.state.get("committed"))
                self.assertNotIn("rolled_back", fake_pool.state)

    def test_actual_import_batch_service_does_not_import_aiosqlite(self) -> None:
        source = Path(batch_module.__file__).read_text(encoding="utf-8")
        self.assertNotIn("aiosqlite", source)


if __name__ == "__main__":
    unittest.main()
