from __future__ import annotations

from io import BytesIO
import sqlite3
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from app.core.config import settings
from app.services.expense_budget_execution_framework import (
    ExpenseFrameworkError,
    FrameworkBudgetDepartmentRow,
    FrameworkProductDepartmentRow,
    FrameworkSubjectRow,
    ParsedFramework,
    build_framework_context,
    canonical_owner_name,
    load_framework_context,
    load_framework_from_db,
    parse_framework_source_bytes,
    persist_framework_snapshot,
    read_sync_meta,
)


def _framework_workbook_bytes() -> bytes:
    wb = Workbook()
    ws_budget = wb.active
    ws_budget.title = "预算部门"
    ws_budget.append(["主体", "事业群", "归属部门", "预算部门"])
    ws_budget.append(["微众银行", "个人金融事业群", "A01 零售部", "零售一组"])
    ws_budget.append(["微众银行", "个人金融事业群", "A01 零售部", "零售一组"])
    ws_budget.append(["", "", "", ""])

    ws_product = wb.create_sheet("产品预算科目")
    ws_product.append(["主体", "事业群", "归属部门", "产品部门"])
    ws_product.append(["微众银行", "个人金融事业群", "A01 零售部", "小微产品"])

    ws_subject = wb.create_sheet("部门预算科目")
    ws_subject.append(["层级", "预算科目", "归口部门", "公式"])
    ws_subject.append(["一级", "业务费用", "A01 零售部", "0"])
    ws_subject.append(["二级", "IT费用", "科技部", "业务费用*0.1"])

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


class ExpenseBudgetExecutionFrameworkTests(unittest.IsolatedAsyncioTestCase):
    def test_parse_framework_workbook_dedupes_rows_and_builds_alias_context(self) -> None:
        parsed = parse_framework_source_bytes("framework.xlsx", _framework_workbook_bytes())

        self.assertEqual(len(parsed.budget_departments), 1)
        self.assertEqual(parsed.budget_departments[0].owner_name, "A01 零售部")
        self.assertEqual(len(parsed.product_departments), 1)
        self.assertEqual([subject.budget_subject for subject in parsed.subjects], ["业务费用", "IT费用"])

        ctx = build_framework_context(parsed)
        self.assertEqual(canonical_owner_name("零售部", ctx), "A01 零售部")
        self.assertEqual(ctx.owner_to_entity["A01 零售部"], "微众银行")
        self.assertEqual(ctx.owner_to_group["A01 零售部"], "个人金融事业群")

    async def test_persist_framework_snapshot_filters_excluded_rows_and_records_meta(self) -> None:
        original_data_dir = settings.data_dir
        with tempfile.TemporaryDirectory() as tmp:
            data_dir = Path(tmp)
            settings.data_dir = data_dir
            conn = sqlite3.connect(data_dir / "common.db")
            try:
                conn.executescript(
                    """
                    CREATE TABLE expense_sync_meta (
                      sync_key TEXT PRIMARY KEY NOT NULL,
                      source_file TEXT NOT NULL,
                      source_mtime TEXT,
                      synced_at TEXT NOT NULL,
                      row_count INTEGER NOT NULL DEFAULT 0,
                      note TEXT
                    );
                    CREATE TABLE expense_framework_budget_department (
                      id INTEGER PRIMARY KEY AUTOINCREMENT,
                      entity_name TEXT NOT NULL DEFAULT '',
                      group_name TEXT NOT NULL,
                      owner_name TEXT NOT NULL,
                      budget_department TEXT NOT NULL
                    );
                    CREATE TABLE expense_framework_product_department (
                      id INTEGER PRIMARY KEY AUTOINCREMENT,
                      entity_name TEXT NOT NULL DEFAULT '',
                      group_name TEXT NOT NULL,
                      owner_name TEXT NOT NULL,
                      product_department TEXT NOT NULL
                    );
                    CREATE TABLE expense_framework_subject (
                      budget_subject TEXT PRIMARY KEY NOT NULL,
                      level_label TEXT,
                      manage_department TEXT,
                      formula_text TEXT,
                      sort_order INTEGER NOT NULL DEFAULT 0
                    );
                    """
                )
                conn.commit()
            finally:
                conn.close()

            parsed = ParsedFramework(
                source_file=Path("framework.xlsx"),
                budget_departments=[
                    FrameworkBudgetDepartmentRow("微众银行", "科技及智能事业群", "科技子平台", "科技子平台"),
                    FrameworkBudgetDepartmentRow("微众银行", "个人金融事业群", "零售部", "零售部"),
                ],
                product_departments=[
                    FrameworkProductDepartmentRow("微众银行", "个人金融事业群", "零售部", "小微产品")
                ],
                subjects=[FrameworkSubjectRow("一级", "业务费用", "零售部", "0", 1)],
            )

            try:
                await persist_framework_snapshot(parsed)
                loaded = await load_framework_from_db()
                meta = await read_sync_meta()
            finally:
                settings.data_dir = original_data_dir

        self.assertIsNotNone(loaded)
        assert loaded is not None
        self.assertEqual([row.owner_name for row in loaded.budget_departments], ["零售部"])
        self.assertEqual([row.product_department for row in loaded.product_departments], ["小微产品"])
        self.assertEqual([subject.budget_subject for subject in loaded.subjects], ["业务费用"])
        self.assertEqual(meta["framework_import"]["row_count"], 3)
        self.assertIn("预算部门1行", meta["framework_import"]["note"])

    async def test_load_framework_context_uses_current_master_data(self) -> None:
        original_data_dir = settings.data_dir
        with tempfile.TemporaryDirectory() as tmp:
            data_dir = Path(tmp)
            settings.data_dir = data_dir
            conn = sqlite3.connect(data_dir / "common.db")
            try:
                conn.executescript(
                    """
                    CREATE TABLE dept_account (
                      dept_code TEXT PRIMARY KEY,
                      dept_name TEXT NOT NULL,
                      parent_code TEXT,
                      level INTEGER NOT NULL,
                      entity_name TEXT
                    );
                    CREATE TABLE budget_subject_catalog (
                      id INTEGER PRIMARY KEY,
                      parent_id INTEGER,
                      level_number INTEGER NOT NULL,
                      subject_name TEXT NOT NULL,
                      manage_department TEXT,
                      formula_text TEXT,
                      sort_order INTEGER NOT NULL
                    );
                    INSERT INTO dept_account(dept_code, dept_name, parent_code, level, entity_name)
                    VALUES
                      ('G01', '个人金融事业群', NULL, 1, '微众银行'),
                      ('D01', '零售部', 'G01', 2, '微众银行');
                    INSERT INTO budget_subject_catalog(
                      id, parent_id, level_number, subject_name, manage_department, formula_text, sort_order
                    ) VALUES (1, NULL, 1, '业务费用', '零售部', '0', 1);
                    """
                )
                conn.commit()
            finally:
                conn.close()

            try:
                ctx, source_mode, source_desc, parsed = await load_framework_context()
            finally:
                settings.data_dir = original_data_dir

        self.assertEqual(source_mode, "master")
        self.assertIn("系统主数据", source_desc)
        self.assertEqual([row.owner_name for row in parsed.budget_departments], ["零售部"])
        self.assertEqual([subject.budget_subject for subject in parsed.subjects], ["业务费用"])
        self.assertEqual(ctx.owner_to_group["零售部"], "个人金融事业群")

    async def test_load_framework_context_does_not_fall_back_to_internal_snapshot(self) -> None:
        original_data_dir = settings.data_dir
        with tempfile.TemporaryDirectory() as tmp:
            data_dir = Path(tmp)
            settings.data_dir = data_dir
            conn = sqlite3.connect(data_dir / "common.db")
            try:
                conn.executescript(
                    """
                    CREATE TABLE dept_account (
                      dept_code TEXT PRIMARY KEY,
                      dept_name TEXT NOT NULL,
                      parent_code TEXT,
                      level INTEGER NOT NULL,
                      entity_name TEXT
                    );
                    CREATE TABLE budget_subject_catalog (
                      id INTEGER PRIMARY KEY,
                      parent_id INTEGER,
                      level_number INTEGER NOT NULL,
                      subject_name TEXT NOT NULL,
                      manage_department TEXT,
                      formula_text TEXT,
                      sort_order INTEGER NOT NULL
                    );
                    CREATE TABLE expense_sync_meta (
                      sync_key TEXT PRIMARY KEY NOT NULL,
                      source_file TEXT NOT NULL,
                      source_mtime TEXT,
                      synced_at TEXT NOT NULL,
                      row_count INTEGER NOT NULL DEFAULT 0,
                      note TEXT
                    );
                    CREATE TABLE expense_framework_budget_department (
                      id INTEGER PRIMARY KEY AUTOINCREMENT,
                      entity_name TEXT NOT NULL DEFAULT '',
                      group_name TEXT NOT NULL,
                      owner_name TEXT NOT NULL,
                      budget_department TEXT NOT NULL
                    );
                    CREATE TABLE expense_framework_product_department (
                      id INTEGER PRIMARY KEY AUTOINCREMENT,
                      entity_name TEXT NOT NULL DEFAULT '',
                      group_name TEXT NOT NULL,
                      owner_name TEXT NOT NULL,
                      product_department TEXT NOT NULL
                    );
                    CREATE TABLE expense_framework_subject (
                      budget_subject TEXT PRIMARY KEY NOT NULL,
                      level_label TEXT,
                      manage_department TEXT,
                      formula_text TEXT,
                      sort_order INTEGER NOT NULL DEFAULT 0
                    );
                    INSERT INTO expense_framework_budget_department(
                      entity_name, group_name, owner_name, budget_department
                    ) VALUES ('微众银行', '个人金融事业群', '零售部', '零售部');
                    INSERT INTO expense_framework_subject(
                      budget_subject, level_label, manage_department, formula_text, sort_order
                    ) VALUES ('业务费用', '一级', '零售部', '0', 1);
                    INSERT INTO expense_sync_meta(sync_key, source_file, synced_at, row_count)
                    VALUES ('framework_import', 'framework.xlsx', '2026-06-03T00:00:00Z', 2);
                    """
                )
                conn.commit()
            finally:
                conn.close()

            try:
                with self.assertRaisesRegex(ExpenseFrameworkError, "当前部门费用主数据"):
                    await load_framework_context()
            finally:
                settings.data_dir = original_data_dir


if __name__ == "__main__":
    unittest.main()
