from __future__ import annotations

from io import BytesIO
import unittest
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from app.services.expense_budget_execution_export import (
    ExpenseBudgetExecutionExportError,
    ExpenseBudgetExecutionExportOptions,
    ExpenseBudgetExecutionExportPlan,
    ExpenseBudgetExecutionWorkbookExport,
    build_expense_budget_execution_export,
    build_expense_budget_execution_workbook,
    expense_budget_execution_workbook_response,
    resolve_expense_budget_execution_export_plan,
)
from app.services import expense_budget_execution_export as export_module


async def _read_response_body(response) -> bytes:
    chunks: list[bytes] = []
    async for chunk in response.body_iterator:
        chunks.append(chunk if isinstance(chunk, bytes) else chunk.encode())
    return b"".join(chunks)


class ExpenseBudgetExecutionExportTests(unittest.TestCase):
    def test_export_plan_centralizes_workbook_kind_and_filename_selection(self) -> None:
        monthly_plan = resolve_expense_budget_execution_export_plan(
            ExpenseBudgetExecutionExportOptions(mode="query", perspective="group")
        )
        template_plan = resolve_expense_budget_execution_export_plan(
            ExpenseBudgetExecutionExportOptions(mode="template", perspective="owner_dept")
        )
        subject_plan = resolve_expense_budget_execution_export_plan(
            ExpenseBudgetExecutionExportOptions(mode="subject", perspective="entity")
        )
        flat_plan = resolve_expense_budget_execution_export_plan(
            ExpenseBudgetExecutionExportOptions(mode="flat", perspective="owner_dept")
        )

        self.assertEqual(monthly_plan.workbook_kind, "monthly")
        self.assertEqual(monthly_plan.filename, "expense_budget_execution_monthly_report.xlsx")
        self.assertEqual(template_plan.workbook_kind, "template")
        self.assertEqual(template_plan.filename, "expense_budget_execution_department_report.xlsx")
        self.assertEqual(subject_plan.workbook_kind, "subject")
        self.assertEqual(subject_plan.filename, "expense_budget_execution_subject_report.xlsx")
        self.assertEqual(flat_plan.workbook_kind, "flat")
        self.assertEqual(flat_plan.filename, "expense_budget_execution_owner_dept.xlsx")

    def test_export_plan_rejects_unknown_mode(self) -> None:
        with self.assertRaisesRegex(
            ExpenseBudgetExecutionExportError,
            "未知费用预算执行导出模式: legacy",
        ):
            resolve_expense_budget_execution_export_plan(
                ExpenseBudgetExecutionExportOptions(mode="legacy", perspective="group")
            )

    def test_export_plan_rejects_unknown_flat_perspective(self) -> None:
        with self.assertRaisesRegex(
            ExpenseBudgetExecutionExportError,
            "未知费用预算执行导出视角: legacy_owner",
        ):
            resolve_expense_budget_execution_export_plan(
                ExpenseBudgetExecutionExportOptions(mode="flat", perspective="legacy_owner")
            )

    def test_export_workbook_builder_centralizes_plan_kind_dispatch(self) -> None:
        report = {"note": "导出"}
        options = ExpenseBudgetExecutionExportOptions(mode="flat", perspective="group")
        plan = ExpenseBudgetExecutionExportPlan(
            workbook_kind="flat",
            filename="expense_budget_execution_group.xlsx",
        )
        workbook = Workbook()

        with patch.object(
            export_module,
            "_build_flat_workbook",
            return_value=workbook,
        ) as flat_builder:
            result = build_expense_budget_execution_workbook(
                report=report,
                options=options,
                plan=plan,
            )

        self.assertIs(result, workbook)
        flat_builder.assert_called_once_with(
            report,
            amount_unit_label="元",
            amount_divisor=1.0,
            options=options,
        )

    def test_export_workbook_builder_rejects_unknown_plan_kind(self) -> None:
        with self.assertRaisesRegex(
            ExpenseBudgetExecutionExportError,
            "未知费用预算执行导出工作簿类型: legacy-flat",
        ):
            build_expense_budget_execution_workbook(
                report={"note": "导出"},
                options=ExpenseBudgetExecutionExportOptions(mode="flat", perspective="group"),
                plan=ExpenseBudgetExecutionExportPlan(
                    workbook_kind="legacy-flat",
                    filename="legacy.xlsx",
                ),
            )

    def test_export_workbook_builder_rejects_unknown_amount_unit(self) -> None:
        with self.assertRaisesRegex(
            ExpenseBudgetExecutionExportError,
            "未知费用预算执行金额单位: legacy_unit",
        ):
            build_expense_budget_execution_workbook(
                report={"note": "导出"},
                options=ExpenseBudgetExecutionExportOptions(
                    mode="flat",
                    perspective="group",
                    amount_unit="legacy_unit",
                ),
                plan=ExpenseBudgetExecutionExportPlan(
                    workbook_kind="flat",
                    filename="expense_budget_execution_group.xlsx",
                ),
            )

    def test_template_export_writes_tree_formulas_sources_and_unit_scaling(self) -> None:
        report = {
            "template_title": "部门费用执行",
            "note": "2026年2月口径",
            "current_month": 2,
            "budget_year": 2026,
            "actual_source_file": "actual.xlsx",
            "previous_actual_source_file": "actual_2025.xlsx",
            "subject_tree": [
                {
                    "level_label": "一级",
                    "level_number": 1,
                    "subject_name": "费用合计",
                    "current_actual": 0,
                    "annual_budget": 0,
                    "last_year_actual": 0,
                    "monthly_actuals": [0, 0],
                    "previous_year_monthly_actuals": [0] * 12,
                    "children": [
                        {
                            "level_label": "二级",
                            "level_number": 2,
                            "subject_name": "业务费用",
                            "current_actual": 20_000,
                            "annual_budget": 40_000,
                            "last_year_actual": 12_000,
                            "month_over_month": 3_000,
                            "month_over_month_rate": 0.15,
                            "monthly_actuals": [10_000, 10_000],
                            "previous_year_monthly_actuals": [1_000] * 12,
                            "children": [],
                        }
                    ],
                }
            ],
        }
        options = ExpenseBudgetExecutionExportOptions(
            mode="template",
            amount_unit="ten_thousand",
            include_monthly_actuals=True,
            include_last_year_monthly_actuals=True,
        )

        export = build_expense_budget_execution_export(report, options)
        ws = export.workbook["部门模式"]

        self.assertEqual(export.filename, "expense_budget_execution_department_report.xlsx")
        self.assertEqual(ws["A3"].value, "单位：万元")
        self.assertEqual(ws["D5"].value, "1月实际")
        self.assertEqual(ws["M5"].value, "25年1月实际")
        self.assertEqual(ws["B7"].value, "  业务费用")
        self.assertEqual(ws["C7"].value, 2.0)
        self.assertEqual(ws["D7"].value, 1.0)
        self.assertEqual(ws["F7"].value, 4.0)
        self.assertEqual(ws["J7"].value, 0.3)
        self.assertEqual(ws["C6"].value, "=C7")
        self.assertEqual(ws["F6"].value, "=F7")
        self.assertEqual(ws["L6"].value, "=L7")
        self.assertEqual(ws["A9"].value, "分月实际来源")
        self.assertEqual(ws["B9"].value, "actual.xlsx")
        self.assertEqual(ws["B10"].value, "actual_2025.xlsx")

    def test_template_export_empty_tree_writes_empty_message(self) -> None:
        report = {
            "template_title": "部门费用执行",
            "note": "空数据口径",
            "current_month": 2,
            "budget_year": 2026,
            "actual_source_file": "",
            "previous_actual_source_file": "",
            "subject_tree": [],
        }

        export = build_expense_budget_execution_export(
            report,
            ExpenseBudgetExecutionExportOptions(mode="template"),
        )
        ws = export.workbook["部门模式"]

        self.assertEqual(ws["A6"].value, "当前条件下没有可展示的数据。")
        self.assertEqual(ws["A8"].value, "分月实际来源")

    def test_flat_export_scales_rows_and_keeps_report_filename_perspective(self) -> None:
        report = {
            "note": "查询导出",
            "rows": [
                {
                    "perspective": "group",
                    "dimension_value": "零售事业群",
                    "entity_name": "微众银行",
                    "group_name": "零售事业群",
                    "owner_dept": "零售部",
                    "budget_subject": "业务费用",
                    "monthly_actuals": [1_000] * 12,
                    "cumulative_actual": 12_000,
                    "annual_budget": 24_000,
                    "execution_rate": 0.5,
                    "month_over_month": 1_500,
                    "month_over_month_rate": 0.1,
                }
            ],
        }
        options = ExpenseBudgetExecutionExportOptions(
            mode="flat",
            perspective="group",
            amount_unit="thousand",
        )

        export = build_expense_budget_execution_export(report, options)
        ws = export.workbook["费用预算执行报表"]

        self.assertEqual(export.filename, "expense_budget_execution_group.xlsx")
        self.assertEqual(ws["A1"].value, "单位：千元")
        self.assertEqual(ws["A3"].value, "事业群")
        self.assertEqual(ws["G3"].value, 1.0)
        self.assertEqual(ws["S3"].value, 12.0)
        self.assertEqual(ws["T3"].value, 24.0)
        self.assertEqual(ws["U3"].value, 0.5)
        self.assertEqual(ws["V3"].value, 1.5)
        self.assertEqual(ws["W3"].number_format, "0.00%")
        self.assertEqual(ws["B5"].value, "查询导出")

    def test_flat_export_empty_rows_writes_empty_message(self) -> None:
        export = build_expense_budget_execution_export(
            {"note": "无查询结果", "rows": []},
            ExpenseBudgetExecutionExportOptions(mode="flat", perspective="group"),
        )
        ws = export.workbook["费用预算执行报表"]

        self.assertEqual(ws["A3"].value, "当前条件下没有可展示的数据。")
        self.assertEqual(ws["A5"].value, "说明")
        self.assertEqual(ws["B5"].value, "无查询结果")


class ExpenseBudgetExecutionWorkbookResponseTests(unittest.IsolatedAsyncioTestCase):
    async def test_workbook_response_uses_common_excel_download_contract(self) -> None:
        wb = Workbook()
        wb.active["A1"] = "费用执行"
        export = ExpenseBudgetExecutionWorkbookExport(
            workbook=wb,
            filename="费用预算执行报表.xlsx",
        )

        response = expense_budget_execution_workbook_response(export)

        self.assertEqual(
            response.media_type,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        disposition = response.headers["content-disposition"]
        self.assertIn("filename=workbook.xlsx", disposition)
        self.assertIn("filename*=UTF-8''%E8%B4%B9%E7%94%A8", disposition)

        body = await _read_response_body(response)
        loaded = load_workbook(BytesIO(body))
        self.assertEqual(loaded.active["A1"].value, "费用执行")


if __name__ == "__main__":
    unittest.main()
