from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any, Awaitable, Callable
import re
import unicodedata

import app.core.aiosqlite_compat as aiosqlite
from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from app.schemas import DeptAccountCreate, DeptAccountRow, DeptAccountUpdate
from app.services.department_expense_contracts import DEPT_GROUP_LEVEL, DEPT_OWNER_LEVEL, MAX_DEPT_LEVEL
from app.services.expense_master_data import sync_expense_dept_name_refs


ENTITY_ORDER = {"微众银行": 0, "科技子": 1, "科技孙": 2}
GROUP_ORDER = {
    "个人金融事业群": 0,
    "企业及机构金融事业群": 1,
    "科技及智能事业群": 2,
    "国际发展部": 3,
    "国际业务": 4,
    "资源管理及管控职能群": 5,
    "其他": 6,
    "历史架构": 7,
    "科技子": 8,
    "科技孙": 9,
    "虚拟架构": 10,
}


@dataclass(frozen=True)
class DeptAccountUpdateResult:
    row: DeptAccountRow
    before_data: dict[str, Any]
    after_data: dict[str, Any]
    sync_counts: dict[str, int]


@dataclass(frozen=True)
class DeptAccountDeletion:
    before_data: dict[str, Any]


@dataclass(frozen=True)
class DeptTreeExportWorkbook:
    content: bytes
    filename: str = "dept_tree_export.xlsx"


@dataclass(frozen=True)
class DeptAccountImportPreview:
    columns: list[str]
    preview_rows: list[dict[str, str]]
    row_count: int


@dataclass(frozen=True)
class DeptAccountImportResultWorkbook:
    content: bytes
    total: int
    success: int
    overwrite: int
    failed: int
    filename: str = "dept_account_import_result.xlsx"


async def list_dept_accounts(common_db: Path | str) -> list[DeptAccountRow]:
    async with aiosqlite.connect(common_db) as db:
        await db.execute("PRAGMA foreign_keys = ON")
        cur = await db.execute(
            """
            SELECT dept_code, dept_name, entity_name, parent_code, level, is_leaf
            FROM dept_account
            ORDER BY entity_name, dept_code
            """
        )
        rows = await cur.fetchall()
    return [_row_from_db(row) for row in rows]


async def create_dept_account(common_db: Path | str, body: DeptAccountCreate) -> DeptAccountRow:
    _validate_level(body.level)
    async with aiosqlite.connect(common_db) as db:
        await db.execute("PRAGMA foreign_keys = ON")
        cur = await db.execute("SELECT 1 FROM dept_account WHERE dept_code = ?", (body.dept_code,))
        if await cur.fetchone():
            raise HTTPException(status_code=409, detail="部门代码已存在")
        if body.parent_code:
            cur = await db.execute("SELECT 1 FROM dept_account WHERE dept_code = ?", (body.parent_code,))
            if not await cur.fetchone():
                raise HTTPException(status_code=400, detail="上级部门不存在")
        await db.execute(
            """
            INSERT INTO dept_account (dept_code, dept_name, entity_name, parent_code, level, is_leaf)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (body.dept_code, body.dept_name, body.entity_name, body.parent_code, body.level, int(body.is_leaf)),
        )
        await db.commit()
    return DeptAccountRow(**body.model_dump())


async def update_dept_account(
    common_db: Path | str,
    code: str,
    body: DeptAccountUpdate,
    *,
    rename_sync: Callable[..., Awaitable[dict[str, int]]] = sync_expense_dept_name_refs,
) -> DeptAccountUpdateResult:
    sync_counts: dict[str, int] = {}
    async with aiosqlite.connect(common_db) as db:
        await db.execute("PRAGMA foreign_keys = ON")
        before = await _fetch_dept_row(db, code)
        if not before:
            raise HTTPException(status_code=404, detail="部门不存在")

        updates: list[str] = []
        vals: list[Any] = []
        if body.dept_name is not None:
            if not body.dept_name.strip():
                raise HTTPException(status_code=400, detail="部门名称不能为空")
            updates.append("dept_name = ?")
            vals.append(body.dept_name.strip())
        if body.parent_code is not None:
            parent_code = body.parent_code.strip() or None
            if parent_code:
                if parent_code == code:
                    raise HTTPException(status_code=400, detail="上级部门不能为自身")
                cur = await db.execute("SELECT 1 FROM dept_account WHERE dept_code = ?", (parent_code,))
                if not await cur.fetchone():
                    raise HTTPException(status_code=400, detail="上级部门不存在")
            updates.append("parent_code = ?")
            vals.append(parent_code)
        if body.entity_name is not None:
            entity_name = body.entity_name.strip()
            if not entity_name:
                raise HTTPException(status_code=400, detail="主体不能为空")
            updates.append("entity_name = ?")
            vals.append(entity_name)
        if body.level is not None:
            _validate_level(body.level)
            updates.append("level = ?")
            vals.append(body.level)
        if body.is_leaf is not None:
            updates.append("is_leaf = ?")
            vals.append(int(body.is_leaf))

        if updates:
            vals.append(code)
            await db.execute(f"UPDATE dept_account SET {', '.join(updates)} WHERE dept_code = ?", vals)
            new_name = body.dept_name.strip() if body.dept_name is not None else str(before[1] or "")
            old_name = str(before[1] or "").strip()
            dept_level = int(before[4] or 0)
            sync_counts = await rename_sync(
                db,
                dept_level=dept_level,
                old_name=old_name,
                new_name=new_name,
            )
            await db.commit()

        row = await _fetch_dept_row(db, code)
    if not row:
        raise HTTPException(status_code=500, detail="更新部门后未找到记录")
    before_data = _audit_data_from_db(before)
    after_data = _audit_data_from_db(row)
    after_data["synced_related_updates"] = sync_counts
    return DeptAccountUpdateResult(
        row=_row_from_db(row),
        before_data=before_data,
        after_data=after_data,
        sync_counts=sync_counts,
    )


async def delete_dept_account(common_db: Path | str, code: str) -> DeptAccountDeletion:
    async with aiosqlite.connect(common_db) as db:
        await db.execute("PRAGMA foreign_keys = ON")
        before = await _fetch_dept_row(db, code)
        if not before:
            raise HTTPException(status_code=404, detail="部门不存在")
        cur = await db.execute("SELECT COUNT(*) FROM dept_account WHERE parent_code = ?", (code,))
        if int((await cur.fetchone())[0] or 0) > 0:
            raise HTTPException(status_code=409, detail="该部门下仍有子节点，无法删除")
        await db.execute("DELETE FROM dept_account WHERE dept_code = ?", (code,))
        await db.commit()
    return DeptAccountDeletion(before_data=_audit_data_from_db(before))


async def build_dept_tree_export_workbook(
    common_db: Path | str,
    *,
    template_path: Path | None = None,
) -> DeptTreeExportWorkbook:
    rows = await _load_dept_tree_rows(common_db)
    workbook = build_dept_tree_workbook(rows, template_path=template_path)
    return DeptTreeExportWorkbook(content=workbook)


async def preview_dept_account_import(
    content: bytes,
    *,
    normalize_cell: Callable[[Any], str],
) -> DeptAccountImportPreview:
    wb = _load_dept_import_workbook(content, data_only=True)
    ws = wb["数据模版"]
    headers = [normalize_cell(cell.value) for cell in ws[1]]
    if not any(headers):
        raise HTTPException(status_code=400, detail="数据模版工作表第一行字段头为空")

    preview_rows: list[dict[str, str]] = []
    total_rows = 0
    for row_idx in range(2, ws.max_row + 1):
        row_values = [normalize_cell(ws.cell(row_idx, col_idx).value) for col_idx in range(1, ws.max_column + 1)]
        if not any(row_values):
            continue
        total_rows += 1
        if len(preview_rows) < 20:
            preview_rows.append(
                {
                    headers[idx]: row_values[idx]
                    for idx in range(min(len(headers), len(row_values)))
                    if headers[idx]
                }
            )

    return DeptAccountImportPreview(
        columns=[header for header in headers if header],
        preview_rows=preview_rows,
        row_count=total_rows,
    )


async def apply_dept_account_import(
    common_db: Path | str,
    content: bytes,
    mappings: dict[str, Any],
    *,
    normalize_cell: Callable[[Any], str],
    color_row: Callable[[Any, int, int, str], None],
    validate_dept_code_with_parent: Callable[[str, int, str | None], str | None],
) -> DeptAccountImportResultWorkbook:
    _validate_import_mappings(mappings)
    wb = _load_dept_import_workbook(content, data_only=False)
    ws = wb["数据模版"]
    headers = [normalize_cell(cell.value) for cell in ws[1]]
    header_col = {header: idx + 1 for idx, header in enumerate(headers) if header}
    for col_name in mappings.values():
        if col_name and col_name not in header_col:
            raise HTTPException(status_code=400, detail=f"映射列不存在：{col_name}")

    fail_reason_col = ws.max_column + 1
    ws.cell(row=1, column=fail_reason_col, value="失败原因")
    imported_rows = _parse_dept_import_rows(ws, header_col, mappings, normalize_cell=normalize_cell)

    dept_code_re = re.compile(r"^Y\d+$")
    success_count = 0
    overwrite_count = 0
    failed_count = 0

    async with aiosqlite.connect(common_db) as db:
        await db.execute("PRAGMA foreign_keys = ON")
        existing_dept = await _load_existing_dept_map(db)

        def mark_failed(row_idx: int, reasons: list[str] | str) -> None:
            reason_text = "；".join(reasons) if isinstance(reasons, list) else str(reasons)
            ws.cell(row=row_idx, column=fail_reason_col, value=reason_text)
            color_row(ws, row_idx, ws.max_column, "FFFF0000")

        for row in imported_rows:
            errors = _validate_import_row(
                row,
                dept_code_re=dept_code_re,
                validate_dept_code_with_parent=validate_dept_code_with_parent,
            )
            if errors:
                failed_count += 1
                mark_failed(int(row["row_idx"]), errors)
                continue

            try:
                changes = [
                    await _upsert_import_dept(
                        db,
                        existing_dept,
                        code=str(row["lvl1_code"]).strip(),
                        name=str(row["lvl1_name"]).strip(),
                        entity_name=str(row["entity_name"]).strip(),
                        parent_code=None,
                        level=DEPT_GROUP_LEVEL,
                        is_leaf=0,
                    )
                ]
                if bool(row.get("row_has_lvl2")):
                    changes.append(
                        await _upsert_import_dept(
                            db,
                            existing_dept,
                            code=str(row["lvl2_code"]).strip(),
                            name=str(row["lvl2_name"]).strip(),
                            entity_name=str(row["entity_name"]).strip(),
                            parent_code=str(row["lvl1_code"]).strip(),
                            level=DEPT_OWNER_LEVEL,
                            is_leaf=1,
                        )
                    )
                success_count += 1
                if "updated" in changes:
                    overwrite_count += 1
                    color_row(ws, int(row["row_idx"]), ws.max_column, "FF0000FF")
                else:
                    color_row(ws, int(row["row_idx"]), ws.max_column, "FF008000")
            except ValueError as exc:
                failed_count += 1
                mark_failed(int(row["row_idx"]), str(exc))

        await db.commit()

    out = BytesIO()
    wb.save(out)
    return DeptAccountImportResultWorkbook(
        content=out.getvalue(),
        total=success_count + failed_count,
        success=success_count,
        overwrite=overwrite_count,
        failed=failed_count,
    )


async def _load_dept_tree_rows(common_db: Path | str) -> list[dict[str, Any]]:
    async with aiosqlite.connect(common_db) as db:
        await db.execute("PRAGMA foreign_keys = ON")
        cur = await db.execute(
            """
            SELECT dept_code, dept_name, entity_name, parent_code, level
            FROM dept_account
            """
        )
        rows = await cur.fetchall()
    return [
        {
            "dept_code": str(row[0]),
            "dept_name": str(row[1] or ""),
            "entity_name": str(row[2] or "微众银行"),
            "parent_code": str(row[3]) if row[3] is not None else None,
            "level": int(row[4] or 0),
        }
        for row in rows
    ]


def build_dept_tree_workbook(
    dept_rows: list[dict[str, Any]],
    *,
    template_path: Path | None = None,
) -> bytes:
    entity_roots = _build_entity_tree(dept_rows)
    headers = [
        "主体",
        "事业群代码",
        "事业群名称",
        "费用归属部门代码",
        "费用归属部门名称",
    ]

    style_cell = None
    if template_path is not None and template_path.exists():
        template_wb = load_workbook(template_path)
        style_ws = template_wb["数据模版"] if "数据模版" in template_wb.sheetnames else template_wb.active
        style_cell = style_ws.cell(row=1, column=1)

    wb = Workbook()
    ws = wb.active
    ws.title = "数据模版"
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        if style_cell is not None:
            cell.font = copy(style_cell.font)
            cell.fill = copy(style_cell.fill)
            cell.border = copy(style_cell.border)
            cell.alignment = copy(style_cell.alignment)
            cell.number_format = style_cell.number_format
            cell.protection = copy(style_cell.protection)

    export_rows: list[list[Any]] = []

    def walk(nodes: list[dict[str, Any]], depth: int, current_entity: str = "") -> None:
        for node in nodes:
            row = [""] * len(headers)
            node_type = str(node.get("type", ""))
            if node_type == "entity":
                row[0] = str(node.get("entity_name", "") or node.get("name", "") or "微众银行")
            elif node_type == "dept":
                row[0] = str(node.get("entity_name", "") or current_entity or "微众银行")
                level = min(max(int(node.get("level", depth) or depth), DEPT_GROUP_LEVEL), MAX_DEPT_LEVEL)
                col_idx = 1 + (level - 1) * 2
                row[col_idx] = str(node.get("code", ""))
                row[col_idx + 1] = str(node.get("name", ""))
            export_rows.append(row)
            children = node.get("children") or []
            if children:
                next_depth = depth if node_type == "entity" else depth + 1
                next_entity = str(node.get("entity_name", "") or current_entity or "微众银行")
                walk(children, next_depth, next_entity)

    walk(entity_roots, 1)
    for row_idx, row in enumerate(export_rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    ws.freeze_panes = "A2"
    for col_idx in (3, 5):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions.group(col_letter, col_letter, outline_level=1, hidden=False)

    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is None:
                continue
            text = str(val)
            part_max = max((_display_width(seg) for seg in text.splitlines()), default=0)
            max_len = max(max_len, part_max)
        ws.column_dimensions[col_letter].width = max(6, min(80, max_len + 2))
        ws.column_dimensions[col_letter].bestFit = True

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _load_dept_import_workbook(content: bytes, *, data_only: bool):
    try:
        wb = load_workbook(filename=BytesIO(content), data_only=data_only)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"无法读取Excel文件：{exc}") from exc
    if "数据模版" not in wb.sheetnames:
        raise HTTPException(status_code=400, detail='上传文件缺失“数据模版”工作表，不能上传数据。')
    return wb


def _validate_import_mappings(mappings: dict[str, Any]) -> None:
    required_fields = [
        "entityName",
        "level1Code",
        "level1Name",
        "level2Code",
        "level2Name",
    ]
    for field_name in required_fields:
        if not mappings.get(field_name):
            raise HTTPException(status_code=400, detail=f"字段映射缺失：{field_name}")


def _parse_dept_import_rows(
    ws,
    header_col: dict[str, int],
    mappings: dict[str, Any],
    *,
    normalize_cell: Callable[[Any], str],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    current_entity_name = ""
    current_lvl1_code = ""
    current_lvl1_name = ""
    for row_idx in range(2, ws.max_row + 1):
        raw_by_header = {header: normalize_cell(ws.cell(row_idx, col_idx).value) for header, col_idx in header_col.items()}
        if not any(raw_by_header.values()):
            continue
        raw_entity_name = raw_by_header.get(str(mappings.get("entityName", "")), "")
        raw_lvl1_code = raw_by_header.get(str(mappings.get("level1Code", "")), "").upper()
        raw_lvl1_name = raw_by_header.get(str(mappings.get("level1Name", "")), "")
        raw_lvl2_code = raw_by_header.get(str(mappings.get("level2Code", "")), "").upper()
        raw_lvl2_name = raw_by_header.get(str(mappings.get("level2Name", "")), "")
        row_has_lvl1 = bool(raw_lvl1_code or raw_lvl1_name)
        row_has_lvl2 = bool(raw_lvl2_code or raw_lvl2_name)

        if raw_entity_name and raw_entity_name != current_entity_name:
            current_entity_name = raw_entity_name
            current_lvl1_code = ""
            current_lvl1_name = ""
        if row_has_lvl1:
            current_lvl1_code = raw_lvl1_code
            current_lvl1_name = raw_lvl1_name
        if not row_has_lvl1 and not row_has_lvl2:
            continue

        rows.append(
            {
                "row_idx": row_idx,
                "entity_name": raw_entity_name or current_entity_name,
                "lvl1_code": raw_lvl1_code or current_lvl1_code,
                "lvl1_name": raw_lvl1_name or current_lvl1_name,
                "lvl2_code": raw_lvl2_code,
                "lvl2_name": raw_lvl2_name,
                "row_has_lvl1": row_has_lvl1,
                "row_has_lvl2": row_has_lvl2,
            }
        )
    return rows


async def _load_existing_dept_map(db: aiosqlite.Connection) -> dict[str, dict[str, Any]]:
    cur = await db.execute("SELECT dept_code, dept_name, entity_name, parent_code, level, is_leaf FROM dept_account")
    rows = await cur.fetchall()
    return {
        str(row[0]): {
            "name": str(row[1] or ""),
            "entity_name": str(row[2] or "微众银行"),
            "parent_code": str(row[3]) if row[3] is not None else None,
            "level": int(row[4]),
            "is_leaf": bool(row[5]),
        }
        for row in rows
    }


def _validate_import_row(
    row: dict[str, Any],
    *,
    dept_code_re: re.Pattern[str],
    validate_dept_code_with_parent: Callable[[str, int, str | None], str | None],
) -> list[str]:
    errors: list[str] = []
    entity_name = str(row["entity_name"] or "").strip()
    lvl1_code = str(row["lvl1_code"] or "").strip()
    lvl1_name = str(row["lvl1_name"] or "").strip()
    lvl2_code = str(row["lvl2_code"] or "").strip()
    lvl2_name = str(row["lvl2_name"] or "").strip()
    row_has_lvl2 = bool(row.get("row_has_lvl2"))
    if not entity_name:
        errors.append("主体不能为空")
    if not lvl1_code:
        errors.append("事业群代码不能为空，且无法从上方事业群行继承")
    elif not dept_code_re.match(lvl1_code):
        errors.append("事业群代码格式错误（示例：Y1、Y11）")
    if not lvl1_name:
        errors.append("事业群名称不能为空，且无法从上方事业群行继承")
    if row_has_lvl2:
        if not lvl2_code:
            errors.append("费用归属部门代码不能为空")
        elif not dept_code_re.match(lvl2_code):
            errors.append("费用归属部门代码格式错误（示例：Y101、Y10101）")
        if not lvl2_name:
            errors.append("费用归属部门名称不能为空")
    lvl1_code_err = validate_dept_code_with_parent(lvl1_code, 1, None)
    if lvl1_code_err:
        errors.append(lvl1_code_err)
    if row_has_lvl2:
        lvl2_code_err = validate_dept_code_with_parent(lvl2_code, 2, lvl1_code)
        if lvl2_code_err:
            errors.append(lvl2_code_err)
    return errors


async def _upsert_import_dept(
    db: aiosqlite.Connection,
    existing_dept: dict[str, dict[str, Any]],
    *,
    code: str,
    name: str,
    entity_name: str,
    parent_code: str | None,
    level: int,
    is_leaf: int,
) -> str:
    existing = existing_dept.get(code)
    if existing:
        if (existing.get("parent_code") or None) != (parent_code or None):
            raise ValueError(f"部门代码 {code} 已存在，但上级部门不匹配")
        if int(existing.get("level") or 0) != level:
            raise ValueError(f"部门代码 {code} 已存在，但层级不匹配")
        changed = (
            existing.get("name") != name
            or existing.get("entity_name") != entity_name
            or bool(existing.get("is_leaf", level >= MAX_DEPT_LEVEL)) != bool(is_leaf)
        )
        if changed:
            await db.execute(
                """
                UPDATE dept_account
                SET dept_name = ?, entity_name = ?, is_leaf = ?
                WHERE dept_code = ?
                """,
                (name, entity_name, is_leaf, code),
            )
            existing_dept[code] = {
                "name": name,
                "entity_name": entity_name,
                "parent_code": parent_code,
                "level": level,
                "is_leaf": bool(is_leaf),
            }
            return "updated"
        return "unchanged"

    await db.execute(
        """
        INSERT INTO dept_account (dept_code, dept_name, entity_name, parent_code, level, is_leaf)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (code, name, entity_name, parent_code, level, is_leaf),
    )
    existing_dept[code] = {
        "name": name,
        "entity_name": entity_name,
        "parent_code": parent_code,
        "level": level,
        "is_leaf": bool(is_leaf),
    }
    return "inserted"


def _build_entity_tree(dept_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    dept_nodes: dict[str, dict[str, Any]] = {}
    for row in dept_rows:
        code = str(row["dept_code"])
        dept_nodes[code] = {
            "type": "dept",
            "code": code,
            "name": str(row.get("dept_name") or ""),
            "entity_name": str(row.get("entity_name") or "微众银行"),
            "parent_code": str(row["parent_code"]) if row.get("parent_code") is not None else None,
            "level": int(row.get("level") or 0),
            "children": [],
        }

    roots: list[dict[str, Any]] = []
    for node in dept_nodes.values():
        parent_code = node["parent_code"]
        if parent_code and parent_code in dept_nodes:
            dept_nodes[parent_code]["children"].append(node)
        else:
            roots.append(node)

    roots_by_entity: dict[str, list[dict[str, Any]]] = {}
    for node in roots:
        entity_name = str(node.get("entity_name", "") or "微众银行")
        roots_by_entity.setdefault(entity_name, []).append(node)

    entity_roots = [
        {
            "type": "entity",
            "code": "",
            "name": entity_name,
            "entity_name": entity_name,
            "children": roots_by_entity[entity_name],
        }
        for entity_name in _sort_entity_names(list(roots_by_entity.keys()))
    ]
    _sort_tree(entity_roots)
    return entity_roots


def _sort_entity_names(names: set[str] | list[str]) -> list[str]:
    return sorted(
        {str(name or "微众银行").strip() or "微众银行" for name in names},
        key=lambda name: (ENTITY_ORDER.get(name, 999), name),
    )


def _group_sort_key(group_name: str) -> tuple[int, str]:
    value = str(group_name or "").strip()
    return (GROUP_ORDER.get(value, len(GROUP_ORDER)), value)


def _sort_tree(nodes: list[dict[str, Any]]) -> None:
    nodes.sort(
        key=lambda node: (
            ENTITY_ORDER.get(str(node.get("entity_name", "") or "微众银行"), 999),
            _group_sort_key(str(node.get("name", ""))) if node.get("level", 0) == 1 else 0,
            len(str(node.get("name", ""))),
            str(node.get("name", "")),
            str(node.get("code", "")),
        )
    )
    for node in nodes:
        children = node.get("children") or []
        if children:
            _sort_tree(children)


def _display_width(value: str) -> int:
    width = 0
    for ch in value:
        width += 2 if unicodedata.east_asian_width(ch) in ("F", "W") else 1
    return width


async def _fetch_dept_row(db: aiosqlite.Connection, code: str) -> tuple[Any, ...] | None:
    cur = await db.execute(
        """
        SELECT dept_code, dept_name, entity_name, parent_code, level, is_leaf
        FROM dept_account
        WHERE dept_code = ?
        """,
        (code,),
    )
    return await cur.fetchone()


def _row_from_db(row: tuple[Any, ...]) -> DeptAccountRow:
    return DeptAccountRow(
        dept_code=str(row[0]),
        dept_name=str(row[1]),
        entity_name=str(row[2] or "微众银行"),
        parent_code=str(row[3]) if row[3] is not None else None,
        level=int(row[4]),
        is_leaf=bool(row[5]),
    )


def _audit_data_from_db(row: tuple[Any, ...]) -> dict[str, Any]:
    return {
        "dept_code": row[0],
        "dept_name": row[1],
        "entity_name": row[2],
        "parent_code": row[3],
        "level": row[4],
        "is_leaf": bool(row[5]),
    }


def _validate_level(level: int) -> None:
    if level < DEPT_GROUP_LEVEL or level > MAX_DEPT_LEVEL:
        raise HTTPException(status_code=400, detail="部门树仅支持两级部门节点：事业群、费用归属部门")
