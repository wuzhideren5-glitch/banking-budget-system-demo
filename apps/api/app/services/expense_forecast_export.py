"""Excel workbook builders for expense forecast exports."""
from __future__ import annotations

from collections import defaultdict
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Protection
from openpyxl.utils import get_column_letter


MONTH_KEYS = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]

AMOUNT_UNIT_DIVISORS: dict[str, tuple[str, float]] = {
    "yuan": ("元", 1.0),
    "thousand": ("千元", 1_000.0),
    "ten_thousand": ("万元", 10_000.0),
    "wan": ("万元", 10_000.0),
    "million": ("百万元", 1_000_000.0),
    "hundred_million": ("亿元", 100_000_000.0),
}


def build_export_columns(exclude_fields: list[str]) -> dict[str, Any]:
    excluded = set(exclude_fields)
    col = 2
    month_cols: dict[int, int] = {}
    for month in range(1, 13):
        key = MONTH_KEYS[month - 1]
        if key not in excluded:
            month_cols[month] = col
            col += 1
    result: dict[str, Any] = {"month_cols": month_cols}
    for key in ["total", "annual_budget", "gap", "rate", "biz", "capital", "capital_gap"]:
        if key not in excluded:
            result[key + "_col"] = col
            col += 1
        else:
            result[key + "_col"] = None
    result["total_data_cols"] = col - 1
    return result


def amount_unit_meta(amount_unit: str | None) -> tuple[str, float]:
    normalized = str(amount_unit or "").strip().lower()
    return AMOUNT_UNIT_DIVISORS.get(normalized, AMOUNT_UNIT_DIVISORS["yuan"])


def scale_amount(value: float | int | None, divisor: float) -> float:
    return round(float(value or 0.0) / divisor, 2)


def _append_header(ws: Any, *, header_label: str, cols: dict[str, Any]) -> None:
    header = [header_label]
    for month in range(1, 13):
        if month in cols["month_cols"]:
            header.append(f"{month}月")
    if cols["total_col"] is not None:
        header.append("全年预测")
    if cols["annual_budget_col"] is not None:
        header.append("年度预算")
    if cols["gap_col"] is not None:
        header.append("全年预测-年度预算")
    if cols["rate_col"] is not None:
        header.append("预算执行率")
    if cols["biz_col"] is not None:
        header.append("业务报送")
    if cols["capital_col"] is not None:
        header.append("资划建议")
    if cols["capital_gap_col"] is not None:
        header.append("资划建议-业务报送")
    ws.append(header)


def _finalize_workbook(wb: Workbook, ws: Any, cols: dict[str, Any]) -> BytesIO:
    ws.protection.sheet = True
    for col_idx in range(1, cols["total_data_cols"] + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def build_expense_forecast_export_workbook(
    *,
    year: int,
    forecast_version: str,
    scope_type: str,
    scope_value: str,
    compile_mode: str,
    amount_unit: str,
    exclude_fields: list[str],
    view: Any | None = None,
    subject_view: Any | None = None,
) -> tuple[BytesIO, str]:
    """Build the regular expense forecast workbook without HTTP concerns."""
    scope_label = {"entity": "主体", "group": "事业群", "owner": "费用归属部门"}[scope_type]
    amount_unit_label, amount_divisor = amount_unit_meta(amount_unit)
    cols = build_export_columns(exclude_fields)
    month_cols = cols["month_cols"]
    total_col = cols["total_col"]
    annual_budget_col = cols["annual_budget_col"]
    gap_col = cols["gap_col"]
    rate_col = cols["rate_col"]
    biz_col = cols["biz_col"]
    capital_col = cols["capital_col"]
    capital_gap_col = cols["capital_gap_col"]

    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    locked = Protection(locked=True)
    unlocked = Protection(locked=False)
    bold_font = Font(bold=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "费用预测表"
    ws.append(["费用预测表"])
    header_label = "预算科目"
    display_file_name = f"费用预测表_{year}_{forecast_version}.xlsx"

    if compile_mode == "subject":
        if subject_view is None:
            raise ValueError("subject_view is required for subject compile mode")
        actual_cutoff_month = int(subject_view.actual_cutoff_month)
        header_label = "费用归属部门"
        display_file_name = f"费用预测表_按预算科目_{year}_{forecast_version}_{subject_view.subject_name}.xlsx"
        ws.append(
            [
                f"年份：{year}",
                f"版本：{forecast_version}",
                "编制维度：按预算科目",
                f"编制口径：{scope_label}",
                f"口径值：{scope_value}",
                f"预算科目：{subject_view.subject_name}",
                f"实际截至月：{actual_cutoff_month}月",
                f"单位：{amount_unit_label}",
            ]
        )
    else:
        if view is None:
            raise ValueError("view is required for scope compile mode")
        actual_cutoff_month = int(view.actual_cutoff_month)
        ws.append(
            [
                f"年份：{year}",
                f"版本：{forecast_version}",
                "编制维度：按费用归属部门",
                f"编制口径：{scope_label}",
                f"口径值：{scope_value}",
                f"实际截至月：{actual_cutoff_month}月",
                f"单位：{amount_unit_label}",
            ]
        )
    _append_header(ws, header_label=header_label, cols=cols)

    data_start_row = 4
    subject_col = 1
    if compile_mode == "subject":
        for row_idx, row in enumerate(subject_view.rows):
            excel_row = data_start_row + row_idx
            ws.cell(row=excel_row, column=subject_col, value=row.owner_name)

            for month in range(1, 13):
                if month not in month_cols:
                    continue
                col = month_cols[month]
                cell_data = row.months[month - 1]
                cell = ws.cell(row=excel_row, column=col)
                cell.value = scale_amount(cell_data.value, amount_divisor) if abs(cell_data.value) > 1e-9 else None
                if cell_data.source == "actual":
                    cell.fill = gray_fill
                    cell.protection = locked
                elif cell_data.editable:
                    cell.protection = unlocked
                else:
                    cell.protection = locked

            month_col_letters = [get_column_letter(month_cols[month]) for month in sorted(month_cols) if month in month_cols]
            if total_col is not None:
                formula_parts = [f"{letter}{excel_row}" for letter in month_col_letters]
                ws.cell(row=excel_row, column=total_col, value=f"={'+'.join(formula_parts)}")

            if annual_budget_col is not None:
                ws.cell(row=excel_row, column=annual_budget_col, value=scale_amount(row.annual_budget, amount_divisor))
                ws.cell(row=excel_row, column=annual_budget_col).fill = gray_fill
                ws.cell(row=excel_row, column=annual_budget_col).protection = locked

            if gap_col is not None and total_col is not None and annual_budget_col is not None:
                cl_total = get_column_letter(total_col)
                cl_budget = get_column_letter(annual_budget_col)
                ws.cell(row=excel_row, column=gap_col, value=f"={cl_total}{excel_row}-{cl_budget}{excel_row}")
                ws.cell(row=excel_row, column=gap_col).fill = gray_fill
                ws.cell(row=excel_row, column=gap_col).protection = locked

            if rate_col is not None and total_col is not None and annual_budget_col is not None:
                cl_total = get_column_letter(total_col)
                cl_budget = get_column_letter(annual_budget_col)
                ws.cell(
                    row=excel_row,
                    column=rate_col,
                    value=f'=IF({cl_budget}{excel_row}=0,"",{cl_total}{excel_row}/{cl_budget}{excel_row})',
                )
                ws.cell(row=excel_row, column=rate_col).fill = gray_fill
                ws.cell(row=excel_row, column=rate_col).protection = locked

            if biz_col is not None:
                ws.cell(row=excel_row, column=biz_col, value=scale_amount(row.business_submission, amount_divisor))
                if row.business_submission_editable:
                    ws.cell(row=excel_row, column=biz_col).protection = unlocked
                else:
                    ws.cell(row=excel_row, column=biz_col).fill = gray_fill
                    ws.cell(row=excel_row, column=biz_col).protection = locked

            if capital_col is not None:
                ws.cell(row=excel_row, column=capital_col, value=scale_amount(row.capital_advice, amount_divisor))
                if row.capital_advice_editable:
                    ws.cell(row=excel_row, column=capital_col).protection = unlocked
                else:
                    ws.cell(row=excel_row, column=capital_col).fill = gray_fill
                    ws.cell(row=excel_row, column=capital_col).protection = locked

            if capital_gap_col is not None and capital_col is not None and biz_col is not None:
                cl_biz = get_column_letter(biz_col)
                cl_cap = get_column_letter(capital_col)
                ws.cell(row=excel_row, column=capital_gap_col, value=f"={cl_cap}{excel_row}-{cl_biz}{excel_row}")
                ws.cell(row=excel_row, column=capital_gap_col).fill = gray_fill
                ws.cell(row=excel_row, column=capital_gap_col).protection = locked
    else:
        child_excel_rows_by_parent: dict[int, list[int]] = defaultdict(list)
        for row_idx, row in enumerate(view.rows):
            if row.parent_id is None:
                continue
            child_excel_rows_by_parent[row.parent_id].append(data_start_row + row_idx)

        for row_idx, row in enumerate(view.rows):
            excel_row = data_start_row + row_idx
            indent = "  " * max(0, int(row.level_number) - 1)
            is_parent = not row.is_leaf or bool(row.formula_text)
            is_actual_month = lambda month: month <= actual_cutoff_month
            child_rows_in_view = child_excel_rows_by_parent.get(row.id, [])

            cell_subject = ws.cell(row=excel_row, column=subject_col, value=f"{indent}{row.subject_name}")
            if is_parent:
                cell_subject.font = bold_font
                cell_subject.fill = gray_fill
                cell_subject.protection = locked

            for month in range(1, 13):
                if month not in month_cols:
                    continue
                col = month_cols[month]
                cell_data = row.months[month - 1]
                is_actual = is_actual_month(month)
                cell = ws.cell(row=excel_row, column=col)

                if is_parent:
                    if child_rows_in_view:
                        col_letter = get_column_letter(col)
                        parts = [f"{col_letter}{child_row}" for child_row in child_rows_in_view]
                        cell.value = f"={'+'.join(parts)}"
                    else:
                        cell.value = scale_amount(cell_data.value, amount_divisor)
                    cell.fill = gray_fill
                    cell.protection = locked
                elif is_actual:
                    cell.value = scale_amount(cell_data.value, amount_divisor)
                    cell.fill = gray_fill
                    cell.protection = locked
                else:
                    if cell_data.editable:
                        cell.value = None
                        cell.protection = unlocked
                    else:
                        cell.value = scale_amount(cell_data.value, amount_divisor) if abs(cell_data.value) > 1e-9 else None
                        cell.protection = locked

            month_col_letters = [get_column_letter(month_cols[month]) for month in sorted(month_cols) if month in month_cols]
            if total_col is not None:
                col_letter_total = get_column_letter(total_col)
                if is_parent:
                    if child_rows_in_view:
                        parts = [f"{col_letter_total}{child_row}" for child_row in child_rows_in_view]
                        ws.cell(row=excel_row, column=total_col, value=f"={'+'.join(parts)}")
                    else:
                        ws.cell(row=excel_row, column=total_col, value=scale_amount(row.total_value, amount_divisor))
                    ws.cell(row=excel_row, column=total_col).fill = gray_fill
                    ws.cell(row=excel_row, column=total_col).protection = locked
                else:
                    formula_parts = [f"{letter}{excel_row}" for letter in month_col_letters]
                    ws.cell(row=excel_row, column=total_col, value=f"={'+'.join(formula_parts)}")
                    if is_actual_month(12):
                        ws.cell(row=excel_row, column=total_col).fill = gray_fill
                        ws.cell(row=excel_row, column=total_col).protection = locked

            if annual_budget_col is not None:
                if is_parent and child_rows_in_view:
                    cl_budget = get_column_letter(annual_budget_col)
                    parts = [f"{cl_budget}{child_row}" for child_row in child_rows_in_view]
                    ws.cell(row=excel_row, column=annual_budget_col, value=f"={'+'.join(parts)}")
                else:
                    ws.cell(row=excel_row, column=annual_budget_col, value=scale_amount(row.annual_budget, amount_divisor))
                ws.cell(row=excel_row, column=annual_budget_col).fill = gray_fill
                ws.cell(row=excel_row, column=annual_budget_col).protection = locked

            if gap_col is not None and total_col is not None and annual_budget_col is not None:
                cl_total = get_column_letter(total_col)
                cl_budget = get_column_letter(annual_budget_col)
                ws.cell(row=excel_row, column=gap_col, value=f"={cl_total}{excel_row}-{cl_budget}{excel_row}")
                ws.cell(row=excel_row, column=gap_col).fill = gray_fill
                ws.cell(row=excel_row, column=gap_col).protection = locked

            if rate_col is not None and total_col is not None and annual_budget_col is not None:
                cl_total = get_column_letter(total_col)
                cl_budget = get_column_letter(annual_budget_col)
                ws.cell(row=excel_row, column=rate_col, value=f'=IF({cl_budget}{excel_row}=0,"",{cl_total}{excel_row}/{cl_budget}{excel_row})')
                ws.cell(row=excel_row, column=rate_col).fill = gray_fill
                ws.cell(row=excel_row, column=rate_col).protection = locked

            if biz_col is not None:
                if is_parent:
                    if child_rows_in_view:
                        cl_biz = get_column_letter(biz_col)
                        parts = [f"{cl_biz}{child_row}" for child_row in child_rows_in_view]
                        ws.cell(row=excel_row, column=biz_col, value=f"={'+'.join(parts)}")
                    else:
                        ws.cell(row=excel_row, column=biz_col, value=scale_amount(row.business_submission, amount_divisor))
                    ws.cell(row=excel_row, column=biz_col).fill = gray_fill
                    ws.cell(row=excel_row, column=biz_col).protection = locked
                else:
                    ws.cell(row=excel_row, column=biz_col, value=scale_amount(row.business_submission, amount_divisor))
                    if row.business_submission_editable:
                        ws.cell(row=excel_row, column=biz_col).protection = unlocked
                    else:
                        ws.cell(row=excel_row, column=biz_col).fill = gray_fill
                        ws.cell(row=excel_row, column=biz_col).protection = locked

            if capital_col is not None:
                if is_parent:
                    if child_rows_in_view:
                        cl_cap = get_column_letter(capital_col)
                        parts = [f"{cl_cap}{child_row}" for child_row in child_rows_in_view]
                        ws.cell(row=excel_row, column=capital_col, value=f"={'+'.join(parts)}")
                    else:
                        ws.cell(row=excel_row, column=capital_col, value=scale_amount(row.capital_advice, amount_divisor))
                    ws.cell(row=excel_row, column=capital_col).fill = gray_fill
                    ws.cell(row=excel_row, column=capital_col).protection = locked
                else:
                    ws.cell(row=excel_row, column=capital_col, value=scale_amount(row.capital_advice, amount_divisor))
                    if row.capital_advice_editable:
                        ws.cell(row=excel_row, column=capital_col).protection = unlocked
                    else:
                        ws.cell(row=excel_row, column=capital_col).fill = gray_fill
                        ws.cell(row=excel_row, column=capital_col).protection = locked

            if capital_gap_col is not None and capital_col is not None and biz_col is not None:
                cl_biz = get_column_letter(biz_col)
                cl_cap = get_column_letter(capital_col)
                ws.cell(row=excel_row, column=capital_gap_col, value=f"={cl_cap}{excel_row}-{cl_biz}{excel_row}")
                ws.cell(row=excel_row, column=capital_gap_col).fill = gray_fill
                ws.cell(row=excel_row, column=capital_gap_col).protection = locked

    return _finalize_workbook(wb, ws, cols), display_file_name


def build_expense_forecast_group_export_workbook(
    *,
    year: int,
    forecast_version: str,
    group_name: str,
    amount_unit: str,
    exclude_fields: list[str],
    actual_cutoff: int,
    owner_names: list[str],
    owner_views: dict[str, Any],
) -> tuple[BytesIO, str]:
    """Build the group-level expense forecast workbook without HTTP concerns."""
    amount_unit_label, amount_divisor = amount_unit_meta(amount_unit)
    cols = build_export_columns(exclude_fields)
    month_cols = cols["month_cols"]
    total_col = cols["total_col"]
    annual_budget_col = cols["annual_budget_col"]
    gap_col = cols["gap_col"]
    rate_col = cols["rate_col"]
    biz_col = cols["biz_col"]
    capital_col = cols["capital_col"]
    capital_gap_col = cols["capital_gap_col"]

    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    locked = Protection(locked=True)
    unlocked = Protection(locked=False)
    bold_font = Font(bold=True)

    wb = Workbook()
    ws = wb.active
    ws.title = f"费用预测表-{group_name}"
    ws.append(["费用预测表（按事业群导出）"])
    ws.append(
        [
            f"年份：{year}",
            f"版本：{forecast_version}",
            f"事业群：{group_name}",
            "编制口径：费用归属部门",
            f"实际截至月：{actual_cutoff}月",
            f"单位：{amount_unit_label}",
        ]
    )
    _append_header(ws, header_label="预算科目", cols=cols)

    data_start_row = 4
    subject_col = 1
    current_row = data_start_row

    group_row_num = current_row
    cell_subject = ws.cell(row=group_row_num, column=subject_col, value=group_name)
    cell_subject.font = bold_font
    cell_subject.fill = gray_fill
    cell_subject.protection = locked
    current_row += 1

    template_view = next(iter(owner_views.values()), None)
    template_rows = template_view.rows if template_view else []

    subject_key_to_owner_excel_rows: dict[int, list[int]] = {}
    subject_key_to_group_excel_row: dict[int, int] = {}

    for row in template_rows:
        excel_row = current_row
        subject_key_to_group_excel_row[row.id] = excel_row

        indent = "  " * max(0, int(row.level_number) + 1)
        cell_subject = ws.cell(row=excel_row, column=subject_col, value=f"{indent}{row.subject_name}")
        cell_subject.font = bold_font
        cell_subject.fill = gray_fill
        cell_subject.protection = locked

        for month in range(1, 13):
            if month not in month_cols:
                continue
            cell = ws.cell(row=excel_row, column=month_cols[month])
            cell.fill = gray_fill
            cell.protection = locked

        for col in [total_col, annual_budget_col, gap_col, rate_col, biz_col, capital_col, capital_gap_col]:
            if col is None:
                continue
            cell = ws.cell(row=excel_row, column=col)
            cell.fill = gray_fill
            cell.protection = locked

        current_row += 1

    group_top_subject_rows = [
        subject_key_to_group_excel_row[row.id]
        for row in template_rows
        if row.level_number <= 1
    ]

    for owner_name in owner_names:
        view = owner_views[owner_name]

        cell_subject = ws.cell(row=current_row, column=subject_col, value=f"  {owner_name}")
        cell_subject.font = bold_font
        cell_subject.fill = gray_fill
        cell_subject.protection = locked

        for month in range(1, 13):
            if month not in month_cols:
                continue
            cell = ws.cell(row=current_row, column=month_cols[month])
            cell.fill = gray_fill
            cell.protection = locked

        for col in [total_col, annual_budget_col, gap_col, rate_col, biz_col, capital_col, capital_gap_col]:
            if col is None:
                continue
            cell = ws.cell(row=current_row, column=col)
            cell.fill = gray_fill
            cell.protection = locked

        current_row += 1

        dept_data_start = current_row
        child_excel_rows_by_parent: dict[int, list[int]] = defaultdict(list)
        for row_idx, row in enumerate(view.rows):
            if row.parent_id is None:
                continue
            child_excel_rows_by_parent[row.parent_id].append(dept_data_start + row_idx)

        for row_idx, row in enumerate(view.rows):
            excel_row = dept_data_start + row_idx
            subject_key_to_owner_excel_rows.setdefault(row.id, []).append(excel_row)

            indent = "  " * max(0, int(row.level_number) + 1)
            is_parent = not row.is_leaf or bool(row.formula_text)
            child_rows_in_view = child_excel_rows_by_parent.get(row.id, [])

            cell_subject = ws.cell(row=excel_row, column=subject_col, value=f"{indent}{row.subject_name}")
            if is_parent:
                cell_subject.font = bold_font
                cell_subject.fill = gray_fill
                cell_subject.protection = locked

            for month in range(1, 13):
                if month not in month_cols:
                    continue
                col = month_cols[month]
                cell_data = row.months[month - 1]
                is_actual = month <= actual_cutoff
                cell = ws.cell(row=excel_row, column=col)

                if is_parent:
                    if child_rows_in_view:
                        col_letter = get_column_letter(col)
                        parts = [f"{col_letter}{child_row}" for child_row in child_rows_in_view]
                        cell.value = f"={'+'.join(parts)}"
                    else:
                        cell.value = scale_amount(cell_data.value, amount_divisor)
                    cell.fill = gray_fill
                    cell.protection = locked
                elif is_actual:
                    cell.value = scale_amount(cell_data.value, amount_divisor)
                    cell.fill = gray_fill
                    cell.protection = locked
                else:
                    if cell_data.editable:
                        cell.value = None
                        cell.protection = unlocked
                    else:
                        cell.value = scale_amount(cell_data.value, amount_divisor) if abs(cell_data.value) > 1e-9 else None
                        cell.protection = locked

            month_col_letters = [get_column_letter(month_cols[month]) for month in sorted(month_cols)]
            if total_col is not None:
                if is_parent:
                    if child_rows_in_view:
                        col_letter_total = get_column_letter(total_col)
                        parts = [f"{col_letter_total}{child_row}" for child_row in child_rows_in_view]
                        ws.cell(row=excel_row, column=total_col, value=f"={'+'.join(parts)}")
                    else:
                        ws.cell(row=excel_row, column=total_col, value=scale_amount(row.total_value, amount_divisor))
                    ws.cell(row=excel_row, column=total_col).fill = gray_fill
                    ws.cell(row=excel_row, column=total_col).protection = locked
                else:
                    formula_parts = [f"{letter}{excel_row}" for letter in month_col_letters]
                    ws.cell(row=excel_row, column=total_col, value=f"={'+'.join(formula_parts)}")
                    if actual_cutoff >= 12:
                        ws.cell(row=excel_row, column=total_col).fill = gray_fill
                        ws.cell(row=excel_row, column=total_col).protection = locked

            if annual_budget_col is not None:
                if is_parent and child_rows_in_view:
                    col_letter_budget = get_column_letter(annual_budget_col)
                    parts = [f"{col_letter_budget}{child_row}" for child_row in child_rows_in_view]
                    ws.cell(row=excel_row, column=annual_budget_col, value=f"={'+'.join(parts)}")
                else:
                    ws.cell(row=excel_row, column=annual_budget_col, value=scale_amount(row.annual_budget, amount_divisor))
                ws.cell(row=excel_row, column=annual_budget_col).fill = gray_fill
                ws.cell(row=excel_row, column=annual_budget_col).protection = locked

            if gap_col is not None and total_col is not None and annual_budget_col is not None:
                cl_total = get_column_letter(total_col)
                cl_budget = get_column_letter(annual_budget_col)
                ws.cell(row=excel_row, column=gap_col, value=f"={cl_total}{excel_row}-{cl_budget}{excel_row}")
                ws.cell(row=excel_row, column=gap_col).fill = gray_fill
                ws.cell(row=excel_row, column=gap_col).protection = locked

            if rate_col is not None and total_col is not None and annual_budget_col is not None:
                cl_total = get_column_letter(total_col)
                cl_budget = get_column_letter(annual_budget_col)
                ws.cell(row=excel_row, column=rate_col, value=f'=IF({cl_budget}{excel_row}=0,"",{cl_total}{excel_row}/{cl_budget}{excel_row})')
                ws.cell(row=excel_row, column=rate_col).fill = gray_fill
                ws.cell(row=excel_row, column=rate_col).protection = locked

            if biz_col is not None:
                if is_parent:
                    if child_rows_in_view:
                        cl_biz = get_column_letter(biz_col)
                        parts = [f"{cl_biz}{child_row}" for child_row in child_rows_in_view]
                        ws.cell(row=excel_row, column=biz_col, value=f"={'+'.join(parts)}")
                    else:
                        ws.cell(row=excel_row, column=biz_col, value=scale_amount(row.business_submission, amount_divisor))
                    ws.cell(row=excel_row, column=biz_col).fill = gray_fill
                    ws.cell(row=excel_row, column=biz_col).protection = locked
                else:
                    ws.cell(row=excel_row, column=biz_col, value=scale_amount(row.business_submission, amount_divisor))
                    if row.business_submission_editable:
                        ws.cell(row=excel_row, column=biz_col).protection = unlocked
                    else:
                        ws.cell(row=excel_row, column=biz_col).fill = gray_fill
                        ws.cell(row=excel_row, column=biz_col).protection = locked

            if capital_col is not None:
                if is_parent:
                    if child_rows_in_view:
                        cl_cap = get_column_letter(capital_col)
                        parts = [f"{cl_cap}{child_row}" for child_row in child_rows_in_view]
                        ws.cell(row=excel_row, column=capital_col, value=f"={'+'.join(parts)}")
                    else:
                        ws.cell(row=excel_row, column=capital_col, value=scale_amount(row.capital_advice, amount_divisor))
                    ws.cell(row=excel_row, column=capital_col).fill = gray_fill
                    ws.cell(row=excel_row, column=capital_col).protection = locked
                else:
                    ws.cell(row=excel_row, column=capital_col, value=scale_amount(row.capital_advice, amount_divisor))
                    if row.capital_advice_editable:
                        ws.cell(row=excel_row, column=capital_col).protection = unlocked
                    else:
                        ws.cell(row=excel_row, column=capital_col).fill = gray_fill
                        ws.cell(row=excel_row, column=capital_col).protection = locked

            if capital_gap_col is not None and capital_col is not None and biz_col is not None:
                cl_biz = get_column_letter(biz_col)
                cl_cap = get_column_letter(capital_col)
                ws.cell(row=excel_row, column=capital_gap_col, value=f"={cl_cap}{excel_row}-{cl_biz}{excel_row}")
                ws.cell(row=excel_row, column=capital_gap_col).fill = gray_fill
                ws.cell(row=excel_row, column=capital_gap_col).protection = locked

        current_row = dept_data_start + len(view.rows)

    for row_idx, row in enumerate(template_rows):
        group_excel_row = subject_key_to_group_excel_row.get(row.id)
        if group_excel_row is None:
            continue
        owner_rows = subject_key_to_owner_excel_rows.get(row.id, [])
        is_parent = not row.is_leaf or bool(row.formula_text)

        if is_parent:
            child_group_rows = [
                subject_key_to_group_excel_row[child_row.id]
                for child_row in template_rows[:row_idx]
                if child_row.parent_id == row.id and child_row.id in subject_key_to_group_excel_row
            ]
            for month in range(1, 13):
                if month not in month_cols:
                    continue
                col = month_cols[month]
                col_letter = get_column_letter(col)
                if child_group_rows:
                    parts = [f"{col_letter}{child_row}" for child_row in child_group_rows]
                    ws.cell(row=group_excel_row, column=col, value=f"={'+'.join(parts)}")
                elif owner_rows:
                    parts = [f"{col_letter}{owner_row}" for owner_row in owner_rows]
                    ws.cell(row=group_excel_row, column=col, value=f"={'+'.join(parts)}")

            for col in [total_col, annual_budget_col, biz_col, capital_col]:
                if col is None:
                    continue
                col_letter = get_column_letter(col)
                if child_group_rows:
                    parts = [f"{col_letter}{child_row}" for child_row in child_group_rows]
                    ws.cell(row=group_excel_row, column=col, value=f"={'+'.join(parts)}")
                elif owner_rows:
                    parts = [f"{col_letter}{owner_row}" for owner_row in owner_rows]
                    ws.cell(row=group_excel_row, column=col, value=f"={'+'.join(parts)}")
        else:
            for month in range(1, 13):
                if month not in month_cols:
                    continue
                col = month_cols[month]
                col_letter = get_column_letter(col)
                if owner_rows:
                    parts = [f"{col_letter}{owner_row}" for owner_row in owner_rows]
                    ws.cell(row=group_excel_row, column=col, value=f"={'+'.join(parts)}")

            for col in [total_col, annual_budget_col, biz_col, capital_col]:
                if col is None:
                    continue
                col_letter = get_column_letter(col)
                if owner_rows:
                    parts = [f"{col_letter}{owner_row}" for owner_row in owner_rows]
                    ws.cell(row=group_excel_row, column=col, value=f"={'+'.join(parts)}")

        if gap_col is not None and total_col is not None and annual_budget_col is not None:
            cl_total = get_column_letter(total_col)
            cl_budget = get_column_letter(annual_budget_col)
            ws.cell(row=group_excel_row, column=gap_col, value=f"={cl_total}{group_excel_row}-{cl_budget}{group_excel_row}")
        if rate_col is not None and total_col is not None and annual_budget_col is not None:
            cl_total = get_column_letter(total_col)
            cl_budget = get_column_letter(annual_budget_col)
            ws.cell(row=group_excel_row, column=rate_col, value=f'=IF({cl_budget}{group_excel_row}=0,"",{cl_total}{group_excel_row}/{cl_budget}{group_excel_row})')
        if capital_gap_col is not None and capital_col is not None and biz_col is not None:
            cl_biz = get_column_letter(biz_col)
            cl_cap = get_column_letter(capital_col)
            ws.cell(row=group_excel_row, column=capital_gap_col, value=f"={cl_cap}{group_excel_row}-{cl_biz}{group_excel_row}")

    for month in range(1, 13):
        if month not in month_cols:
            continue
        col = month_cols[month]
        col_letter = get_column_letter(col)
        if group_top_subject_rows:
            parts = [f"{col_letter}{row}" for row in group_top_subject_rows]
            ws.cell(row=group_row_num, column=col, value=f"={'+'.join(parts)}")
        ws.cell(row=group_row_num, column=col).fill = gray_fill
        ws.cell(row=group_row_num, column=col).protection = locked

    for col in [total_col, annual_budget_col, biz_col, capital_col]:
        if col is None:
            continue
        col_letter = get_column_letter(col)
        if group_top_subject_rows:
            parts = [f"{col_letter}{row}" for row in group_top_subject_rows]
            ws.cell(row=group_row_num, column=col, value=f"={'+'.join(parts)}")
        ws.cell(row=group_row_num, column=col).fill = gray_fill
        ws.cell(row=group_row_num, column=col).protection = locked

    if gap_col is not None and total_col is not None and annual_budget_col is not None:
        cl_total = get_column_letter(total_col)
        cl_budget = get_column_letter(annual_budget_col)
        ws.cell(row=group_row_num, column=gap_col, value=f"={cl_total}{group_row_num}-{cl_budget}{group_row_num}")
        ws.cell(row=group_row_num, column=gap_col).fill = gray_fill
        ws.cell(row=group_row_num, column=gap_col).protection = locked
    if rate_col is not None and total_col is not None and annual_budget_col is not None:
        cl_total = get_column_letter(total_col)
        cl_budget = get_column_letter(annual_budget_col)
        ws.cell(row=group_row_num, column=rate_col, value=f'=IF({cl_budget}{group_row_num}=0,"",{cl_total}{group_row_num}/{cl_budget}{group_row_num})')
        ws.cell(row=group_row_num, column=rate_col).fill = gray_fill
        ws.cell(row=group_row_num, column=rate_col).protection = locked
    if capital_gap_col is not None and capital_col is not None and biz_col is not None:
        cl_biz = get_column_letter(biz_col)
        cl_cap = get_column_letter(capital_col)
        ws.cell(row=group_row_num, column=capital_gap_col, value=f"={cl_cap}{group_row_num}-{cl_biz}{group_row_num}")
        ws.cell(row=group_row_num, column=capital_gap_col).fill = gray_fill
        ws.cell(row=group_row_num, column=capital_gap_col).protection = locked

    display_file_name = f"费用预测表_{year}_{group_name}.xlsx"
    return _finalize_workbook(wb, ws, cols), display_file_name
