"""Excel export builder for expense budget execution reports."""
from __future__ import annotations

from dataclasses import dataclass
import re
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.services.expense_budget_execution_modes import EXPORT_REPORT_MODES, REPORT_PERSPECTIVES
from app.services.export_common import workbook_streaming_response


AMOUNT_UNIT_DIVISORS: dict[str, tuple[str, float]] = {
    "yuan": ("元", 1.0),
    "thousand": ("千元", 1_000.0),
    "ten_thousand": ("万元", 10_000.0),
    "million": ("百万元", 1_000_000.0),
    "hundred_million": ("亿元", 100_000_000.0),
}
EXPORT_WORKBOOK_KINDS = frozenset({"monthly", "template", "subject", "flat"})


class ExpenseBudgetExecutionExportError(Exception):
    """Raised when an expense execution export cannot be built."""


@dataclass(frozen=True)
class ExpenseBudgetExecutionExportOptions:
    mode: str = "query"
    perspective: str = "group"
    amount_unit: str = "yuan"
    include_monthly_actuals: bool = False
    include_last_year_monthly_actuals: bool = False


@dataclass(frozen=True)
class ExpenseBudgetExecutionWorkbookExport:
    workbook: Workbook
    filename: str


@dataclass(frozen=True)
class ExpenseBudgetExecutionExportPlan:
    workbook_kind: str
    filename: str


def amount_unit_meta(amount_unit: str | None) -> tuple[str, float]:
    normalized = str(amount_unit or "").strip().lower()
    if not normalized:
        return AMOUNT_UNIT_DIVISORS["yuan"]
    if normalized not in AMOUNT_UNIT_DIVISORS:
        raise ExpenseBudgetExecutionExportError(f"未知费用预算执行金额单位: {amount_unit}")
    return AMOUNT_UNIT_DIVISORS[normalized]


def _scale_amount(value: float | int | None, divisor: float) -> float:
    digits = 0 if divisor == 1.0 else 2
    return round(float(value or 0.0) / divisor, digits)


def _text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return re.sub(r"\s+", " ", text)


def _sheet_title(raw_title: str, fallback: str) -> str:
    text = _text(raw_title) or fallback
    sanitized = re.sub(r'[\\/*?:\[\]]', "_", text)
    return sanitized[:31] or fallback


def _set_header_row(ws: Any, row_idx: int, headers: list[str], *, min_width: int = 14, padding: int = 4) -> None:
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=row_idx, column=col_idx, value=header)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + padding, min_width)


def _write_metric_rows_sheet(
    ws: Any,
    *,
    sheet_title: str,
    note: str,
    label_header: str,
    rows: list[dict[str, Any]],
    current_month: int,
    amount_unit_label: str,
    amount_divisor: float,
    options: ExpenseBudgetExecutionExportOptions,
) -> None:
    ws.cell(row=1, column=1, value=sheet_title)
    ws.cell(row=2, column=1, value=note)
    ws.cell(row=3, column=1, value=f"单位：{amount_unit_label}")
    headers = [
        label_header,
        "预算科目",
        "本年实际",
    ]
    if options.include_monthly_actuals:
        headers.extend([f"{idx}月实际" for idx in range(1, current_month + 1)])
    headers.extend(["本年预算", "预算进度%", "同比+-", "同比%", "本月环比增减额", "本月环比%", "去年同期"])
    if options.include_last_year_monthly_actuals:
        headers.extend([f"去年{idx}月实际" for idx in range(1, 13)])
    _set_header_row(ws, 5, headers)

    row_idx = 6
    for row in rows:
        ws.cell(row=row_idx, column=1, value=row.get("label"))
        indent = "  " * max(int(row.get("level") or 0), 0)
        ws.cell(row=row_idx, column=2, value=f"{indent}{_text(row.get('subject_name'))}")
        col_idx = 3
        ws.cell(row=row_idx, column=col_idx, value=_scale_amount(row.get("current_actual"), amount_divisor))
        col_idx += 1
        if options.include_monthly_actuals:
            for month_value in list(row.get("monthly_actuals") or [])[:current_month]:
                ws.cell(row=row_idx, column=col_idx, value=_scale_amount(month_value, amount_divisor))
                col_idx += 1
        ws.cell(row=row_idx, column=col_idx, value=_scale_amount(row.get("annual_budget"), amount_divisor))
        progress_cell = ws.cell(row=row_idx, column=col_idx + 1, value=row.get("budget_progress"))
        progress_cell.number_format = "0.00%"
        ws.cell(row=row_idx, column=col_idx + 2, value=_scale_amount(row.get("yoy_change"), amount_divisor))
        yoy_rate_cell = ws.cell(row=row_idx, column=col_idx + 3, value=row.get("yoy_rate"))
        yoy_rate_cell.number_format = "0.00%"
        ws.cell(row=row_idx, column=col_idx + 4, value=_scale_amount(row.get("month_over_month"), amount_divisor))
        month_over_month_rate_cell = ws.cell(row=row_idx, column=col_idx + 5, value=row.get("month_over_month_rate"))
        month_over_month_rate_cell.number_format = "0.00%"
        ws.cell(row=row_idx, column=col_idx + 6, value=_scale_amount(row.get("last_year_actual"), amount_divisor))
        next_col_idx = col_idx + 7
        if options.include_last_year_monthly_actuals:
            for month_value in list(row.get("previous_year_monthly_actuals") or [])[:12]:
                ws.cell(row=row_idx, column=next_col_idx, value=_scale_amount(month_value, amount_divisor))
                next_col_idx += 1
        row_idx += 1

    if row_idx == 6:
        ws.cell(row=6, column=1, value="当前条件下没有可展示的数据。")


def _write_matrix_rows_sheet(
    ws: Any,
    *,
    sheet_title: str,
    note: str,
    columns: list[str],
    rows: list[dict[str, Any]],
    current_month: int,
    amount_unit_label: str,
    amount_divisor: float,
    options: ExpenseBudgetExecutionExportOptions,
) -> None:
    ws.cell(row=1, column=1, value=sheet_title)
    ws.cell(row=2, column=1, value=note)
    ws.cell(row=3, column=1, value=f"单位：{amount_unit_label}")
    headers = ["部门"]
    headers.extend([f"{column}-累计实际" for column in columns])
    if options.include_monthly_actuals:
        for column in columns:
            headers.extend([f"{column}-{idx}月实际" for idx in range(1, current_month + 1)])
    headers.append("累计实际合计")
    headers.extend([f"{column}-预算" for column in columns])
    headers.append("预算合计")
    headers.extend([f"{column}-预算进度%" for column in columns])
    headers.append("预算进度合计%")
    _set_header_row(ws, 5, headers)

    row_idx = 6
    for row in rows:
        indent = "  " * max(int(row.get("level") or 0), 0)
        ws.cell(row=row_idx, column=1, value=f"{indent}{_text(row.get('label'))}")
        col_idx = 2
        actuals = row.get("actuals") or {}
        monthly_actuals_by_subject = row.get("monthly_actuals_by_subject") or {}
        budgets = row.get("budgets") or {}
        progresses = row.get("progresses") or {}
        for column in columns:
            ws.cell(row=row_idx, column=col_idx, value=_scale_amount(actuals.get(column), amount_divisor))
            col_idx += 1
        if options.include_monthly_actuals:
            for column in columns:
                for month_value in list(monthly_actuals_by_subject.get(column) or [])[:current_month]:
                    ws.cell(row=row_idx, column=col_idx, value=_scale_amount(month_value, amount_divisor))
                    col_idx += 1
        ws.cell(row=row_idx, column=col_idx, value=_scale_amount(row.get("actual_total"), amount_divisor))
        col_idx += 1
        for column in columns:
            ws.cell(row=row_idx, column=col_idx, value=_scale_amount(budgets.get(column), amount_divisor))
            col_idx += 1
        ws.cell(row=row_idx, column=col_idx, value=_scale_amount(row.get("budget_total"), amount_divisor))
        col_idx += 1
        for column in columns:
            progress_cell = ws.cell(row=row_idx, column=col_idx, value=progresses.get(column))
            progress_cell.number_format = "0.00%"
            col_idx += 1
        total_progress_cell = ws.cell(row=row_idx, column=col_idx, value=row.get("budget_progress_total"))
        total_progress_cell.number_format = "0.00%"
        row_idx += 1

    if row_idx == 6:
        ws.cell(row=6, column=1, value="当前条件下没有可展示的数据。")


def _write_tree_sheet(
    ws: Any,
    report: dict[str, Any],
    *,
    title: str,
    label_header: str,
    amount_unit_label: str,
    amount_divisor: float,
    options: ExpenseBudgetExecutionExportOptions,
) -> None:
    headers = ["层级", label_header, "本年实际"]
    current_month = int(report["current_month"])
    if options.include_monthly_actuals:
        headers.extend([f"{idx}月实际" for idx in range(1, current_month + 1)])
    headers.extend(["本年预算", "预算进度%", "本年同比增减额", "本年同比%", "本月环比增减额", "本月环比%", "去年同期"])
    if options.include_last_year_monthly_actuals:
        previous_year_short = str(int(report["budget_year"]) - 1)[-2:]
        headers.extend([f"{previous_year_short}年{idx}月实际" for idx in range(1, 13)])
    ws.cell(row=1, column=1, value=title)
    ws.cell(row=2, column=1, value=report["note"])
    ws.cell(row=3, column=1, value=f"单位：{amount_unit_label}")
    _set_header_row(ws, 5, headers)

    def write_nodes(nodes: list[dict[str, Any]], row_idx: int) -> int:
        current_row = row_idx
        for node in nodes:
            node_row = current_row
            ws.cell(row=node_row, column=1, value=node["level_label"])
            indent = "  " * max(int(node["level_number"]) - 1, 0)
            ws.cell(row=node_row, column=2, value=f"{indent}{node['subject_name']}")
            current_row += 1
            current_row = write_nodes(list(node.get("children", [])), current_row)
            child_rows = list(range(node_row + 1, current_row))

            col_idx = 4
            if options.include_monthly_actuals:
                for month_value in list(node.get("monthly_actuals", []))[:current_month]:
                    if child_rows:
                        col_letter = get_column_letter(col_idx)
                        parts = [f"{col_letter}{row}" for row in child_rows]
                        ws.cell(row=node_row, column=col_idx, value=f"={'+'.join(parts)}")
                    else:
                        ws.cell(row=node_row, column=col_idx, value=_scale_amount(month_value, amount_divisor))
                    col_idx += 1
            if child_rows:
                current_actual_col_letter = get_column_letter(3)
                current_actual_parts = [f"{current_actual_col_letter}{row}" for row in child_rows]
                ws.cell(row=node_row, column=3, value=f"={'+'.join(current_actual_parts)}")
                budget_col_letter = get_column_letter(col_idx)
                budget_parts = [f"{budget_col_letter}{row}" for row in child_rows]
                ws.cell(row=node_row, column=col_idx, value=f"={'+'.join(budget_parts)}")
            else:
                ws.cell(row=node_row, column=3, value=_scale_amount(node["current_actual"], amount_divisor))
                ws.cell(row=node_row, column=col_idx, value=_scale_amount(node["annual_budget"], amount_divisor))
            progress_cell = ws.cell(
                row=node_row,
                column=col_idx + 1,
                value=f'=IF({get_column_letter(col_idx)}{node_row}=0,"",{get_column_letter(3)}{node_row}/{get_column_letter(col_idx)}{node_row})',
            )
            progress_cell.number_format = "0.00%"
            ws.cell(row=node_row, column=col_idx + 4, value=_scale_amount(node.get("month_over_month"), amount_divisor))
            month_over_month_rate_cell = ws.cell(row=node_row, column=col_idx + 5, value=node.get("month_over_month_rate"))
            month_over_month_rate_cell.number_format = "0.00%"
            last_year_col = col_idx + 6
            if child_rows:
                last_year_letter = get_column_letter(last_year_col)
                parts = [f"{last_year_letter}{row}" for row in child_rows]
                ws.cell(row=node_row, column=last_year_col, value=f"={'+'.join(parts)}")
            else:
                ws.cell(row=node_row, column=last_year_col, value=_scale_amount(node["last_year_actual"], amount_divisor))
            yoy_change_cell = ws.cell(
                row=node_row,
                column=col_idx + 2,
                value=f"={get_column_letter(3)}{node_row}-{get_column_letter(last_year_col)}{node_row}",
            )
            yoy_change_cell.number_format = "0"
            yoy_rate_cell = ws.cell(
                row=node_row,
                column=col_idx + 3,
                value=f'=IF({get_column_letter(last_year_col)}{node_row}=0,"",{get_column_letter(col_idx + 2)}{node_row}/{get_column_letter(last_year_col)}{node_row})',
            )
            yoy_rate_cell.number_format = "0.00%"
            if options.include_last_year_monthly_actuals:
                last_year_col_idx = col_idx + 7
                for month_value in list(node.get("previous_year_monthly_actuals", []))[:12]:
                    if child_rows:
                        col_letter = get_column_letter(last_year_col_idx)
                        parts = [f"{col_letter}{row}" for row in child_rows]
                        ws.cell(row=node_row, column=last_year_col_idx, value=f"={'+'.join(parts)}")
                    else:
                        ws.cell(row=node_row, column=last_year_col_idx, value=_scale_amount(month_value, amount_divisor))
                    last_year_col_idx += 1
        return current_row

    end_row = write_nodes(list(report.get("subject_tree", [])), 6)
    if end_row == 6:
        ws.cell(row=6, column=1, value="当前条件下没有可展示的数据。")
        end_row += 1
    meta_row = end_row + 1
    ws.cell(row=meta_row, column=1, value="分月实际来源")
    ws.cell(row=meta_row, column=2, value=report["actual_source_file"])
    ws.cell(row=meta_row + 1, column=1, value="去年同期来源")
    ws.cell(row=meta_row + 1, column=2, value=report["previous_actual_source_file"])


def _build_query_workbook(
    report: dict[str, Any],
    *,
    amount_unit_label: str,
    amount_divisor: float,
    options: ExpenseBudgetExecutionExportOptions,
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = _sheet_title("1.费用类型树", "费用类型树")
    _write_tree_sheet(
        ws,
        report,
        title=report["template_title"],
        label_header="费用类型",
        amount_unit_label=amount_unit_label,
        amount_divisor=amount_divisor,
        options=options,
    )

    business_rows = list(report.get("monthly_business_rows", []))
    if business_rows:
        ws_business = wb.create_sheet(title=_sheet_title("2.业务费用", "业务费用"))
        _write_metric_rows_sheet(
            ws_business,
            sheet_title="2.业务费用",
            note=report["note"],
            label_header="部门",
            rows=business_rows,
            current_month=int(report["current_month"]),
            amount_unit_label=amount_unit_label,
            amount_divisor=amount_divisor,
            options=options,
        )

    it_rows = list(report.get("monthly_it_rows", []))
    if it_rows:
        ws_it = wb.create_sheet(title=_sheet_title("2.IT费用", "IT费用"))
        _write_metric_rows_sheet(
            ws_it,
            sheet_title="2.IT费用",
            note=report["note"],
            label_header="归口部门",
            rows=it_rows,
            current_month=int(report["current_month"]),
            amount_unit_label=amount_unit_label,
            amount_divisor=amount_divisor,
            options=options,
        )

    for block in list(report.get("monthly_daily_managed_blocks", [])):
        title = _text(block.get("title")) or "日常费用"
        ws_block = wb.create_sheet(title=_sheet_title(title, "日常费用"))
        _write_metric_rows_sheet(
            ws_block,
            sheet_title=title,
            note=report["note"],
            label_header="归口部门",
            rows=list(block.get("rows", [])),
            current_month=int(report["current_month"]),
            amount_unit_label=amount_unit_label,
            amount_divisor=amount_divisor,
            options=options,
        )

    ws_matrix = wb.create_sheet(title=_sheet_title("3.2 日常费用（其他）", "日常费用其他"))
    _write_matrix_rows_sheet(
        ws_matrix,
        sheet_title="3.2 日常费用（其他）",
        note=report["note"],
        columns=list(report.get("monthly_daily_other_columns", [])),
        rows=list(report.get("monthly_daily_other_rows", [])),
        current_month=int(report["current_month"]),
        amount_unit_label=amount_unit_label,
        amount_divisor=amount_divisor,
        options=options,
    )
    return wb


def _build_template_workbook(
    report: dict[str, Any],
    *,
    amount_unit_label: str,
    amount_divisor: float,
    options: ExpenseBudgetExecutionExportOptions,
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "部门模式"
    _write_tree_sheet(
        ws,
        report,
        title=report["template_title"],
        label_header="费用类型",
        amount_unit_label=amount_unit_label,
        amount_divisor=amount_divisor,
        options=options,
    )
    return wb


def _build_subject_workbook(
    report: dict[str, Any],
    *,
    amount_unit_label: str,
    amount_divisor: float,
    options: ExpenseBudgetExecutionExportOptions,
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "预算科目报表"
    _write_tree_sheet(
        ws,
        report,
        title=report["subject_title"],
        label_header="部门",
        amount_unit_label=amount_unit_label,
        amount_divisor=amount_divisor,
        options=options,
    )
    return wb


def _build_flat_workbook(
    report: dict[str, Any],
    *,
    amount_unit_label: str,
    amount_divisor: float,
    options: ExpenseBudgetExecutionExportOptions,
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "费用预算执行报表"
    headers = [
        "视角",
        "查询维度值",
        "主体",
        "事业群",
        "费用归属部门",
        "部门预算科目",
        "1月实际",
        "2月实际",
        "3月实际",
        "4月实际",
        "5月实际",
        "6月实际",
        "7月实际",
        "8月实际",
        "9月实际",
        "10月实际",
        "11月实际",
        "12月实际",
        "累计实际",
        "年度预算",
        "年度预算执行率",
        "本月环比增减额",
        "本月环比%",
    ]
    _set_header_row(ws, 2, headers, min_width=12, padding=2)
    ws.cell(row=1, column=1, value=f"单位：{amount_unit_label}")

    row_idx = 3
    for row in list(report.get("rows", [])):
        view_label = {
            "entity": "主体",
            "group": "事业群",
            "owner_dept": "费用归属部门",
        }.get(row["perspective"], row["perspective"])
        ws.cell(row=row_idx, column=1, value=view_label)
        ws.cell(row=row_idx, column=2, value=row["dimension_value"])
        ws.cell(row=row_idx, column=3, value=row["entity_name"])
        ws.cell(row=row_idx, column=4, value=row["group_name"])
        ws.cell(row=row_idx, column=5, value=row["owner_dept"])
        ws.cell(row=row_idx, column=6, value=row["budget_subject"])
        for offset, month_value in enumerate(row["monthly_actuals"], start=7):
            ws.cell(row=row_idx, column=offset, value=_scale_amount(month_value, amount_divisor))
        ws.cell(row=row_idx, column=19, value=_scale_amount(row["cumulative_actual"], amount_divisor))
        ws.cell(row=row_idx, column=20, value=_scale_amount(row["annual_budget"], amount_divisor))
        rate_cell = ws.cell(row=row_idx, column=21, value=row["execution_rate"])
        rate_cell.number_format = "0.00%"
        ws.cell(row=row_idx, column=22, value=_scale_amount(row.get("month_over_month"), amount_divisor))
        month_over_month_rate_cell = ws.cell(row=row_idx, column=23, value=row.get("month_over_month_rate"))
        month_over_month_rate_cell.number_format = "0.00%"
        row_idx += 1

    if row_idx == 3:
        ws.cell(row=3, column=1, value="当前条件下没有可展示的数据。")
        row_idx += 1

    meta_row = row_idx + 1
    ws.cell(row=meta_row, column=1, value="说明")
    ws.cell(row=meta_row, column=2, value=report["note"])
    return wb


def resolve_expense_budget_execution_export_plan(
    options: ExpenseBudgetExecutionExportOptions,
) -> ExpenseBudgetExecutionExportPlan:
    if options.mode not in EXPORT_REPORT_MODES:
        raise ExpenseBudgetExecutionExportError(f"未知费用预算执行导出模式: {options.mode}")
    if options.mode == "query":
        return ExpenseBudgetExecutionExportPlan(
            workbook_kind="monthly",
            filename="expense_budget_execution_monthly_report.xlsx",
        )
    if options.mode == "template":
        return ExpenseBudgetExecutionExportPlan(
            workbook_kind="template",
            filename="expense_budget_execution_department_report.xlsx",
        )
    if options.mode == "subject":
        return ExpenseBudgetExecutionExportPlan(
            workbook_kind="subject",
            filename="expense_budget_execution_subject_report.xlsx",
        )
    if options.perspective not in REPORT_PERSPECTIVES:
        raise ExpenseBudgetExecutionExportError(f"未知费用预算执行导出视角: {options.perspective}")
    return ExpenseBudgetExecutionExportPlan(
        workbook_kind="flat",
        filename=f"expense_budget_execution_{options.perspective}.xlsx",
    )


def build_expense_budget_execution_workbook(
    *,
    report: dict[str, Any],
    options: ExpenseBudgetExecutionExportOptions,
    plan: ExpenseBudgetExecutionExportPlan,
) -> Workbook:
    if plan.workbook_kind not in EXPORT_WORKBOOK_KINDS:
        raise ExpenseBudgetExecutionExportError(
            f"未知费用预算执行导出工作簿类型: {plan.workbook_kind}"
        )

    amount_unit_label, amount_divisor = amount_unit_meta(options.amount_unit)

    if plan.workbook_kind == "monthly":
        return _build_query_workbook(
            report,
            amount_unit_label=amount_unit_label,
            amount_divisor=amount_divisor,
            options=options,
        )

    if plan.workbook_kind == "template":
        return _build_template_workbook(
            report,
            amount_unit_label=amount_unit_label,
            amount_divisor=amount_divisor,
            options=options,
        )

    if plan.workbook_kind == "subject":
        return _build_subject_workbook(
            report,
            amount_unit_label=amount_unit_label,
            amount_divisor=amount_divisor,
            options=options,
        )

    return _build_flat_workbook(
        report,
        amount_unit_label=amount_unit_label,
        amount_divisor=amount_divisor,
        options=options,
    )


def build_expense_budget_execution_export(
    report: dict[str, Any],
    options: ExpenseBudgetExecutionExportOptions,
) -> ExpenseBudgetExecutionWorkbookExport:
    plan = resolve_expense_budget_execution_export_plan(options)
    return ExpenseBudgetExecutionWorkbookExport(
        workbook=build_expense_budget_execution_workbook(
            report=report,
            options=options,
            plan=plan,
        ),
        filename=plan.filename,
    )


def expense_budget_execution_workbook_response(
    export: ExpenseBudgetExecutionWorkbookExport,
):
    return workbook_streaming_response(export.workbook, filename=export.filename)
