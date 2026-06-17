from __future__ import annotations

from datetime import datetime
from io import BytesIO
from pathlib import Path
import re
from typing import Any, Awaitable, Callable

import aiosqlite
from fastapi import HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.runtime_metric_identity import product_code_from_runtime_metric_ref
from app.db_bootstrap.budget_version import ensure_budget_version_schema
from app.core.db_paths import common_db_path
from app.formula_refs import RUNTIME_METRIC_REF_CODE_RE, extract_runtime_metric_ref_code
from app.metric_tree_paths import load_metric_tree_with_data_accounts
from app.schemas import BudgetSummaryAggregateRequest, BudgetSummaryExportPivotRequest
from app.services.export_common import autosize_worksheet_columns
from app.services.formula_engine import normalize_formula, prepare_formula_expression
from app.services.runtime_metric_refs import load_org_product_metric_refs_by_runtime_ref_code
from app.services.org_product_runtime_catalog import org_product_runtime_products_cte
from app.services.pivot_aggregate import list_budget_pivot_aggregate_rows
from app.services.pivot_aggregate_export import aggregate_workbook_response, build_pivot_aggregate_workbook

FORMULA_TREE_ORG_PRODUCT_REF_COUNT_COL = 29
FORMULA_TREE_ORG_PRODUCT_REF_SOURCE_COL = 30


def org_product_refs_for_formula_tree_data_code(
    data_code: str | None,
    refs_by_runtime_ref_code: dict[str, list[str]],
) -> list[str]:
    code = str(data_code or "").strip().upper()
    if not code:
        return []
    return list(refs_by_runtime_ref_code.get(code, []))


def month_idx_from_label(label: str | None) -> int | None:
    text = (label or "").strip().upper()
    match = re.search(r"(\d{1,2})", text)
    if not match:
        return None
    month = int(match.group(1))
    if month < 1 or month > 12:
        return None
    return month - 1


class BudgetSummaryExportService:
    def __init__(
        self,
        *,
        editable_context_provider: Callable[[], Awaitable[tuple[Path, int, int]]],
    ) -> None:
        self._editable_context_provider = editable_context_provider

    async def export_budget_pivot_aggregate(
        self,
        body: BudgetSummaryExportPivotRequest,
        output_filename: str,
    ) -> StreamingResponse:
        editable_budget_path, editable_year, _editable_vid = await self._editable_context_provider()
        async with aiosqlite.connect(editable_budget_path) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            cur_ver = await db.execute("SELECT version_id, current_month FROM version")
            version_month_map = {int(r[0]): int(r[1] or 1) for r in await cur_ver.fetchall()}
        aggregate_body = BudgetSummaryAggregateRequest(
            row_field_ids=body.row_field_ids,
            column_field_ids=body.column_field_ids,
            page_field_ids=body.page_field_ids,
            page_selections=body.page_selections,
            pivot_search_text=body.pivot_search_text,
        )
        rows = await list_budget_pivot_aggregate_rows(
            budget_path=editable_budget_path,
            body=aggregate_body,
            current_month_by_version=version_month_map,
        )
        async with aiosqlite.connect(common_db_path()) as common_db:
            await common_db.execute("PRAGMA foreign_keys = ON")
            org_product_refs = await load_org_product_metric_refs_by_runtime_ref_code(common_db)
        wb = build_pivot_aggregate_workbook(
            rows=rows,
            body=body,
            title="当前可编辑年度透视聚合结果",
            source_label=f"预算聚合表 budget_pivot_aggregate（{editable_year}）",
            org_product_refs_by_runtime_ref_code=org_product_refs,
        )
        return aggregate_workbook_response(wb, output_filename)

    async def export_budget_summary_formula_tree_workbook(
        self,
        _body: BudgetSummaryExportPivotRequest,
        version_id: int,
        budget_path: Path,
        budget_year: int,
    ) -> StreamingResponse:
        common_path = common_db_path()
        async with aiosqlite.connect(common_path) as cdb:
            await cdb.execute("PRAGMA foreign_keys = ON")
            roots = await load_metric_tree_with_data_accounts(cdb)
            org_product_refs_by_runtime_ref_code = await load_org_product_metric_refs_by_runtime_ref_code(
                cdb
            )
            cur_data = await cdb.execute(
                """
                SELECT d.data_acct_code,
                       d.data_acct_name,
                       d.budget_formula,
                       d.actual_formula,
                       d.value_type,
                       COALESCE(MIN(NULLIF(b.scope_code, '')), '') AS scope_code
                FROM data_account d
                LEFT JOIN data_account_metric_binding b
                  ON b.data_acct_code = d.data_acct_code
                 AND b.is_active = 1
                GROUP BY d.data_acct_code,
                         d.data_acct_name,
                         d.budget_formula,
                         d.actual_formula,
                         d.value_type
                """
            )
            data_rows = await cur_data.fetchall()
            cur_products = await cdb.execute(
                f"""
                {org_product_runtime_products_cte()}
                SELECT product_code, product_name
                FROM org_product_runtime_products
                WHERE product_code <> '' AND product_name <> ''
                """
            )
            product_rows = await cur_products.fetchall()

        data_meta_map: dict[str, dict[str, Any]] = {}
        product_name_map = {
            str(product_code): str(product_name or "")
            for product_code, product_name in product_rows
        }

        for code, name, budget_formula, actual_formula, value_type, scope_code in data_rows:
            code_text = str(code or "").strip().upper()
            product_code_text = str(scope_code or "").strip().upper()
            if not product_code_text:
                product_code_text = product_code_from_runtime_metric_ref(code_text) or ""
            fallback_product = ""
            fallback_dept_levels = ["", "", ""]
            if product_code_text:
                product_name = product_name_map.get(product_code_text, "")
                fallback_product = f"{product_code_text} {product_name}".strip()
            data_meta_map[code_text] = {
                "name": str(name or ""),
                "budget_formula": normalize_formula(budget_formula),
                "actual_formula": normalize_formula(actual_formula),
                "value_type": str(value_type or ""),
                "fallback_dept_levels": fallback_dept_levels,
                "fallback_product": fallback_product,
            }

        data_values: dict[str, dict[str, Any]] = {}
        current_month = 1
        version_name = ""
        async with aiosqlite.connect(budget_path) as bdb:
            await bdb.execute("PRAGMA foreign_keys = ON")
            await ensure_budget_version_schema(bdb)
            cur_ver = await bdb.execute(
                "SELECT version_name, current_month FROM version WHERE version_id = ?",
                (version_id,),
            )
            ver_row = await cur_ver.fetchone()
            if not ver_row:
                raise HTTPException(status_code=400, detail=f"版本 {version_id} 不存在")
            version_name = str(ver_row[0] or "")
            current_month = int(ver_row[1])
            cur_summary = await bdb.execute(
                """
                SELECT data_code_name, dept_level1, dept_level2, dept_level3, product_code_name,
                       month, budget_actual, value
                FROM budget_summary
                WHERE version_id = ?
                """,
                (version_id,),
            )
            summary_rows = await cur_summary.fetchall()

        for (
            data_code_name,
            dept_level1,
            dept_level2,
            dept_level3,
            product_code_name,
            month_label,
            budget_actual,
            value,
        ) in summary_rows:
            code = extract_runtime_metric_ref_code(str(data_code_name or ""))
            if not code:
                continue
            rec = data_values.setdefault(
                code,
                {
                    "data_display": str(data_code_name or code),
                    "dept_levels": ["", "", ""],
                    "product": "",
                    "budget_months": [0.0] * 12,
                    "actual_months": [0.0] * 12,
                },
            )
            if not rec["data_display"] and data_code_name:
                rec["data_display"] = str(data_code_name)
            for idx, raw in enumerate((dept_level1, dept_level2, dept_level3)):
                if not rec["dept_levels"][idx] and raw:
                    rec["dept_levels"][idx] = str(raw)
            if not rec["product"] and product_code_name:
                rec["product"] = str(product_code_name)
            month_idx = month_idx_from_label(str(month_label or ""))
            if month_idx is None:
                continue
            if int(budget_actual or 0) == 0:
                if abs(rec["budget_months"][month_idx]) < 1e-12:
                    rec["budget_months"][month_idx] = float(value or 0.0)
            else:
                if abs(rec["actual_months"][month_idx]) < 1e-12:
                    rec["actual_months"][month_idx] = float(value or 0.0)

        def _enrich_metric_data_nodes(nodes: list[dict[str, Any]]) -> None:
            for node in nodes:
                if node.get("type") == "data":
                    data_code = str(node.get("code", ""))
                    base = data_values.get(data_code, {})
                    meta = data_meta_map.get(data_code, {})
                    if base.get("data_display"):
                        node["name"] = str(base["data_display"])
                    node["dept_levels"] = list(base.get("dept_levels", meta.get("fallback_dept_levels", ["", "", ""])))
                    node["product"] = str(base.get("product", meta.get("fallback_product", "")))
                    node["budget_months"] = list(base.get("budget_months", [0.0] * 12))
                    node["actual_months"] = list(base.get("actual_months", [0.0] * 12))
                    node["budget_formula"] = str(meta.get("budget_formula", node.get("budget_formula", "")))
                    node["actual_formula"] = str(meta.get("actual_formula", node.get("actual_formula", "")))
                    node["value_type"] = str(meta.get("value_type", node.get("value_type", "")))
                    continue
                _enrich_metric_data_nodes(node.get("children") or [])

        _enrich_metric_data_nodes(roots)

        def _sort_tree(nodes: list[dict[str, Any]]) -> None:
            nodes.sort(key=lambda n: str(n.get("code", "")))
            for node in nodes:
                children = node.get("children") or []
                if children:
                    _sort_tree(children)

        _sort_tree(roots)

        row_entries: list[dict[str, Any]] = []

        def _walk(node: dict[str, Any], metric_level: int) -> None:
            if node.get("type") == "metric":
                entry = {
                    "type": "metric",
                    "node": node,
                    "metric_level": metric_level,
                    "depth": metric_level,
                }
                node["_entry"] = entry
                row_entries.append(entry)
                for child in node.get("children") or []:
                    _walk(child, metric_level + 1)
            else:
                entry = {
                    "type": "data",
                    "node": node,
                    "metric_level": metric_level - 1,
                    "depth": metric_level,
                }
                node["_entry"] = entry
                row_entries.append(entry)

        for root in roots:
            _walk(root, 1)

        x = max(1, min(13, current_month))

        def _month_uses_actual(month_num: int) -> bool:
            if x == 13:
                return True
            if x == 1:
                return False
            return month_num < x

        month_headers = [
            (f"实际M{m:02d}" if _month_uses_actual(m) else f"预算M{m:02d}")
            for m in range(1, 13)
        ]

        wb = Workbook()
        ws = wb.active
        ws.title = "导出带公式树状表"
        ws.cell(
            row=1,
            column=1,
            value=f"当前年度：{budget_year}__当前版本号：{version_id}__版本名称：{version_name}",
        )
        headers = [
            "指标1级",
            "指标2级",
            "指标3级",
            "指标4级",
            "指标5级",
            "机构及产品指标编码",
            "数值类型",
            "部门科目1级",
            "部门科目2级",
            "部门科目3级",
            "机构及产品",
            *month_headers,
            "季度Q1",
            "季度Q2",
            "季度Q3",
            "季度Q4",
            "年度合计",
            "机构产品引用数量",
            "机构产品来源",
        ]
        header_fill = PatternFill(fill_type="solid", fgColor="FF000000")
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)

        for idx, entry in enumerate(row_entries):
            entry["row_num"] = idx + 3

        data_row_by_code: dict[str, int] = {}
        for entry in row_entries:
            if entry["type"] == "data":
                code = str(entry["node"].get("code", ""))
                if code and code not in data_row_by_code:
                    data_row_by_code[code] = int(entry["row_num"])

        def _weighted_sum_formula_for_children(col_idx: int, children: list[dict[str, Any]]) -> str:
            terms: list[tuple[int, str]] = []
            for ch in children:
                if not isinstance(ch, dict) or "_entry" not in ch:
                    continue
                row_ref = f"{get_column_letter(col_idx)}{int(ch['_entry']['row_num'])}"
                if ch.get("type") == "metric":
                    if int(ch.get("is_summary", 0)) != 1:
                        continue
                    sign = -1 if int(ch.get("is_minus", 0)) == 1 else 1
                    terms.append((sign, row_ref))
                else:
                    terms.append((1, row_ref))
            if not terms:
                return "0"
            expr = ""
            for idx, (sign, ref) in enumerate(terms):
                if idx == 0:
                    expr += f"-{ref}" if sign < 0 else ref
                else:
                    expr += f"-{ref}" if sign < 0 else f"+{ref}"
            return f"={expr}"

        gray_fill = PatternFill(fill_type="solid", fgColor="EDEDED")
        top_level_fill = PatternFill(fill_type="solid", fgColor="FFC0C0C0")
        formula_fill = PatternFill(fill_type="solid", fgColor="FFFFF2E6")
        muted_font = Font(color="FFB3B3B3")

        max_metric_level_in_rows = 1
        for entry in row_entries:
            if entry["type"] == "metric":
                max_metric_level_in_rows = max(max_metric_level_in_rows, int(entry["metric_level"]))

        data_outline_level = min(max_metric_level_in_rows, 8)

        for entry in row_entries:
            row_num = int(entry["row_num"])
            if entry["type"] == "metric":
                node = entry["node"]
                level = max(1, min(int(entry["metric_level"]), 5))
                metric_display = f"{node.get('code', '')} {node.get('name', '')}".strip()
                if int(node.get("is_minus", 0)) == 1:
                    metric_display = f"减： {metric_display}"
                ws.cell(row=row_num, column=level, value=metric_display)
                if level == 1:
                    ws.row_dimensions[row_num].height = 17
                    for col_idx in range(1, FORMULA_TREE_ORG_PRODUCT_REF_SOURCE_COL + 1):
                        ws.cell(row=row_num, column=col_idx).fill = top_level_fill
                    for col_idx in range(1, 6):
                        ws.cell(row=row_num, column=col_idx).font = Font(
                            color="FF000000",
                            bold=True,
                            size=12,
                        )
                children = node.get("children") or []

                def _has_descendant_data(n: dict[str, Any]) -> bool:
                    if n.get("type") == "data":
                        return True
                    return any(_has_descendant_data(ch) for ch in (n.get("children") or []))

                has_data_descendant = _has_descendant_data(node)
                if int(node.get("is_summary", 0)) != 1:
                    for col_idx in range(12, 29):
                        c = ws.cell(row=row_num, column=col_idx, value="非汇总")
                        c.font = muted_font
                elif not has_data_descendant:
                    for col_idx in range(12, 29):
                        c = ws.cell(row=row_num, column=col_idx, value="无数据")
                        c.font = muted_font
                else:
                    for col_idx in range(12, 28):
                        ws.cell(
                            row=row_num,
                            column=col_idx,
                            value=_weighted_sum_formula_for_children(col_idx, children),
                        )
                    ws.cell(
                        row=row_num,
                        column=28,
                        value=f"=SUM({get_column_letter(12)}{row_num}:{get_column_letter(23)}{row_num})",
                    )

                for col_idx in range(12, 29):
                    ws.cell(row=row_num, column=col_idx).number_format = "#,##0.00;[Red]-#,##0.00"
            else:
                node = entry["node"]
                value_type = str(node.get("value_type", ""))
                ws.cell(row=row_num, column=6, value=str(node.get("name", "")))
                ws.cell(row=row_num, column=6).fill = gray_fill
                value_type_display = value_type
                if value_type == "金额":
                    value_type_display = "金额_亿元"
                elif value_type == "户数":
                    value_type_display = "户数_万户"
                ws.cell(row=row_num, column=7, value=value_type_display)
                ws.cell(row=row_num, column=7).fill = gray_fill
                dept_levels = list(node.get("dept_levels", ["", "", ""]))
                ws.cell(row=row_num, column=8, value=dept_levels[0] if len(dept_levels) > 0 else "")
                ws.cell(row=row_num, column=9, value=dept_levels[1] if len(dept_levels) > 1 else "")
                ws.cell(row=row_num, column=10, value=dept_levels[2] if len(dept_levels) > 2 else "")
                ws.cell(row=row_num, column=11, value=str(node.get("product", "")))
                ws.cell(row=row_num, column=11).fill = gray_fill

                budget_formula = normalize_formula(str(node.get("budget_formula", "")))
                actual_formula = normalize_formula(str(node.get("actual_formula", "")))
                budget_months = list(node.get("budget_months", [0.0] * 12))
                actual_months = list(node.get("actual_months", [0.0] * 12))

                def _normalize_direct_value(raw_val: float) -> float:
                    if value_type == "百分比":
                        return float(raw_val / 100.0) if abs(raw_val) > 1 else float(raw_val)
                    return float(raw_val)

                def _formula_to_excel(formula: str, col_idx: int) -> str:
                    expr = prepare_formula_expression(formula)
                    if not expr:
                        return "0"

                    def _replace_code(m: re.Match[str]) -> str:
                        code = str(m.group(0)).upper()
                        target_row = data_row_by_code.get(code)
                        if target_row is None:
                            return "0"
                        return f"{get_column_letter(col_idx)}{target_row}"

                    expr = RUNTIME_METRIC_REF_CODE_RE.sub(_replace_code, expr)
                    return f"={expr}"

                for month_idx in range(12):
                    month_num = month_idx + 1
                    month_col = 12 + month_idx
                    mcell = ws.cell(row=row_num, column=month_col)
                    use_actual = _month_uses_actual(month_num)
                    formula_to_use = actual_formula if use_actual else budget_formula
                    if formula_to_use:
                        mcell.value = _formula_to_excel(formula_to_use, month_col)
                        mcell.comment = Comment(
                            f"原始{'实际' if use_actual else '预算'}公式：{formula_to_use}",
                            "System",
                        )
                        mcell.comment.width = 300
                        mcell.fill = formula_fill
                    else:
                        source_vals = actual_months if use_actual else budget_months
                        mcell.value = _normalize_direct_value(float(source_vals[month_idx] or 0.0))
                        mcell.fill = gray_fill

                for q_idx in range(4):
                    q_start = 12 + q_idx * 3
                    q_end = q_start + 2
                    ws.cell(
                        row=row_num,
                        column=24 + q_idx,
                        value=f"=SUM({get_column_letter(q_start)}{row_num}:{get_column_letter(q_end)}{row_num})",
                    )
                ws.cell(
                    row=row_num,
                    column=28,
                    value=f"=SUM({get_column_letter(12)}{row_num}:{get_column_letter(23)}{row_num})",
                )
                org_refs = org_product_refs_for_formula_tree_data_code(
                    str(node.get("code", "")),
                    org_product_refs_by_runtime_ref_code,
                )
                ws.cell(row=row_num, column=FORMULA_TREE_ORG_PRODUCT_REF_COUNT_COL, value=len(org_refs))
                ws.cell(
                    row=row_num,
                    column=FORMULA_TREE_ORG_PRODUCT_REF_SOURCE_COL,
                    value="\n".join(org_refs),
                )

                num_fmt = "0.00%;[Red]-0.00%" if value_type == "百分比" else "#,##0.00;[Red]-#,##0.00"
                for col_idx in range(12, 29):
                    ws.cell(row=row_num, column=col_idx).number_format = num_fmt

            if entry["type"] == "metric":
                outline_level = max(int(entry["metric_level"]) - 1, 0)
            else:
                outline_level = data_outline_level
            ws.row_dimensions[row_num].outlineLevel = min(outline_level, 8)
            ws.row_dimensions[row_num].hidden = outline_level > 0

        ws.freeze_panes = "L3"
        ws.auto_filter.ref = f"A2:{get_column_letter(FORMULA_TREE_ORG_PRODUCT_REF_SOURCE_COL)}2"
        ws.sheet_view.showOutlineSymbols = True
        if ws.sheet_properties.outlinePr is not None:
            ws.sheet_properties.outlinePr.summaryBelow = True

        ws.column_dimensions.group(get_column_letter(8), get_column_letter(10), outline_level=1, hidden=False)
        for col_idx in range(8, 11):
            ws.column_dimensions[get_column_letter(col_idx)].outlineLevel = 1
        ws.column_dimensions.group(get_column_letter(12), get_column_letter(23), outline_level=1, hidden=False)
        ws.column_dimensions.group(get_column_letter(24), get_column_letter(27), outline_level=1, hidden=False)
        for start_col, end_col in ((12, 23), (24, 27)):
            for col_idx in range(start_col, end_col + 1):
                ws.column_dimensions[get_column_letter(col_idx)].outlineLevel = 1

        autosize_worksheet_columns(ws)
        for col_idx in range(1, 6):
            ws.column_dimensions[get_column_letter(col_idx)].width = 8
        for col_idx in range(12, 29):
            letter = get_column_letter(col_idx)
            ws.column_dimensions[letter].width = min(max(ws.column_dimensions[letter].width or 8, 10), 16)
        for row_idx in range(3, ws.max_row + 1):
            for col_idx in range(12, 29):
                ws.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal="right")
            ws.cell(row=row_idx, column=28).fill = top_level_fill
            ws.cell(row=row_idx, column=FORMULA_TREE_ORG_PRODUCT_REF_COUNT_COL).alignment = Alignment(
                horizontal="center"
            )
            ws.cell(row=row_idx, column=FORMULA_TREE_ORG_PRODUCT_REF_SOURCE_COL).alignment = Alignment(
                wrap_text=True,
                vertical="top",
            )

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return StreamingResponse(
            out,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="budget_summary_formula_workbook.xlsx"'},
        )
