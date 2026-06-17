from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from typing import Any

import xlrd
from openpyxl import Workbook, load_workbook

from app.services.expense_actual_import_parser import (
    FrameworkContext,
    normalize_key,
)


class ExpenseBudgetEntryParseError(ValueError):
    """Raised when budget entry workbook cannot be parsed."""


REQUIRED_COLUMNS = ("部门", "预算科目", "预算金额")

HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "部门": ("部门", "费用归属部门", "费用部门", "责任中心", "归口管理部门", "费用发生部门"),
    "预算科目": ("预算科目", "科目", "科目名称", "科目描述", "预算科目名称"),
    "预算金额": ("预算金额", "金额", "预算数", "本年预算", "年度预算"),
}


@dataclass(frozen=True)
class ParsedBudgetEntryRow:
    owner_name_raw: str
    owner_name_mapped: str | None
    budget_subject_raw: str
    budget_subject_mapped: str | None
    amount: float
    owner_matched: bool
    subject_matched: bool
    match_note: str | None


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_amount(value: Any) -> float:
    raw = _text(value)
    if not raw:
        return 0.0
    normalized = raw.replace(",", "").replace("，", "")
    try:
        return float(normalized)
    except ValueError as exc:
        raise ExpenseBudgetEntryParseError(f"无法解析预算金额：{raw}") from exc


def _build_header_idx(headers: list[str]) -> dict[str, int]:
    normalized = [_text(v).replace(" ", "").replace("\n", "") for v in headers]
    result: dict[str, int] = {}
    for canonical, aliases in HEADER_ALIASES.items():
        for idx, header in enumerate(normalized):
            if not header:
                continue
            if header == canonical or header in aliases:
                result[canonical] = idx
                break
    return result


def _resolve_owner(raw: str, ctx: FrameworkContext) -> tuple[str | None, bool]:
    text = _text(raw)
    if not text:
        return None, False
    alias = ctx.owner_alias_map.get(normalize_key(text))
    if alias and alias in ctx.owner_names:
        return alias, True
    if text in ctx.owner_names:
        return text, True
    return None, False


def _resolve_subject(raw: str, ctx: FrameworkContext) -> tuple[str | None, bool]:
    text = _text(raw)
    if not text:
        return None, False
    alias = ctx.subject_alias_map.get(normalize_key(text))
    if alias and alias in ctx.subject_names:
        return alias, True
    if text in ctx.subject_names:
        return text, True
    return None, False


def _build_match_note(*, owner_matched: bool, subject_matched: bool) -> str | None:
    notes: list[str] = []
    if not owner_matched:
        notes.append("部门未匹配")
    if not subject_matched:
        notes.append("预算科目未匹配")
    return "；".join(notes) if notes else None


def _read_sheet_rows(file_name: str, raw: bytes) -> list[list[Any]]:
    lower_name = (file_name or "").lower()
    if lower_name.endswith(".xlsx") or lower_name.endswith(".xlsm"):
        try:
            wb = load_workbook(BytesIO(raw), data_only=True)
        except Exception as exc:
            raise ExpenseBudgetEntryParseError(f"无法读取 Excel 文件：{exc}") from exc
        if not wb.sheetnames:
            raise ExpenseBudgetEntryParseError("导入文件缺少工作表")
        ws = wb.worksheets[0]
        return [[cell for cell in row] for row in ws.iter_rows(values_only=True)]

    try:
        wb = xlrd.open_workbook(file_contents=raw)
    except Exception as exc:
        raise ExpenseBudgetEntryParseError(f"无法读取 Excel 文件：{exc}") from exc
    if not wb.sheet_names():
        raise ExpenseBudgetEntryParseError("导入文件缺少工作表")
    ws = wb.sheet_by_index(0)
    return [ws.row_values(row_idx) for row_idx in range(ws.nrows)]


def parse_budget_entry_file(
    file_name: str,
    raw: bytes,
    ctx: FrameworkContext,
    *,
    amount_unit: str = "yuan",
) -> list[ParsedBudgetEntryRow]:
    sheet_rows = _read_sheet_rows(file_name, raw)
    if not sheet_rows:
        raise ExpenseBudgetEntryParseError("导入文件没有可用明细数据")

    header_idx: dict[str, int] = {}
    data_start_row = 0
    for row_idx in range(min(len(sheet_rows), 5)):
        candidate = _build_header_idx([_text(v) for v in sheet_rows[row_idx]])
        if len(candidate) >= len(REQUIRED_COLUMNS):
            header_idx = candidate
            data_start_row = row_idx + 1
            break

    if len(header_idx) < len(REQUIRED_COLUMNS):
        missing = [name for name in REQUIRED_COLUMNS if name not in header_idx]
        raise ExpenseBudgetEntryParseError(f"导入文件缺少字段：{'、'.join(missing)}")

    rows: list[ParsedBudgetEntryRow] = []

    def get_value(values: list[Any], column_name: str) -> str:
        idx = header_idx[column_name]
        if idx >= len(values):
            return ""
        return _text(values[idx])

    for values in sheet_rows[data_start_row:]:
        owner_raw = get_value(values, "部门")
        subject_raw = get_value(values, "预算科目")
        amount_raw = values[header_idx["预算金额"]] if header_idx["预算金额"] < len(values) else ""
        if not owner_raw and not subject_raw and not _text(amount_raw):
            continue
        from app.services.expense_budget_entry_units import to_base_amount

        amount = to_base_amount(_parse_amount(amount_raw), amount_unit)
        owner_mapped, owner_matched = _resolve_owner(owner_raw, ctx)
        subject_mapped, subject_matched = _resolve_subject(subject_raw, ctx)
        rows.append(
            ParsedBudgetEntryRow(
                owner_name_raw=owner_raw,
                owner_name_mapped=owner_mapped,
                budget_subject_raw=subject_raw,
                budget_subject_mapped=subject_mapped,
                amount=amount,
                owner_matched=owner_matched,
                subject_matched=subject_matched,
                match_note=_build_match_note(owner_matched=owner_matched, subject_matched=subject_matched),
            )
        )

    if not rows:
        raise ExpenseBudgetEntryParseError("导入文件没有可导入的明细行")
    return rows


def build_template_workbook() -> tuple[bytes, str]:
    wb = Workbook()
    ws = wb.active
    ws.title = "预算录入"
    ws.append(list(REQUIRED_COLUMNS))
    ws.append(["个金管理部", "业务及管理费", 100])
    out = BytesIO()
    wb.save(out)
    return out.getvalue(), "预算录入模板.xlsx"


def build_preview_response(
    file_name: str,
    budget_year: int,
    rows: list[ParsedBudgetEntryRow],
    *,
    amount_unit: str = "yuan",
):
    from app.schemas import ExpenseBudgetEntryPreviewResponse, ExpenseBudgetEntryPreviewRow

    preview_rows = [
        ExpenseBudgetEntryPreviewRow(
            owner_name_raw=row.owner_name_raw,
            owner_name_mapped=row.owner_name_mapped,
            budget_subject_raw=row.budget_subject_raw,
            budget_subject_mapped=row.budget_subject_mapped,
            amount=row.amount,
            match_status="已匹配" if row.owner_matched and row.subject_matched else "待确认",
            match_note=row.match_note,
        )
        for row in rows
    ]
    matched_rows = sum(1 for row in rows if row.owner_matched and row.subject_matched)
    unmatched_rows = len(rows) - matched_rows
    matched_preview_rows = [row for row in preview_rows if row.match_status == "已匹配"]
    unmatched_preview_rows = [row for row in preview_rows if row.match_status != "已匹配"]
    return ExpenseBudgetEntryPreviewResponse(
        file_name=file_name,
        budget_year=budget_year,
        amount_unit=amount_unit,
        row_count=len(rows),
        matched_rows=matched_rows,
        unmatched_rows=unmatched_rows,
        preview_rows=matched_preview_rows[:200],
        unmatched_preview_rows=unmatched_preview_rows[:200],
    )
