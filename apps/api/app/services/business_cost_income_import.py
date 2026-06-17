"""Excel template, preview, and import service for business cost-income values."""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any
import re
import sqlite3
import app.core.pymysql_compat  # noqa: F401 -- SQLite->MySQL compat

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.db_bootstrap.business_cost_income import ensure_business_cost_income_schema
from app.core.db_paths import budget_db_path, common_db_path
from app.services.business_cost_income_derived import is_manual_bcir_item_for_mode
from app.services.org_product_runtime_catalog import org_product_runtime_products_cte

ACTUAL_VALUE_SHEET = "实际数"
BUDGET_VALUE_SHEET = "预算数"
FORECAST_VALUE_SHEET = "预测数"
IMPORT_SHEET_SPECS: tuple[tuple[str, str], ...] = (
    (ACTUAL_VALUE_SHEET, "actual"),
    (BUDGET_VALUE_SHEET, "budget"),
    (FORECAST_VALUE_SHEET, "forecast"),
)
FIELD_LABELS = {
    "actual": "实际",
    "budget": "预算",
    "forecast": "预测",
}
SECTION_LABELS = {
    "input": "业务投入",
    "output": "业务产出",
}
SECTION_CODES = {
    "input": "input",
    "output": "output",
    "业务投入": "input",
    "业务产出": "output",
    "投入": "input",
    "产出": "output",
}
DEFAULT_ENTITY_NAME = "微众银行"
DEFAULT_IMPORT_YEAR = 2026
DEFAULT_IMPORT_MONTHS = tuple(range(1, 13))


@dataclass(frozen=True)
class BcirImportableItem:
    item_id: int
    product_code: str
    section: str
    name: str


@dataclass
class BcirActualImportCellResult:
    month: int
    value_text: str
    status: str
    reason: str | None = None


@dataclass
class BcirActualImportRowResult:
    sheet_name: str
    field: str
    excel_row: int
    entity_name: str
    group_name: str
    product_code: str
    section: str
    item_id: int | None
    item_name: str
    months: list[BcirActualImportCellResult] = field(default_factory=list)
    note: str = ""


@dataclass
class BcirActualImportParseResult:
    year: int
    rows: list[BcirActualImportRowResult] = field(default_factory=list)
    saved_cells: int = 0


def clean_text(raw: Any) -> str:
    if raw is None:
        return ""
    return str(raw).strip()


def normalize_header(raw: Any) -> str:
    return re.sub(r"\s+", "", clean_text(raw))


def parse_months(raw: str | None) -> tuple[int, ...]:
    if not raw:
        return DEFAULT_IMPORT_MONTHS
    months: list[int] = []
    seen: set[int] = set()
    for part in re.split(r"[,，、\s]+", str(raw)):
        if not part:
            continue
        month = int(part)
        if month < 1 or month > 12:
            raise ValueError("导入月份必须在 1-12 之间")
        if month not in seen:
            seen.add(month)
            months.append(month)
    return tuple(months or DEFAULT_IMPORT_MONTHS)


def header_map(ws: Any) -> dict[str, int]:
    result: dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 1):
        header = normalize_header(ws.cell(row=1, column=col_idx).value)
        if header and header not in result:
            result[header] = col_idx
    return result


def month_columns(headers: dict[str, int], months: tuple[int, ...]) -> dict[int, int]:
    result: dict[int, int] = {}
    for month in months:
        for candidate in (f"M{month:02d}", f"M{month}", f"{month}月"):
            col = headers.get(candidate)
            if col:
                result[month] = col
                break
    return result


def parse_cell_number(raw: Any) -> tuple[bool, float | None, str | None]:
    if raw is None:
        return False, None, None
    if isinstance(raw, str):
        text = raw.strip().replace(",", "").replace("，", "")
        if not text:
            return False, None, None
        try:
            return True, float(text), None
        except ValueError:
            return True, None, "单元格不是有效数字"
    if isinstance(raw, bool):
        return True, float(raw), None
    if isinstance(raw, (int, float)):
        return True, float(raw), None
    return True, None, "单元格不是有效数字"


def load_product_name_map(product_codes: list[str], *, common_path: Path | None = None) -> dict[str, str]:
    codes = [code.strip().upper() for code in product_codes if code.strip()]
    if not codes:
        return {}
    placeholders = ",".join(["?"] * len(codes))
    conn = sqlite3.connect(common_path or common_db_path())
    try:
        rows = conn.execute(
            f"""
            {org_product_runtime_products_cte()}
            SELECT product_code, product_name
            FROM org_product_runtime_products
            WHERE UPPER(product_code) IN ({placeholders})
            """,
            codes,
        ).fetchall()
    finally:
        conn.close()
    return {str(row[0]).upper(): str(row[1] or row[0]) for row in rows}


def dedupe_importable_items(items: list[BcirImportableItem]) -> list[BcirImportableItem]:
    """Keep one import row per product/section/item name (lowest item id wins)."""
    best_by_key: dict[tuple[str, str, str], BcirImportableItem] = {}
    key_order: list[tuple[str, str, str]] = []
    for item in items:
        key = (item.product_code, item.section, item.name.strip())
        current = best_by_key.get(key)
        if current is None:
            best_by_key[key] = item
            key_order.append(key)
        elif item.item_id < current.item_id:
            best_by_key[key] = item
    return [best_by_key[key] for key in key_order]


def normalize_product_codes(product_codes: list[str]) -> list[str]:
    normalized: list[str] = []
    seen: set[str] = set()
    for raw in product_codes:
        code = clean_text(raw).upper()
        if not code or code in seen:
            continue
        seen.add(code)
        normalized.append(code)
    return normalized


def load_importable_items(*, year: int, product_codes: list[str]) -> list[BcirImportableItem]:
    codes = normalize_product_codes(product_codes)
    if not codes:
        return []
    placeholders = ",".join(["?"] * len(codes))
    conn = sqlite3.connect(budget_db_path(year))
    conn.row_factory = sqlite3.Row
    try:
        ensure_business_cost_income_schema(conn)
        rows = conn.execute(
            f"""
            SELECT id, product_code, section, name, parent_id, sort_order, enabled, manual_entry_mode
            FROM business_cost_income_item
            WHERE enabled = 1 AND UPPER(TRIM(product_code)) IN ({placeholders})
            ORDER BY product_code, section, sort_order, id
            """,
            codes,
        ).fetchall()
        child_parent_ids = {
            int(row["parent_id"])
            for row in rows
            if row["parent_id"] is not None
        }
        items = [
            BcirImportableItem(
                item_id=int(row["id"]),
                product_code=str(row["product_code"]).strip().upper(),
                section=str(row["section"]),
                name=str(row["name"]).strip(),
            )
            for row in rows
            if int(row["id"]) not in child_parent_ids
            and is_manual_bcir_item_for_mode(
                str(row["section"]),
                str(row["name"]),
                has_children=False,
                manual_entry_mode=str(row["manual_entry_mode"] or "disabled"),
            )
        ]
        return dedupe_importable_items(items)
    finally:
        conn.commit()
        conn.close()


def fill_import_sheet(ws: Any, *, product_codes: list[str], items: list[BcirImportableItem], months: tuple[int, ...]) -> None:
    product_name_map = load_product_name_map(product_codes)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    header_font = Font(bold=True)
    month_headers = [f"M{month:02d}" for month in months]
    headers = [
        "主体",
        "事业群",
        "产品编码",
        "产品名称",
        "细项分区",
        "细项ID",
        "细项名称",
        *month_headers,
        "备注",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row_idx = 2
    for product_code in product_codes:
        product_name = product_name_map.get(product_code, product_code)
        product_items = [item for item in items if item.product_code == product_code]
        for item in product_items:
            ws.cell(row=row_idx, column=1, value=DEFAULT_ENTITY_NAME)
            ws.cell(row=row_idx, column=2, value="")
            ws.cell(row=row_idx, column=3, value=product_code)
            ws.cell(row=row_idx, column=4, value=product_name)
            ws.cell(row=row_idx, column=5, value=SECTION_LABELS.get(item.section, item.section))
            ws.cell(row=row_idx, column=6, value=item.item_id)
            ws.cell(row=row_idx, column=7, value=item.name)
            row_idx += 1

    ws.freeze_panes = "A2"
    for col_idx in range(1, len(headers) + 1):
        width = 18 if col_idx <= 7 else 12
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def build_bcir_actual_import_template_workbook(
    *,
    year: int,
    product_codes: list[str],
    months: tuple[int, ...],
) -> Workbook:
    codes = normalize_product_codes(product_codes)
    if not codes:
        raise ValueError("请至少指定一个产品编码")
    items = load_importable_items(year=year, product_codes=codes)
    if not items:
        raise ValueError("业务支出成本收入比维护中暂无可导入细项")
    wb = Workbook()
    first_sheet = True
    for sheet_name, _field in IMPORT_SHEET_SPECS:
        if first_sheet:
            ws = wb.active
            ws.title = sheet_name
            first_sheet = False
        else:
            ws = wb.create_sheet(sheet_name)
        fill_import_sheet(ws, product_codes=codes, items=items, months=months)
    return wb


def validate_import_headers(headers: dict[str, int], *, sheet_name: str) -> None:
    missing: list[str] = []
    for required in ("产品编码", "细项分区"):
        if required not in headers:
            missing.append(required)
    if "细项ID" not in headers and "细项名称" not in headers:
        missing.append("细项ID或细项名称")
    if missing:
        raise ValueError(f"「{sheet_name}」缺少必填列：{', '.join(missing)}")


def resolve_import_item(
    conn: sqlite3.Connection,
    *,
    product_code: str,
    section: str,
    item_id_text: str,
    item_name: str,
) -> tuple[int | None, str | None, str | None]:
    normalized_product = clean_text(product_code).upper()
    if item_id_text:
        try:
            item_id = int(float(item_id_text))
        except ValueError:
            return None, None, "细项ID不是有效数字"
        row = conn.execute(
            """
            SELECT id, name
            FROM business_cost_income_item
            WHERE id = ? AND section = ? AND enabled = 1
              AND UPPER(TRIM(product_code)) = ?
            """,
            (item_id, section, normalized_product),
        ).fetchone()
        if row is None:
            return None, None, "细项ID在当前产品维护表中不存在或未启用"
        return int(row[0]), str(row[1]), None

    normalized_name = clean_text(item_name)
    if not normalized_name:
        return None, None, "细项名称不能为空"
    rows = conn.execute(
        """
        SELECT id, name
        FROM business_cost_income_item
        WHERE section = ? AND TRIM(name) = ? AND enabled = 1
          AND UPPER(TRIM(product_code)) = ?
        ORDER BY id
        """,
        (section, normalized_name, normalized_product),
    ).fetchall()
    if not rows:
        return None, None, "细项名称在当前维护表中不存在或未启用"
    return int(rows[0][0]), str(rows[0][1]), None


def run_bcir_actual_excel_import(
    content: bytes,
    *,
    year: int,
    months: tuple[int, ...],
    apply: bool,
) -> BcirActualImportParseResult:
    try:
        wb = load_workbook(filename=BytesIO(content), data_only=True)
    except Exception as exc:
        raise ValueError(f"无法读取Excel文件：{exc}") from exc

    target_sheets = [(name, field) for name, field in IMPORT_SHEET_SPECS if name in wb.sheetnames]
    if not target_sheets:
        sheet_names = "、".join(name for name, _ in IMPORT_SHEET_SPECS)
        raise ValueError(f"上传文件至少需包含以下工作表之一：{sheet_names}")

    conn = sqlite3.connect(budget_db_path(year))
    conn.row_factory = sqlite3.Row
    now = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    result = BcirActualImportParseResult(year=year)
    try:
        ensure_business_cost_income_schema(conn)
        for sheet_name, field_type in target_sheets:
            ws = wb[sheet_name]
            headers = header_map(ws)
            validate_import_headers(headers, sheet_name=sheet_name)
            month_cols = month_columns(headers, months)
            if not month_cols:
                month_labels = "、".join(f"M{month:02d}" for month in months)
                raise ValueError(f"「{sheet_name}」缺少月度列（{month_labels}）")

            for excel_row in range(2, ws.max_row + 1):
                row_values = [clean_text(ws.cell(excel_row, col).value) for col in range(1, ws.max_column + 1)]
                if not any(row_values):
                    continue

                entity_name = (
                    clean_text(ws.cell(excel_row, headers["主体"]).value)
                    if "主体" in headers
                    else DEFAULT_ENTITY_NAME
                ) or DEFAULT_ENTITY_NAME
                group_name = clean_text(ws.cell(excel_row, headers["事业群"]).value) if "事业群" in headers else ""
                product_code = clean_text(ws.cell(excel_row, headers["产品编码"]).value).upper()
                section_raw = clean_text(ws.cell(excel_row, headers["细项分区"]).value)
                section = SECTION_CODES.get(section_raw, "")
                item_id_text = clean_text(ws.cell(excel_row, headers["细项ID"]).value) if "细项ID" in headers else ""
                item_name = clean_text(ws.cell(excel_row, headers["细项名称"]).value) if "细项名称" in headers else ""

                row_result = BcirActualImportRowResult(
                    sheet_name=sheet_name,
                    field=field_type,
                    excel_row=excel_row,
                    entity_name=entity_name,
                    group_name=group_name,
                    product_code=product_code,
                    section=section_raw,
                    item_id=None,
                    item_name=item_name,
                )
                if not product_code:
                    row_result.note = "产品编码为空"
                    result.rows.append(row_result)
                    continue
                if section not in {"input", "output"}:
                    row_result.note = "细项分区无效，请填写“业务投入”或“业务产出”"
                    result.rows.append(row_result)
                    continue

                item_id, resolved_name, item_error = resolve_import_item(
                    conn,
                    product_code=product_code,
                    section=section,
                    item_id_text=item_id_text,
                    item_name=item_name,
                )
                if item_error or item_id is None:
                    row_result.note = item_error or "无法匹配细项"
                    result.rows.append(row_result)
                    continue
                item_meta = conn.execute(
                    """
                    SELECT name, parent_id
                    FROM business_cost_income_item
                    WHERE id = ?
                    """,
                    (item_id,),
                ).fetchone()
                has_children = conn.execute(
                    """
                    SELECT 1
                    FROM business_cost_income_item
                    WHERE parent_id = ?
                    LIMIT 1
                    """,
                    (item_id,),
                ).fetchone() is not None
                resolved_item_name = str(item_meta[0]) if item_meta is not None else (resolved_name or item_name)
                item_mode_row = conn.execute(
                    """
                    SELECT COALESCE(manual_entry_mode, 'disabled')
                    FROM business_cost_income_item
                    WHERE id = ?
                    """,
                    (item_id,),
                ).fetchone()
                manual_entry_mode = str(item_mode_row[0] or "disabled") if item_mode_row is not None else "disabled"
                if not is_manual_bcir_item_for_mode(
                    section,
                    resolved_item_name,
                    has_children=has_children,
                    manual_entry_mode=manual_entry_mode,
                ):
                    row_result.note = "当前细项不支持手工导入"
                    result.rows.append(row_result)
                    continue
                row_result.item_id = item_id
                row_result.item_name = resolved_item_name

                has_error = False
                for month, col_idx in month_cols.items():
                    raw_value = ws.cell(excel_row, col_idx).value
                    has_value, numeric, error = parse_cell_number(raw_value)
                    value_text = clean_text(raw_value)
                    if not has_value:
                        row_result.months.append(
                            BcirActualImportCellResult(month=month, value_text=value_text, status="empty")
                        )
                        continue
                    if error or numeric is None:
                        has_error = True
                        row_result.months.append(
                            BcirActualImportCellResult(
                                month=month,
                                value_text=value_text,
                                status="error",
                                reason=error,
                            )
                        )
                        continue
                    row_result.months.append(
                        BcirActualImportCellResult(month=month, value_text=value_text, status="ready")
                    )
                    if apply:
                        conn.execute(
                            """
                            INSERT INTO business_cost_income_value (
                              year, month, entity_name, group_name, product_code,
                              item_section, item_id, field, value, update_time
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            ON CONFLICT (
                              year, month,
                              entity_name, group_name, product_code,
                              item_section, item_id, field
                            ) DO UPDATE SET value = excluded.value, update_time = excluded.update_time
                            """,
                            (
                                year,
                                month,
                                entity_name,
                                group_name,
                                product_code,
                                section,
                                item_id,
                                field_type,
                                float(numeric),
                                now,
                            ),
                        )
                        result.saved_cells += 1

                if has_error and not row_result.note:
                    row_result.note = "存在无效月度数值"
                result.rows.append(row_result)

        if apply:
            conn.commit()
    finally:
        conn.close()
    return result
