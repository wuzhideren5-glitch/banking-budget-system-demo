from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from typing import Any

from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.runtime_metric_refs import load_confirmed_org_product_runtime_ref_codes


SHEET_NAME = "预算展示配置"
HEADERS = [
    ("row_key", "展示行编码"),
    ("display_view", "展示视图"),
    ("parent_row_key", "父级展示行编码"),
    ("row_type", "行类型"),
    ("data_acct_code", "机构及产品指标编码"),
    ("display_name", "展示名称"),
    ("value_type", "值类型"),
    ("level", "层级"),
    ("sort_order", "排序"),
    ("is_active", "是否启用"),
    ("org_product_ref_count", "机构产品引用数量"),
    ("org_product_refs", "机构产品来源"),
]
HEADER_ALIASES = {
    canonical: canonical for canonical, _label in HEADERS
} | {
    label: canonical for canonical, label in HEADERS
}


@dataclass(frozen=True)
class BudgetDisplayConfigImportRow:
    row_key: str
    display_view: str
    parent_row_key: str | None
    row_type: str
    data_acct_code: str | None
    display_name: str
    value_type: str | None
    level: int
    sort_order: int
    is_active: int


def _text(raw: Any) -> str:
    return str(raw or "").replace("\u3000", " ").strip()


def _int(raw: Any, *, field_name: str, default: int | None = None) -> int:
    text = _text(raw)
    if not text:
        if default is not None:
            return default
        raise HTTPException(status_code=400, detail=f"{field_name} 不能为空")
    try:
        return int(float(text))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=f"{field_name} 必须是数字：{text}") from exc


def _active_flag(raw: Any) -> int:
    text = _text(raw).lower()
    if not text:
        return 1
    if text in {"1", "true", "yes", "y", "启用", "是"}:
        return 1
    if text in {"0", "false", "no", "n", "停用", "否"}:
        return 0
    raise HTTPException(status_code=400, detail=f"是否启用只支持 1/0 或 是/否：{raw}")


def build_budget_display_config_workbook(
    rows: list[dict[str, Any]],
    *,
    org_product_refs_by_data: dict[str, list[str]] | None = None,
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    fill = PatternFill("solid", fgColor="D9EAF7")
    for col_idx, (_canonical, label) in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = Font(bold=True)
        cell.fill = fill
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(label) + 8)
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, (canonical, _label) in enumerate(HEADERS, start=1):
            data_acct_code = _text(row.get("data_acct_code")).upper()
            org_refs = (org_product_refs_by_data or {}).get(data_acct_code, [])
            if canonical == "org_product_ref_count":
                value = len(org_refs)
            elif canonical == "org_product_refs":
                value = "\n".join(org_refs)
            else:
                value = row.get(canonical)
            ws.cell(row=row_idx, column=col_idx, value=value)
    ws.freeze_panes = "A2"
    return wb


def parse_budget_display_config_workbook(file_name: str, raw: bytes) -> list[BudgetDisplayConfigImportRow]:
    if not file_name.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="预算展示配置导入仅支持 .xlsx / .xlsm 标准模板")
    wb = load_workbook(filename=BytesIO(raw), data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    header_map: dict[str, int] = {}
    for cell in ws[1]:
        canonical = HEADER_ALIASES.get(_text(cell.value))
        if canonical:
            header_map[canonical] = int(cell.column)
    optional_fields = {"value_type", "org_product_ref_count", "org_product_refs"}
    missing = [label for canonical, label in HEADERS if canonical not in header_map and canonical not in optional_fields]
    if missing:
        raise HTTPException(status_code=400, detail=f"导入模板缺少字段：{'、'.join(missing)}")

    rows: list[BudgetDisplayConfigImportRow] = []
    seen: set[str] = set()
    for row_idx in range(2, ws.max_row + 1):
        values = {
            canonical: ws.cell(row=row_idx, column=col_idx).value
            for canonical, col_idx in header_map.items()
        }
        if not any(_text(value) for value in values.values()):
            continue
        row_key = _text(values.get("row_key"))
        if not row_key:
            raise HTTPException(status_code=400, detail=f"第 {row_idx} 行缺少展示行编码")
        if row_key in seen:
            raise HTTPException(status_code=400, detail=f"展示行编码重复：{row_key}")
        seen.add(row_key)
        data_acct_code = _text(values.get("data_acct_code")).upper() or None
        row_type = (_text(values.get("row_type")) or ("METRIC" if data_acct_code else "GROUP")).upper()
        if row_type not in {"GROUP", "METRIC"}:
            raise HTTPException(status_code=400, detail=f"第 {row_idx} 行行类型只支持 GROUP/METRIC")
        if row_type == "METRIC" and not data_acct_code:
            raise HTTPException(status_code=400, detail=f"第 {row_idx} 行 METRIC 必须填写机构及产品指标编码")
        display_name = _text(values.get("display_name"))
        if not display_name:
            raise HTTPException(status_code=400, detail=f"第 {row_idx} 行缺少展示名称")
        rows.append(
            BudgetDisplayConfigImportRow(
                row_key=row_key,
                display_view=(_text(values.get("display_view")) or "TOTAL").upper(),
                parent_row_key=_text(values.get("parent_row_key")) or None,
                row_type="METRIC" if data_acct_code else "GROUP",
                data_acct_code=data_acct_code,
                display_name=display_name,
                value_type=_text(values.get("value_type")) or None,
                level=max(1, _int(values.get("level"), field_name=f"第 {row_idx} 行层级", default=1)),
                sort_order=_int(values.get("sort_order"), field_name=f"第 {row_idx} 行排序", default=row_idx),
                is_active=_active_flag(values.get("is_active")),
            )
        )
    if not rows:
        raise HTTPException(status_code=400, detail="导入文件没有可导入的预算展示配置行")
    return rows


def _sorted_for_parent_insert(rows: list[BudgetDisplayConfigImportRow]) -> list[BudgetDisplayConfigImportRow]:
    by_key = {row.row_key: row for row in rows}
    visiting: set[str] = set()
    depth_cache: dict[str, int] = {}

    def depth(row: BudgetDisplayConfigImportRow) -> int:
        if row.row_key in depth_cache:
            return depth_cache[row.row_key]
        if row.row_key in visiting:
            raise HTTPException(status_code=400, detail=f"展示配置存在循环父级：{row.row_key}")
        visiting.add(row.row_key)
        parent = by_key.get(row.parent_row_key or "")
        value = 1 + (depth(parent) if parent else 0)
        visiting.remove(row.row_key)
        depth_cache[row.row_key] = value
        return value

    return sorted(rows, key=lambda row: (row.display_view, depth(row), row.sort_order, row.row_key))


async def apply_budget_display_config_import(
    db: Any,
    *,
    rows: list[BudgetDisplayConfigImportRow],
    mode: str,
) -> dict[str, int | str]:
    import_mode = _text(mode).lower() or "replace"
    if import_mode not in {"replace", "upsert"}:
        raise HTTPException(status_code=400, detail="导入模式只支持 replace / upsert")

    row_keys = {row.row_key for row in rows}
    parent_keys = {row.parent_row_key for row in rows if row.parent_row_key}
    missing_parents = sorted(parent_keys - row_keys)
    if import_mode == "replace" and missing_parents:
        raise HTTPException(status_code=400, detail=f"父级展示行编码不在导入文件中：{'、'.join(missing_parents[:10])}")
    if import_mode == "upsert" and missing_parents:
        placeholders = ",".join("?" for _ in missing_parents)
        cur = await db.execute(
            f"SELECT row_key FROM budget_output_display_item WHERE row_key IN ({placeholders})",
            tuple(missing_parents),
        )
        existing = {str(row[0]) for row in await cur.fetchall()}
        unresolved = sorted(set(missing_parents) - existing)
        if unresolved:
            raise HTTPException(status_code=400, detail=f"父级展示行编码不存在：{'、'.join(unresolved[:10])}")

    data_codes = sorted({row.data_acct_code for row in rows if row.data_acct_code})
    value_type_by_code: dict[str, str | None] = {}
    if data_codes:
        confirmed_codes = await load_confirmed_org_product_runtime_ref_codes(db)
        unconfirmed_codes = sorted(set(data_codes) - confirmed_codes)
        if unconfirmed_codes:
            raise HTTPException(
                status_code=400,
                detail=(
                    "机构及产品指标编码未在机构及产品指标主表中确认，不能导入预算展示配置："
                    f"{'、'.join(unconfirmed_codes[:10])}"
                ),
            )
        placeholders = ",".join("?" for _ in data_codes)
        cur = await db.execute(
            f"""
            SELECT d.data_acct_code, d.value_type
            FROM data_account d
            JOIN data_account_metric_binding b
              ON b.data_acct_code = d.data_acct_code
             AND b.is_active = 1
            WHERE d.data_acct_code IN ({placeholders})
            GROUP BY d.data_acct_code, d.value_type
            """,
            tuple(data_codes),
        )
        value_type_by_code = {str(row[0]): row[1] for row in await cur.fetchall()}
        missing_codes = sorted(set(data_codes) - set(value_type_by_code))
        if missing_codes:
            raise HTTPException(status_code=400, detail=f"机构及产品指标编码不存在或未绑定指标树：{'、'.join(missing_codes[:10])}")

    if import_mode == "replace":
        await db.execute("DELETE FROM budget_output_display_item")

    saved = 0
    metric_rows = 0
    for row in _sorted_for_parent_insert(rows):
        row_type = "METRIC" if row.data_acct_code else "GROUP"
        if row_type == "METRIC":
            metric_rows += 1
        await db.execute(
            """
            INSERT INTO budget_output_display_item(
              row_key, display_view, parent_row_key, data_acct_code, row_type,
              display_name, value_type, level, sort_order, is_active
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(row_key) DO UPDATE SET
              display_view = excluded.display_view,
              parent_row_key = excluded.parent_row_key,
              data_acct_code = excluded.data_acct_code,
              org_product_ref = NULL,
              org_product_entity_code = NULL,
              org_product_table_name = NULL,
              org_product_metric_code = NULL,
              org_product_metric_name = NULL,
              row_type = excluded.row_type,
              display_name = excluded.display_name,
              value_type = excluded.value_type,
              level = excluded.level,
              sort_order = excluded.sort_order,
              is_active = excluded.is_active,
              updated_at = CURRENT_TIMESTAMP
            """,
            (
                row.row_key,
                row.display_view,
                row.parent_row_key,
                row.data_acct_code,
                row_type,
                row.display_name,
                value_type_by_code.get(row.data_acct_code or "") or row.value_type,
                row.level,
                row.sort_order,
                row.is_active,
            ),
        )
        saved += 1
    await db.commit()
    return {
        "mode": import_mode,
        "saved_rows": saved,
        "metric_rows": metric_rows,
        "group_rows": saved - metric_rows,
    }
