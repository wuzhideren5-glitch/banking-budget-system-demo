from __future__ import annotations

import json
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


SCHEME_TO_LABEL = {
    "MANUAL": "手工/导入",
    "RESIDUAL_ALLOC": "余额分摊",
    "METRIC_EXPR": "指标表达式",
}
LABEL_TO_SCHEME = {value: key for key, value in SCHEME_TO_LABEL.items()}
BOOL_TRUE = {"是", "Y", "YES", "TRUE", "1", "启用"}
BOOL_FALSE = {"否", "N", "NO", "FALSE", "0", "停用"}
BOOL_TO_LABEL = {
    True: "是",
    False: "否",
}
PRIORITY_TO_LABEL = {
    "metric_first": "机构及产品指标编码优先",
    "inline_first": "表内维护优先",
}
LABEL_TO_PRIORITY = {value: key for key, value in PRIORITY_TO_LABEL.items()}
ALLOC_MODE_TO_LABEL = {
    "progressive": "逐月递增",
    "custom": "自定义系数",
}
LABEL_TO_ALLOC_MODE = {value: key for key, value in ALLOC_MODE_TO_LABEL.items()}
CURVE_TO_LABEL = {
    "arithmetic": "等差金额",
    "geometric": "等比比例",
}
LABEL_TO_CURVE = {value: key for key, value in CURVE_TO_LABEL.items()}
HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "年度": ("年度",),
    "版本": ("版本",),
    "费用归属部门": ("费用归属部门", "归属部门"),
    "预算科目": ("预算科目",),
    "预测逻辑": ("预测逻辑", "方案类型"),
    "启用": ("启用",),
    "允许人工覆盖": ("允许人工覆盖",),
    "自动刷新": ("自动刷新",),
    "手动重算": ("手动重算",),
    "生效开始月": ("生效开始月",),
    "生效结束月": ("生效结束月",),
    "规则优先级": ("规则优先级", "优先级"),
    "数据源优先级": ("数据源优先级",),
    "分摊模式": ("分摊模式",),
    "自动反推方式": ("自动反推方式",),
    "是否允许负数": ("是否允许负数",),
    "自定义系数JSON": ("自定义系数JSON",),
    "表达式": ("表达式",),
    "变量映射JSON": ("变量映射JSON",),
    "备注": ("备注",),
}


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _safe_int(value: Any, default: int) -> int:
    try:
        if value is None or _text(value) == "":
            return default
        return int(float(value))
    except Exception:
        return default


def _pick_header(header_map: dict[str, int], canonical_name: str) -> int | None:
    for alias in HEADER_ALIASES.get(canonical_name, (canonical_name,)):
        index = header_map.get(alias)
        if index:
            return index
    return None


def _bool_label(value: bool) -> str:
    return BOOL_TO_LABEL[bool(value)]


def _normalize_scheme(value: Any) -> str:
    raw = _text(value)
    if not raw:
        return ""
    upper = raw.upper()
    if upper in SCHEME_TO_LABEL:
        return upper
    return LABEL_TO_SCHEME.get(raw, "")


def _normalize_priority(value: Any) -> str:
    raw = _text(value)
    if not raw:
        return ""
    if raw in PRIORITY_TO_LABEL:
        return raw
    return LABEL_TO_PRIORITY.get(raw, "")


def _normalize_alloc_mode(value: Any) -> str:
    raw = _text(value)
    if not raw:
        return ""
    lower = raw.lower()
    if lower in ALLOC_MODE_TO_LABEL:
        return lower
    return LABEL_TO_ALLOC_MODE.get(raw, "")


def _normalize_curve(value: Any) -> str:
    raw = _text(value)
    if not raw:
        return ""
    lower = raw.lower()
    if lower in CURVE_TO_LABEL:
        return lower
    return LABEL_TO_CURVE.get(raw, "")


def _truthy(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    return _text(value).lower() in {"1", "true", "yes", "y", "on"}


def _parse_bool(value: Any, default: bool) -> bool:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    raw = _text(value).upper()
    if raw in {item.upper() for item in BOOL_TRUE}:
        return True
    if raw in {item.upper() for item in BOOL_FALSE}:
        return False
    return _truthy(value)


def _parse_org_product_ref_label(label: str) -> tuple[str, str, str, str]:
    source_ref, _, metric_name = _text(label).partition(" ")
    parts = source_ref.split(":", 2)
    if len(parts) != 3:
        return "", "", source_ref, metric_name
    return parts[0], parts[1], parts[2], metric_name


def _variable_code_for_candidate(index: int) -> str:
    return f"metric_{index:03d}"


def build_expense_forecast_rule_template_workbook(
    *,
    default_year: int,
    default_version: str,
    org_product_refs_by_runtime_ref_code: dict[str, tuple[str, ...] | list[str]] | None = None,
) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "预测规则模板"
    headers = [
        "年度", "版本", "费用归属部门", "预算科目", "预测逻辑", "启用", "允许人工覆盖",
        "自动刷新", "手动重算", "生效开始月", "生效结束月", "规则优先级", "数据源优先级",
        "分摊模式", "自动反推方式", "是否允许负数", "自定义系数JSON", "表达式", "变量映射JSON", "备注",
    ]
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=idx, value=header).font = Font(bold=True)
        ws.column_dimensions[get_column_letter(idx)].width = 18 if idx <= 16 else 28
    ws.freeze_panes = "A2"
    ws.cell(row=2, column=1, value=default_year)
    ws.cell(row=2, column=2, value=default_version)
    ws.cell(row=2, column=5, value=SCHEME_TO_LABEL["RESIDUAL_ALLOC"])
    ws.cell(row=2, column=6, value=_bool_label(True))
    ws.cell(row=2, column=7, value=_bool_label(True))
    ws.cell(row=2, column=8, value=_bool_label(True))
    ws.cell(row=2, column=9, value=_bool_label(True))
    ws.cell(row=2, column=10, value=1)
    ws.cell(row=2, column=11, value=12)
    ws.cell(row=2, column=12, value=100)
    ws.cell(row=2, column=13, value=PRIORITY_TO_LABEL["metric_first"])
    ws.cell(row=2, column=14, value=ALLOC_MODE_TO_LABEL["progressive"])
    ws.cell(row=2, column=15, value=CURVE_TO_LABEL["arithmetic"])
    ws.cell(row=2, column=16, value=_bool_label(False))
    ws.cell(row=2, column=20, value="余额分摊示例；指标表达式场景请填写表达式和变量映射JSON")

    option_ws = wb.create_sheet("下拉选项")
    option_sources = [
        ("预测逻辑", list(SCHEME_TO_LABEL.values())),
        ("是否", list(BOOL_TO_LABEL.values())),
        ("数据源优先级", list(PRIORITY_TO_LABEL.values())),
        ("分摊模式", list(ALLOC_MODE_TO_LABEL.values())),
        ("自动反推方式", list(CURVE_TO_LABEL.values())),
    ]
    for col_idx, (title, values) in enumerate(option_sources, start=1):
        option_ws.cell(row=1, column=col_idx, value=title).font = Font(bold=True)
        option_ws.column_dimensions[get_column_letter(col_idx)].width = 18
        for row_idx, value in enumerate(values, start=2):
            option_ws.cell(row=row_idx, column=col_idx, value=value)
    option_ws.sheet_state = "hidden"

    instructions_ws = wb.create_sheet("填写说明")
    instructions = [
        ("字段", "说明"),
        ("预测逻辑", "选择 手工/导入、余额分摊、指标表达式 之一。"),
        ("分摊模式", "仅余额分摊时使用；可选 逐月递增、自定义系数。"),
        ("自动反推方式", "仅余额分摊且分摊模式=逐月递增时使用；可选 等差金额、等比比例。"),
        ("自定义系数JSON", '仅余额分摊且分摊模式=自定义系数时使用，例如 {"7":1,"8":1,"9":1}。'),
        ("表达式", "仅指标表达式时使用，例如 base_amount * (1 + revenue_growth * factor)。"),
        ("变量映射JSON", '仅指标表达式时使用，需填写数组；source_type=metric_tree 表示读取机构及产品指标取数绑定，例如 [{"variable_code":"balance","source_type":"metric_tree","source_key":"A01.01.01.001"}]。'),
        ("机构产品指标候选", "下载模板时会从机构及产品指标表读取已确认的映射候选；可复制候选sheet中的变量映射JSON示例。"),
        ("其余空白字段", "与所选预测逻辑无关的字段可留空，系统会按默认值处理。"),
    ]
    for row_idx, row_values in enumerate(instructions, start=1):
        for col_idx, value in enumerate(row_values, start=1):
            cell = instructions_ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True)
    instructions_ws.column_dimensions["A"].width = 20
    instructions_ws.column_dimensions["B"].width = 90

    validation_map = {
        "E": "$A$2:$A$4",
        "F": "$B$2:$B$3",
        "G": "$B$2:$B$3",
        "H": "$B$2:$B$3",
        "I": "$B$2:$B$3",
        "M": "$C$2:$C$3",
        "N": "$D$2:$D$3",
        "O": "$E$2:$E$3",
        "P": "$B$2:$B$3",
    }
    for column_letter, option_range in validation_map.items():
        validation = DataValidation(
            type="list",
            formula1=f"'下拉选项'!{option_range}",
            allow_blank=True,
        )
        ws.add_data_validation(validation)
        validation.add(f"{column_letter}2:{column_letter}2000")

    candidate_ws = wb.create_sheet("机构产品指标候选")
    candidate_headers = [
        "变量名建议",
        "运行取数编码",
        "机构/产品编码",
        "指标表",
        "机构及产品指标编码",
        "机构及产品指标名称",
        "机构产品引用",
        "变量映射JSON示例",
    ]
    for col_idx, header in enumerate(candidate_headers, start=1):
        candidate_ws.cell(row=1, column=col_idx, value=header).font = Font(bold=True)
        candidate_ws.column_dimensions[get_column_letter(col_idx)].width = 22 if col_idx < 8 else 82
    row_idx = 2
    candidate_index = 1
    for data_acct_code, refs in sorted((org_product_refs_by_runtime_ref_code or {}).items()):
        clean_data_acct_code = _text(data_acct_code).upper()
        if not clean_data_acct_code:
            continue
        for label in refs:
            entity_code, table_name, metric_code, metric_name = _parse_org_product_ref_label(_text(label))
            if not entity_code or not metric_code:
                continue
            variable_code = _variable_code_for_candidate(candidate_index)
            variable_json = json.dumps(
                [
                    {
                        "variable_code": variable_code,
                        "variable_name": metric_name or metric_code,
                        "source_type": "org_product_metric",
                        "source_key": metric_code,
                        "source_subkey": entity_code,
                        "org_product_ref": f"{entity_code}:{table_name}:{metric_code}",
                    }
                ],
                ensure_ascii=False,
            )
            candidate_ws.cell(row=row_idx, column=1, value=variable_code)
            candidate_ws.cell(row=row_idx, column=2, value=clean_data_acct_code)
            candidate_ws.cell(row=row_idx, column=3, value=entity_code)
            candidate_ws.cell(row=row_idx, column=4, value=table_name)
            candidate_ws.cell(row=row_idx, column=5, value=metric_code)
            candidate_ws.cell(row=row_idx, column=6, value=metric_name)
            candidate_ws.cell(row=row_idx, column=7, value=f"{entity_code}:{table_name}:{metric_code}")
            candidate_ws.cell(row=row_idx, column=8, value=variable_json)
            row_idx += 1
            candidate_index += 1

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def parse_expense_forecast_rule_import_workbook(
    raw: bytes,
    *,
    default_year: int,
    default_version: str,
) -> list[dict[str, Any]]:
    wb = load_workbook(BytesIO(raw), data_only=False)
    ws = wb.active
    header_map: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        header_map[_text(ws.cell(1, col).value)] = col
    required = {"年度", "版本", "费用归属部门", "预算科目", "预测逻辑"}
    resolved_headers = {name: _pick_header(header_map, name) for name in HEADER_ALIASES}
    missing = [name for name in required if not resolved_headers.get(name)]
    if missing:
        raise ValueError(f"规则模板缺少列: {','.join(missing)}")

    def cell_value(row_idx: int, canonical_name: str) -> Any:
        col_idx = resolved_headers.get(canonical_name)
        if not col_idx:
            return None
        return ws.cell(row_idx, col_idx).value

    rows: list[dict[str, Any]] = []
    for row_idx in range(2, ws.max_row + 1):
        owner_name = _text(cell_value(row_idx, "费用归属部门"))
        subject_name = _text(cell_value(row_idx, "预算科目"))
        if not owner_name and not subject_name:
            continue
        raw_scheme = cell_value(row_idx, "预测逻辑")
        rows.append(
            {
                "row_number": row_idx,
                "forecast_year": _safe_int(cell_value(row_idx, "年度"), default_year),
                "forecast_version": _text(cell_value(row_idx, "版本")) or default_version,
                "owner_name": owner_name,
                "subject_name": subject_name,
                "scheme_code": _normalize_scheme(raw_scheme),
                "scheme_label": _text(raw_scheme),
                "enabled": _parse_bool(cell_value(row_idx, "启用"), True),
                "allow_manual_override": _parse_bool(cell_value(row_idx, "允许人工覆盖"), False),
                "auto_refresh_enabled": _parse_bool(cell_value(row_idx, "自动刷新"), True),
                "manual_recalc_enabled": _parse_bool(cell_value(row_idx, "手动重算"), True),
                "effective_from_month": _safe_int(cell_value(row_idx, "生效开始月"), 1),
                "effective_to_month": _safe_int(cell_value(row_idx, "生效结束月"), 12),
                "priority": _safe_int(cell_value(row_idx, "规则优先级"), 100),
                "metric_source_priority": _normalize_priority(cell_value(row_idx, "数据源优先级")) or "metric_first",
                "allocation_mode": _normalize_alloc_mode(cell_value(row_idx, "分摊模式")) or "progressive",
                "progressive_curve_type": _normalize_curve(cell_value(row_idx, "自动反推方式")) or "arithmetic",
                "allow_negative": _parse_bool(cell_value(row_idx, "是否允许负数"), False),
                "weight_json": _text(cell_value(row_idx, "自定义系数JSON")),
                "expression": _text(cell_value(row_idx, "表达式")),
                "variables_json": _text(cell_value(row_idx, "变量映射JSON")),
                "remark": _text(cell_value(row_idx, "备注")),
            }
        )
    return rows


def _parse_json_text_or_raise(raw_text: str | None, fallback: Any, field_name: str) -> Any:
    raw = _text(raw_text)
    if not raw:
        return fallback
    try:
        return json.loads(raw)
    except Exception as exc:
        raise ValueError(f"{field_name}不是合法JSON") from exc


def build_expense_forecast_rule_import_payload(
    row: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    scheme_code = _text(row.get("scheme_code"))
    if scheme_code not in {"MANUAL", "RESIDUAL_ALLOC", "METRIC_EXPR"}:
        raise ValueError(f"预测逻辑不支持: {_text(row.get('scheme_label')) or scheme_code or '空值'}")
    if not _text(row.get("owner_name")):
        raise ValueError("费用归属部门不能为空")
    if not _text(row.get("subject_name")):
        raise ValueError("预算科目不能为空")
    from_month = _safe_int(row.get("effective_from_month"), 1)
    to_month = _safe_int(row.get("effective_to_month"), 12)
    if from_month < 1 or from_month > 12 or to_month < 1 or to_month > 12 or from_month > to_month:
        raise ValueError("生效开始月/结束月填写不正确")

    params: list[dict[str, Any]] = []
    variables: list[dict[str, Any]] = []
    if scheme_code == "RESIDUAL_ALLOC":
        params.extend(
            [
                {"param_group": "scheme2", "param_key": "allocation_mode", "param_value": _text(row.get("allocation_mode")) or "progressive", "value_type": "string"},
                {"param_group": "scheme2", "param_key": "progressive_curve_type", "param_value": _text(row.get("progressive_curve_type")) or "arithmetic", "value_type": "string"},
                {"param_group": "scheme2", "param_key": "auto_direction_mode", "param_value": "auto_last_vs_avg", "value_type": "string"},
                {"param_group": "scheme2", "param_key": "last_value_source_mode", "param_value": "actual_first_then_forecast", "value_type": "string"},
                {"param_group": "scheme2", "param_key": "rounding_mode", "param_value": "last_month_adjust", "value_type": "string"},
                {"param_group": "scheme2", "param_key": "allow_negative", "param_value": "true" if bool(row.get("allow_negative")) else "false", "value_type": "string"},
            ]
        )
        if _text(row.get("allocation_mode")) == "custom":
            weight_json = _parse_json_text_or_raise(row.get("weight_json"), {}, "自定义系数JSON")
            if not isinstance(weight_json, (dict, list)):
                raise ValueError("自定义系数JSON需为对象或数组")
            params.append(
                {
                    "param_group": "scheme2",
                    "param_key": "weight_json",
                    "param_value": json.dumps(weight_json, ensure_ascii=False),
                    "value_type": "json",
                }
            )
    if scheme_code == "METRIC_EXPR":
        expression = _text(row.get("expression"))
        if not expression:
            raise ValueError("指标表达式规则必须填写表达式")
        params.append(
            {
                "param_group": "metric_expr",
                "param_key": "expression",
                "param_value": expression,
                "value_type": "string",
            }
        )
        variables_json = _parse_json_text_or_raise(row.get("variables_json"), [], "变量映射JSON")
        if not isinstance(variables_json, list):
            raise ValueError("变量映射JSON需为数组")
        for item in variables_json:
            if not isinstance(item, dict) or not _text(item.get("variable_code")):
                raise ValueError("变量映射JSON中的每一项都必须包含variable_code")
            variables.append(dict(item))
    return params, variables
