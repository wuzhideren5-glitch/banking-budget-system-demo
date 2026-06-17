"""Input-output topic overview built on the current business cost-income tables."""
from __future__ import annotations

from typing import Any, Iterable

import aiosqlite
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.core.db_paths import budget_db_path, common_db_path, list_budget_database_files
from app.services.business_cost_income_ratio import (
    amount_unit_meta,
    amount_unit_options,
    ensure_business_cost_income_tables,
    load_business_cost_income_indicators,
    load_business_cost_income_items,
    norm_dim,
    parse_year_month,
)
from app.services.runtime_metric_refs import load_org_product_metric_refs_by_runtime_ref_code
from app.services.org_product_runtime_catalog import org_product_runtime_products_cte


SECTION_LABELS = {
    "indicator": "评估指标",
    "input": "业务投入细项",
    "output": "业务产出细项",
}

EXPORT_HEADERS = [
    "视图",
    "产品编码",
    "产品名称",
    "分区",
    "层级",
    "名称",
    "机构产品主题指标",
    "机构及产品指标编码",
    "机构产品来源",
    "本年累计实际",
    "全年预算",
    "预算进度",
    "全年预测",
    "预测较预算差额",
    "差额率",
    "同比变化",
    "同比率",
    "环比变化",
    "环比率",
    "上年同期实际",
]


def _scale_amount(value: float | int | None, divisor: float) -> float:
    return round(float(value or 0.0) / divisor, 2)


def _ratio(a: float, b: float) -> float | None:
    if not b:
        return None
    return a / b


def _metrics_from_amounts(
    *,
    current_actual: float,
    annual_budget: float,
    annual_forecast: float,
    last_year_actual: float,
) -> dict[str, Any]:
    current_actual = round(float(current_actual or 0.0), 6)
    annual_budget = round(float(annual_budget or 0.0), 6)
    annual_forecast = round(float(annual_forecast or 0.0), 6)
    last_year_actual = round(float(last_year_actual or 0.0), 6)
    forecast_budget_gap = round(annual_forecast - annual_budget, 6)
    yoy_change = round(current_actual - last_year_actual, 6)
    return {
        "current_actual": current_actual,
        "annual_budget": annual_budget,
        "budget_progress": round(current_actual / annual_budget, 6) if annual_budget else None,
        "annual_forecast": annual_forecast,
        "forecast_budget_gap": forecast_budget_gap,
        "gap_rate": round(forecast_budget_gap / annual_budget, 6) if annual_budget else None,
        "yoy_change": yoy_change,
        "yoy_rate": round(yoy_change / last_year_actual, 6) if last_year_actual else None,
        "last_year_actual": last_year_actual,
    }


def _unique_preserve_order(values: Iterable[str]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        normalized = norm_dim(value).upper()
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        result.append(normalized)
    return result


def _org_product_refs_for_data_acct_code(
    data_acct_code: str | None,
    refs_by_runtime_ref_code: dict[str, list[str]],
) -> list[str]:
    code = norm_dim(data_acct_code).upper()
    if not code:
        return []
    return list(refs_by_runtime_ref_code.get(code, []))


def _parse_org_product_metric_ref(label: str) -> tuple[str, str]:
    source_ref, _, metric_name = str(label or "").partition(" ")
    parts = source_ref.split(":", 2)
    metric_code = parts[2].strip() if len(parts) == 3 else ""
    return metric_code, metric_name.strip()


def _metric_identity_from_org_product_refs(
    refs: list[str],
    *,
    fallback_code: str = "",
    fallback_name: str = "",
) -> tuple[str, str]:
    for ref in refs:
        metric_code, metric_name = _parse_org_product_metric_ref(ref)
        if metric_code:
            return metric_code, metric_name or fallback_name
    return fallback_code, fallback_name


async def _product_has_bcir_template(year: int, product_code: str) -> bool:
    normalized = norm_dim(product_code).upper()
    if not normalized:
        return False
    await ensure_business_cost_income_tables(year)
    async with aiosqlite.connect(budget_db_path(year)) as db:
        cur = await db.execute(
            "SELECT COUNT(*) FROM business_cost_income_item WHERE product_code = ?",
            (normalized,),
        )
        row = await cur.fetchone()
    return int(row[0] or 0) > 0


async def _years_with_bcir_templates() -> list[int]:
    years: list[int] = []
    for path in list_budget_database_files():
        stem = path.stem.removeprefix("budget_")
        if not stem.isdigit():
            continue
        year = int(stem)
        await ensure_business_cost_income_tables(year)
        async with aiosqlite.connect(path) as db:
            cur = await db.execute(
                """
                SELECT COUNT(*)
                FROM business_cost_income_item
                WHERE TRIM(product_code) != '' AND product_code != 'CORP'
                """
            )
            if int((await cur.fetchone())[0] or 0) > 0:
                years.append(year)
    return sorted(years, reverse=True)


async def _resolve_template_year(product_code: str, preferred_year: int) -> int:
    if await _product_has_bcir_template(preferred_year, product_code):
        return preferred_year
    for path in sorted(list_budget_database_files(), key=lambda item: item.stem, reverse=True):
        stem = path.stem.removeprefix("budget_")
        if not stem.isdigit():
            continue
        year = int(stem)
        if await _product_has_bcir_template(year, product_code):
            return year
    return preferred_year


async def _load_product_scope(*, product_code: str, template_year: int) -> dict[str, Any]:
    normalized = norm_dim(product_code).upper()
    await ensure_business_cost_income_tables(template_year)
    async with aiosqlite.connect(budget_db_path(template_year)) as db:
        await db.execute("PRAGMA foreign_keys = ON")
        items = await load_business_cost_income_items(db, product_code=normalized)
        indicators = await load_business_cost_income_indicators(db, product_code=normalized)
    return {
        "items": items,
        "indicators": indicators,
        "children": _children_by_section(items),
        "template_year": template_year,
    }


async def _load_topic_products() -> list[dict[str, str]]:
    async with aiosqlite.connect(common_db_path()) as db:
        await db.execute("PRAGMA foreign_keys = ON")
        cur = await db.execute(
            f"""
            {org_product_runtime_products_cte()}
            SELECT product_code, product_name, parent_code, COALESCE(level, 1) AS level
            FROM org_product_runtime_products
            WHERE product_code <> '' AND product_name <> ''
            ORDER BY product_code
            """
        )
        rows = await cur.fetchall()
    name_by_code = {str(row[0]): str(row[1]) for row in rows}
    parent_codes = {str(row[2]) for row in rows if row[2] is not None and str(row[2]).strip()}
    products: list[dict[str, str]] = []
    for code_raw, name_raw, parent_raw, level_raw in rows:
        code = str(code_raw)
        if code == "CORP" or code in parent_codes:
            continue
        parent_code = str(parent_raw or "")
        level = int(level_raw or 1)
        if level <= 2:
            continue
        products.append(
            {
                "product_code": code,
                "product_name": str(name_raw),
                "group_code": parent_code,
                "group_name": name_by_code.get(parent_code, parent_code),
            }
        )
    return products


async def build_input_output_topic_meta() -> dict[str, Any]:
    async with aiosqlite.connect(common_db_path()) as db:
        await db.execute("PRAGMA foreign_keys = ON")
        cur = await db.execute(
            """
            SELECT DISTINCT entity_name
            FROM dept_account
            WHERE entity_name IS NOT NULL AND TRIM(entity_name) != ''
            ORDER BY entity_name
            """
        )
        entities = [str(row[0]) for row in await cur.fetchall()]
    products = await _load_topic_products()
    group_options = list(
        dict.fromkeys(str(item["group_name"]) for item in products if str(item.get("group_name") or "").strip())
    )
    available_years = await _years_with_bcir_templates()
    if not available_years:
        available_years = [
            int(path.stem.removeprefix("budget_"))
            for path in list_budget_database_files()
            if path.stem.removeprefix("budget_").isdigit()
        ]
        available_years = sorted(available_years, reverse=True)
    return {
        "entity_options": entities,
        "product_options": products,
        "group_options": group_options,
        "amount_unit_options": amount_unit_options(),
        "available_years": sorted(available_years, reverse=True),
    }


async def _load_value_maps(
    *,
    year: int,
    entity_name: str,
    product_codes: list[str],
) -> tuple[
    dict[tuple[str, str, int, str, int], float],
    dict[tuple[str, str, int, int], float],
]:
    await ensure_business_cost_income_tables(year)
    placeholders = ",".join("?" for _ in product_codes)
    entity_filter = "AND entity_name = ?" if entity_name else ""
    entity_params = [entity_name] if entity_name else []
    async with aiosqlite.connect(budget_db_path(year)) as db:
        await db.execute("PRAGMA foreign_keys = ON")
        cur = await db.execute(
            f"""
            SELECT product_code, item_section, item_id, field, month, SUM(value) AS total
            FROM business_cost_income_value
            WHERE year = ?
              {entity_filter}
              AND product_code IN ({placeholders})
            GROUP BY product_code, item_section, item_id, field, month
            """,
            [year, *entity_params, *product_codes],
        )
        current_rows = await cur.fetchall()
        cur = await db.execute(
            f"""
            SELECT product_code, item_section, item_id, month, SUM(value) AS total
            FROM business_cost_income_value
            WHERE year = ?
              {entity_filter}
              AND product_code IN ({placeholders})
              AND field = 'actual'
            GROUP BY product_code, item_section, item_id, month
            """,
            [year - 1, *entity_params, *product_codes],
        )
        last_year_rows = await cur.fetchall()
    current = {
        (str(row[0]), str(row[1]), int(row[2]), str(row[3]), int(row[4])): float(row[5] or 0.0)
        for row in current_rows
    }
    last_year = {
        (str(row[0]), str(row[1]), int(row[2]), int(row[3])): float(row[4] or 0.0)
        for row in last_year_rows
    }
    return current, last_year


def _children_by_section(items: list[dict[str, Any]]) -> dict[tuple[str, int | None], list[dict[str, Any]]]:
    children: dict[tuple[str, int | None], list[dict[str, Any]]] = {}
    for item in items:
        children.setdefault((str(item["section"]), item.get("parent_id")), []).append(item)
    for siblings in children.values():
        siblings.sort(key=lambda item: (int(item.get("sort_order", 0)), int(item["id"])))
    return children


def _leaf_ids(
    children: dict[tuple[str, int | None], list[dict[str, Any]]],
    *,
    section: str,
    item_id: int,
) -> list[int]:
    descendants = children.get((section, item_id), [])
    if not descendants:
        return [item_id]
    leaves: list[int] = []
    for child in descendants:
        leaves.extend(_leaf_ids(children, section=section, item_id=int(child["id"])))
    return leaves


def _tree_order(
    children: dict[tuple[str, int | None], list[dict[str, Any]]],
    *,
    section: str,
) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []

    def walk(node: dict[str, Any]) -> None:
        result.append(node)
        for child in children.get((section, int(node["id"])), []):
            walk(child)

    for root in children.get((section, None), []):
        walk(root)
    return result


def _indicator_tree_order(indicators: list[dict[str, Any]]) -> list[dict[str, Any]]:
    children: dict[int | None, list[dict[str, Any]]] = {}
    for indicator in indicators:
        children.setdefault(indicator.get("parent_id"), []).append(indicator)
    for siblings in children.values():
        siblings.sort(key=lambda item: (int(item.get("sort_order", 0)), int(item["id"])))
    ordered: list[dict[str, Any]] = []

    def walk(node: dict[str, Any]) -> None:
        ordered.append(node)
        for child in children.get(int(node["id"]), []):
            walk(child)

    for root in children.get(None, []):
        walk(root)
    return ordered


def _empty_metrics() -> dict[str, Any]:
    raw = _metrics_from_amounts(
        current_actual=0.0,
        annual_budget=0.0,
        annual_forecast=0.0,
        last_year_actual=0.0,
    )
    return {
        **raw,
        "current_actual": 0.0,
        "annual_budget": 0.0,
        "annual_forecast": 0.0,
        "forecast_budget_gap": 0.0,
        "yoy_change": 0.0,
        "last_year_actual": 0.0,
        "month_over_month": None,
        "month_over_month_rate": None,
    }


def _empty_monthly_series() -> dict[str, list[float]]:
    return {"actual": [0.0] * 12, "last_year_actual": [0.0] * 12}


def _empty_monthly_entry() -> dict[str, float]:
    return {"month_actual": 0.0, "month_budget": 0.0, "month_forecast": 0.0}


def _resolve_total_template_code(
    selected_codes: list[str],
    group: str,
    products: list[dict[str, str]],
) -> str:
    template_code = selected_codes[0]
    if not group:
        return template_code
    for item in products:
        if group not in {str(item.get("group_code") or ""), str(item.get("group_name") or "")}:
            continue
        group_code = str(item.get("group_code") or "").strip()
        if group_code:
            return group_code
    return template_code


def _build_rows_for_scope(
    *,
    items: list[dict[str, Any]],
    indicators: list[dict[str, Any]],
    product_codes: list[str],
    scope_cache: dict[str, dict[str, Any]],
    selected_month: int,
    divisor: float,
    current_values: dict[tuple[str, str, int, str, int], float],
    last_year_values: dict[tuple[str, str, int, int], float],
    org_product_refs_by_runtime_ref_code: dict[str, list[str]],
) -> list[dict[str, Any]]:
    children = _children_by_section(items)
    item_by_key = {(str(item["section"]), int(item["id"])): item for item in items}
    leaf_cache: dict[tuple[str, int], list[int]] = {}

    def leaves(section: str, item_id: int) -> list[int]:
        key = (section, item_id)
        if key not in leaf_cache:
            leaf_cache[key] = _leaf_ids(children, section=section, item_id=item_id)
        return leaf_cache[key]

    def template_leaf_names(section: str, item_id: int) -> list[str]:
        return [str(item_by_key[(section, leaf_id)]["name"]) for leaf_id in leaves(section, item_id)]

    def sum_current(section: str, item_id: int, field_name: str, months: Iterable[int]) -> float:
        month_list = list(months)
        leaf_names = template_leaf_names(section, item_id)
        total = 0.0
        for product_code in product_codes:
            product_items = scope_cache[product_code]["items"]
            name_to_id = {(str(item["section"]), str(item["name"])): int(item["id"]) for item in product_items}
            for leaf_name in leaf_names:
                leaf_id = name_to_id.get((section, leaf_name))
                if leaf_id is None:
                    continue
                for month in month_list:
                    total += current_values.get((product_code, section, leaf_id, field_name, month), 0.0)
        return total

    def sum_last_year(section: str, item_id: int, months: Iterable[int]) -> float:
        month_list = list(months)
        leaf_names = template_leaf_names(section, item_id)
        total = 0.0
        for product_code in product_codes:
            product_items = scope_cache[product_code]["items"]
            name_to_id = {(str(item["section"]), str(item["name"])): int(item["id"]) for item in product_items}
            for leaf_name in leaf_names:
                leaf_id = name_to_id.get((section, leaf_name))
                if leaf_id is None:
                    continue
                for month in month_list:
                    total += last_year_values.get((product_code, section, leaf_id, month), 0.0)
        return total

    def amount_monthly_series(section: str, item_id: int) -> dict[str, list[float]]:
        return {
            "actual": [_scale_amount(sum_current(section, item_id, "actual", [month]), divisor) for month in range(1, 13)],
            "last_year_actual": [
                _scale_amount(sum_last_year(section, item_id, [month]), divisor) for month in range(1, 13)
            ],
        }

    def amount_metrics(section: str, item_id: int) -> dict[str, Any]:
        raw = _metrics_from_amounts(
            current_actual=sum_current(section, item_id, "actual", range(1, selected_month + 1)),
            annual_budget=sum_current(section, item_id, "budget", range(1, 13)),
            annual_forecast=sum_current(section, item_id, "forecast", range(1, 13)),
            last_year_actual=sum_last_year(section, item_id, range(1, selected_month + 1)),
        )
        current_month = sum_current(section, item_id, "actual", [selected_month])
        previous_month = sum_current(section, item_id, "actual", [selected_month - 1]) if selected_month > 1 else 0.0
        month_over_month = current_month - previous_month
        return {
            **raw,
            "current_actual": _scale_amount(raw["current_actual"], divisor),
            "annual_budget": _scale_amount(raw["annual_budget"], divisor),
            "annual_forecast": _scale_amount(raw["annual_forecast"], divisor),
            "forecast_budget_gap": _scale_amount(raw["forecast_budget_gap"], divisor),
            "yoy_change": _scale_amount(raw["yoy_change"], divisor),
            "last_year_actual": _scale_amount(raw["last_year_actual"], divisor),
            "month_over_month": _scale_amount(month_over_month, divisor),
            "month_over_month_rate": round(month_over_month / previous_month, 6) if previous_month else None,
        }

    def amount_monthly_entry(section: str, item_id: int) -> dict[str, float]:
        return {
            "month_actual": _scale_amount(sum_current(section, item_id, "actual", [selected_month]), divisor),
            "month_budget": _scale_amount(sum_current(section, item_id, "budget", [selected_month]), divisor),
            "month_forecast": _scale_amount(sum_current(section, item_id, "forecast", [selected_month]), divisor),
        }

    def indicator_ratio(indicator: dict[str, Any], field_name: str, months: Iterable[int]) -> float:
        month_list = list(months)
        n_section = str(indicator["numerator_section"])
        n_name = str(item_by_key[(n_section, int(indicator["numerator_item_id"]))]["name"])
        d_section = str(indicator["denominator_section"])
        d_name = str(item_by_key[(d_section, int(indicator["denominator_item_id"]))]["name"])
        numerator = 0.0
        denominator = 0.0
        for product_code in product_codes:
            product_items = scope_cache[product_code]["items"]
            name_to_id = {(str(item["section"]), str(item["name"])): int(item["id"]) for item in product_items}
            n_id = name_to_id.get((n_section, n_name))
            d_id = name_to_id.get((d_section, d_name))
            if n_id is None or d_id is None:
                continue
            for month in month_list:
                numerator += current_values.get((product_code, n_section, n_id, field_name, month), 0.0)
                denominator += current_values.get((product_code, d_section, d_id, field_name, month), 0.0)
        return _ratio(numerator, denominator) or 0.0

    def indicator_last_year_ratio(indicator: dict[str, Any], months: Iterable[int]) -> float:
        month_list = list(months)
        n_section = str(indicator["numerator_section"])
        n_name = str(item_by_key[(n_section, int(indicator["numerator_item_id"]))]["name"])
        d_section = str(indicator["denominator_section"])
        d_name = str(item_by_key[(d_section, int(indicator["denominator_item_id"]))]["name"])
        numerator = 0.0
        denominator = 0.0
        for product_code in product_codes:
            product_items = scope_cache[product_code]["items"]
            name_to_id = {(str(item["section"]), str(item["name"])): int(item["id"]) for item in product_items}
            n_id = name_to_id.get((n_section, n_name))
            d_id = name_to_id.get((d_section, d_name))
            if n_id is None or d_id is None:
                continue
            for month in month_list:
                numerator += last_year_values.get((product_code, n_section, n_id, month), 0.0)
                denominator += last_year_values.get((product_code, d_section, d_id, month), 0.0)
        return _ratio(numerator, denominator) or 0.0

    def format_indicator_value(value: float, indicator: dict[str, Any]) -> float:
        if str(indicator.get("format")) == "percent":
            return round(value * 100, 4)
        return round(value, 6)

    def indicator_metrics(indicator: dict[str, Any]) -> dict[str, Any]:
        raw = _metrics_from_amounts(
            current_actual=indicator_ratio(indicator, "actual", range(1, selected_month + 1)),
            annual_budget=indicator_ratio(indicator, "budget", range(1, 13)),
            annual_forecast=indicator_ratio(indicator, "forecast", range(1, 13)),
            last_year_actual=indicator_last_year_ratio(indicator, range(1, selected_month + 1)),
        )
        current_month = indicator_ratio(indicator, "actual", [selected_month])
        previous_month = indicator_ratio(indicator, "actual", [selected_month - 1]) if selected_month > 1 else 0.0
        month_over_month = current_month - previous_month
        if str(indicator.get("format")) == "percent":
            return {
                **raw,
                "current_actual": format_indicator_value(raw["current_actual"], indicator),
                "annual_budget": format_indicator_value(raw["annual_budget"], indicator),
                "annual_forecast": format_indicator_value(raw["annual_forecast"], indicator),
                "forecast_budget_gap": format_indicator_value(raw["forecast_budget_gap"], indicator),
                "yoy_change": format_indicator_value(raw["yoy_change"], indicator),
                "last_year_actual": format_indicator_value(raw["last_year_actual"], indicator),
                "month_over_month": round(month_over_month * 100, 4),
                "month_over_month_rate": round(month_over_month / previous_month, 6) if previous_month else None,
            }
        return {
            **raw,
            "month_over_month": round(month_over_month, 6),
            "month_over_month_rate": round(month_over_month / previous_month, 6) if previous_month else None,
        }

    def indicator_monthly_entry(indicator: dict[str, Any]) -> dict[str, float]:
        return {
            "month_actual": format_indicator_value(indicator_ratio(indicator, "actual", [selected_month]), indicator),
            "month_budget": format_indicator_value(indicator_ratio(indicator, "budget", [selected_month]), indicator),
            "month_forecast": format_indicator_value(indicator_ratio(indicator, "forecast", [selected_month]), indicator),
        }

    def indicator_monthly_series(indicator: dict[str, Any]) -> dict[str, list[float]]:
        return {
            "actual": [
                format_indicator_value(indicator_ratio(indicator, "actual", [month]), indicator)
                for month in range(1, 13)
            ],
            "last_year_actual": [
                format_indicator_value(indicator_last_year_ratio(indicator, [month]), indicator)
                for month in range(1, 13)
            ],
        }

    rows: list[dict[str, Any]] = []
    for indicator in _indicator_tree_order(indicators):
        if int(indicator["enabled"]) != 1:
            continue
        indicator_id = int(indicator["id"])
        if int(indicator.get("display_group") or 0) == 1:
            rows.append(
                {
                    "section": "indicator",
                    "id": indicator_id,
                    "name": str(indicator["name"]),
                    "parent_id": indicator.get("parent_id"),
                    "is_leaf": False,
                    "display_group": True,
                    "display_format": None,
                    "topic_metric_node_code": indicator.get("topic_metric_node_code"),
                    "data_acct_code": "",
                    "org_product_refs": [],
                    "sort_order": int(indicator.get("sort_order", 0)),
                    "enabled": True,
                    "metrics": _empty_metrics(),
                    "monthly_entry": _empty_monthly_entry(),
                    "monthly_series": _empty_monthly_series(),
                }
            )
            continue
        numerator_key = (str(indicator["numerator_section"]), int(indicator["numerator_item_id"]))
        denominator_key = (str(indicator["denominator_section"]), int(indicator["denominator_item_id"]))
        if numerator_key not in item_by_key or denominator_key not in item_by_key:
            continue
        rows.append(
            {
                "section": "indicator",
                "id": indicator_id,
                "name": str(indicator["name"]),
                "parent_id": indicator.get("parent_id"),
                "is_leaf": True,
                "display_group": False,
                "display_format": str(indicator.get("format") or "ratio"),
                "topic_metric_node_code": indicator.get("topic_metric_node_code"),
                "data_acct_code": "",
                "org_product_refs": [],
                "sort_order": int(indicator.get("sort_order", 0)),
                "enabled": True,
                "metrics": indicator_metrics(indicator),
                "monthly_entry": indicator_monthly_entry(indicator),
                "monthly_series": indicator_monthly_series(indicator),
            }
        )

    for section in ("input", "output"):
        for item in _tree_order(children, section=section):
            if int(item["enabled"]) != 1:
                continue
            item_id = int(item["id"])
            data_acct_code = str(item.get("data_acct_code") or "")
            org_product_refs = _org_product_refs_for_data_acct_code(
                data_acct_code,
                org_product_refs_by_runtime_ref_code,
            )
            metric_code, metric_name = _metric_identity_from_org_product_refs(
                org_product_refs,
                fallback_code=data_acct_code,
                fallback_name=str(item["name"]),
            )
            rows.append(
                {
                    "section": section,
                    "id": item_id,
                    "name": str(item["name"]),
                    "parent_id": item.get("parent_id"),
                    "is_leaf": not bool(children.get((section, item_id))),
                    "display_group": int(item.get("display_group") or 0) == 1,
                    "display_format": "number",
                    "topic_metric_node_code": "",
                    "data_acct_code": data_acct_code,
                    "metric_code": metric_code,
                    "metric_name": metric_name,
                    "org_product_refs": org_product_refs,
                    "sort_order": int(item.get("sort_order", 0)),
                    "enabled": True,
                    "metrics": amount_metrics(section, item_id),
                    "monthly_entry": amount_monthly_entry(section, item_id),
                    "monthly_series": amount_monthly_series(section, item_id),
                }
            )
    rows.sort(
        key=lambda row: (
            {"indicator": 0, "input": 1, "output": 2}.get(str(row["section"]), 9),
            int(row["sort_order"]),
            int(row["id"]),
        )
    )
    return rows


async def build_input_output_topic_report(
    *,
    entity_name: str,
    report_month: str,
    group_name: str | None,
    product_codes: list[str] | None,
    amount_unit: str,
) -> dict[str, Any]:
    year, month = parse_year_month(report_month)
    entity = norm_dim(entity_name)
    group = norm_dim(group_name)
    amount_unit_label, divisor = amount_unit_meta(amount_unit)
    products = await _load_topic_products()
    supported_codes = [str(item["product_code"]) for item in products]
    if group:
        grouped_codes = [
            str(item["product_code"])
            for item in products
            if group in {str(item.get("group_code") or ""), str(item.get("group_name") or "")}
        ]
        if grouped_codes:
            supported_codes = grouped_codes
    requested = [code for code in _unique_preserve_order(product_codes or []) if code in supported_codes]
    selected_codes = requested or supported_codes
    if not selected_codes:
        raise ValueError("未找到可用的投入产出专题产品口径")

    total_template_code = _resolve_total_template_code(selected_codes, group, products)
    codes_to_load = list(dict.fromkeys([*selected_codes, total_template_code]))

    await ensure_business_cost_income_tables(year)
    scope_cache: dict[str, dict[str, Any]] = {}
    for code in codes_to_load:
        template_year = await _resolve_template_year(code, year)
        scope_cache[code] = await _load_product_scope(product_code=code, template_year=template_year)

    template_scope = scope_cache[total_template_code]
    current_values, last_year_values = await _load_value_maps(
        year=year,
        entity_name=entity,
        product_codes=selected_codes,
    )
    async with aiosqlite.connect(common_db_path()) as common_db:
        await common_db.execute("PRAGMA foreign_keys = ON")
        org_product_refs_by_runtime_ref_code = await load_org_product_metric_refs_by_runtime_ref_code(common_db)
    product_name_map = {str(item["product_code"]): str(item["product_name"]) for item in products}
    total_rows = _build_rows_for_scope(
        items=template_scope["items"],
        indicators=template_scope["indicators"],
        product_codes=selected_codes,
        scope_cache=scope_cache,
        selected_month=month,
        divisor=divisor,
        current_values=current_values,
        last_year_values=last_year_values,
        org_product_refs_by_runtime_ref_code=org_product_refs_by_runtime_ref_code,
    )
    product_blocks = [
        {
            "product_code": code,
            "product_name": product_name_map.get(code, code),
            "rows": _build_rows_for_scope(
                items=scope_cache[code]["items"],
                indicators=scope_cache[code]["indicators"],
                product_codes=[code],
                scope_cache=scope_cache,
                selected_month=month,
                divisor=divisor,
                current_values=current_values,
                last_year_values=last_year_values,
                org_product_refs_by_runtime_ref_code=org_product_refs_by_runtime_ref_code,
            ),
        }
        for code in selected_codes
    ]
    template_note = ""
    if any(int(scope_cache[code]["template_year"]) != year for code in codes_to_load):
        fallback_years = sorted(
            {int(scope_cache[code]["template_year"]) for code in codes_to_load if int(scope_cache[code]["template_year"]) != year}
        )
        template_note = (
            f"细项结构沿用 {','.join(str(item) for item in fallback_years)} 年模板，数值取 {year} 年。"
        )
    return {
        "report_month": report_month,
        "selected_year": year,
        "entity_name": entity or "全部主体",
        "group_name": group or None,
        "amount_unit": amount_unit,
        "amount_unit_label": amount_unit_label,
        "selected_product_codes": selected_codes,
        "total_rows": total_rows,
        "product_blocks": product_blocks,
        "note": (
            f"主体=全部；费用月份={report_month}；"
            f"产品群={group or '全部'}；"
            f"产品范围={','.join(selected_codes) if selected_codes else '全部'}；"
            f"{template_note}"
            f"当前总览基于 business_cost_income_item / business_cost_income_indicator / "
            f"business_cost_income_value 生成。"
        ),
    }


def _flatten_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    children_by_parent: dict[tuple[str, int | None], list[dict[str, Any]]] = {}
    for row in rows:
        section = str(row.get("section") or "")
        children_by_parent.setdefault((section, row.get("parent_id")), []).append(row)
    for siblings in children_by_parent.values():
        siblings.sort(key=lambda item: (int(item.get("sort_order", 0)), int(item.get("id", 0))))
    result: list[dict[str, Any]] = []

    def walk(section: str, parent_id: int | None, depth: int) -> None:
        for child in children_by_parent.get((section, parent_id), []):
            result.append({**child, "depth": depth})
            walk(section, int(child["id"]), depth + 1)

    for section in ("indicator", "input", "output"):
        section_roots = children_by_parent.get((section, None), [])
        for root in section_roots:
            result.append({**root, "depth": 0})
            walk(section, int(root["id"]), 1)
    return result


def _export_row(view_name: str, product_code: str, product_name: str, row: dict[str, Any]) -> list[Any]:
    metrics = row.get("metrics") or {}
    depth = int(row.get("depth") or 0)
    metric_code, _metric_name = _metric_identity_from_org_product_refs(
        [str(ref) for ref in row.get("org_product_refs") or []],
        fallback_code=str(row.get("metric_code") or row.get("data_acct_code") or ""),
        fallback_name=str(row.get("metric_name") or row.get("name") or ""),
    )
    return [
        view_name,
        product_code,
        product_name,
        SECTION_LABELS.get(str(row.get("section")), str(row.get("section") or "")),
        depth + 1,
        f"{'  ' * depth}{row.get('name') or ''}",
        row.get("topic_metric_node_code") or "",
        metric_code,
        "\n".join(str(ref) for ref in row.get("org_product_refs") or []),
        metrics.get("current_actual"),
        metrics.get("annual_budget"),
        metrics.get("budget_progress"),
        metrics.get("annual_forecast"),
        metrics.get("forecast_budget_gap"),
        metrics.get("gap_rate"),
        metrics.get("yoy_change"),
        metrics.get("yoy_rate"),
        metrics.get("month_over_month"),
        metrics.get("month_over_month_rate"),
        metrics.get("last_year_actual"),
    ]


def build_input_output_topic_workbook(report: dict[str, Any], *, view_mode: str) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "全行总表" if view_mode == "total" else "分产品明细"
    ws.append(["投入产出专题概览"])
    ws.append(
        [
            f"费用月份：{report.get('report_month')}",
            f"产品群：{report.get('group_name') or '全部'}",
            f"产品范围：{','.join(report.get('selected_product_codes') or []) or '全部'}",
            f"单位：{report.get('amount_unit_label') or report.get('amount_unit')}",
        ]
    )
    ws.append([])
    ws.append(EXPORT_HEADERS)
    header_fill = PatternFill("solid", fgColor="E2E8F0")
    for cell in ws[1]:
        cell.font = Font(bold=True, size=14)
    for cell in ws[4]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    if view_mode == "detail":
        for block in report.get("product_blocks") or []:
            for row in _flatten_rows(block.get("rows") or []):
                ws.append(
                    _export_row(
                        "分产品明细",
                        str(block.get("product_code") or ""),
                        str(block.get("product_name") or ""),
                        row,
                    )
                )
    else:
        for row in _flatten_rows(report.get("total_rows") or []):
            ws.append(_export_row("全行总表", "TOTAL", "汇总", row))

    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_length + 2, 10), 36)
    return wb
