"""Excel parser and preview builder for expense actual import files."""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO
import re
from typing import Any

from openpyxl import load_workbook
import xlrd

from app.schemas import (
    ExpenseActualImportManageDepartmentWarning,
    ExpenseActualImportPreviewResponse,
    ExpenseActualImportPreviewRow,
)
from app.services.bi_ai_manage_department import (
    format_manage_departments_display,
    normalize_manage_department_subject_label,
    resolve_manage_department_for_budget_subject,
)
from app.services.expense_budget_execution_framework import norm_key


class ExpenseActualImportParseError(ValueError):
    """Raised when an uploaded expense actual workbook cannot be parsed."""


@dataclass
class ParsedActualDetailRow:
    data_date: str
    period_ym: str
    period_text: str
    org_code: str
    org_name: str
    dep_code: str
    dep_name: str
    subject_code: str
    subject_name: str
    amount: float
    fee_type_code: str
    fee_type_name: str
    bi_ai_source_code: str
    bi_ai_source_name: str
    manage_department_code: str
    journal_name: str
    serial_no: str
    line_desc: str
    owner_name_raw: str
    owner_name_mapped: str | None
    monthly_caliber: str
    budget_subject_raw: str
    budget_subject_mapped: str | None
    fee_major_mapped: str
    fee_category_mapped: str
    budget_release_caliber_mapped: str
    manage_department2: str
    special_control_tag: str
    owner_matched: bool
    subject_matched: bool
    match_note: str | None
    import_matched_department: str = ""
    import_manage_department: str = ""
    mapping_manage_departments: list[str] = field(default_factory=list)
    mapping_manage_department: str = ""


class FrameworkContext:
    def __init__(self) -> None:
        self.owner_alias_map: dict[str, str] = {}
        self.owner_names: set[str] = set()
        self.subject_alias_map: dict[str, str] = {}
        self.subject_names: set[str] = set()
        self.bi_ai_subject_mapping: dict[str, str] = {}
        self.bi_ai_subject_mapping_detail: dict[str, tuple[str, str, str]] = {}
        self.owner_dept_manage_map: dict[str, str] = {}
        self.manage_dept_owner_map: dict[str, str] = {}
        self.subject_manage_department: dict[str, str] = {}
        self.bi_mapping_manage_departments_by_caliber: dict[str, list[str]] = {}


GOVERNANCE_OWNER_ALIASES = {
    "董事会办公室": "公司治理部",
    "监事会办公室": "公司治理部",
    "董监事会办公室": "公司治理部",
}

REQUIRED_IMPORT_COLUMNS = [
    "期间",
    "费用发生部门编码",
    "费用发生部门",
    "责任中心编码",
    "责任中心",
    "科目编码",
    "科目描述",
    "金额",
    "费用类别编码",
    "费用类别",
    "管控口径编码",
    "管控口径名称",
    "归口管理部门编码",
    "费用归属部门",
]

HEADER_ALIASES = {
    "数据日期": ("数据日期", "#dataDate"),
    "期间": ("期间", "日期", "#dataDate"),
    "费用发生部门编码": ("费用发生部门编码", "费用发生部门代码", "费用归属部门编码", "#orgCd"),
    "费用发生部门": ("费用发生部门", "费用部门", "#orgName"),
    "责任中心编码": ("责任中心编码", "责任中心代码", "#depCd"),
    "责任中心": ("责任中心", "#depName"),
    "科目编码": ("科目编码", "科目代码", "#subjectCd"),
    "科目描述": ("科目描述", "科目名称", "#subjectName"),
    "日记帐名": ("日记帐名",),
    "流水号": ("流水号",),
    "行说明": ("行说明",),
    "金额": ("金额", "#amount"),
    "费用类别编码": ("费用类别编码", "费用类别代码", "#feeTypeCd"),
    "费用类别": ("费用类别", "#feeTypeName"),
    "管控口径编码": ("管控口径编码", "管控口径代码", "#controlItem"),
    "管控口径名称": ("管控口径名称", "管控口径", "#controlItemName"),
    "归口管理部门编码": ("归口管理部门编码", "归口管理部门代码", "#controlDepCd"),
    "费用归属部门": ("费用归属部门", "归口管理部门", "#controlDepName"),
    "费用月报口径": ("费用月报口径", "日记帐名"),
    "预算科目": ("预算科目", "部门预算科目"),
}


def _text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return re.sub(r"\s+", " ", text)


def normalize_key(value: str) -> str:
    return re.sub(r"\s+", "", _text(value)).lower()


def strip_leading_code(value: str) -> str:
    text = _text(value)
    stripped = re.sub(r"^[A-Za-z]+\d+\s+", "", text)
    return _text(stripped or text)


def remember_bi_mapping(ctx: FrameworkContext, source_value: str, budget_subject: str) -> None:
    source_key = normalize_key(source_value)
    if source_key and budget_subject and source_key not in ctx.bi_ai_subject_mapping:
        ctx.bi_ai_subject_mapping[source_key] = budget_subject


HEADER_ALIAS_LOOKUP = {
    normalize_key(alias): canonical
    for canonical, aliases in HEADER_ALIASES.items()
    for alias in aliases
}


def _canonical_owner_name(owner_name: str, ctx: FrameworkContext) -> str:
    text = _text(owner_name)
    if not text:
        return ""
    return ctx.owner_alias_map.get(
        normalize_key(text),
        ctx.owner_alias_map.get(normalize_key(strip_leading_code(text)), text),
    )


def _canonical_subject(subject_name: str, ctx: FrameworkContext) -> str:
    text = _text(subject_name)
    if not text:
        return ""
    return ctx.subject_alias_map.get(normalize_key(text), text)


def _value_by_excel_column(values: list[Any], zero_based_index: int) -> str:
    if zero_based_index >= len(values):
        return ""
    return _text(values[zero_based_index])


def _parse_period_ym(value: Any, datemode: int | None = None) -> tuple[str, str]:
    if value is None or value == "":
        return "", ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m"), value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m"), value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float)):
        if datemode is not None:
            dt = xlrd.xldate.xldate_as_datetime(float(value), datemode)
            return dt.strftime("%Y-%m"), dt.strftime("%Y-%m-%d")
        text = _text(value)
        return text, text
    text = _text(value)
    match = re.search(r"(\d{4})[-/年]?(\d{1,2})", text)
    if match:
        year = int(match.group(1))
        month = int(match.group(2))
        return f"{year:04d}-{month:02d}", text
    return text, text


def _build_match_note(
    owner_matched: bool,
    subject_matched: bool,
    manage_department2: str,
    *,
    manage_department_mismatch: str | None = None,
) -> str | None:
    problems: list[str] = []
    if not owner_matched:
        problems.append("费用归属部门未匹配")
    if not subject_matched:
        problems.append("预算科目未匹配")
    if not _text(manage_department2):
        problems.append("归口管理部门2未匹配")
    if manage_department_mismatch:
        problems.append(manage_department_mismatch)
    return "；".join(problems) or None


def _resolve_import_manage_department(
    budget_subject_mapped: str | None,
    ctx: FrameworkContext,
) -> str:
    if not budget_subject_mapped:
        return ""
    return resolve_manage_department_for_budget_subject(
        budget_subject_mapped,
        ctx.subject_manage_department,
    )


def _resolve_mapping_manage_departments(
    budget_release_caliber_mapped: str,
    ctx: FrameworkContext,
    *,
    bi_lookup_keys: list[str] | None = None,
) -> list[str]:
    lookup_keys: list[str] = []
    seen_keys: set[str] = set()
    for raw_key in list(bi_lookup_keys or []) + [budget_release_caliber_mapped]:
        for candidate in (raw_key, normalize_manage_department_subject_label(_text(raw_key))):
            if not candidate:
                continue
            key = norm_key(candidate)
            if key in seen_keys:
                continue
            seen_keys.add(key)
            lookup_keys.append(key)

    for key in lookup_keys:
        departments = ctx.bi_mapping_manage_departments_by_caliber.get(key)
        if departments:
            return list(departments)

    single = resolve_manage_department_for_budget_subject(
        budget_release_caliber_mapped,
        ctx.subject_manage_department,
    )
    return [single] if single else []


def _resolve_import_matched_department(
    owner_name_mapped: str | None,
    owner_name_raw: str,
    ctx: FrameworkContext,
) -> str:
    if _text(owner_name_mapped):
        return _text(owner_name_mapped)
    mapped_owner_from_manage_dept = ctx.manage_dept_owner_map.get(normalize_key(owner_name_raw))
    owner_candidate = _canonical_owner_name(mapped_owner_from_manage_dept or owner_name_raw, ctx)
    if owner_candidate in ctx.owner_names:
        return owner_candidate
    return _text(owner_name_raw)


def _build_manage_department_mismatch_note(
    import_matched_department: str,
    mapping_manage_departments: list[str],
) -> str | None:
    matched = normalize_key(import_matched_department)
    if not matched or not mapping_manage_departments:
        return None
    normalized_mapping = {normalize_key(item) for item in mapping_manage_departments if _text(item)}
    if matched in normalized_mapping:
        return None
    mapping_display = format_manage_departments_display(mapping_manage_departments)
    return (
        f"归口部门不匹配：导入匹配部门={import_matched_department}，"
        f"不在BI-AI映射表归口部门清单内（{mapping_display}）"
    )


def _build_manage_department_warnings(
    rows: list[ParsedActualDetailRow],
) -> list[ExpenseActualImportManageDepartmentWarning]:
    warnings: list[ExpenseActualImportManageDepartmentWarning] = []
    seen: set[str] = set()
    for row in rows:
        mismatch_note = _build_manage_department_mismatch_note(
            row.import_matched_department,
            row.mapping_manage_departments,
        )
        if not mismatch_note:
            continue
        dedupe_key = "|".join(
            [
                row.period_ym,
                normalize_key(row.owner_name_raw),
                normalize_key(row.budget_subject_mapped or ""),
                normalize_key(row.budget_release_caliber_mapped),
                normalize_key(row.import_matched_department),
                "|".join(normalize_key(item) for item in row.mapping_manage_departments),
            ]
        )
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)
        warnings.append(
            ExpenseActualImportManageDepartmentWarning(
                period_ym=row.period_ym,
                owner_name_raw=row.owner_name_raw,
                budget_subject_mapped=row.budget_subject_mapped or "",
                budget_release_caliber_mapped=row.budget_release_caliber_mapped,
                import_manage_department=row.import_matched_department,
                mapping_manage_department=row.mapping_manage_department,
                message=mismatch_note,
            )
        )
    return warnings


def _parse_amount(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = _text(value).replace(",", "").replace("，", "")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError as exc:
        raise ExpenseActualImportParseError(f"金额格式不正确：{value}") from exc


def _build_header_idx(*header_rows: list[str]) -> dict[str, int]:
    header_idx: dict[str, int] = {}
    for headers in header_rows:
        for idx, name in enumerate(headers):
            canonical = HEADER_ALIAS_LOOKUP.get(normalize_key(name))
            if canonical and canonical not in header_idx:
                header_idx[canonical] = idx
    return header_idx


def _resolve_actual_header(sheet_rows: list[list[Any]]) -> tuple[dict[str, int], int]:
    if len(sheet_rows) < 2:
        raise ExpenseActualImportParseError("导入文件缺少表头")

    best_header_idx: dict[str, int] = {}
    best_data_start_row = 2
    scan_rows = min(len(sheet_rows), 5)
    for row_idx in range(scan_rows):
        headers = [_text(v) for v in sheet_rows[row_idx]]
        candidate_idx = _build_header_idx(headers)
        if len(candidate_idx) > len(best_header_idx):
            best_header_idx = candidate_idx
            best_data_start_row = row_idx + 1
        if row_idx + 1 >= scan_rows:
            continue
        next_headers = [_text(v) for v in sheet_rows[row_idx + 1]]
        combined_idx = _build_header_idx(headers, next_headers)
        if len(combined_idx) > len(best_header_idx):
            best_header_idx = combined_idx
            best_data_start_row = row_idx + 2

    missing = [name for name in REQUIRED_IMPORT_COLUMNS if name not in best_header_idx]
    if missing:
        raise ExpenseActualImportParseError(f"导入文件缺少字段：{'、'.join(missing)}")
    return best_header_idx, best_data_start_row


def _read_actual_sheet(file_name: str, raw: bytes) -> tuple[list[list[Any]], int | None]:
    lower_name = (file_name or "").lower()
    if lower_name.endswith(".xlsx") or lower_name.endswith(".xlsm"):
        try:
            wb = load_workbook(BytesIO(raw), data_only=True)
        except Exception as exc:
            raise ExpenseActualImportParseError(f"无法读取 Excel 文件：{exc}") from exc
        if not wb.sheetnames:
            raise ExpenseActualImportParseError("导入文件缺少工作表")
        ws = wb["费用执行表"] if "费用执行表" in wb.sheetnames else wb.worksheets[0]
        return [[cell for cell in row] for row in ws.iter_rows(values_only=True)], None

    try:
        wb = xlrd.open_workbook(file_contents=raw)
    except Exception as exc:
        raise ExpenseActualImportParseError(f"无法读取 Excel 文件：{exc}") from exc
    if not wb.sheet_names():
        raise ExpenseActualImportParseError("导入文件缺少工作表")
    ws = wb.sheet_by_name("费用执行表") if "费用执行表" in wb.sheet_names() else wb.sheet_by_index(0)
    return [ws.row_values(row_idx) for row_idx in range(ws.nrows)], wb.datemode


def parse_actual_file(file_name: str, raw: bytes, ctx: FrameworkContext) -> list[ParsedActualDetailRow]:
    sheet_rows, datemode = _read_actual_sheet(file_name, raw)
    header_idx, data_start_row = _resolve_actual_header(sheet_rows)

    rows: list[ParsedActualDetailRow] = []

    def get_value(values: list[Any], column_name: str) -> str:
        idx = header_idx.get(column_name)
        if idx is None or idx >= len(values):
            return ""
        return _text(values[idx])

    def get_raw(values: list[Any], column_name: str) -> Any:
        idx = header_idx.get(column_name)
        if idx is None or idx >= len(values):
            return ""
        return values[idx]

    for row_idx in range(data_start_row, len(sheet_rows)):
        values = sheet_rows[row_idx]
        if not any(_text(v) for v in values):
            continue
        bi_mapping_candidates = [
            get_value(values, "管控口径编码"),
            get_value(values, "管控口径名称"),
            get_value(values, "费用类别编码"),
            get_value(values, "费用类别"),
            get_value(values, "预算科目"),
            get_value(values, "科目编码"),
            get_value(values, "科目描述"),
        ]
        bi_mapping = next(
            (
                ctx.bi_ai_subject_mapping_detail.get(normalize_key(candidate))
                for candidate in bi_mapping_candidates
                if candidate and ctx.bi_ai_subject_mapping_detail.get(normalize_key(candidate))
            ),
            None,
        )
        fee_major_mapped = bi_mapping[0] if bi_mapping else ""
        fee_category_mapped = bi_mapping[1] if bi_mapping else ""
        budget_release_caliber_mapped = bi_mapping[2] if bi_mapping else ""
        manage_department2 = ctx.manage_dept_owner_map.get(normalize_key(_value_by_excel_column(values, 17)), "")  # R列
        special_control_tag = (
            "抵押"
            if normalize_key(_value_by_excel_column(values, 6)) == normalize_key("其他外包服务费")
            and normalize_key(_value_by_excel_column(values, 15)) == normalize_key("运营其他")
            else ""
        )
        period_source = get_raw(values, "期间") or get_raw(values, "数据日期")
        period_ym, period_text = _parse_period_ym(period_source, datemode)
        owner_name_raw = get_value(values, "费用归属部门")
        mapped_owner_from_manage_dept = ctx.manage_dept_owner_map.get(normalize_key(owner_name_raw))
        owner_candidate = _canonical_owner_name(mapped_owner_from_manage_dept or owner_name_raw, ctx)
        owner_name_mapped = owner_candidate if owner_candidate in ctx.owner_names else None
        budget_subject_raw = get_value(values, "预算科目")
        subject_candidate = _canonical_subject(budget_subject_raw, ctx)
        budget_subject_mapped = subject_candidate if subject_candidate in ctx.subject_names else None
        bi_ai_source_name = get_value(values, "管控口径名称")
        if not budget_subject_mapped:
            mapped_subject = next(
                (
                    ctx.bi_ai_subject_mapping.get(normalize_key(candidate))
                    for candidate in (
                        budget_subject_raw,
                        get_value(values, "管控口径编码"),
                        bi_ai_source_name,
                        get_value(values, "费用类别编码"),
                        get_value(values, "费用类别"),
                        get_value(values, "科目编码"),
                        get_value(values, "科目描述"),
                    )
                    if candidate and ctx.bi_ai_subject_mapping.get(normalize_key(candidate))
                ),
                None,
            )
            if mapped_subject:
                subject_candidate = _canonical_subject(mapped_subject, ctx)
                budget_subject_mapped = subject_candidate if subject_candidate in ctx.subject_names else None
        owner_matched = bool(owner_name_mapped)
        subject_matched = bool(budget_subject_mapped)
        import_matched_department = _resolve_import_matched_department(
            owner_name_mapped,
            owner_name_raw,
            ctx,
        )
        import_manage_department = _resolve_import_manage_department(budget_subject_mapped, ctx)
        mapping_manage_departments = _resolve_mapping_manage_departments(
            budget_release_caliber_mapped,
            ctx,
            bi_lookup_keys=[candidate for candidate in bi_mapping_candidates if _text(candidate)],
        )
        mapping_manage_department = format_manage_departments_display(mapping_manage_departments)
        manage_department_mismatch = _build_manage_department_mismatch_note(
            import_matched_department,
            mapping_manage_departments,
        )
        rows.append(
            ParsedActualDetailRow(
                data_date=get_value(values, "数据日期"),
                period_ym=period_ym,
                period_text=period_text,
                org_code=get_value(values, "费用发生部门编码"),
                org_name=get_value(values, "费用发生部门"),
                dep_code=get_value(values, "责任中心编码"),
                dep_name=get_value(values, "责任中心"),
                subject_code=get_value(values, "科目编码"),
                subject_name=get_value(values, "科目描述"),
                amount=_parse_amount(get_raw(values, "金额")),
                fee_type_code=get_value(values, "费用类别编码"),
                fee_type_name=get_value(values, "费用类别"),
                bi_ai_source_code=get_value(values, "管控口径编码"),
                bi_ai_source_name=bi_ai_source_name,
                manage_department_code=get_value(values, "归口管理部门编码"),
                journal_name=get_value(values, "日记帐名"),
                serial_no=get_value(values, "流水号"),
                line_desc=get_value(values, "行说明"),
                owner_name_raw=owner_name_raw,
                owner_name_mapped=owner_name_mapped,
                monthly_caliber=get_value(values, "费用月报口径"),
                budget_subject_raw=budget_subject_raw,
                budget_subject_mapped=budget_subject_mapped,
                fee_major_mapped=fee_major_mapped,
                fee_category_mapped=fee_category_mapped,
                budget_release_caliber_mapped=budget_release_caliber_mapped,
                manage_department2=manage_department2,
                special_control_tag=special_control_tag,
                owner_matched=owner_matched,
                subject_matched=subject_matched,
                match_note=_build_match_note(
                    owner_matched,
                    subject_matched,
                    manage_department2,
                    manage_department_mismatch=manage_department_mismatch,
                ),
                import_matched_department=import_matched_department,
                import_manage_department=import_manage_department,
                mapping_manage_departments=mapping_manage_departments,
                mapping_manage_department=mapping_manage_department,
            )
        )
    if not rows:
        raise ExpenseActualImportParseError("导入文件没有可用明细数据")
    return rows


def _resolve_match_status(row: ParsedActualDetailRow) -> str:
    if not row.match_note:
        return "已匹配"
    if not row.owner_matched and not row.subject_matched:
        return "未匹配"
    return "部分匹配"


def _row_has_validation_warning(row: ParsedActualDetailRow) -> bool:
    return bool(row.match_note)


def build_preview_response(
    file_name: str,
    rows: list[ParsedActualDetailRow],
) -> ExpenseActualImportPreviewResponse:
    periods = sorted({row.period_ym for row in rows if row.period_ym})
    matched_owner_rows = sum(1 for row in rows if row.owner_matched)
    matched_subject_rows = sum(1 for row in rows if row.subject_matched)
    warning_rows = sum(1 for row in rows if _row_has_validation_warning(row))
    unmatched_rows = warning_rows

    def to_preview_row(row: ParsedActualDetailRow) -> ExpenseActualImportPreviewRow:
        return ExpenseActualImportPreviewRow(
            data_date=row.data_date,
            period_ym=row.period_ym,
            org_code=row.org_code,
            org_name=row.org_name,
            dep_code=row.dep_code,
            dep_name=row.dep_name,
            subject_code=row.subject_code,
            subject_name=row.subject_name,
            journal_name=row.journal_name,
            serial_no=row.serial_no,
            line_desc=row.line_desc,
            fee_type_code=row.fee_type_code,
            fee_type_name=row.fee_type_name,
            bi_ai_source_code=row.bi_ai_source_code,
            bi_ai_source_name=row.bi_ai_source_name,
            manage_department_code=row.manage_department_code,
            owner_name_raw=row.owner_name_raw,
            monthly_caliber=row.monthly_caliber,
            owner_name_mapped=row.owner_name_mapped,
            budget_subject_raw=row.budget_subject_raw,
            budget_subject_mapped=row.budget_subject_mapped,
            fee_major_mapped=row.fee_major_mapped,
            fee_category_mapped=row.fee_category_mapped,
            budget_release_caliber_mapped=row.budget_release_caliber_mapped,
            manage_department2=row.manage_department2,
            special_control_tag=row.special_control_tag,
            amount=round(float(row.amount), 2),
            match_status=_resolve_match_status(row),
            match_note=row.match_note,
        )

    warning_row_items = [row for row in rows if _row_has_validation_warning(row)]
    matched_row_items = [row for row in rows if not _row_has_validation_warning(row)]
    preview_limit = 500
    ordered_rows = warning_row_items + matched_row_items
    preview_rows = [to_preview_row(row) for row in ordered_rows[:preview_limit]]
    unmatched_preview_rows = [to_preview_row(row) for row in warning_row_items[:preview_limit]]
    return ExpenseActualImportPreviewResponse(
        file_name=file_name,
        row_count=len(rows),
        periods=periods,
        matched_owner_rows=matched_owner_rows,
        matched_subject_rows=matched_subject_rows,
        unmatched_rows=unmatched_rows,
        preview_rows=preview_rows,
        unmatched_preview_rows=unmatched_preview_rows,
        manage_department_warnings=_build_manage_department_warnings(rows),
    )
