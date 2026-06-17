from __future__ import annotations

from datetime import datetime
import re
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.utils import get_column_letter

from app.formula_refs import extract_runtime_metric_ref_code
from app.schemas import BudgetSummaryExportPivotRequest, BudgetSummaryRowDto, CompareSummaryRowDto
from app.services.export_common import workbook_streaming_response


FIELD_LABELS: dict[str, str] = {
    "show_level": "展示层级",
    "data_file_id": "数据文件",
    "source_year": "来源年度",
    "metric_level1": "指标1级",
    "metric_level2": "指标2级",
    "metric_level3": "指标3级",
    "metric_level4": "指标4级",
    "metric_level5": "指标5级",
    "dept_level1": "部门科目1级",
    "dept_level2": "部门科目2级",
    "dept_level3": "部门科目3级",
    "data_code_name": "机构及产品指标编码",
    "product_code_name": "机构及产品",
    "year": "年度",
    "month": "月份",
    "quarter": "季度",
    "budget_actual": "预算/实际",
    "version_display": "版本号及名称",
    "value_type": "数值类型",
}


def _field_ids(ids: list[str]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for raw in ids:
        field_id = str(raw or "").strip()
        if not field_id or field_id in seen or field_id not in FIELD_LABELS:
            continue
        seen.add(field_id)
        out.append(field_id)
    return out


def _field_value(row: BudgetSummaryRowDto | CompareSummaryRowDto, field_id: str) -> str:
    if field_id == "show_level":
        return str(getattr(row, "show_level", "未设置") or "未设置")
    if field_id == "data_file_id":
        return str(getattr(row, "data_file_id", "未设置") or "未设置")
    if field_id == "source_year":
        return str(getattr(row, "source_year", "未设置") or "未设置")
    if field_id == "budget_actual":
        return "预算" if int(getattr(row, "budget_actual", 0) or 0) == 0 else "实际"
    if field_id == "version_display":
        if hasattr(row, "source_version_id"):
            show_level = int(getattr(row, "show_level", 0) or 0)
            prefix = f"展示版本第{show_level}级 " if show_level else ""
            version_id = getattr(row, "source_version_id", None)
            version_name = getattr(row, "source_version_name", None) or "未设置"
            return f"{prefix}版本号：{version_id if version_id is not None else '未设置'} 版本名称：{version_name}"
        version_id = getattr(row, "version_id", None)
        version_name = getattr(row, "version_name", None) or "未设置"
        return f"版本号：{version_id if version_id is not None else '未设置'} 版本名称：{version_name}"
    raw = getattr(row, field_id, None)
    text = str(raw or "").strip()
    return text or "未设置"


def _is_percent(value_type: str | None) -> bool:
    return bool(re.search(r"(%|百分|占比|比率|比例|收益率|利率|费率|率)", str(value_type or "")))


def _cell_number(value: float, percent_only: bool) -> float:
    if percent_only:
        return value / 100.0 if abs(value) > 1 else value
    return value


def _sort_tuple(values: tuple[str, ...]) -> tuple[str, ...]:
    return tuple(str(v or "") for v in values)


def build_pivot_aggregate_workbook(
    *,
    rows: list[BudgetSummaryRowDto | CompareSummaryRowDto],
    body: BudgetSummaryExportPivotRequest,
    title: str,
    source_label: str,
    org_product_refs_by_runtime_ref_code: dict[str, list[str]] | None = None,
) -> Workbook:
    row_fields = _field_ids(body.row_field_ids)
    column_fields = _field_ids(body.column_field_ids)
    page_fields = _field_ids(body.page_field_ids)
    row_defs = row_fields or ["__all__"]
    col_defs = column_fields or ["__value__"]
    refs_by_data = {
        str(code or "").strip().upper(): list(refs or [])
        for code, refs in (org_product_refs_by_runtime_ref_code or {}).items()
        if str(code or "").strip()
    }
    data_code_row_index = row_defs.index("data_code_name") if "data_code_name" in row_defs else -1
    include_org_product_refs = bool(refs_by_data) and data_code_row_index >= 0

    values: dict[tuple[tuple[str, ...], tuple[str, ...]], float] = {}
    value_types: dict[tuple[tuple[str, ...], tuple[str, ...]], set[str]] = {}
    row_keys: set[tuple[str, ...]] = set()
    col_keys: set[tuple[str, ...]] = set()
    org_product_refs_by_row: dict[tuple[str, ...], list[str]] = {}

    for row in rows:
        row_key = tuple("全部" if field == "__all__" else _field_value(row, field) for field in row_defs)
        col_key = tuple("预算数值" if field == "__value__" else _field_value(row, field) for field in col_defs)
        key = (row_key, col_key)
        values[key] = values.get(key, 0.0) + float(getattr(row, "value", 0.0) or 0.0)
        value_types.setdefault(key, set()).add(str(getattr(row, "value_type", "") or ""))
        row_keys.add(row_key)
        col_keys.add(col_key)
        if include_org_product_refs:
            data_code = extract_runtime_metric_ref_code(row_key[data_code_row_index])
            if data_code:
                refs = refs_by_data.get(data_code.upper(), [])
                if refs:
                    existing = org_product_refs_by_row.setdefault(row_key, [])
                    for ref in refs:
                        if ref not in existing:
                            existing.append(ref)

    sorted_rows = sorted(row_keys, key=_sort_tuple) or [("全部",)]
    sorted_cols = sorted(col_keys, key=_sort_tuple) or [("预算数值",)]

    wb = Workbook()
    ws = wb.active
    ws.title = "当前透视聚合结果"

    title_fill = PatternFill(fill_type="solid", fgColor="FF1F2937")
    title_font = Font(color="FFFFFFFF", bold=True, size=13)
    header_fill = PatternFill(fill_type="solid", fgColor="FFE5E7EB")
    subheader_fill = PatternFill(fill_type="solid", fgColor="FFF3F4F6")
    total_fill = PatternFill(fill_type="solid", fgColor="FFFFF7ED")
    thin = Side(style="thin", color="FFD1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    lineage_col_count = 2 if include_org_product_refs else 0
    last_value_col = len(row_defs) + len(sorted_cols) + (1 if body.show_column_total else 0)
    last_col = last_value_col + lineage_col_count
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(last_col, 4))
    c = ws.cell(row=1, column=1, value=title)
    c.fill = title_fill
    c.font = title_font
    c.alignment = Alignment(horizontal="left")

    exported_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.cell(row=2, column=1, value=f"数据来源：{source_label}")
    ws.cell(row=2, column=2, value=f"导出时间：{exported_at}")
    ws.cell(row=2, column=3, value=f"聚合行数：{len(rows)}")

    ws.cell(row=3, column=1, value="行字段")
    ws.cell(row=3, column=2, value="、".join(FIELD_LABELS.get(f, f) for f in row_fields) or "全部")
    ws.cell(row=4, column=1, value="列字段")
    ws.cell(row=4, column=2, value="、".join(FIELD_LABELS.get(f, f) for f in column_fields) or "预算数值")
    page_text = []
    for field in page_fields:
        selected = str(body.page_selections.get(field) or "全部").strip() or "全部"
        page_text.append(f"{FIELD_LABELS.get(field, field)}={selected}")
    ws.cell(row=5, column=1, value="页字段")
    ws.cell(row=5, column=2, value="；".join(page_text) or "无")
    ws.cell(row=6, column=1, value="搜索")
    ws.cell(row=6, column=2, value=body.pivot_search_text or "无")

    header_row = 8
    for idx, field in enumerate(row_defs, start=1):
        ws.cell(row=header_row, column=idx, value="行" if field == "__all__" else FIELD_LABELS.get(field, field))
    for offset, col_key in enumerate(sorted_cols, start=len(row_defs) + 1):
        ws.cell(row=header_row, column=offset, value=" / ".join(col_key))
    if body.show_column_total:
        ws.cell(row=header_row, column=last_value_col, value="行合计")
    if include_org_product_refs:
        ws.cell(row=header_row, column=last_value_col + 1, value="机构产品引用数量")
        ws.cell(row=header_row, column=last_value_col + 2, value="机构产品来源")

    for col_idx in range(1, last_col + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    data_start = header_row + 1
    for row_offset, row_key in enumerate(sorted_rows):
        excel_row = data_start + row_offset
        for idx, label in enumerate(row_key, start=1):
            cell = ws.cell(row=excel_row, column=idx, value=label)
            cell.fill = subheader_fill
            cell.border = border
        row_total = 0.0
        row_percent_flags: list[bool] = []
        for col_offset, col_key in enumerate(sorted_cols, start=len(row_defs) + 1):
            key = (row_key, col_key)
            raw_value = values.get(key, 0.0)
            row_total += raw_value
            percent_only = bool(value_types.get(key)) and all(_is_percent(v) for v in value_types.get(key, set()))
            row_percent_flags.append(percent_only)
            cell = ws.cell(row=excel_row, column=col_offset, value=_cell_number(raw_value, percent_only))
            cell.number_format = "0.00%;[Red]-0.00%" if percent_only else "#,##0.00;[Red]-#,##0.00"
            cell.border = border
            cell.alignment = Alignment(horizontal="right")
        if body.show_column_total:
            percent_total = bool(row_percent_flags) and all(row_percent_flags)
            cell = ws.cell(row=excel_row, column=last_value_col, value=_cell_number(row_total, percent_total))
            cell.fill = total_fill
            cell.font = Font(bold=True)
            cell.number_format = "0.00%;[Red]-0.00%" if percent_total else "#,##0.00;[Red]-#,##0.00"
            cell.border = border
            cell.alignment = Alignment(horizontal="right")
        if include_org_product_refs:
            org_refs = sorted(org_product_refs_by_row.get(row_key, []))
            count_cell = ws.cell(row=excel_row, column=last_value_col + 1, value=len(org_refs))
            refs_cell = ws.cell(row=excel_row, column=last_value_col + 2, value="\n".join(org_refs))
            for cell in (count_cell, refs_cell):
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    if body.show_row_total and sorted_rows:
        total_row = data_start + len(sorted_rows)
        ws.cell(row=total_row, column=1, value="列合计")
        for idx in range(1, len(row_defs) + 1):
            cell = ws.cell(row=total_row, column=idx)
            cell.fill = total_fill
            cell.font = Font(bold=True)
            cell.border = border
        grand_total = 0.0
        grand_percent_flags: list[bool] = []
        for col_offset, col_key in enumerate(sorted_cols, start=len(row_defs) + 1):
            col_value = sum(values.get((row_key, col_key), 0.0) for row_key in sorted_rows)
            grand_total += col_value
            col_types = set()
            for row_key in sorted_rows:
                col_types.update(value_types.get((row_key, col_key), set()))
            percent_only = bool(col_types) and all(_is_percent(v) for v in col_types)
            grand_percent_flags.append(percent_only)
            cell = ws.cell(row=total_row, column=col_offset, value=_cell_number(col_value, percent_only))
            cell.fill = total_fill
            cell.font = Font(bold=True)
            cell.number_format = "0.00%;[Red]-0.00%" if percent_only else "#,##0.00;[Red]-#,##0.00"
            cell.border = border
            cell.alignment = Alignment(horizontal="right")
        if body.show_column_total:
            percent_total = bool(grand_percent_flags) and all(grand_percent_flags)
            cell = ws.cell(row=total_row, column=last_value_col, value=_cell_number(grand_total, percent_total))
            cell.fill = total_fill
            cell.font = Font(bold=True)
            cell.number_format = "0.00%;[Red]-0.00%" if percent_total else "#,##0.00;[Red]-#,##0.00"
            cell.border = border
            cell.alignment = Alignment(horizontal="right")
        if include_org_product_refs:
            all_refs: list[str] = []
            for row_key in sorted_rows:
                for ref in org_product_refs_by_row.get(row_key, []):
                    if ref not in all_refs:
                        all_refs.append(ref)
            sorted_refs = sorted(all_refs)
            count_cell = ws.cell(row=total_row, column=last_value_col + 1, value=len(sorted_refs))
            refs_cell = ws.cell(row=total_row, column=last_value_col + 2, value="\n".join(sorted_refs))
            for cell in (count_cell, refs_cell):
                cell.fill = total_fill
                cell.font = Font(bold=True)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = ws.cell(row=data_start, column=len(row_defs) + 1)
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=2, max_row=6, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, min(ws.max_row, 2000) + 1):
            text = str(ws.cell(row=row_idx, column=col_idx).value or "")
            max_len = max(max_len, max((len(part) for part in text.splitlines()), default=0))
        ws.column_dimensions[col_letter].width = max(10, min(42, max_len + 2))

    return wb


def aggregate_workbook_response(wb: Workbook, filename: str):
    return workbook_streaming_response(wb, filename=filename)
