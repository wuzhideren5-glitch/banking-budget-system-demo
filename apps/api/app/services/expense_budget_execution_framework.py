"""Expense budget execution framework parsing and context service."""
from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
import re
import sqlite3
import tempfile
from typing import Any

from openpyxl import load_workbook
import xlrd

from app.core.config import settings
from app.core.database import get_pool
from app.core.db_paths import common_db_path
from app.services.department_expense_contracts import (
    BUDGET_SUBJECT_LEVEL_NUMBER_TO_NAME,
    DEPT_OWNER_LEVEL,
)


class ExpenseFrameworkError(ValueError):
    """Raised when the expense framework cannot be parsed or loaded."""


@dataclass(frozen=True)
class FrameworkBudgetDepartmentRow:
    entity_name: str
    group_name: str
    owner_name: str
    budget_department: str


@dataclass(frozen=True)
class FrameworkProductDepartmentRow:
    entity_name: str
    group_name: str
    owner_name: str
    product_department: str


@dataclass(frozen=True)
class FrameworkSubjectRow:
    level_label: str
    budget_subject: str
    manage_department: str
    formula_text: str
    sort_order: int


@dataclass(frozen=True)
class ParsedFramework:
    source_file: Path
    budget_departments: list[FrameworkBudgetDepartmentRow]
    product_departments: list[FrameworkProductDepartmentRow]
    subjects: list[FrameworkSubjectRow]


class FrameworkContext:
    def __init__(self) -> None:
        self.owner_alias_map: dict[str, str] = {}
        self.owner_to_entity: dict[str, str] = {}
        self.owner_to_group: dict[str, str] = {}
        self.group_to_entity: dict[str, str] = {}
        self.product_department_to_entity: dict[str, str] = {}
        self.product_department_to_owner: dict[str, str] = {}
        self.product_department_to_group: dict[str, str] = {}
        self.subject_alias_map: dict[str, str] = {}


GOVERNANCE_OWNER_ALIASES = {
    "董事会办公室": "公司治理部",
    "监事会办公室": "公司治理部",
    "董监事会办公室": "公司治理部",
}
MANAGE_DEPARTMENT_ALIASES = {
    "董事会办公室": "公司治理部",
    "监事会办公室": "公司治理部",
    "董监事会办公室": "公司治理部",
}
ENTITY_ORDER = ["微众银行", "科技子", "科技孙"]
GROUP_ORDER = [
    "个人金融事业群", "企业及机构金融事业群", "科技及智能事业群", "国际发展部", "国际业务",
    "资源管理及管控职能群", "其他", "历史架构", "科技子", "科技孙", "虚拟架构",
]


def text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def norm_key(value: str) -> str:
    return re.sub(r"\s+", "", text(value)).lower()


def strip_leading_code(value: str) -> str:
    raw_text = text(value)
    stripped = re.sub(r"^[A-Za-z]+\d+\s+", "", raw_text)
    return text(stripped or raw_text)


def canonical_subject(raw_name: str, ctx: FrameworkContext) -> str:
    raw_text = text(raw_name)
    if not raw_text:
        return ""
    return ctx.subject_alias_map.get(norm_key(raw_text), raw_text)


def canonical_owner_name(owner_name: str, ctx: FrameworkContext) -> str:
    raw_text = text(owner_name)
    if not raw_text:
        return ""
    return ctx.owner_alias_map.get(norm_key(raw_text), ctx.owner_alias_map.get(norm_key(strip_leading_code(raw_text)), raw_text))


def default_entity_name() -> str:
    return "科技子"


def entity_sort_key(entity_name: str) -> tuple[int, str]:
    raw_text = text(entity_name)
    try:
        return (ENTITY_ORDER.index(raw_text), raw_text)
    except ValueError:
        return (len(ENTITY_ORDER), raw_text)


def group_sort_key(group_name: str) -> tuple[int, str]:
    raw_text = text(group_name)
    try:
        return (GROUP_ORDER.index(raw_text), raw_text)
    except ValueError:
        return (len(GROUP_ORDER), raw_text)


def list_available_entities(ctx: FrameworkContext) -> list[str]:
    values = {
        text(value)
        for value in (
            list(ctx.owner_to_entity.values())
            + list(ctx.group_to_entity.values())
            + list(ctx.product_department_to_entity.values())
        )
        if text(value)
    }
    return sorted(values, key=entity_sort_key)


def entity_for_group(group_name: str, ctx: FrameworkContext) -> str:
    return ctx.group_to_entity.get(group_name, default_entity_name())


def entity_for_owner(owner_name: str, ctx: FrameworkContext) -> str:
    return ctx.owner_to_entity.get(owner_name, default_entity_name())


def default_group_name(entity_name: str) -> str:
    return entity_name or default_entity_name()


def should_exclude_budget_department_row(row: FrameworkBudgetDepartmentRow) -> bool:
    return row.entity_name == "微众银行" and (
        row.owner_name.startswith("科技子") or row.owner_name.startswith("科技孙")
    )


def normalize_leaf_department_name(
    entity_name: str,
    group_name: str,
    owner_name: str,
    department_name: str,
) -> str:
    raw_text = text(department_name)
    if raw_text:
        return raw_text
    if owner_name and (owner_name == group_name or owner_name == entity_name):
        return owner_name
    return ""


def compose_subject_remark(subject: FrameworkSubjectRow, existing_remark: str | None = None) -> str | None:
    parts: list[str] = []
    base = text(existing_remark)
    if base:
        parts.append(base)
    if subject.manage_department:
        token = f"框架归口管理部门:{subject.manage_department}"
        if token not in parts:
            parts.append(token)
    if subject.formula_text and subject.formula_text != "0":
        token = f"框架公式:{subject.formula_text}"
        if token not in parts:
            parts.append(token)
    return "；".join(parts) or None


def _iso_now() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _mtime_text(path: Path) -> str | None:
    try:
        return datetime.fromtimestamp(path.stat().st_mtime, timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    except Exception:
        return None


def _row_value(row: Any, key: str, index: int) -> Any:
    if isinstance(row, dict):
        return row.get(key)
    keys = getattr(row, "keys", None)
    if callable(keys) and key in keys():
        return row[key]
    return row[index]


def _uses_mysql_path(path: Path | str) -> bool:
    try:
        candidate = Path(path).expanduser().resolve()
    except TypeError:
        return False
    temp_root = Path(tempfile.gettempdir()).expanduser().resolve()
    try:
        candidate.relative_to(temp_root)
        return False
    except ValueError:
        pass
    data_dir = Path(settings.data_dir).expanduser().resolve()
    try:
        candidate.relative_to(data_dir)
    except ValueError:
        return False
    return candidate.name == "common.db"


def _mysql_sql(sql: str) -> str:
    return sql.replace("?", "%s")


async def _fetch_all_for_path(db_path: Path, sql: str, params: tuple[Any, ...] = ()) -> list[Any]:
    if _uses_mysql_path(db_path):
        return await get_pool().fetch_all(_mysql_sql(sql), params)
    with sqlite3.connect(db_path) as db:
        db.execute("PRAGMA foreign_keys = ON")
        return db.execute(sql, params).fetchall()


async def _execute_for_path(db_path: Path, sql: str, params: tuple[Any, ...] = ()) -> int:
    if _uses_mysql_path(db_path):
        return await get_pool().execute(_mysql_sql(sql), params)
    with sqlite3.connect(db_path) as db:
        db.execute("PRAGMA foreign_keys = ON")
        cur = db.execute(sql, params)
        db.commit()
        return cur.rowcount


def parse_framework_source_bytes(source_file: str, raw: bytes) -> ParsedFramework:
    source_name = text(source_file).lower()
    is_xls = source_name.endswith(".xls")
    if is_xls:
        wb = xlrd.open_workbook(file_contents=raw)
        if "预算部门" not in wb.sheet_names() or "部门预算科目" not in wb.sheet_names():
            raise ExpenseFrameworkError("上传文件缺少“预算部门”或“部门预算科目”工作表")
    else:
        wb = load_workbook(BytesIO(raw), data_only=True)
        if not {"预算部门", "部门预算科目"}.issubset(set(wb.sheetnames)):
            raise ExpenseFrameworkError("上传文件缺少“预算部门”或“部门预算科目”工作表")

    def _cell_value(sheet: Any, row: int, col: int) -> Any:
        if is_xls:
            return sheet.cell_value(row, col)
        return sheet.cell(row + 1, col + 1).value

    def _sheet_rows(sheet: Any) -> range:
        return range(1, sheet.nrows) if is_xls else range(1, sheet.max_row)

    budget_rows: list[FrameworkBudgetDepartmentRow] = []
    seen_budget: set[tuple[str, str, str, str]] = set()
    ws_budget = wb.sheet_by_name("预算部门") if is_xls else wb["预算部门"]
    for row_idx in _sheet_rows(ws_budget):
        entity_name = text(_cell_value(ws_budget, row_idx, 0))
        group_name = text(_cell_value(ws_budget, row_idx, 1))
        owner_name = text(_cell_value(ws_budget, row_idx, 2))
        row = FrameworkBudgetDepartmentRow(
            entity_name=entity_name,
            group_name=group_name,
            owner_name=owner_name,
            budget_department=normalize_leaf_department_name(
                entity_name,
                group_name,
                owner_name,
                text(_cell_value(ws_budget, row_idx, 3)),
            ),
        )
        key = (row.entity_name, row.group_name, row.owner_name, row.budget_department)
        if not all(key) or key in seen_budget:
            continue
        seen_budget.add(key)
        budget_rows.append(row)

    product_rows: list[FrameworkProductDepartmentRow] = []
    if is_xls:
        if "产品部门" in wb.sheet_names():
            product_sheet_name = "产品部门"
        elif "产品预算科目" in wb.sheet_names():
            product_sheet_name = "产品预算科目"
        else:
            product_sheet_name = ""
        ws_product = wb.sheet_by_name(product_sheet_name) if product_sheet_name else None
    else:
        product_sheet_name = "产品部门" if "产品部门" in wb.sheetnames else "产品预算科目" if "产品预算科目" in wb.sheetnames else ""
        ws_product = wb[product_sheet_name] if product_sheet_name else None
    if ws_product is not None:
        seen_product: set[tuple[str, str, str, str]] = set()
        for row_idx in _sheet_rows(ws_product):
            entity_name = text(_cell_value(ws_product, row_idx, 0))
            group_name = text(_cell_value(ws_product, row_idx, 1))
            owner_name = text(_cell_value(ws_product, row_idx, 2))
            row = FrameworkProductDepartmentRow(
                entity_name=entity_name,
                group_name=group_name,
                owner_name=owner_name,
                product_department=normalize_leaf_department_name(
                    entity_name,
                    group_name,
                    owner_name,
                    text(_cell_value(ws_product, row_idx, 3)),
                ),
            )
            key = (row.entity_name, row.group_name, row.owner_name, row.product_department)
            if not all(key) or key in seen_product:
                continue
            seen_product.add(key)
            product_rows.append(row)

    subjects: list[FrameworkSubjectRow] = []
    seen_subject: set[str] = set()
    ws_subject = wb.sheet_by_name("部门预算科目") if is_xls else wb["部门预算科目"]
    for row_idx in _sheet_rows(ws_subject):
        subject_name = text(_cell_value(ws_subject, row_idx, 1))
        if not subject_name or subject_name in seen_subject:
            continue
        seen_subject.add(subject_name)
        subjects.append(
            FrameworkSubjectRow(
                level_label=text(_cell_value(ws_subject, row_idx, 0)),
                budget_subject=subject_name,
                manage_department=text(_cell_value(ws_subject, row_idx, 2)),
                formula_text=text(_cell_value(ws_subject, row_idx, 3)),
                sort_order=len(subjects) + 1,
            )
        )

    return ParsedFramework(
        source_file=Path(source_file),
        budget_departments=budget_rows,
        product_departments=product_rows,
        subjects=subjects,
    )


async def read_sync_meta() -> dict[str, dict[str, Any]]:
    rows = await _fetch_all_for_path(
        common_db_path(),
        "SELECT sync_key, source_file, source_mtime, synced_at, row_count, note FROM expense_sync_meta",
    )
    return {
        str(_row_value(row, "sync_key", 0)): {
            "source_file": str(_row_value(row, "source_file", 1)),
            "source_mtime": (
                str(_row_value(row, "source_mtime", 2))
                if _row_value(row, "source_mtime", 2) is not None
                else None
            ),
            "synced_at": str(_row_value(row, "synced_at", 3)),
            "row_count": int(_row_value(row, "row_count", 4) or 0),
            "note": str(_row_value(row, "note", 5)) if _row_value(row, "note", 5) is not None else None,
        }
        for row in rows
    }


async def upsert_sync_meta(sync_key: str, source_file: str, row_count: int, note: str | None = None) -> None:
    db_path = common_db_path()
    params = (
        sync_key,
        source_file,
        _mtime_text(Path(source_file)) if source_file else None,
        _iso_now(),
        int(row_count),
        note,
    )
    if _uses_mysql_path(db_path):
        await get_pool().execute(
            """
            INSERT INTO expense_sync_meta(sync_key, source_file, source_mtime, synced_at, row_count, note)
            VALUES (%s, %s, %s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE
              source_file = VALUES(source_file),
              source_mtime = VALUES(source_mtime),
              synced_at = VALUES(synced_at),
              row_count = VALUES(row_count),
              note = VALUES(note)
            """,
            params,
        )
        return
    await _execute_for_path(
        db_path,
        """
        INSERT INTO expense_sync_meta(sync_key, source_file, source_mtime, synced_at, row_count, note)
        VALUES (?, ?, ?, ?, ?, ?)
        ON CONFLICT(sync_key) DO UPDATE SET
          source_file = excluded.source_file,
          source_mtime = excluded.source_mtime,
          synced_at = excluded.synced_at,
          row_count = excluded.row_count,
          note = excluded.note
        """,
        params,
    )


async def _persist_framework_snapshot_mysql(
    filtered_budget_departments: list[FrameworkBudgetDepartmentRow],
    parsed: ParsedFramework,
) -> None:
    async with get_pool().acquire() as conn:
        await conn.begin()
        try:
            async with conn.cursor() as cur:
                await cur.execute("DELETE FROM expense_framework_budget_department")
                await cur.execute("DELETE FROM expense_framework_product_department")
                await cur.execute("DELETE FROM expense_framework_subject")
                if filtered_budget_departments:
                    await cur.executemany(
                        """
                        INSERT INTO expense_framework_budget_department(entity_name, group_name, owner_name, budget_department)
                        VALUES (%s, %s, %s, %s)
                        """,
                        [(r.entity_name, r.group_name, r.owner_name, r.budget_department) for r in filtered_budget_departments],
                    )
                if parsed.product_departments:
                    await cur.executemany(
                        """
                        INSERT INTO expense_framework_product_department(entity_name, group_name, owner_name, product_department)
                        VALUES (%s, %s, %s, %s)
                        """,
                        [(r.entity_name, r.group_name, r.owner_name, r.product_department) for r in parsed.product_departments],
                    )
                if parsed.subjects:
                    await cur.executemany(
                        """
                        INSERT INTO expense_framework_subject(budget_subject, level_label, manage_department, formula_text, sort_order)
                        VALUES (%s, %s, %s, %s, %s)
                        """,
                        [(r.budget_subject, r.level_label, r.manage_department, r.formula_text, r.sort_order) for r in parsed.subjects],
                    )
            await conn.commit()
        except Exception:
            await conn.rollback()
            raise


async def _persist_framework_snapshot_sqlite(
    db_path: Path,
    filtered_budget_departments: list[FrameworkBudgetDepartmentRow],
    parsed: ParsedFramework,
) -> None:
    with sqlite3.connect(db_path) as db:
        db.execute("PRAGMA foreign_keys = ON")
        db.execute("DELETE FROM expense_framework_budget_department")
        db.execute("DELETE FROM expense_framework_product_department")
        db.execute("DELETE FROM expense_framework_subject")
        if filtered_budget_departments:
            db.executemany(
                """
                INSERT INTO expense_framework_budget_department(entity_name, group_name, owner_name, budget_department)
                VALUES (?, ?, ?, ?)
                """,
                [(r.entity_name, r.group_name, r.owner_name, r.budget_department) for r in filtered_budget_departments],
            )
        if parsed.product_departments:
            db.executemany(
                """
                INSERT INTO expense_framework_product_department(entity_name, group_name, owner_name, product_department)
                VALUES (?, ?, ?, ?)
                """,
                [(r.entity_name, r.group_name, r.owner_name, r.product_department) for r in parsed.product_departments],
            )
        if parsed.subjects:
            db.executemany(
                """
                INSERT INTO expense_framework_subject(budget_subject, level_label, manage_department, formula_text, sort_order)
                VALUES (?, ?, ?, ?, ?)
                """,
                [(r.budget_subject, r.level_label, r.manage_department, r.formula_text, r.sort_order) for r in parsed.subjects],
            )
        db.commit()


async def persist_framework_snapshot(parsed: ParsedFramework) -> None:
    filtered_budget_departments = [
        row for row in parsed.budget_departments if not should_exclude_budget_department_row(row)
    ]
    db_path = common_db_path()
    if _uses_mysql_path(db_path):
        await _persist_framework_snapshot_mysql(filtered_budget_departments, parsed)
    else:
        await _persist_framework_snapshot_sqlite(db_path, filtered_budget_departments, parsed)
    total_rows = len(filtered_budget_departments) + len(parsed.product_departments) + len(parsed.subjects)
    await upsert_sync_meta(
        "framework_import",
        str(parsed.source_file),
        total_rows,
        note=(
            f"预算部门{len(filtered_budget_departments)}行；"
            f"产品部门{len(parsed.product_departments)}行；"
            f"部门预算科目{len(parsed.subjects)}行"
        ),
    )


async def load_framework_from_db() -> ParsedFramework | None:
    db_path = common_db_path()
    budget_db_rows = await _fetch_all_for_path(
        db_path,
        "SELECT entity_name, group_name, owner_name, budget_department FROM expense_framework_budget_department ORDER BY id",
    )
    budget_rows = [
        FrameworkBudgetDepartmentRow(
            entity_name=text(_row_value(row, "entity_name", 0)),
            group_name=text(_row_value(row, "group_name", 1)),
            owner_name=text(_row_value(row, "owner_name", 2)),
            budget_department=text(_row_value(row, "budget_department", 3)),
        )
        for row in budget_db_rows
    ]
    product_db_rows = await _fetch_all_for_path(
        db_path,
        "SELECT entity_name, group_name, owner_name, product_department FROM expense_framework_product_department ORDER BY id",
    )
    product_rows = [
        FrameworkProductDepartmentRow(
            entity_name=text(_row_value(row, "entity_name", 0)),
            group_name=text(_row_value(row, "group_name", 1)),
            owner_name=text(_row_value(row, "owner_name", 2)),
            product_department=text(_row_value(row, "product_department", 3)),
        )
        for row in product_db_rows
    ]
    subject_db_rows = await _fetch_all_for_path(
        db_path,
        """
        SELECT budget_subject, level_label, manage_department, formula_text, sort_order
        FROM expense_framework_subject
        ORDER BY sort_order, budget_subject
        """,
    )
    subject_rows = [
        FrameworkSubjectRow(
            level_label=text(_row_value(row, "level_label", 1)),
            budget_subject=text(_row_value(row, "budget_subject", 0)),
            manage_department=text(_row_value(row, "manage_department", 2)),
            formula_text=text(_row_value(row, "formula_text", 3)),
            sort_order=int(_row_value(row, "sort_order", 4) or 0),
        )
        for row in subject_db_rows
    ]
    if not budget_rows and not product_rows and not subject_rows:
        return None
    meta = await read_sync_meta()
    source_file = meta.get("framework_import", {}).get("source_file") or "expense_framework_internal"
    return ParsedFramework(
        source_file=Path(source_file),
        budget_departments=budget_rows,
        product_departments=product_rows,
        subjects=subject_rows,
    )


def derive_product_departments_from_budget_departments(
    budget_rows: list[FrameworkBudgetDepartmentRow],
) -> list[FrameworkProductDepartmentRow]:
    seen: set[tuple[str, str, str, str]] = set()
    product_rows: list[FrameworkProductDepartmentRow] = []
    for row in budget_rows:
        key = (row.entity_name, row.group_name, row.owner_name, row.budget_department)
        if not all(key) or key in seen:
            continue
        seen.add(key)
        product_rows.append(
            FrameworkProductDepartmentRow(
                entity_name=row.entity_name,
                group_name=row.group_name,
                owner_name=row.owner_name,
                product_department=row.budget_department,
            )
        )
    return product_rows


async def merge_framework_with_existing(parsed: ParsedFramework) -> ParsedFramework:
    existing = await load_framework_from_db()
    budget_departments = parsed.budget_departments or (existing.budget_departments if existing else [])
    return ParsedFramework(
        source_file=parsed.source_file,
        budget_departments=budget_departments,
        product_departments=derive_product_departments_from_budget_departments(budget_departments),
        subjects=parsed.subjects or (existing.subjects if existing else []),
    )


def build_framework_context(parsed: ParsedFramework) -> FrameworkContext:
    ctx = FrameworkContext()
    for row in parsed.budget_departments:
        if row.owner_name:
            ctx.owner_alias_map[norm_key(row.owner_name)] = row.owner_name
            ctx.owner_alias_map[norm_key(strip_leading_code(row.owner_name))] = row.owner_name
        if row.owner_name and row.entity_name and row.owner_name not in ctx.owner_to_entity:
            ctx.owner_to_entity[row.owner_name] = row.entity_name
        if row.group_name and row.entity_name and row.group_name not in ctx.group_to_entity:
            ctx.group_to_entity[row.group_name] = row.entity_name
        if row.owner_name and row.group_name and row.owner_name not in ctx.owner_to_group:
            ctx.owner_to_group[row.owner_name] = row.group_name

    for row in parsed.product_departments:
        if row.owner_name:
            ctx.owner_alias_map[norm_key(row.owner_name)] = row.owner_name
            ctx.owner_alias_map[norm_key(strip_leading_code(row.owner_name))] = row.owner_name
        if row.product_department:
            ctx.product_department_to_entity[row.product_department] = row.entity_name
            ctx.product_department_to_owner[row.product_department] = row.owner_name
            ctx.product_department_to_group[row.product_department] = row.group_name
        if row.group_name and row.entity_name and row.group_name not in ctx.group_to_entity:
            ctx.group_to_entity[row.group_name] = row.entity_name
        if row.owner_name and row.entity_name and row.owner_name not in ctx.owner_to_entity:
            ctx.owner_to_entity[row.owner_name] = row.entity_name
        if row.owner_name and row.group_name and row.owner_name not in ctx.owner_to_group:
            ctx.owner_to_group[row.owner_name] = row.group_name

    for subject in parsed.subjects:
        if subject.budget_subject:
            ctx.subject_alias_map[norm_key(subject.budget_subject)] = subject.budget_subject

    for alias_name, canonical_name in GOVERNANCE_OWNER_ALIASES.items():
        ctx.owner_alias_map[norm_key(alias_name)] = canonical_name
        ctx.owner_alias_map[norm_key(strip_leading_code(alias_name))] = canonical_name

    return ctx


async def load_runtime_context_from_master_data() -> ParsedFramework | None:
    db_path = common_db_path()
    dept_rows = await _fetch_all_for_path(
        db_path,
        """
        SELECT
            COALESCE(NULLIF(TRIM(child.entity_name), ''), '微众银行') AS entity_name,
            COALESCE(NULLIF(TRIM(parent.dept_name), ''), '') AS group_name,
            COALESCE(NULLIF(TRIM(child.dept_name), ''), '') AS owner_name
        FROM dept_account child
        LEFT JOIN dept_account parent
          ON parent.dept_code = child.parent_code
        WHERE child.level = ?
        ORDER BY entity_name, group_name, owner_name
        """,
        (DEPT_OWNER_LEVEL,),
    )
    subject_rows = await _fetch_all_for_path(
        db_path,
        """
        SELECT level_number, subject_name, manage_department, formula_text, sort_order
        FROM budget_subject_catalog
        ORDER BY sort_order, id
        """,
    )

    budget_rows = []
    for row in dept_rows:
        entity_name = text(_row_value(row, "entity_name", 0))
        group_name = text(_row_value(row, "group_name", 1))
        owner_name = text(_row_value(row, "owner_name", 2))
        if not entity_name or not group_name or not owner_name:
            continue
        budget_row = FrameworkBudgetDepartmentRow(
            entity_name=entity_name,
            group_name=group_name,
            owner_name=owner_name,
            budget_department=owner_name,
        )
        if should_exclude_budget_department_row(budget_row):
            continue
        budget_rows.append(budget_row)
    subjects = [
        FrameworkSubjectRow(
            level_label=BUDGET_SUBJECT_LEVEL_NUMBER_TO_NAME.get(
                int(_row_value(row, "level_number", 0) or 0),
                f"{int(_row_value(row, 'level_number', 0) or 0)}级",
            ),
            budget_subject=text(_row_value(row, "subject_name", 1)),
            manage_department=text(_row_value(row, "manage_department", 2)),
            formula_text=text(_row_value(row, "formula_text", 3)),
            sort_order=int(_row_value(row, "sort_order", 4) or 0),
        )
        for row in subject_rows
        if text(_row_value(row, "subject_name", 1))
    ]
    if not budget_rows and not subjects:
        return None
    return ParsedFramework(
        source_file=Path("system-master-data"),
        budget_departments=budget_rows,
        product_departments=[],
        subjects=subjects,
    )


async def load_framework_context() -> tuple[FrameworkContext, str, str, ParsedFramework]:
    master_parsed = await load_runtime_context_from_master_data()
    if master_parsed is not None and master_parsed.budget_departments:
        return build_framework_context(master_parsed), "master", "系统主数据（部门科目维护/部门预算科目维护）", master_parsed
    raise ExpenseFrameworkError("费用预算执行报表缺少当前部门费用主数据，请先在“部门科目维护”和“部门预算科目维护”中维护或同步当前目录。")


def build_template_scope_options(parsed: ParsedFramework) -> list[dict[str, str]]:
    seen: set[tuple[str, str, str]] = set()
    items: list[dict[str, str]] = []
    for row in list(parsed.budget_departments) + list(parsed.product_departments):
        entity_name = text(row.entity_name)
        group_name = text(row.group_name)
        owner_name = text(row.owner_name)
        if not entity_name or not group_name or not owner_name:
            continue
        key = (entity_name, group_name, owner_name)
        if key in seen:
            continue
        seen.add(key)
        items.append(
            {
                "entity_name": entity_name,
                "group_name": group_name,
                "owner_dept": owner_name,
            }
        )
    return sorted(items, key=lambda item: (entity_sort_key(item["entity_name"]), group_sort_key(item["group_name"]), len(item["owner_dept"]), item["owner_dept"]))


def matches_template_scope(
    *,
    ctx: FrameworkContext,
    owner_name: str,
    selected_entity: str,
    selected_group: str,
    selected_owner: str,
) -> bool:
    normalized_owner = text(owner_name)
    if not normalized_owner:
        return False
    resolved_entity = entity_for_owner(normalized_owner, ctx)
    resolved_group = ctx.owner_to_group.get(
        normalized_owner,
        default_group_name(resolved_entity),
    )
    if selected_entity and resolved_entity != selected_entity:
        return False
    if selected_group and resolved_group != selected_group:
        return False
    if selected_owner and normalized_owner != selected_owner:
        return False
    return True


def normalized_manage_department(raw_manage_department: str) -> str:
    value = text(raw_manage_department)
    if not value or value == "使用部门":
        return ""
    return MANAGE_DEPARTMENT_ALIASES.get(value, value)


def effective_manage_departments(
    subject_rows: list[dict[str, Any]],
) -> tuple[dict[int, str], dict[str, list[str]]]:
    row_by_id = {int(row["id"]): row for row in subject_rows}
    children_by_parent: dict[int | None, list[int]] = defaultdict(list)
    for row in subject_rows:
        children_by_parent[row["parent_id"]].append(int(row["id"]))

    effective_by_id: dict[int, str] = {}
    effective_by_name: dict[str, list[str]] = defaultdict(list)

    def _walk(node_id: int, inherited_department: str) -> None:
        row = row_by_id[node_id]
        current_department = normalized_manage_department(text(row.get("manage_department"))) or inherited_department
        effective_by_id[node_id] = current_department
        subject_name = text(row.get("subject_name"))
        if subject_name and current_department and current_department not in effective_by_name[subject_name]:
            effective_by_name[subject_name].append(current_department)
        for child_id in children_by_parent.get(node_id, []):
            _walk(child_id, current_department)

    for root_id in children_by_parent.get(None, []):
        _walk(root_id, "")
    return effective_by_id, effective_by_name


def resolve_department_scope(
    *,
    ctx: FrameworkContext,
    department_name: str,
) -> tuple[str, str, str]:
    normalized_department = normalized_manage_department(department_name)
    if not normalized_department:
        return "", "", ""
    if normalized_department in ctx.owner_to_entity:
        entity_name = entity_for_owner(normalized_department, ctx)
        group_name = ctx.owner_to_group.get(
            normalized_department,
            default_group_name(entity_name),
        )
        return entity_name, group_name, normalized_department
    if normalized_department in ctx.group_to_entity:
        entity_name = entity_for_group(normalized_department, ctx)
        return entity_name, normalized_department, ""
    return entity_for_owner(normalized_department, ctx), "", normalized_department


def scope_allows_manage_department(
    *,
    ctx: FrameworkContext,
    manage_department: str,
    selected_entity: str,
    selected_group: str,
    selected_owner: str,
) -> bool:
    normalized_department = normalized_manage_department(manage_department)
    if not normalized_department:
        return True
    department_entity, department_group, department_owner = resolve_department_scope(
        ctx=ctx,
        department_name=normalized_department,
    )
    if selected_entity and department_entity != selected_entity:
        return False
    if selected_group and department_group != selected_group:
        return False
    if selected_owner and department_owner != selected_owner:
        return False
    return True


def subject_visible_for_scope(
    *,
    ctx: FrameworkContext,
    manage_departments: list[str],
    selected_entity: str,
    selected_group: str,
    selected_owner: str,
) -> bool:
    if not manage_departments:
        return True
    return any(
        scope_allows_manage_department(
            ctx=ctx,
            manage_department=manage_department,
            selected_entity=selected_entity,
            selected_group=selected_group,
            selected_owner=selected_owner,
        )
        for manage_department in manage_departments
    )
