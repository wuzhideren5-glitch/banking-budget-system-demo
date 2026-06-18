from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
import sqlite3
import tempfile
from typing import Any, Awaitable, Callable

from fastapi import HTTPException
from openpyxl import Workbook

from app.core.config import settings
from app.core.database import get_pool
from app.db_bootstrap.expense import (
    ensure_department_expense_master_schema,
    ensure_department_expense_master_schema_sync,
)
from app.schemas import (
    BudgetSubjectCatalogCreate,
    BudgetSubjectCatalogRow,
    BudgetSubjectCatalogUpdate,
)
from app.services.department_expense_contracts import (
    BUDGET_SUBJECT_LEVEL_NAME_TO_NUMBER,
    BUDGET_SUBJECT_LEVEL_NUMBER_TO_NAME,
    MAX_BUDGET_SUBJECT_LEVEL,
    MIN_BUDGET_SUBJECT_LEVEL,
)


EnsureSchema = Callable[[Any], Awaitable[None]]

LEVEL_NAME_TO_NUMBER = BUDGET_SUBJECT_LEVEL_NAME_TO_NUMBER
LEVEL_NUMBER_TO_NAME = BUDGET_SUBJECT_LEVEL_NUMBER_TO_NAME


@dataclass(frozen=True)
class BudgetSubjectCatalogDeletion:
    subject_name: str


@dataclass(frozen=True)
class BudgetSubjectCatalogWorkbook:
    content: bytes
    filename: str = "budget_subject_catalog.xlsx"


def text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _uses_mysql_path(path: Path | str) -> bool:
    try:
        candidate = Path(path).expanduser().resolve()
    except TypeError:
        return False
    temp_root = Path(tempfile.gettempdir()).expanduser().resolve()
    try:
        candidate.relative_to(temp_root)
        return False
    except ValueError:
        pass

    data_dir = Path(settings.data_dir).expanduser().resolve()
    try:
        candidate.relative_to(data_dir)
    except ValueError:
        return False
    return candidate.name == "common.db"


def _row_value(row: Any, key: str, index: int) -> Any:
    if isinstance(row, dict):
        return row.get(key)
    try:
        return row[key]
    except (TypeError, KeyError, IndexError):
        return row[index]


async def ensure_budget_subject_catalog_bootstrapped(
    common_db: Path | str,
    *,
    ensure_schema: EnsureSchema = ensure_department_expense_master_schema,
) -> None:
    if _uses_mysql_path(common_db):
        existing_count = await get_pool().fetch_val("SELECT COUNT(*) FROM budget_subject_catalog")
        if int(existing_count or 0) > 0:
            return

        framework_rows = await get_pool().fetch_all(
            """
            SELECT level_label, budget_subject, manage_department, formula_text, sort_order
            FROM expense_framework_subject
            ORDER BY sort_order, budget_subject
            """
        )
        rows_to_insert = build_bootstrap_rows(
            [
                {
                    "level_label": text(_row_value(row, "level_label", 0)),
                    "subject_name": text(_row_value(row, "budget_subject", 1)),
                    "manage_department": text(_row_value(row, "manage_department", 2)) or None,
                    "formula_text": text(_row_value(row, "formula_text", 3)) or None,
                    "sort_order": int(_row_value(row, "sort_order", 4) or 0),
                }
                for row in framework_rows
            ]
        )
        if not rows_to_insert:
            return

        async with get_pool().acquire() as db:
            try:
                await db.begin()
                async with db.cursor() as cur:
                    parent_id_by_virtual_id: dict[int, int] = {}
                    for item in rows_to_insert:
                        await cur.execute(
                            """
                            INSERT INTO budget_subject_catalog(
                              parent_id, level_number, subject_name, manage_department, formula_text, sort_order
                            )
                            VALUES (%s, %s, %s, %s, %s, %s)
                            """,
                            (
                                parent_id_by_virtual_id.get(int(item["parent_virtual_id"]))
                                if item["parent_virtual_id"] is not None
                                else None,
                                int(item["level_number"]),
                                text(item["subject_name"]),
                                text(item.get("manage_department")) or None,
                                text(item["formula_text"]) or None,
                                int(item["sort_order"]),
                            ),
                        )
                        parent_id_by_virtual_id[int(item["virtual_id"])] = int(cur.lastrowid)
                await db.commit()
            except Exception:
                await db.rollback()
                raise
        return

    with sqlite3.connect(common_db) as db:
        db.execute("PRAGMA foreign_keys = ON")
        try:
            if ensure_schema is ensure_department_expense_master_schema:
                ensure_department_expense_master_schema_sync(db)
            else:
                await ensure_schema(db)
        except RuntimeError as exc:
            raise HTTPException(status_code=500, detail=str(exc)) from exc
        cur = db.execute("SELECT COUNT(*) FROM budget_subject_catalog")
        if int((cur.fetchone())[0] or 0) > 0:
            return

        cur = db.execute(
            """
            SELECT level_label, budget_subject, manage_department, formula_text, sort_order
            FROM expense_framework_subject
            ORDER BY sort_order, budget_subject
            """
        )
        framework_rows = cur.fetchall()
        rows_to_insert = build_bootstrap_rows(
            [
                {
                    "level_label": text(_row_value(row, "level_label", 0)),
                    "subject_name": text(_row_value(row, "budget_subject", 1)),
                    "manage_department": text(_row_value(row, "manage_department", 2)) or None,
                    "formula_text": text(_row_value(row, "formula_text", 3)) or None,
                    "sort_order": int(_row_value(row, "sort_order", 4) or 0),
                }
                for row in framework_rows
            ]
        )

        parent_id_by_virtual_id: dict[int, int] = {}
        for item in rows_to_insert:
            cur = db.execute(
                """
                INSERT INTO budget_subject_catalog(
                  parent_id, level_number, subject_name, manage_department, formula_text, sort_order
                )
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (
                    parent_id_by_virtual_id.get(int(item["parent_virtual_id"]))
                    if item["parent_virtual_id"] is not None
                    else None,
                    int(item["level_number"]),
                    text(item["subject_name"]),
                    text(item.get("manage_department")) or None,
                    text(item["formula_text"]) or None,
                    int(item["sort_order"]),
                ),
            )
            parent_id_by_virtual_id[int(item["virtual_id"])] = int(cur.lastrowid)
        if rows_to_insert:
            db.commit()


def build_bootstrap_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    stack_by_level: dict[int, int] = {}
    next_virtual_id = 1
    result: list[dict[str, Any]] = []
    for item in rows:
        level_number = LEVEL_NAME_TO_NUMBER.get(text(item.get("level_label")))
        subject_name = text(item.get("subject_name"))
        if not level_number or not subject_name:
            continue
        parent_virtual_id = stack_by_level.get(level_number - 1) if level_number > 1 else None
        result.append(
            {
                "virtual_id": next_virtual_id,
                "parent_virtual_id": parent_virtual_id,
                "level_number": level_number,
                "subject_name": subject_name,
                "manage_department": text(item.get("manage_department")) or None,
                "formula_text": text(item.get("formula_text")) or None,
                "sort_order": int(item.get("sort_order") or len(result) + 1),
            }
        )
        stack_by_level[level_number] = next_virtual_id
        next_virtual_id += 1
        for key in list(stack_by_level.keys()):
            if key > level_number:
                stack_by_level.pop(key, None)
    return result


async def list_budget_subject_catalog(
    common_db: Path | str,
    *,
    ensure_schema: EnsureSchema = ensure_department_expense_master_schema,
) -> list[BudgetSubjectCatalogRow]:
    await ensure_budget_subject_catalog_bootstrapped(common_db, ensure_schema=ensure_schema)
    if _uses_mysql_path(common_db):
        rows = await get_pool().fetch_all(
            """
            SELECT c.id, c.parent_id, c.level_number, c.subject_name, c.manage_department, c.formula_text, c.sort_order,
                   EXISTS(SELECT 1 FROM budget_subject_catalog child WHERE child.parent_id = c.id) AS has_children
            FROM budget_subject_catalog c
            ORDER BY COALESCE(c.parent_id, 0), c.sort_order, c.id
            """
        )
    else:
        with sqlite3.connect(common_db) as db:
            db.execute("PRAGMA foreign_keys = ON")
            rows = db.execute(
                """
                SELECT c.id, c.parent_id, c.level_number, c.subject_name, c.manage_department, c.formula_text, c.sort_order,
                       EXISTS(SELECT 1 FROM budget_subject_catalog child WHERE child.parent_id = c.id) AS has_children
                FROM budget_subject_catalog c
                ORDER BY COALESCE(c.parent_id, 0), c.sort_order, c.id
                """
            ).fetchall()
    return [_row_from_db(row) for row in rows]


async def create_budget_subject_catalog(
    common_db: Path | str,
    body: BudgetSubjectCatalogCreate,
    *,
    ensure_schema: EnsureSchema = ensure_department_expense_master_schema,
) -> BudgetSubjectCatalogRow:
    await ensure_budget_subject_catalog_bootstrapped(common_db, ensure_schema=ensure_schema)
    parent_id = body.parent_id
    level_number = 1
    if _uses_mysql_path(common_db):
        if parent_id is not None:
            parent_row = await get_pool().fetch_one(
                "SELECT level_number FROM budget_subject_catalog WHERE id = %s",
                (parent_id,),
            )
            if not parent_row:
                raise HTTPException(status_code=404, detail="上级预算科目不存在")
            level_number = int(_row_value(parent_row, "level_number", 0) or 0) + 1
        if level_number < MIN_BUDGET_SUBJECT_LEVEL or level_number > MAX_BUDGET_SUBJECT_LEVEL:
            raise HTTPException(status_code=400, detail=f"预算科目最多支持 {MAX_BUDGET_SUBJECT_LEVEL} 层")
        if parent_id is None:
            max_sort_order = await get_pool().fetch_val(
                "SELECT COALESCE(MAX(sort_order), 0) FROM budget_subject_catalog WHERE parent_id IS NULL"
            )
        else:
            max_sort_order = await get_pool().fetch_val(
                "SELECT COALESCE(MAX(sort_order), 0) FROM budget_subject_catalog WHERE parent_id = %s",
                (parent_id,),
            )
        next_sort_order = int(max_sort_order or 0) + 1
        async with get_pool().acquire() as db:
            async with db.cursor() as cur:
                await cur.execute(
                    """
                    INSERT INTO budget_subject_catalog(parent_id, level_number, subject_name, manage_department, formula_text, sort_order)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    """,
                    (
                        parent_id,
                        level_number,
                        body.subject_name.strip(),
                        text(body.manage_department) or None,
                        text(body.formula_text) or None,
                        next_sort_order,
                    ),
                )
                new_id = int(cur.lastrowid)
        return await get_budget_subject_catalog_row(common_db, new_id, ensure_schema=ensure_schema)

    with sqlite3.connect(common_db) as db:
        db.execute("PRAGMA foreign_keys = ON")
        if parent_id is not None:
            parent_row = db.execute(
                "SELECT level_number FROM budget_subject_catalog WHERE id = ?",
                (parent_id,),
            ).fetchone()
            if not parent_row:
                raise HTTPException(status_code=404, detail="上级预算科目不存在")
            level_number = int(_row_value(parent_row, "level_number", 0) or 0) + 1
        if level_number < MIN_BUDGET_SUBJECT_LEVEL or level_number > MAX_BUDGET_SUBJECT_LEVEL:
            raise HTTPException(status_code=400, detail=f"预算科目最多支持 {MAX_BUDGET_SUBJECT_LEVEL} 层")
        if parent_id is None:
            max_sort_order = db.execute(
                "SELECT COALESCE(MAX(sort_order), 0) FROM budget_subject_catalog WHERE parent_id IS NULL"
            ).fetchone()[0]
        else:
            max_sort_order = db.execute(
                "SELECT COALESCE(MAX(sort_order), 0) FROM budget_subject_catalog WHERE parent_id = ?",
                (parent_id,),
            ).fetchone()[0]
        next_sort_order = int(max_sort_order or 0) + 1
        cur = db.execute(
            """
            INSERT INTO budget_subject_catalog(parent_id, level_number, subject_name, manage_department, formula_text, sort_order)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                parent_id,
                level_number,
                body.subject_name.strip(),
                text(body.manage_department) or None,
                text(body.formula_text) or None,
                next_sort_order,
            ),
        )
        new_id = int(cur.lastrowid)
        db.commit()
    return await get_budget_subject_catalog_row(common_db, new_id, ensure_schema=ensure_schema)


async def update_budget_subject_catalog(
    common_db: Path | str,
    row_id: int,
    body: BudgetSubjectCatalogUpdate,
    *,
    ensure_schema: EnsureSchema = ensure_department_expense_master_schema,
) -> BudgetSubjectCatalogRow:
    await ensure_budget_subject_catalog_bootstrapped(common_db, ensure_schema=ensure_schema)
    updates: list[str] = []
    values: list[Any] = []
    if body.subject_name is not None:
        updates.append("subject_name = ?")
        values.append(body.subject_name.strip())
    if body.manage_department is not None:
        updates.append("manage_department = ?")
        values.append(text(body.manage_department) or None)
    if body.formula_text is not None:
        updates.append("formula_text = ?")
        values.append(text(body.formula_text) or None)
    if not updates:
        raise HTTPException(status_code=400, detail="没有可更新的内容")

    if _uses_mysql_path(common_db):
        row = await get_pool().fetch_one("SELECT id FROM budget_subject_catalog WHERE id = %s", (row_id,))
        if not row:
            raise HTTPException(status_code=404, detail="预算科目不存在")
        mysql_updates = [part.replace("?", "%s") for part in updates]
        await get_pool().execute(
            f"UPDATE budget_subject_catalog SET {', '.join(mysql_updates)} WHERE id = %s",
            [*values, row_id],
        )
        return await get_budget_subject_catalog_row(common_db, row_id, ensure_schema=ensure_schema)

    with sqlite3.connect(common_db) as db:
        db.execute("PRAGMA foreign_keys = ON")
        row = db.execute("SELECT id FROM budget_subject_catalog WHERE id = ?", (row_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="预算科目不存在")
        values.append(row_id)
        db.execute(f"UPDATE budget_subject_catalog SET {', '.join(updates)} WHERE id = ?", values)
        db.commit()
    return await get_budget_subject_catalog_row(common_db, row_id, ensure_schema=ensure_schema)


async def delete_budget_subject_catalog(
    common_db: Path | str,
    row_id: int,
    *,
    ensure_schema: EnsureSchema = ensure_department_expense_master_schema,
) -> BudgetSubjectCatalogDeletion:
    await ensure_budget_subject_catalog_bootstrapped(common_db, ensure_schema=ensure_schema)
    if _uses_mysql_path(common_db):
        row = await get_pool().fetch_one(
            "SELECT subject_name FROM budget_subject_catalog WHERE id = %s",
            (row_id,),
        )
        if not row:
            raise HTTPException(status_code=404, detail="预算科目不存在")
        child_count = await get_pool().fetch_val(
            "SELECT COUNT(*) FROM budget_subject_catalog WHERE parent_id = %s",
            (row_id,),
        )
        if int(child_count or 0) > 0:
            raise HTTPException(status_code=409, detail="当前预算科目下仍有下级，无法删除")
        await get_pool().execute("DELETE FROM budget_subject_catalog WHERE id = %s", (row_id,))
        return BudgetSubjectCatalogDeletion(subject_name=text(_row_value(row, "subject_name", 0)))

    with sqlite3.connect(common_db) as db:
        db.execute("PRAGMA foreign_keys = ON")
        row = db.execute("SELECT subject_name FROM budget_subject_catalog WHERE id = ?", (row_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="预算科目不存在")
        child_count = db.execute(
            "SELECT COUNT(*) FROM budget_subject_catalog WHERE parent_id = ?",
            (row_id,),
        ).fetchone()[0]
        if int(child_count or 0) > 0:
            raise HTTPException(status_code=409, detail="当前预算科目下仍有下级，无法删除")
        db.execute("DELETE FROM budget_subject_catalog WHERE id = ?", (row_id,))
        db.commit()
    return BudgetSubjectCatalogDeletion(subject_name=text(_row_value(row, "subject_name", 0)))


async def get_budget_subject_catalog_row(
    common_db: Path | str,
    row_id: int,
    *,
    ensure_schema: EnsureSchema = ensure_department_expense_master_schema,
) -> BudgetSubjectCatalogRow:
    rows = await list_budget_subject_catalog(common_db, ensure_schema=ensure_schema)
    for row in rows:
        if row.id == row_id:
            return row
    raise HTTPException(status_code=500, detail="预算科目保存后未找到记录")


def build_budget_subject_catalog_workbook(rows: list[BudgetSubjectCatalogRow]) -> BudgetSubjectCatalogWorkbook:
    children_by_parent: dict[int | None, list[BudgetSubjectCatalogRow]] = {}
    for row in rows:
        children_by_parent.setdefault(row.parent_id, []).append(row)
    for value in children_by_parent.values():
        value.sort(key=lambda item: (item.sort_order, item.id))

    wb = Workbook()
    ws = wb.active
    ws.title = "部门预算科目"
    headers = ["层级", "预算科目", "归口管理部门", "公式"]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    export_rows: list[list[str | None]] = []

    def walk(parent_id: int | None) -> None:
        for row in children_by_parent.get(parent_id, []):
            export_rows.append([row.level_label, row.subject_name, row.manage_department, row.formula_text])
            walk(row.id)

    walk(None)
    for row_idx, row in enumerate(export_rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 30

    buffer = BytesIO()
    wb.save(buffer)
    return BudgetSubjectCatalogWorkbook(content=buffer.getvalue())


def _row_from_db(row: Any) -> BudgetSubjectCatalogRow:
    level_number = int(_row_value(row, "level_number", 2))
    return BudgetSubjectCatalogRow(
        id=int(_row_value(row, "id", 0)),
        parent_id=(
            int(_row_value(row, "parent_id", 1))
            if _row_value(row, "parent_id", 1) is not None
            else None
        ),
        level_number=level_number,
        level_label=LEVEL_NUMBER_TO_NAME.get(level_number, f"{level_number}级"),
        subject_name=text(_row_value(row, "subject_name", 3)),
        manage_department=text(_row_value(row, "manage_department", 4)) or None,
        formula_text=text(_row_value(row, "formula_text", 5)) or None,
        sort_order=int(_row_value(row, "sort_order", 6) or 0),
        is_leaf=not bool(_row_value(row, "has_children", 7)),
    )
