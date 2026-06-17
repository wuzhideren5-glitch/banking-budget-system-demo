"""Excel export builder for budget output display reports."""
from __future__ import annotations

import re
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Awaitable, Callable

import aiosqlite
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.formula_refs import OFFICIAL_RUNTIME_METRIC_REF_CODE, extract_formula_codes
from app.schemas import (
    BudgetOutputDisplayReportResponse,
    BudgetOutputReportNodeDto,
    BudgetOutputReportRowDto,
    BudgetOutputVersionDto,
    BudgetOutputVersionMetricDto,
)
from app.services.budget_output_display import build_budget_output_display_report
from app.services.runtime_metric_refs import load_org_product_metric_refs_by_runtime_ref_code

REPORT_GROUP_ROW = 4
REPORT_VERSION_ROW = 5
REPORT_HEADER_ROW = 6
REPORT_DATA_START_ROW = 7
TITLE_FILL = "17365D"
GROUP_FILL = "1F4E78"
VERSION_FILL = "5B9BD5"
HEADER_FILL = "D9EAF7"
BAND_FILL = "F3F6FA"
SUBTLE_FILL = "F8FAFC"


@dataclass(frozen=True)
class ReportColumnSpec:
    col: int
    label: str
    kind: str
    version_key: str | None = None
    month: int | None = None
    related_version_key: str | None = None
    product_code: str | None = None
    hidden_when_collapsed: bool = False


@dataclass(frozen=True)
class BudgetOutputWorkbookExport:
    workbook: Workbook
    filename: str


def _display_code_for_export(row_id: str) -> str:
    if ":" not in row_id:
        return row_id
    return row_id.split(":", 1)[1]


def _flatten_report_nodes(nodes: list[BudgetOutputReportNodeDto]) -> list[BudgetOutputReportNodeDto]:
    result: list[BudgetOutputReportNodeDto] = []
    for node in nodes:
        result.append(node)
        result.extend(_flatten_report_nodes(node.children))
    return result


def _formula_string(value: str) -> str:
    return '"' + value.replace('"', '""') + '"'


def _formula_ref_text(row: BudgetOutputReportRowDto) -> str:
    refs = extract_formula_codes(row.budget_formula) | extract_formula_codes(row.actual_formula)
    return "、".join(sorted(refs))


def _runtime_ref_code_candidates(row: BudgetOutputReportRowDto) -> list[str]:
    codes: list[str] = []
    if row.data_acct_code:
        codes.append(row.data_acct_code)
    if row.row_key.startswith("FORMULA_REF."):
        codes.append(row.row_key.removeprefix("FORMULA_REF."))
    return codes


def _org_product_refs_for_row(
    row: BudgetOutputReportRowDto | None,
    refs_by_runtime_ref_code: dict[str, list[str]],
) -> list[str]:
    if row is None:
        return []
    if row.org_product_ref:
        direct_ref = row.org_product_ref.strip()
        if row.org_product_metric_name and row.org_product_metric_name not in direct_ref:
            direct_ref = f"{direct_ref} {row.org_product_metric_name}"
        return [direct_ref]
    refs: list[str] = []
    for code in _runtime_ref_code_candidates(row):
        for ref in refs_by_runtime_ref_code.get(str(code or "").strip().upper(), []):
            if ref not in refs:
                refs.append(ref)
    return sorted(refs)


def _month_values(metric: BudgetOutputVersionMetricDto) -> list[float]:
    return (metric.monthly_values + [0.0] * 12)[:12]


def _version_metric(row: BudgetOutputReportRowDto | None, version_key: str | None) -> BudgetOutputVersionMetricDto:
    if row is None or not version_key:
        return BudgetOutputVersionMetricDto()
    return row.values_by_version.get(version_key, BudgetOutputVersionMetricDto())


def _metric_value(row: BudgetOutputReportRowDto | None, spec: ReportColumnSpec) -> float:
    metric = _version_metric(row, spec.version_key)
    if spec.kind == "month" and spec.month is not None:
        values = _month_values(metric)
        return float(values[spec.month - 1] if 1 <= spec.month <= 12 else 0)
    return float(metric.annual_value or 0)


def _is_percent_row(row: BudgetOutputReportRowDto) -> bool:
    text = f"{row.value_type or ''}{row.display_name}"
    return "百分比" in text or "率" in text or "占比" in text or "%" in text


def _numeric_cell_format(row: BudgetOutputReportRowDto | None, spec: ReportColumnSpec) -> str:
    if spec.kind in {"ratio", "yoy"}:
        return "0.00%;[Red]-0.00%;-"
    if row and _is_percent_row(row):
        return "0.00%;[Red]-0.00%;-"
    return '#,##0.00;[Red]-#,##0.00;-'


_DATA_ACCOUNT_REF_RE = re.compile(
    rf"<\s*(?P<bracket>{OFFICIAL_RUNTIME_METRIC_REF_CODE})(?:\s+[^>]*)?>"
    rf"|(?<![A-Z0-9.])(?P<bare>{OFFICIAL_RUNTIME_METRIC_REF_CODE})(?![A-Z0-9.])"
)


def _formula_for_version(row: BudgetOutputReportRowDto | None, version: BudgetOutputVersionDto) -> str | None:
    if row is None:
        return None
    if version.source == "show":
        return row.actual_formula or row.budget_formula
    return row.budget_formula or row.actual_formula


def _excel_note_formula(value: float, note: str) -> str:
    safe_note = note.replace('"', '""')[:240]
    return f"={value}+N(\"{safe_note}\")"


def _excel_runtime_ref_formula(
    formula: str | None,
    *,
    target_col: str,
    runtime_ref_row_numbers: dict[str, int],
    fallback_value: float,
) -> str | None:
    if not formula or not extract_formula_codes(formula):
        return None
    missing_codes: list[str] = []

    def replace_ref(match: re.Match[str]) -> str:
        code = match.group("bracket") or match.group("bare")
        row_number = runtime_ref_row_numbers.get(code)
        if row_number is not None:
            return f"{target_col}{row_number}"
        missing_codes.append(code)
        return "0"

    converted = _DATA_ACCOUNT_REF_RE.sub(replace_ref, formula)
    if missing_codes:
        return _excel_note_formula(
            fallback_value,
            f"底层公式含当前展示结构未展示的引用：{formula}",
        )
    return "=" + converted


def _active_versions(report: BudgetOutputDisplayReportResponse) -> list[BudgetOutputVersionDto]:
    return [
        version
        for version in report.versions
        if version.selected_by_default
    ]


def _top_header_label(version: BudgetOutputVersionDto) -> str:
    if version.source == "show":
        return "历史实际"
    return f"{str(version.year)[-2:]}年预算/预测"


def _split_versions(versions: list[BudgetOutputVersionDto]) -> tuple[list[BudgetOutputVersionDto], BudgetOutputVersionDto | None, list[BudgetOutputVersionDto]]:
    show_versions = [version for version in versions if version.source == "show"]
    budget_version = next((version for version in versions if version.source == "budget"), None)
    forecast_versions = [version for version in versions if version.source == "forecast"]
    return show_versions, budget_version, forecast_versions


def _build_detail_column_specs(
    versions: list[BudgetOutputVersionDto],
    *,
    start_col: int = 3,
) -> tuple[list[ReportColumnSpec], list[tuple[int, int, str]], list[tuple[int, int, str]], BudgetOutputVersionDto | None, BudgetOutputVersionDto | None]:
    show_versions, budget_version, forecast_versions = _split_versions(versions)
    last_actual_version = show_versions[-1] if show_versions else None
    specs: list[ReportColumnSpec] = []
    top_spans: list[tuple[int, int, str]] = []
    version_spans: list[tuple[int, int, str]] = []
    col = start_col

    def add_monthly_version(version: BudgetOutputVersionDto, *, annual_label: str, top_label: str) -> None:
        nonlocal col
        span_start = col
        specs.append(ReportColumnSpec(col=col, label=annual_label, kind="annual", version_key=version.key))
        col += 1
        for month in range(1, 13):
            specs.append(
                ReportColumnSpec(
                    col=col,
                    label=f"    {month}月",
                    kind="month",
                    version_key=version.key,
                    month=month,
                    hidden_when_collapsed=True,
                )
            )
            col += 1
        span_end = col - 1
        top_spans.append((span_start, span_end, top_label))
        version_spans.append((span_start, span_end, version.version_name))

    for version in show_versions:
        add_monthly_version(version, annual_label="实际", top_label="历史实际")

    if budget_version is not None:
        add_monthly_version(budget_version, annual_label="预算", top_label=f"{str(budget_version.year)[-2:]}年预算")

    for version in forecast_versions:
        span_start = col
        specs.append(ReportColumnSpec(col=col, label="预测", kind="annual", version_key=version.key))
        col += 1
        specs.append(
            ReportColumnSpec(
                col=col,
                label="预测-预算",
                kind="diff",
                version_key=version.key,
                related_version_key=budget_version.key if budget_version else None,
            )
        )
        col += 1
        specs.append(
            ReportColumnSpec(
                col=col,
                label="预算达成",
                kind="ratio",
                version_key=version.key,
                related_version_key=budget_version.key if budget_version else None,
            )
        )
        col += 1
        specs.append(
            ReportColumnSpec(
                col=col,
                label="同比",
                kind="yoy",
                version_key=version.key,
                related_version_key=last_actual_version.key if last_actual_version else None,
            )
        )
        col += 1
        for month in range(1, 13):
            specs.append(
                ReportColumnSpec(
                    col=col,
                    label=f"    {month}月",
                    kind="month",
                    version_key=version.key,
                    month=month,
                    hidden_when_collapsed=True,
                )
            )
            col += 1
        span_end = col - 1
        top_spans.append((span_start, span_end, f"{str(version.year)[-2:]}年预算/预测"))
        version_spans.append((span_start, span_end, version.version_name))

    return specs, top_spans, version_spans, budget_version, last_actual_version


def _build_product_overview_column_specs(
    versions: list[BudgetOutputVersionDto],
    product_blocks: list[Any],
    *,
    start_col: int = 3,
) -> tuple[list[ReportColumnSpec], list[tuple[int, int, str]]]:
    show_versions, _, forecast_versions = _split_versions(versions)
    last_actual_version = show_versions[-1] if show_versions else None
    specs: list[ReportColumnSpec] = []
    product_spans: list[tuple[int, int, str]] = []
    col = start_col
    for block in product_blocks:
        span_start = col
        for version in versions:
            specs.append(
                ReportColumnSpec(
                    col=col,
                    label=version.version_name,
                    kind="annual",
                    version_key=version.key,
                    product_code=block.product_code,
                )
            )
            col += 1
        first_forecast = forecast_versions[0] if forecast_versions else None
        if first_forecast and last_actual_version:
            specs.append(
                ReportColumnSpec(
                    col=col,
                    label="同比",
                    kind="yoy",
                    version_key=first_forecast.key,
                    related_version_key=last_actual_version.key,
                    product_code=block.product_code,
                )
            )
            col += 1
        product_spans.append((span_start, col - 1, f"{block.product_code} {block.product_name}"))
    return specs, product_spans


def _style_cell_font(cell: Any, *, name: str = "微软雅黑", size: float | None = None) -> None:
    font = copy(cell.font)
    font.name = name
    if size is not None:
        font.sz = size
    elif font.sz is None:
        font.sz = 10
    cell.font = font


def _write_report_title(ws: Any, *, title: str, subtitle: str, max_col: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    title_cell = ws.cell(1, 1, title)
    title_cell.fill = PatternFill("solid", fgColor=TITLE_FILL)
    title_cell.font = Font(name="微软雅黑", size=15, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    note_cell = ws.cell(2, 1, subtitle)
    note_cell.fill = PatternFill("solid", fgColor=SUBTLE_FILL)
    note_cell.font = Font(name="微软雅黑", size=9, color="44546A")
    note_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 34


def _merge_same_labels(ws: Any, spans: list[tuple[int, int, str]], *, row: int) -> None:
    start_col: int | None = None
    end_col: int | None = None
    current_label: str | None = None
    for span_start, span_end, label in spans:
        if current_label is None:
            start_col, end_col, current_label = span_start, span_end, label
            continue
        if label == current_label and end_col is not None and span_start == end_col + 1:
            end_col = span_end
            continue
        if start_col is not None and end_col is not None:
            ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
            ws.cell(row, start_col, current_label)
        start_col, end_col, current_label = span_start, span_end, label
    if start_col is not None and end_col is not None and current_label is not None:
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
        ws.cell(row, start_col, current_label)


def _style_header_block(ws: Any, *, max_col: int) -> None:
    fills = {
        REPORT_GROUP_ROW: PatternFill("solid", fgColor=GROUP_FILL),
        REPORT_VERSION_ROW: PatternFill("solid", fgColor=VERSION_FILL),
        REPORT_HEADER_ROW: PatternFill("solid", fgColor=HEADER_FILL),
    }
    colors = {
        REPORT_GROUP_ROW: "FFFFFF",
        REPORT_VERSION_ROW: "FFFFFF",
        REPORT_HEADER_ROW: "1F2937",
    }
    for row in (REPORT_GROUP_ROW, REPORT_VERSION_ROW, REPORT_HEADER_ROW):
        ws.row_dimensions[row].height = 22
        for col in range(1, max_col + 1):
            cell = ws.cell(row, col)
            cell.fill = fills[row]
            cell.font = Font(name="微软雅黑", size=9, bold=True, color=colors[row])
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(
                top=Side(style="thin", color="D9E2EC"),
                bottom=Side(style="thin", color="D9E2EC"),
                left=Side(style="thin", color="D9E2EC"),
                right=Side(style="thin", color="D9E2EC"),
            )


def _merge_fixed_columns_header(ws: Any) -> None:
    for col, value in ((1, "展示行编码"), (2, "预算展示科目")):
        ws.merge_cells(start_row=REPORT_GROUP_ROW, start_column=col, end_row=REPORT_HEADER_ROW, end_column=col)
        cell = ws.cell(REPORT_GROUP_ROW, col, value)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_budget_workbook(wb: Workbook) -> None:
    for ws in wb.worksheets:
        ws.freeze_panes = "C7"
        ws.sheet_view.showGridLines = False
        ws.sheet_view.zoomScale = 85
        for row in ws.iter_rows():
            for cell in row:
                _style_cell_font(cell)
                if cell.alignment:
                    cell.alignment = Alignment(
                        horizontal=cell.alignment.horizontal,
                        vertical="center",
                        text_rotation=cell.alignment.textRotation,
                        wrap_text=cell.alignment.wrap_text,
                        shrink_to_fit=cell.alignment.shrink_to_fit,
                        indent=cell.alignment.indent,
                    )
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 11
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 34
        ws.sheet_properties.outlinePr.summaryBelow = False
        ws.sheet_properties.outlinePr.summaryRight = False


def _write_metric_sheet(
    wb: Workbook,
    *,
    title: str,
    sheet_name: str,
    scope_code: str,
    tree: list[BudgetOutputReportNodeDto],
    rows: list[BudgetOutputReportRowDto],
    dependency_rows: list[BudgetOutputReportRowDto] | None = None,
    versions: list[BudgetOutputVersionDto],
    org_product_refs_by_runtime_ref_code: dict[str, list[str]] | None = None,
) -> None:
    ws = wb.create_sheet(sheet_name[:31])
    row_dto_by_key = {row.row_key: row for row in rows}
    if rows:
        nodes = [
            BudgetOutputReportNodeDto(
                row_key=row.row_key,
                display_name=row.display_name,
                parent_row_key=row.parent_row_key,
                level=row.level,
                is_summary=row.is_summary,
                is_minus=row.is_minus,
            )
            for row in rows
        ]
        node_by_code = {node.row_key: node for node in nodes}
        for node in nodes:
            if node.parent_row_key and node.parent_row_key in node_by_code:
                node_by_code[node.parent_row_key].children.append(node)
    else:
        nodes = _flatten_report_nodes(tree)
    dependency_rows = dependency_rows or []
    dependency_nodes = [
        BudgetOutputReportNodeDto(
            row_key=row.row_key,
            display_name=row.display_name,
            parent_row_key=None,
            level=8,
            is_summary=False,
            is_minus=row.is_minus,
        )
        for row in dependency_rows
    ]
    row_dto_by_key.update({row.row_key: row for row in dependency_rows})
    org_refs_by_data = org_product_refs_by_runtime_ref_code or {}
    version_by_key = {version.key: version for version in versions}
    column_specs, top_spans, version_spans, _, _ = _build_detail_column_specs(versions)
    value_max_col = max((spec.col for spec in column_specs), default=2)
    org_ref_count_col = value_max_col + 1
    org_ref_source_col = value_max_col + 2
    max_col = org_ref_source_col
    _write_report_title(
        ws,
        title=title,
        subtitle="单位：元；导出列与预算展示页面一致，分月列默认折叠，可展开检查月份明细和单元格公式。",
        max_col=max_col,
    )
    _merge_fixed_columns_header(ws)
    _merge_same_labels(ws, top_spans, row=REPORT_GROUP_ROW)
    for span_start, span_end, label in version_spans:
        ws.merge_cells(start_row=REPORT_VERSION_ROW, start_column=span_start, end_row=REPORT_VERSION_ROW, end_column=span_end)
        ws.cell(REPORT_VERSION_ROW, span_start, label)
    for spec in column_specs:
        ws.cell(REPORT_HEADER_ROW, spec.col, spec.label)
    ws.merge_cells(start_row=REPORT_GROUP_ROW, start_column=org_ref_count_col, end_row=REPORT_VERSION_ROW, end_column=org_ref_source_col)
    lineage_group_cell = ws.cell(REPORT_GROUP_ROW, org_ref_count_col, "机构产品指标追溯")
    lineage_group_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.cell(REPORT_HEADER_ROW, org_ref_count_col, "机构产品引用数量")
    ws.cell(REPORT_HEADER_ROW, org_ref_source_col, "机构产品来源")
    hidden_month_groups: list[tuple[int, int]] = []
    group_start: int | None = None
    last_col: int | None = None
    for spec in column_specs:
        if spec.hidden_when_collapsed:
            if group_start is None:
                group_start = spec.col
            last_col = spec.col
            continue
        if group_start is not None and last_col is not None:
            hidden_month_groups.append((group_start, last_col))
            group_start = None
            last_col = None
    if group_start is not None and last_col is not None:
        hidden_month_groups.append((group_start, last_col))
    for start_col, end_col in hidden_month_groups:
        ws.column_dimensions.group(
            get_column_letter(start_col),
            get_column_letter(end_col),
            outline_level=1,
            hidden=True,
        )
        for hidden_col in range(start_col, end_col + 1):
            dim = ws.column_dimensions[get_column_letter(hidden_col)]
            dim.hidden = True
            dim.outlineLevel = 1
        if start_col > 1:
            ws.column_dimensions[get_column_letter(start_col - 1)].collapsed = True

    row_by_code: dict[str, int] = {}
    row_no = REPORT_DATA_START_ROW
    for node in [*nodes, *dependency_nodes]:
        row_by_code[node.row_key] = row_no
        ws.cell(row_no, 1, _display_code_for_export(node.row_key))
        outline_level = max(0, min(7, node.level - 1))
        label_text = f"{'  ' * outline_level}{node.display_name}"
        if node in dependency_nodes:
            label_text = f"{'  ' * outline_level}公式引用底稿：{node.display_name}"
        label = ws.cell(row_no, 2, label_text)
        ws.row_dimensions[row_no].outlineLevel = outline_level
        ws.row_dimensions[row_no].height = 21
        if node in dependency_nodes:
            ws.row_dimensions[row_no].hidden = True
        label.alignment = Alignment(indent=max(0, outline_level // 2), vertical="center")
        if node.children:
            label.font = Font(name="微软雅黑", size=10, bold=True)
        row_no += 1

    runtime_ref_row_numbers: dict[str, int] = {}
    for row in [*rows, *dependency_rows]:
        if row.row_key not in row_by_code:
            continue
        for code in _runtime_ref_code_candidates(row):
            runtime_ref_row_numbers[code] = row_by_code[row.row_key]
    for node in [*nodes, *dependency_nodes]:
        excel_row = row_by_code[node.row_key]
        row_dto = row_dto_by_key.get(node.row_key)
        row_is_group = bool(row_dto and row_dto.row_type == "GROUP")
        for spec in column_specs:
            col_letter = get_column_letter(spec.col)
            cell = ws.cell(excel_row, spec.col)
            if spec.kind in {"diff", "ratio", "yoy"}:
                if row_is_group:
                    cell.value = None
                    continue
                base_spec = next(
                    (
                        candidate
                        for candidate in column_specs
                        if candidate.version_key == spec.version_key and candidate.kind == "annual"
                    ),
                    None,
                )
                base_col = get_column_letter(base_spec.col) if base_spec else None
                related_col: str | None = None
                if spec.related_version_key:
                    related_spec = next(
                        (
                            candidate
                            for candidate in column_specs
                            if candidate.version_key == spec.related_version_key and candidate.kind == "annual"
                        ),
                        None,
                    )
                    if related_spec:
                        related_col = get_column_letter(related_spec.col)
                if spec.kind == "diff" and base_col and related_col:
                    cell.value = f"={base_col}{excel_row}-{related_col}{excel_row}"
                elif spec.kind == "ratio" and base_col and related_col:
                    cell.value = f'=IFERROR({base_col}{excel_row}/{related_col}{excel_row},"-")'
                elif spec.kind == "yoy" and base_col and related_col:
                    cell.value = f'=IFERROR(({base_col}{excel_row}-{related_col}{excel_row})/{related_col}{excel_row},"-")'
                else:
                    cell.value = "-"
                cell.number_format = _numeric_cell_format(row_dto, spec)
                continue

            value = _metric_value(row_dto, spec)
            version = version_by_key.get(spec.version_key or "")
            metric_formula = _formula_for_version(row_dto, version) if version else None
            formula_value = _excel_runtime_ref_formula(
                metric_formula,
                target_col=col_letter,
                runtime_ref_row_numbers=runtime_ref_row_numbers,
                fallback_value=value,
            )
            if row_is_group:
                cell.value = None
            elif formula_value:
                cell.value = formula_value
            else:
                cell.value = formula_value or value
            cell.number_format = _numeric_cell_format(row_dto, spec)
        org_refs = _org_product_refs_for_row(row_dto, org_refs_by_data)
        ws.cell(excel_row, org_ref_count_col, len(org_refs))
        ws.cell(excel_row, org_ref_source_col, "\n".join(org_refs))
        ws.cell(excel_row, org_ref_count_col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(excel_row, org_ref_source_col).alignment = Alignment(vertical="center", wrap_text=True)

    _format_report_sheet(
        ws,
        max_row=row_no - 1,
        max_col=max_col,
        text_cols={org_ref_source_col},
        center_cols={org_ref_count_col},
    )


def _format_report_sheet(
    ws: Any,
    *,
    max_row: int,
    max_col: int,
    text_cols: set[int] | None = None,
    center_cols: set[int] | None = None,
) -> None:
    text_cols = text_cols or set()
    center_cols = center_cols or set()
    thin = Side(style="thin", color="D9E2EC")
    border = Border(bottom=thin)
    _style_header_block(ws, max_col=max_col)
    for row in range(REPORT_DATA_START_ROW, max_row + 1):
        outline_level = int(ws.row_dimensions[row].outlineLevel or 0)
        if outline_level <= 1:
            for col in range(1, max_col + 1):
                ws.cell(row, col).fill = PatternFill("solid", fgColor=BAND_FILL)
                ws.cell(row, col).font = Font(name="微软雅黑", size=10, bold=True)
        elif row % 2 == 0:
            for col in range(1, max_col + 1):
                ws.cell(row, col).fill = PatternFill("solid", fgColor="FBFCFE")
        for col in range(1, max_col + 1):
            cell = ws.cell(row, col)
            cell.border = border
            if col in text_cols:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                continue
            if col in center_cols:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                continue
            if col >= 3:
                if cell.number_format == "General":
                    cell.number_format = '#,##0.00;[Red]-#,##0.00;-'
                cell.alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row, 1).font = Font(name="微软雅黑", size=9, color="667085")
        ws.cell(row, 2).alignment = Alignment(
            indent=int(ws.row_dimensions[row].outlineLevel or 0),
            vertical="center",
        )
    ws.auto_filter.ref = f"A{REPORT_HEADER_ROW}:{get_column_letter(max_col)}{max_row}"
    ws.print_title_rows = f"$1:${REPORT_HEADER_ROW}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def _write_product_overview_sheet(
    wb: Workbook,
    report: BudgetOutputDisplayReportResponse,
    versions: list[BudgetOutputVersionDto],
    *,
    org_product_refs_by_runtime_ref_code: dict[str, list[str]] | None = None,
) -> None:
    ws = wb.create_sheet("预算分产品表")
    nodes = _flatten_report_nodes(report.product_overview_tree)
    product_blocks = report.product_overview_blocks
    dependency_rows_by_code: dict[str, BudgetOutputReportRowDto] = {}
    for block in product_blocks:
        for row in block.formula_dependency_rows:
            if row.data_acct_code and row.data_acct_code not in dependency_rows_by_code:
                dependency_rows_by_code[row.data_acct_code] = row
    dependency_rows = list(dependency_rows_by_code.values())
    dependency_nodes = [
        BudgetOutputReportNodeDto(
            row_key=row.row_key,
            display_name=row.display_name,
            parent_row_key=None,
            level=8,
            is_summary=False,
            is_minus=row.is_minus,
        )
        for row in dependency_rows
    ]
    version_by_key = {version.key: version for version in versions}
    column_specs, product_spans = _build_product_overview_column_specs(versions, product_blocks)
    value_max_col = max((spec.col for spec in column_specs), default=2)
    org_ref_count_col = value_max_col + 1
    org_ref_source_col = value_max_col + 2
    max_col = org_ref_source_col
    _write_report_title(
        ws,
        title=f"{report.selected_year}年度预算分产品表",
        subtitle="单位：元；导出列与分产品概览页面一致，横向为产品，纵向为预算展示结构。",
        max_col=max_col,
    )
    _merge_fixed_columns_header(ws)
    for span_start, span_end, label in product_spans:
        ws.merge_cells(start_row=REPORT_GROUP_ROW, start_column=span_start, end_row=REPORT_GROUP_ROW, end_column=span_end)
        ws.cell(REPORT_GROUP_ROW, span_start, label)
    for spec in column_specs:
        ws.cell(REPORT_VERSION_ROW, spec.col, spec.label if spec.kind == "annual" else "")
        ws.cell(REPORT_HEADER_ROW, spec.col, "全年" if spec.kind == "annual" else spec.label)
    ws.merge_cells(start_row=REPORT_GROUP_ROW, start_column=org_ref_count_col, end_row=REPORT_VERSION_ROW, end_column=org_ref_source_col)
    lineage_group_cell = ws.cell(REPORT_GROUP_ROW, org_ref_count_col, "机构产品指标追溯")
    lineage_group_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.cell(REPORT_HEADER_ROW, org_ref_count_col, "机构产品引用数量")
    ws.cell(REPORT_HEADER_ROW, org_ref_source_col, "机构产品来源")

    row_by_code: dict[str, int] = {}
    row_no = REPORT_DATA_START_ROW
    for node in [*nodes, *dependency_nodes]:
        row_by_code[node.row_key] = row_no
        ws.cell(row_no, 1, _display_code_for_export(node.row_key))
        outline_level = max(0, min(7, node.level - 1))
        label_text = f"{'  ' * outline_level}{node.display_name}"
        if node in dependency_nodes:
            label_text = f"{'  ' * outline_level}公式引用底稿：{node.display_name}"
        label = ws.cell(row_no, 2, label_text)
        ws.row_dimensions[row_no].outlineLevel = outline_level
        ws.row_dimensions[row_no].height = 21
        if node in dependency_nodes:
            ws.row_dimensions[row_no].hidden = True
        label.alignment = Alignment(indent=max(0, outline_level // 2), vertical="center")
        if node.children:
            label.font = Font(name="微软雅黑", size=10, bold=True)
        row_no += 1

    for node in [*nodes, *dependency_nodes]:
        excel_row = row_by_code[node.row_key]
        refs_for_node: list[str] = []
        for block in product_blocks:
            block_rows_by_key = {row.row_key: row for row in [*block.rows, *block.formula_dependency_rows]}
            runtime_ref_row_numbers: dict[str, int] = {}
            for row in [*block.rows, *block.formula_dependency_rows]:
                if row.row_key not in row_by_code:
                    continue
                for code in _runtime_ref_code_candidates(row):
                    runtime_ref_row_numbers[code] = row_by_code[row.row_key]
            block_specs = [spec for spec in column_specs if spec.product_code == block.product_code]
            for spec in block_specs:
                col_letter = get_column_letter(spec.col)
                row_dto = block_rows_by_key.get(node.row_key)
                cell = ws.cell(excel_row, spec.col)
                if row_dto and row_dto.row_type == "GROUP":
                    cell.value = None
                    continue
                if spec.kind == "yoy":
                    forecast_spec = next(
                        (
                            candidate
                            for candidate in block_specs
                            if candidate.version_key == spec.version_key and candidate.kind == "annual"
                        ),
                        None,
                    )
                    actual_spec = next(
                        (
                            candidate
                            for candidate in block_specs
                            if candidate.version_key == spec.related_version_key and candidate.kind == "annual"
                        ),
                        None,
                    )
                    if forecast_spec and actual_spec:
                        forecast_col = get_column_letter(forecast_spec.col)
                        actual_col = get_column_letter(actual_spec.col)
                        cell.value = f'=IFERROR(({forecast_col}{excel_row}-{actual_col}{excel_row})/{actual_col}{excel_row},"-")'
                    else:
                        cell.value = "-"
                    cell.number_format = _numeric_cell_format(row_dto, spec)
                else:
                    value = _metric_value(row_dto, spec)
                    version = version_by_key.get(spec.version_key or "")
                    metric_formula = _formula_for_version(row_dto, version) if version else None
                    formula_value = _excel_runtime_ref_formula(
                        metric_formula,
                        target_col=col_letter,
                        runtime_ref_row_numbers=runtime_ref_row_numbers,
                        fallback_value=value,
                    )
                    if formula_value:
                        cell.value = formula_value
                    else:
                        cell.value = value
                    cell.number_format = _numeric_cell_format(row_dto, spec)
                if row_dto:
                    for ref in _org_product_refs_for_row(row_dto, org_product_refs_by_runtime_ref_code or {}):
                        if ref not in refs_for_node:
                            refs_for_node.append(ref)
        refs_for_node = sorted(refs_for_node)
        ws.cell(excel_row, org_ref_count_col, len(refs_for_node))
        ws.cell(excel_row, org_ref_source_col, "\n".join(refs_for_node))
        ws.cell(excel_row, org_ref_count_col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(excel_row, org_ref_source_col).alignment = Alignment(vertical="center", wrap_text=True)
    _format_report_sheet(
        ws,
        max_row=row_no - 1,
        max_col=max_col,
        text_cols={org_ref_source_col},
        center_cols={org_ref_count_col},
    )


def build_budget_output_formula_workbook(
    report: BudgetOutputDisplayReportResponse,
    *,
    org_product_refs_by_runtime_ref_code: dict[str, list[str]] | None = None,
) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    versions = _active_versions(report)
    _write_metric_sheet(
        wb,
        title=f"{report.selected_year}年度预算全行总表",
        sheet_name="预算全行总表",
        scope_code="TOTAL",
        tree=report.report_tree,
        rows=report.total_rows,
        dependency_rows=report.total_formula_dependency_rows,
        versions=versions,
        org_product_refs_by_runtime_ref_code=org_product_refs_by_runtime_ref_code,
    )
    _write_product_overview_sheet(
        wb,
        report,
        versions,
        org_product_refs_by_runtime_ref_code=org_product_refs_by_runtime_ref_code,
    )
    for block in report.product_detail_blocks:
        _write_metric_sheet(
            wb,
            title=f"{block.product_name}（{block.product_code}）预算明细表",
            sheet_name=f"{block.product_code}_{block.product_name}"[:31],
            scope_code=f"DETAIL:{block.product_code}",
            tree=report.product_detail_tree,
            rows=block.rows,
            dependency_rows=block.formula_dependency_rows,
            versions=versions,
            org_product_refs_by_runtime_ref_code=org_product_refs_by_runtime_ref_code,
        )
    _style_budget_workbook(wb)
    return wb


async def build_budget_output_display_report_export(
    *,
    year: int | None,
    budget_version_id: int | None,
    forecast_version_ids: list[int] | None,
    editable_context_provider: Callable[[], Awaitable[tuple[Path, int, int]]],
    data_dir: Path,
) -> BudgetOutputWorkbookExport:
    org_product_refs: dict[str, list[str]] = {}
    common_path = data_dir / "common.db"
    if common_path.exists():
        async with aiosqlite.connect(common_path) as common_db:
            await common_db.execute("PRAGMA foreign_keys = ON")
            org_product_refs = await load_org_product_metric_refs_by_runtime_ref_code(common_db)
    report = await build_budget_output_display_report(
        year=year,
        budget_version_id=budget_version_id,
        forecast_version_ids=forecast_version_ids,
        product_codes=None,
        editable_context_provider=editable_context_provider,
        data_dir=data_dir,
    )
    return BudgetOutputWorkbookExport(
        workbook=build_budget_output_formula_workbook(
            report,
            org_product_refs_by_runtime_ref_code=org_product_refs,
        ),
        filename=f"{report.selected_year}年度预算展示全套报表.xlsx",
    )
