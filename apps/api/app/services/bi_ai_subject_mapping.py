from __future__ import annotations

from contextlib import asynccontextmanager
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
import re
import sqlite3
import tempfile
from typing import Any

from openpyxl import load_workbook

from app.core.config import settings
from app.core.database import get_pool
from app.db_bootstrap.expense import ensure_bi_ai_subject_mapping_schema
from app.services.bi_ai_manage_department import (
    build_caliber_to_catalog_subject_map,
    build_effective_manage_department_by_subject,
    format_manage_departments_display,
    load_all_expense_departments,
    load_budget_subject_catalog_manage_rows,
    normalize_manage_department_subject_label,
    parse_manage_department_override,
    resolve_effective_manage_departments_for_bi_mapping_row,
    serialize_manage_department_override,
)


EXPECTED_HEADERS = [
    "二级名称",
    "三级编码",
    "三级名称",
    "四级编码",
    "四级名称",
    "五级编码",
    "五级名称",
    "六级编码",
    "六级名称",
    "预算发布口径（二级）",
    "费用类别（一级）",
    "费用大类",
]


class BiAiSubjectMappingSourceMissingError(FileNotFoundError):
    pass


class BiAiSubjectMappingHeaderError(ValueError):
    pass


class BiAiSubjectMappingUpdateError(ValueError):
    pass


class BiAiSubjectMappingNotFoundError(LookupError):
    pass


SKIPPED_LEVEL3_CODES = {"YS0104", "YS0105"}


@dataclass(frozen=True)
class BiAiSubjectMappingReloadResult:
    row_count: int
    source_file: str


def _uses_mysql_common_path(path: Path | str) -> bool:
    try:
        candidate = Path(path).expanduser().resolve()
        data_dir = Path(settings.data_dir).expanduser().resolve()
        temp_root = Path(tempfile.gettempdir()).expanduser().resolve()
    except (TypeError, OSError):
        return False
    try:
        candidate.relative_to(temp_root)
        return False
    except ValueError:
        pass
    try:
        candidate.relative_to(data_dir)
    except ValueError:
        return False
    return candidate.name == "common.db"


def _mysql_sql(sql: str) -> str:
    return sql.replace("?", "%s")


class _CursorAdapter:
    def __init__(self, rows: list[Any] | None = None, *, rowcount: int = 0, lastrowid: int | None = None):
        self._rows = rows or []
        self.rowcount = rowcount
        self.lastrowid = lastrowid

    async def fetchone(self) -> Any | None:
        return self._rows[0] if self._rows else None

    async def fetchall(self) -> list[Any]:
        return list(self._rows)


class _SQLiteConnection:
    def __init__(self, path: Path):
        self._path = path
        self._conn: sqlite3.Connection | None = None

    async def __aenter__(self) -> "_SQLiteConnection":
        self._conn = sqlite3.connect(self._path)
        self._conn.execute("PRAGMA foreign_keys = ON")
        return self

    async def __aexit__(self, exc_type, exc, tb) -> None:
        if self._conn is not None:
            if exc_type is not None:
                self._conn.rollback()
            self._conn.close()
            self._conn = None

    async def execute(self, sql: str, params: tuple[Any, ...] | list[Any] = ()) -> _CursorAdapter:
        assert self._conn is not None
        cur = self._conn.execute(sql, tuple(params))
        return _CursorAdapter(
            cur.fetchall() if cur.description else [],
            rowcount=max(0, int(cur.rowcount or 0)),
            lastrowid=cur.lastrowid,
        )

    async def commit(self) -> None:
        assert self._conn is not None
        self._conn.commit()


class _MySQLConnection:
    def __init__(self):
        self._ctx: Any = None
        self._conn: Any = None

    async def __aenter__(self) -> "_MySQLConnection":
        self._ctx = get_pool().acquire()
        self._conn = await self._ctx.__aenter__()
        return self

    async def __aexit__(self, exc_type, exc, tb) -> None:
        if exc_type is not None and self._conn is not None:
            rollback = getattr(self._conn, "rollback", None)
            if rollback is not None:
                await rollback()
        if self._ctx is not None:
            await self._ctx.__aexit__(exc_type, exc, tb)
            self._ctx = None
            self._conn = None

    async def execute(self, sql: str, params: tuple[Any, ...] | list[Any] = ()) -> _CursorAdapter:
        assert self._conn is not None
        stripped = " ".join(sql.strip().split())
        if stripped.lower().startswith("pragma foreign_keys"):
            return _CursorAdapter([(1,)])
        pragma_match = re.fullmatch(r'PRAGMA table_info\("([^"]+)"\)', stripped, flags=re.IGNORECASE)
        if pragma_match:
            table_name = pragma_match.group(1)
            async with self._conn.cursor() as cur:
                await cur.execute(f"SHOW COLUMNS FROM `{table_name}`")
                rows = await cur.fetchall()
            return _CursorAdapter([(idx, row[0]) for idx, row in enumerate(rows)])

        sqlite_master_match = re.search(
            r"FROM\s+sqlite_master\s+WHERE\s+type\s*=\s*'table'\s+AND\s+name\s*=\s*'([^']+)'",
            stripped,
            flags=re.IGNORECASE,
        )
        if sqlite_master_match:
            table_name = sqlite_master_match.group(1)
            async with self._conn.cursor() as cur:
                await cur.execute("SHOW TABLES LIKE %s", (table_name,))
                rows = await cur.fetchall()
            return _CursorAdapter([(1,)] if rows else [])

        async with self._conn.cursor() as cur:
            await cur.execute(_mysql_sql(sql), tuple(params))
            rows = await cur.fetchall() if cur.description else []
            return _CursorAdapter(
                list(rows),
                rowcount=max(0, int(cur.rowcount or 0)),
                lastrowid=getattr(cur, "lastrowid", None),
            )

    async def commit(self) -> None:
        if self._conn is not None:
            await self._conn.commit()


@asynccontextmanager
async def _connect_db(db_path: Path):
    if _uses_mysql_common_path(db_path):
        async with _MySQLConnection() as db:
            yield db
    else:
        async with _SQLiteConnection(db_path) as db:
            yield db


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _iso_now() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def candidate_source_workbooks(repo_root: Path) -> list[Path]:
    names = ("BI科目匹配表.xlsx", "BI科目mapping.xlsx")
    roots = (
        repo_root,
        repo_root / "resources" / "business_inputs",
        repo_root / "resources" / "download_template",
    )
    return [root / name for root in roots for name in names]


def source_workbook_path(repo_root: Path) -> Path | None:
    for candidate in candidate_source_workbooks(repo_root):
        if candidate.exists():
            return candidate
    return None


def read_workbook_rows(path: Path) -> list[dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [_text(cell.value) for cell in ws[2]]
    if headers[: len(EXPECTED_HEADERS)] != EXPECTED_HEADERS:
        raise BiAiSubjectMappingHeaderError("BI科目mapping源文件表头不符合预期")

    rows: list[dict[str, str]] = []
    for excel_row in ws.iter_rows(min_row=3, values_only=True):
        values = [_text(value) for value in excel_row[: len(EXPECTED_HEADERS)]]
        if not any(values):
            continue
        if values[1] in SKIPPED_LEVEL3_CODES:
            continue
        rows.append(
            {
                "level5_code": values[5],
                "level5_name": values[6],
                "level6_code": values[7],
                "level6_name": values[8],
                "budget_release_caliber": values[9],
                "fee_category": values[10],
                "fee_major": values[11],
            }
        )
    return rows


def _attach_manage_departments(
    rows: list[dict[str, Any]],
    manage_by_subject: dict[str, str],
    all_expense_departments: list[str],
    *,
    catalog_names: set[str],
) -> list[dict[str, Any]]:
    caliber_to_catalog = build_caliber_to_catalog_subject_map(rows, catalog_names)
    enriched: list[dict[str, Any]] = []
    for row in rows:
        item = dict(row)
        departments, source = resolve_effective_manage_departments_for_bi_mapping_row(
            item,
            manage_by_subject,
            all_expense_departments,
            catalog_names=catalog_names,
            caliber_to_catalog=caliber_to_catalog,
        )
        override = parse_manage_department_override(item.get("manage_department_override"))
        item["manage_departments"] = departments
        item["manage_department"] = format_manage_departments_display(
            departments,
            default_all=source == "default_all",
        )
        item["manage_department_source"] = source
        item["manage_department_override"] = override
        item["manage_department_is_default_all"] = source == "default_all"
        enriched.append(item)
    return enriched


async def ensure_bi_ai_subject_mapping_table(db_path: Path) -> None:
    if _uses_mysql_common_path(db_path):
        return
    async with _connect_db(db_path) as db:
        await ensure_bi_ai_subject_mapping_schema(db)
        await db.commit()


async def ensure_bi_ai_subject_mapping_seeded(
    db_path: Path,
    repo_root: Path,
    *,
    force_reload: bool = False,
) -> BiAiSubjectMappingReloadResult:
    await ensure_bi_ai_subject_mapping_table(db_path)
    async with _connect_db(db_path) as db:
        cur = await db.execute("SELECT COUNT(*) FROM bi_ai_subject_mapping")
        existing_count = int((await cur.fetchone())[0] or 0)
    if existing_count > 0 and not force_reload:
        return BiAiSubjectMappingReloadResult(row_count=existing_count, source_file="")

    source_path = source_workbook_path(repo_root)
    if source_path is None:
        if force_reload:
            raise BiAiSubjectMappingSourceMissingError("BI科目mapping源文件不存在")
        return BiAiSubjectMappingReloadResult(row_count=existing_count, source_file="")

    rows = read_workbook_rows(source_path)
    now = _iso_now()
    async with _connect_db(db_path) as db:
        await db.execute("PRAGMA foreign_keys = ON")
        await db.execute("DELETE FROM bi_ai_subject_mapping")
        for index, row in enumerate(rows, start=1):
            await db.execute(
                """
                INSERT INTO bi_ai_subject_mapping(
                  level5_code, level5_name, level6_code, level6_name,
                  budget_release_caliber, fee_category, fee_major,
                  sort_order, source_file, created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    row["level5_code"],
                    row["level5_name"],
                    row["level6_code"],
                    row["level6_name"],
                    row["budget_release_caliber"],
                    row["fee_category"],
                    row["fee_major"],
                    index,
                    source_path.name,
                    now,
                    now,
                ),
            )
        await db.commit()
    return BiAiSubjectMappingReloadResult(row_count=len(rows), source_file=source_path.name)


async def query_bi_ai_subject_mapping_rows(
    db_path: Path,
    repo_root: Path,
) -> list[dict[str, Any]]:
    async with _connect_db(db_path) as db:
        cur = await db.execute(
            """
            SELECT id, level5_code, level5_name, level6_code, level6_name,
                   budget_release_caliber, fee_category, fee_major,
                   manage_department_override, sort_order, source_file
            FROM bi_ai_subject_mapping
            ORDER BY sort_order, id
            """
        )
        rows = await cur.fetchall()
        catalog_rows = []
        catalog_cur = await db.execute(
            "SELECT 1 FROM sqlite_master WHERE type = 'table' AND name = 'budget_subject_catalog'"
        )
        if await catalog_cur.fetchone():
            catalog_rows = await load_budget_subject_catalog_manage_rows(db)
        all_expense_departments = await load_all_expense_departments(db)

    manage_by_subject = build_effective_manage_department_by_subject(catalog_rows)
    catalog_names = {
        normalize_manage_department_subject_label(_text(row.get("subject_name")))
        for row in catalog_rows
    }
    catalog_names.discard("")
    base_rows: list[dict[str, Any]] = [
        {
            "id": int(row[0]),
            "level5_code": _text(row[1]),
            "level5_name": _text(row[2]),
            "level6_code": _text(row[3]),
            "level6_name": _text(row[4]),
            "budget_release_caliber": _text(row[5]),
            "fee_category": _text(row[6]),
            "fee_major": _text(row[7]),
            "manage_department_override": _text(row[8]),
            "sort_order": int(row[9] or 0),
            "source_file": _text(row[10]),
        }
        for row in rows
    ]
    return _attach_manage_departments(
        base_rows,
        manage_by_subject,
        all_expense_departments,
        catalog_names=catalog_names,
    )


async def get_bi_ai_subject_mapping_reference_data(db_path: Path) -> dict[str, Any]:
    await ensure_bi_ai_subject_mapping_table(db_path)
    async with _connect_db(db_path) as db:
        expense_departments = await load_all_expense_departments(db)
    return {"expense_departments": expense_departments}


async def update_bi_ai_subject_mapping_manage_departments(
    mapping_id: int,
    manage_departments: list[str] | None,
    *,
    db_path: Path,
    repo_root: Path,
) -> dict[str, Any]:
    await ensure_bi_ai_subject_mapping_table(db_path)
    async with _connect_db(db_path) as db:
        expense_departments = await load_all_expense_departments(db)
        allowed = set(expense_departments)
        if manage_departments is not None:
            invalid = [_text(item) for item in manage_departments if _text(item) not in allowed]
            if invalid:
                raise BiAiSubjectMappingUpdateError(
                    f"归口部门不在部门科目维护范围内：{', '.join(invalid[:5])}"
                )
        override_value = serialize_manage_department_override(manage_departments)
        cur = await db.execute(
            """
            UPDATE bi_ai_subject_mapping
            SET manage_department_override = ?, updated_at = ?
            WHERE id = ?
            """,
            (override_value, _iso_now(), mapping_id),
        )
        await db.commit()
        if int(cur.rowcount or 0) == 0:
            raise BiAiSubjectMappingNotFoundError(f"BI-AI映射记录不存在：{mapping_id}")

    rows = await query_bi_ai_subject_mapping_rows(db_path, repo_root)
    for row in rows:
        if int(row["id"]) == mapping_id:
            return row
    raise BiAiSubjectMappingNotFoundError(f"BI-AI映射记录不存在：{mapping_id}")


async def create_bi_ai_subject_mapping_row(
    payload: dict[str, Any],
    *,
    db_path: Path,
    repo_root: Path,
) -> dict[str, Any]:
    await ensure_bi_ai_subject_mapping_table(db_path)
    level5_code = _text(payload.get("level5_code"))
    level5_name = _text(payload.get("level5_name"))
    level6_code = _text(payload.get("level6_code"))
    level6_name = _text(payload.get("level6_name"))
    budget_release_caliber = _text(payload.get("budget_release_caliber"))
    fee_category = _text(payload.get("fee_category"))
    fee_major = _text(payload.get("fee_major"))
    manage_departments = payload.get("manage_departments")
    if not level5_code or not level5_name or not level6_code or not level6_name:
        raise BiAiSubjectMappingUpdateError("五级编码、五级名称、六级编码、六级名称不能为空")
    if not budget_release_caliber:
        raise BiAiSubjectMappingUpdateError("预算发布口径（二级）不能为空")
    if manage_departments is not None and not isinstance(manage_departments, list):
        raise BiAiSubjectMappingUpdateError("归口部门必须是字符串数组或空值")

    now = _iso_now()
    async with _connect_db(db_path) as db:
        expense_departments = await load_all_expense_departments(db)
        allowed = set(expense_departments)
        if manage_departments is not None:
            invalid = [_text(item) for item in manage_departments if _text(item) not in allowed]
            if invalid:
                raise BiAiSubjectMappingUpdateError(
                    f"归口部门不在部门科目维护范围内：{', '.join(invalid[:5])}"
                )
        override_value = serialize_manage_department_override(manage_departments)
        cur = await db.execute("SELECT COALESCE(MAX(sort_order), 0) FROM bi_ai_subject_mapping")
        sort_order = int((await cur.fetchone())[0] or 0) + 1
        cur = await db.execute(
            """
            INSERT INTO bi_ai_subject_mapping(
              level5_code, level5_name, level6_code, level6_name,
              budget_release_caliber, fee_category, fee_major,
              manage_department_override, sort_order, source_file, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                level5_code,
                level5_name,
                level6_code,
                level6_name,
                budget_release_caliber,
                fee_category,
                fee_major,
                override_value,
                sort_order,
                "manual",
                now,
                now,
            ),
        )
        await db.commit()
        row_id = int(cur.lastrowid)

    rows = await query_bi_ai_subject_mapping_rows(db_path, repo_root)
    for row in rows:
        if int(row["id"]) == row_id:
            return row
    raise BiAiSubjectMappingNotFoundError(f"BI-AI映射记录不存在：{row_id}")


async def list_bi_ai_subject_mapping_rows(
    db_path: Path,
    repo_root: Path,
) -> list[dict[str, Any]]:
    await ensure_bi_ai_subject_mapping_seeded(db_path, repo_root)
    return await query_bi_ai_subject_mapping_rows(db_path, repo_root)
