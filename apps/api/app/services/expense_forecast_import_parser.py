"""Excel parsers for expense forecast import workbooks."""
from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import load_workbook


class ExpenseForecastImportParseError(ValueError):
    """Raised when an import workbook does not match the current import contract."""


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _field_label(field_name: str, month: int | None = None) -> str:
    if field_name == "month_forecast":
        return f"M{int(month or 0)}"
    if field_name == "business_submission":
        return "业务报送"
    return "资划建议"


def _scan_import_header(ws: Any) -> tuple[int, dict[str, int]]:
    header_row = None
    header_map: dict[str, int] = {}
    scan_limit = min(ws.max_row, 10)
    for row_idx in range(1, scan_limit + 1):
        candidate_header_map: dict[str, int] = {}
        for col in range(1, ws.max_column + 1):
            key = _text(ws.cell(row_idx, col).value)
            if key:
                candidate_header_map[key] = col
        if "预算科目" in candidate_header_map:
            header_row = row_idx
            header_map = candidate_header_map
            break
    subject_col = header_map.get("预算科目")
    if subject_col is None:
        raise ExpenseForecastImportParseError("导入模板缺少“预算科目”列")
    return int(header_row or 1), header_map


def _scan_subject_import_header(ws: Any) -> tuple[int, dict[str, int]]:
    header_row = None
    header_map: dict[str, int] = {}
    scan_limit = min(ws.max_row, 10)
    for row_idx in range(1, scan_limit + 1):
        candidate_header_map: dict[str, int] = {}
        for col in range(1, ws.max_column + 1):
            key = _text(ws.cell(row_idx, col).value)
            if key:
                candidate_header_map[key] = col
        if "费用归属部门" in candidate_header_map:
            header_row = row_idx
            header_map = candidate_header_map
            break
    owner_col = header_map.get("费用归属部门")
    if owner_col is None:
        raise ExpenseForecastImportParseError("导入模板缺少“费用归属部门”列")
    return int(header_row or 1), header_map


def _parse_import_number(raw_value: Any) -> float:
    if raw_value is None or raw_value == "":
        raise ValueError("empty")
    if isinstance(raw_value, str):
        text = raw_value.strip()
        if text in {"-", "—", "－"}:
            raise ValueError("empty")
        text = text.replace(",", "").replace("，", "")
        if text == "":
            raise ValueError("empty")
        return float(text)
    return float(raw_value)


def _month_columns(header_map: dict[str, int]) -> dict[int, int]:
    month_cols: dict[int, int] = {}
    for key, col in header_map.items():
        normalized = key.upper().replace("月", "").replace("M", "").strip()
        if normalized.isdigit():
            month = int(normalized)
            if 1 <= month <= 12:
                month_cols[month] = col
    return month_cols


def _annual_columns(header_map: dict[str, int]) -> dict[str, int]:
    annual_cols: dict[str, int] = {}
    if header_map.get("业务报送") is not None:
        annual_cols["business_submission"] = int(header_map["业务报送"])
    if header_map.get("资划建议") is not None:
        annual_cols["capital_advice"] = int(header_map["资划建议"])
    return annual_cols


def _ensure_import_value_columns(month_cols: dict[int, int], annual_cols: dict[str, int]) -> None:
    if not month_cols and not annual_cols:
        raise ExpenseForecastImportParseError(
            "导入模板缺少可识别字段，请提供 M1~M12 / 1月~12月 / 业务报送 / 资划建议 列"
        )


def _parse_import_rows(
    ws: Any,
    *,
    header_row: int,
    header_map: dict[str, int],
    owner_name_by_row: dict[int, str] | None = None,
) -> list[dict[str, Any]]:
    subject_col = int(header_map["预算科目"])
    month_cols = _month_columns(header_map)
    annual_cols = _annual_columns(header_map)
    _ensure_import_value_columns(month_cols, annual_cols)

    rows: list[dict[str, Any]] = []
    data_start_row = int(header_row) + 1
    for row_idx in range(data_start_row, ws.max_row + 1):
        budget_subject = _text(ws.cell(row_idx, subject_col).value)
        if not budget_subject:
            continue
        owner_name = _text((owner_name_by_row or {}).get(row_idx))
        for month, col in month_cols.items():
            raw_value = ws.cell(row_idx, col).value
            if raw_value in (None, ""):
                continue
            try:
                value = _parse_import_number(raw_value)
            except (TypeError, ValueError):
                rows.append(
                    {
                        "row_number": row_idx,
                        "owner_name": owner_name,
                        "budget_subject": budget_subject,
                        "field_name": "month_forecast",
                        "field_label": _field_label("month_forecast", month),
                        "month": month,
                        "value": 0.0,
                        "error": f"月份 M{month} 不是有效数字",
                    }
                )
                continue
            rows.append(
                {
                    "row_number": row_idx,
                    "owner_name": owner_name,
                    "budget_subject": budget_subject,
                    "field_name": "month_forecast",
                    "field_label": _field_label("month_forecast", month),
                    "month": month,
                    "value": value,
                    "error": None,
                }
            )
        for field_name, col in annual_cols.items():
            raw_value = ws.cell(row_idx, col).value
            if raw_value in (None, ""):
                continue
            try:
                value = _parse_import_number(raw_value)
            except (TypeError, ValueError):
                rows.append(
                    {
                        "row_number": row_idx,
                        "owner_name": owner_name,
                        "budget_subject": budget_subject,
                        "field_name": field_name,
                        "field_label": _field_label(field_name),
                        "month": None,
                        "value": 0.0,
                        "error": f"{_field_label(field_name)}不是有效数字",
                    }
                )
                continue
            rows.append(
                {
                    "row_number": row_idx,
                    "owner_name": owner_name,
                    "budget_subject": budget_subject,
                    "field_name": field_name,
                    "field_label": _field_label(field_name),
                    "month": None,
                    "value": value,
                    "error": None,
                }
            )
    return rows


def parse_expense_forecast_import_file(raw: bytes) -> list[dict[str, Any]]:
    if not raw:
        raise ExpenseForecastImportParseError("上传文件为空")
    wb = load_workbook(BytesIO(raw), data_only=True)
    ws = wb[wb.sheetnames[0]]
    title_text = _text(ws.cell(1, 1).value)
    if "按事业群导出" in title_text:
        raise ExpenseForecastImportParseError(
            "当前文件为按事业群导出的汇总文件，不支持导入Excel；请切换到单个费用归属部门后重新导出并填写导入。"
        )
    header_row, header_map = _scan_import_header(ws)
    return _parse_import_rows(ws, header_row=header_row, header_map=header_map)


def parse_expense_forecast_group_import_file(raw: bytes, *, owner_names: list[str]) -> list[dict[str, Any]]:
    if not raw:
        raise ExpenseForecastImportParseError("上传文件为空")
    wb = load_workbook(BytesIO(raw), data_only=True)
    ws = wb[wb.sheetnames[0]]
    title_text = _text(ws.cell(1, 1).value)
    if "按事业群导出" not in title_text:
        raise ExpenseForecastImportParseError("“全部部门”导入请使用按事业群导出的Excel模板。")
    header_row, header_map = _scan_import_header(ws)
    owner_name_set = {_text(name) for name in owner_names if _text(name)}
    owner_name_by_row: dict[int, str] = {}
    current_owner = ""
    for row_idx in range(header_row + 1, ws.max_row + 1):
        first_col_text = _text(ws.cell(row_idx, int(header_map["预算科目"])).value)
        if not first_col_text:
            continue
        if first_col_text in owner_name_set:
            current_owner = first_col_text
            continue
        if current_owner:
            owner_name_by_row[row_idx] = current_owner
    parsed_rows = _parse_import_rows(ws, header_row=header_row, header_map=header_map, owner_name_by_row=owner_name_by_row)
    return [row for row in parsed_rows if _text(row.get("owner_name"))]


def parse_expense_forecast_subject_import_file(
    raw: bytes,
    *,
    subject_name: str,
    default_owner_name: str = "",
) -> list[dict[str, Any]]:
    if not raw:
        raise ExpenseForecastImportParseError("上传文件为空")
    wb = load_workbook(BytesIO(raw), data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row, header_map = _scan_subject_import_header(ws)
    owner_col = int(header_map["费用归属部门"])
    month_cols = _month_columns(header_map)
    annual_cols = _annual_columns(header_map)
    _ensure_import_value_columns(month_cols, annual_cols)

    rows: list[dict[str, Any]] = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        owner_name = _text(ws.cell(row_idx, owner_col).value) or default_owner_name
        if not owner_name:
            continue
        for month, col in month_cols.items():
            raw_value = ws.cell(row_idx, col).value
            if raw_value in (None, ""):
                continue
            try:
                value = _parse_import_number(raw_value)
            except (TypeError, ValueError):
                rows.append(
                    {
                        "row_number": row_idx,
                        "owner_name": owner_name,
                        "budget_subject": subject_name,
                        "field_name": "month_forecast",
                        "field_label": _field_label("month_forecast", month),
                        "month": month,
                        "value": 0.0,
                        "error": f"月份 M{month} 不是有效数字",
                    }
                )
                continue
            rows.append(
                {
                    "row_number": row_idx,
                    "owner_name": owner_name,
                    "budget_subject": subject_name,
                    "field_name": "month_forecast",
                    "field_label": _field_label("month_forecast", month),
                    "month": month,
                    "value": value,
                    "error": None,
                }
            )
        for field_name, col in annual_cols.items():
            raw_value = ws.cell(row_idx, col).value
            if raw_value in (None, ""):
                continue
            try:
                value = _parse_import_number(raw_value)
            except (TypeError, ValueError):
                rows.append(
                    {
                        "row_number": row_idx,
                        "owner_name": owner_name,
                        "budget_subject": subject_name,
                        "field_name": field_name,
                        "field_label": _field_label(field_name),
                        "month": None,
                        "value": 0.0,
                        "error": f"{_field_label(field_name)}不是有效数字",
                    }
                )
                continue
            rows.append(
                {
                    "row_number": row_idx,
                    "owner_name": owner_name,
                    "budget_subject": subject_name,
                    "field_name": field_name,
                    "field_label": _field_label(field_name),
                    "month": None,
                    "value": value,
                    "error": None,
                }
            )
    return rows
