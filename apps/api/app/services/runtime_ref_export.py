from __future__ import annotations

import asyncio
import sqlite3
import tempfile
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font

from app.core.config import settings
from app.core.database import get_pool
from app.core.db_paths import common_db_path
from app.services.runtime_metric_refs import (
    load_org_product_metric_refs_by_runtime_ref_code,
    load_org_product_metric_refs_by_runtime_ref_code_sync,
)
from app.services.org_product_runtime_catalog import org_product_runtime_products_cte_for_db
from app.services.org_product_runtime_catalog import org_product_runtime_products_cte


RUNTIME_REF_EXPORT_HEADERS = (
    "指标路径",
    "机构及产品指标编码",
    "产品代码",
    "机构及产品指标主键",
    "机构及产品指标名称",
    "预算数计算公式",
    "实际数计算公式",
    "所属机构及产品代码",
    "所属机构及产品名称",
    "数值类型",
    "备注",
    "是否公式计算",
    "是否允许手工补录",
    "机构产品引用数量",
    "机构产品来源",
)

RUNTIME_REF_INTRO_ROWS = (
    ("定位", "机构及产品指标编码直接来自机构及产品指标体系主键，不作为独立配置入口"),
    ("唯一配置入口", "机构及产品指标"),
    ("主键规则", "兼容读模型编码与机构及产品指标编码保持同一业务含义"),
    ("导入限制", "本文件用于查看和核对机构及产品指标编码，不作为独立配置导入模板"),
)


def _row_value(row: Any, key: str, index: int) -> Any:
    if isinstance(row, dict):
        return row.get(key)
    try:
        return row[key]
    except (TypeError, KeyError, IndexError):
        return row[index]


def _uses_mysql_path(path: Path | str) -> bool:
    try:
        candidate = Path(path).expanduser().resolve()
    except TypeError:
        return False
    temp_root = Path(tempfile.gettempdir()).expanduser().resolve()
    try:
        candidate.relative_to(temp_root)
        return False
    except ValueError:
        pass
    data_dir = Path(settings.data_dir).expanduser().resolve()
    try:
        candidate.relative_to(data_dir)
    except ValueError:
        return False
    return candidate.name == "common.db" or candidate.name == "compare.db" or (
        candidate.name.startswith("budget_") and candidate.suffix == ".db"
    )


def _metric_path_labels_from_rows(rows: list[Any]) -> dict[str, str]:
    by_code = {
        str(_row_value(r, "node_code", 0)): {
            "name": str(_row_value(r, "node_name", 1) or ""),
            "parent": (
                str(_row_value(r, "parent_code", 2))
                if _row_value(r, "parent_code", 2) is not None
                else None
            ),
        }
        for r in rows
    }
    memo: dict[str, str] = {}

    def path_for(code: str) -> str:
        if code in memo:
            return memo[code]
        node = by_code.get(code)
        if not node:
            memo[code] = ""
            return ""
        parent = node["parent"]
        parts = []
        if parent:
            parent_path = path_for(parent)
            if parent_path:
                parts.append(parent_path)
        parts.append(node["name"])
        memo[code] = " / ".join([p for p in parts if p])
        return memo[code]

    return {code: path_for(code) for code in by_code}


async def _load_metric_path_labels(db: Any) -> dict[str, str]:
    cur = await db.execute(
        """
        SELECT node_code, node_name, parent_code
        FROM data_account_metric_node
        """
    )
    return _metric_path_labels_from_rows(await cur.fetchall())


async def build_runtime_ref_export_workbook(db: Any) -> BytesIO:
    cur = await db.execute(
        f"""
        {org_product_runtime_products_cte_for_db(db)}
        SELECT d.data_acct_code, d.data_acct_name,
               d.budget_formula, d.actual_formula, d.value_type, d.remark,
               d.need_calc, d.formula_calc_mode, d.allow_manual_entry,
               b.metric_node_code, b.scope_code, p.product_name
        FROM data_account d
        LEFT JOIN data_account_metric_binding b ON b.data_acct_code = d.data_acct_code AND b.is_active = 1
        LEFT JOIN data_account_metric_node n ON n.node_code = b.metric_node_code
        LEFT JOIN org_product_runtime_products p ON b.scope_type = 'PRODUCT' AND p.product_code = b.scope_code
        ORDER BY d.data_acct_code
        """
    )
    rows = await cur.fetchall()
    metric_path_labels = await _load_metric_path_labels(db)
    cur = await db.execute(
        """
        SELECT b.data_acct_code, b.metric_node_code, b.scope_code,
               b.data_acct_code
        FROM data_account_metric_binding b
        ORDER BY b.data_acct_code
        """
    )
    binding_rows = await cur.fetchall()
    org_product_refs_by_data = await load_org_product_metric_refs_by_runtime_ref_code(db)
    bindings_by_data: dict[str, list[tuple[str, str, str, str]]] = {}
    for br in binding_rows:
        binding_code = str(br[0] or "")
        metric_node_code = str(br[1] or "")
        scope_code = str(br[2] or "")
        data_code = str(br[3] or "")
        if not data_code:
            continue
        bindings_by_data.setdefault(data_code, []).append(
            (
                binding_code,
                metric_path_labels.get(metric_node_code, ""),
                metric_node_code,
                scope_code,
            )
        )

    return _build_workbook(rows, bindings_by_data, org_product_refs_by_data)


async def export_runtime_refs_workbook(common_db: Path | str | None = None) -> BytesIO:
    path = common_db if common_db is not None else common_db_path()
    if _uses_mysql_path(path):
        return await _build_runtime_ref_export_workbook_mysql()
    return await asyncio.to_thread(_build_runtime_ref_export_workbook_sqlite, Path(path))


async def _build_runtime_ref_export_workbook_mysql() -> BytesIO:
    rows = await get_pool().fetch_all(
        f"""
        {org_product_runtime_products_cte(dialect="mysql")}
        SELECT d.data_acct_code AS data_acct_code, d.data_acct_name AS data_acct_name,
               d.budget_formula AS budget_formula, d.actual_formula AS actual_formula,
               d.value_type AS value_type, d.remark AS remark,
               d.need_calc AS need_calc, d.formula_calc_mode AS formula_calc_mode,
               d.allow_manual_entry AS allow_manual_entry,
               b.metric_node_code AS metric_node_code, b.scope_code AS scope_code,
               p.product_name AS product_name
        FROM data_account d
        LEFT JOIN data_account_metric_binding b ON b.data_acct_code = d.data_acct_code AND b.is_active = 1
        LEFT JOIN data_account_metric_node n ON n.node_code = b.metric_node_code
        LEFT JOIN org_product_runtime_products p ON b.scope_type = 'PRODUCT' AND p.product_code = b.scope_code
        ORDER BY d.data_acct_code
        """
    )
    metric_path_rows = await get_pool().fetch_all(
        """
        SELECT node_code, node_name, parent_code
        FROM data_account_metric_node
        """
    )
    metric_path_labels = _metric_path_labels_from_rows(metric_path_rows)
    binding_rows = await get_pool().fetch_all(
        """
        SELECT b.data_acct_code AS data_acct_code,
               b.metric_node_code AS metric_node_code,
               b.scope_code AS scope_code,
               b.data_acct_code AS binding_code
        FROM data_account_metric_binding b
        ORDER BY b.data_acct_code
        """
    )
    org_product_refs_by_data = await _load_org_product_metric_refs_by_runtime_ref_code_mysql()
    return _build_workbook(
        rows,
        _bindings_by_data(binding_rows, metric_path_labels),
        org_product_refs_by_data,
    )


async def _load_org_product_metric_refs_by_runtime_ref_code_mysql() -> dict[str, list[str]]:
    try:
        rows = await get_pool().fetch_all(
            """
            SELECT node_code, node_name, product_code, metric_table_name
            FROM data_account_metric_node
            WHERE is_active = 1
              AND runtime_account_enabled = 1
              AND COALESCE(product_code, '') <> ''
              AND COALESCE(metric_table_name, '') <> ''
            """
        )
    except Exception:
        return {}
    refs_by_code: dict[str, list[str]] = {}
    seen: set[tuple[str, str]] = set()
    for row in rows:
        data_acct_code = str(_row_value(row, "node_code", 0) or "").strip().upper()
        metric_name = str(_row_value(row, "node_name", 1) or "").strip()
        entity_code = str(_row_value(row, "product_code", 2) or "").strip().upper()
        table_name = str(_row_value(row, "metric_table_name", 3) or "").strip()
        if not data_acct_code or not entity_code or not table_name:
            continue
        source_ref = f"{entity_code}:{table_name}:{data_acct_code}"
        label = f"{source_ref} {metric_name}".strip()
        dedupe_key = (data_acct_code, source_ref)
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)
        refs_by_code.setdefault(data_acct_code, []).append(label)
    return {code: sorted(refs) for code, refs in refs_by_code.items()}


def _build_runtime_ref_export_workbook_sqlite(path: Path) -> BytesIO:
    with sqlite3.connect(path) as conn:
        rows = conn.execute(
            f"""
            {org_product_runtime_products_cte(dialect="sqlite")}
            SELECT d.data_acct_code, d.data_acct_name,
                   d.budget_formula, d.actual_formula, d.value_type, d.remark,
                   d.need_calc, d.formula_calc_mode, d.allow_manual_entry,
                   b.metric_node_code, b.scope_code, p.product_name
            FROM data_account d
            LEFT JOIN data_account_metric_binding b ON b.data_acct_code = d.data_acct_code AND b.is_active = 1
            LEFT JOIN data_account_metric_node n ON n.node_code = b.metric_node_code
            LEFT JOIN org_product_runtime_products p ON b.scope_type = 'PRODUCT' AND p.product_code = b.scope_code
            ORDER BY d.data_acct_code
            """
        ).fetchall()
        metric_path_rows = conn.execute(
            """
            SELECT node_code, node_name, parent_code
            FROM data_account_metric_node
            """
        ).fetchall()
        binding_rows = conn.execute(
            """
            SELECT b.data_acct_code, b.metric_node_code, b.scope_code,
                   b.data_acct_code
            FROM data_account_metric_binding b
            ORDER BY b.data_acct_code
            """
        ).fetchall()
        metric_path_labels = _metric_path_labels_from_rows(metric_path_rows)
        org_product_refs_by_data = {
            code: list(refs)
            for code, refs in load_org_product_metric_refs_by_runtime_ref_code_sync(conn).items()
        }
    return _build_workbook(
        rows,
        _bindings_by_data(binding_rows, metric_path_labels),
        org_product_refs_by_data,
    )


def _bindings_by_data(
    binding_rows: list[Any],
    metric_path_labels: dict[str, str],
) -> dict[str, list[tuple[str, str, str, str]]]:
    bindings_by_data: dict[str, list[tuple[str, str, str, str]]] = {}
    for br in binding_rows:
        binding_code = str(_row_value(br, "binding_code", 0) or "")
        metric_node_code = str(_row_value(br, "metric_node_code", 1) or "")
        scope_code = str(_row_value(br, "scope_code", 2) or "")
        data_code = str(_row_value(br, "data_acct_code", 3) or "")
        if not data_code:
            continue
        bindings_by_data.setdefault(data_code, []).append(
            (
                binding_code,
                metric_path_labels.get(metric_node_code, ""),
                metric_node_code,
                scope_code,
            )
        )
    return bindings_by_data


def _build_workbook(
    rows: list[Any],
    bindings_by_data: dict[str, list[tuple[str, str, str, str]]],
    org_product_refs_by_data: dict[str, list[str]] | None = None,
) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "机构及产品指标编码清单"
    intro_ws = wb.create_sheet("运行说明", 0)
    intro_ws.cell(row=1, column=1, value="项目").font = Font(bold=True)
    intro_ws.cell(row=1, column=2, value="说明").font = Font(bold=True)
    for row_idx, (item, note) in enumerate(RUNTIME_REF_INTRO_ROWS, start=2):
        intro_ws.cell(row=row_idx, column=1, value=item)
        intro_ws.cell(row=row_idx, column=2, value=note)
    intro_ws.column_dimensions["A"].width = 18
    intro_ws.column_dimensions["B"].width = 88

    header_to_col = {h: idx for idx, h in enumerate(RUNTIME_REF_EXPORT_HEADERS, start=1)}
    for idx, header in enumerate(RUNTIME_REF_EXPORT_HEADERS, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.font = Font(bold=True)

    row_idx = 2
    for r in rows:
        data_code = str(_row_value(r, "data_acct_code", 0) or "")
        org_product_refs = (org_product_refs_by_data or {}).get(data_code.upper(), [])
        row_bindings = bindings_by_data.get(data_code) or [("", "", "", "")]
        for _binding_code, metric_path, metric_node_code, scope_code in row_bindings:
            ws.cell(row=row_idx, column=header_to_col["指标路径"], value=metric_path)
            ws.cell(row=row_idx, column=header_to_col["机构及产品指标编码"], value=metric_node_code)
            ws.cell(row=row_idx, column=header_to_col["产品代码"], value=scope_code)
            ws.cell(row=row_idx, column=header_to_col["机构及产品指标主键"], value=data_code)
            ws.cell(
                row=row_idx,
                column=header_to_col["机构及产品指标名称"],
                value=_row_value(r, "data_acct_name", 1),
            )
            ws.cell(
                row=row_idx,
                column=header_to_col["预算数计算公式"],
                value=_row_value(r, "budget_formula", 2),
            )
            ws.cell(
                row=row_idx,
                column=header_to_col["实际数计算公式"],
                value=_row_value(r, "actual_formula", 3),
            )
            prod_code = str(scope_code or "").strip()
            prod_name = (
                "全行"
                if prod_code == "CORP"
                else str(_row_value(r, "product_name", 11) or "")
            )
            ws.cell(row=row_idx, column=header_to_col["所属机构及产品代码"], value=prod_code)
            ws.cell(row=row_idx, column=header_to_col["所属机构及产品名称"], value=prod_name)
            ws.cell(row=row_idx, column=header_to_col["数值类型"], value=_row_value(r, "value_type", 4))
            ws.cell(row=row_idx, column=header_to_col["备注"], value=_row_value(r, "remark", 5))
            ws.cell(
                row=row_idx,
                column=header_to_col["是否公式计算"],
                value=int(_row_value(r, "formula_calc_mode", 7) or 0),
            )
            allow_manual_entry = _row_value(r, "allow_manual_entry", 8)
            ws.cell(
                row=row_idx,
                column=header_to_col["是否允许手工补录"],
                value=int(1 if allow_manual_entry is None else allow_manual_entry),
            )
            ws.cell(row=row_idx, column=header_to_col["机构产品引用数量"], value=len(org_product_refs))
            ws.cell(row=row_idx, column=header_to_col["机构产品来源"], value="\n".join(org_product_refs))
            row_idx += 1

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
