"""Workbook export builder for budget simulation results."""
from __future__ import annotations

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.schemas import SimulationBaselineRow, SimulationInputItem, SimulationResultRow


HEADER_FILL = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFFFF")


def _style_header(ws) -> None:
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for col_idx, _ in enumerate(ws[1], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18


def build_budget_simulation_export_buffer(
    *,
    params: list[SimulationInputItem],
    result_rows: list[SimulationResultRow],
    baseline_rows: list[SimulationBaselineRow] | None = None,
) -> tuple[BytesIO, str]:
    """Build the budget simulation workbook without knowing HTTP details."""
    baseline_by_key = {
        (
            str(row.indicator_code or "").strip().upper(),
            str(row.product_code or "").strip().upper(),
        ): row
        for row in (baseline_rows or [])
    }
    wb = Workbook()
    ws_params = wb.active
    ws_params.title = "模拟参数"
    ws_params.append(["指标编码", "产品编码", "模拟值", "机构及产品指标编码", "机构产品来源"])
    for item in params:
        indicator_code = str(item.indicator_code or "").strip().upper()
        product_code = str(item.product_code or "").strip().upper()
        baseline = baseline_by_key.get((indicator_code, product_code))
        ws_params.append(
            [
                indicator_code,
                product_code,
                float(item.simulate_value),
                "\n".join(baseline.source_metric_codes) if baseline else "",
                "\n".join(baseline.source_org_product_refs) if baseline else "",
            ]
        )

    ws_result = wb.create_sheet("测算结果")
    ws_result.append(["指标组", "指标编码", "指标名称", "值类型", "2025基准", "2026基准", "2026模拟"])
    for row in result_rows:
        ws_result.append(
            [
                row.metric_group,
                row.indicator_code,
                row.indicator_name,
                row.value_type,
                row.baseline_2025,
                row.baseline_2026,
                row.simulation_2026,
            ]
        )

    for ws in (ws_params, ws_result):
        _style_header(ws)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"budget_simulation_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    return buffer, filename
