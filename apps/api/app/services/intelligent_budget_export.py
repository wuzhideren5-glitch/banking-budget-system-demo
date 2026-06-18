"""Excel export for intelligent budget simulation results."""
from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFFFF")


def _style(ws) -> None:
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20


def build_intelligent_budget_simulation_export(task: dict[str, Any]) -> tuple[BytesIO, str]:
    wb = Workbook()
    ws_target = wb.active
    ws_target.title = "目标与约束"
    parsed = task.get("parsed_target", {})
    ws_target.append(["项目", "内容"])
    ws_target.append(["原始目标", task.get("target_text", "")])
    ws_target.append(["净利润目标", parsed.get("min_net_profit_growth", "")])
    ws_target.append(["不良率目标", parsed.get("max_npl_ratio", "")])
    ws_target.append(["软偏好", "、".join(parsed.get("soft_preferences", []))])
    ws_target.append(["可调因子", "、".join(parsed.get("adjustable_factors", []))])

    ws_steps = wb.create_sheet("步长摘要")
    ws_steps.append(["项目", "内容"])
    ws_steps.append(["步长摘要", task.get("step_summary", "")])

    ws_snapshot = wb.create_sheet("预算结果快照")
    ws_snapshot.append(["方案", "角色", "贷款余额", "生息资产", "营业收入", "净利息收入", "费用", "拨备/减值", "净利润", "净利润增长", "不良余额", "不良率", "风险成本率", "拨备余额", "超额拨备"])
    snapshot_solutions = [task.get("baseline_solution", {})] + list(task.get("solutions", []))
    for solution in snapshot_solutions:
        snap = solution.get("budget_snapshot", {}) if isinstance(solution, dict) else {}
        ws_snapshot.append(
            [
                solution.get("name"),
                solution.get("display_role"),
                snap.get("loan_balance"),
                snap.get("interest_earning_assets"),
                snap.get("operating_income"),
                snap.get("net_interest_income"),
                snap.get("operating_expense"),
                snap.get("impairment_loss"),
                snap.get("net_profit"),
                snap.get("net_profit_growth"),
                snap.get("npl_balance"),
                snap.get("npl_ratio"),
                snap.get("risk_cost_rate"),
                snap.get("provision_balance"),
                snap.get("excess_provision"),
            ]
        )

    ws_top = wb.create_sheet("Top10方案")
    ws_top.append(["排名", "方案名称", "推荐角色", "推荐理由", "数学评分", "净利润增长", "不良率", "核心动作"])
    for solution in task.get("solutions", []):
        core = solution.get("core_actions", {})
        ws_top.append(
            [
                solution.get("rank"),
                solution.get("name"),
                solution.get("display_role"),
                solution.get("recommendation_reason"),
                solution.get("math_score"),
                solution.get("net_profit_growth"),
                solution.get("npl_ratio"),
                "\n".join(f"{key}: {value}" for key, value in core.items()),
            ]
        )

    ws_factors = wb.create_sheet("二层因子")
    ws_factors.append(["方案", "规模", "收益率bp", "费用", "新生成不良控制", "回收/清收提升", "拨备调节"])
    for solution in task.get("solutions", []):
        factors = solution.get("factor_movements", {})
        ws_factors.append(
            [
                solution.get("name"),
                factors.get("scale_growth"),
                factors.get("yield_bp"),
                factors.get("expense_growth"),
                factors.get("new_npl_control"),
                factors.get("recovery_improvement"),
                factors.get("provision_delta"),
            ]
        )

    ws_products = wb.create_sheet("产品拆解")
    ws_products.append(["方案", "产品编码", "产品名称", "规模", "收益率bp", "风险动作", "费用", "边际贡献", "是否Top5"])
    for solution in task.get("solutions", []):
        for product in solution.get("top_product_contributions", []):
            ws_products.append(
                [
                    solution.get("name"),
                    product.get("product_code"),
                    product.get("product_name"),
                    product.get("scale_growth"),
                    product.get("yield_bp"),
                    product.get("risk_action"),
                    product.get("expense_growth"),
                    product.get("marginal_contribution"),
                    "是",
                ]
            )
        if solution.get("other_product_contribution", 0):
            ws_products.append([solution.get("name"), "OTHER", "其他产品", "", "", "", "", solution.get("other_product_contribution"), "否"])

    ws_risk = wb.create_sheet("风险传导")
    ws_risk.append(["方案", "期初不良余额", "基准新生成不良", "新生成不良压降", "方案后新生成不良", "压降金额", "基准回收清收", "回收清收提升", "方案后回收清收", "提升金额", "核销处置", "期末不良余额", "期末贷款规模", "推导不良率"])
    for solution in snapshot_solutions:
        bridge = solution.get("risk_bridge", {}) if isinstance(solution, dict) else {}
        ws_risk.append(
            [
                solution.get("name"),
                bridge.get("opening_npl_balance"),
                bridge.get("baseline_new_npl_amount"),
                bridge.get("new_npl_control_rate"),
                bridge.get("after_new_npl_amount"),
                bridge.get("new_npl_reduction_amount"),
                bridge.get("baseline_recovery_amount"),
                bridge.get("recovery_improvement_rate"),
                bridge.get("after_recovery_amount"),
                bridge.get("recovery_increment_amount"),
                bridge.get("writeoff_disposal_amount"),
                bridge.get("ending_npl_balance"),
                bridge.get("ending_loan_scale"),
                bridge.get("derived_npl_ratio"),
            ]
        )

    ws_negotiation = wb.create_sheet("协商记录")
    ws_negotiation.append(["项目", "内容"])
    ws_negotiation.append(["状态", task.get("status", "")])
    ws_negotiation.append(["协商提示", task.get("negotiation_message", "")])
    ws_negotiation.append(["建议", "\n".join(task.get("negotiation_suggestions") or [])])

    for ws in wb.worksheets:
        _style(ws)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    filename = f"intelligent_budget_simulation_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    return out, filename
