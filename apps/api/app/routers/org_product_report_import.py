from __future__ import annotations

import app.core.pymysql_compat  # noqa: F401 -- SQLite->MySQL compat
from io import BytesIO

from fastapi import APIRouter, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook

from app.routers.org_product_helpers import *

router = APIRouter()

@router.post("/api/org-product-metrics/import-report")
async def import_org_product_metrics_report(
    file: UploadFile = File(...),
    candidates_json: str = Form(""),
    strict_import: str = Form("true"),
):
    content = await file.read()
    strict_flag = _parse_strict_import_flag(strict_import)
    return parse_metric_workbook_import(
        content,
        file.filename or "report.xlsx",
        candidates_json=candidates_json,
        strict_import=strict_flag,
    )

@router.post("/api/org-product-metrics/export")
async def export_org_product_metrics(payload: MetricExportPayload):
    wb = Workbook()
    ws = wb.active
    sheet_name = f"{payload.entity_code}{payload.entity_name}{payload.table_name}".replace("/", "_").replace("\\", "_")
    ws.title = (sheet_name or "机构及产品指标")[:31]

    headers = metric_export_headers_v04(include_admin_columns=True)
    ws.append(headers)

    rows: list[dict[str, str]] = []
    _build_metric_rows([item.model_dump() for item in payload.metrics], rows, payload.entity_code)
    for row in rows:
        ws.append(metric_export_row_values(row, include_admin_columns=True))

    _style_metric_export_header_row(ws, len(headers))
    _apply_metric_export_v04_column_widths(ws, include_admin_columns=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"{payload.entity_code}_{payload.entity_name}_{payload.table_name}.xlsx".replace(" ", "")
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

@router.post("/api/org-product-metrics/export-report")
async def export_org_product_metrics_report(payload: MetricReportExportPayload):
    if not payload.sheets:
        raise HTTPException(status_code=400, detail="sheets 不能为空")

    wb = Workbook()
    wb.remove(wb.active)

    used_titles: set[str] = set()
    for sheet in payload.sheets:
        entity_code = _normalize_text(sheet.entity_code).upper()
        table_name = _normalize_text(sheet.table_name)
        if not entity_code or not table_name:
            continue
        title = _unique_sheet_title(used_titles, f"{entity_code}{table_name}")
        ws = wb.create_sheet(title=title)

        headers = metric_export_headers_v04(include_admin_columns=True)
        ws.append(headers)

        rows: list[dict[str, str]] = []
        _build_metric_rows([item.model_dump() for item in sheet.metrics], rows, sheet.entity_code)
        for row in rows:
            ws.append(metric_export_row_values(row, include_admin_columns=True))

        _style_metric_export_header_row(ws, len(headers))
        _apply_metric_export_v04_column_widths(ws, include_admin_columns=True)

    if not wb.sheetnames:
        raise HTTPException(status_code=400, detail="没有可导出的工作表")

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = "机构及产品指标.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": _attachment_content_disposition(filename, ascii_fallback="org_product_metrics.xlsx")},
    )

