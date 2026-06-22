from __future__ import annotations

import calendar
from datetime import datetime, timezone
from io import BytesIO
import json
from pathlib import Path
import re
import sqlite3
import app.core.pymysql_compat  # noqa: F401 -- SQLite->MySQL compat
import unicodedata
from collections import defaultdict
from typing import Any
from urllib.parse import quote

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from pydantic import BaseModel

from app.core.config import settings
from app.core.db_paths import budget_db_path, common_db_path
from app.org_product_excel_formula import (
    ExcelFormulaConvertError,
    SheetFormulaContext,
    build_sheet_formula_context,
    convert_excel_formula_to_system,
    index_sheet_contexts,
    normalize_sheet_lookup_key,
)
from app.services.budget_fact_periods import load_budget_fact_period_month_map_sync
from app.services.budget_fact_versions import BudgetFactVersionNotFound, load_budget_fact_version_identity_sync
from app.services.org_product_budget_sync import (
    OrgProductBudgetSyncPlan,
    apply_org_product_budget_sync_plan,
    plan_org_product_budget_sync,
)
from app.services.export_common import workbook_streaming_response
from app.services.org_product_metric_runtime_sync import (
    dedupe_org_product_metric_payload_nodes,
    OrgProductMetricRuntimeSyncError,
    sync_org_product_metric_runtime_refs,
)
from app.services.org_product_metric_runtime_snapshot import (
    load_org_product_metric_payload_from_runtime_tree,
    load_org_product_metric_table_rows_from_runtime_tree,
)
from app.services.org_product_runtime_catalog import sync_org_product_runtime_catalog_from_tree
from app.services.runtime_metric_refs import (
    compact_org_product_metric_code,
    normalize_org_product_metric_code,
)

ORG_METRIC_FILE = settings.business_inputs_dir / "机构汇总指标.xlsx"
PRODUCT_METRIC_FILE = settings.business_inputs_dir / "产品指标.xlsx"
LEVEL_RANK = {
    "一级": 1,
    "二级": 2,
    "三级": 3,
    "四级": 4,
    "五级": 5,
    "六级": 6,
}
PROFIT_NAMES = {"分摊前利润", "税前利润", "净利润"}
DEFAULT_METRIC_TABLE_NAME = "业务状况表"
METRIC_SHEET_TABLE_ALIASES = {
    "净利息收入表": "利息净收入表",
}
METRIC_SHEET_TITLE_TABLE_ALIASES = {
    "资产负债表日均": "资产负债表（日均）",
    "资产负债表(日均)": "资产负债表（日均）",
    "资产负债表余额": "资产负债表（余额）",
    "资产负债表(余额)": "资产负债表（余额）",
    "资产质量": "资产质量表",
}
METRIC_HEADER_ALIASES = {
    "层级": "科目层级",
    "科目级别": "科目层级",
    "级别": "科目层级",
    "性质": "科目性质",
    "代码": "科目代码",
    "科目编码": "科目代码",
    "编码": "科目代码",
    "指标代码": "科目代码",
    "名称": "科目名称",
    "科目": "科目名称",
    "指标名称": "科目名称",
    "公式": "取数公式",
    "取数公式": "取数公式",
    "年预算公式": "年预算公式",
    "年度预算公式": "年预算公式",
    "全年预算公式": "年预算公式",
    "预算年公式": "年预算公式",
    "年预测公式": "年预测公式",
    "年度预测公式": "年预测公式",
    "全年预测公式": "年预测公式",
    "预测年公式": "年预测公式",
    "取数公式(实际)": "实际月公式",
    "取数公式（实际）": "实际月公式",
    "月度实际公式": "实际月公式",
    "实际月公式": "实际月公式",
    "取数公式(预测)": "预测月公式",
    "取数公式（预测）": "预测月公式",
    "月度预测公式": "预测月公式",
    "预测月公式": "预测月公式",
    "公式说明": "公式说明",
    "录入粒度": "录入粒度",
    "时间粒度": "录入粒度",
    "数值类型": "数值类型",
    "值类型": "数值类型",
    "数据类型": "数值类型",
    "是否允许手工录入": "允许手工录入",
    "允许手工录入": "允许手工录入",
    "是否允许手工补录": "允许手工录入",
    "允许手工补录": "允许手工录入",
    "机构及产品指标编码": "机构及产品指标编码",
    "横向汇总": "横向汇总",
    "是否横向汇总": "横向汇总",
    "横向汇总标识": "横向汇总",
    "纵向汇总": "纵向汇总",
    "是否纵向汇总": "纵向汇总",
    "纵向汇总标识": "纵向汇总",
    "逻辑码": "逻辑码",
    "逻辑代码": "逻辑码",
    "指标逻辑码": "逻辑码",
    "科目层级※仅展示": "科目层级",
    "科目层级（仅展示）": "科目层级",
    "科目层级(仅展示)": "科目层级",
    "逻辑码※仅展示": "逻辑码",
    "逻辑码（仅展示）": "逻辑码",
    "逻辑码(仅展示)": "逻辑码",
    "规则": "规则",
    "年度规则": "规则",
    "时间汇总规则": "规则",
}
METRIC_EXPORT_HEADER_LEVEL = "科目层级※仅展示"
METRIC_EXPORT_HEADER_LOGIC = "逻辑码※仅展示"
METRIC_DISPLAY_ONLY_HEADER_SUFFIXES = ("※仅展示", "（仅展示）", "(仅展示)")
NATURE_VALUE_SET = {"收入", "支出", "利润", "其他", "其它"}
ENTRY_GRANULARITY_ANNUAL = {"annual", "year", "yearly", "按年", "按年录入", "仅年度", "仅年度录入", "年度", "全年", "仅全年录入"}
ENTRY_GRANULARITY_MONTHLY = {"monthly", "month", "按月", "按月录入", "月度", "按月录入（默认）"}
ORG_PRODUCT_OUTPUT_EXPORT_HEADERS = (
    ["科目层级", "科目性质", "科目代码", "科目名称"]
    + [f"{i}月" for i in range(1, 13)]
    + ["年度汇总"]
)


class MetricNodePayload(BaseModel):
    id: str
    levelLabel: str
    nature: str
    code: str
    name: str
    value_type: str = ""
    allow_manual_entry: int | bool | str = 1
    note: str = ""
    formula: str = ""
    formula_budget_annual: str = ""
    formula_forecast_annual: str = ""
    formula_actual: str = ""
    formula_forecast: str = ""
    formula_note: str = ""
    entry_granularity: str = "monthly"
    horizontal_rollup: int | bool | str = 0
    vertical_rollup: int | bool | str = 0
    logic_code: str = ""
    annual_agg_rule: str = ""
    children: list["MetricNodePayload"] = []


class MetricExportPayload(BaseModel):
    entity_code: str
    entity_name: str
    table_name: str = "业务状况表"
    metrics: list[MetricNodePayload]


class MetricReportExportSheetPayload(BaseModel):
    entity_code: str
    table_name: str
    metrics: list[MetricNodePayload] = []


class MetricReportExportPayload(BaseModel):
    sheets: list[MetricReportExportSheetPayload] = []


class MetricTablePayload(BaseModel):
    id: str
    name: str
    metrics: list[MetricNodePayload]


class MetricSaveEntityPayload(BaseModel):
    entity_code: str
    entity_name: str
    tables: list[MetricTablePayload]


class MetricSavePayload(BaseModel):
    entities: list[MetricSaveEntityPayload]


class MetricSaveTablePayload(BaseModel):
    entity_code: str
    entity_name: str
    table_id: str
    table_name: str
    metrics: list[MetricNodePayload]


class DataEntryMetricValuesPayload(BaseModel):
    prev_actual: str = ""
    prev_budget: str = ""
    prev_forecast: str = ""
    year_forecast: str = ""
    months: dict[str, str] = {}


class DataEntryMetricRowPayload(BaseModel):
    metric_id: str
    metric_code: str = ""
    metric_name: str = ""
    levelLabel: str = ""
    nature: str = ""
    values: DataEntryMetricValuesPayload


class DataEntrySavePayload(BaseModel):
    entity_code: str
    entity_name: str = ""
    year: int
    version_id: int | None = None
    version_name: str = ""
    month_index: int | None = None
    table_id: str = ""
    table_name: str = ""
    entry_status: str = "draft"
    metrics: list[DataEntryMetricRowPayload] = []


class OrgProductDataEntryCommitRequest(BaseModel):
    entity_code: str
    year: int
    table_name: str
    user_ids: list[int] | None = None
    version_id: int | None = None
    version_name: str = ""
    force: bool = False


class OrgProductDataEntryBudgetSyncRequest(BaseModel):
    entity_code: str
    year: int
    table_name: str
    entry_version_id: int
    budget_version_id: int
    budget_actuals: list[int] | None = None


class OrgProductOutputRunRequest(BaseModel):
    entity_code: str
    year: int
    version_id: int
    table_name: str | None = None
    include_children: bool = False


class OrgProductOutputCommitRequest(BaseModel):
    entity_code: str
    year: int
    input_version_id: int
    table_name: str
    output_version_id: int | None = None
    output_version_name: str = ""
    force: bool = False


class OrgProductNodePayload(BaseModel):
    id: str
    code: str
    name: str
    type: str
    children: list["OrgProductNodePayload"] = []


class OrgProductTreeSavePayload(BaseModel):
    tree: OrgProductNodePayload


class MetricTableCatalogCreatePayload(BaseModel):
    entity_scope: str
    table_name: str
    sort_order: int | None = None
    remark: str = ""


class MetricTableCatalogPatchPayload(BaseModel):
    sort_order: int | None = None
    status: str | None = None
    remark: str | None = None


METRIC_TABLE_CATALOG_SCOPES = frozenset({"AA", "AB", "PRODUCT"})
METRIC_TABLE_CATALOG_STATUSES = frozenset({"active", "inactive"})
METRIC_TABLE_CATALOG_SCOPE_ENTITY_NAMES = {"AA": "微众银行", "AB": "微众科技"}
DEFAULT_METRIC_TABLE_CATALOG_ROWS: list[tuple[str, str, int]] = [
    ("AA", "业务状况表", 10),
    ("AA", "损益表", 20),
    ("AA", "资产负债表（余额）", 30),
    ("AA", "资产负债表（日均）", 40),
    ("AA", "资产质量表", 50),
    ("AA", "利息净收入表", 60),
    ("AA", "业务支出评估", 70),
    ("AB", "业务状况表", 10),
    ("AB", "损益表", 20),
    ("AB", "业务支出评估", 30),
    ("PRODUCT", "业务状况表", 10),
    ("PRODUCT", "业务支出评估", 20),
]


def _now_iso() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _attachment_content_disposition(filename: str, *, ascii_fallback: str = "download.xlsx") -> str:
    """HTTP 头仅支持 latin-1；中文文件名须用 filename* (RFC 5987)。"""
    safe_ascii = re.sub(r"[^\x20-\x7E]", "_", (ascii_fallback or "download.xlsx").strip()) or "download.xlsx"
    encoded = quote((filename or safe_ascii).strip() or safe_ascii)
    return f"attachment; filename=\"{safe_ascii}\"; filename*=UTF-8''{encoded}"



def _ensure_metric_table_catalog(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS org_product_metric_table_catalog (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            entity_scope TEXT NOT NULL,
            table_name TEXT NOT NULL,
            sort_order INTEGER NOT NULL DEFAULT 0,
            status TEXT NOT NULL DEFAULT 'active',
            remark TEXT,
            updated_at TEXT NOT NULL,
            UNIQUE(entity_scope, table_name)
        )
        """
    )
    conn.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_org_product_metric_table_catalog_scope
        ON org_product_metric_table_catalog(entity_scope, sort_order)
        """
    )


def _seed_metric_table_catalog(conn: sqlite3.Connection) -> None:
    cur = conn.execute("SELECT COUNT(*) FROM org_product_metric_table_catalog")
    if int(cur.fetchone()[0] or 0) > 0:
        return
    now = _now_iso()
    for entity_scope, table_name, sort_order in DEFAULT_METRIC_TABLE_CATALOG_ROWS:
        conn.execute(
            """
            INSERT INTO org_product_metric_table_catalog
            (entity_scope, table_name, sort_order, status, remark, updated_at)
            VALUES (?, ?, ?, 'active', '', ?)
            """,
            (entity_scope, table_name, sort_order, now),
        )


def _metric_table_catalog_row_to_dict(row: tuple[Any, ...]) -> dict[str, Any]:
    return {
        "id": int(row[0]),
        "entity_scope": str(row[1] or "").strip(),
        "table_name": str(row[2] or "").strip(),
        "sort_order": int(row[3] or 0),
        "status": str(row[4] or "active").strip() or "active",
        "remark": str(row[5] or "").strip() if row[5] is not None else "",
        "updated_at": str(row[6] or "").strip(),
    }


def _load_metric_table_catalog_rows(
    conn: sqlite3.Connection,
    *,
    entity_scope: str | None = None,
    active_only: bool = False,
) -> list[dict[str, Any]]:
    _ensure_metric_table_catalog(conn)
    _seed_metric_table_catalog(conn)
    clauses: list[str] = []
    params: list[Any] = []
    scope = _normalize_text(entity_scope).upper() if entity_scope else ""
    if scope:
        if scope not in METRIC_TABLE_CATALOG_SCOPES:
            raise ValueError(f"entity_scope 无效：{scope}")
        clauses.append("entity_scope = ?")
        params.append(scope)
    if active_only:
        clauses.append("status = 'active'")
    where_sql = f" WHERE {' AND '.join(clauses)}" if clauses else ""
    cur = conn.execute(
        f"""
        SELECT id, entity_scope, table_name, sort_order, status, remark, updated_at
        FROM org_product_metric_table_catalog
        {where_sql}
        ORDER BY entity_scope, sort_order, table_name
        """,
        params,
    )
    return [_metric_table_catalog_row_to_dict(r) for r in cur.fetchall()]


def _active_catalog_table_names_by_scope(conn: sqlite3.Connection) -> dict[str, list[str]]:
    rows = _load_metric_table_catalog_rows(conn, active_only=True)
    grouped: dict[str, list[str]] = defaultdict(list)
    for row in rows:
        scope = str(row.get("entity_scope") or "").strip()
        name = str(row.get("table_name") or "").strip()
        if scope and name:
            grouped[scope].append(name)
    return grouped


def _canonical_import_table_names(conn: sqlite3.Connection | None = None) -> list[str]:
    names: list[str] = []
    seen: set[str] = set()
    if conn is not None:
        try:
            for scope_rows in _active_catalog_table_names_by_scope(conn).values():
                for name in scope_rows:
                    canonical = _canonical_import_table_name(name)
                    key = _canonical_metric_table_key(canonical)
                    if key in seen:
                        continue
                    seen.add(key)
                    names.append(canonical)
        except Exception:
            pass
    if names:
        return names
    return [
        "业务状况表",
        "损益表",
        "资产负债表（余额）",
        "资产负债表（日均）",
        "资产质量表",
        "利息净收入表",
    ]


def _import_report_catalog_candidates(conn: sqlite3.Connection) -> list[tuple[str, str, str]]:
    grouped = _active_catalog_table_names_by_scope(conn)
    out: list[tuple[str, str, str]] = []
    for scope, table_names in grouped.items():
        if scope == "PRODUCT":
            continue
        entity_code = scope
        entity_name = METRIC_TABLE_CATALOG_SCOPE_ENTITY_NAMES.get(scope, scope)
        for table_name in table_names:
            out.append((entity_code, entity_name, table_name))
            if table_name == "利息净收入表":
                out.append((entity_code, entity_name, "净利息收入表"))
    return out


def _import_report_saved_metric_candidates(conn: sqlite3.Connection) -> list[tuple[str, str, str]]:
    rows = load_org_product_metric_table_rows_from_runtime_tree(conn)
    return [
        (
            _normalize_text(row["entity_code"]).upper(),
            _normalize_text(row["entity_name"]),
            _normalize_text(row["table_name"]),
        )
        for row in rows
        if _normalize_text(row.get("entity_code")) and _normalize_text(row.get("table_name"))
    ]


def _ensure_data_entry_snapshot_table(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS org_product_data_entry_snapshot (
            entity_code TEXT NOT NULL,
            entity_name TEXT NOT NULL,
            year INTEGER NOT NULL,
            month_index INTEGER,
            table_id TEXT,
            table_name TEXT,
            payload_json LONGTEXT NOT NULL,
            updated_at TEXT NOT NULL,
            PRIMARY KEY (entity_code, year)
        )
        """
    )
    conn.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_org_product_data_entry_snapshot_entity
        ON org_product_data_entry_snapshot(entity_code)
        """
    )


def _ensure_data_entry_snapshot_table_v2(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS org_product_data_entry_snapshot_v2 (
            entity_code TEXT NOT NULL,
            entity_name TEXT NOT NULL,
            year INTEGER NOT NULL,
            version_id INTEGER NOT NULL,
            version_name TEXT NOT NULL,
            table_name TEXT NOT NULL,
            month_index INTEGER,
            table_id TEXT,
            payload_json LONGTEXT NOT NULL,
            updated_at TEXT NOT NULL,
            PRIMARY KEY (entity_code, year, version_id, table_name)
        )
        """
    )
    conn.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_org_product_data_entry_snapshot_v2_entity_year
        ON org_product_data_entry_snapshot_v2(entity_code, year)
        """
    )


def _ensure_data_entry_draft_table(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS org_product_data_entry_draft (
            user_id INTEGER NOT NULL,
            user_display_name TEXT NOT NULL,
            entity_code TEXT NOT NULL,
            entity_name TEXT NOT NULL,
            year INTEGER NOT NULL,
            table_name TEXT NOT NULL,
            payload_json LONGTEXT NOT NULL,
            updated_at TEXT NOT NULL,
            PRIMARY KEY (user_id, entity_code, year, table_name)
        )
        """
    )
    conn.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_org_product_data_entry_draft_entity_year
        ON org_product_data_entry_draft(entity_code, year)
        """
    )


def _ensure_org_product_output_snapshot_table(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS org_product_output_snapshot_v1 (
            entity_code TEXT NOT NULL,
            entity_name TEXT NOT NULL,
            year INTEGER NOT NULL,
            input_version_id INTEGER NOT NULL,
            output_version_id INTEGER NOT NULL,
            output_version_name TEXT NOT NULL,
            table_name TEXT NOT NULL,
            payload_json LONGTEXT NOT NULL,
            updated_at TEXT NOT NULL,
            PRIMARY KEY (entity_code, year, input_version_id, output_version_id, table_name)
        )
        """
    )
    conn.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_org_product_output_snapshot_entity_year
        ON org_product_output_snapshot_v1(entity_code, year)
        """
    )


def _norm_text(value: Any) -> str:
    return str(value or "").strip()


def _is_non_empty(value: Any) -> bool:
    return bool(_norm_text(value))


def _next_data_entry_version_id(conn: sqlite3.Connection, entity_code: str, year: int, table_name: str) -> int:
    _ensure_data_entry_snapshot_table_v2(conn)
    cur = conn.execute(
        """
        SELECT MAX(version_id)
        FROM org_product_data_entry_snapshot_v2
        WHERE entity_code=? AND year=? AND table_name=?
        """,
        (entity_code, int(year), table_name),
    )
    row = cur.fetchone()
    max_id = int(row[0]) if row and row[0] is not None else 0
    return max_id + 1


def _sanitize_sheet_title(name: str) -> str:
    safe = (name or "").replace("/", "_").replace("\\", "_").strip()
    return safe[:31] if safe else "Sheet"


def _unique_sheet_title(existing: set[str], desired: str) -> str:
    base = _sanitize_sheet_title(desired)
    if base not in existing:
        existing.add(base)
        return base
    for idx in range(2, 1000):
        candidate = _sanitize_sheet_title(f"{base[:28]}_{idx}")
        if candidate not in existing:
            existing.add(candidate)
            return candidate
    fallback = _sanitize_sheet_title(f"{base[:24]}_{len(existing) + 1}")
    existing.add(fallback)
    return fallback


def _parse_data_entry_month_value(values: dict[str, Any], month: int) -> float | None:
    if month < 1 or month > 12:
        return None
    months = values.get("months") if isinstance(values, dict) else None
    if not isinstance(months, dict):
        return None
    raw_a = months.get(f"a{month}")
    raw_f = months.get(f"f{month}")
    raw = raw_a if str(raw_a or "").strip() else raw_f
    text = str(raw or "").strip()
    if not text:
        return None
    normalized = text.replace(",", "").replace("，", "").replace(" ", "").replace("%", "").replace("％", "")
    try:
        val = float(normalized)
    except Exception:
        return None
    if "%" in text or "％" in text:
        return val / 100.0
    return val


def _prepare_metric_formula_expression(formula: str | None) -> str:
    expr = str(formula or "").strip()
    if not expr:
        return ""
    translate_map = str.maketrans(
        {
            "（": "(",
            "）": ")",
            "，": ",",
            "＋": "+",
            "－": "-",
            "×": "*",
            "÷": "/",
        }
    )
    return expr.translate(translate_map)


_LOCAL_METRIC_CODE_RE = re.compile(r"\b[A-Za-z0-9]+(?:\.\d+)+\b")
_CROSS_ENTITY_REF_RE = re.compile(
    r"([A-Za-z0-9]{1,6})/([^/\s()+\-*.,]+)/([A-Za-z0-9]+(?:\.\d+)+)"
)
_CROSS_TABLE_REF_RE = re.compile(r"([^/\s()+\-*.,]+表)/([A-Za-z0-9]+(?:\.\d+)+)")


def _extract_metric_formula_refs(expression: str) -> list[str]:
    expr = str(expression or "")
    if not expr:
        return []
    refs: list[str] = []
    scrubbed = expr
    for m in _CROSS_ENTITY_REF_RE.finditer(expr):
        refs.append(m.group(0))
        scrubbed = scrubbed.replace(m.group(0), " ")
    for m in _CROSS_TABLE_REF_RE.finditer(scrubbed):
        refs.append(m.group(0))
        scrubbed = scrubbed.replace(m.group(0), " ")
    refs.extend(_LOCAL_METRIC_CODE_RE.findall(scrubbed))
    seen: set[str] = set()
    ordered: list[str] = []
    for r in refs:
        if r in seen:
            continue
        seen.add(r)
        ordered.append(r)
    return ordered


def _ref_to_identifier(ref: str) -> str:
    safe = re.sub(r"[^A-Za-z0-9_]+", "_", str(ref or ""))
    if not safe:
        return "R0"
    if safe[0].isdigit():
        return f"R_{safe}"
    return f"R_{safe}"


def _eval_formula_ast(node, values: dict[str, float]) -> float:
    import ast

    if isinstance(node, ast.Expression):
        return _eval_formula_ast(node.body, values)
    if isinstance(node, ast.Constant):
        if isinstance(node.value, (int, float)):
            return float(node.value)
        raise ValueError("公式包含非法常量")
    if isinstance(node, ast.Name):
        return float(values.get(node.id, 0.0))
    if isinstance(node, ast.BinOp):
        left = _eval_formula_ast(node.left, values)
        right = _eval_formula_ast(node.right, values)
        if isinstance(node.op, ast.Add):
            return left + right
        if isinstance(node.op, ast.Sub):
            return left - right
        if isinstance(node.op, ast.Mult):
            return left * right
        if isinstance(node.op, ast.Div):
            if right == 0:
                raise ZeroDivisionError("division by zero")
            return left / right
        raise ValueError("公式仅支持 + - * / 运算")
    if isinstance(node, ast.UnaryOp):
        val = _eval_formula_ast(node.operand, values)
        if isinstance(node.op, ast.UAdd):
            return +val
        if isinstance(node.op, ast.USub):
            return -val
        raise ValueError("公式一元运算不合法")
    if isinstance(node, ast.Call):
        if getattr(node, "keywords", None):
            raise ValueError("函数调用不支持命名参数")
        if not isinstance(node.func, ast.Name):
            raise ValueError("函数调用不合法")
        fn = node.func.id.upper()
        args = [_eval_formula_ast(arg, values) for arg in node.args]
        if fn == "SUM":
            return float(sum(args))
        if fn == "AVG":
            return float(sum(args) / len(args)) if args else 0.0
        if fn == "MAX":
            return float(max(args)) if args else 0.0
        if fn == "MIN":
            return float(min(args)) if args else 0.0
        raise ValueError("仅支持 SUM/AVG/MAX/MIN 函数")
    raise ValueError("公式语法不合法")


def _try_calculate_metric_formula_value(formula: str | None, ref_values: dict[str, float]) -> tuple[float, str | None]:
    import ast

    expression = _prepare_metric_formula_expression(formula)
    if not expression:
        return 0.0, None
    refs = _extract_metric_formula_refs(expression)
    values: dict[str, float] = {}
    replaced = expression
    for ref in sorted(refs, key=len, reverse=True):
        ident = _ref_to_identifier(ref)
        values[ident] = float(ref_values.get(ref, 0.0))
        replaced = re.sub(re.escape(ref), ident, replaced)
    try:
        parsed = ast.parse(replaced, mode="eval")
        return float(_eval_formula_ast(parsed, values)), None
    except ZeroDivisionError:
        return 0.0, "#DIV/0!"
    except Exception:
        return 0.0, "#ERROR!"


def _flatten_metric_nodes(nodes: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []

    def walk(n: dict[str, Any]) -> None:
        out.append(n)
        for c in list(n.get("children") or []):
            walk(c)

    for n in nodes:
        walk(n)
    return out


def _autosize_worksheet(ws) -> None:
    for column_cells in ws.columns:
        max_len = 0
        letter = column_cells[0].column_letter
        for cell in column_cells:
            max_len = max(max_len, len(str(cell.value or "")))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 70)


def _append_org_product_output_export_sheet(ws: Any, rows: list[dict[str, Any]], entity_code: str = "") -> None:
    ws.append(list(ORG_PRODUCT_OUTPUT_EXPORT_HEADERS))
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True) if hasattr(cell, "font") else Font(bold=True)
    for row in rows:
        months = list(row.get("months") or [])
        ws.append(
            [
                str(row.get("levelLabel") or ""),
                str(row.get("nature") or ""),
                str(row.get("code") or ""),
                str(row.get("name") or ""),
                *[("" if v is None else float(v)) for v in (months + [None] * 12)[:12]],
                "" if row.get("annual") is None else float(row.get("annual")),
            ]
        )


def _days_in_year(year: int) -> int:
    y = int(year)
    return 366 if calendar.isleap(y) else 365


def _days_in_month(year: int, month: int) -> int:
    return calendar.monthrange(int(year), int(month))[1]


def _is_rate_like_nature(nature: str) -> bool:
    n = str(nature or "").strip()
    if n in {"收入", "支出", "利润", "资产余额", "负债余额", "资产日均", "负债日均"}:
        return False
    return ("率" in n) or ("占比" in n) or ("比例" in n)


VALID_ANNUAL_AGG_RULES = frozenset({"SUM", "AVG", "LAST", "WGT", "CALC"})


def _normalize_annual_agg_rule(raw: Any) -> str:
    text = str(raw or "").strip().upper()
    if text in VALID_ANNUAL_AGG_RULES:
        return text
    return ""


def _should_annual_recompute_via_formula(
    nature: str,
    formula: str,
    annual_agg_rule: str = "",
) -> bool:
    rule = str(annual_agg_rule or "").strip().upper()
    if rule == "CALC":
        return bool(str(formula or "").strip())
    return bool(str(formula or "").strip()) and _is_rate_like_nature(nature)


def _resolve_annual_formula_recompute(
    *,
    nature: str,
    formula: str,
    formula_budget_annual: str = "",
    formula_forecast_annual: str = "",
    annual_agg_rule: str = "",
) -> tuple[bool, str]:
    rule = str(annual_agg_rule or "").strip().upper()
    if rule == "CALC":
        annual_formula = str(formula_budget_annual or formula or "").strip()
        return bool(annual_formula), annual_formula
    annual_formula = (
        str(formula_budget_annual or "").strip()
        or str(formula_forecast_annual or "").strip()
    )
    if annual_formula:
        return True, annual_formula
    if _should_annual_recompute_via_formula(nature, formula, annual_agg_rule):
        return True, str(formula or "").strip()
    return False, ""


def _should_use_vertical_rollup_annual(meta: dict[str, Any]) -> bool:
    """全年列：仅当无年规则、无 CALC/年公式时，才用纵向汇总子节点。"""
    if not _normalize_rollup_flag(meta.get("vertical_rollup")):
        return False
    annual_agg_rule = str(meta.get("annual_agg_rule") or "").strip().upper()
    if annual_agg_rule in VALID_ANNUAL_AGG_RULES:
        return False
    if str(meta.get("formula_budget_annual") or "").strip() or str(
        meta.get("formula_forecast_annual") or ""
    ).strip():
        return False
    use_formula, _annual_formula = _resolve_annual_formula_recompute(
        nature=str(meta.get("nature") or ""),
        formula=str(meta.get("formula") or "").strip(),
        formula_budget_annual=str(meta.get("formula_budget_annual") or "").strip(),
        formula_forecast_annual=str(meta.get("formula_forecast_annual") or "").strip(),
        annual_agg_rule=annual_agg_rule,
    )
    return not use_formula


def _annual_summary_by_nature(
    nature: str, months: list[float | None], year: int = 2026
) -> float | None:
    """月序列 → 全年列（阶段 1：按科目性质默认规则）。"""
    n = str(nature or "").strip()
    if not months or len(months) != 12:
        return None
    if n in {"资产余额", "负债余额"}:
        v = months[11]
        return float(v) if v is not None else None
    if n in {"资产日均", "负债日均"}:
        total = 0.0
        weight = 0
        for idx, val in enumerate(months):
            if val is None:
                continue
            d = _days_in_month(year, idx + 1)
            total += float(val) * d
            weight += d
        return float(total / weight) if weight > 0 else None
    if n in {"收入", "支出", "利润"}:
        xs = [float(x) for x in months if x is not None]
        return float(sum(xs)) if xs else None
    if _is_rate_like_nature(n):
        v = months[11]
        return float(v) if v is not None else None
    xs = [float(x) for x in months if x is not None]
    return float(sum(xs) / len(xs)) if xs else None


def _annual_method_label(
    nature: str,
    formula: str,
    annual_agg_rule: str = "",
) -> str:
    n = str(nature or "").strip()
    rule = str(annual_agg_rule or "").strip().upper()
    if rule == "CALC":
        return "按公式重算（CALC：引用项为全年口径）"
    if _should_annual_recompute_via_formula(n, formula, annual_agg_rule):
        return "按公式重算（引用项为全年口径）"
    if n in {"资产余额", "负债余额"}:
        return "取 12 月"
    if n in {"资产日均", "负债日均"}:
        return "按当月天数加权全年日均"
    if n in {"收入", "支出", "利润"}:
        return "12 个月合计"
    if _is_rate_like_nature(n):
        return "取 12 月（未配公式时）"
    return "12 个月算术平均"


def _merge_data_entry_drafts(
    draft_items: list[dict[str, Any]],
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    """
    Returns (merged_payload, conflicts)
    - merged_payload shape aligns with DataEntrySavePayload (entity_code/entity_name/year/table_id/table_name/month_index/metrics)
    - conflicts entries include metric_code/metric_name/key/user_values
    """

    def pick_latest_non_empty(candidates: list[tuple[str, str, str]]) -> str:
        # (updated_at, user_id, value)
        non_empty = [(ts, uid, v) for ts, uid, v in candidates if _is_non_empty(v)]
        if not non_empty:
            return ""
        non_empty.sort(key=lambda x: x[0], reverse=True)
        return non_empty[0][2]

    def collect_value_candidates(
        value_by_user: dict[str, tuple[str, str]]
    ) -> tuple[str, dict[str, str], bool]:
        # value_by_user[user_id] = (updated_at, value)
        candidates = [(ts, uid, v) for uid, (ts, v) in value_by_user.items()]
        chosen = pick_latest_non_empty(candidates)
        non_empty_values = sorted({str(v).strip() for _, _, v in candidates if _is_non_empty(v)})
        has_conflict = len(non_empty_values) > 1
        user_values = {uid: v for uid, (_, v) in value_by_user.items() if _is_non_empty(v)}
        return chosen, user_values, has_conflict

    entity_code = ""
    entity_name = ""
    year = 0
    table_id = ""
    table_name = ""
    month_index: int | None = None

    metric_meta: dict[str, dict[str, Any]] = {}
    cell_values: dict[str, dict[str, dict[str, tuple[str, str]]]] = {}
    # metric_key -> cell_key -> user_id -> (updated_at, value)

    for item in draft_items:
        user_id = str(item.get("user_id") or "")
        updated_at = str(item.get("updated_at") or "")
        payload = item.get("payload") if isinstance(item.get("payload"), dict) else None
        if not payload:
            continue
        entity_code = entity_code or _norm_text(payload.get("entity_code"))
        entity_name = entity_name or _norm_text(payload.get("entity_name"))
        try:
            year = year or int(payload.get("year") or 0)
        except Exception:
            year = year or 0
        table_id = table_id or _norm_text(payload.get("table_id"))
        table_name = table_name or _norm_text(payload.get("table_name"))
        if month_index is None and payload.get("month_index") is not None:
            try:
                month_index = int(payload.get("month_index"))
            except Exception:
                month_index = None

        metrics = payload.get("metrics")
        if not isinstance(metrics, list):
            continue
        for m in metrics:
            if not isinstance(m, dict):
                continue
            metric_id = _norm_text(m.get("metric_id"))
            metric_code = _norm_text(m.get("metric_code"))
            metric_key = metric_id or metric_code
            if not metric_key:
                continue
            meta = metric_meta.setdefault(
                metric_key,
                {
                    "metric_id": metric_id,
                    "metric_code": metric_code,
                    "metric_name": _norm_text(m.get("metric_name")),
                    "levelLabel": _norm_text(m.get("levelLabel")),
                    "nature": _norm_text(m.get("nature")),
                },
            )
            if not meta.get("metric_id") and metric_id:
                meta["metric_id"] = metric_id
            if not meta.get("metric_code") and metric_code:
                meta["metric_code"] = metric_code
            if not meta.get("metric_name") and _norm_text(m.get("metric_name")):
                meta["metric_name"] = _norm_text(m.get("metric_name"))

            values = m.get("values") if isinstance(m.get("values"), dict) else {}
            metric_cells = cell_values.setdefault(metric_key, {})
            for k in ("prev_actual", "prev_budget", "prev_forecast", "year_forecast"):
                v = _norm_text(values.get(k))
                cell = metric_cells.setdefault(k, {})
                cell[user_id] = (updated_at, v)
            months = values.get("months") if isinstance(values.get("months"), dict) else {}
            for mk, mv in months.items():
                if not isinstance(mk, str):
                    continue
                v = _norm_text(mv)
                cell = metric_cells.setdefault(f"months.{mk}", {})
                cell[user_id] = (updated_at, v)

    conflicts: list[dict[str, Any]] = []
    merged_metrics: list[dict[str, Any]] = []
    for metric_key, meta in metric_meta.items():
        metric_cells = cell_values.get(metric_key, {})
        merged_values: dict[str, Any] = {
            "prev_actual": "",
            "prev_budget": "",
            "prev_forecast": "",
            "year_forecast": "",
            "months": {},
        }
        for k in ("prev_actual", "prev_budget", "prev_forecast", "year_forecast"):
            chosen, user_values, has_conflict = collect_value_candidates(metric_cells.get(k, {}))
            merged_values[k] = chosen
            if has_conflict:
                conflicts.append(
                    {
                        "metric_code": meta.get("metric_code") or meta.get("metric_id") or metric_key,
                        "metric_name": meta.get("metric_name") or "",
                        "key": k,
                        "user_values": user_values,
                    }
                )
        months_out: dict[str, str] = {}
        for cell_key, per_user in metric_cells.items():
            if not isinstance(cell_key, str) or not cell_key.startswith("months."):
                continue
            mk = cell_key.split(".", 1)[1]
            chosen, user_values, has_conflict = collect_value_candidates(per_user)
            if _is_non_empty(chosen):
                months_out[mk] = chosen
            if has_conflict:
                conflicts.append(
                    {
                        "metric_code": meta.get("metric_code") or meta.get("metric_id") or metric_key,
                        "metric_name": meta.get("metric_name") or "",
                        "key": f"months.{mk}",
                        "user_values": user_values,
                    }
                )
        merged_values["months"] = months_out
        has_any = (
            _is_non_empty(merged_values.get("prev_actual"))
            or _is_non_empty(merged_values.get("prev_budget"))
            or _is_non_empty(merged_values.get("year_forecast"))
            or any(_is_non_empty(v) for v in months_out.values())
        )
        if not has_any:
            continue
        merged_metrics.append(
            {
                "metric_id": meta.get("metric_id") or "",
                "metric_code": meta.get("metric_code") or "",
                "metric_name": meta.get("metric_name") or "",
                "levelLabel": meta.get("levelLabel") or "",
                "nature": meta.get("nature") or "",
                "values": merged_values,
            }
        )

    merged_payload = {
        "entity_code": entity_code,
        "entity_name": entity_name,
        "year": year,
        "month_index": month_index,
        "table_id": table_id,
        "table_name": table_name,
        "metrics": merged_metrics,
    }
    return merged_payload, conflicts


def _ensure_org_product_tree_table(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS org_product_tree_snapshot (
            id INTEGER PRIMARY KEY CHECK (id = 1),
            payload_json LONGTEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )


def ensure_org_product_schema(conn: sqlite3.Connection) -> None:
    _ensure_org_product_tree_table(conn)
    # _ensure_metrics_table removed (retired no-op)
    from app.services.org_product_metric_runtime_snapshot import ensure_org_product_metric_table_payload

    ensure_org_product_metric_table_payload(conn)
    _ensure_metric_table_catalog(conn)
    _seed_metric_table_catalog(conn)
    _ensure_data_entry_snapshot_table(conn)
    _ensure_data_entry_snapshot_table_v2(conn)
    _ensure_data_entry_draft_table(conn)
    _ensure_org_product_output_snapshot_table(conn)



class MetricBatchImportTablePayload(BaseModel):
    table_name: str
    row_count: int


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        text = str(value).strip()
    elif isinstance(value, int):
        text = str(value).strip()
    elif isinstance(value, float):
        text = str(int(value)).strip() if value.is_integer() else str(value).strip()
    else:
        text = str(value).strip()
    text = text.replace("\ufeff", "").replace("\u3000", " ").strip()
    return unicodedata.normalize("NFKC", text)


def _normalize_metric_code(entity_code: str, raw_code: Any) -> str:
    code = _normalize_text(raw_code).upper().replace(" ", "")
    owner = _normalize_text(entity_code).upper()
    if not code:
        return ""
    return _normalize_runtime_metric_code(owner, code)


def _compact_runtime_metric_code(code: Any) -> str:
    return compact_org_product_metric_code(code)


def _normalize_runtime_metric_code(entity_code: str, raw_code: Any) -> str:
    return normalize_org_product_metric_code(entity_code, raw_code)


def _normalize_rollup_flag(value: Any) -> int:
    if value is None:
        return 0
    if isinstance(value, bool):
        return 1 if value else 0
    if isinstance(value, (int, float)):
        return 0 if float(value) == 0 else 1
    text = _normalize_text(value).lower()
    if not text:
        return 0
    if text in {"0", "false", "否", "不", "no", "n", "不汇总", "无需汇总"}:
        return 0
    if text in {"1", "true", "是", "y", "yes", "汇总", "需要汇总"}:
        return 1
    return 1


def _rollup_flag_label(value: Any) -> str:
    return "是" if _normalize_rollup_flag(value) else "否"


def _derive_metric_logic_code(entity_code: str, code: str, raw_logic_code: Any = "") -> str:
    explicit = _normalize_text(raw_logic_code).upper().replace(" ", "")
    if explicit:
        return explicit
    metric_code = _normalize_text(code).upper().replace(" ", "")
    owner = _normalize_text(entity_code).upper().replace(" ", "")
    if not metric_code:
        return ""
    if owner and metric_code == owner:
        return ""
    if owner and metric_code.startswith(f"{owner}."):
        return metric_code[len(owner) + 1 :]
    if owner and metric_code.startswith(owner) and len(metric_code) > len(owner):
        rest = metric_code[len(owner) :]
        return rest[1:] if rest.startswith(".") else rest
    first, sep, rest = metric_code.partition(".")
    if sep and re.fullmatch(r"[A-Z]{1,2}\d{0,2}", first):
        return rest
    return metric_code


def _drop_legacy_runtime_identity_fields(item: dict[str, Any]) -> dict[str, Any]:
    item.pop("mapping_status", None)
    item.pop("metric_node_code", None)
    item.pop("data_acct_code", None)
    return item


def _ensure_metric_node_has_name(item: dict[str, Any], entity_code: str) -> None:
    """Auto-fill empty node names so that save never fails due to blank names.

    Rules (same as the sync-layer ``_node_name`` logic):
    - If ``name`` is present and non-empty → keep it.
    - If ``name`` is empty but ``code`` exists → use ``code`` as the name.
    - If both are empty → derive from first child name → ``f"[{child_name}的父节点]"``.
    - Final fallback → ``"(未命名)"``.
    """
    name = _normalize_text(item.get("name"))
    if name:
        return
    code = _normalize_text(item.get("code"))
    if code:
        item["name"] = code
        return
    children = item.get("children")
    if isinstance(children, list):
        for child in children:
            if not isinstance(child, dict):
                continue
            child_name = _normalize_text(child.get("name"))
            if child_name:
                item["name"] = f"[{child_name}的父节点]"
                return
    item["name"] = "(未命名)"


def _sanitize_metric_nodes_for_save(entity_code: str, nodes: list[MetricNodePayload]) -> list[dict[str, Any]]:
    sanitized: list[dict[str, Any]] = []
    for node in nodes:
        item = node.model_dump()
        item["children"] = _sanitize_metric_nodes_for_save(entity_code, list(node.children or []))
        _drop_legacy_runtime_identity_fields(item)
        item["code"] = _normalize_metric_code(entity_code, item.get("code"))
        item["horizontal_rollup"] = _normalize_rollup_flag(item.get("horizontal_rollup"))
        item["vertical_rollup"] = _normalize_rollup_flag(item.get("vertical_rollup"))
        item["logic_code"] = _derive_metric_logic_code(entity_code, item.get("code"), item.get("logic_code"))
        _ensure_metric_node_has_name(item, entity_code)
        sanitized.append(item)
    return dedupe_org_product_metric_payload_nodes(sanitized)


def _sanitize_metric_node_dicts_for_response(entity_code: str, nodes: list[dict[str, Any]]) -> list[dict[str, Any]]:
    sanitized: list[dict[str, Any]] = []
    for node in nodes:
        if not isinstance(node, dict):
            continue
        item = dict(node)
        children = item.get("children") if isinstance(item.get("children"), list) else []
        item["children"] = _sanitize_metric_node_dicts_for_response(
            entity_code,
            [child for child in children if isinstance(child, dict)],
        )
        _drop_legacy_runtime_identity_fields(item)
        item["code"] = _normalize_metric_code(entity_code, item.get("code"))
        item["horizontal_rollup"] = _normalize_rollup_flag(item.get("horizontal_rollup"))
        item["vertical_rollup"] = _normalize_rollup_flag(item.get("vertical_rollup"))
        item["logic_code"] = _derive_metric_logic_code(entity_code, item.get("code"), item.get("logic_code"))
        sanitized.append(item)
    return dedupe_org_product_metric_payload_nodes(sanitized)


def _sanitize_org_product_output_entity_for_snapshot(entity: dict[str, Any]) -> dict[str, Any]:
    cleaned = dict(entity)
    rows: list[dict[str, Any]] = []
    allowed_row_keys = {
        "id",
        "levelLabel",
        "nature",
        "code",
        "name",
        "value_type",
        "formula",
        "months",
        "month_errors",
        "annual",
        "annual_method",
    }
    for row in list(cleaned.get("rows") or []):
        if not isinstance(row, dict):
            continue
        item = {key: row.get(key) for key in allowed_row_keys if key in row}
        rows.append(item)
    cleaned["rows"] = rows
    return cleaned


def _data_entry_metric_dedupe_key(entity_code: str, row: dict[str, Any]) -> str:
    code = _normalize_runtime_metric_code(entity_code, row.get("metric_code"))
    derived = _derive_metric_logic_code(entity_code, code, "")
    owner = _normalize_text(entity_code).upper()
    if owner and code.startswith(owner):
        compact = code.replace(".", "")
        return f"code:{compact}"
    if derived:
        return f"code:{owner}{derived.replace('.', '')}"
    metric_id = _normalize_text(row.get("metric_id"))
    return f"id:{metric_id}" if metric_id else ""


def _data_entry_row_rank(row: dict[str, Any]) -> int:
    metric_id = _normalize_text(row.get("metric_id"))
    if metric_id.startswith("canonical-"):
        return 3
    return 1


def _merge_data_entry_metric_row(target: dict[str, Any], source: dict[str, Any]) -> None:
    for key in (
        "metric_id",
        "metric_code",
        "metric_name",
        "levelLabel",
        "nature",
    ):
        if target.get(key) in (None, "") and source.get(key) not in (None, ""):
            target[key] = source[key]
    target_values = target.get("values") if isinstance(target.get("values"), dict) else {}
    source_values = source.get("values") if isinstance(source.get("values"), dict) else {}
    merged_values = dict(target_values)
    for key in ("prev_actual", "prev_budget", "prev_forecast", "year_forecast"):
        if merged_values.get(key) in (None, "") and source_values.get(key) not in (None, ""):
            merged_values[key] = source_values[key]
    target_months = merged_values.get("months") if isinstance(merged_values.get("months"), dict) else {}
    source_months = source_values.get("months") if isinstance(source_values.get("months"), dict) else {}
    merged_months = dict(target_months)
    for key, value in source_months.items():
        if merged_months.get(key) in (None, "") and value not in (None, ""):
            merged_months[key] = value
    merged_values["months"] = merged_months
    target["values"] = merged_values


def _dedupe_data_entry_metric_rows(entity_code: str, rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    deduped: list[dict[str, Any]] = []
    seen: dict[str, dict[str, Any]] = {}
    for row in rows:
        key = _data_entry_metric_dedupe_key(entity_code, row)
        if key and key in seen:
            existing = seen[key]
            if _data_entry_row_rank(row) > _data_entry_row_rank(existing):
                _merge_data_entry_metric_row(row, existing)
                idx = deduped.index(existing)
                deduped[idx] = row
                seen[key] = row
            else:
                _merge_data_entry_metric_row(existing, row)
            continue
        deduped.append(row)
        if key:
            seen[key] = row
    return deduped


def _sanitize_data_entry_payload_mapping_refs(conn: sqlite3.Connection, payload: dict[str, Any]) -> dict[str, Any]:
    cleaned = dict(payload)
    entity_code = _normalize_text(cleaned.get("entity_code"))
    table_name = _normalize_text(cleaned.get("table_name"))
    metric_by_key: dict[str, dict[str, Any]] = {}

    if entity_code and table_name:
        table_obj = load_org_product_metric_payload_from_runtime_tree(
            conn,
            entity_code=entity_code,
            table_name=table_name,
        )
        if table_obj:
            metrics = table_obj.get("metrics") if isinstance(table_obj, dict) else []
            if isinstance(metrics, list):
                for metric in _flatten_metric_nodes([m for m in metrics if isinstance(m, dict)]):
                    metric_id = _normalize_text(metric.get("id"))
                    metric_code = _normalize_text(metric.get("code"))
                    if metric_id:
                        metric_by_key[metric_id] = metric
                    if metric_code:
                        metric_by_key[metric_code] = metric
                        metric_by_key[_compact_runtime_metric_code(metric_code)] = metric

    rows: list[dict[str, Any]] = []
    for row in list(cleaned.get("metrics") or []):
        if not isinstance(row, dict):
            continue
        item = dict(row)
        metric_id = _normalize_text(item.get("metric_id"))
        metric_code = _normalize_text(item.get("metric_code"))
        normalized_metric_code = _normalize_runtime_metric_code(entity_code, metric_code)
        meta = (
            metric_by_key.get(metric_code)
            or metric_by_key.get(normalized_metric_code)
            or metric_by_key.get(_compact_runtime_metric_code(metric_code))
            or metric_by_key.get(metric_id)
        )
        if meta:
            item["metric_code"] = _normalize_text(meta.get("code")) or metric_code
            item["metric_name"] = _normalize_text(meta.get("name")) or _normalize_text(item.get("metric_name"))
        elif normalized_metric_code:
            item["metric_code"] = normalized_metric_code
        _drop_legacy_runtime_identity_fields(item)
        rows.append(item)
    cleaned["metrics"] = _dedupe_data_entry_metric_rows(entity_code, rows)
    return cleaned


def _sanitize_data_entry_payload_for_response(
    conn: sqlite3.Connection,
    payload: Any,
    *,
    entity_code: Any = "",
    table_name: Any = "",
) -> Any:
    if not isinstance(payload, dict):
        return payload
    item = dict(payload)
    if entity_code and not _normalize_text(item.get("entity_code")):
        item["entity_code"] = _normalize_text(entity_code)
    if table_name and not _normalize_text(item.get("table_name")):
        item["table_name"] = _normalize_text(table_name)
    return _sanitize_data_entry_payload_mapping_refs(conn, item)


def _sanitize_output_payload_for_response(payload: Any, *, entity_code: Any = "", table_name: Any = "") -> Any:
    if not isinstance(payload, dict):
        return payload
    item = dict(payload)
    if entity_code and not _normalize_text(item.get("entity_code")):
        item["entity_code"] = _normalize_text(entity_code)
    if table_name and not _normalize_text(item.get("table_name")):
        item["table_name"] = _normalize_text(table_name)
    return _sanitize_org_product_output_entity_for_snapshot(item)


def _data_entry_export_columns(year: int, forecast_month: int) -> list[tuple[str, str]]:
    prev_yy = (int(year) - 1) % 100
    yy = int(year) % 100
    month = max(1, min(12, int(forecast_month or 1)))
    columns: list[tuple[str, str]] = [
        ("levelLabel", "科目层级"),
        ("nature", "科目性质"),
        ("metric_code", "科目代码"),
        ("metric_name", "科目名称"),
        ("prev_actual", f"{prev_yy}年实际"),
        ("prev_budget", f"{yy}年预算"),
        ("prev_forecast", f"{yy}年预测"),
    ]
    for mi in range(1, month + 1):
        columns.append((f"a{mi}", f"{yy}年{mi}月实际"))
    for mi in range(month + 1, 13):
        columns.append((f"f{mi}", f"{yy}年{mi}月预测"))
    return columns


def _get_data_entry_metric_value(values: dict[str, Any], key: str) -> str:
    if key in {"prev_actual", "prev_budget", "prev_forecast", "year_forecast"}:
        return _normalize_text(values.get(key))
    months = values.get("months") if isinstance(values.get("months"), dict) else {}
    return _normalize_text(months.get(key))


def _build_data_entry_export_workbook(payload: dict[str, Any]) -> Workbook:
    year = int(payload.get("year") or datetime.now(timezone.utc).year)
    forecast_month = int(payload.get("month_index") or 1)
    entity_code = _normalize_text(payload.get("entity_code"))
    entity_name = _normalize_text(payload.get("entity_name"))
    table_name = _normalize_text(payload.get("table_name")) or DEFAULT_METRIC_TABLE_NAME
    columns = _data_entry_export_columns(year, forecast_month)
    wb = Workbook()
    ws = wb.active
    ws.title = _unique_sheet_title(set(), f"{entity_code}{entity_name}{table_name}" or "data_entry")
    ws.append([label for _key, label in columns])
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True) if hasattr(cell, "font") else Font(bold=True)

    for metric in list(payload.get("metrics") or []):
        if not isinstance(metric, dict):
            continue
        values = metric.get("values") if isinstance(metric.get("values"), dict) else {}
        row: list[str] = []
        for key, _label in columns:
            if key in {"prev_actual", "prev_budget", "prev_forecast", "year_forecast"} or re.fullmatch(r"[af]\d{1,2}", key):
                row.append(_get_data_entry_metric_value(values, key))
            else:
                row.append(_normalize_text(metric.get(key)))
        ws.append(row)

    widths = {
        "A": 12,
        "B": 12,
        "C": 18,
        "D": 32,
        "E": 24,
        "F": 20,
        "G": 20,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    return wb


def _normalize_entry_granularity(value: Any) -> str:
    text = _normalize_text(value)
    if not text:
        return "monthly"
    compact = re.sub(r"\s+", "", text).lower()
    if compact in ENTRY_GRANULARITY_ANNUAL or text in ENTRY_GRANULARITY_ANNUAL:
        return "annual"
    if compact in ENTRY_GRANULARITY_MONTHLY or text in ENTRY_GRANULARITY_MONTHLY:
        return "monthly"
    if "按年录入" in text or "按年" in text or "仅年度" in text or "仅全年" in text or text == "年度":
        return "annual"
    return "monthly"


def _entry_granularity_label(value: Any) -> str:
    return "按年录入" if _normalize_entry_granularity(value) == "annual" else "按月录入"


def _normalize_metric_value_type(value: Any, nature: Any = "") -> str:
    text = _normalize_text(value)
    if text in {"金额", "百分比", "户数"}:
        return text
    nature_text = _normalize_nature(nature)
    if nature_text in {"比例", "百分比", "率"}:
        return "百分比"
    if nature_text in {"户数", "人数", "笔数"}:
        return "户数"
    return "金额"


def _normalize_allow_manual_entry(value: Any, default: int = 1) -> int:
    if value is None:
        return 1 if default else 0
    if isinstance(value, bool):
        return 1 if value else 0
    if isinstance(value, (int, float)):
        return 0 if int(value) == 0 else 1
    text = _normalize_text(value).lower()
    if not text:
        return 1 if default else 0
    if text in {"0", "false", "否", "不允许", "no", "n"}:
        return 0
    if text in {"1", "true", "是", "允许", "yes", "y"}:
        return 1
    return 1 if default else 0


def _allow_manual_entry_label(value: Any) -> str:
    return "允许" if _normalize_allow_manual_entry(value) else "不允许"


def _normalize_nature(value: Any) -> str:
    text = _normalize_text(value)
    if text in {"其它", "其他"}:
        return "其他"
    if text == "住处":
        return "支出"
    return text


def _apply_profit_nature(name: str, nature: str) -> str:
    normalized_name = re.sub(r"\s+", "", name or "")
    if normalized_name in PROFIT_NAMES:
        return "利润"
    return nature


def _extract_owner_code(sheet_name: str, rows: list[dict[str, Any]]) -> str | None:
    normalized_name = re.sub(r"\s+", "", sheet_name or "")
    matched = re.match(r"^([A-Z]+[0-9]*)", normalized_name)
    if matched:
        return matched.group(1)
    for row in rows:
        code = _normalize_text(row.get("code"))
        code_match = re.match(r"^([A-Z]+[0-9]*)", code)
        if code_match:
            code_text = code_match.group(1)
            if len(code_text) <= 3:
                return code_text
    return None


def _parse_metric_workbook(path: Path) -> dict[str, list[dict[str, Any]]]:
    table_items = _parse_metric_workbook_tables(path)
    result: dict[str, list[dict[str, Any]]] = {}
    for owner_code, tables in table_items.items():
        first_table = next((table for table in tables if table.get("metrics")), None)
        if first_table:
            result[owner_code] = list(first_table.get("metrics") or [])
    return result


def _metric_table_name_from_sheet(sheet_name: str, owner_code: str) -> str:
    compact_sheet = _normalize_metric_sheet_title_key(sheet_name)
    owner = _normalize_text(owner_code).upper()
    suffix = compact_sheet[len(owner) :] if owner and compact_sheet.upper().startswith(owner) else ""
    if owner in {"AA", "AB"} and suffix:
        return _canonical_import_table_name(suffix)
    if suffix.endswith("表"):
        return _canonical_import_table_name(suffix)
    return DEFAULT_METRIC_TABLE_NAME


def _parse_metric_workbook_tables(path: Path) -> dict[str, list[dict[str, Any]]]:
    if not path.is_file():
        raise FileNotFoundError(str(path))

    workbook = load_workbook(path, data_only=True)
    result: dict[str, list[dict[str, Any]]] = {}

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        header_row_idx, header_map, _header_mode, _header_source = _find_header_row(ws)
        if header_row_idx is None:
            continue
        level_col = header_map.get("科目层级")
        nature_col = header_map.get("科目性质")
        code_col = header_map.get("科目代码")
        name_col = header_map.get("科目名称")
        formula_col = header_map.get("取数公式")
        if not code_col or not name_col:
            continue

        roots: list[dict[str, Any]] = []
        stack: dict[int, dict[str, Any]] = {}
        row_count = 0

        for row_idx in range(header_row_idx + 1, _sheet_scan_row_limit(ws, header_row_idx) + 1):
            level_label = _normalize_level_label(ws.cell(row_idx, level_col).value) if level_col else ""
            nature = _normalize_nature(ws.cell(row_idx, nature_col).value) if nature_col else ""
            code = _normalize_text(ws.cell(row_idx, code_col).value)
            name = _normalize_text(ws.cell(row_idx, name_col).value)
            formula = _normalize_text(ws.cell(row_idx, formula_col).value) if formula_col else ""
            if not any([level_label, nature, code, name, formula]):
                continue
            if not code or not name:
                continue
            level_label = level_label or _infer_level_label_from_code(code)
            rank = LEVEL_RANK.get(level_label)
            if not rank:
                continue
            nature = _apply_profit_nature(name, nature or "其他")

            node = {
                "id": f"metric-{sheet_name}-{row_idx}",
                "levelLabel": level_label,
                "nature": nature,
                "code": code,
                "name": name,
                "note": "",
                "formula": formula,
                "children": [],
            }
            if rank == 1 or stack.get(rank - 1) is None:
                roots.append(node)
            else:
                stack[rank - 1]["children"].append(node)

            stack[rank] = node
            for key in list(stack.keys()):
                if key > rank:
                    del stack[key]
            row_count += 1

        owner_code = _extract_owner_code(sheet_name, roots)
        if owner_code and roots and row_count > 0:
            table_name = _metric_table_name_from_sheet(sheet_name, owner_code)
            result.setdefault(owner_code, []).append(
                {
                    "id": f"table-{table_name}",
                    "name": table_name,
                    "metrics": roots,
                }
            )

    return result


def _resolve_metric_formula_for_month(meta: dict[str, Any], month_idx: int, rolling_month: int) -> str:
    legacy = _normalize_text(meta.get("formula"))
    actual = _normalize_text(meta.get("formula_actual"))
    forecast = _normalize_text(meta.get("formula_forecast"))
    has_month_specific_formula = bool(actual or forecast)
    m = max(1, min(12, int(rolling_month)))
    if month_idx <= m:
        if has_month_specific_formula:
            return actual
        return legacy
    if has_month_specific_formula:
        return forecast
    return legacy


def _build_metric_rows(nodes: list[dict[str, Any]], rows: list[dict[str, str]], entity_code: str = "") -> None:
    for node in nodes:
        rows.append(
            {
                "levelLabel": _normalize_text(node.get("levelLabel")),
                "nature": _normalize_nature(node.get("nature")),
                "code": _normalize_text(node.get("code")),
                "name": _normalize_text(node.get("name")),
                "value_type": _normalize_metric_value_type(node.get("value_type"), node.get("nature")),
                "allow_manual_entry": _normalize_allow_manual_entry(node.get("allow_manual_entry"), 1),
                "note": _normalize_text(node.get("note")),
                "formula": _normalize_text(node.get("formula")),
                "formula_budget_annual": _normalize_text(node.get("formula_budget_annual")),
                "formula_forecast_annual": _normalize_text(node.get("formula_forecast_annual")),
                "formula_actual": _normalize_text(node.get("formula_actual")),
                "formula_forecast": _normalize_text(node.get("formula_forecast")),
                "formula_note": _normalize_text(node.get("formula_note")),
                "entry_granularity": _normalize_entry_granularity(node.get("entry_granularity")),
                "horizontal_rollup": _normalize_rollup_flag(node.get("horizontal_rollup")),
                "vertical_rollup": _normalize_rollup_flag(node.get("vertical_rollup")),
                "logic_code": _derive_metric_logic_code(entity_code, node.get("code"), node.get("logic_code")),
                "annual_agg_rule": _normalize_annual_agg_rule(node.get("annual_agg_rule")),
            }
        )
        _build_metric_rows(list(node.get("children") or []), rows, entity_code)


def _normalize_level_label(value: Any) -> str:
    text = _normalize_text(value)
    if not text:
        return ""
    if text in LEVEL_RANK:
        return text
    compact = re.sub(r"\s+", "", text)
    digit_map = {"1": "一级", "2": "二级", "3": "三级", "4": "四级", "5": "五级", "6": "六级"}
    digit_match = re.fullmatch(r"(\d)[级層层]?", compact)
    if digit_match:
        return digit_map.get(digit_match.group(1), "")
    for label in LEVEL_RANK:
        if label in compact:
            return label
    return text


def _normalize_metric_sheet_title_key(title: str) -> str:
    key = re.sub(r"[\s_\u200b-\u200d\ufeff]+", "", _normalize_text(title))
    key = key.replace("（", "(").replace("）", ")")
    for old, new in METRIC_SHEET_TABLE_ALIASES.items():
        key = key.replace(old, new)
    return key


def _canonical_metric_table_key(table_name: str) -> str:
    key = re.sub(r"[\s_]+", "", _normalize_text(table_name))
    return key.replace("（", "(").replace("）", ")")


def _canonical_import_table_name(table_name: str, *, known_names: list[str] | None = None) -> str:
    name = _normalize_text(table_name)
    aliased = METRIC_SHEET_TABLE_ALIASES.get(name, name)
    key = _canonical_metric_table_key(aliased)
    display_names = known_names or [
        "业务状况表",
        "损益表",
        "资产负债表（余额）",
        "资产负债表（日均）",
        "资产质量表",
        "利息净收入表",
    ]
    for display in display_names:
        if _canonical_metric_table_key(display) == key:
            return display
    return aliased


def _apply_sheet_title_table_aliases(title: str, entity_code: str) -> str:
    code = _normalize_text(entity_code).upper()
    if code and title.upper().startswith(code):
        suffix = title[len(code) :]
    else:
        suffix = title
    for old, new in METRIC_SHEET_TITLE_TABLE_ALIASES.items():
        old_key = old.replace("（", "(").replace("）", ")")
        new_key = new.replace("（", "(").replace("）", ")")
        if suffix == old_key or suffix == old:
            suffix = new_key
            break
    return f"{code}{suffix}" if code else suffix


def _infer_level_label_from_code(code: str) -> str:
    c = code.strip().upper()
    if not c:
        return ""
    if "." in c:
        segments = [seg for seg in c.split(".") if seg]
        depth = min(max(len(segments), 1), 6)
        return list(LEVEL_RANK.keys())[depth - 1]
    digits = re.sub(r"^[A-Z]{1,3}", "", c)
    digits = re.sub(r"[^0-9]", "", digits)
    if not digits:
        return "一级"
    n = len(digits)
    if n <= 2:
        return "一级"
    if n <= 4:
        return "二级"
    if n <= 6:
        return "三级"
    if n <= 8:
        return "四级"
    if n <= 10:
        return "五级"
    return "六级"


def _ws_cell_value(ws, row_idx: int, col_idx: int) -> Any:
    """读取单元格；优先 _cells、合并格左上角，避免 dimension 过小漏读。"""
    if row_idx < 1 or col_idx < 1:
        return None
    try:
        cells = getattr(ws, "_cells", None)
        if cells:
            cell = cells.get((row_idx, col_idx))
            if cell is not None:
                val = cell.value
                if _normalize_text(val):
                    return val
    except Exception:
        pass
    val = ws.cell(row_idx, col_idx).value
    if _normalize_text(val):
        return val
    merged_cells = getattr(ws, "merged_cells", None)
    if merged_cells is not None:
        for merged_range in merged_cells.ranges:
            if (
                merged_range.min_row <= row_idx <= merged_range.max_row
                and merged_range.min_col <= col_idx <= merged_range.max_col
            ):
                return ws.cell(merged_range.min_row, merged_range.min_col).value
    return val


def _prepare_metric_worksheet(ws) -> None:
    if hasattr(ws, "calculate_dimension"):
        try:
            ws.calculate_dimension()
        except Exception:
            pass


def _worksheet_effective_max_row(ws, header_row_idx: int = 1) -> int:
    dim_row = int(ws.max_row or 0)
    cell_row = 0
    try:
        cells = getattr(ws, "_cells", None)
        if cells:
            cell_row = max((int(pos[0]) for pos in cells.keys()), default=0)
    except Exception:
        cell_row = 0
    return max(dim_row, cell_row, header_row_idx + 1)


def _sheet_scan_col_limit(ws, probe_row: int = 1) -> int:
    limit = max(int(ws.max_column or 0), 40)
    for col_idx in range(limit + 1, limit + 21):
        if _normalize_text(_ws_cell_value(ws, probe_row, col_idx)):
            limit = col_idx
    return limit


def _sheet_scan_row_limit(ws, header_row_idx: int) -> int:
    return max(_worksheet_effective_max_row(ws, header_row_idx), header_row_idx + 500)


def _looks_like_metric_code(entity_code: str, raw: Any) -> bool:
    text = _normalize_metric_code(entity_code, raw)
    owner = _normalize_text(entity_code).upper()
    if not text or re.search(r"[\u4e00-\u9fff]", text):
        return False
    if owner and not text.startswith(owner):
        return False
    if owner and len(text) <= len(owner):
        return False
    return bool(re.fullmatch(r"[A-Z]{1,3}[0-9][A-Z0-9.]*", text))


def _looks_like_metric_name(raw: Any) -> bool:
    text = _normalize_text(raw)
    if not text:
        return False
    if text in LEVEL_RANK or text in NATURE_VALUE_SET:
        return False
    return bool(re.search(r"[\u4e00-\u9fff]{2,}", text))


def _maybe_swap_metric_code_name_columns(
    ws,
    *,
    header_row_idx: int,
    entity_code: str,
    code_col: int | None,
    name_col: int | None,
) -> tuple[int | None, int | None]:
    if not code_col or not name_col or code_col == name_col:
        return code_col, name_col

    normal_score = 0
    swapped_score = 0
    row_limit = min(_sheet_scan_row_limit(ws, header_row_idx), header_row_idx + 80)
    for row_idx in range(header_row_idx + 1, row_limit + 1):
        code_raw = _ws_cell_value(ws, row_idx, code_col)
        name_raw = _ws_cell_value(ws, row_idx, name_col)
        if not _normalize_text(code_raw) and not _normalize_text(name_raw):
            continue
        if _looks_like_metric_code(entity_code, code_raw):
            normal_score += 2
        if _looks_like_metric_name(name_raw):
            normal_score += 1
        if _looks_like_metric_code(entity_code, name_raw):
            swapped_score += 2
        if _looks_like_metric_name(code_raw):
            swapped_score += 1

    if swapped_score >= normal_score + 3:
        return name_col, code_col
    return code_col, name_col


def _balance_sheet_standard_layout_fallback(ws, sheet_title: str = "") -> tuple[int | None, dict[str, int]]:
    """资产负债表常见版式：A层级 B性质 C名称 D代码（与截图一致）。"""
    title_key = _normalize_metric_sheet_title_key(sheet_title)
    if title_key and "资产负债" not in title_key:
        return None, {}
    standard = {"科目层级": 1, "科目性质": 2, "科目名称": 3, "科目代码": 4}
    row1_ok = all(
        _canon_header_label(_normalize_text(_ws_cell_value(ws, 1, col))) == label
        for label, col in standard.items()
    )
    if row1_ok:
        return 1, dict(standard)
    code_probe = _normalize_metric_code(
        "AA",
        _ws_cell_value(ws, 2, standard["科目代码"]),
    )
    if code_probe.startswith("AA") and len(code_probe) > 2:
        return 1, dict(standard)
    return None, {}


def _known_metric_header_labels() -> set[str]:
    return {
        "科目层级",
        "科目性质",
        "科目代码",
        "科目名称",
        "机构及产品指标编码",
        "取数公式",
        "年预算公式",
        "年预测公式",
        "实际月公式",
        "预测月公式",
        "公式说明",
        "录入粒度",
        "数值类型",
        "允许手工录入",
        "横向汇总",
        "纵向汇总",
        "逻辑码",
        "规则",
    }


def _header_map_scan_rows(ws, max_scan_row: int = 12) -> tuple[int | None, dict[str, int]]:
    """逐行扫描表头（合并格 + iter_rows），用于 hits 漏检时的兜底。"""
    full_headers = {"科目层级", "科目性质", "科目代码", "科目名称"}
    known_headers = _known_metric_header_labels()
    scan_col = _sheet_scan_col_limit(ws)
    for row_idx in range(1, max_scan_row + 1):
        row_map: dict[str, int] = {}
        for col_idx in range(1, scan_col + 1):
            raw = _normalize_text(_ws_cell_value(ws, row_idx, col_idx))
            if not raw:
                continue
            canon = _canon_header_label(raw)
            if canon in known_headers:
                row_map[canon] = col_idx
        if full_headers.issubset(row_map):
            return row_idx, row_map
    return None, {}


def _strip_display_only_header_suffix(text: str) -> str:
    out = str(text or "").strip()
    for suffix in METRIC_DISPLAY_ONLY_HEADER_SUFFIXES:
        out = out.replace(suffix, "")
    return out.strip()


def _canon_header_label(raw: str) -> str:
    text = _strip_display_only_header_suffix(_normalize_text(raw))
    if text in METRIC_HEADER_ALIASES:
        return METRIC_HEADER_ALIASES[text]
    if text in _known_metric_header_labels():
        return text
    return text


def metric_export_headers_v04(*, include_admin_columns: bool = False) -> list[str]:
    headers = [
        METRIC_EXPORT_HEADER_LEVEL,
        "科目性质",
        "科目代码",
        "科目名称",
    ]
    if include_admin_columns:
        headers.extend(["数值类型", "允许手工录入"])
    headers.extend(
        [
            "录入粒度",
            "年预算公式",
            "年预测公式",
            "实际月公式",
            "预测月公式",
            "公式说明",
            "横向汇总",
            "纵向汇总",
            METRIC_EXPORT_HEADER_LOGIC,
            "规则",
        ]
    )
    return headers


def metric_export_row_values(row: dict[str, Any], *, include_admin_columns: bool = False) -> list[Any]:
    actual = str(row.get("formula_actual") or row.get("formula") or "").strip()
    forecast = str(row.get("formula_forecast") or row.get("formula") or "").strip()
    values: list[Any] = [
        row.get("levelLabel") or "",
        row.get("nature") or "",
        row.get("code") or "",
        row.get("name") or "",
    ]
    if include_admin_columns:
        values.extend(
            [
                row.get("value_type") or "",
                _allow_manual_entry_label(row.get("allow_manual_entry")),
            ]
        )
    values.extend(
        [
            _entry_granularity_label(row.get("entry_granularity")),
            row.get("formula_budget_annual") or "",
            row.get("formula_forecast_annual") or "",
            actual,
            forecast,
            row.get("formula_note") or "",
            _rollup_flag_label(row.get("horizontal_rollup")),
            _rollup_flag_label(row.get("vertical_rollup")),
            row.get("logic_code") or "",
            str(row.get("annual_agg_rule") or "").strip().upper(),
        ]
    )
    return values


def _style_metric_export_header_row(ws, header_count: int) -> None:
    for cell in ws[1][:header_count]:
        cell.font = Font(bold=True)


def _apply_metric_export_v04_column_widths(ws, *, include_admin_columns: bool = False) -> None:
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 36
    if include_admin_columns:
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 14
        ws.column_dimensions["G"].width = 12
        ws.column_dimensions["H"].width = 48
        ws.column_dimensions["I"].width = 48
        ws.column_dimensions["J"].width = 48
        ws.column_dimensions["K"].width = 48
        ws.column_dimensions["L"].width = 16
        ws.column_dimensions["M"].width = 10
        ws.column_dimensions["N"].width = 10
        ws.column_dimensions["O"].width = 14
        ws.column_dimensions["P"].width = 10
    else:
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 48
        ws.column_dimensions["G"].width = 48
        ws.column_dimensions["H"].width = 48
        ws.column_dimensions["I"].width = 48
        ws.column_dimensions["J"].width = 16
        ws.column_dimensions["K"].width = 10
        ws.column_dimensions["L"].width = 10
        ws.column_dimensions["M"].width = 14
        ws.column_dimensions["N"].width = 10


def build_org_product_metric_import_template_workbook() -> BytesIO:
    """v04 导入模板：表头 + 示例 sheet + 填写说明。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "AA业务状况表"

    headers = metric_export_headers_v04(include_admin_columns=False)
    ws.append(headers)
    _style_metric_export_header_row(ws, len(headers))
    ws.append(
        [
            "一级",
            "收入",
            "AA.01",
            "营业收入",
            "按月",
            "",
            "",
            "=H3+H6+H9",
            "=I3+I6+I9",
            "收支月化指标",
            "是",
            "",
            "01",
            "SUM",
        ]
    )
    ws.append(
        [
            "二级",
            "收入",
            "AA.01.01",
            "  利息净收入",
            "",
            "",
            "",
            "=H4-H5",
            "=I4-I5",
            "",
            "是",
            "",
            "01.01",
            "SUM",
        ]
    )
    ws.append(["三级", "收入", "AA.14", "    利息收入", "", "", "", "", "", "", "", "", "01.01.14", ""])
    ws.append(["三级", "支出", "AA.16", "    利息支出", "", "", "", "", "", "", "", "", "01.01.16", ""])
    ws.append(["二级", "收入", "AA.01.02", "  净手续费收入", "", "", "", "=H7-H8", "=I7-I8", "", "是", "", "01.02", "SUM"])
    ws.append(["三级", "收入", "AA.18", "    手续费收入", "", "", "", "", "", "", "", "", "01.02.18", ""])
    ws.append(["三级", "支出", "AA.19", "    手续费支出", "", "", "", "", "", "", "", "", "01.02.19", ""])
    _apply_metric_export_v04_column_widths(ws, include_admin_columns=False)

    guide = wb.create_sheet("填写说明")
    guide.append(["机构及产品指标 Excel 导入模板（v04）"])
    guide.append([])
    guide.append(["1. 工作表命名", "机构代码 + 指标表名称，如 AA业务状况表、A01泛微粒贷（默认业务状况表）"])
    guide.append(["2. 必填列", "科目性质、科目代码、科目名称"])
    guide.append(["3. 公式列", "实际月公式 / 预测月公式可写 Excel 单元格公式（如 =H3+H6），导入时自动转为科目代码公式"])
    guide.append(["4. 仅展示列", "科目层级※仅展示、逻辑码※仅展示：导出供查看，导入时系统按科目代码重算，可留空"])
    guide.append(["5. 规则列", "SUM / AVG / LAST / WGT / CALC；率类全年重算填 CALC"])
    guide.append(["6. 导入后", "在页面点击「保存刷新」写入数据库"])
    guide.column_dimensions["A"].width = 18
    guide.column_dimensions["B"].width = 72

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _collect_header_hits(ws, max_row: int = 120, max_col: int = 40) -> list[tuple[int, int, str]]:
    known_headers = _known_metric_header_labels()
    hits: list[tuple[int, int, str]] = []
    scan_col = max(max_col, _sheet_scan_col_limit(ws))
    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, scan_col + 1):
            raw = _normalize_text(_ws_cell_value(ws, row_idx, col_idx))
            if not raw:
                continue
            canon = _canon_header_label(raw)
            if canon in known_headers:
                hits.append((row_idx, col_idx, canon))
    return hits


def _score_data_infer_nature_column(text: str) -> int:
    if text in NATURE_VALUE_SET:
        return 2
    if len(text) >= 2 and re.search(r"[\u4e00-\u9fff]{2,}", text) and text not in LEVEL_RANK:
        return 1
    return 0


def _header_map_from_hits(hits: list[tuple[int, int, str]], start_row: int, span: int = 6) -> dict[str, int]:
    row_map: dict[str, int] = {}
    end_row = start_row + span
    for row_idx, col_idx, canon in hits:
        if start_row <= row_idx <= end_row:
            row_map[canon] = col_idx
    return row_map


def _infer_header_map_from_data(ws, entity_code: str) -> tuple[int | None, dict[str, int]]:
    """表头行读不到时，根据 AA 科目代码等数据形态推断列。"""
    owner = _normalize_text(entity_code).upper()
    if not owner:
        return None, {}

    code_scores: dict[int, int] = defaultdict(int)
    name_scores: dict[int, int] = defaultdict(int)
    level_scores: dict[int, int] = defaultdict(int)
    nature_scores: dict[int, int] = defaultdict(int)
    first_hit_row: int | None = None

    for row_idx in range(1, 201):
        for col_idx in range(1, 41):
            raw = _ws_cell_value(ws, row_idx, col_idx)
            text = _normalize_text(raw)
            if not text:
                continue
            code = _normalize_metric_code(owner, raw)
            if owner and code.startswith(owner) and len(code) >= len(owner) + 1:
                code_scores[col_idx] += 3
                first_hit_row = first_hit_row or row_idx
                continue
            if text in LEVEL_RANK:
                level_scores[col_idx] += 2
                first_hit_row = first_hit_row or row_idx
                continue
            nature_score = _score_data_infer_nature_column(text)
            if nature_score:
                nature_scores[col_idx] += nature_score
                first_hit_row = first_hit_row or row_idx
                continue
            if re.search(r"[\u4e00-\u9fff]{2,}", text) and not re.fullmatch(r"[A-Z0-9.]+", code):
                name_scores[col_idx] += 1

    if not code_scores:
        return None, {}

    code_col = max(code_scores, key=code_scores.get)
    name_col = max(
        (col for col in name_scores if col != code_col),
        key=lambda c: name_scores[c],
        default=0,
    )
    if not name_col:
        for col_idx in range(1, 41):
            if col_idx == code_col:
                continue
            name_scores[col_idx] += sum(
                1
                for row_idx in range(1, 201)
                if re.search(
                    r"[\u4e00-\u9fff]{2,}",
                    _normalize_text(ws.cell(row_idx, col_idx).value),
                )
                and _normalize_text(ws.cell(row_idx, col_idx).value) not in NATURE_VALUE_SET
                and _normalize_text(ws.cell(row_idx, col_idx).value) not in LEVEL_RANK
            )
        name_col = max(
            (col for col in name_scores if col != code_col),
            key=lambda c: name_scores[c],
            default=0,
        )

    if not name_col:
        return None, {}

    level_col = max(level_scores, key=level_scores.get) if level_scores else 0
    nature_col = max(nature_scores, key=nature_scores.get) if nature_scores else 0
    header_map: dict[str, int] = {"科目代码": code_col, "科目名称": name_col}
    if level_col:
        header_map["科目层级"] = level_col
    if nature_col:
        header_map["科目性质"] = nature_col
    header_row = max(1, (first_hit_row or 2) - 1)
    return header_row, header_map


def _parse_strict_import_flag(value: str) -> bool:
    text = _normalize_text(value).lower()
    if not text:
        return True
    return text not in {"0", "false", "no", "off", "宽松", "lenient"}


def _find_header_row(
    ws,
    entity_code: str = "",
    *,
    strict: bool = False,
    sheet_title: str = "",
) -> tuple[int | None, dict[str, int], str, str]:
    """
    按表头文字定位列号，与 Excel 列左右顺序无关。
    返回：(表头行, 列映射, 模式, 来源 header|data_infer)
    strict=True：须四列表头齐全，禁止 data_infer。
    """
    full_headers = {"科目层级", "科目性质", "科目代码", "科目名称"}
    hits = _collect_header_hits(ws)

    if strict:
        scanned_row, scanned_map = _header_map_scan_rows(ws)
        if scanned_row and full_headers.issubset(scanned_map):
            return scanned_row, scanned_map, "full", "header"

        for start_row in range(1, 31):
            row_map = _header_map_from_hits(hits, start_row, span=5)
            if not full_headers.issubset(row_map):
                continue
            header_row = max((r for r, _, _ in hits if start_row <= r <= start_row + 5), default=start_row)
            return header_row, row_map, "full", "header"
        inferred_row, inferred_map = _infer_header_map_from_data(ws, entity_code)
        if inferred_row and full_headers.issubset(inferred_map):
            return inferred_row, inferred_map, "full", "header_infer"
        layout_row, layout_map = _balance_sheet_standard_layout_fallback(ws, sheet_title)
        if layout_row and full_headers.issubset(layout_map):
            return layout_row, layout_map, "full", "layout_abcd"
        return None, {}, "none", ""

    best_row: int | None = None
    best_map: dict[str, int] = {}
    best_score = 0

    for row_idx in range(1, 121):
        row_map = _header_map_from_hits(hits, row_idx, span=0)
        score = len(row_map)
        if "科目代码" in row_map and "科目名称" in row_map:
            score += 10
        if score > best_score:
            best_score = score
            best_row = row_idx
            best_map = dict(row_map)

    for start_row in range(1, 31):
        row_map = _header_map_from_hits(hits, start_row, span=5)
        score = len(row_map)
        if "科目代码" in row_map and "科目名称" in row_map:
            score += 10
            header_row = max((r for r, _, _ in hits if start_row <= r <= start_row + 5), default=start_row)
            mode = "full" if full_headers.issubset(row_map) else "minimal"
            return header_row, row_map, mode, "header"

    if best_row and "科目代码" in best_map and "科目名称" in best_map:
        mode = "full" if full_headers.issubset(best_map) else "minimal"
        return best_row, best_map, mode, "header"

    inferred_row, inferred_map = _infer_header_map_from_data(ws, entity_code)
    if inferred_row and inferred_map.get("科目代码") and inferred_map.get("科目名称"):
        mode = "full" if full_headers.issubset(inferred_map) else "minimal"
        return inferred_row, inferred_map, mode, "data_infer"

    return None, {}, "none", ""


def _parse_metric_worksheet(ws, prefix: str) -> tuple[list[dict[str, Any]], int]:
    header_row_idx, header_map, header_mode, _header_source = _find_header_row(ws)
    if header_row_idx is None:
        raise HTTPException(status_code=400, detail=f"工作表“{ws.title}”缺少表头：科目层级、科目性质、科目代码、科目名称")

    note_col = header_map.get("指标解释") or header_map.get("备注")
    formula_col = header_map.get("取数公式")
    level_col = header_map.get("科目层级")
    nature_col = header_map.get("科目性质")
    code_col = header_map["科目代码"]
    name_col = header_map["科目名称"]
    roots: list[dict[str, Any]] = []
    stack: dict[int, dict[str, Any]] = {}
    row_count = 0

    for row_idx in range(header_row_idx + 1, _sheet_scan_row_limit(ws, header_row_idx) + 1):
        level_label = _normalize_level_label(ws.cell(row_idx, level_col).value) if level_col else ""
        nature = _normalize_nature(ws.cell(row_idx, nature_col).value) if nature_col else ""
        code = _normalize_text(ws.cell(row_idx, code_col).value)
        name = _normalize_text(ws.cell(row_idx, name_col).value)
        if header_mode == "minimal" or not level_label:
            level_label = level_label or _infer_level_label_from_code(code)
        if not nature:
            nature = "其他"
        note = _normalize_text(ws.cell(row_idx, note_col).value) if note_col else ""
        formula = _normalize_text(ws.cell(row_idx, formula_col).value) if formula_col else ""
        if not any([level_label, nature, code, name, note, formula]):
            continue
        nature = _apply_profit_nature(name, nature)
        rank = LEVEL_RANK.get(level_label)
        if not rank:
            raise HTTPException(status_code=400, detail=f"第 {row_idx} 行科目层级不合法：{level_label}")
        if not code or not name:
            raise HTTPException(status_code=400, detail=f"第 {row_idx} 行科目代码或科目名称为空")

        node = {
            "id": f"import-{prefix}-{row_idx}",
            "levelLabel": level_label,
            "nature": nature,
            "code": code,
            "name": name,
            "note": note,
            "formula": formula,
            "children": [],
        }
        if rank == 1 or stack.get(rank - 1) is None:
            roots.append(node)
        else:
            stack[rank - 1]["children"].append(node)

        stack[rank] = node
        for key in list(stack.keys()):
            if key > rank:
                del stack[key]
        row_count += 1

    return roots, row_count


def _resolve_metric_level_label(level_label: str, code: str, *, strict: bool) -> tuple[str, int | None]:
    label = _normalize_level_label(level_label)
    rank = LEVEL_RANK.get(label)
    if rank:
        return label, rank
    if strict and code:
        inferred = _infer_level_label_from_code(code)
        rank = LEVEL_RANK.get(inferred)
        if rank:
            return inferred, rank
    if not strict and code:
        inferred = _infer_level_label_from_code(code)
        return inferred, LEVEL_RANK.get(inferred)
    return label, rank


def _resolve_import_sheet_entity_table(
    sheet_name: str,
    candidates: list[tuple[str, str, str]],
    *,
    strict: bool = True,
) -> tuple[str, str, str] | None:
    entity_code = ""
    table_name = ""
    entity_name = ""
    if candidates:
        matched = _match_sheet_title_to_entity_table(sheet_name, candidates, strict=strict)
        if not matched:
            return None
        entity_code, table_name = matched
        table_name = _canonical_import_table_name(table_name)
        for code, name, tn in candidates:
            if code == entity_code and _canonical_import_table_name(tn) == table_name:
                entity_name = name
                break
    else:
        owner_code = _extract_owner_code(sheet_name, [])
        if not owner_code:
            return None
        entity_code = owner_code
        compact_sheet = _normalize_metric_sheet_title_key(sheet_name)
        entity_name = compact_sheet[len(owner_code) :] if compact_sheet.upper().startswith(owner_code.upper()) else sheet_name
        table_name = entity_name if entity_name.endswith("表") else DEFAULT_METRIC_TABLE_NAME
        if table_name == entity_name and not table_name.endswith("表"):
            table_name = DEFAULT_METRIC_TABLE_NAME
    if not table_name:
        return None
    return entity_code, table_name, entity_name


def _read_formula_cell_text(ws, row_idx: int, col_idx: int) -> str:
    raw = _ws_cell_value(ws, row_idx, col_idx)
    if raw is None:
        return ""
    if isinstance(raw, str):
        return raw.strip()
    return _normalize_text(raw)


def _parse_formula_cell_value(
    ws,
    row_idx: int,
    col_idx: int | None,
    *,
    sheet_formula_context: SheetFormulaContext | None,
    all_sheet_contexts: dict[str, SheetFormulaContext] | None,
    formula_convert_errors: list[dict[str, Any]] | None,
) -> str:
    if not col_idx:
        return ""
    raw_formula = _read_formula_cell_text(ws, row_idx, col_idx)
    if not raw_formula:
        return ""
    if raw_formula.startswith("="):
        if sheet_formula_context and all_sheet_contexts:
            try:
                return convert_excel_formula_to_system(
                    raw_formula,
                    current=sheet_formula_context,
                    all_contexts=all_sheet_contexts,
                )
            except ExcelFormulaConvertError as exc:
                if formula_convert_errors is not None:
                    formula_convert_errors.append(
                        {
                            "sheet_name": str(getattr(ws, "title", "") or ""),
                            "row": row_idx,
                            "excel_formula": raw_formula,
                            "reason": str(exc),
                        }
                    )
                return raw_formula[1:].strip()
        return raw_formula[1:].strip()
    return raw_formula


def _parse_metric_worksheet_basic(
    ws,
    prefix: str,
    entity_code: str = "",
    *,
    strict: bool = False,
    sheet_formula_context: SheetFormulaContext | None = None,
    all_sheet_contexts: dict[str, SheetFormulaContext] | None = None,
    formula_convert_errors: list[dict[str, Any]] | None = None,
) -> tuple[list[dict[str, Any]], int, str | None, dict[str, int]]:
    full_headers = {"科目层级", "科目性质", "科目代码", "科目名称"}
    header_row_idx, header_map, header_mode, header_source = _find_header_row(
        ws, entity_code, strict=strict, sheet_title=str(getattr(ws, "title", "") or "")
    )
    if header_row_idx is None:
        if strict:
            return [], 0, "严格模式：缺少四列表头（科目层级、科目性质、科目代码、科目名称）", {}
        return [], 0, "缺少表头（至少需：科目代码、科目名称；推荐含科目层级、科目性质）", {}

    if strict and header_source not in {"header", "header_infer", "layout_abcd"}:
        return [], 0, "严格模式：须识别四列表头（表头文字或四列数据形态推断）", header_map

    if strict and not full_headers.issubset(header_map):
        missing = "、".join(sorted(full_headers - set(header_map.keys())))
        found = "、".join(sorted(header_map.keys())) or "无"
        return [], 0, f"严格模式：四列表头须齐全（缺少：{missing}；已识别：{found}）", header_map

    level_col = header_map.get("科目层级")
    nature_col = header_map.get("科目性质")
    code_col = header_map.get("科目代码")
    name_col = header_map.get("科目名称")
    formula_col = header_map.get("取数公式")
    formula_budget_annual_col = header_map.get("年预算公式")
    formula_forecast_annual_col = header_map.get("年预测公式")
    formula_actual_col = header_map.get("实际月公式")
    formula_forecast_col = header_map.get("预测月公式")
    formula_note_col = header_map.get("公式说明")
    entry_granularity_col = header_map.get("录入粒度")
    horizontal_rollup_col = header_map.get("横向汇总")
    vertical_rollup_col = header_map.get("纵向汇总")
    logic_code_col = header_map.get("逻辑码")
    value_type_col = header_map.get("数值类型")
    allow_manual_entry_col = header_map.get("允许手工录入")
    annual_agg_rule_col = header_map.get("规则")
    if not code_col or not name_col:
        found = "、".join(sorted(header_map.keys())) or "无"
        return [], 0, f"表头中未找到「科目代码/科目名称」列（已识别：{found}）", header_map
    code_col, name_col = _maybe_swap_metric_code_name_columns(
        ws,
        header_row_idx=header_row_idx,
        entity_code=entity_code,
        code_col=code_col,
        name_col=name_col,
    )
    if code_col:
        header_map["科目代码"] = code_col
    if name_col:
        header_map["科目名称"] = name_col
    if strict and (not level_col or not nature_col):
        return [], 0, "严格模式：四列表头须齐全（科目层级、科目性质、科目代码、科目名称）", header_map

    roots: list[dict[str, Any]] = []
    stack: dict[int, dict[str, Any]] = {}
    row_count = 0
    skipped_rows = 0
    empty_streak = 0
    row_limit = _sheet_scan_row_limit(ws, header_row_idx)

    for row_idx in range(header_row_idx + 1, row_limit + 1):
        level_raw = _ws_cell_value(ws, row_idx, level_col) if level_col else None
        nature_raw = _ws_cell_value(ws, row_idx, nature_col) if nature_col else None
        code_raw = _ws_cell_value(ws, row_idx, code_col) if code_col else None
        name_raw = _ws_cell_value(ws, row_idx, name_col) if name_col else None
        code = _normalize_metric_code(entity_code, code_raw)
        name = _normalize_text(name_raw)
        level_label = _infer_level_label_from_code(code)
        rank = LEVEL_RANK.get(level_label)
        if not rank:
            level_label, rank = _resolve_metric_level_label(_normalize_level_label(level_raw), code, strict=strict)
        nature = _normalize_nature(nature_raw)
        formula_legacy = _parse_formula_cell_value(
            ws,
            row_idx,
            formula_col,
            sheet_formula_context=sheet_formula_context,
            all_sheet_contexts=all_sheet_contexts,
            formula_convert_errors=formula_convert_errors,
        )
        formula_budget_annual = _parse_formula_cell_value(
            ws,
            row_idx,
            formula_budget_annual_col,
            sheet_formula_context=sheet_formula_context,
            all_sheet_contexts=all_sheet_contexts,
            formula_convert_errors=formula_convert_errors,
        )
        formula_forecast_annual = _parse_formula_cell_value(
            ws,
            row_idx,
            formula_forecast_annual_col,
            sheet_formula_context=sheet_formula_context,
            all_sheet_contexts=all_sheet_contexts,
            formula_convert_errors=formula_convert_errors,
        )
        formula_actual = _parse_formula_cell_value(
            ws,
            row_idx,
            formula_actual_col,
            sheet_formula_context=sheet_formula_context,
            all_sheet_contexts=all_sheet_contexts,
            formula_convert_errors=formula_convert_errors,
        )
        formula_forecast = _parse_formula_cell_value(
            ws,
            row_idx,
            formula_forecast_col,
            sheet_formula_context=sheet_formula_context,
            all_sheet_contexts=all_sheet_contexts,
            formula_convert_errors=formula_convert_errors,
        )
        formula = formula_legacy or formula_forecast or formula_actual
        formula_note = _normalize_text(_ws_cell_value(ws, row_idx, formula_note_col)) if formula_note_col else ""
        entry_granularity = (
            _normalize_entry_granularity(_ws_cell_value(ws, row_idx, entry_granularity_col))
            if entry_granularity_col
            else "monthly"
        )
        value_type = (
            _normalize_metric_value_type(_ws_cell_value(ws, row_idx, value_type_col), nature)
            if value_type_col
            else _normalize_metric_value_type("", nature)
        )
        allow_manual_entry = (
            _normalize_allow_manual_entry(_ws_cell_value(ws, row_idx, allow_manual_entry_col), 1)
            if allow_manual_entry_col
            else 1
        )
        annual_agg_rule = (
            _normalize_annual_agg_rule(_ws_cell_value(ws, row_idx, annual_agg_rule_col))
            if annual_agg_rule_col
            else ""
        )
        horizontal_rollup = _normalize_rollup_flag(_ws_cell_value(ws, row_idx, horizontal_rollup_col)) if horizontal_rollup_col else 0
        vertical_rollup = _normalize_rollup_flag(_ws_cell_value(ws, row_idx, vertical_rollup_col)) if vertical_rollup_col else 0
        logic_code = _derive_metric_logic_code(entity_code, code, "")
        if strict:
            if not any(
                [
                    level_label,
                    nature,
                    code,
                    name,
                    formula,
                    formula_budget_annual,
                    formula_forecast_annual,
                    formula_actual,
                    formula_forecast,
                    formula_note,
                    entry_granularity_col and _normalize_text(_ws_cell_value(ws, row_idx, entry_granularity_col)),
                    value_type_col and _normalize_text(_ws_cell_value(ws, row_idx, value_type_col)),
                    allow_manual_entry_col and _normalize_text(_ws_cell_value(ws, row_idx, allow_manual_entry_col)),
                    horizontal_rollup,
                    vertical_rollup,
                    logic_code,
                ]
            ):
                empty_streak += 1
                if empty_streak >= 40:
                    break
                continue
            empty_streak = 0
            if not code or not name:
                continue
            if not nature:
                nature = "其他"
            nature = _apply_profit_nature(name, nature)
            level_label, rank = _resolve_metric_level_label(level_label, code, strict=True)
            if not rank:
                skipped_rows += 1
                continue
        else:
            if header_source == "data_infer" and not level_label:
                level_label = _infer_level_label_from_code(code)
            if not code and not name:
                continue
            if not code or not name:
                skipped_rows += 1
                continue
            if not nature:
                nature = "其他"
            nature = _apply_profit_nature(name, nature)
            level_label, rank = _resolve_metric_level_label(level_label, code, strict=False)
            if not rank:
                skipped_rows += 1
                continue

        node = {
            "id": f"import-{prefix}-{row_idx}",
            "levelLabel": level_label,
            "nature": nature,
            "code": code,
            "name": name,
            "value_type": value_type,
            "allow_manual_entry": allow_manual_entry,
            "note": "",
            "formula": formula,
            "formula_budget_annual": formula_budget_annual,
            "formula_forecast_annual": formula_forecast_annual,
            "formula_actual": formula_actual,
            "formula_forecast": formula_forecast,
            "formula_note": formula_note,
            "entry_granularity": entry_granularity,
            "horizontal_rollup": horizontal_rollup,
            "vertical_rollup": vertical_rollup,
            "logic_code": logic_code,
            "annual_agg_rule": annual_agg_rule,
            "children": [],
        }
        if rank == 1 or stack.get(rank - 1) is None:
            roots.append(node)
        else:
            stack[rank - 1]["children"].append(node)

        stack[rank] = node
        for key in list(stack.keys()):
            if key > rank:
                del stack[key]
        row_count += 1

    if row_count <= 0:
        if strict:
            sample_code = _normalize_text(_ws_cell_value(ws, header_row_idx + 1, code_col))
            sample_name = _normalize_text(_ws_cell_value(ws, header_row_idx + 1, name_col))
            detail = (
                "严格模式：未解析到有效科目行（请确认「科目代码」「科目名称」列有值；"
                "公式列请在本机 Excel 打开并保存后再导入）"
            )
            if sample_code or sample_name:
                detail += f"；第{header_row_idx + 1}行读取示例：代码={sample_code or '空'}，名称={sample_name or '空'}"
            if skipped_rows > 0:
                detail += f"；已跳过 {skipped_rows} 行"
            return [], 0, detail, header_map
        reason = "未解析到有效科目行（需含科目代码、科目名称；层级可留空由代码推断）"
        if skipped_rows > 0:
            reason += f"；已跳过 {skipped_rows} 行"
        if header_mode == "minimal":
            reason += "；当前表头仅识别到「科目代码、科目名称」"
        return [], 0, reason, header_map
    return roots, row_count, None, header_map


def _default_metric_import_candidates() -> list[tuple[str, str, str]]:
    return [
        ("AA", "微众银行", table_name)
        for table_name in (
            "业务状况表",
            "损益表",
            "资产负债表（余额）",
            "资产负债表（日均）",
            "资产质量表",
            "利息净收入表",
            "净利息收入表",
        )
    ]


def _merge_metric_import_candidates(candidates_json: str = "") -> list[tuple[str, str, str]]:
    candidates: list[tuple[str, str, str]] = []
    seen_candidate_keys: set[str] = set()

    def _add_candidate(code: str, name: str, table_name: str) -> None:
        entity_code = _normalize_text(code).upper()
        canonical_table = _normalize_text(table_name)
        if not entity_code or not canonical_table:
            return
        key = f"{entity_code}::{canonical_table}"
        if key in seen_candidate_keys:
            return
        seen_candidate_keys.add(key)
        candidates.append((entity_code, _normalize_text(name), canonical_table))

    if candidates_json.strip():
        try:
            parsed = json.loads(candidates_json)
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"candidates_json 无效：{exc}") from exc
        if isinstance(parsed, list):
            for item in parsed:
                if not isinstance(item, dict):
                    continue
                _add_candidate(
                    str(item.get("entity_code") or ""),
                    str(item.get("entity_name") or ""),
                    str(item.get("table_name") or ""),
                )

    path = common_db_path()
    if path.exists():
        try:
            with sqlite3.connect(path) as conn:
                _ensure_metric_table_catalog(conn)
                _seed_metric_table_catalog(conn)
                known_names = _canonical_import_table_names(conn)
                for entity_code, entity_name, table_name in _import_report_catalog_candidates(conn):
                    _add_candidate(
                        entity_code,
                        entity_name,
                        _canonical_import_table_name(table_name, known_names=known_names),
                    )
                for entity_code, entity_name, table_name in _import_report_saved_metric_candidates(conn):
                    _add_candidate(
                        entity_code,
                        entity_name,
                        _canonical_import_table_name(table_name, known_names=known_names),
                    )
        except Exception:
            for entity_code, entity_name, table_name in _default_metric_import_candidates():
                _add_candidate(entity_code, entity_name, table_name)
    else:
        for entity_code, entity_name, table_name in _default_metric_import_candidates():
            _add_candidate(entity_code, entity_name, table_name)

    return candidates


def parse_metric_workbook_import(
    content: bytes,
    filename: str,
    *,
    candidates_json: str = "",
    strict_import: bool = True,
    table_names_filter: list[str] | None = None,
    only_first_sheet: bool = False,
) -> dict[str, Any]:
    try:
        workbook = load_workbook(filename=BytesIO(content), data_only=False, read_only=False)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"无法读取Excel文件：{exc}") from exc

    candidates = _merge_metric_import_candidates(candidates_json)
    imported_entities: list[dict[str, Any]] = []
    ignored_sheets: list[str] = []
    ignored_details: list[dict[str, str]] = []
    formula_convert_errors: list[dict[str, Any]] = []
    prefix = re.sub(r"[^A-Za-z0-9]+", "-", Path(filename or "metrics.xlsx").stem) or "import"
    strict_flag = _parse_strict_import_flag(str(strict_import))
    filter_keys = {
        _canonical_metric_table_key(name)
        for name in (table_names_filter or [])
        if _normalize_text(name)
    }

    def _ignore_sheet(sheet: str, reason: str) -> None:
        ignored_sheets.append(sheet)
        ignored_details.append({"sheet_name": sheet, "reason": reason})

    sheet_names = workbook.sheetnames[:1] if only_first_sheet else list(workbook.sheetnames)

    sheet_contexts: list[SheetFormulaContext] = []
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        _prepare_metric_worksheet(ws)
        resolved = _resolve_import_sheet_entity_table(sheet_name, candidates, strict=not only_first_sheet)
        if not resolved and only_first_sheet and sheet_name == workbook.sheetnames[0]:
            owner_code = _extract_owner_code(sheet_name, [])
            if owner_code:
                resolved = (owner_code, DEFAULT_METRIC_TABLE_NAME, sheet_name)
        if not resolved:
            continue
        entity_code, table_name, _entity_name = resolved
        header_row_idx, header_map, _header_mode, _header_source = _find_header_row(
            ws, entity_code, strict=strict_flag, sheet_title=sheet_name
        )
        code_col = header_map.get("科目代码") if header_row_idx else None
        name_col = header_map.get("科目名称") if header_row_idx else None
        code_col, _name_col = _maybe_swap_metric_code_name_columns(
            ws,
            header_row_idx=header_row_idx or 1,
            entity_code=entity_code,
            code_col=code_col,
            name_col=name_col,
        )
        if not header_row_idx or not code_col:
            continue
        row_limit = _sheet_scan_row_limit(ws, header_row_idx)
        sheet_contexts.append(
            build_sheet_formula_context(
                sheet_name,
                entity_code,
                table_name,
                header_row_idx,
                code_col,
                lambda r, c, _ws=ws: _ws_cell_value(_ws, r, c),
                _normalize_metric_code,
                row_limit,
            )
        )
    all_sheet_contexts = index_sheet_contexts(sheet_contexts)

    for sheet_name in sheet_names:
        ws = workbook[sheet_name]
        _prepare_metric_worksheet(ws)

        resolved = _resolve_import_sheet_entity_table(sheet_name, candidates, strict=not only_first_sheet)
        if not resolved and only_first_sheet:
            owner_code = _extract_owner_code(sheet_name, [])
            if owner_code:
                resolved = (owner_code, DEFAULT_METRIC_TABLE_NAME, sheet_name)
        if not resolved:
            _ignore_sheet(
                sheet_name,
                "工作表名未匹配到机构/指标表（标准格式：代码+表名，如 AA资产质量表）",
            )
            continue
        entity_code, table_name, entity_name = resolved
        table_key = _canonical_metric_table_key(table_name)
        if filter_keys and table_key not in filter_keys:
            continue

        sheet_ctx = all_sheet_contexts.get(normalize_sheet_lookup_key(sheet_name))

        metrics, row_count, parse_error, header_map = _parse_metric_worksheet_basic(
            ws,
            f"{prefix}-{sheet_name}",
            entity_code=entity_code,
            strict=strict_flag,
            sheet_formula_context=sheet_ctx,
            all_sheet_contexts=all_sheet_contexts,
            formula_convert_errors=formula_convert_errors,
        )
        if parse_error:
            detail = parse_error
            if header_map:
                detail += f"；列映射={header_map}"
            _ignore_sheet(sheet_name, detail)
            continue

        if row_count > 0 and len(metrics) <= 0:
            _ignore_sheet(sheet_name, "解析到行数但科目树为空，请检查科目层级/代码列")
            continue

        imported_entities.append(
            {
                "sheet_name": sheet_name,
                "entity_code": entity_code,
                "entity_name": entity_name,
                "table_name": table_name,
                "row_count": row_count,
                "has_formula_column": bool(
                    header_map.get("取数公式")
                    or header_map.get("年预算公式")
                    or header_map.get("年预测公式")
                    or header_map.get("实际月公式")
                    or header_map.get("预测月公式")
                ),
                "metrics": metrics,
            }
        )

    return {
        "imported_entities": imported_entities,
        "ignored_sheets": ignored_sheets,
        "ignored_details": ignored_details,
        "formula_convert_errors": formula_convert_errors,
    }


def _parse_metric_upload(content: bytes, filename: str) -> tuple[list[dict[str, Any]], int]:
    result = parse_metric_workbook_import(
        content,
        filename or "metrics.xlsx",
        strict_import=False,
        only_first_sheet=True,
    )
    if result["imported_entities"]:
        first = result["imported_entities"][0]
        return list(first["metrics"]), int(first["row_count"])
    detail = ""
    if result["ignored_details"]:
        detail = str(result["ignored_details"][0].get("reason") or "")
    raise HTTPException(
        status_code=400,
        detail=detail or "无法解析 Excel：请确认表头含科目代码/科目名称，且公式列使用 Excel 原生公式或系统文本",
    )


def _normalize_sheet_key(value: str) -> str:
    return re.sub(r"\s+", "", value or "")


def _match_sheet_to_table(sheet_name: str, table_name: str) -> bool:
    normalized_sheet = _normalize_sheet_key(sheet_name)
    normalized_table = _normalize_sheet_key(table_name)
    return bool(normalized_sheet) and bool(normalized_table) and (
        normalized_sheet == normalized_table
        or normalized_sheet.endswith(normalized_table)
        or normalized_table in normalized_sheet
    )


def _parse_metric_batch_upload(content: bytes, filename: str, table_names: list[str]) -> tuple[list[dict[str, Any]], list[str], list[str]]:
    result = parse_metric_workbook_import(
        content,
        filename or "metrics.xlsx",
        strict_import=False,
        table_names_filter=[name.strip() for name in table_names if name.strip()],
    )
    imported_tables = [
        {
            "table_name": item["table_name"],
            "row_count": item["row_count"],
            "metrics": item["metrics"],
        }
        for item in result["imported_entities"]
    ]
    imported_table_keys = {_canonical_metric_table_key(item["table_name"]) for item in imported_tables}
    missing_tables = [
        name
        for name in table_names
        if _canonical_metric_table_key(name) not in imported_table_keys
    ]
    matched_sheet_names = {str(item.get("sheet_name") or "") for item in result["imported_entities"]}
    ignored_sheets = [
        sheet_name
        for sheet_name in result["ignored_sheets"]
        if sheet_name not in matched_sheet_names
    ]
    return imported_tables, missing_tables, ignored_sheets


def _find_table_header_row(ws, required_header: str) -> tuple[int | None, list[str], dict[str, int]]:
    header_row_idx: int | None = None
    headers: list[str] = []
    header_map: dict[str, int] = {}

    for row_idx in range(1, min(ws.max_row, 20) + 1):
        row_headers: list[str] = []
        row_map: dict[str, int] = {}
        for col_idx in range(1, ws.max_column + 1):
            h = _normalize_text(ws.cell(row_idx, col_idx).value)
            row_headers.append(h)
            if h and h not in row_map:
                row_map[h] = col_idx
        if required_header in row_map:
            header_row_idx = row_idx
            headers = row_headers
            header_map = row_map
            break

    headers = [h for h in headers if h]
    return header_row_idx, headers, header_map


def _parse_simple_excel_table(content: bytes, filename: str, required_header: str) -> tuple[list[str], list[dict[str, str]]]:
    try:
        workbook = load_workbook(filename=BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"无法读取Excel文件：{exc}") from exc

    ws = workbook[workbook.sheetnames[0]]
    header_row_idx, headers, header_map = _find_table_header_row(ws, required_header)
    if header_row_idx is None:
        raise HTTPException(status_code=400, detail=f"工作表“{ws.title}”缺少表头：{required_header}")

    rows: list[dict[str, str]] = []
    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        row_obj: dict[str, str] = {}
        any_value = False
        for h in headers:
            col_idx = header_map.get(h)
            if not col_idx:
                continue
            v = _normalize_text(ws.cell(row_idx, col_idx).value)
            if v:
                any_value = True
            row_obj[h] = v
        if not any_value:
            continue
        rows.append(row_obj)
    return headers, rows


def _map_data_entry_header_to_key(header: str, year: int, forecast_month: int) -> str | None:
    h = _normalize_text(header)
    if not h:
        return None
    prev_yy = (int(year) - 1) % 100
    yy = int(year) % 100
    if h in {f"{prev_yy}年实际", f"{year - 1}年实际"} or h.endswith("年实际") and str(year - 1) in h:
        return "prev_actual"
    if h in {f"{yy}年预算", f"{year}年预算", f"{prev_yy}年预算", f"{year - 1}年预算"} or (
        h.endswith("年预算")
        and "月" not in h
        and (str(year) in h or str(yy) in h or str(year - 1) in h or str(prev_yy) in h)
    ):
        return "prev_budget"
    # 全年预测列由前端按滚动月汇总计算，Excel 中该列即使存在也不在导入时写入
    if h.endswith("年预测") and "月" not in h and (str(year) in h or str(yy) in h):
        return None
    if h in {f"{prev_yy}年预测", f"{year - 1}年预测"} or (
        h.endswith("年预测") and "月" not in h and (str(year - 1) in h or str(prev_yy) in h)
    ):
        return None
    fm = max(1, min(12, int(forecast_month)))
    m_actual = re.search(rf"(?:{yy}年)?(\d{{1,2}})月\s*实际", h)
    if m_actual:
        month = int(m_actual.group(1))
        if 1 <= month <= fm:
            return f"a{month}"
    m_forecast = re.search(rf"(?:{yy}年)?(\d{{1,2}})月\s*预测", h)
    if m_forecast:
        month = int(m_forecast.group(1))
        if fm < month <= 12:
            return f"f{month}"
    if h in {"实际", "预算", "预测"}:
        return None
    return None


def _normalize_data_entry_metric_code(entity_code: str, raw_code: str) -> str:
    owner = _normalize_text(entity_code).upper()
    cleaned = _normalize_text(raw_code).upper().replace(" ", "")
    if not cleaned:
        return ""
    return _normalize_runtime_metric_code(owner, cleaned)


def _match_sheet_title_to_entity_table(
    sheet_title: str,
    candidates: list[tuple[str, str, str]],
    *,
    strict: bool = False,
) -> tuple[str, str] | None:
    """工作表名标准：机构及产品代码 + 指标表名称，如「AA业务状况表」。"""
    title = _normalize_metric_sheet_title_key(sheet_title)
    if not title:
        return None

    sorted_candidates = sorted(
        candidates,
        key=lambda item: len(_canonical_metric_table_key(item[2])),
        reverse=True,
    )

    def _try_match(candidate_title: str) -> tuple[str, str] | None:
        for entity_code, _entity_name, table_name in sorted_candidates:
            code = _normalize_text(entity_code).upper()
            tn = _canonical_metric_table_key(table_name)
            if not code or not tn:
                continue
            if candidate_title == f"{code}{tn}":
                return entity_code, table_name
            aliased = _apply_sheet_title_table_aliases(candidate_title, code)
            if aliased != candidate_title and aliased == f"{code}{tn}":
                return entity_code, table_name
        return None

    matched = _try_match(title)
    if matched:
        return matched

    # v01 用户上传模板：产品/群组 sheet 常命名为「代码+名称」
    # （例如 A01泛微粒贷），这种 sheet 默认导入为业务状况表。
    for entity_code, entity_name, table_name in sorted_candidates:
        code = _normalize_text(entity_code).upper()
        name = re.sub(r"[\s_]+", "", _normalize_text(entity_name))
        if not code or not name or not table_name:
            continue
        if title == f"{code}{name}" and _canonical_metric_table_key(table_name) == _canonical_metric_table_key(DEFAULT_METRIC_TABLE_NAME):
            return entity_code, table_name

    if not title.endswith("表"):
        matched = _try_match(f"{title}表")
        if matched:
            return matched

    if strict:
        return None

    best: tuple[str, str, int] | None = None
    for entity_code, entity_name, table_name in sorted_candidates:
        code = _normalize_text(entity_code).upper()
        name = re.sub(r"[\s_]+", "", _normalize_text(entity_name))
        tn = _canonical_metric_table_key(table_name)
        if not tn:
            continue
        scored_patterns: list[tuple[str, int]] = []
        if code and tn:
            scored_patterns.append((f"{code}{tn}", 200))
        if code and name and tn:
            scored_patterns.append((f"{code}{name}{tn}", 120))
        for pattern, base_score in scored_patterns:
            if pattern and pattern in title:
                score = base_score + len(pattern)
                if best is None or score > best[2]:
                    best = (entity_code, table_name, score)
    if best:
        return best[0], best[1]
    return None


def _rows_to_data_entry_values(
    headers: list[str],
    row: dict[str, str],
    year: int,
    forecast_month: int,
) -> dict[str, Any]:
    values: dict[str, Any] = {
        "prev_actual": "",
        "prev_budget": "",
        "prev_forecast": "",
        "year_forecast": "",
        "months": {},
    }
    for h in headers:
        key = _map_data_entry_header_to_key(h, year, forecast_month)
        if not key:
            continue
        cell_val = _normalize_text(row.get(h))
        if not cell_val:
            continue
        if key in {"prev_actual", "prev_budget", "prev_forecast", "year_forecast"}:
            values[key] = cell_val
        else:
            months = values.setdefault("months", {})
            if isinstance(months, dict):
                months[key] = cell_val
    return values


def _parse_data_entry_workbook(
    content: bytes,
    filename: str,
    year: int,
    forecast_month: int,
    candidates: list[tuple[str, str, str]],
) -> list[dict[str, Any]]:
    try:
        workbook = load_workbook(filename=BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"无法读取Excel文件：{exc}") from exc

    sheets_out: list[dict[str, Any]] = []
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        header_row_idx, headers, header_map = _find_table_header_row(ws, "科目代码")
        if header_row_idx is None:
            continue
        rows: list[dict[str, str]] = []
        for row_idx in range(header_row_idx + 1, ws.max_row + 1):
            row_obj: dict[str, str] = {}
            any_value = False
            for h in headers:
                col_idx = header_map.get(h)
                if not col_idx:
                    continue
                v = _normalize_text(ws.cell(row_idx, col_idx).value)
                if v:
                    any_value = True
                row_obj[h] = v
            if not any_value:
                continue
            rows.append(row_obj)
        matched = _match_sheet_title_to_entity_table(sheet_name, candidates)
        entity_code = matched[0] if matched else ""
        table_name = matched[1] if matched else ""
        metric_rows: list[dict[str, Any]] = []
        for row in rows:
            raw_code = _normalize_text(row.get("科目代码"))
            if not raw_code:
                continue
            code = _normalize_data_entry_metric_code(entity_code, raw_code) if entity_code else raw_code.upper()
            if not code:
                continue
            metric_rows.append(
                {
                    "metric_code": code,
                    "metric_name": _normalize_text(row.get("科目名称")),
                    "levelLabel": _normalize_text(row.get("科目层级")),
                    "nature": _normalize_text(row.get("科目性质")),
                    "values": _rows_to_data_entry_values(headers, row, year, forecast_month),
                }
            )
        sheets_out.append(
            {
                "sheet_name": sheet_name,
                "entity_code": entity_code,
                "table_name": table_name,
                "matched": bool(entity_code and table_name),
                "row_count": len(metric_rows),
                "metrics": metric_rows,
            }
        )
    return sheets_out


def _normalize_org_level(value: Any) -> int | None:
    text = _normalize_text(value)
    if not text:
        return None
    if text in {"1", "一级"}:
        return 1
    if text in {"2", "二级"}:
        return 2
    if text in {"3", "三级"}:
        return 3
    if text in {"4", "四级"}:
        return 4
    if "一级" in text:
        return 1
    if "二级" in text:
        return 2
    if "三级" in text:
        return 3
    if "四级" in text:
        return 4
    parsed = None
    try:
        parsed = int(text)
    except Exception:
        parsed = None
    if parsed in {1, 2, 3, 4}:
        return parsed
    return None


def _org_node_type_from_excel_level(level_rank: int) -> str:
    return {
        1: "level0",
        2: "level1",
        3: "level2",
        4: "level3",
    }.get(int(level_rank), "level3")


def _org_excel_level_label(node_type: str) -> str:
    mapping = {
        "level0": "一级",
        "level1": "二级",
        "level2": "三级",
        "level3": "四级",
    }
    return mapping.get(str(node_type or "").strip(), "四级")


def _default_org_product_tree_excel_paths() -> list[Path]:
    return [
        settings.business_inputs_dir / "机构及产品.xlsx",
        settings.resources_dir / "机构及产品.xlsx",
    ]


def _resolve_org_product_tree_excel_path() -> Path | None:
    for candidate in _default_org_product_tree_excel_paths():
        if candidate.exists():
            return candidate
    return None


def _parse_org_product_tree_excel(content: bytes, filename: str) -> dict[str, Any]:
    try:
        workbook = load_workbook(filename=BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"无法读取Excel文件：{exc}") from exc
    ws = workbook[workbook.sheetnames[0]]

    header_row_idx = None
    header_map: dict[str, int] = {}
    required_any = [
        {"层级", "编码", "名称"},
        {"层级", "机构及产品编码", "机构及产品名称"},
        {"层级", "机构及产品代码", "机构及产品名称"},
        {"level", "code", "name"},
    ]

    def _match_org_tree_columns(row_map: dict[str, int]) -> tuple[int | None, int | None, int | None, int | None]:
        """按表头文字模糊匹配：层级 / 编码 / 名称 / 上级编码。"""
        level_col = code_col = name_col = parent_code_col = None
        for header, col_idx in row_map.items():
            h = header.strip()
            if not h:
                continue
            hl = h.lower()
            if level_col is None and (h == "层级" or hl in {"level", "level_rank"} or h.endswith("层级")):
                level_col = col_idx
            if parent_code_col is None and ("上级" in h and "编码" in h):
                parent_code_col = col_idx
            if code_col is None and h in {
                "编码",
                "code",
                "Code",
                "层级编码",
                "机构及产品编码",
                "机构及产品代码",
                "机构编码",
                "产品编码",
            }:
                code_col = col_idx
            if name_col is None and h in {"名称", "name", "Name", "层级名称", "机构及产品名称", "机构名称", "产品名称"}:
                name_col = col_idx
        if code_col is None:
            for header, col_idx in row_map.items():
                h = header.strip()
                if "上级" in h:
                    continue
                if ("编码" in h or "代码" in h) and "层级" not in h:
                    code_col = col_idx
                    break
        if name_col is None:
            for header, col_idx in row_map.items():
                h = header.strip()
                if "名称" in h and "上级" not in h:
                    name_col = col_idx
                    break
        return level_col, code_col, name_col, parent_code_col

    for row_idx in range(1, min(ws.max_row, 20) + 1):
        row_map: dict[str, int] = {}
        for col_idx in range(1, ws.max_column + 1):
            h = _normalize_text(ws.cell(row_idx, col_idx).value)
            if h and h not in row_map:
                row_map[h] = col_idx
        if any(req.issubset(set(row_map.keys())) for req in required_any):
            header_row_idx = row_idx
            header_map = row_map
            break
        level_col, code_col, name_col, _parent = _match_org_tree_columns(row_map)
        if level_col and code_col and name_col:
            header_row_idx = row_idx
            header_map = row_map
            break

    def col(*names: str) -> int | None:
        for n in names:
            if n in header_map:
                return header_map[n]
        return None

    level_col = code_col = name_col = parent_code_col = None
    if header_row_idx is not None:
        level_col = col("层级", "level", "Level")
        code_col = col(
            "编码",
            "code",
            "Code",
            "层级编码",
            "机构及产品编码",
            "机构及产品代码",
            "机构编码",
            "产品编码",
        )
        name_col = col(
            "名称",
            "name",
            "Name",
            "层级名称",
            "机构及产品名称",
            "机构名称",
            "产品名称",
        )
        parent_code_col = col("上级编码", "parent_code", "ParentCode")
        if not (level_col and code_col and name_col):
            level_col, code_col, name_col, parent_code_col = _match_org_tree_columns(header_map)
    nodes_by_code: dict[str, dict[str, Any]] = {}
    parent_by_code: dict[str, str | None] = {}

    if level_col and code_col and name_col:
        stack: dict[int, dict[str, Any]] = {}
        for row_idx in range((header_row_idx or 0) + 1, ws.max_row + 1):
            level_rank = _normalize_org_level(ws.cell(row_idx, level_col).value)
            code = _normalize_text(ws.cell(row_idx, code_col).value).upper()
            name = _normalize_text(ws.cell(row_idx, name_col).value)
            if not any([level_rank, code, name]):
                continue
            if not level_rank or not code or not name:
                raise HTTPException(status_code=400, detail=f"第 {row_idx} 行数据不完整：层级/编码/名称不能为空")

            node_type = _org_node_type_from_excel_level(level_rank)
            node_id = "node-root" if level_rank == 1 else f"node-{code.lower()}"
            node = {"id": node_id, "code": code, "name": name, "type": node_type, "children": []}
            nodes_by_code[code] = node

            if parent_code_col:
                parent_code = _normalize_text(ws.cell(row_idx, parent_code_col).value).upper() or None
                parent_by_code[code] = parent_code
            else:
                parent_node = stack.get(level_rank - 1)
                parent_by_code[code] = parent_node["code"] if parent_node else None

            stack[level_rank] = node
            for k in list(stack.keys()):
                if k > level_rank:
                    del stack[k]

        roots = [node for code, node in nodes_by_code.items() if parent_by_code.get(code) in {None, ""}]
        root = None
        if roots:
            roots_sorted = sorted(
                roots,
                key=lambda n: (
                    0 if n.get("type") == "level0" else 1 if n.get("type") == "level1" else 2,
                    n.get("code") or "",
                ),
            )
            root = roots_sorted[0]
        if not root:
            raise HTTPException(status_code=400, detail="未识别到集团根节点（层级=1/一级，编码建议 AAA）")

        for code, node in nodes_by_code.items():
            parent_code = parent_by_code.get(code)
            if not parent_code:
                continue
            parent = nodes_by_code.get(parent_code)
            if not parent:
                continue
            parent.setdefault("children", []).append(node)

        def sort_children(n: dict[str, Any]) -> None:
            children = list(n.get("children") or [])
            children.sort(key=lambda x: x.get("code") or "")
            n["children"] = children
            for c in children:
                sort_children(c)

        sort_children(root)
        return root

    headers, rows = _parse_simple_excel_table(content, filename, "二级机构编码")
    if not rows:
        raise HTTPException(status_code=400, detail="未读取到可用于导入的机构及产品数据")

    root_code = "AA"
    root_name = "微众银行"
    for r in rows:
        root_code = (_normalize_text(r.get("一级主体编码")) or root_code).upper()
        root_name = _normalize_text(r.get("一级主体名称")) or root_name
        break

    root = {"id": "node-root", "code": root_code, "name": root_name, "type": "level1", "children": []}
    org_map: dict[str, dict[str, Any]] = {}
    for r in rows:
        org_code = _normalize_text(r.get("二级机构编码")).upper()
        org_name = _normalize_text(r.get("二级机构名称"))
        prod_code = _normalize_text(r.get("三级产品编码")).upper()
        prod_name = _normalize_text(r.get("三级产品名称"))
        if not org_code or not org_name:
            continue
        org = org_map.get(org_code)
        if not org:
            org = {"id": f"node-{org_code.lower()}", "code": org_code, "name": org_name, "type": "level2", "children": []}
            org_map[org_code] = org
            root["children"].append(org)
        if prod_code and prod_name:
            org["children"].append({"id": f"node-{prod_code.lower()}", "code": prod_code, "name": prod_name, "type": "level3", "children": []})

    root["children"].sort(key=lambda x: x.get("code") or "")
    for org in root["children"]:
        org["children"].sort(key=lambda x: x.get("code") or "")
    return root


def _flatten_org_product_tree(root: dict[str, Any]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []

    def walk(node: dict[str, Any], parent: dict[str, Any] | None) -> None:
        t = str(node.get("type") or "")
        level_label = _org_excel_level_label(t)
        rows.append(
            {
                "层级": level_label,
                "机构及产品代码": _normalize_text(node.get("code")),
                "机构及产品名称": _normalize_text(node.get("name")),
                "上级代码": _normalize_text(parent.get("code")) if parent else "",
                "上级名称": _normalize_text(parent.get("name")) if parent else "",
            }
        )
        for c in list(node.get("children") or []):
            walk(c, node)

    walk(root, None)
    return rows


# Several org/product routers intentionally share this helper module via
# `from ...org_product_helpers import *`. Python excludes single-underscore
# names from star imports unless `__all__` is provided, so expose the shared
# private helpers explicitly to keep the split router modules wired together.
__all__ = [name for name in globals() if not name.startswith("__")]
