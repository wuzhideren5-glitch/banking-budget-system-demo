#!/usr/bin/env python3
"""Restore contextual org-product metric nodes from v03 Excel into MySQL.

Only inserts v03 rows that:
- are missing from ``data_account_metric_node``
- are NOT legacy/retired branches (second segment .05/.99/.90/.91)
- have an existing parent in MySQL, or a parent inserted earlier in the same run

See ``.scratch/sqlite-to-mysql-migration/v03_authority_notes_20260618.md``.

Usage:
    cd apps/api && . .venv/bin/activate
    python scripts/restore_v03_context_nodes_to_mysql.py --dry-run
    python scripts/restore_v03_context_nodes_to_mysql.py --apply
    python scripts/restore_v03_context_nodes_to_mysql.py --verify-only
"""
from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pymysql
from openpyxl import load_workbook

API_ROOT = Path(__file__).resolve().parents[1]
if str(API_ROOT) not in sys.path:
    sys.path.insert(0, str(API_ROOT))

from app.core.config import settings  # noqa: E402
from app.db_bootstrap.runtime_metric_tree import (  # noqa: E402
    _ensure_metric_node_v02_columns,
    _sync_derived_metric_node_identity,
)
from app.org_product_excel_formula import (  # noqa: E402
    build_sheet_formula_context,
    index_sheet_contexts,
    normalize_sheet_lookup_key,
)
from app.routers.org_product_helpers import (  # noqa: E402
    _find_header_row,
    _maybe_swap_metric_code_name_columns,
    _normalize_allow_manual_entry,
    _normalize_metric_code,
    _normalize_metric_value_type,
    _normalize_nature,
    _normalize_rollup_flag,
    _normalize_text,
    _parse_metric_worksheet_basic,
    _prepare_metric_worksheet,
    _resolve_import_sheet_entity_table,
    _sheet_scan_row_limit,
    _ws_cell_value,
)
from scripts.import_v03_formulas_to_mysql import (  # noqa: E402
    DEFAULT_WORKBOOK,
    resolve_v03_db_formulas,
)

VALID_ANNUAL_RULES = frozenset({"SUM", "AVG", "LAST", "WGT", "CALC"})


def _norm(value: Any) -> str:
    return str(value or "").strip()


def _now_iso() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


from app.services.v03_metric_node_catalog import (  # noqa: E402
    IMPLICIT_GROUP_PARENTS,
    ImplicitGroupSpec,
    code_depth,
    infer_implicit_groups_for_codes,
    is_v03_mirror_duplicate_row,
    is_v03_stale_node_code,
    local_metric_code,
    merge_v03_node_payload,
    parent_code,
    product_code,
    repair_implicit_group_nodes,
)

@dataclass(frozen=True)
class V03NodeRecord:
    node_code: str
    node_name: str
    entity_code: str
    table_name: str
    sheet_name: str
    nature: str
    value_type: str
    allow_manual_entry: int
    horizontal_rollup: int
    vertical_rollup: int
    budget_formula: str
    actual_formula: str
    annual_agg_rule: str
    has_children: bool


@dataclass(frozen=True)
class GroupInsertPlan:
    spec: ImplicitGroupSpec
    sort_order: int


@dataclass(frozen=True)
class NodeInsertPlan:
    record: V03NodeRecord
    parent_code: str | None
    metric_table_name: str
    functional_group_code: str
    product_code: str
    local_metric_code: str
    logic_code: str
    level: int
    node_type: str
    sort_order: int


def _iter_metric_nodes(nodes: list[dict[str, Any]]):
    stack = list(nodes)
    while stack:
        node = stack.pop(0)
        if not isinstance(node, dict):
            continue
        yield node
        children = node.get("children")
        if isinstance(children, list):
            stack[0:0] = [child for child in children if isinstance(child, dict)]


def _build_workbook_formula_contexts(wb) -> dict[str, Any]:
    contexts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        _prepare_metric_worksheet(ws)
        resolved = _resolve_import_sheet_entity_table(sheet_name, [], strict=False)
        if not resolved:
            continue
        entity_code, table_name, _entity_name = resolved
        header_row_idx, header_map, _header_mode, _header_source = _find_header_row(
            ws,
            entity_code,
            strict=False,
            sheet_title=sheet_name,
        )
        code_col = header_map.get("科目代码") if header_row_idx else None
        name_col = header_map.get("科目名称") if header_row_idx else None
        if header_row_idx and code_col and name_col:
            code_col, _name_col = _maybe_swap_metric_code_name_columns(
                ws,
                header_row_idx=header_row_idx,
                entity_code=entity_code,
                code_col=code_col,
                name_col=name_col,
            )
        if not header_row_idx or not code_col:
            continue
        row_limit = _sheet_scan_row_limit(ws, header_row_idx)
        contexts.append(
            build_sheet_formula_context(
                sheet_name,
                entity_code,
                table_name,
                header_row_idx,
                code_col,
                lambda r, c, _ws=ws: _ws_cell_value(_ws, r, c),
                _normalize_metric_code,
                row_limit,
            )
        )
    return index_sheet_contexts(contexts)


def load_v03_restore_candidates(workbook_path: Path) -> dict[str, V03NodeRecord]:
    if not workbook_path.exists():
        raise FileNotFoundError(f"v03 workbook not found: {workbook_path}")

    wb = load_workbook(workbook_path, data_only=False, read_only=False)
    all_sheet_contexts = _build_workbook_formula_contexts(wb)
    rules_by_code: dict[str, str] = {}
    parsed_nodes: dict[str, dict[str, Any]] = {}
    meta_by_code: dict[str, dict[str, str]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        _prepare_metric_worksheet(ws)
        resolved = _resolve_import_sheet_entity_table(sheet_name, [], strict=False)
        if not resolved:
            continue
        entity_code, table_name, _entity_name = resolved
        header_row_idx, header_map, _header_mode, _header_source = _find_header_row(
            ws,
            entity_code,
            strict=False,
            sheet_title=sheet_name,
        )
        code_col = header_map.get("科目代码") if header_row_idx else None
        rule_col = header_map.get("规则") if header_row_idx else None
        if header_row_idx and code_col and rule_col:
            for row_idx in range(header_row_idx + 1, _sheet_scan_row_limit(ws, header_row_idx) + 1):
                code = _normalize_metric_code(entity_code, _ws_cell_value(ws, row_idx, code_col)).upper()
                rule = _norm(_ws_cell_value(ws, row_idx, rule_col)).upper()
                if code and rule in VALID_ANNUAL_RULES:
                    rules_by_code[code] = rule

        sheet_ctx = all_sheet_contexts.get(normalize_sheet_lookup_key(sheet_name))
        metrics, _row_count, parse_error, _header_map = _parse_metric_worksheet_basic(
            ws,
            f"restore-{sheet_name}",
            entity_code=entity_code,
            strict=False,
            sheet_formula_context=sheet_ctx,
            all_sheet_contexts=all_sheet_contexts,
        )
        if parse_error or not metrics:
            continue
        for node in _iter_metric_nodes(metrics):
            code = _norm(node.get("code")).upper()
            name = _norm(node.get("name"))
            if not code or not name:
                continue
            if is_v03_mirror_duplicate_row(sheet_name, code):
                continue
            parsed_nodes[code] = node
            incoming_meta = {
                "entity_code": entity_code,
                "table_name": table_name,
                "sheet_name": sheet_name,
            }
            existing_meta = meta_by_code.get(code)
            meta_by_code[code] = (
                merge_v03_node_payload(existing_meta, incoming_meta)
                if existing_meta
                else incoming_meta
            )

    child_codes: set[str] = set()
    for code in parsed_nodes:
        parent = parent_code(code)
        if parent:
            child_codes.add(parent)

    records: dict[str, V03NodeRecord] = {}
    for code, node in parsed_nodes.items():
        if is_v03_stale_node_code(code):
            continue
        if code in IMPLICIT_GROUP_PARENTS:
            continue
        meta = meta_by_code[code]
        nature = _normalize_nature(node.get("nature"))
        budget_formula, actual_formula = resolve_v03_db_formulas(node)
        records[code] = V03NodeRecord(
            node_code=code,
            node_name=_norm(node.get("name")),
            entity_code=meta["entity_code"],
            table_name=meta["table_name"],
            sheet_name=meta["sheet_name"],
            nature=nature,
            value_type=_normalize_metric_value_type(node.get("value_type"), nature),
            allow_manual_entry=_normalize_allow_manual_entry(node.get("allow_manual_entry"), 1),
            horizontal_rollup=_normalize_rollup_flag(node.get("horizontal_rollup")),
            vertical_rollup=_normalize_rollup_flag(node.get("vertical_rollup")),
            budget_formula=budget_formula,
            actual_formula=actual_formula,
            annual_agg_rule=rules_by_code.get(code, ""),
            has_children=code in child_codes,
        )
    return records


def _mysql_connect() -> pymysql.connections.Connection:
    return pymysql.connect(
        host=settings.MYSQL_HOST,
        port=int(settings.MYSQL_PORT),
        user=settings.MYSQL_USER,
        password=settings.MYSQL_PASSWORD or "",
        database=settings.MYSQL_DATABASE,
        charset="utf8mb4",
        autocommit=False,
    )


def load_active_node_codes(conn: pymysql.connections.Connection) -> set[str]:
    with conn.cursor() as cur:
        cur.execute("SELECT node_code FROM data_account_metric_node WHERE is_active = 1")
        return {str(row[0]).strip().upper() for row in cur.fetchall() if str(row[0] or "").strip()}


def _load_parent_template(
    conn: pymysql.connections.Connection,
    parent_code: str,
) -> dict[str, Any] | None:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT metric_table_name, functional_group_code, product_code, sort_order
            FROM data_account_metric_node
            WHERE node_code = %s AND is_active = 1
            """,
            (parent_code,),
        )
        row = cur.fetchone()
    if not row:
        return None
    return {
        "metric_table_name": _norm(row[0]),
        "functional_group_code": _norm(row[1]),
        "product_code": _norm(row[2]),
        "sort_order": int(row[3] or 0),
    }


def _next_sort_order(conn: pymysql.connections.Connection, parent_code: str | None, level: int) -> int:
    with conn.cursor() as cur:
        if parent_code:
            cur.execute(
                """
                SELECT COALESCE(MAX(sort_order), 0)
                FROM data_account_metric_node
                WHERE parent_code = %s
                """,
                (parent_code,),
            )
        else:
            cur.execute(
                """
                SELECT COALESCE(MAX(sort_order), 0)
                FROM data_account_metric_node
                WHERE parent_code IS NULL
                """
            )
        current = int(cur.fetchone()[0] or 0)
    return max(current + 10, level * 10)


def _resolve_parent_template(
    conn: pymysql.connections.Connection,
    parent: str,
    *,
    planned_by_code: dict[str, NodeInsertPlan],
    planned_groups: dict[str, GroupInsertPlan],
    fallback_table_name: str,
) -> dict[str, Any] | None:
    template = _load_parent_template(conn, parent)
    if template:
        return template
    group_plan = planned_groups.get(parent)
    if group_plan:
        table_name = group_plan.spec.metric_table_name or fallback_table_name
        return {
            "metric_table_name": table_name,
            "functional_group_code": table_name,
            "product_code": product_code(parent),
            "sort_order": group_plan.sort_order,
        }
    parent_plan = planned_by_code.get(parent)
    if not parent_plan:
        return None
    return {
        "metric_table_name": parent_plan.metric_table_name or fallback_table_name,
        "functional_group_code": parent_plan.functional_group_code or fallback_table_name,
        "product_code": parent_plan.product_code,
        "sort_order": parent_plan.sort_order,
    }


def build_insert_plans(
    conn: pymysql.connections.Connection,
    candidates: dict[str, V03NodeRecord],
    active_codes: set[str],
) -> tuple[list[GroupInsertPlan], list[NodeInsertPlan], list[str]]:
    missing_codes = sorted(
        [code for code in candidates if code not in active_codes],
        key=lambda code: (code_depth(code), code),
    )
    known_codes = set(active_codes)
    metric_plans: list[NodeInsertPlan] = []
    planned_by_code: dict[str, NodeInsertPlan] = {}
    skipped: list[str] = []

    target_codes = set(missing_codes)
    group_specs = infer_implicit_groups_for_codes(target_codes)
    group_plans: list[GroupInsertPlan] = []
    for spec in group_specs:
        if spec.node_code in known_codes:
            continue
        group_plans.append(
            GroupInsertPlan(
                spec=spec,
                sort_order=_next_sort_order(conn, spec.parent_code, spec.level),
            )
        )
        known_codes.add(spec.node_code)
    planned_groups = {plan.spec.node_code: plan for plan in group_plans}

    for code in missing_codes:
        record = candidates[code]
        parent = parent_code(code)
        if parent and parent not in known_codes:
            skipped.append(f"{code}: parent {parent} missing")
            continue

        if parent:
            template = _resolve_parent_template(
                conn,
                parent,
                planned_by_code=planned_by_code,
                planned_groups=planned_groups,
                fallback_table_name=record.table_name,
            )
            if not template:
                skipped.append(f"{code}: parent template missing for {parent}")
                continue
            metric_table_name = template["metric_table_name"] or record.table_name
            functional_group_code = template["functional_group_code"] or record.table_name
        else:
            metric_table_name = record.table_name
            functional_group_code = record.table_name

        product_code_value = product_code(code)
        local_code = local_metric_code(code)
        level = code_depth(code)
        node_type = "GROUP" if record.has_children else "METRIC"
        sort_order = _next_sort_order(conn, parent, level)
        plan = NodeInsertPlan(
            record=record,
            parent_code=parent,
            metric_table_name=metric_table_name,
            functional_group_code=functional_group_code,
            product_code=product_code_value,
            local_metric_code=local_code,
            logic_code=local_code,
            level=level,
            node_type=node_type,
            sort_order=sort_order,
        )
        metric_plans.append(plan)
        planned_by_code[code] = plan
        known_codes.add(code)
    return group_plans, metric_plans, skipped


def apply_group_insert_plans(conn: pymysql.connections.Connection, plans: list[GroupInsertPlan]) -> int:
    inserted = 0
    now = _now_iso()
    with conn.cursor() as cur:
        for plan in plans:
            spec = plan.spec
            cur.execute(
                """
                INSERT INTO data_account_metric_node (
                  node_code, node_name, parent_code, product_code, local_metric_code, logic_code,
                  functional_group_code, metric_table_name, level, node_type,
                  horizontal_rollup, vertical_rollup, runtime_account_enabled,
                  budget_formula, actual_formula, need_calc, formula_calc_mode,
                  allow_manual_entry, value_type, nature, annual_agg_rule, sort_order,
                  is_active, remark, created_at, updated_at
                ) VALUES (
                  %s, %s, %s, %s, %s, %s,
                  %s, %s, %s, %s,
                  0, 0, 0,
                  NULL, NULL, 0, 0,
                  1, '金额', '其他', '', %s,
                  1, %s, %s, %s
                )
                """,
                (
                    spec.node_code,
                    spec.node_name,
                    spec.parent_code,
                    product_code(spec.node_code),
                    local_metric_code(spec.node_code),
                    local_metric_code(spec.node_code),
                    spec.metric_table_name,
                    spec.metric_table_name,
                    spec.level,
                    spec.node_type,
                    plan.sort_order,
                    f"来源：v03隐式分组节点恢复；{spec.metric_table_name}",
                    now,
                    now,
                ),
            )
            inserted += int(cur.rowcount or 0)
    conn.commit()
    return inserted


def apply_insert_plans(conn: pymysql.connections.Connection, plans: list[NodeInsertPlan]) -> int:
    inserted = 0
    now = _now_iso()
    with conn.cursor() as cur:
        for plan in plans:
            record = plan.record
            cur.execute(
                """
                INSERT INTO data_account_metric_node (
                  node_code, node_name, parent_code, product_code, local_metric_code, logic_code,
                  functional_group_code, metric_table_name, level, node_type,
                  horizontal_rollup, vertical_rollup, runtime_account_enabled,
                  budget_formula, actual_formula, need_calc, formula_calc_mode,
                  allow_manual_entry, value_type, nature, annual_agg_rule, sort_order,
                  is_active, remark, created_at, updated_at
                ) VALUES (
                  %s, %s, %s, %s, %s, %s,
                  %s, %s, %s, %s,
                  %s, %s, 1,
                  %s, %s, 0, 0,
                  %s, %s, %s, %s, %s,
                  1, %s, %s, %s
                )
                """,
                (
                    record.node_code,
                    record.node_name,
                    plan.parent_code,
                    plan.product_code,
                    plan.local_metric_code,
                    plan.logic_code,
                    plan.functional_group_code,
                    plan.metric_table_name,
                    plan.level,
                    plan.node_type,
                    record.horizontal_rollup,
                    record.vertical_rollup,
                    record.budget_formula or None,
                    record.actual_formula or None,
                    record.allow_manual_entry,
                    record.value_type,
                    record.nature,
                    record.annual_agg_rule,
                    plan.sort_order,
                    f"来源：v03上下文节点恢复；{record.entity_code}/{record.table_name}/{record.sheet_name}",
                    now,
                    now,
                ),
            )
            inserted += int(cur.rowcount or 0)
    conn.commit()
    return inserted


def verify_restore(
    candidates: dict[str, V03NodeRecord],
    active_codes: set[str],
) -> tuple[list[str], list[str]]:
    stale_present = sorted(code for code in active_codes if is_v03_stale_node_code(code))
    still_missing = sorted(
        code
        for code, record in candidates.items()
        if not is_v03_stale_node_code(code) and code not in active_codes
    )
    return still_missing, stale_present


def backfill_nature_from_v03(
    conn: pymysql.connections.Connection,
    records: dict[str, V03NodeRecord],
) -> int:
    """Write v03 科目性质 into data_account_metric_node.nature for all matching active nodes."""
    _ensure_metric_node_v02_columns(conn)
    now = _now_iso()
    updated = 0
    with conn.cursor() as cur:
        for code, record in records.items():
            if is_v03_stale_node_code(code):
                continue
            cur.execute(
                """
                UPDATE data_account_metric_node
                SET nature = %s, updated_at = %s
                WHERE node_code = %s AND is_active = 1
                """,
                (record.nature or "其他", now, code),
            )
            updated += int(cur.rowcount or 0)
    conn.commit()
    return updated


def verify_nature_backfill(
    conn: pymysql.connections.Connection,
    records: dict[str, V03NodeRecord],
) -> tuple[int, int, list[str]]:
    """Return (matched, other_default, sample_mismatches)."""
    _ensure_metric_node_v02_columns(conn)
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT node_code, COALESCE(nature, '')
            FROM data_account_metric_node
            WHERE is_active = 1
            """
        )
        db_nature = {str(row[0]).strip().upper(): _norm(row[1]) for row in cur.fetchall()}
    other_count = sum(1 for value in db_nature.values() if value in {"", "其他"})
    mismatches: list[str] = []
    for code, record in records.items():
        if is_v03_stale_node_code(code):
            continue
        actual = db_nature.get(code)
        if actual is None:
            continue
        expected = record.nature or "其他"
        if actual != expected:
            mismatches.append(f"{code}: db={actual or '-'} v03={expected}")
    matched = len(records) - len(mismatches)
    return matched, other_count, mismatches


def _print_plans(
    group_plans: list[GroupInsertPlan],
    metric_plans: list[NodeInsertPlan],
    skipped: list[str],
    *,
    limit: int = 15,
) -> None:
    print(f"group insert plans: {len(group_plans)}")
    for plan in group_plans[:limit]:
        spec = plan.spec
        print(f"  [GROUP] {spec.node_code} parent={spec.parent_code or '-'} table={spec.metric_table_name}")
    print(f"metric insert plans: {len(metric_plans)}")
    for plan in metric_plans[:limit]:
        record = plan.record
        print(
            f"  {record.node_code}  {record.node_name}  "
            f"parent={plan.parent_code or '-'}  rule={record.annual_agg_rule or '-'}"
        )
    if len(metric_plans) > limit:
        print(f"  ... and {len(metric_plans) - limit} more metrics")
    if skipped:
        print(f"skipped: {len(skipped)}")
        for item in skipped[:limit]:
            print(f"  {item}")
        if len(skipped) > limit:
            print(f"  ... and {len(skipped) - limit} more skipped")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    mode = parser.add_mutually_exclusive_group()
    mode.add_argument("--dry-run", action="store_true", help="Show planned inserts only (default)")
    mode.add_argument("--apply", action="store_true", help="Insert missing nodes into MySQL")
    mode.add_argument("--verify-only", action="store_true", help="Verify eligible v03 nodes exist in MySQL")
    mode.add_argument(
        "--backfill-nature",
        action="store_true",
        help="Backfill 科目性质 from v03 into data_account_metric_node.nature",
    )
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK, help="Path to v03 workbook")
    args = parser.parse_args(argv)

    if not args.apply and not args.verify_only:
        args.dry_run = True

    candidates = load_v03_restore_candidates(args.workbook)
    eligible = {code: record for code, record in candidates.items() if not is_v03_stale_node_code(code)}
    print(f"v03 parsed nodes: {len(candidates)}")
    print(f"eligible restore candidates: {len(eligible)}")

    with _mysql_connect() as conn:
        active_before = load_active_node_codes(conn)
        print(f"mysql active nodes before: {len(active_before)}")

        if args.backfill_nature:
            updated = backfill_nature_from_v03(conn, candidates)
            matched, other_count, mismatches = verify_nature_backfill(conn, eligible)
            print(f"nature rows updated: {updated}")
            print(f"v03 eligible nodes still at default 其他/empty in db: {other_count} total nodes with 其他")
            print(f"v03 nature mismatches after backfill: {len(mismatches)}")
            if mismatches[:20]:
                print("sample mismatches:", "; ".join(mismatches[:20]))
            return 1 if mismatches else 0

        if args.verify_only:
            still_missing, stale_present = verify_restore(eligible, active_before)
            print(f"still missing eligible nodes: {len(still_missing)}")
            if still_missing[:20]:
                print("sample missing:", ", ".join(still_missing[:20]))
            print(f"stale branch nodes present (informational): {len(stale_present)}")
            return 1 if still_missing else 0

        group_plans, metric_plans, skipped = build_insert_plans(conn, eligible, active_before)
        if args.dry_run:
            _print_plans(group_plans, metric_plans, skipped)
            return 0

        total_inserted = 0
        pass_no = 0
        last_skipped: list[str] = []
        while True:
            pass_no += 1
            active_codes = load_active_node_codes(conn)
            group_plans, metric_plans, skipped = build_insert_plans(conn, eligible, active_codes)
            last_skipped = skipped
            if not group_plans and not metric_plans:
                break
            batch_inserted = 0
            if group_plans:
                batch_inserted += apply_group_insert_plans(conn, group_plans)
            if metric_plans:
                batch_inserted += apply_insert_plans(conn, metric_plans)
            total_inserted += batch_inserted
            print(
                f"pass {pass_no}: inserted {batch_inserted} "
                f"(groups={len(group_plans)}, metrics={len(metric_plans)})"
            )
            if pass_no >= 8:
                break

        _sync_derived_metric_node_identity(conn)
        repaired = repair_implicit_group_nodes(conn)
        if repaired:
            print(f"repaired implicit group rows: {repaired}")
        conn.commit()
        active_after = load_active_node_codes(conn)
        still_missing, _stale_present = verify_restore(eligible, active_after)
        print(f"rows inserted total: {total_inserted}")
        print(f"mysql active nodes after: {len(active_after)}")
        print(f"still missing eligible nodes: {len(still_missing)}")
        if still_missing[:15]:
            print("still missing sample:", ", ".join(still_missing[:15]))
        if last_skipped:
            print(f"skipped (last pass): {len(last_skipped)}")
        return 1 if still_missing else 0


if __name__ == "__main__":
    raise SystemExit(main())
