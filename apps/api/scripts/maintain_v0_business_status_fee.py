#!/usr/bin/env python3
"""Maintain v0 workbook: AA业务状况表 fee/投入产出 block + copy to product sheets.

Usage:
    cd apps/api && . .venv/bin/activate
    python scripts/maintain_v0_business_status_fee.py --dry-run
    python scripts/maintain_v0_business_status_fee.py --apply
    python scripts/maintain_v0_business_status_fee.py --reconfigure-90-formulas --apply
"""
from __future__ import annotations

import argparse
import re
import shutil
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

API_ROOT = Path(__file__).resolve().parents[1]
if str(API_ROOT) not in sys.path:
    sys.path.insert(0, str(API_ROOT))

from app.core.config import settings  # noqa: E402
from app.routers.org_product_helpers import (  # noqa: E402
    _find_header_row,
    _normalize_metric_code,
    _resolve_import_sheet_entity_table,
    _sheet_scan_row_limit,
    _ws_cell_value,
)

DEFAULT_WORKBOOK = settings.business_inputs_dir / "机构及产品指标（公式配置） - v0.xlsx"
SOURCE_SHEET = "AA业务状况表"
SOURCE_ENTITY = "AA"
LEVEL_LABELS = ("一级", "二级", "三级", "四级", "五级", "六级")

SKIP_SHEETS = frozenset(
    {
        "AA业务状况表",
        "AA资产负债表（余额）",
        "AA资产负债表（日均）",
        "AA资产质量表",
        "AA利息净收入表",
        "AB微众科技",
    }
)

# Columns in standard header (1-based): A=层级 B=性质 C=代码 D=名称 E=录入 F=年预算 ...
COL_LEVEL = 1
COL_NATURE = 2
COL_CODE = 3
COL_NAME = 4
COL_BUDGET_FORMULA = 6
COL_HROLLUP = 11
COL_LOGIC = 13
COL_RULE = 14
MAX_COL = 14


def _norm(value: Any) -> str:
    return str(value or "").strip()


def _level_from_code(code: str) -> str:
    parts = [p for p in _norm(code).upper().split(".") if p]
    if len(parts) < 2:
        return "一级"
    depth = min(max(len(parts) - 1, 1), len(LEVEL_LABELS))
    return LEVEL_LABELS[depth - 1]


def _read_section_rows(ws, entity_code: str, start_row: int, end_row: int) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for row_idx in range(start_row, end_row + 1):
        row = [
            ws.cell(row_idx, col).value
            for col in range(1, MAX_COL + 1)
        ]
        rows.append(row)
    return rows


def _codes_in_rows(rows: list[list[Any]]) -> list[str]:
    codes: list[str] = []
    for row in rows:
        code = _norm(row[COL_CODE - 1]).upper()
        if code and code not in {"科目代码"}:
            codes.append(code)
    return codes


def _build_children_map(codes: list[str]) -> dict[str, list[str]]:
    code_set = set(codes)
    children: dict[str, list[str]] = defaultdict(list)
    for code in codes:
        if "." not in code:
            continue
        parent = code.rsplit(".", 1)[0]
        if parent in code_set:
            children[parent].append(code)
    return children


def _child_sum_formula(children: list[str]) -> str:
    return "+".join(children)


def _replace_entity_prefix(text: Any, entity_code: str) -> Any:
    if text is None:
        return None
    if isinstance(text, (int, float)):
        return text
    text = str(text)
    return re.sub(rf"\b{re.escape(SOURCE_ENTITY)}\.", f"{entity_code}.", text)


def _transform_row(row: list[Any], entity_code: str) -> list[Any]:
    out = list(row)
    out[COL_CODE - 1] = _replace_entity_prefix(_norm(out[COL_CODE - 1]), entity_code)
    out[COL_BUDGET_FORMULA - 1] = _replace_entity_prefix(out[COL_BUDGET_FORMULA - 1], entity_code)
    code = _norm(out[COL_CODE - 1]).upper()
    if code:
        out[COL_LEVEL - 1] = _level_from_code(code)
    out[COL_HROLLUP - 1] = "是"
    return out


def _enrich_template_rows(
    rows: list[list[Any]],
    *,
    overwrite_formulas: bool = False,
) -> tuple[list[list[Any]], dict[str, int]]:
    stats = {"levels_set": 0, "formulas_set": 0, "formulas_cleared": 0, "hrollup_set": 0}
    codes = _codes_in_rows(rows)
    children = _build_children_map(codes)
    enriched: list[list[Any]] = []
    for row in rows:
        new_row = list(row)
        code = _norm(new_row[COL_CODE - 1]).upper()
        if not code:
            enriched.append(new_row)
            continue
        level = _level_from_code(code)
        if _norm(new_row[COL_LEVEL - 1]) != level:
            stats["levels_set"] += 1
        new_row[COL_LEVEL - 1] = level
        if _norm(new_row[COL_HROLLUP - 1]) != "是":
            stats["hrollup_set"] += 1
        new_row[COL_HROLLUP - 1] = "是"
        child_codes = children.get(code, [])
        if child_codes:
            expected = _child_sum_formula(child_codes)
            if overwrite_formulas or not _norm(new_row[COL_BUDGET_FORMULA - 1]):
                if _norm(new_row[COL_BUDGET_FORMULA - 1]) != expected:
                    stats["formulas_set"] += 1
                new_row[COL_BUDGET_FORMULA - 1] = expected
        elif overwrite_formulas and _norm(new_row[COL_BUDGET_FORMULA - 1]):
            new_row[COL_BUDGET_FORMULA - 1] = None
            stats["formulas_cleared"] += 1
        enriched.append(new_row)
    return enriched, stats


def _find_fee_start_row(ws, entity_code: str) -> int | None:
    hr, hm, *_ = _find_header_row(ws, entity_code, strict=False, sheet_title=str(getattr(ws, "title", "")))
    code_col = hm.get("科目代码") if hr else None
    if not hr or not code_col:
        return None
    for row_idx in range(hr + 1, _sheet_scan_row_limit(ws, hr) + 1):
        code = _normalize_metric_code(entity_code, _ws_cell_value(ws, row_idx, code_col)).upper()
        parts = code.split(".") if code else []
        if len(parts) >= 2 and parts[1] == "90":
            return row_idx
    return None


def _write_rows(ws, start_row: int, rows: list[list[Any]]) -> None:
    for offset, row in enumerate(rows):
        row_idx = start_row + offset
        for col, value in enumerate(row, start=1):
            ws.cell(row_idx, col, value)


def _find_fee_block(
    ws,
    entity_code: str,
) -> tuple[int, int, list[tuple[int, str]]] | None:
    hr, hm, *_ = _find_header_row(ws, entity_code, strict=False, sheet_title=str(getattr(ws, "title", "")))
    code_col = hm.get("科目代码") if hr else None
    if not hr or not code_col:
        return None
    start_row = _find_fee_start_row(ws, entity_code)
    if start_row is None:
        return None
    root_prefix = f"{_norm(entity_code).upper()}.90"
    block: list[tuple[int, str]] = []
    for row_idx in range(start_row, _sheet_scan_row_limit(ws, hr) + 1):
        code = _normalize_metric_code(entity_code, _ws_cell_value(ws, row_idx, code_col)).upper()
        if not code:
            if block:
                break
            continue
        if not code.startswith(root_prefix):
            if block:
                break
            continue
        block.append((row_idx, code))
    if not block:
        return None
    return start_row, block[-1][0], block


def reconfigure_90_formulas(workbook_path: Path, *, apply: bool) -> dict[str, Any]:
    if not workbook_path.exists():
        raise FileNotFoundError(workbook_path)

    wb = load_workbook(workbook_path, data_only=False)
    report: dict[str, Any] = {"sheets": {}}

    target_sheets = [
        sn
        for sn in wb.sheetnames
        if _resolve_import_sheet_entity_table(sn, [], strict=False)
        and _resolve_import_sheet_entity_table(sn, [], strict=False)[1] == "业务状况表"
    ]

    for sheet_name in target_sheets:
        resolved = _resolve_import_sheet_entity_table(sheet_name, [], strict=False)
        if not resolved:
            continue
        entity_code = resolved[0]
        ws = wb[sheet_name]
        block_info = _find_fee_block(ws, entity_code)
        if not block_info:
            report["sheets"][sheet_name] = {"status": "skip_no_90"}
            continue
        start_row, end_row, block = block_info
        rows = _read_section_rows(ws, entity_code, start_row, end_row)
        enriched, stats = _enrich_template_rows(rows, overwrite_formulas=True)
        if apply:
            _write_rows(ws, start_row, enriched)
        report["sheets"][sheet_name] = {
            "entity": entity_code,
            "start_row": start_row,
            "end_row": end_row,
            "nodes": len(block),
            **stats,
        }

    if apply:
        backup_dir = settings.data_dir / "backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = backup_dir / f"v0_metric_workbook_before_90_formulas_{stamp}.xlsx"
        shutil.copy2(workbook_path, backup_path)
        wb.save(workbook_path)
        report["backup"] = str(backup_path)

    return report


def maintain_workbook(workbook_path: Path, *, apply: bool) -> dict[str, Any]:
    if not workbook_path.exists():
        raise FileNotFoundError(workbook_path)

    wb = load_workbook(workbook_path, data_only=False)
    if SOURCE_SHEET not in wb.sheetnames:
        raise KeyError(SOURCE_SHEET)

    ws_aa = wb[SOURCE_SHEET]
    template_rows = _read_section_rows(ws_aa, SOURCE_ENTITY, 68, 137)
    template_rows, aa_stats = _enrich_template_rows(template_rows)

    report: dict[str, Any] = {
        "aa_stats": aa_stats,
        "sheets": {},
    }

    if apply:
        _write_rows(ws_aa, 68, template_rows)
        # Row 67 is a visual separator; keep it blank.
        for col in range(1, MAX_COL + 1):
            ws_aa.cell(67, col).value = None

    target_sheets = [
        sn
        for sn in wb.sheetnames
        if sn not in SKIP_SHEETS and _resolve_import_sheet_entity_table(sn, [], strict=False)
    ]

    for sheet_name in target_sheets:
        resolved = _resolve_import_sheet_entity_table(sheet_name, [], strict=False)
        if not resolved:
            continue
        entity_code = resolved[0]
        ws = wb[sheet_name]
        start_row = _find_fee_start_row(ws, entity_code)
        if start_row is None:
            report["sheets"][sheet_name] = {"status": "skip_no_90"}
            continue
        old_end = ws.max_row or start_row
        old_count = old_end - start_row + 1
        new_rows = [_transform_row(row, entity_code) for row in template_rows]
        # re-apply formula enrichment on transformed rows (prefix already replaced)
        new_rows, child_stats = _enrich_template_rows(new_rows)
        if apply:
            delete_count = old_end - start_row + 1
            if delete_count > 0:
                ws.delete_rows(start_row, delete_count)
            _write_rows(ws, start_row, new_rows)
        report["sheets"][sheet_name] = {
            "entity": entity_code,
            "start_row": start_row,
            "old_rows": old_count,
            "new_rows": len(new_rows),
            "formulas_set": child_stats["formulas_set"],
            "levels_set": child_stats["levels_set"],
        }

    if apply:
        backup_dir = settings.data_dir / "backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = backup_dir / f"v0_metric_workbook_before_fee_maintain_{stamp}.xlsx"
        shutil.copy2(workbook_path, backup_path)
        wb.save(workbook_path)
        report["backup"] = str(backup_path)

    return report


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    mode = parser.add_mutually_exclusive_group()
    mode.add_argument("--dry-run", action="store_true")
    mode.add_argument("--apply", action="store_true")
    parser.add_argument(
        "--reconfigure-90-formulas",
        action="store_true",
        help="Rewrite .90 block 年预算公式 from direct-child tree structure",
    )
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    args = parser.parse_args(argv)
    if not args.apply:
        args.dry_run = True

    if args.reconfigure_90_formulas:
        report = reconfigure_90_formulas(args.workbook, apply=args.apply)
        print("reconfigure .90 formulas:")
        for sheet, info in sorted(report["sheets"].items()):
            print(f"  {sheet}: {info}")
        if args.dry_run:
            print("(dry-run: no file written)")
        elif report.get("backup"):
            print(f"backup: {report['backup']}")
        return 0

    report = maintain_workbook(args.workbook, apply=args.apply)
    print("AA template enrich:", report["aa_stats"])
    for sheet, info in sorted(report["sheets"].items()):
        print(f"{sheet}: {info}")
    if args.dry_run:
        print("(dry-run: no file written)")
    elif report.get("backup"):
        print(f"backup: {report['backup']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
