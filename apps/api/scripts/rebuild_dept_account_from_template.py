#!/usr/bin/env python3
"""Rebuild dept_account from the confirmed department hierarchy template.

This is a data-only maintenance script. It intentionally keeps the Panpan
department-maintenance structure as:

    entity -> business group -> expense owner department

Expense occurrence departments remain in expense_framework_budget_department
and are not written into dept_account.
"""

from __future__ import annotations

import argparse
import json
import shutil
import sqlite3
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class DeptRow:
    dept_code: str
    dept_name: str
    entity_name: str
    parent_code: str | None
    level: int
    is_leaf: int
    source_row: int
    source_type: str


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _backup_db(db_path: Path, backup_dir: Path) -> Path:
    backup_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = backup_dir / f"{db_path.stem}_before_dept_account_rebuild_{timestamp}{db_path.suffix}"
    shutil.copy2(db_path, backup_path)
    return backup_path


def _load_template(template_path: Path) -> list[DeptRow]:
    if not template_path.exists():
        raise FileNotFoundError(template_path)
    wb = load_workbook(template_path, data_only=True, read_only=True)
    if "数据模版" not in wb.sheetnames:
        raise ValueError("部门架构模板缺少“数据模版”工作表")
    ws = wb["数据模版"]
    headers = [_text(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_idx = {header: idx for idx, header in enumerate(headers) if header}
    required = {"主体", "事业群代码", "事业群名称", "费用归属部门代码", "费用归属部门名称"}
    missing = sorted(required - set(header_idx))
    if missing:
        raise ValueError(f"部门架构模板缺少字段: {', '.join(missing)}")

    by_code: dict[str, DeptRow] = {}
    current_entity = ""
    current_group_code = ""
    current_group_name = ""
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        entity_name = _text(row[header_idx["主体"]])
        group_code = _text(row[header_idx["事业群代码"]]).upper()
        group_name = _text(row[header_idx["事业群名称"]])
        owner_code = _text(row[header_idx["费用归属部门代码"]]).upper()
        owner_name = _text(row[header_idx["费用归属部门名称"]])

        if entity_name:
            current_entity = entity_name
        if group_code or group_name:
            current_group_code = group_code
            current_group_name = group_name
            if not group_code or not group_name:
                raise ValueError(f"第 {row_idx} 行事业群代码/名称不完整")
            by_code[group_code] = DeptRow(
                dept_code=group_code,
                dept_name=group_name,
                entity_name=current_entity or "微众银行",
                parent_code=None,
                level=1,
                is_leaf=0,
                source_row=row_idx,
                source_type="事业群",
            )
        if owner_code or owner_name:
            if not owner_code or not owner_name:
                raise ValueError(f"第 {row_idx} 行费用归属部门代码/名称不完整")
            if not current_group_code:
                raise ValueError(f"第 {row_idx} 行费用归属部门无法继承事业群")
            by_code[owner_code] = DeptRow(
                dept_code=owner_code,
                dept_name=owner_name,
                entity_name=current_entity or "微众银行",
                parent_code=current_group_code,
                level=2,
                is_leaf=1,
                source_row=row_idx,
                source_type="费用归属部门",
            )

    return sorted(by_code.values(), key=lambda item: (item.entity_name, item.level, item.dept_code))


def _load_current(conn: sqlite3.Connection) -> dict[str, dict[str, Any]]:
    rows = conn.execute(
        """
        SELECT dept_code, dept_name, entity_name, parent_code, level, is_leaf
        FROM dept_account
        ORDER BY dept_code
        """
    ).fetchall()
    return {
        str(row["dept_code"]): {
            "dept_code": str(row["dept_code"]),
            "dept_name": str(row["dept_name"]),
            "entity_name": str(row["entity_name"] or "微众银行"),
            "parent_code": row["parent_code"],
            "level": int(row["level"]),
            "is_leaf": int(row["is_leaf"]),
        }
        for row in rows
    }


def _build_impact(current: dict[str, dict[str, Any]], target_rows: list[DeptRow]) -> list[dict[str, Any]]:
    target = {row.dept_code: row for row in target_rows}
    impact: list[dict[str, Any]] = []
    for code in sorted(set(current) | set(target), key=lambda value: (len(value), value)):
        current_row = current.get(code)
        target_row = target.get(code)
        if current_row and target_row:
            diffs = [
                field
                for field in ("dept_name", "entity_name", "parent_code", "level", "is_leaf")
                if current_row[field] != getattr(target_row, field)
            ]
            impact.append(
                {
                    "action": "UPDATE" if diffs else "UNCHANGED",
                    "dept_code": code,
                    "current_name": current_row["dept_name"],
                    "target_name": target_row.dept_name,
                    "diffs": ",".join(diffs),
                }
            )
        elif target_row:
            impact.append(
                {
                    "action": "INSERT",
                    "dept_code": code,
                    "current_name": "",
                    "target_name": target_row.dept_name,
                    "diffs": "新增主键",
                }
            )
        elif current_row:
            impact.append(
                {
                    "action": "DELETE",
                    "dept_code": code,
                    "current_name": current_row["dept_name"],
                    "target_name": "",
                    "diffs": "当前主键不在部门架构模板中",
                }
            )
    return impact


def _write_report(report_path: Path, summary: dict[str, Any], target_rows: list[DeptRow], impact: list[dict[str, Any]]) -> None:
    report_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "执行摘要"
    ws.append(["项目", "值"])
    for key, value in summary.items():
        ws.append([key, json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value])

    ws = wb.create_sheet("执行后dept_account")
    ws.append(["dept_code", "dept_name", "entity_name", "parent_code", "level", "is_leaf", "source_row", "source_type"])
    for row in target_rows:
        ws.append(
            [
                row.dept_code,
                row.dept_name,
                row.entity_name,
                row.parent_code or "",
                row.level,
                row.is_leaf,
                row.source_row,
                row.source_type,
            ]
        )

    ws = wb.create_sheet("执行影响")
    ws.append(["action", "dept_code", "current_name", "target_name", "diffs"])
    for item in impact:
        ws.append([item["action"], item["dept_code"], item["current_name"], item["target_name"], item["diffs"]])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in sheet.columns:
            letter = get_column_letter(column_cells[0].column)
            width = max(len(str(cell.value or "")) for cell in column_cells[:80]) + 2
            sheet.column_dimensions[letter].width = min(max(width, 10), 48)

    wb.save(report_path)


def rebuild(
    *,
    db_path: Path,
    template_path: Path,
    backup_dir: Path,
    report_path: Path,
    dry_run: bool,
    no_backup: bool,
) -> dict[str, Any]:
    target_rows = _load_template(template_path)
    backup_path: Path | None = None
    if not dry_run and not no_backup:
        backup_path = _backup_db(db_path, backup_dir)

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    try:
        current = _load_current(conn)
        impact = _build_impact(current, target_rows)
        counts: dict[str, int] = {}
        for item in impact:
            counts[item["action"]] = counts.get(item["action"], 0) + 1

        if not dry_run:
            conn.execute("PRAGMA foreign_keys = OFF")
            conn.execute("BEGIN")
            conn.execute("DELETE FROM dept_account")
            conn.executemany(
                """
                INSERT INTO dept_account(dept_code, dept_name, entity_name, parent_code, level, is_leaf)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                [
                    (
                        row.dept_code,
                        row.dept_name,
                        row.entity_name,
                        row.parent_code,
                        row.level,
                        row.is_leaf,
                    )
                    for row in target_rows
                ],
            )
            conn.commit()
            conn.execute("PRAGMA foreign_keys = ON")
            fk_errors = conn.execute("PRAGMA foreign_key_check").fetchall()
            if fk_errors:
                raise RuntimeError(f"foreign key check failed: {fk_errors}")

        summary: dict[str, Any] = {
            "db_path": str(db_path),
            "template_path": str(template_path),
            "backup_path": str(backup_path) if backup_path else "",
            "report_path": str(report_path),
            "target_count": len(target_rows),
            "impact_counts": counts,
            "dry_run": dry_run,
        }
        _write_report(report_path, summary, target_rows, impact)
        return summary
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Rebuild dept_account from department hierarchy template.")
    parser.add_argument("--db", default="var/data/common.db", help="Path to common.db")
    parser.add_argument("--template", default="resources/business_inputs/部门架构维护模版.xlsx", help="Department hierarchy template")
    parser.add_argument("--backup-dir", default="var/data/backups", help="Directory for database backups")
    parser.add_argument(
        "--report",
        default=f"var/output/acceptance/部门科目维护更新执行结果_{datetime.now().strftime('%Y%m%d')}.xlsx",
        help="Execution report path",
    )
    parser.add_argument("--dry-run", action="store_true", help="Only preview, do not write DB")
    parser.add_argument("--no-backup", action="store_true", help="Skip backup")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    summary = rebuild(
        db_path=Path(args.db),
        template_path=Path(args.template),
        backup_dir=Path(args.backup_dir),
        report_path=Path(args.report),
        dry_run=bool(args.dry_run),
        no_backup=bool(args.no_backup),
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
