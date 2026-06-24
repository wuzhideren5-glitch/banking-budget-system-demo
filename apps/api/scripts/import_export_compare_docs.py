#!/usr/bin/env python3
"""Import docs test data, export prediction output, compare with standard workbooks."""
from __future__ import annotations

import asyncio
import json
import sqlite3
import sys
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path

ROOT = Path(__file__).resolve().parents[3]
API_ROOT = Path(__file__).resolve().parents[1]
DOCS = ROOT / "docs"
OUT_DIR = ROOT / ".scratch" / "org-product-import-export-compare"
sys.path.insert(0, str(API_ROOT))

import app.core.pymysql_compat  # noqa: F401

from openpyxl import Workbook, load_workbook

from app.core.db_paths import common_db_path
from app.routers.org_product_data_entry import _persist_data_entry_snapshot
from app.routers.org_product_helpers import (
    OrgProductOutputRunRequest,
    _append_org_product_output_export_sheet,
    _ensure_data_entry_snapshot_table,
    _ensure_data_entry_snapshot_table_v2,
    _parse_data_entry_workbook,
    _resolve_data_entry_entity_name,
    _sanitize_data_entry_payload_mapping_refs,
    _unique_sheet_title,
    load_org_product_metric_table_rows_from_runtime_tree,
)
from app.routers import org_product_output as output_mod

TEST_ENTRY = DOCS / "机构产品数据录入_2026_测试数据.xlsx"
STD_ENTRY = DOCS / "机构产品数据录入_2026_完整版.xlsx"
STD_OUTPUT = DOCS / "预测输出_全量_AA_2026_v1 (1).xlsx"

YEAR = 2026
MONTH = 3
VERSION_ID = 1
VERSION_NAME = "202603v1"
ENTITY_ROOT = "AA"


def _parse_amount(raw) -> float | None:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    text = str(raw).strip()
    if not text or text in {"-", "—"}:
        return None
    try:
        return float(text.replace(",", ""))
    except Exception:
        return None


def _load_output_workbook(path: Path) -> dict[str, dict[str, dict[str, float | None]]]:
    """sheet_name -> metric_code -> {annual, m1..m12}"""
    wb = load_workbook(path, data_only=True, read_only=True)
    out: dict[str, dict[str, dict[str, float | None]]] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(min_row=1, values_only=True)
        headers = list(next(rows_iter, []) or [])
        if len(headers) < 5:
            continue
        sheet_rows: dict[str, dict[str, float | None]] = {}
        for row in rows_iter:
            if not row or len(row) < 5:
                continue
            code = str(row[2] or "").strip()
            if not code:
                continue
            vals: dict[str, float | None] = {"annual": _parse_amount(row[4])}
            for i in range(12):
                vals[f"m{i + 1}"] = _parse_amount(row[5 + i] if len(row) > 5 + i else None)
            sheet_rows[code] = vals
        out[sheet_name] = sheet_rows
    wb.close()
    return out


def _load_entry_workbook(path: Path) -> dict[str, dict[str, dict[str, float | None]]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    out: dict[str, dict[str, dict[str, float | None]]] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(min_row=1, values_only=True)
        headers = list(next(rows_iter, []) or [])
        if len(headers) < 8:
            continue
        sheet_rows: dict[str, dict[str, float | None]] = {}
        for row in rows_iter:
            if not row or len(row) < 8:
                continue
            code = str(row[2] or "").strip()
            if not code:
                continue
            vals: dict[str, float | None] = {}
            for i, h in enumerate(headers[4:], start=4):
                key = str(h or "")
                vals[key] = _parse_amount(row[i] if len(row) > i else None)
            sheet_rows[code] = vals
        out[sheet_name] = sheet_rows
    wb.close()
    return out


def _match_sheet(actual_sheets: list[str], std_sheet: str) -> str | None:
    std_norm = std_sheet.replace("_", "").replace(" ", "").lower()
    for s in actual_sheets:
        if s == std_sheet:
            return s
    for s in actual_sheets:
        if std_norm in s.replace("_", "").replace(" ", "").lower():
            return s
    return None


def _compare_numeric(
    actual: dict[str, float | None],
    expected: dict[str, float | None],
    *,
    keys: list[str],
    tol: float = 1.0,
) -> list[dict]:
    diffs: list[dict] = []
    for key in keys:
        a = actual.get(key)
        e = expected.get(key)
        if a is None and e is None:
            continue
        if a is None or e is None:
            diffs.append({"field": key, "actual": a, "expected": e, "delta": None})
            continue
        delta = abs(float(a) - float(e))
        if delta > tol:
            diffs.append({"field": key, "actual": a, "expected": e, "delta": delta})
    return diffs


def import_test_data() -> dict:
    if not TEST_ENTRY.exists():
        raise FileNotFoundError(TEST_ENTRY)
    raw = TEST_ENTRY.read_bytes()
    path = common_db_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    saved: list[dict] = []
    unmatched: list[dict] = []
    with sqlite3.connect(path) as conn:
        _ensure_data_entry_snapshot_table(conn)
        _ensure_data_entry_snapshot_table_v2(conn)
        candidates = [
            (
                str(row.get("entity_code") or "").strip(),
                str(row.get("entity_name") or "").strip(),
                str(row.get("table_name") or "").strip(),
            )
            for row in load_org_product_metric_table_rows_from_runtime_tree(conn)
            if str(row.get("entity_code") or "").strip() and str(row.get("table_name") or "").strip()
        ]
        sheets = _parse_data_entry_workbook(raw, TEST_ENTRY.name, YEAR, MONTH, candidates)
        for sheet in sheets:
            if not sheet.get("matched"):
                unmatched.append(
                    {
                        "sheet_name": sheet.get("sheet_name"),
                        "entity_code": sheet.get("entity_code"),
                        "table_name": sheet.get("table_name"),
                    }
                )
                continue
            entity_code = str(sheet.get("entity_code") or "").strip()
            table_name = str(sheet.get("table_name") or "").strip()
            payload_obj = {
                "entity_code": entity_code,
                "entity_name": _resolve_data_entry_entity_name(conn, entity_code),
                "year": YEAR,
                "month_index": MONTH,
                "version_id": VERSION_ID,
                "version_name": VERSION_NAME,
                "table_name": table_name,
                "entry_status": "draft",
                "metrics": list(sheet.get("metrics") or []),
            }
            payload_obj = _sanitize_data_entry_payload_mapping_refs(conn, payload_obj)
            updated_at = _persist_data_entry_snapshot(conn, payload_obj)
            metrics_with_data = 0
            for m in payload_obj.get("metrics") or []:
                months = (m.get("values") or {}).get("months") or {}
                if any(str(v).strip() for v in months.values()):
                    metrics_with_data += 1
            saved.append(
                {
                    "entity_code": entity_code,
                    "table_name": table_name,
                    "sheet_name": sheet.get("sheet_name"),
                    "metrics_with_data": metrics_with_data,
                    "updated_at": updated_at,
                }
            )
        conn.commit()
    return {"saved": saved, "unmatched": unmatched, "saved_count": len(saved)}


async def export_prediction_output(export_path: Path) -> dict:
    payload = OrgProductOutputRunRequest(
        entity_code=ENTITY_ROOT,
        year=YEAR,
        version_id=VERSION_ID,
        table_name="",
        include_children=True,
    )
    result = await output_mod.run_org_product_output(payload)
    entities = list(result.get("entities") or [])
    wb = Workbook()
    wb.remove(wb.active)
    used_titles: set[str] = set()
    for ent in entities:
        ec = str(ent.get("entity_code") or "")
        en = str(ent.get("entity_name") or "")
        tn = str(ent.get("table_name") or "")
        title = _unique_sheet_title(used_titles, f"{ec}{en}_{tn}")
        ws = wb.create_sheet(title=title)
        _append_org_product_output_export_sheet(ws, list(ent.get("rows") or []), ec)
    export_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(export_path)
    return {"entity_count": len(entities), "path": str(export_path)}


def compare_outputs(actual_path: Path, std_path: Path) -> dict:
    actual = _load_output_workbook(actual_path)
    expected = _load_output_workbook(std_path)
    keys = ["annual"] + [f"m{i}" for i in range(1, 13)]
    sheet_reports: list[dict] = []
    total_diff_cells = 0
    p0_issues: list[dict] = []

    for std_sheet, std_rows in expected.items():
        act_sheet = _match_sheet(list(actual.keys()), std_sheet)
        if not act_sheet:
            sheet_reports.append({"sheet": std_sheet, "status": "missing_sheet", "diff_count": 0})
            continue
        act_rows = actual[act_sheet]
        code_diffs: list[dict] = []
        for code, exp_vals in std_rows.items():
            act_vals = act_rows.get(code)
            if act_vals is None:
                if any(v is not None for v in exp_vals.values()):
                    code_diffs.append({"code": code, "issue": "missing_code", "diffs": []})
                continue
            diffs = _compare_numeric(act_vals, exp_vals, keys=keys, tol=1.0)
            if not diffs:
                continue
            code_diffs.append({"code": code, "issue": "value_mismatch", "diffs": diffs})
            total_diff_cells += len(diffs)
            for d in diffs:
                a = d.get("actual")
                if isinstance(a, (int, float)) and abs(float(a)) >= 1e15:
                    p0_issues.append(
                        {
                            "sheet": std_sheet,
                            "code": code,
                            "field": d.get("field"),
                            "actual": a,
                            "expected": d.get("expected"),
                        }
                    )
        sheet_reports.append(
            {
                "sheet": std_sheet,
                "matched_sheet": act_sheet,
                "status": "ok" if not code_diffs else "diff",
                "diff_metric_count": len(code_diffs),
                "top_diffs": code_diffs[:5],
            }
        )

    return {
        "actual_sheets": len(actual),
        "standard_sheets": len(expected),
        "total_diff_cells": total_diff_cells,
        "p0_extreme_values": p0_issues[:20],
        "sheets": sheet_reports,
    }


def compare_entry_import(actual_db_metrics: int, std_path: Path) -> dict:
    """Compare imported test data coverage vs 完整版 entry standard."""
    std = _load_entry_workbook(std_path)
    month_cols = [h for h in next(iter(next(iter(std.values())).values())).keys()] if std else []
    coverage: list[dict] = []
    path = common_db_path()
    with sqlite3.connect(path) as conn:
        for sheet_name, std_rows in std.items():
            entity_code = table_name = None
            for row in load_org_product_metric_table_rows_from_runtime_tree(conn):
                ec = str(row.get("entity_code") or "").strip()
                en = str(row.get("entity_name") or "").strip()
                tn = str(row.get("table_name") or "").strip()
                guess = f"{ec}{en}{tn}".replace(" ", "")
                if guess and guess in sheet_name.replace(" ", ""):
                    entity_code, table_name = ec, tn
                    break
            db_rows: dict[str, dict] = {}
            if entity_code and table_name:
                cur = conn.execute(
                    """
                    SELECT payload_json FROM org_product_data_entry_snapshot_v2
                    WHERE entity_code=? AND year=? AND version_id=? AND table_name=?
                    """,
                    (entity_code, YEAR, VERSION_ID, table_name),
                )
                row = cur.fetchone()
                if row and row[0]:
                    obj = json.loads(row[0])
                    for m in obj.get("metrics") or []:
                        code = str(m.get("metric_code") or "").strip()
                        if code:
                            db_rows[code] = m.get("values") or {}

            std_with_values = sum(
                1 for code, cols in std_rows.items() if any(v is not None for v in cols.values())
            )
            db_with_values = 0
            mismatch_vs_std = 0
            for code, std_cols in std_rows.items():
                db_vals = db_rows.get(code) or {}
                db_months = db_vals.get("months") if isinstance(db_vals.get("months"), dict) else {}
                db_any = any(str(v).strip() for v in db_months.values()) or str(db_vals.get("year_forecast") or "").strip()
                if db_any:
                    db_with_values += 1
            coverage.append(
                {
                    "sheet": sheet_name,
                    "entity_code": entity_code,
                    "table_name": table_name,
                    "std_metrics_with_values": std_with_values,
                    "imported_metrics_with_values": db_with_values,
                    "in_db": bool(db_rows),
                }
            )
    return {"sheets": coverage}


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    export_path = OUT_DIR / f"预测输出_全量_AA_{YEAR}_v{VERSION_ID}_after_import_{ts}.xlsx"
    report_path = OUT_DIR / f"compare_report_{ts}.json"

    print("=== 1) 导入测试数据 ===")
    import_result = import_test_data()
    print(f"saved {import_result['saved_count']} sheets, unmatched {len(import_result['unmatched'])}")
    for item in import_result["saved"]:
        print(
            f"  {item['entity_code']}/{item['table_name']}: "
            f"{item['metrics_with_data']} metrics with values"
        )

    print("\n=== 2) 导出预测输出 ===")
    export_result = asyncio.run(export_prediction_output(export_path))
    print(f"entities {export_result['entity_count']} -> {export_path}")

    print("\n=== 3) 与标准预测输出对比 ===")
    if not STD_OUTPUT.exists():
        print(f"标准文件不存在: {STD_OUTPUT}")
        output_compare = {"error": "standard output missing"}
    else:
        output_compare = compare_outputs(export_path, STD_OUTPUT)
        diff_sheets = [s for s in output_compare["sheets"] if s.get("status") != "ok"]
        print(f"diff cells: {output_compare['total_diff_cells']}, sheets with diff: {len(diff_sheets)}")
        for s in diff_sheets[:8]:
            print(
                f"  {s.get('sheet')}: {s.get('status')} "
                f"diff_metrics={s.get('diff_metric_count', 0)}"
            )
            for top in s.get("top_diffs") or []:
                code = top.get("code")
                for d in (top.get("diffs") or [])[:2]:
                    print(f"    {code} {d.get('field')}: actual={d.get('actual')} expected={d.get('expected')}")

    print("\n=== 4) 录入覆盖 vs 完整版 ===")
    entry_compare = compare_entry_import(import_result["saved_count"], STD_ENTRY)
    for s in entry_compare["sheets"]:
        if not s.get("in_db"):
            print(f"  MISSING DB: {s['sheet']}")
        elif s["imported_metrics_with_values"] < s["std_metrics_with_values"]:
            print(
                f"  {s['sheet']}: imported {s['imported_metrics_with_values']} "
                f"vs std {s['std_metrics_with_values']}"
            )

    report = {
        "timestamp": ts,
        "import": import_result,
        "export": export_result,
        "output_compare": output_compare,
        "entry_coverage_compare": entry_compare,
    }
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\nReport: {report_path}")


if __name__ == "__main__":
    main()
