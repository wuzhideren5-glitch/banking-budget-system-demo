#!/usr/bin/env python3
"""Seed default expense forecast rules for an existing banking-budget DB.

The script is intentionally data-only: it does not create tables or change
primary keys. It expects the expense framework, budget subject catalog, and
actual expense detail tables to have already been initialized.
"""

from __future__ import annotations

import argparse
import json
import shutil
import sqlite3
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any


DEFAULT_PARAMS: tuple[tuple[str, str, str, str], ...] = (
    ("scheme2", "allocation_mode", "progressive", "string"),
    ("scheme2", "progressive_curve_type", "arithmetic", "string"),
    ("scheme2", "auto_direction_mode", "auto_last_vs_avg", "string"),
    ("scheme2", "last_value_source_mode", "actual_first_then_forecast", "string"),
    ("scheme2", "rounding_mode", "last_month_adjust", "string"),
    ("scheme2", "allow_negative", "false", "string"),
)


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _default_version() -> str:
    return datetime.now().strftime("%y%m%d") + "v1"


def _connect(db_path: Path) -> sqlite3.Connection:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def _backup_db(db_path: Path, backup_dir: Path) -> Path:
    backup_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = backup_dir / f"{db_path.stem}_before_expense_forecast_defaults_{timestamp}{db_path.suffix}"
    shutil.copy2(db_path, backup_path)
    return backup_path


def _fetch_framework_subjects(conn: sqlite3.Connection) -> dict[str, sqlite3.Row]:
    rows = conn.execute(
        """
        SELECT budget_subject, manage_department, formula_text
        FROM expense_framework_subject
        """
    ).fetchall()
    return {_text(row["budget_subject"]): row for row in rows if _text(row["budget_subject"])}


def _fetch_leaf_subjects(conn: sqlite3.Connection) -> dict[str, sqlite3.Row]:
    rows = conn.execute(
        """
        WITH child AS (
          SELECT parent_id, COUNT(*) AS child_count
          FROM budget_subject_catalog
          GROUP BY parent_id
        )
        SELECT b.id, b.subject_name, b.manage_department, b.formula_text
        FROM budget_subject_catalog b
        LEFT JOIN child c ON c.parent_id = b.id
        WHERE COALESCE(c.child_count, 0) = 0
        """
    ).fetchall()
    return {_text(row["subject_name"]): row for row in rows if _text(row["subject_name"])}


def _fetch_owner_names(conn: sqlite3.Connection) -> set[str]:
    rows = conn.execute(
        """
        SELECT DISTINCT owner_name
        FROM expense_framework_budget_department
        WHERE COALESCE(owner_name, '') <> ''
        """
    ).fetchall()
    return {_text(row["owner_name"]) for row in rows}


def _fetch_actual_pairs(conn: sqlite3.Connection) -> set[tuple[str, str]]:
    rows = conn.execute(
        """
        SELECT DISTINCT owner_name_mapped AS owner_name, budget_subject_mapped AS subject_name
        FROM expense_actual_detail_raw
        WHERE COALESCE(owner_name_mapped, '') <> ''
          AND COALESCE(budget_subject_mapped, '') <> ''
        """
    ).fetchall()
    return {(_text(row["owner_name"]), _text(row["subject_name"])) for row in rows}


def _actual_cutoff_month(conn: sqlite3.Connection, year: int) -> int:
    row = conn.execute(
        """
        SELECT MAX(CAST(substr(period_ym, 6, 2) AS INTEGER)) AS cutoff_month
        FROM expense_actual_detail_raw
        WHERE substr(period_ym, 1, 4) = ?
        """,
        (str(year),),
    ).fetchone()
    return int(row["cutoff_month"] or 0)


def _sync_budget_subject_catalog(conn: sqlite3.Connection) -> list[dict[str, Any]]:
    framework = _fetch_framework_subjects(conn)
    catalog_rows = conn.execute(
        """
        SELECT id, subject_name, manage_department, formula_text
        FROM budget_subject_catalog
        """
    ).fetchall()
    changes: list[dict[str, Any]] = []
    for row in catalog_rows:
        subject_name = _text(row["subject_name"])
        framework_row = framework.get(subject_name)
        if not framework_row:
            continue
        old_manage_department = _text(row["manage_department"])
        old_formula_text = _text(row["formula_text"])
        new_manage_department = _text(framework_row["manage_department"]) or old_manage_department
        new_formula_text = _text(framework_row["formula_text"]) or old_formula_text
        if new_manage_department == old_manage_department and new_formula_text == old_formula_text:
            continue
        conn.execute(
            """
            UPDATE budget_subject_catalog
            SET manage_department = ?, formula_text = ?
            WHERE id = ?
            """,
            (new_manage_department or None, new_formula_text or None, int(row["id"])),
        )
        changes.append(
            {
                "subject_id": int(row["id"]),
                "subject_name": subject_name,
                "old_manage_department": old_manage_department,
                "new_manage_department": new_manage_department,
                "old_formula_text": old_formula_text,
                "new_formula_text": new_formula_text,
            }
        )
    return changes


def _build_rule_plan(conn: sqlite3.Connection) -> tuple[list[dict[str, Any]], list[dict[str, str]]]:
    framework = _fetch_framework_subjects(conn)
    leaf_subjects = _fetch_leaf_subjects(conn)
    owners = _fetch_owner_names(conn)
    actual_pairs = _fetch_actual_pairs(conn)
    rule_by_key: dict[tuple[str, int], dict[str, Any]] = {}
    skipped: list[dict[str, str]] = []

    for subject_name, subject in leaf_subjects.items():
        framework_row = framework.get(subject_name)
        manage_department = _text(framework_row["manage_department"] if framework_row else "")
        if not manage_department:
            continue
        if manage_department not in owners:
            skipped.append(
                {
                    "owner_name": manage_department,
                    "subject_name": subject_name,
                    "reason": "manage_department_not_in_framework_owner_tree",
                }
            )
            continue
        rule_by_key[(manage_department, int(subject["id"]))] = {
            "owner_name": manage_department,
            "subject_id": int(subject["id"]),
            "subject_name": subject_name,
            "reason": "manage_department",
        }

    for owner_name, subject_name in actual_pairs:
        subject = leaf_subjects.get(subject_name)
        if not subject:
            skipped.append({"owner_name": owner_name, "subject_name": subject_name, "reason": "subject_not_leaf"})
            continue
        if owner_name not in owners:
            skipped.append({"owner_name": owner_name, "subject_name": subject_name, "reason": "owner_not_in_framework"})
            continue
        framework_row = framework.get(subject_name)
        manage_department = _text(framework_row["manage_department"] if framework_row else "")
        if manage_department:
            continue
        rule_by_key.setdefault(
            (owner_name, int(subject["id"])),
            {
                "owner_name": owner_name,
                "subject_id": int(subject["id"]),
                "subject_name": subject_name,
                "reason": "actual_pair_for_unmanaged_subject",
            },
        )

    plan = sorted(rule_by_key.values(), key=lambda item: (item["owner_name"], item["subject_id"]))
    return plan, skipped


def _existing_rule_keys(conn: sqlite3.Connection, year: int, version: str) -> set[tuple[str, int]]:
    rows = conn.execute(
        """
        SELECT owner_name, subject_id
        FROM expense_forecast_rule
        WHERE forecast_year = ? AND forecast_version = ?
        """,
        (year, version),
    ).fetchall()
    return {(_text(row["owner_name"]), int(row["subject_id"])) for row in rows}


def _upsert_rule(conn: sqlite3.Connection, year: int, version: str, item: dict[str, Any], now: str) -> int:
    remark = "系统默认初始化：余额分摊；来源=" + item["reason"]
    conn.execute(
        """
        INSERT INTO expense_forecast_rule(
          forecast_year, forecast_version, owner_name, subject_id, scheme_code,
          enabled, allow_manual_override, auto_refresh_enabled, manual_recalc_enabled,
          metric_source_priority, effective_from_month, effective_to_month,
          priority, remark, created_at, updated_at
        ) VALUES (?, ?, ?, ?, 'RESIDUAL_ALLOC', 1, 1, 1, 1, 'metric_first', 1, 12, 100, ?, ?, ?)
        ON CONFLICT(forecast_year, forecast_version, owner_name, subject_id)
        DO UPDATE SET
          scheme_code = excluded.scheme_code,
          enabled = excluded.enabled,
          allow_manual_override = excluded.allow_manual_override,
          auto_refresh_enabled = excluded.auto_refresh_enabled,
          manual_recalc_enabled = excluded.manual_recalc_enabled,
          metric_source_priority = excluded.metric_source_priority,
          effective_from_month = excluded.effective_from_month,
          effective_to_month = excluded.effective_to_month,
          priority = excluded.priority,
          remark = excluded.remark,
          updated_at = excluded.updated_at
        """,
        (year, version, item["owner_name"], int(item["subject_id"]), remark, now, now),
    )
    row = conn.execute(
        """
        SELECT id
        FROM expense_forecast_rule
        WHERE forecast_year = ? AND forecast_version = ? AND owner_name = ? AND subject_id = ?
        """,
        (year, version, item["owner_name"], int(item["subject_id"])),
    ).fetchone()
    if row is None:
        raise RuntimeError(f"failed to upsert rule for {item['owner_name']} / {item['subject_name']}")
    return int(row["id"])


def _upsert_params(conn: sqlite3.Connection, rule_id: int) -> None:
    for param_group, param_key, param_value, value_type in DEFAULT_PARAMS:
        conn.execute(
            """
            INSERT INTO expense_forecast_rule_param(rule_id, param_group, param_key, param_value, value_type)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(rule_id, param_group, param_key)
            DO UPDATE SET param_value = excluded.param_value, value_type = excluded.value_type
            """,
            (rule_id, param_group, param_key, param_value, value_type),
        )


def _upsert_annual_inputs(
    conn: sqlite3.Connection,
    *,
    year: int,
    version: str,
    owner_name: str,
    subject_id: int,
    now: str,
) -> None:
    for field_name in ("business_submission", "capital_advice"):
        conn.execute(
            """
            INSERT INTO expense_forecast_annual_entry(
              forecast_year, forecast_version, scope_type, scope_value, subject_id,
              field_name, field_value, create_time, update_time
            ) VALUES (?, ?, 'owner', ?, ?, ?, 0, ?, ?)
            ON CONFLICT(forecast_year, forecast_version, scope_type, scope_value, subject_id, field_name)
            DO UPDATE SET update_time = excluded.update_time
            """,
            (year, version, owner_name, subject_id, field_name, now, now),
        )


def _upsert_monthly_defaults(
    conn: sqlite3.Connection,
    *,
    year: int,
    version: str,
    owner_name: str,
    subject_id: int,
    subject_name: str,
    rule_id: int,
    actual_cutoff_month: int,
    now: str,
) -> int:
    months = range(max(1, actual_cutoff_month + 1), 13)
    count = 0
    for month in months:
        basis = json.dumps(
            {
                "scheme_code": "RESIDUAL_ALLOC",
                "init_reason": "default_rule_seed",
                "actual_cutoff_month": actual_cutoff_month,
                "capital_advice": 0,
                "subject_name": subject_name,
            },
            ensure_ascii=False,
            sort_keys=True,
        )
        conn.execute(
            """
            INSERT INTO expense_forecast_calc_result(
              forecast_year, forecast_version, owner_name, subject_id, month, rule_id,
              calc_value, calc_basis_json, calc_status, calc_time
            ) VALUES (?, ?, ?, ?, ?, ?, 0, ?, 'ok', ?)
            ON CONFLICT(forecast_year, forecast_version, owner_name, subject_id, month)
            DO UPDATE SET
              rule_id = excluded.rule_id,
              calc_basis_json = excluded.calc_basis_json,
              calc_status = excluded.calc_status,
              calc_time = excluded.calc_time
            """,
            (year, version, owner_name, subject_id, month, rule_id, basis, now),
        )
        conn.execute(
            """
            INSERT INTO expense_forecast_entry(
              forecast_year, forecast_version, scope_type, scope_value, subject_id,
              month, forecast_value, create_time, update_time
            ) VALUES (?, ?, 'owner', ?, ?, ?, 0, ?, ?)
            ON CONFLICT(forecast_year, forecast_version, scope_type, scope_value, subject_id, month)
            DO UPDATE SET update_time = excluded.update_time
            """,
            (year, version, owner_name, subject_id, month, now, now),
        )
        count += 1
    return count


def _write_report(
    *,
    report_path: Path,
    summary: dict[str, Any],
    rule_plan: list[dict[str, Any]],
    subject_changes: list[dict[str, Any]],
    skipped: list[dict[str, str]],
) -> str | None:
    try:
        from openpyxl import Workbook
    except ImportError:
        return None

    report_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    ws = wb.active
    ws.title = "摘要"
    ws.append(["项目", "值"])
    for key, value in summary.items():
        ws.append([key, json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value])

    ws = wb.create_sheet("默认规则")
    ws.append(["owner_name", "subject_id", "subject_name", "reason"])
    for item in rule_plan:
        ws.append([item["owner_name"], item["subject_id"], item["subject_name"], item["reason"]])

    ws = wb.create_sheet("科目同步")
    ws.append(
        [
            "subject_id",
            "subject_name",
            "old_manage_department",
            "new_manage_department",
            "old_formula_text",
            "new_formula_text",
        ]
    )
    for item in subject_changes:
        ws.append(
            [
                item["subject_id"],
                item["subject_name"],
                item["old_manage_department"],
                item["new_manage_department"],
                item["old_formula_text"],
                item["new_formula_text"],
            ]
        )

    ws = wb.create_sheet("跳过项")
    ws.append(["owner_name", "subject_name", "reason"])
    for item in skipped:
        ws.append([item["owner_name"], item["subject_name"], item["reason"]])

    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 48)

    wb.save(report_path)
    return str(report_path)


def seed_defaults(
    *,
    db_path: Path,
    backup_dir: Path,
    report_path: Path,
    year: int,
    version: str,
    dry_run: bool,
    no_backup: bool,
) -> dict[str, Any]:
    if not db_path.exists():
        raise FileNotFoundError(db_path)

    backup_path: Path | None = None
    if not dry_run and not no_backup:
        backup_path = _backup_db(db_path, backup_dir)

    now = datetime.now().isoformat(timespec="seconds")
    conn = _connect(db_path)
    try:
        subject_changes = _sync_budget_subject_catalog(conn)
        rule_plan, skipped = _build_rule_plan(conn)
        actual_cutoff_month = _actual_cutoff_month(conn, year)
        existing_keys = _existing_rule_keys(conn, year, version)

        reason_counts = Counter(item["reason"] for item in rule_plan)
        owner_counts = Counter(item["owner_name"] for item in rule_plan)
        monthly_rows = 0

        if dry_run:
            conn.rollback()
        else:
            with conn:
                for item in rule_plan:
                    rule_id = _upsert_rule(conn, year, version, item, now)
                    _upsert_params(conn, rule_id)
                    _upsert_annual_inputs(
                        conn,
                        year=year,
                        version=version,
                        owner_name=item["owner_name"],
                        subject_id=int(item["subject_id"]),
                        now=now,
                    )
                    monthly_rows += _upsert_monthly_defaults(
                        conn,
                        year=year,
                        version=version,
                        owner_name=item["owner_name"],
                        subject_id=int(item["subject_id"]),
                        subject_name=item["subject_name"],
                        rule_id=rule_id,
                        actual_cutoff_month=actual_cutoff_month,
                        now=now,
                    )

        summary: dict[str, Any] = {
            "db_path": str(db_path),
            "backup_path": str(backup_path) if backup_path else "",
            "forecast_year": year,
            "forecast_version": version,
            "actual_cutoff_month": actual_cutoff_month,
            "subject_catalog_updates": len(subject_changes),
            "rule_plan_count": len(rule_plan),
            "rules_created": len([item for item in rule_plan if (item["owner_name"], int(item["subject_id"])) not in existing_keys]),
            "rules_updated": len([item for item in rule_plan if (item["owner_name"], int(item["subject_id"])) in existing_keys]),
            "rule_reason_counts": dict(reason_counts),
            "rule_owner_counts": dict(owner_counts),
            "annual_entry_rows_expected": len(rule_plan) * 2,
            "monthly_rows_expected_per_table": len(rule_plan) * max(0, 12 - actual_cutoff_month),
            "monthly_rows_written_per_table": monthly_rows,
            "skipped_count": len(skipped),
            "dry_run": dry_run,
        }

        report = _write_report(
            report_path=report_path,
            summary=summary,
            rule_plan=rule_plan,
            subject_changes=subject_changes,
            skipped=skipped,
        )
        if report:
            summary["report_path"] = report
        return summary
    finally:
        conn.close()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Seed default expense forecast rules.")
    parser.add_argument("--db", default="var/data/common.db", help="Path to common.db")
    parser.add_argument("--year", type=int, default=2026, help="Forecast year")
    parser.add_argument("--version", default=_default_version(), help="Forecast version, defaults to yymmddv1")
    parser.add_argument("--backup-dir", default="var/data/backups", help="Directory for DB backup")
    parser.add_argument(
        "--report",
        default=f"var/output/acceptance/费用预测默认规则导入报告_{datetime.now().strftime('%Y%m%d')}.xlsx",
        help="XLSX report path",
    )
    parser.add_argument("--dry-run", action="store_true", help="Plan only, do not write DB")
    parser.add_argument("--no-backup", action="store_true", help="Skip DB backup")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    summary = seed_defaults(
        db_path=Path(args.db),
        backup_dir=Path(args.backup_dir),
        report_path=Path(args.report),
        year=args.year,
        version=_text(args.version) or _default_version(),
        dry_run=bool(args.dry_run),
        no_backup=bool(args.no_backup),
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
