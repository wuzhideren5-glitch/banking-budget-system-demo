#!/usr/bin/env python3
"""Maintain v03 org-product metric workbook structure.

- Inserts implicit GROUP rows (AA.24 / AA.25 / AA.14.02 / …) when absent
- Removes mirror-duplicate rows whose canonical definition lives on another sheet

Usage:
    cd apps/api && . .venv/bin/activate
    python scripts/maintain_v03_metric_workbook.py --dry-run
    python scripts/maintain_v03_metric_workbook.py --apply
"""
from __future__ import annotations

import argparse
import shutil
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

API_ROOT = Path(__file__).resolve().parents[1]
if str(API_ROOT) not in sys.path:
    sys.path.insert(0, str(API_ROOT))

from app.core.config import settings  # noqa: E402
from app.routers.org_product_helpers import (  # noqa: E402
    _find_header_row,
    _normalize_metric_code,
    _normalize_text,
    _prepare_metric_worksheet,
    _resolve_import_sheet_entity_table,
    _sheet_scan_row_limit,
    _ws_cell_value,
)
from app.services.v03_metric_node_catalog import (  # noqa: E402
    IMPLICIT_GROUP_PARENTS,
    V03_MIRROR_DUPLICATE_ROWS,
    is_v03_mirror_duplicate_row,
    implicit_group_spec,
)

DEFAULT_WORKBOOK = settings.business_inputs_dir / "机构及产品指标（公式配置） - v03.xlsx"

IMPLICIT_GROUP_INSERTS: tuple[tuple[str, str, str], ...] = (
    ("AA资产负债表（余额）", "AA.24", "AA.24.01"),
    ("AA资产负债表（余额）", "AA.26", "AA.26.01"),
    ("AA资产负债表（日均）", "AA.25", "AA.25.01"),
    ("AA资产负债表（日均）", "AA.27", "AA.27.01"),
    ("AA利息净收入表", "AA.14.02", "AA.14.02.01"),
    ("AA利息净收入表", "AA.16.02", "AA.16.02.01"),
)


def _codes_in_sheet(ws, entity_code: str) -> set[str]:
    hr, hm, *_ = _find_header_row(ws, entity_code, strict=False, sheet_title=str(getattr(ws, "title", "")))
    code_col = hm.get("科目代码") if hr else None
    if not hr or not code_col:
        return set()
    codes: set[str] = set()
    for row_idx in range(hr + 1, _sheet_scan_row_limit(ws, hr) + 1):
        code = _normalize_metric_code(entity_code, _ws_cell_value(ws, row_idx, code_col)).upper()
        if code:
            codes.add(code)
    return codes


def _find_row_for_code(ws, entity_code: str, target_code: str) -> int | None:
    hr, hm, *_ = _find_header_row(ws, entity_code, strict=False, sheet_title=str(getattr(ws, "title", "")))
    code_col = hm.get("科目代码") if hr else None
    if not hr or not code_col:
        return None
    target = target_code.upper()
    for row_idx in range(hr + 1, _sheet_scan_row_limit(ws, hr) + 1):
        code = _normalize_metric_code(entity_code, _ws_cell_value(ws, row_idx, code_col)).upper()
        if code == target:
            return row_idx
    return None


def _insert_implicit_group_row(ws, entity_code: str, group_code: str, before_code: str) -> bool:
    hr, hm, *_ = _find_header_row(ws, entity_code, strict=False, sheet_title=str(getattr(ws, "title", "")))
    if not hr:
        return False
    code_col = hm.get("科目代码")
    name_col = hm.get("科目名称")
    level_col = hm.get("科目层级")
    if not code_col or not name_col:
        return False
    before_row = _find_row_for_code(ws, entity_code, before_code)
    if before_row is None:
        return False
    spec = implicit_group_spec(group_code)
    if not spec:
        return False
    ws.insert_rows(before_row)
    ws.cell(before_row, code_col, group_code)
    ws.cell(before_row, name_col, spec.node_name)
    if level_col:
        anchor_level = _normalize_text(_ws_cell_value(ws, before_row + 1, level_col))
        ws.cell(before_row, level_col, anchor_level or "二级")
    return True


def _delete_mirror_rows(ws, entity_code: str) -> list[str]:
    hr, hm, *_ = _find_header_row(ws, entity_code, strict=False, sheet_title=str(getattr(ws, "title", "")))
    code_col = hm.get("科目代码") if hr else None
    if not hr or not code_col:
        return []
    removed: list[str] = []
    sheet_name = str(getattr(ws, "title", "") or "")
    row_idx = hr + 1
    while row_idx <= _sheet_scan_row_limit(ws, hr):
        code = _normalize_metric_code(entity_code, _ws_cell_value(ws, row_idx, code_col)).upper()
        if code and is_v03_mirror_duplicate_row(sheet_name, code):
            ws.delete_rows(row_idx, 1)
            removed.append(code)
            continue
        row_idx += 1
    return removed


def maintain_workbook(workbook_path: Path, *, apply: bool) -> dict[str, list[str]]:
    if not workbook_path.exists():
        raise FileNotFoundError(workbook_path)

    wb = load_workbook(workbook_path, data_only=False)
    inserted: list[str] = []
    removed: list[str] = []

    for sheet_name, group_code, before_code in IMPLICIT_GROUP_INSERTS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        _prepare_metric_worksheet(ws)
        resolved = _resolve_import_sheet_entity_table(sheet_name, [], strict=False)
        if not resolved:
            continue
        entity_code = resolved[0]
        codes = _codes_in_sheet(ws, entity_code)
        if group_code in codes:
            continue
        if before_code not in codes:
            continue
        if apply:
            if _insert_implicit_group_row(ws, entity_code, group_code, before_code):
                inserted.append(f"{sheet_name}:{group_code}")
        else:
            inserted.append(f"{sheet_name}:{group_code}")

    for sheet_name, mirror_code in V03_MIRROR_DUPLICATE_ROWS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        _prepare_metric_worksheet(ws)
        resolved = _resolve_import_sheet_entity_table(sheet_name, [], strict=False)
        if not resolved:
            continue
        entity_code = resolved[0]
        codes = _codes_in_sheet(ws, entity_code)
        if mirror_code not in codes:
            continue
        if apply:
            removed.extend(f"{sheet_name}:{code}" for code in _delete_mirror_rows(ws, entity_code))
        else:
            removed.append(f"{sheet_name}:{mirror_code}")

    if apply:
        backup_dir = settings.data_dir / "backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = backup_dir / f"v03_metric_workbook_before_maintain_{stamp}.xlsx"
        shutil.copy2(workbook_path, backup_path)
        wb.save(workbook_path)
        print(f"backup: {backup_path}")

    return {"inserted": inserted, "removed": removed}


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    mode = parser.add_mutually_exclusive_group()
    mode.add_argument("--dry-run", action="store_true")
    mode.add_argument("--apply", action="store_true")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    args = parser.parse_args(argv)
    if not args.apply:
        args.dry_run = True

    # dry-run uses a temp copy in memory only via reload after simulated checks
    result = maintain_workbook(args.workbook, apply=args.apply)
    print(f"implicit groups to insert: {len(result['inserted'])}")
    for item in result["inserted"]:
        print(f"  + {item}")
    print(f"mirror duplicates to remove: {len(result['removed'])}")
    for item in result["removed"]:
        print(f"  - {item}")
    if args.dry_run and (result["inserted"] or result["removed"]):
        print("(dry-run: no file written)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
