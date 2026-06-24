#!/usr/bin/env python3
"""Import full test data, run prediction output, compare with 完整版(含公式)."""
from __future__ import annotations

import asyncio
import json
import re
import sqlite3
import sys
from datetime import datetime, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parents[3]
API_ROOT = Path(__file__).resolve().parents[1]
DOCS = ROOT / "docs"
OUT_DIR = ROOT / ".scratch" / "org-product-import-export-compare"
sys.path.insert(0, str(API_ROOT))

import app.core.pymysql_compat  # noqa: F401

from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from app.core.db_paths import common_db_path
from app.routers.org_product_data_entry import _persist_data_entry_snapshot
from app.routers.org_product_helpers import (
    OrgProductOutputRunRequest,
    _append_org_product_output_export_sheet,
    _ensure_data_entry_snapshot_table,
    _ensure_data_entry_snapshot_table_v2,
    _parse_data_entry_workbook,
    _resolve_data_entry_entity_name,
    _sanitize_data_entry_payload_mapping_refs,
    _unique_sheet_title,
    load_org_product_metric_table_rows_from_runtime_tree,
)
from app.routers import org_product_output as output_mod

TEST_ENTRY = DOCS / "机构产品数据录入_2026_模拟数据(仅叶子行).xlsx"
STD_FORMULA = DOCS / "机构产品数据录入_2026_完整版(含公式-两级汇总).xlsx"
# 回退路径（旧对比）
TEST_ENTRY_FALLBACK = DOCS / "机构产品数据录入_2026_测试数据-全部.xlsx"
STD_FORMULA_FALLBACK = DOCS / "机构产品数据录入_2026_完整版(含公式).xlsx"

YEAR = 2026
MONTH = 3
VERSION_ID = 1
VERSION_NAME = "202603v1"
ENTITY_ROOT = "AA"

CELL_REF_RE = re.compile(r"([A-Z]+)(\d+)", re.IGNORECASE)
MONTH_COLS = list(range(8, 20))  # H..S : 26年1月实际 .. 26年12月预测


def _parse_amount(raw) -> float | None:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    text = str(raw).strip()
    if not text or text in {"-", "—"}:
        return None
    try:
        return float(text.replace(",", ""))
    except Exception:
        return None


def _eval_sum_formula(formula: str, sheet_values: dict[tuple[int, int], float | None]) -> float | None:
    expr = str(formula or "").strip()
    if not expr.startswith("="):
        return _parse_amount(expr)
    body = expr[1:].replace(" ", "")
    if not body:
        return None
    total = 0.0
    found = False
    pos = 0
    sign = 1
    while pos < len(body):
        if body[pos] in "+-":
            sign = -1 if body[pos] == "-" else 1
            pos += 1
            continue
        m = CELL_REF_RE.match(body, pos)
        if not m:
            return None
        col = column_index_from_string(m.group(1).upper())
        row = int(m.group(2))
        val = sheet_values.get((row, col))
        if val is None:
            return None
        total += sign * float(val)
        found = True
        sign = 1
        pos = m.end()
    return total if found else None


def _resolve_sheet_month_grid(ws) -> dict[tuple[int, int], float | None]:
    """Evaluate month columns (H..S) including simple =A1+A2 sum formulas."""
    max_row = ws.max_row or 0
    raw: dict[tuple[int, int], str | float | None] = {}
    for row_idx in range(2, max_row + 1):
        for col_idx in MONTH_COLS:
            raw[(row_idx, col_idx)] = ws.cell(row_idx, col_idx).value

    resolved: dict[tuple[int, int], float | None] = {}
    for col_idx in MONTH_COLS:
        for _ in range(max(max_row, 8)):
            changed = False
            for row_idx in range(2, max_row + 1):
                key = (row_idx, col_idx)
                val = raw.get(key)
                if val is None or val == "":
                    if key not in resolved:
                        resolved[key] = None
                    continue
                if isinstance(val, (int, float)):
                    new_val = float(val)
                elif isinstance(val, str) and val.startswith("="):
                    new_val = _eval_sum_formula(val, resolved)
                    if new_val is None:
                        continue
                else:
                    new_val = _parse_amount(val)
                prev = resolved.get(key, object())
                if prev != new_val:
                    resolved[key] = new_val
                    changed = True
            if not changed:
                break
    return resolved


def _load_standard_formula_workbook(path: Path) -> dict[str, dict[str, dict[str, float | None]]]:
    """sheet_name -> metric_code -> {m1..m12, year_forecast}"""
    wb = load_workbook(path, data_only=False)
    out: dict[str, dict[str, dict[str, float | None]]] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        grid = _resolve_sheet_month_grid(ws)
        code_by_row: dict[int, str] = {}
        for row_idx in range(2, (ws.max_row or 0) + 1):
            code = str(ws.cell(row_idx, 3).value or "").strip()
            if code:
                code_by_row[row_idx] = code
        sheet_rows: dict[str, dict[str, float | None]] = {}
        for row_idx, code in code_by_row.items():
            vals: dict[str, float | None] = {
                "year_forecast": _parse_amount(ws.cell(row_idx, 7).value),
            }
            for month in range(1, 13):
                col_idx = 7 + month
                vals[f"m{month}"] = grid.get((row_idx, col_idx))
            sheet_rows[code] = vals
        out[sheet_name] = sheet_rows
    wb.close()
    return out


def _load_output_workbook(path: Path) -> dict[str, dict[str, dict[str, float | None]]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    out: dict[str, dict[str, dict[str, float | None]]] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_rows: dict[str, dict[str, float | None]] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 5:
                continue
            code = str(row[2] or "").strip()
            if not code:
                continue
            vals: dict[str, float | None] = {"annual": _parse_amount(row[4])}
            for i in range(12):
                vals[f"m{i + 1}"] = _parse_amount(row[5 + i] if len(row) > 5 + i else None)
            sheet_rows[code] = vals
        out[sheet_name] = sheet_rows
    wb.close()
    return out


def _sheet_entity_table(conn: sqlite3.Connection) -> dict[str, tuple[str, str, str]]:
    mapping: dict[str, tuple[str, str, str]] = {}
    for row in load_org_product_metric_table_rows_from_runtime_tree(conn):
        ec = str(row.get("entity_code") or "").strip()
        en = str(row.get("entity_name") or "").strip()
        tn = str(row.get("table_name") or "").strip()
        if not ec or not tn:
            continue
        key = f"{ec}{en}{tn}".replace(" ", "")
        mapping[key] = (ec, en, tn)
    return mapping


def _match_output_sheet(std_sheet: str, output_sheets: list[str]) -> str | None:
    """Map entry sheet name (A01泛微粒贷业务状况表) to output sheet (A01泛微粒贷_业务状况表)."""
    std_norm = std_sheet.replace("_", "").replace(" ", "").lower()
    for s in output_sheets:
        if s.replace("_", "").replace(" ", "").lower() == std_norm:
            return s
    # fallback: match entity prefix + table suffix
    if "业务状况表" in std_sheet:
        suffix = "业务状况表"
    elif "业务支出评估" in std_sheet:
        suffix = "业务支出评估"
    elif "损益表" in std_sheet:
        suffix = "损益表"
    elif "资产质量表" in std_sheet:
        suffix = "资产质量表"
    elif "利息净收入表" in std_sheet:
        suffix = "利息净收入表"
    else:
        suffix = ""
    prefix = std_sheet.split(suffix)[0] if suffix else std_sheet[:6]
    prefix_norm = prefix.replace(" ", "")
    for s in output_sheets:
        if suffix in s and prefix_norm in s.replace("_", "").replace(" ", ""):
            return s
    return None


def _compare_numeric(
    actual: dict[str, float | None],
    expected: dict[str, float | None],
    *,
    keys: list[str],
    tol: float = 1.0,
) -> list[dict]:
    diffs: list[dict] = []
    for key in keys:
        a = actual.get(key)
        e = expected.get(key)
        if a is None and e is None:
            continue
        if a is None or e is None:
            diffs.append({"field": key, "actual": a, "expected": e, "delta": None})
            continue
        delta = abs(float(a) - float(e))
        rel = delta / max(abs(float(e)), 1.0)
        if delta > tol and rel > 1e-6:
            diffs.append({"field": key, "actual": a, "expected": e, "delta": delta})
    return diffs


def import_test_data(path: Path) -> dict:
    raw = path.read_bytes()
    db_path = common_db_path()
    db_path.parent.mkdir(parents=True, exist_ok=True)
    saved: list[dict] = []
    unmatched: list[dict] = []
    with sqlite3.connect(db_path) as conn:
        _ensure_data_entry_snapshot_table(conn)
        _ensure_data_entry_snapshot_table_v2(conn)
        candidates = [
            (
                str(row.get("entity_code") or "").strip(),
                str(row.get("entity_name") or "").strip(),
                str(row.get("table_name") or "").strip(),
            )
            for row in load_org_product_metric_table_rows_from_runtime_tree(conn)
            if str(row.get("entity_code") or "").strip() and str(row.get("table_name") or "").strip()
        ]
        sheets = _parse_data_entry_workbook(raw, path.name, YEAR, MONTH, candidates)
        for sheet in sheets:
            if not sheet.get("matched"):
                unmatched.append({"sheet_name": sheet.get("sheet_name")})
                continue
            entity_code = str(sheet.get("entity_code") or "").strip()
            table_name = str(sheet.get("table_name") or "").strip()
            payload_obj = {
                "entity_code": entity_code,
                "entity_name": _resolve_data_entry_entity_name(conn, entity_code),
                "year": YEAR,
                "month_index": MONTH,
                "version_id": VERSION_ID,
                "version_name": VERSION_NAME,
                "table_name": table_name,
                "entry_status": "draft",
                "metrics": list(sheet.get("metrics") or []),
            }
            payload_obj = _sanitize_data_entry_payload_mapping_refs(conn, payload_obj)
            updated_at = _persist_data_entry_snapshot(conn, payload_obj)
            metrics_with_data = sum(
                1
                for m in payload_obj.get("metrics") or []
                if any(
                    str(v).strip()
                    for v in ((m.get("values") or {}).get("months") or {}).values()
                )
            )
            saved.append(
                {
                    "entity_code": entity_code,
                    "table_name": table_name,
                    "sheet_name": sheet.get("sheet_name"),
                    "metrics_with_data": metrics_with_data,
                    "updated_at": updated_at,
                }
            )
        conn.commit()
    return {"saved": saved, "unmatched": unmatched, "saved_count": len(saved)}


async def export_prediction_output(export_path: Path) -> dict:
    payload = OrgProductOutputRunRequest(
        entity_code=ENTITY_ROOT,
        year=YEAR,
        version_id=VERSION_ID,
        table_name="",
        include_children=True,
    )
    result = await output_mod.run_org_product_output(payload)
    entities = list(result.get("entities") or [])
    wb = Workbook()
    wb.remove(wb.active)
    used_titles: set[str] = set()
    for ent in entities:
        ec = str(ent.get("entity_code") or "")
        en = str(ent.get("entity_name") or "")
        tn = str(ent.get("table_name") or "")
        title = _unique_sheet_title(used_titles, f"{ec}{en}_{tn}")
        ws = wb.create_sheet(title=title)
        _append_org_product_output_export_sheet(ws, list(ent.get("rows") or []), ec)
    export_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(export_path)
    return {"entity_count": len(entities), "path": str(export_path)}


def compare_with_formula_standard(actual_path: Path, std_path: Path) -> dict:
    actual = _load_output_workbook(actual_path)
    expected_by_entry_sheet = _load_standard_formula_workbook(std_path)
    keys_monthly = [f"m{i}" for i in range(1, 13)]
    sheet_reports: list[dict] = []
    total_diff_cells = 0
    p0_extreme: list[dict] = []
    perfect_sheets: list[str] = []

    output_sheets = list(actual.keys())
    for std_sheet, std_rows in expected_by_entry_sheet.items():
        act_sheet = _match_output_sheet(std_sheet, output_sheets)
        if not act_sheet:
            sheet_reports.append({"entry_sheet": std_sheet, "status": "missing_output_sheet"})
            continue
        act_rows = actual[act_sheet]
        code_diffs: list[dict] = []
        compared = 0
        matched = 0
        for code, exp_vals in std_rows.items():
            exp_has = any(v is not None for k, v in exp_vals.items() if k.startswith("m"))
            if not exp_has:
                continue
            act_vals = act_rows.get(code)
            if act_vals is None:
                code_diffs.append({"code": code, "issue": "missing_code", "diffs": []})
                continue
            diffs = _compare_numeric(
                act_vals,
                exp_vals,
                keys=keys_monthly,
                tol=1.0,
            )
            annual_diff: list[dict] = []
            if exp_vals.get("year_forecast") is not None:
                annual_diff = _compare_numeric(
                    act_vals,
                    {"annual": exp_vals.get("year_forecast")},
                    keys=["annual"],
                    tol=1.0,
                )
            diffs = diffs + annual_diff
            compared += 1
            if not diffs:
                matched += 1
                continue
            code_diffs.append({"code": code, "issue": "value_mismatch", "diffs": diffs})
            total_diff_cells += len(diffs)
            for d in diffs:
                a = d.get("actual")
                if isinstance(a, (int, float)) and abs(float(a)) >= 1e15:
                    p0_extreme.append(
                        {
                            "entry_sheet": std_sheet,
                            "output_sheet": act_sheet,
                            "code": code,
                            "field": d.get("field"),
                            "actual": a,
                            "expected": d.get("expected"),
                        }
                    )
        status = "ok" if not code_diffs else "diff"
        if status == "ok" and compared > 0:
            perfect_sheets.append(std_sheet)
        sheet_reports.append(
            {
                "entry_sheet": std_sheet,
                "output_sheet": act_sheet,
                "status": status,
                "compared_metrics": compared,
                "matched_metrics": matched,
                "diff_metric_count": len(code_diffs),
                "top_diffs": code_diffs[:8],
            }
        )

    return {
        "actual_output_sheets": len(actual),
        "standard_entry_sheets": len(expected_by_entry_sheet),
        "total_diff_cells": total_diff_cells,
        "perfect_sheets": perfect_sheets,
        "p0_extreme_values": p0_extreme[:30],
        "sheets": sheet_reports,
    }


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    export_path = OUT_DIR / f"预测输出_全量_AA_{YEAR}_v{VERSION_ID}_全部测试数据_{ts}.xlsx"
    report_path = OUT_DIR / f"compare_all_test_vs_formula_std_{ts}.json"
    summary_path = OUT_DIR / f"compare_all_test_vs_formula_std_{ts}.md"

    if not TEST_ENTRY.exists():
        if TEST_ENTRY_FALLBACK.exists():
            test_path = TEST_ENTRY_FALLBACK
        else:
            raise FileNotFoundError(TEST_ENTRY)
    else:
        test_path = TEST_ENTRY
    if not STD_FORMULA.exists():
        if STD_FORMULA_FALLBACK.exists():
            std_path = STD_FORMULA_FALLBACK
        else:
            raise FileNotFoundError(STD_FORMULA)
    else:
        std_path = STD_FORMULA

    print("=== 1) 导入：", test_path.name, "===")
    import_result = import_test_data(test_path)
    print(f"saved {import_result['saved_count']} sheets")
    for item in import_result["saved"]:
        print(f"  {item['entity_code']}/{item['table_name']}: {item['metrics_with_data']} metrics")

    print("\n=== 2) 跑批 + 导出预测输出 ===")
    export_result = asyncio.run(export_prediction_output(export_path))
    print(f"entities {export_result['entity_count']} -> {export_path}")

    print("\n=== 3) 对比：预测输出 vs", std_path.name, "===")
    compare_result = compare_with_formula_standard(export_path, std_path)
    diff_sheets = [s for s in compare_result["sheets"] if s.get("status") != "ok"]
    print(f"diff cells: {compare_result['total_diff_cells']}")
    print(f"perfect sheets: {len(compare_result['perfect_sheets'])} / {len(compare_result['sheets'])}")
    for s in diff_sheets[:10]:
        print(
            f"  {s.get('entry_sheet')}: compared={s.get('compared_metrics')} "
            f"matched={s.get('matched_metrics')} diff={s.get('diff_metric_count')}"
        )
        for top in s.get("top_diffs") or []:
            code = top.get("code")
            for d in (top.get("diffs") or [])[:2]:
                print(f"    {code} {d.get('field')}: out={d.get('actual')} std={d.get('expected')}")

    summary_lines = [
        "# 全部测试数据 → 预测输出 vs 完整版(含公式) 对比",
        "",
        f"- 导入文件: `{test_path.name}`",
        f"- 标准文件: `{std_path.name}`",
        f"- 导出文件: `{export_path.relative_to(ROOT)}`",
        f"- 导入 sheet: {import_result['saved_count']}",
        f"- 输出 sheet: {export_result['entity_count']}",
        f"- 差异单元格: {compare_result['total_diff_cells']}",
        f"- 完全一致 sheet: {len(compare_result['perfect_sheets'])} / {len(compare_result['sheets'])}",
        "",
        "## 完全一致 sheet",
        "",
    ]
    for name in compare_result["perfect_sheets"]:
        summary_lines.append(f"- {name}")
    summary_lines.extend(["", "## 存在差异 sheet", ""])
    for s in diff_sheets:
        summary_lines.append(
            f"- **{s.get('entry_sheet')}** → {s.get('output_sheet')}: "
            f"对比 {s.get('compared_metrics')} 项，一致 {s.get('matched_metrics')}，"
            f"差异 {s.get('diff_metric_count')}"
        )
    if compare_result["p0_extreme_values"]:
        summary_lines.extend(["", "## P0 极值", ""])
        for item in compare_result["p0_extreme_values"][:10]:
            summary_lines.append(
                f"- {item['entry_sheet']} {item['code']} {item['field']}: "
                f"{item['actual']} vs {item['expected']}"
            )

    report = {
        "timestamp": ts,
        "import_file": str(test_path),
        "standard_file": str(std_path),
        "import": import_result,
        "export": export_result,
        "compare": compare_result,
    }
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    summary_path.write_text("\n".join(summary_lines) + "\n", encoding="utf-8")
    print(f"\nReport: {report_path}")
    print(f"Summary: {summary_path}")


if __name__ == "__main__":
    main()
