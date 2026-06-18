#!/usr/bin/env python3
"""Import org-product metric formulas from v03 Excel into MySQL.

Maps v03 worksheet columns to ``data_account_metric_node``:

- budget_formula: 年预算公式 > 预测月公式 > 取数公式 > 年预测公式
- actual_formula: 实际月公式

Usage:
    cd apps/api && . .venv/bin/activate
    python scripts/import_v03_formulas_to_mysql.py --dry-run
    python scripts/import_v03_formulas_to_mysql.py --apply
    python scripts/import_v03_formulas_to_mysql.py --verify-only
"""
from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pymysql
from openpyxl import load_workbook

API_ROOT = Path(__file__).resolve().parents[1]
REPO_ROOT = API_ROOT.parents[1]
if str(API_ROOT) not in sys.path:
    sys.path.insert(0, str(API_ROOT))

from app.core.config import settings  # noqa: E402
from app.services.v03_metric_node_catalog import (  # noqa: E402
    is_v03_mirror_duplicate_row,
    merge_v03_node_payload,
)
from app.org_product_excel_formula import (  # noqa: E402
    build_sheet_formula_context,
    index_sheet_contexts,
    normalize_sheet_lookup_key,
)
from app.routers.org_product_helpers import (  # noqa: E402
    _extract_owner_code,
    _find_header_row,
    _maybe_swap_metric_code_name_columns,
    _normalize_metric_code,
    _parse_metric_worksheet_basic,
    _prepare_metric_worksheet,
    _resolve_import_sheet_entity_table,
    _sheet_scan_row_limit,
    _ws_cell_value,
)

DEFAULT_WORKBOOK = (
    settings.business_inputs_dir / "机构及产品指标（公式配置） - v03.xlsx"
)


@dataclass(frozen=True)
class FormulaPatch:
    node_code: str
    budget_formula: str | None
    actual_formula: str | None
    old_budget_formula: str | None
    old_actual_formula: str | None

    @property
    def budget_changed(self) -> bool:
        return _norm(self.budget_formula) != _norm(self.old_budget_formula)

    @property
    def actual_changed(self) -> bool:
        return _norm(self.actual_formula) != _norm(self.old_actual_formula)

    @property
    def changed(self) -> bool:
        return self.budget_changed or self.actual_changed


def _norm(value: str | None) -> str:
    return str(value or "").strip()


def _clean_formula_value(value: Any) -> str:
    if value is None:
        return ""
    if not isinstance(value, str):
        value = str(value)
    text = _norm(value)
    if text.startswith("<openpyxl"):
        return ""
    return text


def resolve_v03_db_formulas(node: dict[str, Any]) -> tuple[str, str]:
    budget_annual = _clean_formula_value(node.get("formula_budget_annual"))
    forecast = _clean_formula_value(node.get("formula_forecast"))
    legacy = _clean_formula_value(node.get("formula"))
    forecast_annual = _clean_formula_value(node.get("formula_forecast_annual"))
    actual = _clean_formula_value(node.get("formula_actual"))
    budget_formula = budget_annual or forecast or legacy or forecast_annual
    return budget_formula, actual


def _iter_metric_nodes(nodes: list[dict[str, Any]]):
    stack = list(nodes)
    while stack:
        node = stack.pop(0)
        if not isinstance(node, dict):
            continue
        yield node
        children = node.get("children")
        if isinstance(children, list):
            stack[0:0] = [child for child in children if isinstance(child, dict)]


def _build_workbook_formula_contexts(wb) -> dict[str, Any]:
    contexts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        _prepare_metric_worksheet(ws)
        resolved = _resolve_import_sheet_entity_table(sheet_name, [], strict=False)
        if not resolved:
            entity_code = _extract_owner_code(sheet_name, []) or ""
            if not entity_code:
                continue
            table_name = sheet_name
        else:
            entity_code, table_name, _entity_name = resolved
        header_row_idx, header_map, _header_mode, _header_source = _find_header_row(
            ws,
            entity_code,
            strict=False,
            sheet_title=sheet_name,
        )
        code_col = header_map.get("科目代码") if header_row_idx else None
        name_col = header_map.get("科目名称") if header_row_idx else None
        if header_row_idx and code_col and name_col:
            code_col, _name_col = _maybe_swap_metric_code_name_columns(
                ws,
                header_row_idx=header_row_idx,
                entity_code=entity_code,
                code_col=code_col,
                name_col=name_col,
            )
        if not header_row_idx or not code_col:
            continue
        row_limit = _sheet_scan_row_limit(ws, header_row_idx)
        contexts.append(
            build_sheet_formula_context(
                sheet_name,
                entity_code,
                table_name,
                header_row_idx,
                code_col,
                lambda r, c, _ws=ws: _ws_cell_value(_ws, r, c),
                _normalize_metric_code,
                row_limit,
            )
        )
    return index_sheet_contexts(contexts)


def load_v03_formula_rows(workbook_path: Path) -> dict[str, dict[str, str]]:
    if not workbook_path.exists():
        raise FileNotFoundError(f"v03 workbook not found: {workbook_path}")

    wb = load_workbook(workbook_path, data_only=False, read_only=False)
    all_sheet_contexts = _build_workbook_formula_contexts(wb)
    merged: dict[str, dict[str, str]] = {}
    formula_convert_errors: list[dict[str, Any]] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        _prepare_metric_worksheet(ws)
        resolved = _resolve_import_sheet_entity_table(sheet_name, [], strict=False)
        entity_code = (
            resolved[0]
            if resolved
            else (_extract_owner_code(sheet_name, []) or "")
        )
        if not entity_code:
            continue
        sheet_ctx = all_sheet_contexts.get(normalize_sheet_lookup_key(sheet_name))
        metrics, _row_count, parse_error, _header_map = _parse_metric_worksheet_basic(
            ws,
            f"v03-{sheet_name}",
            entity_code=entity_code,
            strict=False,
            sheet_formula_context=sheet_ctx,
            all_sheet_contexts=all_sheet_contexts,
            formula_convert_errors=formula_convert_errors,
        )
        if parse_error or not metrics:
            continue
        for node in _iter_metric_nodes(metrics):
            code = _norm(node.get("code")).upper()
            if not code:
                continue
            if is_v03_mirror_duplicate_row(sheet_name, code):
                continue
            budget_formula, actual_formula = resolve_v03_db_formulas(node)
            if not budget_formula and not actual_formula:
                continue
            payload = {
                "formula_budget_annual": _clean_formula_value(node.get("formula_budget_annual")),
                "formula_forecast": _clean_formula_value(node.get("formula_forecast")),
                "formula": _clean_formula_value(node.get("formula")),
                "formula_forecast_annual": _clean_formula_value(node.get("formula_forecast_annual")),
                "formula_actual": _clean_formula_value(node.get("formula_actual")),
                "budget_formula": budget_formula,
                "actual_formula": actual_formula,
            }
            existing = merged.get(code)
            if existing is None:
                merged[code] = payload
                continue
            for key, value in payload.items():
                if value and not existing.get(key):
                    existing[key] = value
    if formula_convert_errors:
        print(f"formula convert warnings: {len(formula_convert_errors)}")
    return merged


def _mysql_connect() -> pymysql.connections.Connection:
    return pymysql.connect(
        host=settings.MYSQL_HOST,
        port=int(settings.MYSQL_PORT),
        user=settings.MYSQL_USER,
        password=settings.MYSQL_PASSWORD or "",
        database=settings.MYSQL_DATABASE,
        charset="utf8mb4",
        autocommit=False,
    )


def load_mysql_formula_state(conn: pymysql.connections.Connection) -> dict[str, tuple[str, str]]:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT node_code, COALESCE(budget_formula, ''), COALESCE(actual_formula, '')
            FROM data_account_metric_node
            WHERE is_active = 1
            """
        )
        rows = cur.fetchall()
    return {
        str(row[0] or "").strip().upper(): (str(row[1] or ""), str(row[2] or ""))
        for row in rows
        if str(row[0] or "").strip()
    }


def build_patches(
    v03_rows: dict[str, dict[str, str]],
    db_state: dict[str, tuple[str, str]],
) -> list[FormulaPatch]:
    patches: list[FormulaPatch] = []
    for node_code, db_values in sorted(db_state.items()):
        v03 = v03_rows.get(node_code)
        if not v03:
            continue
        budget_formula = v03.get("budget_formula") or None
        actual_formula = v03.get("actual_formula") or None
        if not budget_formula and not actual_formula:
            continue
        old_budget, old_actual = db_values
        patch = FormulaPatch(
            node_code=node_code,
            budget_formula=budget_formula or None,
            actual_formula=actual_formula or None,
            old_budget_formula=old_budget or None,
            old_actual_formula=old_actual or None,
        )
        if patch.changed:
            patches.append(patch)
    return patches


def apply_patches(conn: pymysql.connections.Connection, patches: list[FormulaPatch]) -> int:
    updated = 0
    with conn.cursor() as cur:
        for patch in patches:
            sets: list[str] = []
            params: list[Any] = []
            if patch.budget_changed:
                sets.append("budget_formula = %s")
                params.append(patch.budget_formula)
            if patch.actual_changed:
                sets.append("actual_formula = %s")
                params.append(patch.actual_formula)
            if not sets:
                continue
            sets.append("updated_at = CURRENT_TIMESTAMP")
            params.append(patch.node_code)
            cur.execute(
                f"UPDATE data_account_metric_node SET {', '.join(sets)} WHERE node_code = %s",
                params,
            )
            updated += cur.rowcount
    conn.commit()
    return updated


def count_non_empty_formulas(conn: pymysql.connections.Connection) -> tuple[int, int]:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT COUNT(*)
            FROM data_account_metric_node
            WHERE is_active = 1 AND TRIM(COALESCE(budget_formula, '')) <> ''
            """
        )
        budget_count = int(cur.fetchone()[0])
        cur.execute(
            """
            SELECT COUNT(*)
            FROM data_account_metric_node
            WHERE is_active = 1 AND TRIM(COALESCE(actual_formula, '')) <> ''
            """
        )
        actual_count = int(cur.fetchone()[0])
    return budget_count, actual_count


def verify_against_v03(
    v03_rows: dict[str, dict[str, str]],
    db_state: dict[str, tuple[str, str]],
) -> tuple[int, list[str]]:
    mismatches: list[str] = []
    checked = 0
    for node_code, v03 in v03_rows.items():
        if node_code not in db_state:
            continue
        expected_budget = v03.get("budget_formula") or ""
        expected_actual = v03.get("actual_formula") or ""
        if not expected_budget and not expected_actual:
            continue
        checked += 1
        actual_budget, actual_actual = db_state[node_code]
        if _norm(actual_budget) != _norm(expected_budget) or _norm(actual_actual) != _norm(expected_actual):
            mismatches.append(node_code)
    return checked, mismatches


def _print_summary(
    *,
    patches: list[FormulaPatch],
    budget_before: int,
    actual_before: int,
    budget_after: int | None = None,
    actual_after: int | None = None,
) -> None:
    budget_updates = sum(1 for patch in patches if patch.budget_changed)
    actual_updates = sum(1 for patch in patches if patch.actual_changed)
    print(f"planned patches: {len(patches)} (budget={budget_updates}, actual={actual_updates})")
    print(f"budget_formula non-empty: {budget_before}" + (f" -> {budget_after}" if budget_after is not None else ""))
    print(f"actual_formula non-empty: {actual_before}" + (f" -> {actual_after}" if actual_after is not None else ""))
    for patch in patches[:10]:
        print(f"  {patch.node_code}")
        if patch.budget_changed:
            print(f"    budget: { _short(patch.old_budget_formula) } -> { _short(patch.budget_formula) }")
        if patch.actual_changed:
            print(f"    actual: { _short(patch.old_actual_formula) } -> { _short(patch.actual_formula) }")
    if len(patches) > 10:
        print(f"  ... and {len(patches) - 10} more")


def _short(value: str | None, limit: int = 72) -> str:
    text = _norm(value)
    if len(text) <= limit:
        return text or "(empty)"
    return text[: limit - 3] + "..."


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    mode = parser.add_mutually_exclusive_group()
    mode.add_argument("--dry-run", action="store_true", help="Show planned updates only (default)")
    mode.add_argument("--apply", action="store_true", help="Write updates to MySQL")
    mode.add_argument("--verify-only", action="store_true", help="Verify DB formulas against v03 mapping")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK, help="Path to v03 workbook")
    args = parser.parse_args(argv)

    if not args.apply and not args.verify_only:
        args.dry_run = True

    v03_rows = load_v03_formula_rows(args.workbook)
    print(f"loaded v03 formula rows: {len(v03_rows)}")

    with _mysql_connect() as conn:
        db_state = load_mysql_formula_state(conn)
        budget_before, actual_before = count_non_empty_formulas(conn)

        if args.verify_only:
            checked, mismatches = verify_against_v03(v03_rows, db_state)
            print(f"verified codes: {checked}")
            print(f"mismatches: {len(mismatches)}")
            if mismatches[:20]:
                print("sample:", ", ".join(mismatches[:20]))
            return 1 if mismatches else 0

        patches = build_patches(v03_rows, db_state)
        if args.dry_run:
            _print_summary(
                patches=patches,
                budget_before=budget_before,
                actual_before=actual_before,
            )
            return 0

        updated = apply_patches(conn, patches)
        db_state = load_mysql_formula_state(conn)
        budget_after, actual_after = count_non_empty_formulas(conn)
        checked, mismatches = verify_against_v03(v03_rows, db_state)
        print(f"rows updated: {updated}")
        _print_summary(
            patches=patches,
            budget_before=budget_before,
            actual_before=actual_before,
            budget_after=budget_after,
            actual_after=actual_after,
        )
        print(f"verify checked={checked} mismatches={len(mismatches)}")
        if mismatches[:10]:
            print("sample mismatches:", ", ".join(mismatches[:10]))
        return 1 if mismatches else 0


if __name__ == "__main__":
    raise SystemExit(main())
