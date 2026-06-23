#!/usr/bin/env python3
"""Generate AA 业务状况表 test data (~300亿营收) into the data-entry workbook."""
from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
V04_PATH = ROOT / "机构及产品指标表 v04 0622.xlsx"
DATA_PATH = ROOT / "机构产品数据录入_AA_业务状况表_2026.xlsx"

# 300亿元 = 3,000,000 万元
WAN = 1.0
REVENUE_WAN = 3_000_000


def parse_v04_formulas() -> dict[str, list[tuple[int, str]]]:
    ws = load_workbook(V04_PATH, data_only=False)["AA业务状况表"]
    row_code = {
        r: str(ws.cell(r, 3).value).strip()
        for r in range(2, ws.max_row + 1)
        if ws.cell(r, 3).value
    }
    formulas: dict[str, list[tuple[int, str]]] = {}
    for r, code in row_code.items():
        raw = ws.cell(r, 8).value
        if not raw or not str(raw).startswith("="):
            continue
        refs = [int(x) for x in re.findall(r"H(\d+)", str(raw))]
        ops = re.findall(r"([+\-])", str(raw)[1:])
        terms: list[tuple[int, str]] = []
        for i, ref in enumerate(refs):
            sign = 1 if i == 0 else (1 if ops[i - 1] == "+" else -1)
            terms.append((sign, row_code.get(ref, f"R{ref}")))
        formulas[code] = terms
    # 数据录入表与 v04 在 AA.05 结构上略有差异，按录入表口径修正
    formulas["AA.05"] = [
        (1, "AA.05.01"),
        (1, "AA.05.02"),
        (1, "AA.05.03"),
    ]
    return formulas


def metric_kind(code: str, nature: str, rule: str | None, has_formula: bool) -> str:
    if has_formula:
        return "formula"
    nature = (nature or "").strip()
    rule = (rule or "").strip().upper()
    if rule == "CALC" or nature in {"其他"} and code.startswith("AA.4"):
        return "ratio"
    if nature in {"资产余额", "负债余额"} or rule == "LAST":
        return "balance"
    if nature in {"投入产出"} and any(x in code for x in (".101", ".102", ".103")):
        return "balance"
    if nature in {"投入产出"} and ("客户数" in code or code.endswith(".003") or code.endswith(".004") or code.endswith(".009")):
        return "count"
    if nature in {"投入产出"}:
        return "amount"
    return "amount"


def build_v04_meta() -> dict[str, dict[str, str]]:
    ws = load_workbook(V04_PATH, data_only=False)["AA业务状况表"]
    meta: dict[str, dict[str, str]] = {}
    formulas = parse_v04_formulas()
    for r in range(2, ws.max_row + 1):
        code = ws.cell(r, 3).value
        if not code:
            continue
        code = str(code).strip()
        meta[code] = {
            "nature": str(ws.cell(r, 2).value or "").strip(),
            "rule": str(ws.cell(r, 14).value or "").strip(),
            "kind": metric_kind(
                code,
                str(ws.cell(r, 2).value or ""),
                str(ws.cell(r, 14).value or ""),
                code in formulas,
            ),
        }
    return meta


def leaf_annual_values() -> dict[str, float]:
    """Driver values for leaf metrics (annual, 万元 unless ratio/count)."""
    r = REVENUE_WAN
    v: dict[str, float] = {}

    # 营业收入构成：目标 AA.01 = 300亿 = 3,000,000 万元
    # 利息净收入 210亿 / 净手续费 60亿 / 其他 30亿
    v["AA.14"] = 2_800_000.0
    v["AA.16"] = 700_000.0
    v["AA.18"] = 750_000.0
    v["AA.19"] = 150_000.0
    v["AA.01.03"] = 300_000.0

    # 减值（约占营收 20%）
    v["AA.02.01.01"] = 480_000.0
    v["AA.02.01.02"] = 75_000.0
    v["AA.02.02"] = 45_000.0

    # 税金 / 营业外
    v["AA.04"] = 45_000.0
    v["AA.06"] = 15_000.0

    # 业务及管理费（约占营收 27%）
    v["AA.05.01.01"] = 360_000.0
    v["AA.05.01.02"] = 75_000.0
    v["AA.05.01.03"] = 15_000.0
    v["AA.05.02.01"] = 135_000.0
    v["AA.05.02.02"] = 165_000.0
    v["AA.05.03"] = 60_000.0

    v["AA.30"] = 165_000.0  # 所得税

    # 资产负债（时点，万元）
    loan = 3_500_000.0
    v["AA.10.01.01"] = loan * 0.72
    v["AA.10.01.02"] = loan * 0.08
    v["AA.10.02"] = loan * 0.12
    v["AA.10.03"] = loan * 0.05
    v["AA.10.04"] = loan * 0.03

    deposit = 2_800_000.0
    v["AA.12.01"] = deposit * 0.82
    v["AA.12.02"] = deposit * 0.10
    v["AA.12.03"] = deposit * 0.05
    v["AA.12.04"] = deposit * 0.03

    v["AA.47.01.01"] = loan * 0.025
    v["AA.47.02"] = loan * 0.004
    v["AA.54.01"] = loan * 0.024
    v["AA.54.02"] = loan * 0.0038

    # 比率类（直接填展示值：%/比率）
    v["AA.45"] = 238.5
    v["AA.46"] = 2.85
    v["AA.38"] = 1.02
    v["AA.48.01"] = 0.18
    v["AA.49.01"] = 20.0
    v["AA.49.02"] = 28.0
    v["AA.49.03"] = 1.05
    v["AA.49.04"] = 18.5
    v["AA.49.05"] = 32.0

    # AA.90 费用分解
    expense = 810_000.0
    direct = expense * 0.72
    indirect = expense * 0.28
    v["AA.90.01.01.01"] = direct * 0.28
    v["AA.90.01.01.02"] = direct * 0.07
    v["AA.90.01.02.01"] = direct * 0.12
    v["AA.90.01.02.02"] = direct * 0.10
    v["AA.90.01.03.01.001"] = direct * 0.10
    v["AA.90.01.03.01.002"] = direct * 0.02
    v["AA.90.01.03.02"] = direct * 0.04
    v["AA.90.01.03.03"] = direct * 0.03
    v["AA.90.01.03.04"] = direct * 0.02
    v["AA.90.01.04"] = direct * 0.02
    v["AA.90.01.05"] = direct * 0.02
    v["AA.90.02.01.01"] = indirect * 0.30
    v["AA.90.02.01.02"] = indirect * 0.08
    v["AA.90.02.02.01"] = indirect * 0.10
    v["AA.90.02.02.02"] = indirect * 0.08
    v["AA.90.02.03.01.001"] = indirect * 0.12
    v["AA.90.02.03.01.002"] = indirect * 0.03
    v["AA.90.02.03.02"] = indirect * 0.05
    v["AA.90.02.03.03"] = indirect * 0.04
    v["AA.90.02.03.04"] = indirect * 0.03
    v["AA.90.02.04"] = indirect * 0.03
    v["AA.90.02.05"] = indirect * 0.04

    # 业务支出评估 / 客户经营
    v["AA.91.01.01.001"] = 54_000.0
    v["AA.91.01.01.002"] = 36_000.0
    v["AA.91.01.01.003"] = 24_000.0
    v["AA.91.01.01.004"] = 30_000.0
    v["AA.91.01.02.001"] = 850_000.0
    v["AA.91.01.02.002"] = 2_100_000.0
    v["AA.91.01.02.005"] = 120_000.0
    v["AA.91.01.02.006"] = 95_000.0
    v["AA.91.01.02.010"] = 2_950_000.0
    v["AA.91.01.02.007"] = 680_000.0
    v["AA.91.01.02.008"] = 45_000.0
    v["AA.91.01.02.011"] = 725_000.0
    v["AA.91.01.02.003"] = 1_250.0  # 万户
    v["AA.91.01.02.004"] = 3_800.0
    v["AA.91.01.02.009"] = 2_100.0
    v["AA.91.01.03.001"] = 66_000.0
    v["AA.91.01.03.002"] = 24_000.0
    v["AA.91.01.03.003"] = 18_000.0

    return v


SEASONAL = [0.92, 0.88, 1.05, 1.02, 1.00, 0.98, 0.96, 0.97, 1.04, 1.06, 1.08, 1.04]


def period_keys() -> list[str]:
    keys = ["prev_actual"]
    keys.extend(f"a{m}" for m in range(1, 4))
    keys.extend(f"f{m}" for m in range(4, 13))
    return keys


def month_index(key: str) -> int | None:
    if key == "prev_actual":
        return None
    if key.startswith("a"):
        return int(key[1:])
    if key.startswith("f"):
        return int(key[1:])
    return None


def scale_for_period(base: float, key: str, kind: str) -> float:
    mi = month_index(key)
    if key == "prev_actual":
        return base
    if kind == "ratio":
        return base * (1.0 + (mi or 0) * 0.001)
    if kind == "count":
        if key.startswith("a"):
            return base * (mi or 1) / 12.0
        return base * (mi or 1) / 12.0 * 1.05
    if kind == "balance":
        growth = 1.0 + (mi or 0) * 0.004
        return base * growth
    # amount (flow): monthly with seasonality
    factor = SEASONAL[(mi or 1) - 1]
    return base / 12.0 * factor


def evaluate_formulas(
    values: dict[str, float],
    formulas: dict[str, list[tuple[int, str]]],
    codes: set[str],
) -> None:
    for _ in range(8):
        changed = False
        for code, terms in formulas.items():
            if code not in codes:
                continue
            total = 0.0
            ok = True
            for sign, term_code in terms:
                if term_code not in values:
                    ok = False
                    break
                total += sign * values[term_code]
            if ok:
                new_val = round(total, 2)
                if values.get(code) != new_val:
                    values[code] = new_val
                    changed = True
        if not changed:
            break


def fmt_num(value: float, kind: str) -> str | int | float:
    if kind == "ratio":
        return round(value, 4)
    if kind == "count":
        return round(value, 2)
    return round(value, 2)


def generate_all_periods(
    codes: list[str],
    meta: dict[str, dict[str, str]],
    formulas: dict[str, list[tuple[int, str]]],
    leaves: dict[str, float],
) -> dict[str, dict[str, float]]:
    code_set = set(codes)
    out: dict[str, dict[str, float]] = {c: {} for c in codes}
    growth_2026 = 1.06

    for key in period_keys():
        period_values: dict[str, float] = {}
        for code in codes:
            info = meta.get(code, {"kind": "amount"})
            kind = info.get("kind", "amount")
            if code in formulas:
                continue
            base = leaves.get(code)
            if base is None:
                continue
            if key != "prev_actual" and kind == "amount":
                base = base * growth_2026
            period_values[code] = scale_for_period(base, key, kind)

        evaluate_formulas(period_values, formulas, code_set)
        for code in codes:
            if code in period_values:
                out[code][key] = period_values[code]

    return out


def main() -> None:
    formulas = parse_v04_formulas()
    meta = build_v04_meta()
    leaves = leaf_annual_values()

    wb = load_workbook(DATA_PATH)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col_by_header = {str(h): i for i, h in enumerate(headers, start=1) if h}
    code_col = col_by_header["科目代码"]

    header_to_key = {
        "25年实际": "prev_actual",
        "26年预算": "prev_budget",
        "26年预测": "prev_forecast",
    }
    for m in range(1, 4):
        header_to_key[f"26年{m}月实际"] = f"a{m}"
    for m in range(4, 13):
        header_to_key[f"26年{m}月预测"] = f"f{m}"

    codes: list[str] = []
    row_by_code: dict[str, int] = {}
    for r in range(2, ws.max_row + 1):
        code = ws.cell(r, code_col).value
        if not code:
            continue
        code = str(code).strip()
        codes.append(code)
        row_by_code[code] = r

    all_values = generate_all_periods(codes, meta, formulas, leaves)

    for code, row in row_by_code.items():
        period_vals = all_values.get(code, {})
        for header, key in header_to_key.items():
            col = col_by_header.get(header)
            if not col:
                continue
            val = period_vals.get(key)
            if val is None:
                continue
            kind = meta.get(code, {}).get("kind", "amount")
            ws.cell(row, col).value = fmt_num(val, kind)

        # 26年预算：在 prev_actual 基础上 +8%
        budget_col = col_by_header.get("26年预算")
        prev_col = col_by_header.get("25年实际")
        if budget_col and prev_col:
            prev = ws.cell(row, prev_col).value
            if prev is not None and str(prev).strip() != "":
                try:
                    ws.cell(row, budget_col).value = fmt_num(float(prev) * 1.08, meta.get(code, {}).get("kind", "amount"))
                except (TypeError, ValueError):
                    pass

        # 26年预测：实际月 + 预测月汇总
        fc_col = col_by_header.get("26年预测")
        if fc_col:
            total = 0.0
            has = False
            for m in range(1, 4):
                v = period_vals.get(f"a{m}")
                if v is not None:
                    total += v
                    has = True
            for m in range(4, 13):
                v = period_vals.get(f"f{m}")
                if v is not None:
                    total += v
                    has = True
            if has:
                kind = meta.get(code, {}).get("kind", "amount")
                if kind in {"ratio", "balance"}:
                    # 比率/余额类预测列填 3 月实际或最新时点
                    ws.cell(row, fc_col).value = fmt_num(period_vals.get("a3") or period_vals.get("f12") or total, kind)
                else:
                    ws.cell(row, fc_col).value = fmt_num(total, kind)

    wb.save(DATA_PATH)

    aa01_prev = all_values.get("AA.01", {}).get("prev_actual")
    aa01_fc = 0.0
    for k in period_keys():
        if k == "prev_actual":
            continue
        v = all_values.get("AA.01", {}).get(k)
        if v:
            aa01_fc += v
    print(f"已写入: {DATA_PATH}")
    print(f"AA.01 2025年实际: {aa01_prev:,.2f} 万元 ({(aa01_prev or 0)/10000:,.2f} 亿元)")
    print(f"AA.01 2026年月度合计: {aa01_fc:,.2f} 万元 ({aa01_fc/10000:,.2f} 亿元)")


if __name__ == "__main__":
    main()
