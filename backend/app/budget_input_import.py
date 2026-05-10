"""预算基础数据 Excel 导入（同步解析与写库）。"""
from __future__ import annotations

import re
import sqlite3
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.budget_window import budget_actual_allowed_for_month
from app.db_paths import budget_db_path, common_db_path
from app.months import parse_month_index
from app.schemas import (
    BudgetInputImportMonthResult,
    BudgetInputImportResponse,
    BudgetInputImportResultRow,
)


def budget_input_template_path() -> Path:
    return Path(__file__).resolve().parent.parent.parent / "download_template" / "budget_data_temp.xlsx"


def _code_prefix_five(raw: Any) -> str:
    """模版第6列数据科目、第8列产品科目：单元格常为「代码+名称」全文，只取前五个字符与主数据代码对齐。"""
    if raw is None:
        return ""
    s = str(raw).strip().upper()
    if not s:
        return ""
    return s[:5]


def _excel_to_stored(value: float, value_type: str) -> float:
    if value_type == "百分比":
        return value / 100.0
    return float(value)


def _parse_cell_number(raw: Any) -> tuple[bool, float | None, str | None]:
    """has_value, numeric_value_or_none, error_message"""
    if raw is None:
        return False, None, None
    if isinstance(raw, str):
        s = raw.strip().replace(",", "").replace("，", "")
        if not s:
            return False, None, None
        try:
            return True, float(s), None
        except ValueError:
            return True, None, "单元格不是有效数字"
    if isinstance(raw, bool):
        return True, float(raw), None
    if isinstance(raw, (int, float)):
        return True, float(raw), None
    return True, None, "单元格不是有效数字"


def _formula_locked_for_sheet(
    cconn: sqlite3.Connection,
    data_acct_code: str,
    *,
    is_budget_sheet: bool,
) -> bool:
    """
    预算数据工作表只看 budget_formula；实际数据工作表只看 actual_formula。
    二者互不混用。
    """
    cur = cconn.execute(
        "SELECT budget_formula, actual_formula FROM data_account WHERE data_acct_code = ?",
        (data_acct_code.strip().upper(),),
    )
    row = cur.fetchone()
    if not row:
        return False
    raw = row[0] if is_budget_sheet else row[1]
    return bool((raw or "").strip())


def run_budget_excel_import(
    file_bytes: bytes,
    version_id: int,
    budget_year: int,
) -> tuple[BudgetInputImportResponse, set[str]]:
    """返回导入结果与需要重算公式的 product_code 集合。"""
    common_p = common_db_path()
    budget_p = budget_db_path(budget_year)

    cconn = sqlite3.connect(str(common_p))
    cconn.execute("PRAGMA foreign_keys = ON")
    bconn = sqlite3.connect(str(budget_p))
    bconn.execute("PRAGMA foreign_keys = ON")

    try:
        cur_v = bconn.execute(
            "SELECT current_month FROM version WHERE version_id = ?",
            (version_id,),
        )
        vrow = cur_v.fetchone()
        if not vrow:
            raise ValueError(f"版本 {version_id} 不存在")
        current_month = int(vrow[0] or 1)
        if current_month < 1 or current_month > 13:
            current_month = 1

        year_label = f"Y{budget_year}"
        cur_p = cconn.execute(
            """
            SELECT period_id, month FROM period WHERE year = ? ORDER BY period_id
            """,
            (year_label,),
        )
        period_month_map: dict[int, int] = {}
        for pid, mlab in cur_p.fetchall():
            mi = parse_month_index(str(mlab or ""))
            if 1 <= mi <= 12:
                period_month_map[int(pid)] = mi
        month_to_period: dict[int, int] = {m: pid for pid, m in period_month_map.items()}

        cur_da = cconn.execute(
            """
            SELECT data_acct_code, value_type, product_code, product_codes
            FROM data_account
            """
        )
        da_map: dict[str, dict[str, Any]] = {}
        for r in cur_da.fetchall():
            da_map[str(r[0]).strip().upper()] = {
                "value_type": str(r[1] or "金额"),
                "product_code": str(r[2] or "").strip().upper() if r[2] else "",
                "is_all_products": (r[3] is None or (r[3] or "").strip() == ""),
            }

        cur_pt = cconn.execute("SELECT product_code FROM product_type")
        product_codes = {str(r[0]).strip().upper() for r in cur_pt.fetchall() if r[0]}

        wb = load_workbook(BytesIO(file_bytes), data_only=True)
        result_rows: list[BudgetInputImportResultRow] = []
        saved_cells = 0
        products_touched: set[str] = set()
        x = max(1, min(13, current_month))

        sheet_specs = [("预算数据", 0), ("实际数据", 1)]

        for sheet_name, budget_actual in sheet_specs:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            max_r = ws.max_row or 0

            sheet_whole_block = (sheet_name == "预算数据" and x == 13) or (
                sheet_name == "实际数据" and x == 1
            )
            block_reason = ""
            if sheet_name == "预算数据" and x == 13:
                block_reason = "当前月份窗口为13（年度已结束），不导入预算数"
            elif sheet_name == "实际数据" and x == 1:
                block_reason = "当前月份窗口为1，不导入实际数"

            is_budget_sheet = sheet_name == "预算数据"

            for r in range(2, max_r + 1):
                dc = ws.cell(row=r, column=6).value
                pc = ws.cell(row=r, column=8).value
                data_s = _code_prefix_five(dc)
                prod_s = _code_prefix_five(pc)
                if not data_s and not prod_s:
                    continue
                months_out: list[BudgetInputImportMonthResult] = []
                row_notes: list[str] = []

                if not data_s or not prod_s:
                    note = "数据科目代码或产品科目代码为空"
                    for m in range(1, 13):
                        months_out.append(
                            BudgetInputImportMonthResult(
                                month=m,
                                value_text="",
                                status="error",
                                reason=note,
                            )
                        )
                    result_rows.append(
                        BudgetInputImportResultRow(
                            sheet_name=sheet_name,
                            excel_row=r,
                            data_acct_code=data_s or "—",
                            product_code=prod_s or "—",
                            months=months_out,
                            note=note,
                        )
                    )
                    continue

                ok_master = True
                master_note = ""
                if data_s not in da_map:
                    ok_master = False
                    master_note = "数据科目代码不在主数据中"
                elif prod_s not in product_codes:
                    ok_master = False
                    master_note = "产品科目代码不在主数据中"
                else:
                    acct = da_map[data_s]
                    if (acct.get("product_codes") or "").strip() != "":
                        apc = acct["product_code"] or ""
                        if apc != prod_s:
                            ok_master = False
                            master_note = "数据科目与产品科目组合与主数据不匹配"

                locked = False
                if ok_master:
                    locked = _formula_locked_for_sheet(
                        cconn,
                        data_s,
                        is_budget_sheet=is_budget_sheet,
                    )

                if sheet_whole_block:
                    for m in range(1, 13):
                        months_out.append(
                            BudgetInputImportMonthResult(
                                month=m,
                                value_text="",
                                status="error",
                                reason=block_reason,
                            )
                        )
                    result_rows.append(
                        BudgetInputImportResultRow(
                            sheet_name=sheet_name,
                            excel_row=r,
                            data_acct_code=data_s,
                            product_code=prod_s,
                            months=months_out,
                            note=block_reason,
                        )
                    )
                    continue

                if not ok_master:
                    for m in range(1, 13):
                        months_out.append(
                            BudgetInputImportMonthResult(
                                month=m,
                                value_text="",
                                status="error",
                                reason=master_note,
                            )
                        )
                    result_rows.append(
                        BudgetInputImportResultRow(
                            sheet_name=sheet_name,
                            excel_row=r,
                            data_acct_code=data_s,
                            product_code=prod_s,
                            months=months_out,
                            note=master_note,
                        )
                    )
                    continue

                if locked:
                    fnote = (
                        "该数据科目已配置预算计算公式，不可手工导入预算数据"
                        if is_budget_sheet
                        else "该数据科目已配置实际计算公式，不可手工导入实际数据"
                    )
                    vt_locked = da_map[data_s]["value_type"]
                    for m in range(1, 13):
                        col = 9 + m
                        cell_val = ws.cell(row=r, column=col).value
                        has_v, num_v, perr = _parse_cell_number(cell_val)
                        if not has_v:
                            months_out.append(
                                BudgetInputImportMonthResult(month=m, value_text="", status="empty")
                            )
                            continue
                        if perr or num_v is None:
                            months_out.append(
                                BudgetInputImportMonthResult(
                                    month=m,
                                    value_text=str(cell_val) if cell_val is not None else "",
                                    status="error",
                                    reason=fnote,
                                )
                            )
                            continue
                        disp = f"{num_v:.4g}" if vt_locked != "百分比" else f"{num_v:.4g}%"
                        months_out.append(
                            BudgetInputImportMonthResult(
                                month=m,
                                value_text=disp,
                                status="error",
                                reason=fnote,
                            )
                        )
                    result_rows.append(
                        BudgetInputImportResultRow(
                            sheet_name=sheet_name,
                            excel_row=r,
                            data_acct_code=data_s,
                            product_code=prod_s,
                            months=months_out,
                            note=fnote,
                        )
                    )
                    continue

                vt = da_map[data_s]["value_type"]
                row_now = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

                for m in range(1, 13):
                    col = 9 + m
                    cell_val = ws.cell(row=r, column=col).value
                    has_v, num_v, perr = _parse_cell_number(cell_val)
                    if not has_v:
                        months_out.append(
                            BudgetInputImportMonthResult(month=m, value_text="", status="empty")
                        )
                        continue
                    if perr or num_v is None:
                        msg = perr or "无效数字"
                        months_out.append(
                            BudgetInputImportMonthResult(
                                month=m,
                                value_text=str(cell_val) if cell_val is not None else "",
                                status="error",
                                reason=msg,
                            )
                        )
                        row_notes.append(f"M{m:02d}:{msg}")
                        continue

                    if not budget_actual_allowed_for_month(budget_actual, m, current_month):
                        kind = "实际" if budget_actual == 1 else "预算"
                        rs = f"当前月份窗口不接收该月的{kind}数"
                        months_out.append(
                            BudgetInputImportMonthResult(
                                month=m,
                                value_text=str(num_v),
                                status="skipped",
                                reason=rs,
                            )
                        )
                        row_notes.append(f"M{m:02d}:{rs}")
                        continue

                    pid = month_to_period.get(m)
                    if pid is None:
                        rs = "系统未找到对应会计期间"
                        months_out.append(
                            BudgetInputImportMonthResult(
                                month=m,
                                value_text=str(num_v),
                                status="error",
                                reason=rs,
                            )
                        )
                        row_notes.append(f"M{m:02d}:{rs}")
                        continue

                    stored = _excel_to_stored(num_v, vt)
                    cur_ex = bconn.execute(
                        """
                        SELECT 1 FROM budget_data
                        WHERE data_acct_code = ? AND product_code = ? AND period_id = ?
                          AND version_id = ? AND budget_actual = ?
                        """,
                        (data_s, prod_s, pid, version_id, budget_actual),
                    )
                    existed = cur_ex.fetchone() is not None

                    bconn.execute(
                        """
                        INSERT INTO budget_data (
                          data_acct_code, product_code, period_id, budget_actual, version_id, value, need_calc, create_time, update_time
                        ) VALUES (?, ?, ?, ?, ?, ?, 1, ?, ?)
                        ON CONFLICT(data_acct_code, product_code, period_id, version_id, budget_actual)
                        DO UPDATE SET value = excluded.value, need_calc = 1, update_time = excluded.update_time
                        """,
                        (data_s, prod_s, pid, budget_actual, version_id, stored, row_now, row_now),
                    )

                    disp = f"{num_v:.4g}" if vt != "百分比" else f"{num_v:.4g}%"
                    months_out.append(
                        BudgetInputImportMonthResult(
                            month=m,
                            value_text=disp,
                            status="updated" if existed else "inserted",
                        )
                    )
                    saved_cells += 1
                    products_touched.add(prod_s)

                note = "；".join(dict.fromkeys(row_notes))
                result_rows.append(
                    BudgetInputImportResultRow(
                        sheet_name=sheet_name,
                        excel_row=r,
                        data_acct_code=data_s,
                        product_code=prod_s,
                        months=months_out,
                        note=note,
                    )
                )

        bconn.commit()
        wb.close()

        resp = BudgetInputImportResponse(
            budget_year=budget_year,
            version_id=version_id,
            current_month=current_month,
            rows=result_rows,
            saved_cells=saved_cells,
        )
        return resp, products_touched
    finally:
        cconn.close()
        bconn.close()
