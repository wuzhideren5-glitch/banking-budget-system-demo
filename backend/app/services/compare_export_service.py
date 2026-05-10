from __future__ import annotations

import asyncio
from datetime import datetime
from io import BytesIO
from pathlib import Path
import re
from typing import Any, Awaitable, Callable

import aiosqlite
from fastapi import HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.db_paths import common_db_path, compare_db_path
from app.schemas import BudgetSummaryExportPivotRequest


class CompareExportService:
    def __init__(
        self,
        *,
        data_dir: Path,
        compare_formula_export_jobs: dict[str, dict[str, Any]],
        compare_formula_export_jobs_lock: asyncio.Lock,
        sync_compare_budget_summary: Callable[..., Awaitable[Any]],
        extract_data_acct_code_from_name: Callable[[str], str | None],
        extract_product_code_from_summary_name: Callable[[str | None], str | None],
        month_idx_from_label: Callable[[str | None], int | None],
        normalize_formula: Callable[[str | None], str],
        prepare_formula_expression: Callable[[str | None], str],
        autosize_worksheet_columns: Callable[[Any], None],
        normalize_summary_value: Callable[[str, Any], str],
        budget_summary_field_meta: Callable[[], dict[str, tuple[str, str]]],
        write_template_pivot_data_area: Callable[..., None],
        build_export_versions_info_text: Callable[[list[tuple[Any, Any, Any]]], str],
    ) -> None:
        self._data_dir = data_dir
        self._jobs = compare_formula_export_jobs
        self._jobs_lock = compare_formula_export_jobs_lock
        self._sync_compare_budget_summary = sync_compare_budget_summary
        self._extract_data_acct_code_from_name = extract_data_acct_code_from_name
        self._extract_product_code_from_summary_name = extract_product_code_from_summary_name
        self._month_idx_from_label = month_idx_from_label
        self._normalize_formula = normalize_formula
        self._prepare_formula_expression = prepare_formula_expression
        self._autosize_worksheet_columns = autosize_worksheet_columns
        self._normalize_summary_value = normalize_summary_value
        self._budget_summary_field_meta = budget_summary_field_meta
        self._write_template_pivot_data_area = write_template_pivot_data_area
        self._build_export_versions_info_text = build_export_versions_info_text

    async def _compare_formula_export_job_update(
        self,
        job_id: str,
        *,
        status: str | None = None,
        processed_sheets: int | None = None,
        total_sheets: int | None = None,
        message: str | None = None,
        file_bytes: bytes | None = None,
        filename: str | None = None,
        error: str | None = None,
    ) -> None:
        async with self._jobs_lock:
            job = self._jobs.get(job_id)
            if not job:
                return
            if status is not None:
                job["status"] = status
            if processed_sheets is not None:
                job["processed_sheets"] = processed_sheets
            if total_sheets is not None:
                job["total_sheets"] = total_sheets
            if message is not None:
                job["message"] = message
            if file_bytes is not None:
                job["file_bytes"] = file_bytes
            if filename is not None:
                job["filename"] = filename
            if error is not None:
                job["error"] = error

    @staticmethod
    def _compare_use_actual_for_month(current_month: int, month_num: int) -> bool:
        x = max(1, min(13, int(current_month)))
        if x == 13:
            return True
        if x == 1:
            return False
        return month_num < x

    @staticmethod
    def _safe_excel_sheet_name(base_name: str, used_names: set[str]) -> str:
        cleaned = re.sub(r"[\\/*?:\[\]]", "_", (base_name or "").strip())
        cleaned = cleaned.replace("'", "’")
        if not cleaned:
            cleaned = "Sheet"
        cleaned = cleaned[:31]
        if cleaned not in used_names:
            used_names.add(cleaned)
            return cleaned
        suffix = 2
        while True:
            suffix_text = f"_{suffix}"
            candidate = f"{cleaned[: max(1, 31 - len(suffix_text))]}{suffix_text}"
            if candidate not in used_names:
                used_names.add(candidate)
                return candidate
            suffix += 1

    async def _build_compare_formula_workbook_bytes(
        self,
        *,
        job_id: str,
        body: BudgetSummaryExportPivotRequest,
    ) -> tuple[bytes, str]:
        del body
        await self._compare_formula_export_job_update(job_id, message="步骤 1/5：同步展示版本汇总数据...")
        await self._sync_compare_budget_summary(trigger_source="auto_before_compare_formula_export")
        await self._compare_formula_export_job_update(job_id, message="步骤 2/5：读取展示版本与基础科目...")

        common_path = common_db_path()
        selected_levels: list[dict[str, Any]] = []
        async with aiosqlite.connect(common_path) as cdb:
            await cdb.execute("PRAGMA foreign_keys = ON")
            cur = await cdb.execute(
                """
                SELECT e.edit_show_sign, d.data_file_name, d.year, e.version_id
                FROM edit_show_version e
                JOIN databases d ON d.id = e.data_file_id
                WHERE e.edit_show_sign BETWEEN 1 AND 5
                ORDER BY e.edit_show_sign DESC
                """
            )
            selected = await cur.fetchall()
            if not selected:
                raise HTTPException(status_code=400, detail="请先在系统设置中选择展示版本")
            for show_level, file_name, source_year, version_id in selected:
                bpath = self._data_dir / str(file_name)
                if not bpath.exists():
                    continue
                async with aiosqlite.connect(bpath) as bdb:
                    await bdb.execute("PRAGMA foreign_keys = ON")
                    cur_ver = await bdb.execute(
                        "SELECT version_name, current_month FROM version WHERE version_id = ?",
                        (int(version_id),),
                    )
                    ver_row = await cur_ver.fetchone()
                if not ver_row:
                    continue
                selected_levels.append(
                    {
                        "show_level": int(show_level),
                        "year": int(source_year),
                        "version_id": int(version_id),
                        "version_name": str(ver_row[0] or ""),
                        "current_month": int(ver_row[1] or 1),
                    }
                )
            cur_acct = await cdb.execute(
                """
                SELECT data_acct_code, data_acct_name, budget_formula, actual_formula, value_type
                FROM data_account
                """
            )
            data_acct_rows = await cur_acct.fetchall()
            cur_report = await cdb.execute(
                """
                SELECT report_acct_code, report_acct_name, parent_code, is_summary, is_minus
                FROM report_account
                """
            )
            report_rows = await cur_report.fetchall()
            cur_mapping = await cdb.execute(
                """
                SELECT report_acct_code, data_acct_code
                FROM report_data_mapping
                """
            )
            mapping_rows = await cur_mapping.fetchall()
            cur_dept = await cdb.execute(
                """
                SELECT dept_code, dept_name, parent_code
                FROM dept_account
                """
            )
            dept_rows = await cur_dept.fetchall()
            cur_dp = await cdb.execute(
                """
                SELECT dept_code, product_code
                FROM dept_product_mapping
                """
            )
            dept_product_rows = await cur_dp.fetchall()
            cur_prod = await cdb.execute(
                """
                SELECT product_code, product_name
                FROM product_type
                """
            )
            product_rows = await cur_prod.fetchall()

        if not selected_levels:
            raise HTTPException(status_code=400, detail="展示版本配置无效，请检查版本与文件")

        level_by_show = {int(item["show_level"]): item for item in selected_levels}

        await self._compare_formula_export_job_update(job_id, message="步骤 3/5：装载对比汇总明细...")
        compare_path = compare_db_path()
        value_map: dict[tuple[int, str, str, int, int], float] = {}
        if compare_path.exists():
            async with aiosqlite.connect(compare_path) as cdb:
                await cdb.execute("PRAGMA foreign_keys = ON")
                cur = await cdb.execute(
                    """
                    SELECT show_level, data_code_name, product_code_name, month, budget_actual, value
                    FROM compare_budget_summary
                    ORDER BY show_level
                    """
                )
                for row in await cur.fetchall():
                    show_level = int(row[0] or 0)
                    if show_level not in level_by_show:
                        continue
                    data_code = self._extract_data_acct_code_from_name(str(row[1] or ""))
                    product_code = self._extract_product_code_from_summary_name(row[2])
                    month_idx = self._month_idx_from_label(str(row[3] or ""))
                    if not data_code or not product_code or month_idx is None:
                        continue
                    budget_actual = int(row[4] or 0)
                    key = (show_level, product_code, data_code, month_idx + 1, budget_actual)
                    value_map[key] = value_map.get(key, 0.0) + float(row[5] or 0.0)

        data_meta_map: dict[str, dict[str, Any]] = {}
        for code, name, budget_formula, actual_formula, value_type in data_acct_rows:
            data_meta_map[str(code)] = {
                "name": str(name or ""),
                "budget_formula": self._normalize_formula(str(budget_formula or "")),
                "actual_formula": self._normalize_formula(str(actual_formula or "")),
                "value_type": str(value_type or ""),
            }

        report_nodes: dict[str, dict[str, Any]] = {}
        for report_code, report_name, parent_code, is_summary, is_minus in report_rows:
            code = str(report_code)
            report_nodes[code] = {
                "type": "report",
                "code": code,
                "name": str(report_name or ""),
                "parent_code": str(parent_code) if parent_code is not None else None,
                "is_summary": int(is_summary or 0),
                "is_minus": int(is_minus or 0),
                "children": [],
            }
        roots: list[dict[str, Any]] = []
        for node in report_nodes.values():
            parent_code = node.get("parent_code")
            if parent_code and parent_code in report_nodes:
                report_nodes[parent_code]["children"].append(node)
            else:
                roots.append(node)

        for idx, (report_code_raw, data_code_raw) in enumerate(mapping_rows):
            report_code = str(report_code_raw or "")
            data_code = str(data_code_raw or "")
            parent = report_nodes.get(report_code)
            if not parent or not data_code:
                continue
            meta = data_meta_map.get(data_code, {})
            display_name = str(meta.get("name", "")).strip()
            data_display = f"{data_code} {display_name}".strip() if display_name else data_code
            parent["children"].append(
                {
                    "type": "data",
                    "id": f"{report_code}:{data_code}:{idx}",
                    "code": data_code,
                    "name": data_display,
                    "value_type": str(meta.get("value_type", "")),
                    "budget_formula": str(meta.get("budget_formula", "")),
                    "actual_formula": str(meta.get("actual_formula", "")),
                }
            )

        def _sort_tree(nodes: list[dict[str, Any]]) -> None:
            nodes.sort(key=lambda n: str(n.get("code", "")))
            for n in nodes:
                children = n.get("children") or []
                if children:
                    _sort_tree(children)

        _sort_tree(roots)

        row_entries: list[dict[str, Any]] = []

        def _walk(node: dict[str, Any], report_level: int) -> None:
            if node.get("type") == "report":
                entry = {
                    "type": "report",
                    "node": node,
                    "report_level": report_level,
                }
                node["_entry"] = entry
                row_entries.append(entry)
                for child in node.get("children") or []:
                    _walk(child, report_level + 1)
            else:
                entry = {
                    "type": "data",
                    "node": node,
                    "report_level": report_level - 1,
                }
                node["_entry"] = entry
                row_entries.append(entry)

        for root in roots:
            _walk(root, 1)

        for idx, entry in enumerate(row_entries):
            entry["row_num"] = idx + 3

        data_row_by_code: dict[str, int] = {}
        for entry in row_entries:
            if entry["type"] != "data":
                continue
            code = str(entry["node"].get("code", ""))
            if code and code not in data_row_by_code:
                data_row_by_code[code] = int(entry["row_num"])

        dept_nodes: dict[str, dict[str, Any]] = {}
        for dept_code, dept_name, parent_code in dept_rows:
            code = str(dept_code)
            dept_nodes[code] = {
                "code": code,
                "name": str(dept_name or ""),
                "parent": str(parent_code) if parent_code is not None else None,
                "children": [],
                "products": [],
            }
        for node in dept_nodes.values():
            parent = str(node.get("parent") or "")
            if parent and parent in dept_nodes:
                dept_nodes[parent]["children"].append(node["code"])

        product_name_by_code = {str(code): str(name or "") for code, name in product_rows}
        for dept_code_raw, product_code_raw in dept_product_rows:
            dept_code = str(dept_code_raw or "")
            product_code = str(product_code_raw or "")
            if not dept_code or not product_code:
                continue
            if dept_code not in dept_nodes or product_code not in product_name_by_code:
                continue
            dept_nodes[dept_code]["products"].append(product_code)
        for node in dept_nodes.values():
            node["products"] = sorted(set(node["products"]))
            node["children"] = sorted(set(node["children"]))

        dept_depth_cache: dict[str, int] = {}

        def _dept_depth(code: str) -> int:
            if code in dept_depth_cache:
                return dept_depth_cache[code]
            node = dept_nodes.get(code)
            if not node:
                return 0
            parent = str(node.get("parent") or "")
            depth = 1 + (_dept_depth(parent) if parent else 0)
            dept_depth_cache[code] = depth
            return depth

        def _dept_path_codes(code: str) -> list[str]:
            path: list[str] = []
            seen: set[str] = set()
            cur = code
            while cur and cur not in seen and cur in dept_nodes:
                seen.add(cur)
                path.append(cur)
                cur = str(dept_nodes[cur].get("parent") or "")
            path.reverse()
            return path

        block_defs: list[dict[str, Any]] = []
        start_col = 12
        for info in selected_levels:
            month_headers = [
                (
                    f"实际M{m:02d}"
                    if self._compare_use_actual_for_month(info["current_month"], m)
                    else f"预算M{m:02d}"
                )
                for m in range(1, 13)
            ]
            block_defs.append(
                {
                    "show_level": int(info["show_level"]),
                    "year": int(info["year"]),
                    "version_id": int(info["version_id"]),
                    "version_name": str(info["version_name"] or ""),
                    "current_month": int(info["current_month"] or 1),
                    "start_col": start_col,
                    "month_headers": month_headers,
                }
            )
            start_col += 17
        total_last_col = start_col - 1

        used_sheet_names: set[str] = set()
        sheet_name_by_dept: dict[str, str] = {}
        for code, node in sorted(dept_nodes.items(), key=lambda x: x[0]):
            base = f"{code}_{str(node.get('name') or '')}".strip("_")
            sheet_name_by_dept[code] = self._safe_excel_sheet_name(base, used_names=used_sheet_names)
        sheet_name_by_product: dict[str, str] = {}
        for pcode, pname in sorted(product_name_by_code.items(), key=lambda x: x[0]):
            mapped = any(pcode in node.get("products", []) for node in dept_nodes.values())
            if not mapped:
                continue
            base = f"{pcode}_{pname}".strip("_")
            sheet_name_by_product[pcode] = self._safe_excel_sheet_name(base, used_names=used_sheet_names)
        summary_sheet_name = self._safe_excel_sheet_name("展示年度汇总", used_names=used_sheet_names)

        total_sheets = len(sheet_name_by_dept) + len(sheet_name_by_product) + 1
        await self._compare_formula_export_job_update(
            job_id,
            total_sheets=total_sheets,
            message=f"步骤 4/5：准备生成 {total_sheets} 张工作表...",
        )

        wb = Workbook()
        wb.remove(wb.active)
        header_fill = PatternFill(fill_type="solid", fgColor="FF000000")
        gray_fill = PatternFill(fill_type="solid", fgColor="EDEDED")
        formula_fill = PatternFill(fill_type="solid", fgColor="FFFFF2E6")
        top_level_fill = PatternFill(fill_type="solid", fgColor="FFC0C0C0")
        muted_font = Font(color="FFB3B3B3")

        version_info_parts: list[str] = []
        for block in block_defs:
            version_info_parts.append(
                f"L{block['show_level']} 导出年份：{block['year']}  版本号：{block['version_id']}  版本名称：{block['version_name'] or '未设置'}"
            )
        exported_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        version_info_text = "  |  ".join(version_info_parts) + f"  |  导出日期和时间：{exported_at}"

        def _fill_headers(ws) -> None:
            ws.cell(row=1, column=1, value=version_info_text)
            fixed_headers = [
                "报告科目1级",
                "报告科目2级",
                "报告科目3级",
                "报告科目4级",
                "报告科目5级",
                "数据科目",
                "数值类型",
                "部门科目1级",
                "部门科目2级",
                "部门科目3级",
                "产品科目",
            ]
            for col_idx, header in enumerate(fixed_headers, start=1):
                c = ws.cell(row=2, column=col_idx, value=header)
                c.fill = header_fill
                c.font = Font(color="FFFFFF", bold=True)
            for block in block_defs:
                col = int(block["start_col"])
                prefix = f"L{block['show_level']}-{block['year']}-V{block['version_id']}"
                for h in block["month_headers"]:
                    # 需求：列表头中预算/实际+月份信息前置，其次展示版本块信息。
                    c = ws.cell(row=2, column=col, value=f"{h}-{prefix}")
                    c.fill = header_fill
                    c.font = Font(color="FFFFFF", bold=True)
                    col += 1
                for q in ("季度Q1", "季度Q2", "季度Q3", "季度Q4", "年度合计"):
                    c = ws.cell(row=2, column=col, value=f"{q}-{prefix}")
                    c.fill = header_fill
                    c.font = Font(color="FFFFFF", bold=True)
                    col += 1
            ws.freeze_panes = "L3"
            ws.auto_filter.ref = f"A2:{get_column_letter(total_last_col)}2"

        def _weighted_sum_formula_for_children(col_idx: int, children: list[dict[str, Any]]) -> str:
            terms: list[tuple[int, str]] = []
            for ch in children:
                if not isinstance(ch, dict) or "_entry" not in ch:
                    continue
                row_ref = f"{get_column_letter(col_idx)}{int(ch['_entry']['row_num'])}"
                if ch.get("type") == "report":
                    if int(ch.get("is_summary", 0)) != 1:
                        continue
                    sign = -1 if int(ch.get("is_minus", 0)) == 1 else 1
                    terms.append((sign, row_ref))
                else:
                    terms.append((1, row_ref))
            if not terms:
                return "0"
            expr = ""
            for idx, (sign, ref) in enumerate(terms):
                if idx == 0:
                    expr += f"-{ref}" if sign < 0 else ref
                else:
                    expr += f"-{ref}" if sign < 0 else f"+{ref}"
            return f"={expr}"

        def _formula_to_excel(formula: str, col_idx: int) -> str:
            expr = self._prepare_formula_expression(formula)
            if not expr:
                return "0"

            def _replace_code(m: re.Match[str]) -> str:
                code = str(m.group(1))
                row_num = data_row_by_code.get(code)
                if row_num is None:
                    return "0"
                return f"{get_column_letter(col_idx)}{row_num}"

            expr = re.sub(r"\b([A-Z]\d{4})\b", _replace_code, expr)
            return f"={expr}"

        def _fill_common_rows(
            ws,
            *,
            dept_levels: list[str],
            product_text: str,
            value_getter: Any,
            data_formula_getter: Any,
            data_fill_by_formula: bool,
        ) -> None:
            max_report_level_in_rows = 1
            for entry in row_entries:
                if entry["type"] == "report":
                    max_report_level_in_rows = max(max_report_level_in_rows, int(entry["report_level"]))
            data_outline_level = min(max_report_level_in_rows, 8)
            for entry in row_entries:
                row_num = int(entry["row_num"])
                if entry["type"] == "report":
                    node = entry["node"]
                    level = max(1, min(int(entry["report_level"]), 5))
                    report_display = f"{node.get('code', '')} {node.get('name', '')}".strip()
                    if int(node.get("is_minus", 0)) == 1:
                        report_display = f"减： {report_display}"
                    ws.cell(row=row_num, column=level, value=report_display)
                    if level == 1:
                        for c in range(1, total_last_col + 1):
                            ws.cell(row=row_num, column=c).fill = top_level_fill
                        for c in range(1, 6):
                            ws.cell(row=row_num, column=c).font = Font(color="FF000000", bold=True, size=12)
                    children = node.get("children") or []
                    if int(node.get("is_summary", 0)) != 1:
                        for col_idx in range(12, total_last_col + 1):
                            c = ws.cell(row=row_num, column=col_idx, value="非汇总")
                            c.font = muted_font
                    else:
                        for block in block_defs:
                            start = int(block["start_col"])
                            for _m in range(12):
                                ws.cell(
                                    row=row_num,
                                    column=start + _m,
                                    value=_weighted_sum_formula_for_children(start + _m, children),
                                )
                            for q in range(4):
                                q_start = start + q * 3
                                q_end = q_start + 2
                                ws.cell(
                                    row=row_num,
                                    column=start + 12 + q,
                                    value=f"=SUM({get_column_letter(q_start)}{row_num}:{get_column_letter(q_end)}{row_num})",
                                )
                            ws.cell(
                                row=row_num,
                                column=start + 16,
                                value=f"=SUM({get_column_letter(start)}{row_num}:{get_column_letter(start + 11)}{row_num})",
                            )
                    for col_idx in range(12, total_last_col + 1):
                        ws.cell(row=row_num, column=col_idx).number_format = "#,##0.00;[Red]-#,##0.00"
                else:
                    node = entry["node"]
                    data_code = str(node.get("code", ""))
                    value_type = str(node.get("value_type", ""))
                    ws.cell(row=row_num, column=6, value=str(node.get("name", ""))).fill = gray_fill
                    value_type_display = value_type
                    if value_type == "金额":
                        value_type_display = "金额_亿元"
                    elif value_type == "户数":
                        value_type_display = "户数_万户"
                    ws.cell(row=row_num, column=7, value=value_type_display).fill = gray_fill
                    ws.cell(row=row_num, column=8, value=dept_levels[0] if len(dept_levels) > 0 else "")
                    ws.cell(row=row_num, column=9, value=dept_levels[1] if len(dept_levels) > 1 else "")
                    ws.cell(row=row_num, column=10, value=dept_levels[2] if len(dept_levels) > 2 else "")
                    ws.cell(row=row_num, column=11, value=product_text).fill = gray_fill

                    for block in block_defs:
                        show_level = int(block["show_level"])
                        current_month = int(block["current_month"])
                        start = int(block["start_col"])
                        for month_num in range(1, 13):
                            col_idx = start + month_num - 1
                            use_actual = self._compare_use_actual_for_month(current_month, month_num)
                            budget_actual = 1 if use_actual else 0
                            formula_text = str(data_formula_getter(data_code, budget_actual) or "")
                            cell = ws.cell(row=row_num, column=col_idx)
                            if data_fill_by_formula:
                                if formula_text:
                                    cell.value = _formula_to_excel(formula_text, col_idx)
                                    cell.fill = formula_fill
                                else:
                                    cell.value = value_getter(show_level, data_code, month_num, budget_actual)
                                    cell.fill = gray_fill
                            else:
                                cell.value = value_getter(show_level, data_code, month_num, budget_actual)
                                cell.fill = gray_fill
                        for q in range(4):
                            q_start = start + q * 3
                            q_end = q_start + 2
                            ws.cell(
                                row=row_num,
                                column=start + 12 + q,
                                value=f"=SUM({get_column_letter(q_start)}{row_num}:{get_column_letter(q_end)}{row_num})",
                            )
                        ws.cell(
                            row=row_num,
                            column=start + 16,
                            value=f"=SUM({get_column_letter(start)}{row_num}:{get_column_letter(start + 11)}{row_num})",
                        )
                    num_fmt = "0.00%;[Red]-0.00%" if value_type == "百分比" else "#,##0.00;[Red]-#,##0.00"
                    for col_idx in range(12, total_last_col + 1):
                        ws.cell(row=row_num, column=col_idx).number_format = num_fmt
                if entry["type"] == "report":
                    outline_level = max(int(entry["report_level"]) - 1, 0)
                else:
                    outline_level = data_outline_level
                ws.row_dimensions[row_num].outlineLevel = min(outline_level, 8)
                ws.row_dimensions[row_num].hidden = outline_level > 0
            self._autosize_worksheet_columns(ws)
            for col_idx in range(1, 6):
                ws.column_dimensions[get_column_letter(col_idx)].width = 8
            for col_idx in range(12, total_last_col + 1):
                letter = get_column_letter(col_idx)
                ws.column_dimensions[letter].width = min(max(ws.column_dimensions[letter].width or 8, 10), 18)
            for block in block_defs:
                start = int(block["start_col"])
                ws.column_dimensions.group(
                    get_column_letter(start),
                    get_column_letter(start + 15),
                    outline_level=1,
                    hidden=False,
                )
                for col_idx in range(start, start + 16):
                    ws.column_dimensions[get_column_letter(col_idx)].outlineLevel = 1
            ws.sheet_view.showOutlineSymbols = True
            if ws.sheet_properties.outlinePr is not None:
                ws.sheet_properties.outlinePr.summaryBelow = True
            for row_idx in range(3, ws.max_row + 1):
                for col_idx in range(12, total_last_col + 1):
                    ws.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal="right")

        processed_sheets = 0
        product_items: list[tuple[str, str]] = []
        for dept_code, node in dept_nodes.items():
            for pcode in node.get("products", []):
                if pcode in sheet_name_by_product:
                    product_items.append((dept_code, pcode))
        product_items.sort(key=lambda x: (x[0], x[1]))

        for dept_code, product_code in product_items:
            ws = wb.create_sheet(title=sheet_name_by_product[product_code])
            _fill_headers(ws)
            path_codes = _dept_path_codes(dept_code)
            dept_levels = [f"{c} {dept_nodes[c]['name']}".strip() for c in path_codes[:3]]
            while len(dept_levels) < 3:
                dept_levels.append("")
            product_text = f"{product_code} {product_name_by_code.get(product_code, '')}".strip()

            def _value_getter(show_level: int, data_code: str, month_num: int, budget_actual: int) -> float:
                return float(value_map.get((show_level, product_code, data_code, month_num, budget_actual), 0.0))

            def _formula_getter(data_code: str, budget_actual: int) -> str:
                meta = data_meta_map.get(data_code, {})
                return str(meta.get("budget_formula", "") if budget_actual == 0 else meta.get("actual_formula", ""))

            _fill_common_rows(
                ws,
                dept_levels=dept_levels,
                product_text=product_text,
                value_getter=_value_getter,
                data_formula_getter=_formula_getter,
                data_fill_by_formula=True,
            )
            processed_sheets += 1
            await self._compare_formula_export_job_update(
                job_id,
                processed_sheets=processed_sheets,
                message=f"已生成 {processed_sheets}/{total_sheets}：{sheet_name_by_product[product_code]}",
            )

        dept_codes_by_depth = sorted(dept_nodes.keys(), key=lambda c: _dept_depth(c), reverse=True)
        for dept_code in dept_codes_by_depth:
            ws = wb.create_sheet(title=sheet_name_by_dept[dept_code])
            _fill_headers(ws)
            path_codes = _dept_path_codes(dept_code)
            dept_levels = [f"{c} {dept_nodes[c]['name']}".strip() for c in path_codes[:3]]
            while len(dept_levels) < 3:
                dept_levels.append("")
            child_sheet_names: list[str] = []
            for child_code in dept_nodes[dept_code].get("children", []):
                child_sheet_names.append(sheet_name_by_dept[child_code])
            for pcode in dept_nodes[dept_code].get("products", []):
                sname = sheet_name_by_product.get(pcode)
                if sname:
                    child_sheet_names.append(sname)
            child_sheet_names = [n for n in child_sheet_names if n]
            _fill_common_rows(
                ws,
                dept_levels=dept_levels,
                product_text="",
                value_getter=lambda _a, _b, _c, _d: 0.0,
                data_formula_getter=lambda _a, _b: "",
                data_fill_by_formula=False,
            )
            for entry in row_entries:
                if entry["type"] != "data":
                    continue
                row_num = int(entry["row_num"])
                for block in block_defs:
                    start = int(block["start_col"])
                    for month_num in range(12):
                        col_idx = start + month_num
                        if child_sheet_names:
                            refs = ",".join(
                                f"'{sn}'!{get_column_letter(col_idx)}{row_num}" for sn in child_sheet_names
                            )
                            ws.cell(row=row_num, column=col_idx, value=f"=SUM({refs})")
                        else:
                            ws.cell(row=row_num, column=col_idx, value=0)
            processed_sheets += 1
            await self._compare_formula_export_job_update(
                job_id,
                processed_sheets=processed_sheets,
                message=f"已生成 {processed_sheets}/{total_sheets}：{sheet_name_by_dept[dept_code]}",
            )

        ws_summary = wb.create_sheet(title=summary_sheet_name, index=0)
        _fill_headers(ws_summary)
        top_level_depts = sorted(
            [
                code
                for code, node in dept_nodes.items()
                if not str(node.get("parent") or "") or str(node.get("parent")) not in dept_nodes
            ]
        )
        top_sheet_names = [sheet_name_by_dept[c] for c in top_level_depts if c in sheet_name_by_dept]
        _fill_common_rows(
            ws_summary,
            dept_levels=["", "", ""],
            product_text="",
            value_getter=lambda _a, _b, _c, _d: 0.0,
            data_formula_getter=lambda _a, _b: "",
            data_fill_by_formula=False,
        )
        for entry in row_entries:
            if entry["type"] != "data":
                continue
            row_num = int(entry["row_num"])
            for block in block_defs:
                start = int(block["start_col"])
                for month_num in range(12):
                    col_idx = start + month_num
                    if top_sheet_names:
                        refs = ",".join(f"'{sn}'!{get_column_letter(col_idx)}{row_num}" for sn in top_sheet_names)
                        ws_summary.cell(row=row_num, column=col_idx, value=f"=SUM({refs})")
                    else:
                        ws_summary.cell(row=row_num, column=col_idx, value=0)

        processed_sheets += 1
        await self._compare_formula_export_job_update(
            job_id,
            processed_sheets=processed_sheets,
            message=f"步骤 5/5：已生成 {processed_sheets}/{total_sheets}：{summary_sheet_name}",
        )

        top_level_depts = sorted(
            [
                code
                for code, node in dept_nodes.items()
                if not str(node.get("parent") or "") or str(node.get("parent")) not in dept_nodes
            ]
        )
        ordered_sheet_names: list[str] = [summary_sheet_name]
        visited_depts: set[str] = set()

        def _walk_sheet_order(dept_code: str) -> None:
            if dept_code in visited_depts or dept_code not in sheet_name_by_dept:
                return
            visited_depts.add(dept_code)
            ordered_sheet_names.append(sheet_name_by_dept[dept_code])
            child_items: list[tuple[str, str, str]] = []
            for child in dept_nodes[dept_code].get("children", []):
                child_items.append(("dept", str(child), str(child)))
            for pcode in dept_nodes[dept_code].get("products", []):
                if pcode in sheet_name_by_product:
                    child_items.append(("product", str(pcode), str(pcode)))
            child_items.sort(key=lambda item: item[2])
            for kind, code, _sort_key in child_items:
                if kind == "dept":
                    _walk_sheet_order(code)
                else:
                    ordered_sheet_names.append(sheet_name_by_product[code])

        for top_dept in top_level_depts:
            _walk_sheet_order(top_dept)
        for dept_code in sorted(dept_nodes.keys()):
            _walk_sheet_order(dept_code)

        existing_names = list(wb.sheetnames)
        for sname in existing_names:
            if sname not in ordered_sheet_names:
                ordered_sheet_names.append(sname)
        wb._sheets = [wb[sname] for sname in ordered_sheet_names if sname in wb.sheetnames]

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        filename = f"compare_summary_formula_workbook_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return out.getvalue(), filename

    async def run_compare_formula_export_job(
        self,
        job_id: str,
        body: BudgetSummaryExportPivotRequest,
    ) -> None:
        try:
            await self._compare_formula_export_job_update(job_id, status="running", message="开始生成工作簿")
            file_bytes, filename = await self._build_compare_formula_workbook_bytes(job_id=job_id, body=body)
            await self._compare_formula_export_job_update(
                job_id,
                status="done",
                file_bytes=file_bytes,
                filename=filename,
                message="导出完成",
            )
        except Exception as exc:
            detail = exc.detail if isinstance(exc, HTTPException) else str(exc)
            await self._compare_formula_export_job_update(job_id, status="error", error=str(detail), message="导出失败")

    async def export_compare_summary_full_pivot(self) -> StreamingResponse:
        await self._sync_compare_budget_summary(trigger_source="auto_before_compare_export")
        path = compare_db_path()
        if not path.exists():
            raise HTTPException(status_code=404, detail="compare.db 不存在")
        async with aiosqlite.connect(path) as db:
            cur = await db.execute(
                """
                SELECT source_year, source_version_id, source_version_name,
                       report_level1, report_level2, report_level3, report_level4, report_level5,
                       dept_level1, dept_level2, dept_level3, data_code_name, product_code_name,
                       year, month, quarter, budget_actual, value, value_type, sync_time
                FROM compare_budget_summary
                ORDER BY source_year, source_version_id,
                         report_level1, report_level2, report_level3, data_code_name, month, budget_actual
                """
            )
            rows = await cur.fetchall()
        template_path = Path(__file__).resolve().parents[3] / "download_template" / "pivot_export_temp.xlsx"
        if not template_path.exists():
            raise HTTPException(status_code=404, detail="缺少导出模板 pivot_export_temp.xlsx")
        wb = load_workbook(template_path)
        if "Pivot数据区" not in wb.sheetnames:
            raise HTTPException(status_code=400, detail="模板缺少工作表：Pivot数据区")
        if "数据透视表" not in wb.sheetnames:
            raise HTTPException(status_code=400, detail="模板缺少工作表：数据透视表")
        pivot_count = sum(len(getattr(ws, "_pivots", [])) for ws in wb.worksheets)
        if pivot_count == 0:
            raise HTTPException(
                status_code=400,
                detail="模板未检测到Excel原生数据透视表实例，请在“数据透视表”中先创建透视表并保存",
            )
        ws_data = wb["Pivot数据区"]

        records: list[dict[str, Any]] = []
        version_rows: list[tuple[Any, Any, Any]] = []
        for row in rows:
            source_year = row[0]
            source_version_id = row[1]
            source_version_name = row[2]
            records.append(
                {
                    "report_level1": row[3],
                    "report_level2": row[4],
                    "report_level3": row[5],
                    "report_level4": row[6],
                    "report_level5": row[7],
                    "dept_level1": row[8],
                    "dept_level2": row[9],
                    "dept_level3": row[10],
                    "data_code_name": row[11],
                    "product_code_name": row[12],
                    "year": row[13],
                    "month": row[14],
                    "quarter": row[15],
                    "budget_actual": "预算" if int(row[16] or 0) == 0 else "实际",
                    "version_display": self._normalize_summary_value(
                        "version_display",
                        {"version_id": source_version_id, "version_name": source_version_name},
                    ),
                    "value": row[17],
                    "value_type": row[18],
                    "update_time": row[19],
                    "calc_formula_raw": "",
                    "calc_formula_values": "",
                }
            )
            version_rows.append((source_year, source_version_id, source_version_name))

        field_meta = self._budget_summary_field_meta()
        chinese_value_by_header = {
            field_meta["report_level1"][0]: lambda r: r["report_level1"] or "",
            field_meta["report_level2"][0]: lambda r: r["report_level2"] or "",
            field_meta["report_level3"][0]: lambda r: r["report_level3"] or "",
            field_meta["report_level4"][0]: lambda r: r["report_level4"] or "",
            field_meta["report_level5"][0]: lambda r: r["report_level5"] or "",
            field_meta["dept_level1"][0]: lambda r: r["dept_level1"] or "",
            field_meta["dept_level2"][0]: lambda r: r["dept_level2"] or "",
            field_meta["dept_level3"][0]: lambda r: r["dept_level3"] or "",
            field_meta["data_code_name"][0]: lambda r: r["data_code_name"] or "",
            field_meta["product_code_name"][0]: lambda r: r["product_code_name"] or "",
            field_meta["month"][0]: lambda r: r["month"] or "",
            field_meta["quarter"][0]: lambda r: r["quarter"] or "",
            field_meta["budget_actual"][0]: lambda r: r["budget_actual"] or "",
            field_meta["version_display"][0]: lambda r: r["version_display"] or "",
            "年度": lambda r: r.get("year") or "",
            field_meta["value"][0]: lambda r: float(r["value"] or 0.0),
            field_meta["value_type"][0]: lambda r: r["value_type"] or "",
            field_meta["update_time"][0]: lambda r: r["update_time"] or "",
            "计算公式原文": lambda r: r["calc_formula_raw"] or "",
            "计算公式数值": lambda r: r["calc_formula_values"] or "",
        }
        self._write_template_pivot_data_area(
            wb=wb,
            ws_data=ws_data,
            records=records,
            chinese_value_by_header=chinese_value_by_header,
            versions_info_text=self._build_export_versions_info_text(version_rows),
        )

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return StreamingResponse(
            out,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="compare_summary_full_pivot.xlsx"'},
        )
