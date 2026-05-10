from __future__ import annotations

from datetime import datetime
import re
from typing import Any
import unicodedata

from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def budget_summary_field_meta() -> dict[str, tuple[str, str]]:
    return {
        "report_level1": ("报告科目1级", "report_level1"),
        "report_level2": ("报告科目2级", "report_level2"),
        "report_level3": ("报告科目3级", "report_level3"),
        "report_level4": ("报告科目4级", "report_level4"),
        "report_level5": ("报告科目5级", "report_level5"),
        "dept_level1": ("部门科目1级", "dept_level1"),
        "dept_level2": ("部门科目2级", "dept_level2"),
        "dept_level3": ("部门科目3级", "dept_level3"),
        "data_code_name": ("数据科目", "data_code_name"),
        "product_code_name": ("产品科目", "product_code_name"),
        "month": ("月份", "month"),
        "quarter": ("季度", "quarter"),
        "budget_actual": ("预算/实际", "budget_actual"),
        "version_display": ("版本号及名称", "version_display"),
        "value_type": ("数值类型", "value_type"),
        "value": ("预算数值", "value"),
        "update_time": ("更新时间", "update_time"),
    }


def normalize_summary_value(field_id: str, raw: Any) -> str:
    if field_id == "budget_actual":
        return "预算" if int(raw or 0) == 0 else "实际"
    if field_id == "version_display":
        version_id = ""
        version_name = ""
        if isinstance(raw, dict):
            version_id = str(raw.get("version_id") or "").strip()
            version_name = str(raw.get("version_name") or "").strip()
        if version_id and version_name:
            return f"版本号{version_id}：{version_name}"
        if version_id:
            return f"版本号{version_id}"
        if version_name:
            return version_name
        return "未设置"
    if raw is None:
        return "未设置"
    text = str(raw).strip()
    return text or "未设置"


def autosize_worksheet_columns(ws) -> None:
    def _display_width(text: str) -> int:
        width = 0
        for ch in text:
            width += 2 if unicodedata.east_asian_width(ch) in ("F", "W") else 1
        return width

    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is None:
                continue
            text = str(val)
            part_max = max((_display_width(seg) for seg in text.splitlines()), default=0)
            max_len = max(max_len, part_max)
        ws.column_dimensions[col_letter].width = max(8, min(80, max_len + 2))
        ws.column_dimensions[col_letter].bestFit = True


def mark_pivots_refresh_on_open(wb: Workbook) -> None:
    for ws in wb.worksheets:
        for pivot in getattr(ws, "_pivots", []):
            try:
                if getattr(pivot, "cache", None) is not None:
                    pivot.cache.refreshOnLoad = True
            except Exception:
                continue


def build_export_versions_info_text(version_rows: list[tuple[Any, Any, Any]]) -> str:
    parts: list[str] = []
    seen: set[tuple[str, str, str]] = set()
    deduped: list[tuple[str, str, str]] = []
    for export_year, version_id, version_name in version_rows:
        year_text = str(export_year or "").strip() or "未设置"
        vid_text = str(version_id or "").strip() or "未设置"
        vname_text = str(version_name or "").strip() or "未设置"
        key = (year_text, vid_text, vname_text)
        if key in seen:
            continue
        seen.add(key)
        deduped.append(key)

    def _sort_key(item: tuple[str, str, str]) -> tuple[int, int, str, int, str, str]:
        y, v, n = item
        y_is_num = 0 if y.isdigit() else 1
        v_is_num = 0 if v.isdigit() else 1
        y_num = int(y) if y.isdigit() else 0
        v_num = int(v) if v.isdigit() else 0
        return (y_is_num, y_num, y, v_is_num, v_num, n)

    for year_text, vid_text, vname_text in sorted(deduped, key=_sort_key):
        parts.append(f"导出年份：{year_text} __ 版本号：{vid_text} __ 版本名称：{vname_text} | ")
    if not parts:
        return "导出年份：未设置 __ 版本号：未设置 __ 版本名称：未设置 | "
    return "".join(parts)


def build_export_year_datetime_text(export_years: list[Any]) -> str:
    years: list[str] = []
    for raw in export_years:
        text = str(raw or "").strip()
        if not text:
            continue
        m = re.search(r"(\d{4})", text)
        years.append(m.group(1) if m else text)
    unique_years = sorted(set(years))
    year_text = "、".join(unique_years) if unique_years else "未设置"
    exported_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return f"导出年份：{year_text} 导出日期和时间：{exported_at}"


def write_template_pivot_data_area(
    wb: Workbook,
    ws_data,
    records: list[dict[str, Any]],
    chinese_value_by_header: dict[str, Any],
    versions_info_text: str,
) -> None:
    ws_data.cell(row=1, column=1, value=versions_info_text)
    header_map = {
        str(ws_data.cell(row=2, column=idx).value).strip(): idx
        for idx in range(1, 21)
        if ws_data.cell(row=2, column=idx).value is not None
        and str(ws_data.cell(row=2, column=idx).value).strip()
    }
    if not header_map:
        raise HTTPException(status_code=400, detail="模板 Pivot数据区 第2行缺少表头")
    if ws_data.max_row > 2:
        ws_data.delete_rows(3, ws_data.max_row - 2)
    excel_row = 3
    for rec in records:
        for header, col_idx in header_map.items():
            getter = chinese_value_by_header.get(header)
            if getter is None:
                continue
            ws_data.cell(row=excel_row, column=col_idx, value=getter(rec))
        excel_row += 1
    autosize_worksheet_columns(ws_data)
    mark_pivots_refresh_on_open(wb)
