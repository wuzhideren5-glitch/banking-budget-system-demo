from __future__ import annotations

from io import BytesIO
from typing import Any, Awaitable, Callable

import aiosqlite
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook

from app.db_paths import common_db_path
from app.schemas import (
    BudgetSubjectCatalogCreate,
    BudgetSubjectCatalogRow,
    BudgetSubjectCatalogUpdate,
)


def build_budget_subject_catalog_router(
    *,
    write_operation_log: Callable[..., Awaitable[None]],
) -> APIRouter:
    router = APIRouter()

    level_name_to_number = {"一级": 1, "二级": 2, "三级": 3, "四级": 4, "五级": 5}
    level_number_to_name = {value: key for key, value in level_name_to_number.items()}

    def _text(value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip()

    async def _ensure_bootstrapped() -> None:
        async with aiosqlite.connect(common_db_path()) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            await db.execute(
                """
                CREATE TABLE IF NOT EXISTS budget_subject_catalog (
                  id INTEGER PRIMARY KEY AUTOINCREMENT,
                  parent_id INTEGER REFERENCES budget_subject_catalog(id) ON DELETE RESTRICT,
                  level_number INTEGER NOT NULL CHECK (level_number BETWEEN 1 AND 5),
                  subject_name TEXT NOT NULL,
                  formula_text TEXT,
                  sort_order INTEGER NOT NULL DEFAULT 0
                )
                """
            )
            cur = await db.execute("SELECT COUNT(*) FROM budget_subject_catalog")
            if int((await cur.fetchone())[0] or 0) > 0:
                return

            cur = await db.execute(
                """
                SELECT level_label, budget_subject, formula_text, sort_order
                FROM expense_framework_subject
                ORDER BY sort_order, budget_subject
                """
            )
            framework_rows = await cur.fetchall()

            rows_to_insert: list[dict[str, Any]] = []
            if framework_rows:
                rows_to_insert = _build_bootstrap_rows(
                    [
                        {
                            "level_label": _text(row[0]),
                            "subject_name": _text(row[1]),
                            "formula_text": _text(row[2]) or None,
                            "sort_order": int(row[3] or 0),
                        }
                        for row in framework_rows
                    ]
                )

            if rows_to_insert:
                parent_id_by_virtual_id: dict[int, int] = {}
                for item in rows_to_insert:
                    cur = await db.execute(
                        """
                        INSERT INTO budget_subject_catalog(parent_id, level_number, subject_name, formula_text, sort_order)
                        VALUES (?, ?, ?, ?, ?)
                        """,
                        (
                            parent_id_by_virtual_id.get(int(item["parent_virtual_id"])) if item["parent_virtual_id"] is not None else None,
                            int(item["level_number"]),
                            _text(item["subject_name"]),
                            _text(item["formula_text"]) or None,
                            int(item["sort_order"]),
                        ),
                    )
                    parent_id_by_virtual_id[int(item["virtual_id"])] = int(cur.lastrowid)
                await db.commit()

    def _build_bootstrap_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
        stack_by_level: dict[int, int] = {}
        next_virtual_id = 1
        result: list[dict[str, Any]] = []
        for item in rows:
            level_number = level_name_to_number.get(_text(item.get("level_label")))
            subject_name = _text(item.get("subject_name"))
            if not level_number or not subject_name:
                continue
            parent_virtual_id = stack_by_level.get(level_number - 1) if level_number > 1 else None
            result.append(
                {
                    "virtual_id": next_virtual_id,
                    "parent_virtual_id": parent_virtual_id,
                    "level_number": level_number,
                    "subject_name": subject_name,
                    "formula_text": _text(item.get("formula_text")) or None,
                    "sort_order": int(item.get("sort_order") or len(result) + 1),
                }
            )
            stack_by_level[level_number] = next_virtual_id
            next_virtual_id += 1
            for key in list(stack_by_level.keys()):
                if key > level_number:
                    stack_by_level.pop(key, None)
        return result

    async def _list_rows() -> list[BudgetSubjectCatalogRow]:
        await _ensure_bootstrapped()
        async with aiosqlite.connect(common_db_path()) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            cur = await db.execute(
                """
                SELECT c.id, c.parent_id, c.level_number, c.subject_name, c.formula_text, c.sort_order,
                       EXISTS(SELECT 1 FROM budget_subject_catalog child WHERE child.parent_id = c.id) AS has_children
                FROM budget_subject_catalog c
                ORDER BY COALESCE(c.parent_id, 0), c.sort_order, c.id
                """
            )
            rows = await cur.fetchall()
        return [
            BudgetSubjectCatalogRow(
                id=int(row[0]),
                parent_id=int(row[1]) if row[1] is not None else None,
                level_number=int(row[2]),
                level_label=level_number_to_name.get(int(row[2]), f"{int(row[2])}级"),
                subject_name=_text(row[3]),
                formula_text=_text(row[4]) or None,
                sort_order=int(row[5] or 0),
                is_leaf=not bool(row[6]),
            )
            for row in rows
        ]

    @router.get("/api/budget-subject-catalog", response_model=list[BudgetSubjectCatalogRow])
    async def list_budget_subject_catalog():
        return await _list_rows()

    @router.post("/api/budget-subject-catalog", response_model=BudgetSubjectCatalogRow)
    async def create_budget_subject_catalog(body: BudgetSubjectCatalogCreate):
        await _ensure_bootstrapped()
        async with aiosqlite.connect(common_db_path()) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            parent_id = body.parent_id
            level_number = 1
            if parent_id is not None:
                cur = await db.execute(
                    "SELECT level_number FROM budget_subject_catalog WHERE id = ?",
                    (parent_id,),
                )
                parent_row = await cur.fetchone()
                if not parent_row:
                    raise HTTPException(status_code=404, detail="上级预算科目不存在")
                level_number = int(parent_row[0] or 0) + 1
            if level_number < 1 or level_number > 5:
                raise HTTPException(status_code=400, detail="预算科目最多支持 5 层")
            if parent_id is None:
                cur = await db.execute(
                    "SELECT COALESCE(MAX(sort_order), 0) FROM budget_subject_catalog WHERE parent_id IS NULL"
                )
            else:
                cur = await db.execute(
                    "SELECT COALESCE(MAX(sort_order), 0) FROM budget_subject_catalog WHERE parent_id = ?",
                    (parent_id,),
                )
            next_sort_order = int((await cur.fetchone())[0] or 0) + 1
            cur = await db.execute(
                """
                INSERT INTO budget_subject_catalog(parent_id, level_number, subject_name, formula_text, sort_order)
                VALUES (?, ?, ?, ?, ?)
                """,
                (
                    parent_id,
                    level_number,
                    body.subject_name.strip(),
                    _text(body.formula_text) or None,
                    next_sort_order,
                ),
            )
            new_id = int(cur.lastrowid)
            await db.commit()
        await write_operation_log(
            action_type="INSERT",
            action_desc=f"新增部门预算科目 {body.subject_name.strip()}",
            target_table="budget_subject_catalog",
            affected_rows=1,
            after_data=body.model_dump(),
        )
        rows = await _list_rows()
        for row in rows:
            if row.id == new_id:
                return row
        raise HTTPException(status_code=500, detail="新增预算科目后未找到记录")

    @router.patch("/api/budget-subject-catalog/{row_id}", response_model=BudgetSubjectCatalogRow)
    async def update_budget_subject_catalog(row_id: int, body: BudgetSubjectCatalogUpdate):
        await _ensure_bootstrapped()
        updates: list[str] = []
        values: list[Any] = []
        if body.subject_name is not None:
            updates.append("subject_name = ?")
            values.append(body.subject_name.strip())
        if body.formula_text is not None:
            updates.append("formula_text = ?")
            values.append(_text(body.formula_text) or None)
        if not updates:
            raise HTTPException(status_code=400, detail="没有可更新的内容")

        async with aiosqlite.connect(common_db_path()) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            cur = await db.execute(
                "SELECT id FROM budget_subject_catalog WHERE id = ?",
                (row_id,),
            )
            if not await cur.fetchone():
                raise HTTPException(status_code=404, detail="预算科目不存在")
            values.append(row_id)
            await db.execute(
                f"UPDATE budget_subject_catalog SET {', '.join(updates)} WHERE id = ?",
                values,
            )
            await db.commit()
        await write_operation_log(
            action_type="UPDATE",
            action_desc=f"更新部门预算科目 {row_id}",
            target_table="budget_subject_catalog",
            affected_rows=1,
            after_data=body.model_dump(),
        )
        rows = await _list_rows()
        for row in rows:
            if row.id == row_id:
                return row
        raise HTTPException(status_code=500, detail="更新预算科目后未找到记录")

    @router.delete("/api/budget-subject-catalog/{row_id}")
    async def delete_budget_subject_catalog(row_id: int):
        await _ensure_bootstrapped()
        async with aiosqlite.connect(common_db_path()) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            cur = await db.execute(
                "SELECT subject_name FROM budget_subject_catalog WHERE id = ?",
                (row_id,),
            )
            row = await cur.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="预算科目不存在")
            cur = await db.execute(
                "SELECT COUNT(*) FROM budget_subject_catalog WHERE parent_id = ?",
                (row_id,),
            )
            if int((await cur.fetchone())[0] or 0) > 0:
                raise HTTPException(status_code=409, detail="当前预算科目下仍有下级，无法删除")
            await db.execute("DELETE FROM budget_subject_catalog WHERE id = ?", (row_id,))
            await db.commit()
        await write_operation_log(
            action_type="DELETE",
            action_desc=f"删除部门预算科目 {row_id}",
            target_table="budget_subject_catalog",
            affected_rows=1,
            before_data={"id": row_id, "subject_name": _text(row[0])},
            after_data=None,
        )
        return {"ok": True}

    @router.get("/api/budget-subject-catalog/export")
    async def export_budget_subject_catalog():
        rows = await _list_rows()
        children_by_parent: dict[int | None, list[BudgetSubjectCatalogRow]] = {}
        for row in rows:
            children_by_parent.setdefault(row.parent_id, []).append(row)
        for value in children_by_parent.values():
            value.sort(key=lambda item: (item.sort_order, item.id))

        wb = Workbook()
        ws = wb.active
        ws.title = "部门预算科目"
        headers = ["层级", "预算科目", "公式"]
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        export_rows: list[list[str | None]] = []

        def walk(parent_id: int | None) -> None:
            for row in children_by_parent.get(parent_id, []):
                export_rows.append([row.level_label, row.subject_name, row.formula_text])
                walk(row.id)

        walk(None)
        for row_idx, row in enumerate(export_rows, start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 30

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="budget_subject_catalog.xlsx"'},
        )

    return router
