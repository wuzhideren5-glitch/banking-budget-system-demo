"""预算预测驱动参数模块：驱动因素分类/指标管理、Excel 模板下载与导入计算。"""
from __future__ import annotations

import asyncio
import re
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any, Awaitable, Callable

import aiosqlite
from fastapi import APIRouter, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.db_paths import common_db_path
from app.formula_refs import extract_formula_codes
from app.schemas import (
    DriverAccountMappingUpsert,
    DriverCategoryTree,
    DriverDataAccountOption,
    DriverImportMonthlyItem,
    DriverImportPreviewResponse,
    DriverImportPreviewRow,
    DriverImportRequest,
    DriverImportResponse,
    DriverIndicatorTree,
    DriverMappedDataAccount,
    DriverProductRow,
)


HEADER_FILL = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFFFF")
BODY_FONT = Font(name="微软雅黑", size=9)


def _iso_now() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _parse_percentage_for_storage(raw: Any, is_pct_type: bool) -> float:
    """将 Excel 单元格值转为存储值；百分比类型除以 100 后存储。"""
    if raw is None:
        return 0.0
    if isinstance(raw, str):
        s = raw.strip().replace(",", "").replace("，", "").replace("%", "")
        if not s:
            return 0.0
        try:
            return float(s) / 100.0 if is_pct_type else float(s)
        except ValueError:
            return 0.0
    if isinstance(raw, (int, float)):
        v = float(raw)
        return v / 100.0 if is_pct_type else v
    return 0.0


def _match_indicator_from_name(sheet_name: str, name_cell: str, all_indicators: list[dict]) -> dict | None:
    """按指标名匹配 driver_indicator。"""
    n = name_cell.strip()
    for ind in all_indicators:
        if ind["indicator_name"] == n or ind["indicator_code"] == n:
            return ind
    return None


def _match_product_from_name(name_cell: str, all_products: list[dict]) -> str | None:
    """按产品名或代码匹配 product_code，'全行'返回空字符串（不限定产品）。"""
    n = name_cell.strip()
    if not n or n in ("全行", "公司级", "全部"):
        return ""  # 公司级/全行
    for p in all_products:
        if p["product_code"] == n or p["product_name"] == n:
            return p["product_code"]
    # 模糊匹配：产品名包含输入词
    for p in all_products:
        if n in (p["product_name"] or ""):
            return p["product_code"]
    return None


def _header_index(headers: list[str], aliases: set[str]) -> int | None:
    normalized = {h.strip(): idx for idx, h in enumerate(headers)}
    for alias in aliases:
        if alias in normalized:
            return normalized[alias]
    return None


def _month_to_period_id(period_month_map: dict[int, int]) -> dict[int, int]:
    """Convert shared pid->month map into month->pid for driver imports."""
    return {month: period_id for period_id, month in period_month_map.items()}


async def _load_current_month(budget_path: Path, version_id: int) -> int:
    async with aiosqlite.connect(budget_path) as db:
        cur = await db.execute(
            "SELECT current_month FROM version WHERE version_id = ?",
            (version_id,),
        )
        row = await cur.fetchone()
    if not row:
        return 1
    return max(1, min(13, int(row[0] or 1)))


def _mapping_key(indicator_code: str, product_code: str | None) -> tuple[str, str]:
    return (indicator_code.strip().upper(), (product_code or "").strip().upper())


def _product_hierarchy(product_code: str) -> list[str]:
    codes = [product_code.strip().upper()]
    code = codes[0]
    while len(code) > 3:
        code = code[:-2]
        if len(code) >= 3:
            codes.append(code)
    return codes


def _data_account_applies_to_product(product_codes: Any, product_code: str) -> bool:
    if product_codes is None:
        return True
    raw = str(product_codes).strip()
    if raw.upper() == "ALL":
        return True
    if not raw:
        return False
    allowed = {p.strip().upper() for p in raw.split(",") if p.strip()}
    return bool(allowed & set(_product_hierarchy(product_code)))


def _build_report_path(
    report_code: str, report_name_map: dict[str, str], report_parent_map: dict[str, str | None]
) -> list[str]:
    path: list[str] = []
    seen: set[str] = set()
    cur = report_code
    while cur and cur not in seen:
        seen.add(cur)
        name = report_name_map.get(cur, "")
        path.append(f"{cur} {name}".strip())
        cur = report_parent_map.get(cur) or ""
    path.reverse()
    return path


async def _load_report_paths_for_data_accounts(
    db: aiosqlite.Connection,
    data_acct_codes: list[str] | set[str],
) -> dict[str, list[dict[str, Any]]]:
    codes = sorted({str(c).strip().upper() for c in data_acct_codes if str(c).strip()})
    if not codes:
        return {}

    placeholders = ",".join(["?"] * len(codes))
    cur = await db.execute(
        f"""
        SELECT report_acct_code, data_acct_code
        FROM report_data_mapping
        WHERE data_acct_code IN ({placeholders})
        ORDER BY report_acct_code, data_acct_code
        """,
        codes,
    )
    mapping_rows = await cur.fetchall()
    report_codes = sorted({str(r[0]) for r in mapping_rows})
    if not report_codes:
        return {}

    report_name_map: dict[str, str] = {}
    report_parent_map: dict[str, str | None] = {}
    unresolved = set(report_codes)
    while unresolved:
        report_placeholders = ",".join(["?"] * len(unresolved))
        cur = await db.execute(
            f"""
            SELECT report_acct_code, report_acct_name, parent_code
            FROM report_account
            WHERE report_acct_code IN ({report_placeholders})
            """,
            tuple(unresolved),
        )
        fetched = await cur.fetchall()
        if not fetched:
            break
        unresolved = set()
        for rr in fetched:
            code = str(rr[0])
            report_name_map[code] = str(rr[1])
            parent = str(rr[2]) if rr[2] is not None else None
            report_parent_map[code] = parent
            if parent and parent not in report_name_map:
                unresolved.add(parent)

    result: dict[str, list[dict[str, Any]]] = {}
    for report_code, data_code in mapping_rows:
        code = str(data_code)
        rc = str(report_code)
        result.setdefault(code, []).append({
            "report_code": rc,
            "report_path": _build_report_path(rc, report_name_map, report_parent_map),
        })
    return result


async def _load_actual_values_by_key(
    budget_path: Path,
    version_id: int,
    period_month_map: dict[int, int],
    keys: set[tuple[str, str]],
) -> dict[tuple[str, str], list[float]]:
    result = {(data_code, product_code): [0.0] * 12 for data_code, product_code in keys}
    if not keys:
        return result
    data_codes = sorted({data_code for data_code, _ in keys})
    product_codes = sorted({product_code for _, product_code in keys})
    period_ids = sorted(period_month_map.keys())
    if not data_codes or not product_codes or not period_ids:
        return result

    async with aiosqlite.connect(budget_path) as bdb:
        await bdb.execute("PRAGMA foreign_keys = ON")
        cur = await bdb.execute(
            f"""
            SELECT data_acct_code, product_code, period_id, value
            FROM budget_data
            WHERE version_id = ?
              AND budget_actual = 1
              AND data_acct_code IN ({",".join(["?"] * len(data_codes))})
              AND product_code IN ({",".join(["?"] * len(product_codes))})
              AND period_id IN ({",".join(["?"] * len(period_ids))})
            """,
            (version_id, *data_codes, *product_codes, *period_ids),
        )
        for data_code, product_code, period_id, value in await cur.fetchall():
            key = (str(data_code).strip().upper(), str(product_code).strip().upper())
            month = int(period_month_map.get(int(period_id), 0) or 0)
            if key in result and 1 <= month <= 12:
                result[key][month - 1] = float(value or 0.0)
    return result


def _resolve_data_acct_codes(
    indicator: dict,
    product_code: str | None,
    account_mappings: dict[tuple[str, str], list[str]],
    requested_data_acct_code: str | None = None,
) -> list[str]:
    codes = account_mappings.get(_mapping_key(str(indicator["indicator_code"]), product_code))
    requested = (requested_data_acct_code or "").strip().upper()
    if requested:
        if codes and requested not in codes:
            return []
        fallback = str(indicator.get("data_acct_code") or "").strip().upper()
        if not codes and fallback and requested != fallback:
            return []
        return [requested]
    if codes:
        return codes
    fallback = str(indicator.get("data_acct_code") or "").strip().upper()
    return [fallback] if fallback else []


def _non_empty_cell_count(ws: Any, row_index: int, month_cols: dict[int, int]) -> int:
    count = 0
    for ci in month_cols:
        raw_val = ws.cell(row_index, ci + 1).value
        if raw_val is None:
            continue
        if isinstance(raw_val, str) and not raw_val.strip():
            continue
        count += 1
    return count


def _collect_driver_import_plan(
    wb: Any,
    *,
    all_indicators: list[dict[str, Any]],
    all_products: list[dict[str, Any]],
    account_mappings: dict[tuple[str, str], list[str]],
    current_month: int = 1,
) -> tuple[list[dict[str, Any]], list[DriverImportPreviewRow], list[str], list[str], int]:
    ind_by_name: dict[str, dict[str, Any]] = {}
    ind_by_code: dict[str, dict[str, Any]] = {}
    for ind in all_indicators:
        ind_by_name[str(ind["indicator_name"])] = ind
        ind_by_code[str(ind["indicator_code"])] = ind

    plan_rows: list[dict[str, Any]] = []
    preview_rows: list[DriverImportPreviewRow] = []
    errors: list[str] = []
    warnings: list[str] = []
    total_rows = 0

    for ws in wb.worksheets:
        if ws.max_row < 2:
            continue

        header_row = [str(ws.cell(1, c).value or "").strip() for c in range(1, min(ws.max_column + 1, 20))]
        indicator_col = _header_index(header_row, {"指标名称", "指标编码", "驱动指标", "indicator_code", "indicator"})
        product_col = _header_index(header_row, {"产品", "产品名称", "产品编码", "product_code", "product"})
        data_code_col = _header_index(header_row, {"数据科目编码", "数据科目", "data_acct_code", "data_account"})
        if indicator_col is None:
            indicator_col = 0
        if product_col is None:
            product_col = 1

        month_cols: dict[int, int] = {}
        for ci, h in enumerate(header_row):
            m = re.match(r"M(\d{2})", h)
            if m:
                month_cols[ci] = int(m.group(1))

        if not month_cols:
            warnings.append(f"工作表 {ws.title} 未识别到 M01-M12 月份列，已跳过")
            continue

        for ridx in range(2, ws.max_row + 1):
            name_cell = str(ws.cell(ridx, indicator_col + 1).value or "").strip()
            product_cell = str(ws.cell(ridx, product_col + 1).value or "").strip()
            requested_data_code = (
                str(ws.cell(ridx, data_code_col + 1).value or "").strip().upper()
                if data_code_col is not None
                else ""
            )

            if not name_cell and not requested_data_code:
                continue

            total_rows += 1
            recognized_value_cells = _non_empty_cell_count(ws, ridx, month_cols)
            indicator = ind_by_name.get(name_cell) or ind_by_code.get(name_cell)
            product_code = ""
            resolved_data_codes: list[str] = []
            status = "ok"
            message = ""

            if not indicator:
                status = "error"
                message = f"行 {ridx}: 未识别驱动指标 '{name_cell or requested_data_code}'"
                errors.append(message)
            else:
                product_code = _match_product_from_name(product_cell, all_products)
                if product_code is None and indicator["has_product_detail"]:
                    status = "error"
                    message = f"行 {ridx}: 未找到产品 '{product_cell}'"
                    errors.append(message)
                else:
                    if not indicator["has_product_detail"]:
                        product_code = ""
                    resolved_data_codes = _resolve_data_acct_codes(
                        indicator, product_code, account_mappings, requested_data_code
                    )
                    if not resolved_data_codes:
                        status = "error"
                        if requested_data_code:
                            message = (
                                f"行 {ridx}: 数据科目 {requested_data_code} 未绑定到指标 "
                                f"{indicator['indicator_name']} 产品 {product_code or '全行'}"
                            )
                        else:
                            message = (
                                f"行 {ridx}: 指标 {indicator['indicator_name']} 产品 "
                                f"{product_code or '全行'} 未绑定数据科目"
                            )
                        errors.append(message)
                    elif recognized_value_cells == 0:
                        status = "warning"
                        message = f"行 {ridx}: 未检测到可导入的月份数值"
                        warnings.append(message)

            preview_rows.append(
                DriverImportPreviewRow(
                    sheet_name=ws.title,
                    excel_row=ridx,
                    indicator_text=name_cell,
                    product_text=product_cell,
                    requested_data_acct_code=requested_data_code or None,
                    matched_indicator_code=(indicator or {}).get("indicator_code"),
                    matched_indicator_name=(indicator or {}).get("indicator_name"),
                    matched_product_code=product_code or None,
                    resolved_data_acct_codes=resolved_data_codes,
                    recognized_value_cells=recognized_value_cells,
                    status=status,
                    message=message or None,
                )
            )

            if status != "ok":
                continue

            is_pct = indicator["value_type"] == "百分比"
            month_values: list[tuple[int, float]] = []
            for ci, mi in month_cols.items():
                raw_val = ws.cell(ridx, ci + 1).value
                if raw_val is None or (isinstance(raw_val, str) and not raw_val.strip()):
                    continue
                if mi < current_month:
                    warnings.append(
                        f"行 {ridx}: M{mi:02d} 已进入实际数区间，不允许写入预测值，已跳过"
                    )
                    continue
                month_values.append((mi, _parse_percentage_for_storage(raw_val, is_pct)))
            if not month_values:
                continue

            plan_rows.append(
                {
                    "sheet_name": ws.title,
                    "excel_row": ridx,
                    "indicator": indicator,
                    "product_code": product_code,
                    "data_acct_codes": resolved_data_codes,
                    "month_values": month_values,
                }
            )

    return plan_rows, preview_rows, errors, warnings, total_rows


def _generate_excel_template(categories: list[dict], indicators: list[dict], products: list[dict]) -> BytesIO:
    """动态生成驱动参数 Excel 模板。"""
    wb = Workbook()
    wb.remove(wb.active)

    # 构建 indicator_lookup: indicator_code → indicator dict
    ind_by_code = {i["indicator_code"]: i for i in indicators}
    # 构建 products for indicator: 每个有产品拆分的指标对应哪些产品
    prod_by_indicator: dict[str, list[dict]] = {}
    for i in indicators:
        if i["has_product_detail"]:
            prod_by_indicator[i["indicator_code"]] = [
                p for p in products if p.get("indicator_code") == i["indicator_code"]
            ]
            if not prod_by_indicator[i["indicator_code"]]:
                prod_by_indicator[i["indicator_code"]] = products  # fallback all

    for cat in sorted(categories, key=lambda x: x["sort_order"]):
        ws = wb.create_sheet(title=cat["category_name"])
        cat_indicators = [i for i in indicators if i["category_code"] == cat["category_code"]]

        # Header
        headers = ["指标名称", "产品", "报告科目", "数据科目编码", "数据科目名称"] + [f"M{m:02d}" for m in range(1, 13)]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        row = 2
        for ind in sorted(cat_indicators, key=lambda x: x["sort_order"]):
            if ind["has_product_detail"]:
                prods = prod_by_indicator.get(ind["indicator_code"], products)
                for p in sorted(prods, key=lambda x: x.get("product_code", "")):
                    accounts = p.get("data_accounts") or [None]
                    for account in accounts:
                        ws.cell(row=row, column=1, value=ind["indicator_name"]).font = BODY_FONT
                        ws.cell(row=row, column=2, value=p.get("product_name", p.get("product_code", ""))).font = BODY_FONT
                        if account:
                            ws.cell(row=row, column=3, value=" > ".join(account.get("report_path") or [])).font = BODY_FONT
                            ws.cell(row=row, column=4, value=account.get("data_acct_code")).font = BODY_FONT
                            ws.cell(row=row, column=5, value=account.get("data_acct_name")).font = BODY_FONT
                        row += 1
            else:
                ws.cell(row=row, column=1, value=ind["indicator_name"]).font = BODY_FONT
                ws.cell(row=row, column=2, value="全行").font = BODY_FONT
                row += 1

        # Auto-width
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 42
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions["E"].width = 28
        for m in range(1, 13):
            ws.column_dimensions[get_column_letter(m + 5)].width = 12

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def build_budget_driver_router(
    *,
    editable_context_provider: Callable[[], Awaitable[tuple[Path, int, int]]],
    recalculate_product_formula_rows: Callable[..., Awaitable[int]],
    rebuild_budget_summary: Callable[[int, Path], Awaitable[int]],
    get_year_period_months: Callable[[int], Awaitable[dict[int, int]]],
    write_operation_log: Callable[..., Awaitable[None]],
) -> APIRouter:
    router = APIRouter()

    @router.get("/api/driver/data-account-options", response_model=list[DriverDataAccountOption])
    async def list_driver_data_account_options(q: str = Query("", max_length=80)):
        keyword = q.strip().lower()
        common_path = common_db_path()
        async with aiosqlite.connect(common_path) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            cur = await db.execute(
                """
                SELECT data_acct_code, data_acct_name, value_type, product_codes
                FROM data_account
                WHERE COALESCE(TRIM(budget_formula), '') = ''
                ORDER BY data_acct_code
                """
            )
            rows = await cur.fetchall()
            if keyword:
                rows = [
                    r for r in rows
                    if keyword in str(r[0]).lower() or keyword in str(r[1]).lower()
                ]
            data_codes = {str(r[0]).strip().upper() for r in rows}
            report_paths = await _load_report_paths_for_data_accounts(db, data_codes)

        result: list[DriverDataAccountOption] = []
        for r in rows[:300]:
            code = str(r[0]).strip().upper()
            first_path = (report_paths.get(code) or [{}])[0]
            result.append(DriverDataAccountOption(
                data_acct_code=code,
                data_acct_name=str(r[1]),
                value_type=str(r[2]),
                product_codes=str(r[3]) if r[3] is not None else None,
                report_code=first_path.get("report_code"),
                report_path=first_path.get("report_path", ["未映射数据科目"]),
            ))
        return result

    @router.post("/api/driver/account-mappings", response_model=DriverMappedDataAccount)
    async def upsert_driver_account_mapping(body: DriverAccountMappingUpsert):
        data_acct_code = body.data_acct_code.strip().upper()
        indicator_code = (body.indicator_code or data_acct_code).strip().upper()
        category_code = (body.category_code or "").strip().upper()
        product_code = body.product_code.strip().upper()
        common_path = common_db_path()
        async with aiosqlite.connect(common_path) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            if category_code:
                cur = await db.execute(
                    "SELECT 1 FROM driver_category WHERE category_code = ?",
                    (category_code,),
                )
                if not await cur.fetchone():
                    raise HTTPException(status_code=404, detail=f"驱动因素分类不存在：{category_code}")
            cur = await db.execute(
                "SELECT 1 FROM product_type WHERE product_code = ?",
                (product_code,),
            )
            if not await cur.fetchone():
                raise HTTPException(status_code=404, detail=f"产品科目不存在：{product_code}")
            cur = await db.execute(
                "SELECT data_acct_name, value_type, product_codes FROM data_account WHERE data_acct_code = ?",
                (data_acct_code,),
            )
            data_row = await cur.fetchone()
            if not data_row:
                raise HTTPException(status_code=404, detail=f"数据科目不存在：{data_acct_code}")
            cur = await db.execute(
                "SELECT COALESCE(TRIM(budget_formula), '') FROM data_account WHERE data_acct_code = ?",
                (data_acct_code,),
            )
            formula_row = await cur.fetchone()
            if formula_row and str(formula_row[0]).strip():
                raise HTTPException(status_code=400, detail=f"数据科目 {data_acct_code} 为公式计算科目，不允许绑定为预测驱动底层输入")
            cur = await db.execute(
                "SELECT 1 FROM driver_indicator WHERE indicator_code = ?",
                (indicator_code,),
            )
            if not await cur.fetchone():
                if not category_code:
                    raise HTTPException(status_code=404, detail=f"驱动指标不存在：{indicator_code}")
                await db.execute(
                    """
                    INSERT INTO driver_indicator(
                      indicator_code, category_code, indicator_name, value_type,
                      data_acct_code, has_product_detail, has_monthly_detail, sort_order
                    ) VALUES (?, ?, ?, ?, ?, 1, 1, ?)
                    """,
                    (
                        indicator_code,
                        category_code,
                        str(data_row[0]),
                        str(data_row[1]),
                        data_acct_code,
                        int(body.sort_order or 0),
                    ),
                )
            elif category_code:
                await db.execute(
                    """
                    UPDATE driver_indicator
                    SET category_code = ?, indicator_name = ?, value_type = ?, data_acct_code = ?,
                        has_product_detail = 1, has_monthly_detail = 1
                    WHERE indicator_code = ?
                    """,
                    (category_code, str(data_row[0]), str(data_row[1]), data_acct_code, indicator_code),
                )
            if not _data_account_applies_to_product(data_row[2], product_code):
                raise HTTPException(
                    status_code=400,
                    detail=f"数据科目 {data_acct_code} 不适用于产品 {product_code}，请先在数据科目维护中调整适用产品范围",
                )
            await db.execute(
                """
                INSERT OR IGNORE INTO driver_product(indicator_code, product_code, sort_order)
                VALUES (?, ?, 0)
                """,
                (indicator_code, product_code),
            )
            await db.execute(
                """
                INSERT INTO driver_account_mapping(indicator_code, product_code, data_acct_code, sort_order)
                VALUES (?, ?, ?, ?)
                ON CONFLICT(indicator_code, product_code, data_acct_code)
                DO UPDATE SET sort_order = excluded.sort_order
                """,
                (indicator_code, product_code, data_acct_code, int(body.sort_order or 0)),
            )
            await db.commit()
            report_paths = await _load_report_paths_for_data_accounts(db, {data_acct_code})

        first_path = (report_paths.get(data_acct_code) or [{}])[0]
        await write_operation_log(
            action_type="UPSERT",
            action_desc=f"维护驱动数据科目绑定 {indicator_code}/{product_code}/{data_acct_code}",
            target_table="driver_account_mapping",
            affected_rows=1,
            after_data=body.model_dump(),
        )
        return DriverMappedDataAccount(
            data_acct_code=data_acct_code,
            data_acct_name=str(data_row[0]),
            value_type=str(data_row[1]),
            report_code=first_path.get("report_code"),
            report_path=first_path.get("report_path", ["未映射数据科目"]),
            sort_order=int(body.sort_order or 0),
        )

    @router.delete("/api/driver/account-mappings")
    async def delete_driver_account_mapping(
        indicator_code: str = Query(..., min_length=1),
        product_code: str = Query(..., min_length=1),
        data_acct_code: str = Query(..., min_length=1),
    ):
        ic = indicator_code.strip().upper()
        pc = product_code.strip().upper()
        dc = data_acct_code.strip().upper()
        common_path = common_db_path()
        async with aiosqlite.connect(common_path) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            cur = await db.execute(
                """
                SELECT COUNT(*)
                FROM driver_account_mapping
                WHERE indicator_code = ? AND product_code = ? AND data_acct_code = ?
                """,
                (ic, pc, dc),
            )
            deleted_count = int((await cur.fetchone())[0] or 0)
            await db.execute(
                """
                DELETE FROM driver_account_mapping
                WHERE indicator_code = ? AND product_code = ? AND data_acct_code = ?
                """,
                (ic, pc, dc),
            )
            await db.commit()
        await write_operation_log(
            action_type="DELETE",
            action_desc=f"删除驱动数据科目绑定 {ic}/{pc}/{dc}",
            target_table="driver_account_mapping",
            affected_rows=deleted_count,
            after_data={"indicator_code": ic, "product_code": pc, "data_acct_code": dc},
        )
        return {"deleted": deleted_count}

    # ── GET 驱动分类树 ──
    @router.get("/api/driver/categories", response_model=list[DriverCategoryTree])
    async def list_driver_categories():
        budget_path, budget_year, version_id = await editable_context_provider()
        current_month = await _load_current_month(budget_path, version_id)
        period_month_map = await get_year_period_months(budget_year)
        common_path = common_db_path()
        async with aiosqlite.connect(common_path) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            cur = await db.execute(
                """
                SELECT node_code, node_name, parent_code, level, sort_order
                FROM data_account_metric_node
                WHERE is_active = 1
                ORDER BY level, sort_order, node_code
                """
            )
            node_rows = await cur.fetchall()
            nodes = {
                str(r[0]): {
                    "code": str(r[0]),
                    "name": str(r[1] or ""),
                    "parent": str(r[2]) if r[2] is not None else None,
                    "level": int(r[3] or 0),
                    "sort_order": int(r[4] or 0),
                }
                for r in node_rows
            }

            def ancestors(code: str) -> list[dict[str, Any]]:
                result: list[dict[str, Any]] = []
                seen: set[str] = set()
                cur_code = code
                while cur_code and cur_code not in seen:
                    seen.add(cur_code)
                    node = nodes.get(cur_code)
                    if not node:
                        break
                    result.append(node)
                    cur_code = str(node.get("parent") or "")
                result.reverse()
                return result

            cur = await db.execute(
                """
                SELECT b.binding_code, b.metric_node_code, n.node_name,
                       b.scope_type, b.scope_code, b.product_code, pt.product_name,
                       b.data_acct_code, da.data_acct_name, da.value_type,
                       b.sort_order
                FROM data_account_metric_binding b
                JOIN data_account_metric_node n ON n.node_code = b.metric_node_code
                JOIN data_account da ON da.data_acct_code = b.data_acct_code
                LEFT JOIN product_type pt ON pt.product_code = b.product_code
                WHERE COALESCE(TRIM(da.budget_formula), '') = ''
                  AND b.is_active = 1
                ORDER BY n.sort_order, b.metric_node_code, b.sort_order, b.scope_code, b.data_acct_code
                """
            )
            binding_rows = await cur.fetchall()
            mapped_data_codes: set[str] = set()
            actual_keys: set[tuple[str, str]] = set()
            for r in binding_rows:
                scope_type = str(r[3] or "").strip().upper()
                data_code = str(r[7]).strip().upper()
                product_code = "" if scope_type == "CORP" else str(r[4] or r[5] or "").strip().upper()
                mapped_data_codes.add(data_code)
                actual_keys.add((data_code, product_code))
            report_paths_by_data_code = await _load_report_paths_for_data_accounts(db, mapped_data_codes)
        actual_values_by_key = await _load_actual_values_by_key(
            budget_path, version_id, period_month_map, actual_keys
        )

        categories: dict[str, dict[str, Any]] = {}
        seq = 1
        for r in binding_rows:
            metric_node_code = str(r[1]).strip()
            path_nodes = ancestors(metric_node_code)
            if path_nodes:
                root = path_nodes[0]
                indicator_name = " / ".join(n["name"] for n in path_nodes[1:] if n.get("name")) or str(r[2] or metric_node_code)
            else:
                root = {"code": "UNBOUND", "name": "未绑定指标树", "sort_order": 999999}
                indicator_name = str(r[2] or metric_node_code)
            category_code = str(root["code"])
            scope_type = str(r[3] or "").strip().upper()
            product_code = "" if scope_type == "CORP" else str(r[4] or r[5] or "").strip().upper()
            product_name = "全行" if scope_type == "CORP" else str(r[6] or product_code)
            data_code = str(r[7]).strip().upper()
            data_name = str(r[8])
            value_type = str(r[9])
            sort_order = int(r[10] or 0)
            first_path = (report_paths_by_data_code.get(data_code) or [{}])[0]
            category = categories.setdefault(category_code, {
                "category_code": category_code,
                "category_name": str(root["name"]),
                "sort_order": int(root.get("sort_order") or 0),
                "indicators": {},
            })
            item = category["indicators"].setdefault(metric_node_code, {
                "indicator_code": metric_node_code,
                "indicator_name": indicator_name,
                "value_type": value_type,
                "sort_order": sort_order,
                "data_acct_code": data_code,
                "products": [],
            })
            item["products"].append(DriverProductRow(
                id=seq,
                indicator_code=metric_node_code,
                product_code=product_code,
                product_name=product_name,
                sort_order=sort_order,
                data_accounts=[
                    DriverMappedDataAccount(
                        data_acct_code=data_code,
                        data_acct_name=data_name,
                        value_type=value_type,
                        report_code=first_path.get("report_code"),
                        report_path=first_path.get("report_path", ["未映射数据科目"]),
                        actual_values=actual_values_by_key.get((data_code, product_code), [0.0] * 12),
                        sort_order=sort_order,
                    )
                ],
            ))
            seq += 1

        result: list[DriverCategoryTree] = []
        for cat in sorted(categories.values(), key=lambda x: (x["sort_order"], x["category_code"])):
            indicators = [
                DriverIndicatorTree(
                    indicator_code=ind["indicator_code"],
                    indicator_name=ind["indicator_name"],
                    value_type=ind["value_type"],
                    data_acct_code=ind["data_acct_code"],
                    has_product_detail=1,
                    has_monthly_detail=1,
                    sort_order=ind["sort_order"],
                    products=sorted(ind["products"], key=lambda p: (p.sort_order, p.product_code, p.data_accounts[0].data_acct_code if p.data_accounts else "")),
                )
                for ind in sorted(cat["indicators"].values(), key=lambda x: (x["sort_order"], x["indicator_code"]))
            ]
            result.append(DriverCategoryTree(
                category_code=cat["category_code"],
                category_name=cat["category_name"],
                current_month=current_month,
                sort_order=cat["sort_order"],
                indicators=indicators,
            ))

        return result

    # ── GET 模板下载 ──
    @router.get("/api/driver/template")
    async def download_driver_template():
        category_tree = await list_driver_categories()
        categories: list[dict[str, Any]] = []
        indicators: list[dict[str, Any]] = []
        products: list[dict[str, Any]] = []
        for cat in category_tree:
            categories.append({
                "category_code": cat.category_code,
                "category_name": cat.category_name,
                "sort_order": cat.sort_order,
            })
            for ind in cat.indicators:
                indicators.append({
                    "indicator_code": ind.indicator_code,
                    "category_code": cat.category_code,
                    "indicator_name": ind.indicator_name,
                    "value_type": ind.value_type,
                    "data_acct_code": ind.data_acct_code,
                    "has_product_detail": ind.has_product_detail,
                    "has_monthly_detail": ind.has_monthly_detail,
                    "sort_order": ind.sort_order,
                })
                for product in ind.products:
                    products.append({
                        "indicator_code": ind.indicator_code,
                        "product_code": product.product_code,
                        "product_name": product.product_name,
                        "data_accounts": [account.model_dump() for account in product.data_accounts],
                    })

        stream = await asyncio.to_thread(_generate_excel_template, categories, indicators, products)
        return StreamingResponse(
            stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=driver_prediction_temp.xlsx"},
        )

    # ── POST 导入预览 ──
    @router.post("/api/driver/preview", response_model=DriverImportPreviewResponse)
    async def preview_driver_import(file: UploadFile = File(...)):
        if not file.filename or not str(file.filename).lower().endswith((".xlsx", ".xlsm")):
            raise HTTPException(status_code=400, detail="请上传 .xlsx 或 .xlsm 文件")
        content = await file.read()
        if len(content) > 15 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="文件过大（上限 15MB）")
        try:
            wb = await asyncio.to_thread(lambda: __import__("openpyxl").load_workbook(filename=BytesIO(content), data_only=True))
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"无法读取 Excel 文件：{exc}")
        category_tree = await list_driver_categories()
        current_month = category_tree[0].current_month if category_tree else 1
        all_indicators: list[dict[str, Any]] = []
        product_by_code: dict[str, dict[str, Any]] = {}
        account_mappings: dict[tuple[str, str], list[str]] = {}
        for cat in category_tree:
            for ind in cat.indicators:
                all_indicators.append({
                    "indicator_code": ind.indicator_code,
                    "indicator_name": ind.indicator_name,
                    "value_type": ind.value_type,
                    "data_acct_code": ind.data_acct_code,
                    "category_code": cat.category_code,
                    "has_product_detail": int(ind.has_product_detail or 0),
                })
                for product in ind.products:
                    product_by_code[product.product_code] = {
                        "product_code": product.product_code,
                        "product_name": product.product_name,
                    }
                    for account in product.data_accounts:
                        account_mappings.setdefault(
                            _mapping_key(ind.indicator_code, product.product_code),
                            [],
                        ).append(account.data_acct_code.strip().upper())
        all_products = sorted(product_by_code.values(), key=lambda x: x["product_code"])

        plan_rows, preview_rows, errors, warnings, total_rows = _collect_driver_import_plan(
            wb,
            all_indicators=all_indicators,
            all_products=all_products,
            account_mappings=account_mappings,
            current_month=current_month,
        )
        return DriverImportPreviewResponse(
            row_count=total_rows,
            ready_rows=len(plan_rows),
            error_rows=len(errors),
            preview_rows=preview_rows[:50],
            errors=errors[:50],
            warnings=warnings[:50],
        )

    # ── POST 导入计算 ──
    @router.post("/api/driver/import", response_model=DriverImportResponse)
    async def import_driver_data(file: UploadFile = File(...)):
        if not file.filename or not str(file.filename).lower().endswith((".xlsx", ".xlsm")):
            raise HTTPException(status_code=400, detail="请上传 .xlsx 或 .xlsm 文件")

        raw = await file.read()
        if len(raw) > 15 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="文件过大（上限 15MB）")

        budget_path, budget_year, version_id = await editable_context_provider()
        current_month = await _load_current_month(budget_path, version_id)
        common_path = common_db_path()

        category_tree = await list_driver_categories()
        all_indicators: list[dict[str, Any]] = []
        product_by_code: dict[str, dict[str, Any]] = {}
        account_mappings: dict[tuple[str, str], list[str]] = {}
        for cat in category_tree:
            for ind in cat.indicators:
                all_indicators.append({
                    "indicator_code": ind.indicator_code,
                    "indicator_name": ind.indicator_name,
                    "value_type": ind.value_type,
                    "data_acct_code": ind.data_acct_code,
                    "category_code": cat.category_code,
                    "has_product_detail": int(ind.has_product_detail or 0),
                })
                for product in ind.products:
                    product_by_code[product.product_code] = {
                        "product_code": product.product_code,
                        "product_name": product.product_name,
                    }
                    for account in product.data_accounts:
                        account_mappings.setdefault(
                            _mapping_key(ind.indicator_code, product.product_code),
                            [],
                        ).append(account.data_acct_code.strip().upper())
        all_products = sorted(product_by_code.values(), key=lambda x: x["product_code"])

        # Load data account formula metadata
        async with aiosqlite.connect(common_path) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            cur = await db.execute(
                "SELECT data_acct_code, value_type, budget_formula FROM data_account ORDER BY data_acct_code"
            )
            data_accounts = {str(r[0]): {"value_type": r[1], "budget_formula": r[2]} for r in await cur.fetchall()}

        # Period map: month index → period_id
        month_to_period = _month_to_period_id(await get_year_period_months(budget_year))

        saved_cells = 0
        affected_products: set[str] = set()
        written_data_accts: set[str] = set()
        errors: list[str] = []
        warnings: list[str] = []

        # Parse Excel
        try:
            wb = await asyncio.to_thread(lambda: __import__("openpyxl").load_workbook(filename=BytesIO(raw), data_only=True))
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"无法读取 Excel：{exc}")

        plan_rows, _preview_rows, errors, warnings, _total_rows = _collect_driver_import_plan(
            wb,
            all_indicators=all_indicators,
            all_products=all_products,
            account_mappings=account_mappings,
            current_month=current_month,
        )

        async with aiosqlite.connect(budget_path) as bdb:
            await bdb.execute("PRAGMA foreign_keys = ON")
            now = _iso_now()
            for row_plan in plan_rows:
                product_code = str(row_plan["product_code"] or "")
                for mi, val in row_plan["month_values"]:
                    period_id = month_to_period.get(mi)
                    if not period_id:
                        warnings.append(f"行 {row_plan['excel_row']}: 月份 M{mi:02d} 未匹配到期间，已跳过")
                        continue
                    if mi < current_month:
                        warnings.append(
                            f"行 {row_plan['excel_row']}: M{mi:02d} 已进入实际数区间，不允许写入预测值，已跳过"
                        )
                        continue
                    for data_acct_code in row_plan["data_acct_codes"]:
                        await bdb.execute(
                            """
                            INSERT INTO budget_data (
                                data_acct_code, product_code, period_id, budget_actual,
                                version_id, value, need_calc, create_time, update_time
                            ) VALUES (?, ?, ?, 0, ?, ?, 0, ?, ?)
                            ON CONFLICT(data_acct_code, product_code, period_id, version_id, budget_actual)
                            DO UPDATE SET value = excluded.value, need_calc = 0, update_time = excluded.update_time
                            """,
                            (data_acct_code, product_code, period_id, version_id, val, now, now),
                        )
                        saved_cells += 1
                        written_data_accts.add(data_acct_code)
                if product_code:
                    affected_products.add(product_code)
            await bdb.commit()

        # Recalculate formulas for affected products
        for pc in sorted(affected_products):
            await recalculate_product_formula_rows(
                pc, version_id, 0,
                budget_path=budget_path, budget_year=budget_year,
            )

        # Rebuild budget_summary
        summary_rows = await rebuild_budget_summary(version_id, budget_path)

        # Compute summary: query key data accounts from budget_summary
        summary = {
            "version_id": version_id,
            "budget_year": budget_year,
            "saved_cells": saved_cells,
            "rebuilt_summary_rows": summary_rows,
            "affected_products": sorted(affected_products),
        }

        # Build simple monthly summary
        async with aiosqlite.connect(common_path) as cdb:
            await cdb.execute("PRAGMA foreign_keys = ON")
            cur = await cdb.execute("""
                SELECT data_acct_code, data_acct_name, value_type, budget_formula, actual_formula
                FROM data_account ORDER BY data_acct_code
            """)
            all_das = {str(r[0]): {"name": r[1], "value_type": r[2], "budget_formula": r[3], "actual_formula": r[4]}
                       for r in await cur.fetchall()}
            formula_data_accts = {
                code for code, info in all_das.items()
                if extract_formula_codes(str(info.get("budget_formula") or "")) & written_data_accts
            }

        monthly = []
        async with aiosqlite.connect(budget_path) as bdb:
            await bdb.execute("PRAGMA foreign_keys = ON")
            for mi in range(1, 13):
                month_label = f"M{mi:02d}"
                period_id = month_to_period.get(mi)
                if not period_id:
                    continue
                cur = await bdb.execute(
                    """
                    SELECT bd.data_acct_code, SUM(bd.value) as total
                    FROM budget_data bd
                    WHERE bd.version_id = ? AND bd.budget_actual = 0 AND bd.period_id = ?
                      AND (
                        bd.data_acct_code IN ({})
                        OR (bd.product_code IN ({}) AND bd.data_acct_code IN ({}))
                      )
                    GROUP BY bd.data_acct_code
                    ORDER BY bd.data_acct_code
                    """.format(
                        ",".join(["?"] * len(written_data_accts)) if written_data_accts else "''",
                        ",".join(["?"] * len(affected_products)) if affected_products else "''",
                        ",".join(["?"] * len(formula_data_accts)) if formula_data_accts else "''",
                    ),
                    (
                        version_id,
                        period_id,
                        *sorted(written_data_accts),
                        *sorted(affected_products),
                        *sorted(formula_data_accts),
                    ),
                )
                row_data = {"month": month_label}
                for r in await cur.fetchall():
                    code = str(r[0])
                    da_info = all_das.get(code, {})
                    name = da_info.get("name", code)
                    row_data[name] = round(float(r[1] or 0), 2)
                monthly.append(row_data)

        await write_operation_log(
            action_type="IMPORT",
            action_desc=f"驱动预测试算导入，写入 {saved_cells} 个单元格",
            target_table="budget_data",
            affected_rows=saved_cells,
            after_data={"version_id": version_id, "budget_year": budget_year, "saved_cells": saved_cells},
        )

        return DriverImportResponse(
            version_id=version_id,
            budget_year=budget_year,
            saved_cells=saved_cells,
            summary=summary,
            monthly=monthly,
            errors=errors,
            warnings=warnings,
        )

    # ── POST 直接 JSON 导入（用于测试）──
    @router.post("/api/driver/import-json", response_model=DriverImportResponse)
    async def import_driver_json(
        body: list[DriverImportRequest],
        recalculate: bool = Query(True),
    ):
        budget_path, budget_year, version_id = await editable_context_provider()
        current_month = await _load_current_month(budget_path, version_id)
        common_path = common_db_path()

        month_to_period = _month_to_period_id(await get_year_period_months(budget_year))
        saved_cells = 0
        affected_products: set[str] = set()
        written_data_accts: set[str] = set()
        errors: list[str] = []

        async with aiosqlite.connect(common_path) as cdb:
            await cdb.execute("PRAGMA foreign_keys = ON")
            cur = await cdb.execute(
                "SELECT indicator_code, value_type, data_acct_code FROM driver_indicator"
            )
            ind_map = {str(r[0]): {"indicator_code": str(r[0]), "value_type": r[1], "data_acct_code": r[2]}
                       for r in await cur.fetchall()}

            cur = await cdb.execute(
                """
                SELECT indicator_code, product_code, data_acct_code
                FROM driver_account_mapping
                ORDER BY indicator_code, product_code, sort_order, data_acct_code
                """
            )
            account_mappings: dict[tuple[str, str], list[str]] = {}
            for r in await cur.fetchall():
                account_mappings.setdefault(_mapping_key(str(r[0]), str(r[1])), []).append(str(r[2]).strip().upper())
            cur = await cdb.execute(
                "SELECT data_acct_code, data_acct_name, value_type, budget_formula FROM data_account ORDER BY data_acct_code"
            )
            data_account_meta = {
                str(r[0]).strip().upper(): {
                    "name": str(r[1]),
                    "value_type": str(r[2]),
                    "budget_formula": r[3],
                    "is_formula": bool(str(r[3] or "").strip()),
                }
                for r in await cur.fetchall()
            }

        async with aiosqlite.connect(budget_path) as bdb:
            await bdb.execute("PRAGMA foreign_keys = ON")
            now = _iso_now()

            for item in body:
                product_code = (item.product_code or "").strip().upper()
                requested_data_code = (item.data_acct_code or item.indicator_code).strip().upper()
                if requested_data_code in data_account_meta:
                    if data_account_meta[requested_data_code]["is_formula"]:
                        errors.append(f"计算科目 {requested_data_code} 由公式生成，不允许在预测驱动中手工录入")
                        continue
                    data_acct_codes = [requested_data_code]
                    value_type = str(data_account_meta[requested_data_code]["value_type"])
                else:
                    ind_info = ind_map.get(item.indicator_code)
                    if not ind_info:
                        errors.append(f"未知数据科目: {item.indicator_code}")
                        continue
                    data_acct_codes = _resolve_data_acct_codes(
                        ind_info,
                        product_code,
                        account_mappings,
                        item.data_acct_code,
                    )
                    value_type = str(ind_info["value_type"])
                if not data_acct_codes:
                    if item.data_acct_code:
                        errors.append(
                            f"产品 {product_code or '全行'} 未绑定数据科目 {item.data_acct_code}"
                        )
                    else:
                        errors.append(f"产品 {product_code or '全行'} 未绑定数据科目")
                    continue
                is_pct = value_type == "百分比"

                for mv in item.monthly_values:
                    mi = int(mv.month.replace("M", "").lstrip("0") or "1")
                    period_id = month_to_period.get(mi)
                    if not period_id:
                        continue
                    if mi < current_month:
                        errors.append(f"M{mi:02d} 已进入实际数区间，不允许写入预测值")
                        continue

                    val = mv.value / 100.0 if is_pct else mv.value

                    for data_acct_code in data_acct_codes:
                        await bdb.execute(
                            """
                            INSERT INTO budget_data (
                                data_acct_code, product_code, period_id, budget_actual,
                                version_id, value, need_calc, create_time, update_time
                            ) VALUES (?, ?, ?, 0, ?, ?, 0, ?, ?)
                            ON CONFLICT(data_acct_code, product_code, period_id, version_id, budget_actual)
                            DO UPDATE SET value = excluded.value, need_calc = 0, update_time = excluded.update_time
                            """,
                            (data_acct_code, product_code, period_id, version_id, val, now, now),
                        )
                        saved_cells += 1
                        written_data_accts.add(data_acct_code)

                if product_code:
                    affected_products.add(product_code)

            await bdb.commit()

        summary_rows = 0
        if recalculate:
            for pc in sorted(affected_products):
                await recalculate_product_formula_rows(
                    pc, version_id, 0,
                    budget_path=budget_path, budget_year=budget_year,
                )
            summary_rows = await rebuild_budget_summary(version_id, budget_path)

        formula_data_accts = {
            code for code, info in data_account_meta.items()
            if str(info.get("budget_formula") or "").strip()
        }

        # Monthly summary
        async with aiosqlite.connect(budget_path) as bdb:
            await bdb.execute("PRAGMA foreign_keys = ON")
            monthly = []
            for mi in range(1, 13):
                period_id = month_to_period.get(mi)
                if not period_id:
                    continue
                cur = await bdb.execute(
                    """
                    SELECT bd.data_acct_code, bd.value
                    FROM budget_data bd
                    WHERE bd.version_id = ? AND bd.budget_actual = 0 AND bd.period_id = ?
                      AND (
                        bd.data_acct_code IN ({})
                        OR (bd.product_code IN ({}) AND bd.data_acct_code IN ({}))
                      )
                    ORDER BY bd.data_acct_code
                    """.format(
                        ",".join(["?"] * len(written_data_accts)) if written_data_accts else "''",
                        ",".join(["?"] * len(affected_products)) if affected_products else "''",
                        ",".join(["?"] * len(formula_data_accts)) if formula_data_accts else "''",
                    ),
                    (
                        version_id,
                        period_id,
                        *sorted(written_data_accts),
                        *sorted(affected_products),
                        *sorted(formula_data_accts),
                    ),
                )
                row_data = {"month": f"M{mi:02d}"}
                for r in await cur.fetchall():
                    code = str(r[0])
                    name = data_account_meta.get(code, {}).get("name", code)
                    row_data[str(name)] = float(r[1] or 0)
                monthly.append(row_data)

        await write_operation_log(
            action_type="IMPORT",
            action_desc=f"驱动预测保存{'并重算' if recalculate else ''}，写入 {saved_cells} 个单元格",
            target_table="budget_data",
            affected_rows=saved_cells,
            after_data={"version_id": version_id, "budget_year": budget_year, "saved_cells": saved_cells},
        )

        return DriverImportResponse(
            version_id=version_id,
            budget_year=budget_year,
            saved_cells=saved_cells,
            summary={
                "rebuilt_summary_rows": summary_rows,
                "affected_products": sorted(affected_products),
                "recalculated": recalculate,
                "written_data_accts": sorted(written_data_accts),
            },
            monthly=monthly,
            errors=errors,
            warnings=[],
        )

    return router
