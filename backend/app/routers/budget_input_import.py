from __future__ import annotations

import asyncio
from io import BytesIO
from pathlib import Path
from typing import Any, Awaitable, Callable

import aiosqlite
from fastapi import APIRouter, File, HTTPException, Query, UploadFile
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import load_workbook
from openpyxl.styles import Font

from app.budget_input_import import budget_input_template_path, run_budget_excel_import
from app.schemas import BudgetInputImportResponse


def build_budget_input_import_router(
    *,
    editable_context_provider: Callable[[], Awaitable[tuple[Path, int, int]]],
    recalculate_product_formula_rows: Callable[..., Awaitable[int]],
    get_year_period_months: Callable[[int], Awaitable[dict[int, int]]],
    get_version_current_month: Callable[[int, Path], Awaitable[int]],
    purge_disallowed_budget_data_for_version: Callable[..., Awaitable[None]],
    normalize_cell: Callable[[Any], str],
    write_operation_log: Callable[..., Awaitable[None]],
) -> APIRouter:
    router = APIRouter()

    @router.get("/api/budget-input/template")
    async def download_budget_input_template():
        path = budget_input_template_path()
        if not path.is_file():
            raise HTTPException(status_code=404, detail="预算上传模版文件不存在：download_template/budget_data_temp.xlsx")
        return FileResponse(
            path,
            filename="budget_data_temp.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @router.post("/api/budget-input/import-preview")
    async def preview_budget_input_import(file: UploadFile = File(...)):
        content = await file.read()
        try:
            wb = load_workbook(filename=BytesIO(content), data_only=True)
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"无法读取Excel文件：{exc}") from exc

        target_sheets = [s for s in ("预算数据", "实际数据") if s in wb.sheetnames]
        if not target_sheets:
            raise HTTPException(status_code=400, detail='上传文件缺失“预算数据/实际数据”工作表，不能上传数据。')

        # 使用第一个目标工作表的表头，并在预览中额外标注工作表来源。
        ws0 = wb[target_sheets[0]]
        headers = [normalize_cell(c.value) for c in ws0[1]]
        headers = [h for h in headers if h]
        if not headers:
            raise HTTPException(status_code=400, detail="模板表头为空")

        preview_rows: list[dict[str, str]] = []
        total_rows = 0
        for sheet_name in target_sheets:
            ws = wb[sheet_name]
            for ridx in range(2, ws.max_row + 1):
                row_values = [normalize_cell(ws.cell(ridx, c).value) for c in range(1, ws.max_column + 1)]
                if not any(row_values):
                    continue
                total_rows += 1
                if len(preview_rows) < 20:
                    preview_rows.append(
                        {"工作表": sheet_name, **{headers[i]: row_values[i] for i in range(min(len(headers), len(row_values))) if headers[i]}}
                    )

        return {
            "columns": ["工作表", *headers],
            "preview_rows": preview_rows,
            "row_count": total_rows,
        }

    @router.post("/api/budget-input/import-apply")
    async def apply_budget_input_import(
        file: UploadFile = File(...),
        version_id: int | None = Query(None),
    ):
        if not file.filename or not str(file.filename).lower().endswith((".xlsx", ".xlsm")):
            raise HTTPException(status_code=400, detail="请上传 .xlsx 或 .xlsm 文件（模版见「下载模版」）")

        raw = await file.read()
        if len(raw) > 15 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="文件过大（上限 15MB）")
        editable_budget_path, editable_year, editable_vid = await editable_context_provider()
        vid = int(version_id if version_id is not None else editable_vid)
        if vid != editable_vid:
            raise HTTPException(status_code=409, detail="当前可编辑版本已变更，请刷新页面后重试")

        try:
            result, products = await asyncio.to_thread(
                run_budget_excel_import,
                raw,
                vid,
                editable_year,
            )
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e)) from e
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"导入解析失败：{e}") from e

        # 按导入影响的产品做预算/实际全局公式重算，保证导入后一致性。
        for pc in sorted(products):
            await recalculate_product_formula_rows(
                pc,
                vid,
                0,
                budget_path=editable_budget_path,
                budget_year=editable_year,
            )

        period_month_map = await get_year_period_months(editable_year)
        current_month = await get_version_current_month(vid, editable_budget_path)
        async with aiosqlite.connect(editable_budget_path) as bdb:
            await bdb.execute("PRAGMA foreign_keys = ON")
            await purge_disallowed_budget_data_for_version(
                bdb, vid, current_month, period_month_map
            )
            await bdb.commit()
            await recalculate_product_formula_rows(
                pc,
                vid,
                1,
                budget_path=editable_budget_path,
                budget_year=editable_year,
            )

        try:
            wb = load_workbook(filename=BytesIO(raw))
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"无法回写导入结果文件：{exc}") from exc

        inserted = 0
        updated = 0
        failed = 0

        for sheet_name in ("预算数据", "实际数据"):
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            reason_col = ws.max_column + 1
            ws.cell(row=1, column=reason_col, value="导入评估说明")
            rows_by_excel = {
                int(r.excel_row): r
                for r in result.rows
                if r.sheet_name == sheet_name
            }
            for excel_row, row in rows_by_excel.items():
                reasons: list[str] = []
                for m in row.months:
                    col = 9 + int(m.month)  # 第10列-M1 ... 第21列-M12
                    if col < 1 or col > ws.max_column:
                        continue
                    cell = ws.cell(row=excel_row, column=col)
                    base_font = cell.font
                    if m.status == "inserted":
                        inserted += 1
                        color = "FF008000"
                    elif m.status == "updated":
                        updated += 1
                        color = "FF0000FF"
                    elif m.status in {"error", "skipped"}:
                        failed += 1
                        color = "FFFF0000"
                        if m.reason:
                            reasons.append(f"M{int(m.month):02d}:{m.reason}")
                    else:
                        continue
                    cell.font = Font(
                        name=base_font.name,
                        size=base_font.size,
                        bold=base_font.bold,
                        italic=base_font.italic,
                        underline=base_font.underline,
                        strike=base_font.strike,
                        color=color,
                    )
                if row.note:
                    reasons.append(row.note)
                if reasons:
                    ws.cell(
                        row=excel_row,
                        column=reason_col,
                        value="；".join(dict.fromkeys(r for r in reasons if r)),
                    )

        total = inserted + updated + failed
        out = BytesIO()
        wb.save(out)
        out.seek(0)

        await write_operation_log(
            action_type="IMPORT",
            action_desc=f"Excel 导入预算基础数据，写入 {inserted + updated} 个单元格",
            target_table="budget_data",
            affected_rows=inserted + updated,
            after_data={
                "version_id": vid,
                "budget_year": editable_year,
                "inserted_cells": inserted,
                "updated_cells": updated,
                "failed_cells": failed,
            },
        )

        return StreamingResponse(
            out,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": 'attachment; filename="budget_input_import_result.xlsx"',
                "X-Import-Total": str(total),
                "X-Import-Success": str(inserted + updated),
                "X-Import-Overwrite": str(updated),
                "X-Import-Failed": str(failed),
                "Access-Control-Expose-Headers": (
                    "Content-Disposition,"
                    "X-Import-Total,X-Import-Success,X-Import-Overwrite,X-Import-Failed"
                ),
            },
        )

    @router.post("/api/budget-input/import", response_model=BudgetInputImportResponse)
    async def import_budget_input_excel(
        file: UploadFile = File(...),
        version_id: int | None = Query(None),
    ):
        if not file.filename or not str(file.filename).lower().endswith((".xlsx", ".xlsm")):
            raise HTTPException(status_code=400, detail="请上传 .xlsx 或 .xlsm 文件（模版见「下载模版」）")

        raw = await file.read()
        if len(raw) > 15 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="文件过大（上限 15MB）")

        editable_budget_path, editable_year, editable_vid = await editable_context_provider()
        vid = int(version_id if version_id is not None else editable_vid)
        if vid != editable_vid:
            raise HTTPException(status_code=409, detail="当前可编辑版本已变更，请刷新页面后重试")

        try:
            result, products = await asyncio.to_thread(
                run_budget_excel_import,
                raw,
                vid,
                editable_year,
            )
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e)) from e
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"导入解析失败：{e}") from e

        for pc in sorted(products):
            await recalculate_product_formula_rows(
                pc,
                vid,
                0,
                budget_path=editable_budget_path,
                budget_year=editable_year,
            )

        period_month_map = await get_year_period_months(editable_year)
        current_month = await get_version_current_month(vid, editable_budget_path)
        async with aiosqlite.connect(editable_budget_path) as bdb:
            await bdb.execute("PRAGMA foreign_keys = ON")
            await purge_disallowed_budget_data_for_version(
                bdb, vid, current_month, period_month_map
            )
            await bdb.commit()
            await recalculate_product_formula_rows(
                pc,
                vid,
                1,
                budget_path=editable_budget_path,
                budget_year=editable_year,
            )

        await write_operation_log(
            action_type="IMPORT",
            action_desc=f"Excel 导入预算基础数据，写入 {result.saved_cells} 个单元格",
            target_table="budget_data",
            affected_rows=result.saved_cells,
            after_data={
                "version_id": vid,
                "budget_year": editable_year,
                "saved_cells": result.saved_cells,
            },
        )

        return result

    return router
