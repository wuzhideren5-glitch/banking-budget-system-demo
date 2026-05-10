"""费用实际数模块 ― Excel 模板生成、解析与导入。

约定：
  - 模板为归一化行格式，每行 = 一个(费用科目, 产品, 费用类型)组合
  - 费用类型（直接/间接）通过 data_acct_code 区分，不与 budget_data 表结构冲突
  - 导入目标：budget_data（budget_actual=1）
"""
from __future__ import annotations

import re
import sqlite3
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.months import parse_month_index

# ── 常量 ──────────────────────────────────────────────
TEMPLATE_SHEET_NAME = "费用实际数据"

# 必填导出列
FEE_TEMPLATE_COLUMNS = [
    ("费用科目代码", 18),
    ("费用科目名称", 22),
    ("产品代码", 14),
    ("产品名称", 20),
    ("费用类型", 12),
]

# 月份列 (M01-M12)
MONTH_COLS = [(f"M{m:02d}", 14) for m in range(1, 13)]

ALL_COLUMNS = FEE_TEMPLATE_COLUMNS + MONTH_COLS

# 样式
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
DATA_FONT = Font(name="微软雅黑", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


# ── 模板生成 ──────────────────────────────────────────

def _make_header_style() -> dict:
    return {"font": HEADER_FONT, "fill": HEADER_FILL, "border": THIN_BORDER, "alignment": CENTER}


def _make_data_style() -> dict:
    return {"font": DATA_FONT, "border": THIN_BORDER, "alignment": LEFT}


def generate_fee_actual_template(
    data_accounts: list[dict],
    products: list[dict],
) -> bytes:
    """生成费用实际导入模板（xlsx 二进制）。

    data_accounts: [{"data_acct_code": "C7001", "data_acct_name": "非IT人力基础薪酬"}, ...]
    products: [{"product_code": "Z0001", "product_name": "微粒贷产品"}, ...]
    """
    wb = Workbook()
    ws = wb.active
    ws.title = TEMPLATE_SHEET_NAME

    # ── 表头行 ──
    for ci, (title, width) in enumerate(ALL_COLUMNS, start=1):
        cell = ws.cell(row=1, column=ci, value=title)
        for k, v in _make_header_style().items():
            setattr(cell, k, v)
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 24

    # ── 数据行（预填代码）──
    row_idx = 2
    for acct in data_accounts:
        for prod in products:
            ws.cell(row=row_idx, column=1, value=acct["data_acct_code"])
            ws.cell(row=row_idx, column=2, value=acct["data_acct_name"])
            ws.cell(row=row_idx, column=3, value=prod["product_code"])
            ws.cell(row=row_idx, column=4, value=prod["product_name"])
            # 费用类型列留空，由用户填写"直接"或"间接"
            for ci in range(1, len(ALL_COLUMNS) + 1):
                cell = ws.cell(row=row_idx, column=ci)
                for k, v in _make_data_style().items():
                    setattr(cell, k, v)
            row_idx += 1

    # 冻结表头
    ws.freeze_panes = "A2"

    # 添加数据验证 - 费用类型列(E)只能填"直接"/"间接"
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(type="list", formula1='"直接,间接"', allow_blank=True)
    dv.error = '费用类型只能填\u201c直接\u201d或\u201c间接\u201d'
    dv.errorTitle = "输入错误"
    ws.add_data_validation(dv)
    dv.add(f"E2:E{row_idx - 1}")

    wb.active = 0
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


# ── 解析辅助 ──────────────────────────────────────────

def _normalize_cell(raw: Any) -> str:
    if raw is None:
        return ""
    return str(raw).strip()


def _parse_number(raw: Any) -> tuple[bool, float | None, str | None]:
    """(has_value, numeric, error_msg)"""
    if raw is None:
        return False, None, None
    if isinstance(raw, str):
        s = raw.strip().replace(",", "").replace("，", "")
        if not s:
            return False, None, None
        try:
            return True, float(s), None
        except ValueError:
            return True, None, "单元格不是有效数字"
    if isinstance(raw, bool):
        return True, float(raw), None
    if isinstance(raw, (int, float)):
        return True, float(raw), None
    return True, None, "单元格不是有效数字"


# ── 预览 / 导入结果 ───────────────────────────────────

class FeeActualImportMonthResult:
    def __init__(self, month: int, value_text: str = "", status: str = "empty", reason: str | None = None):
        self.month = month
        self.value_text = value_text
        self.status = status  # empty | inserted | updated | skipped | error
        self.reason = reason


class FeeActualImportRowResult:
    def __init__(
        self,
        excel_row: int,
        data_acct_code: str,
        product_code: str,
        fee_type: str,
        months: list[FeeActualImportMonthResult],
        note: str = "",
    ):
        self.excel_row = excel_row
        self.data_acct_code = data_acct_code
        self.product_code = product_code
        self.fee_type = fee_type
        self.months = months
        self.note = note


class FeeActualImportResponse:
    def __init__(
        self,
        budget_year: int,
        year_month: str,
        version_id: int,
        rows: list[FeeActualImportRowResult],
        saved_cells: int = 0,
    ):
        self.budget_year = budget_year
        self.year_month = year_month
        self.version_id = version_id
        self.rows = rows
        self.saved_cells = saved_cells


# ── 主入口 ────────────────────────────────────────────

def run_fee_actual_import(
    file_bytes: bytes,
    year_month: str,      # "2025-05"
    version_id: int,
    budget_year: int,
    common_conn: sqlite3.Connection,
    budget_conn: sqlite3.Connection,
    *,
    preview_only: bool = False,
) -> FeeActualImportResponse:
    """解析费用实际模板并写入/预览。

    返回 FeeActualImportResponse（preview_only=True 时 saved_cells=0）。
    """
    # ── 加载数据 ──
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    if TEMPLATE_SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"上传文件缺少「{TEMPLATE_SHEET_NAME}」工作表")

    ws = wb[TEMPLATE_SHEET_NAME]

    # ── 读取主数据 ──
    da_map: dict[str, dict] = {}
    for r in common_conn.execute(
        "SELECT data_acct_code, data_acct_name, value_type FROM data_account"
    ).fetchall():
        da_map[str(r[0]).strip().upper()] = {
            "name": str(r[1]),
            "value_type": str(r[2] or "金额"),
        }

    pt_set: set[str] = {
        str(r[0]).strip().upper()
        for r in common_conn.execute("SELECT product_code FROM product_type").fetchall()
    }

    # ── 期间映射 ──
    year_label = f"Y{budget_year}"
    period_month_map: dict[int, int] = {}
    month_to_period: dict[int, int] = {}
    for pid, mlab in common_conn.execute(
        "SELECT period_id, month FROM period WHERE year = ? ORDER BY period_id",
        (year_label,),
    ).fetchall():
        mi = parse_month_index(str(mlab or ""))
        if 1 <= mi <= 12:
            period_month_map[int(pid)] = mi
            month_to_period[mi] = int(pid)

    # 验证 year_month
    if not re.match(r"^\d{4}-(0[1-9]|1[0-2])$", year_month):
        raise ValueError("year_month 格式必须为 YYYY-MM，如 2025-05")

    # ── 检查当前月份窗口 ──
    cur_v = budget_conn.execute(
        "SELECT current_month FROM version WHERE version_id = ?", (version_id,)
    ).fetchone()
    if not cur_v:
        raise ValueError(f"版本 {version_id} 不存在")
    current_month = int(cur_v[0] or 1)

    # ── 解析表头验证 ──
    header_row = 1
    headers = [_normalize_cell(ws.cell(row=header_row, column=c).value) for c in range(1, 18)]
    expected_headers = [c[0] for c in ALL_COLUMNS]
    for i, (h, e) in enumerate(zip(headers, expected_headers)):
        if h != e:
            raise ValueError(f"表头列{i+1}应为「{e}」，实际为「{h}」")

    # ── 遍历数据行 ──
    max_row = ws.max_row or 0
    result_rows: list[FeeActualImportRowResult] = []
    saved_cells = 0
    row_now = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    for r in range(2, max_row + 1):
        data_acct_code = _normalize_cell(ws.cell(row=r, column=1).value).upper()
        product_code = _normalize_cell(ws.cell(row=r, column=3).value).upper()
        fee_type = _normalize_cell(ws.cell(row=r, column=5).value)

        if not data_acct_code and not product_code:
            continue  # 空行跳过

        months_out: list[FeeActualImportMonthResult] = []
        row_notes: list[str] = []

        # ── 校验主数据 ──
        ok = True
        if not data_acct_code:
            ok = False
            row_notes.append("费用科目代码为空")
        elif data_acct_code not in da_map:
            ok = False
            row_notes.append(f"费用科目代码「{data_acct_code}」不在主数据中")

        if not product_code:
            ok = False
            row_notes.append("产品代码为空")
        elif product_code not in pt_set:
            ok = False
            row_notes.append(f"产品代码「{product_code}」不在主数据中")

        if fee_type and fee_type not in ("直接", "间接"):
            row_notes.append(f"费用类型「{fee_type}」无效，应为\u201c直接\u201d或\u201c间接\u201d")

        if not ok:
            for m in range(1, 13):
                months_out.append(FeeActualImportMonthResult(month=m, status="error", reason=row_notes[0] if row_notes else "主数据校验失败"))
            result_rows.append(
                FeeActualImportRowResult(
                    excel_row=r,
                    data_acct_code=data_acct_code or "—",
                    product_code=product_code or "—",
                    fee_type=fee_type or "—",
                    months=months_out,
                    note="；".join(row_notes),
                )
            )
            continue

        vt = da_map[data_acct_code]["value_type"]

        for m in range(1, 13):
            col = 5 + m  # M01 在第 6 列（F列），M12 在第 17 列（Q列）
            cell_val = ws.cell(row=r, column=col).value
            has_v, num_v, perr = _parse_number(cell_val)

            if not has_v:
                months_out.append(FeeActualImportMonthResult(month=m, status="empty"))
                continue

            if perr or num_v is None:
                months_out.append(
                    FeeActualImportMonthResult(
                        month=m,
                        value_text=str(cell_val) if cell_val is not None else "",
                        status="error",
                        reason=perr or "无效数字",
                    )
                )
                row_notes.append(f"M{m:02d}:{perr or '无效数字'}")
                continue

            # 月份窗口限制
            if current_month < 1 or current_month > 13:
                current_month = 1
            # 实际数据：仅允许导入 current_month 及之前的月份
            if m > current_month:
                months_out.append(
                    FeeActualImportMonthResult(
                        month=m,
                        value_text=str(num_v),
                        status="skipped",
                        reason=f"M{m:02d}超出当前月份窗口（{current_month}月）",
                    )
                )
                row_notes.append(f"M{m:02d}:超出月份窗口")
                continue

            pid = month_to_period.get(m)
            if pid is None:
                months_out.append(
                    FeeActualImportMonthResult(
                        month=m,
                        value_text=str(num_v),
                        status="error",
                        reason="系统未找到对应会计期间",
                    )
                )
                row_notes.append(f"M{m:02d}:无对应期间")
                continue

            # 百分比按小数存储
            stored = num_v / 100.0 if vt == "百分比" else float(num_v)

            if not preview_only:
                cur_ex = budget_conn.execute(
                    """
                    SELECT 1 FROM budget_data
                    WHERE data_acct_code = ? AND product_code = ? AND period_id = ?
                      AND version_id = ? AND budget_actual = 1
                    """,
                    (data_acct_code, product_code, pid, version_id),
                )
                existed = cur_ex.fetchone() is not None

                budget_conn.execute(
                    """
                    INSERT INTO budget_data (
                      data_acct_code, product_code, period_id, budget_actual,
                      version_id, value, need_calc, create_time, update_time
                    ) VALUES (?, ?, ?, 1, ?, ?, 1, ?, ?)
                    ON CONFLICT(data_acct_code, product_code, period_id, version_id, budget_actual)
                    DO UPDATE SET value = excluded.value, need_calc = 1, update_time = excluded.update_time
                    """,
                    (data_acct_code, product_code, pid, version_id, stored, row_now, row_now),
                )

            disp = f"{num_v:.4g}" if vt != "百分比" else f"{num_v:.4g}%"
            months_out.append(
                FeeActualImportMonthResult(
                    month=m,
                    value_text=disp,
                    status="updated" if not preview_only and existed else ("inserted" if not preview_only else "preview_ok"),
                )
            )
            if not preview_only:
                saved_cells += 1

        note = "；".join(dict.fromkeys(row_notes))
        result_rows.append(
            FeeActualImportRowResult(
                excel_row=r,
                data_acct_code=data_acct_code,
                product_code=product_code,
                fee_type=fee_type,
                months=months_out,
                note=note,
            )
        )

    if not preview_only:
        budget_conn.commit()

    wb.close()
    return FeeActualImportResponse(
        budget_year=budget_year,
        year_month=year_month,
        version_id=version_id,
        rows=result_rows,
        saved_cells=saved_cells,
    )
