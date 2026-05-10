from __future__ import annotations

import argparse
from collections import Counter
from datetime import datetime
from pathlib import Path
import shutil
import sqlite3
from typing import Any

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[2]
COMMON_DB = ROOT / "data" / "common.db"
SOURCE_XLSX = ROOT / "report_tree_export2.0_05051000.xlsx"

SEMANTIC_MAPPING_OVERRIDES: dict[str, tuple[str, ...]] = {
    # These four are credit external interest income accounts. The source Excel
    # currently places them below FTP interest expense nodes, which is a
    # semantic error for report presentation and summary calculation.
    "C9015": ("X0301010103",),
    "C9016": ("X0301010103",),
    "C9017": ("X0301010103",),
    "C9018": ("X0301010103",),
}

SEMANTIC_REJECT_PAIRS: set[tuple[str, str]] = {
    ("X0301010102", "C9015"),
    ("X0301010102", "C9016"),
    ("X0301010102", "C9017"),
    ("X0301010102", "C9018"),
    ("X0301010104", "C9015"),
    ("X0301010104", "C9016"),
    ("X0301010104", "C9017"),
    ("X0301010104", "C9018"),
}


def _now_tag() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _parse_report_tree(path: Path) -> tuple[dict[str, dict[str, Any]], list[dict[str, str]]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    if "数据模版" not in wb.sheetnames:
        raise RuntimeError(f"{path.name} 缺少“数据模版”工作表")
    ws = wb["数据模版"]

    nodes: dict[str, dict[str, Any]] = {}
    mappings: list[dict[str, str]] = []
    current_path: list[str | None] = [None] * 5

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None and str(v).strip() for v in row):
            continue

        for idx in range(5):
            code_raw = row[idx * 2]
            name_raw = row[idx * 2 + 1]
            if not code_raw:
                continue
            code = str(code_raw).strip().upper()
            name = str(name_raw or "").strip()
            parent_code = current_path[idx - 1] if idx > 0 else None
            nodes[code] = {
                "report_acct_code": code,
                "report_acct_name": name,
                "parent_code": parent_code,
                "level": idx + 1,
                "is_summary": int(row[12] or 0),
                "is_minus": int(row[13] or 0),
                "remark": str(row[14]).strip() if row[14] is not None else None,
            }
            current_path[idx] = code
            for clear_idx in range(idx + 1, 5):
                current_path[clear_idx] = None

        data_code_raw = row[10] if len(row) > 10 else None
        if data_code_raw:
            report_code = next((current_path[i] for i in range(4, -1, -1) if current_path[i]), None)
            if report_code:
                mappings.append(
                    {
                        "report_acct_code": report_code,
                        "data_acct_code": str(data_code_raw).strip().upper(),
                        "data_acct_name": str(row[11]).strip() if row[11] is not None else "",
                    }
                )

    return nodes, mappings


def _apply_semantic_mapping_rules(mappings: list[dict[str, str]]) -> list[dict[str, str]]:
    normalized: dict[tuple[str, str], dict[str, str]] = {}
    data_name_by_code = {item["data_acct_code"].upper(): item.get("data_acct_name", "") for item in mappings}

    for item in mappings:
        report_code = item["report_acct_code"].upper()
        data_code = item["data_acct_code"].upper()
        if (report_code, data_code) in SEMANTIC_REJECT_PAIRS:
            continue
        if data_code in SEMANTIC_MAPPING_OVERRIDES:
            continue
        normalized[(report_code, data_code)] = {
            **item,
            "report_acct_code": report_code,
            "data_acct_code": data_code,
        }

    for data_code, report_codes in SEMANTIC_MAPPING_OVERRIDES.items():
        for report_code in report_codes:
            normalized[(report_code, data_code)] = {
                "report_acct_code": report_code,
                "data_acct_code": data_code,
                "data_acct_name": data_name_by_code.get(data_code, ""),
            }

    return [normalized[key] for key in sorted(normalized)]


def _backup_common_db(db_path: Path) -> Path:
    backup_dir = db_path.parent / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup_path = backup_dir / f"common_before_report_catalog_maintenance_{_now_tag()}.db"
    shutil.copy2(db_path, backup_path)
    return backup_path


def _refresh_leaf_flags(conn: sqlite3.Connection) -> int:
    before = dict(conn.execute("SELECT report_acct_code, is_leaf FROM report_account").fetchall())
    conn.execute(
        """
        UPDATE report_account
        SET is_leaf = CASE
            WHEN EXISTS (
                SELECT 1
                FROM report_account child
                WHERE child.parent_code = report_account.report_acct_code
            )
            THEN 0 ELSE 1
        END
        """
    )
    after = dict(conn.execute("SELECT report_acct_code, is_leaf FROM report_account").fetchall())
    return sum(1 for code, old_value in before.items() if after.get(code) != old_value)


def _write_audit(
    *,
    output_path: Path,
    summary: list[tuple[str, Any]],
    node_updates: list[dict[str, Any]],
    mapping_adds: list[dict[str, Any]],
    mapping_skips: list[dict[str, Any]],
    leaf_counts: list[tuple[Any, Any]],
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "汇总"
    ws.append(["指标", "数量/内容"])
    for row in summary:
        ws.append(list(row))

    def add_sheet(title: str, rows: list[dict[str, Any]]) -> None:
        sheet = wb.create_sheet(title)
        headers = sorted({k for row in rows for k in row.keys()})
        if not headers:
            sheet.append(["无记录"])
            return
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(h, "") for h in headers])

    add_sheet("节点新增或更新", node_updates)
    add_sheet("新增映射", mapping_adds)
    add_sheet("跳过映射", mapping_skips)

    ws_leaf = wb.create_sheet("叶子标识统计")
    ws_leaf.append(["is_leaf", "数量"])
    for row in leaf_counts:
        ws_leaf.append(list(row))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def maintain(*, source_path: Path, db_path: Path, dry_run: bool = False) -> dict[str, Any]:
    if not source_path.exists():
        raise FileNotFoundError(f"找不到报告树源文件：{source_path}")
    if not db_path.exists():
        raise FileNotFoundError(f"找不到数据库文件：{db_path}")

    source_nodes, raw_source_mappings = _parse_report_tree(source_path)
    source_mappings = _apply_semantic_mapping_rules(raw_source_mappings)
    backup_path = None if dry_run else _backup_common_db(db_path)

    conn = sqlite3.connect(db_path)
    try:
        conn.execute("PRAGMA foreign_keys = ON")
        conn.execute("BEGIN")

        existing_nodes = {
            row[0]: {
                "report_acct_name": row[1],
                "parent_code": row[2],
                "is_summary": int(row[3] or 0),
                "is_minus": int(row[4] or 0),
                "level": int(row[5]),
                "remark": row[7],
            }
            for row in conn.execute(
                """
                SELECT report_acct_code, report_acct_name, parent_code, is_summary,
                       is_minus, level, is_leaf, remark
                FROM report_account
                """
            )
        }
        data_codes = {str(row[0]).upper() for row in conn.execute("SELECT data_acct_code FROM data_account")}
        existing_mappings = {
            (str(row[0]).upper(), str(row[1]).upper())
            for row in conn.execute("SELECT report_acct_code, data_acct_code FROM report_data_mapping")
        }

        node_updates: list[dict[str, Any]] = []
        for code, node in sorted(source_nodes.items()):
            current = existing_nodes.get(code)
            if current is None:
                conn.execute(
                    """
                    INSERT INTO report_account (
                        report_acct_code, report_acct_name, parent_code,
                        is_summary, is_minus, level, is_leaf, remark
                    ) VALUES (?, ?, ?, ?, ?, ?, 0, ?)
                    """,
                    (
                        code,
                        node["report_acct_name"],
                        node["parent_code"],
                        node["is_summary"],
                        node["is_minus"],
                        node["level"],
                        node["remark"],
                    ),
                )
                node_updates.append({**node, "action": "insert"})
                continue

            changed = {
                key: node[key]
                for key in ("report_acct_name", "parent_code", "is_summary", "is_minus", "level", "remark")
                if (current.get(key) or None) != (node.get(key) or None)
            }
            if changed:
                conn.execute(
                    """
                    UPDATE report_account
                    SET report_acct_name = ?, parent_code = ?, is_summary = ?,
                        is_minus = ?, level = ?, remark = ?
                    WHERE report_acct_code = ?
                    """,
                    (
                        node["report_acct_name"],
                        node["parent_code"],
                        node["is_summary"],
                        node["is_minus"],
                        node["level"],
                        node["remark"],
                        code,
                    ),
                )
                node_updates.append({**node, "action": "update", "changed_fields": ",".join(changed)})

        mapping_adds: list[dict[str, Any]] = []
        mapping_skips: list[dict[str, Any]] = []
        for item in source_mappings:
            report_code = item["report_acct_code"].upper()
            data_code = item["data_acct_code"].upper()
            if report_code not in source_nodes and report_code not in existing_nodes:
                mapping_skips.append({**item, "reason": "报告科目不存在"})
                continue
            if data_code not in data_codes:
                mapping_skips.append({**item, "reason": "数据科目不存在"})
                continue
            if (report_code, data_code) in existing_mappings:
                continue
            conn.execute(
                """
                INSERT INTO report_data_mapping (report_acct_code, data_acct_code)
                VALUES (?, ?)
                """,
                (report_code, data_code),
            )
            existing_mappings.add((report_code, data_code))
            mapping_adds.append({**item, "report_acct_code": report_code, "data_acct_code": data_code})

        leaf_changed = _refresh_leaf_flags(conn)

        duplicate_count = conn.execute(
            """
            SELECT COUNT(*)
            FROM (
                SELECT report_acct_code, data_acct_code, COUNT(*) c
                FROM report_data_mapping
                GROUP BY report_acct_code, data_acct_code
                HAVING c > 1
            )
            """
        ).fetchone()[0]
        orphan_report_count = conn.execute(
            """
            SELECT COUNT(*)
            FROM report_data_mapping m
            LEFT JOIN report_account r ON r.report_acct_code = m.report_acct_code
            WHERE r.report_acct_code IS NULL
            """
        ).fetchone()[0]
        orphan_data_count = conn.execute(
            """
            SELECT COUNT(*)
            FROM report_data_mapping m
            LEFT JOIN data_account d ON d.data_acct_code = m.data_acct_code
            WHERE d.data_acct_code IS NULL
            """
        ).fetchone()[0]
        leaf_counts = list(conn.execute("SELECT is_leaf, COUNT(*) FROM report_account GROUP BY is_leaf"))

        if dry_run:
            conn.rollback()
        else:
            conn.commit()

        output_path = ROOT / "data" / f"report_catalog_maintenance_{_now_tag()}.xlsx"
        summary = [
            ("源文件", str(source_path)),
            ("数据库", str(db_path)),
            ("备份文件", str(backup_path) if backup_path else "dry-run 未备份"),
            ("源文件报告节点", len(source_nodes)),
            ("源文件映射", len(source_mappings)),
            ("新增/更新节点", len(node_updates)),
            ("新增映射", len(mapping_adds)),
            ("跳过映射", len(mapping_skips)),
            ("叶子标识变化", leaf_changed),
            ("重复映射", duplicate_count),
            ("孤儿报告映射", orphan_report_count),
            ("孤儿数据映射", orphan_data_count),
            ("模式", "dry-run" if dry_run else "applied"),
        ]
        _write_audit(
            output_path=output_path,
            summary=summary,
            node_updates=node_updates,
            mapping_adds=mapping_adds,
            mapping_skips=mapping_skips,
            leaf_counts=leaf_counts,
        )
        return {
            "backup_path": str(backup_path) if backup_path else None,
            "audit_path": str(output_path),
            "source_nodes": len(source_nodes),
            "source_mappings": len(source_mappings),
            "node_updates": len(node_updates),
            "mapping_adds": len(mapping_adds),
            "mapping_skips": len(mapping_skips),
            "leaf_changed": leaf_changed,
            "duplicates": duplicate_count,
            "orphan_reports": orphan_report_count,
            "orphan_data": orphan_data_count,
            "leaf_counts": dict(Counter(dict(leaf_counts))),
            "dry_run": dry_run,
        }
    finally:
        conn.close()


def main() -> None:
    parser = argparse.ArgumentParser(description="维护报告科目树、报告-数据映射和叶子标识")
    parser.add_argument("--source", type=Path, default=SOURCE_XLSX, help="报告树 Excel 源文件")
    parser.add_argument("--db", type=Path, default=COMMON_DB, help="common.db 路径")
    parser.add_argument("--dry-run", action="store_true", help="只生成审计，不提交数据库")
    args = parser.parse_args()
    result = maintain(source_path=args.source, db_path=args.db, dry_run=args.dry_run)
    for key, value in result.items():
        print(f"{key}: {value}")


if __name__ == "__main__":
    main()
