#!/usr/bin/env python3
from __future__ import annotations

import argparse
import re
import shutil
import sqlite3
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
DATA_DIR = ROOT / "data"
COMMON_DB = DATA_DIR / "common.db"
CODE_RE = re.compile(r"^A(\d{4})$")


def ts() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def connect(path: Path) -> sqlite3.Connection:
    conn = sqlite3.connect(path)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def next_a_code(conn: sqlite3.Connection, start: int) -> tuple[str, int]:
    n = start
    while n <= 9999:
        code = f"A{n:04d}"
        exists = conn.execute(
            "SELECT 1 FROM data_account WHERE data_acct_code = ?",
            (code,),
        ).fetchone()
        if not exists:
            return code, n + 1
        n += 1
    raise RuntimeError("A#### 自动编码已耗尽")


def product_descendants(conn: sqlite3.Connection) -> dict[str, set[str]]:
    rows = conn.execute("SELECT product_code, parent_code FROM product_type").fetchall()
    children: dict[str, list[str]] = {}
    codes = set()
    for r in rows:
        code = str(r["product_code"])
        codes.add(code)
        parent = r["parent_code"]
        if parent:
            children.setdefault(str(parent), []).append(code)

    memo: dict[str, set[str]] = {}

    def collect(code: str) -> set[str]:
        if code in memo:
            return memo[code]
        result = {code} if code in codes else set()
        for child in children.get(code, []):
            result |= collect(child)
        memo[code] = result
        return result

    for code in codes:
        collect(code)
    return memo


def product_label(conn: sqlite3.Connection, code: str) -> str:
    row = conn.execute(
        "SELECT product_name FROM product_type WHERE product_code = ?",
        (code,),
    ).fetchone()
    return str(row["product_name"]) if row and row["product_name"] else code


def make_name(base_name: str, scope_code: str, label: str) -> str:
    suffix = label or scope_code
    if suffix and suffix not in base_name:
        return f"{suffix}_{base_name}"
    return base_name


def table_columns(conn: sqlite3.Connection, table: str) -> list[str]:
    return [str(r["name"]) for r in conn.execute(f"PRAGMA table_info({table})")]


def clone_data_account(
    conn: sqlite3.Connection,
    source: sqlite3.Row,
    new_code: str,
    new_name: str,
    scope_code: str,
) -> None:
    columns = table_columns(conn, "data_account")
    values = {col: source[col] for col in columns}
    values["data_acct_code"] = new_code
    values["data_acct_name"] = new_name
    values["product_code"] = scope_code or None
    values["product_codes"] = scope_code or None
    if "applies_to_all_products" in values:
        values["applies_to_all_products"] = 0
    placeholders = ",".join(["?"] * len(columns))
    conn.execute(
        f"INSERT INTO data_account({','.join(columns)}) VALUES ({placeholders})",
        [values[col] for col in columns],
    )


def update_budget_db(
    budget_db: Path,
    old_code: str,
    new_code: str,
    product_codes: set[str],
) -> int:
    if not budget_db.exists() or not product_codes:
        return 0
    conn = connect(budget_db)
    try:
        placeholders = ",".join(["?"] * len(product_codes))
        rows = conn.execute(
            f"""
            SELECT data_acct_code, product_code, period_id, budget_actual, version_id,
                   value, need_calc, create_time, update_time
            FROM budget_data
            WHERE data_acct_code = ?
              AND product_code IN ({placeholders})
            """,
            [old_code, *sorted(product_codes)],
        ).fetchall()
        moved = 0
        conn.execute("BEGIN")
        for r in rows:
            conn.execute(
                """
                INSERT INTO budget_data(
                  data_acct_code, product_code, period_id, budget_actual,
                  version_id, value, need_calc, create_time, update_time
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(data_acct_code, product_code, period_id, version_id, budget_actual)
                DO UPDATE SET value = excluded.value,
                              need_calc = excluded.need_calc,
                              update_time = excluded.update_time
                """,
                (
                    new_code,
                    r["product_code"],
                    r["period_id"],
                    r["budget_actual"],
                    r["version_id"],
                    r["value"],
                    r["need_calc"],
                    r["create_time"],
                    r["update_time"],
                ),
            )
            moved += 1
        conn.execute(
            f"""
            DELETE FROM budget_data
            WHERE data_acct_code = ?
              AND product_code IN ({placeholders})
            """,
            [old_code, *sorted(product_codes)],
        )
        conn.commit()
        return moved
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def write_audit(rows: list[dict], formula_refs: list[dict], out: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)

    def sheet(name: str, headers: list[str], data: list[dict]) -> None:
        ws = wb.create_sheet(name[:31])
        for idx, h in enumerate(headers, 1):
            cell = ws.cell(1, idx, h)
            cell.fill = fill
            cell.font = font
        for ridx, row in enumerate(data, 2):
            for cidx, h in enumerate(headers, 1):
                ws.cell(ridx, cidx, row.get(h))
        ws.freeze_panes = "A2"
        for cidx, h in enumerate(headers, 1):
            width = min(max(len(str(ws.cell(r, cidx).value or "")) for r in range(1, min(ws.max_row, 100) + 1)) + 2, 60)
            ws.column_dimensions[get_column_letter(cidx)].width = width

    sheet(
        "拆分结果",
        [
            "old_code",
            "old_name",
            "binding_code",
            "metric_node_code",
            "scope_code",
            "product_name",
            "new_code",
            "new_name",
            "kept_original",
            "budget_rows_moved",
            "report_mappings_added",
        ],
        rows,
    )
    sheet(
        "需人工复核公式",
        ["formula_code", "formula_name", "referenced_old_code", "formula_text"],
        formula_refs,
    )
    wb.save(out)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true", help="执行修复；默认只生成预览")
    args = parser.parse_args()

    run_ts = ts()
    backup_dir = DATA_DIR / "backups"
    backup_dir.mkdir(exist_ok=True)
    if args.apply:
        for db in [COMMON_DB, *sorted(DATA_DIR.glob("budget_*.db")), DATA_DIR / "compare.db"]:
            if db.exists():
                shutil.copy2(db, backup_dir / f"{db.stem}_before_split_multi_binding_{run_ts}{db.suffix}")

    conn = connect(COMMON_DB)
    desc = product_descendants(conn)
    max_num = 1
    for r in conn.execute("SELECT data_acct_code FROM data_account WHERE data_acct_code GLOB 'A[0-9][0-9][0-9][0-9]'"):
        m = CODE_RE.match(str(r["data_acct_code"]))
        if m:
            max_num = max(max_num, int(m.group(1)) + 1)

    multi_codes = [
        str(r["data_acct_code"])
        for r in conn.execute(
            """
            SELECT data_acct_code
            FROM data_account_metric_binding
            WHERE is_active = 1
            GROUP BY data_acct_code
            HAVING COUNT(*) > 1
            ORDER BY data_acct_code
            """
        )
    ]

    result_rows: list[dict] = []
    formula_refs: list[dict] = []
    split_map: dict[tuple[str, str], str] = {}
    budget_dbs = sorted(DATA_DIR.glob("budget_*.db"))

    try:
        if args.apply:
            conn.execute("BEGIN")

        for old_code in multi_codes:
            source = conn.execute(
                "SELECT * FROM data_account WHERE data_acct_code = ?",
                (old_code,),
            ).fetchone()
            if not source:
                continue
            bindings = conn.execute(
                """
                SELECT binding_code, metric_node_code, scope_code, product_code, sort_order
                FROM data_account_metric_binding
                WHERE data_acct_code = ? AND is_active = 1
                ORDER BY sort_order, binding_code
                """,
                (old_code,),
            ).fetchall()
            if len(bindings) <= 1:
                continue

            for idx, b in enumerate(bindings):
                scope = str(b["scope_code"] or "").strip().upper()
                label = product_label(conn, scope) if scope else "全行"
                kept = idx == 0
                new_code = old_code if kept else ""
                new_name = make_name(str(source["data_acct_name"]), scope, label)
                if not kept:
                    new_code, max_num = next_a_code(conn, max_num)
                    split_map[(old_code, str(b["binding_code"]))] = new_code
                    if args.apply:
                        clone_data_account(conn, source, new_code, new_name, scope)
                        conn.execute(
                            """
                            UPDATE data_account_metric_binding
                            SET data_acct_code = ?, updated_at = CURRENT_TIMESTAMP
                            WHERE binding_code = ?
                            """,
                            (new_code, b["binding_code"]),
                        )
                else:
                    if args.apply:
                        conn.execute(
                            """
                            UPDATE data_account
                            SET data_acct_name = ?, product_code = ?, product_codes = ?, applies_to_all_products = 0
                            WHERE data_acct_code = ?
                            """,
                            (new_name, scope or None, scope or None, old_code),
                        )

                moved = 0
                if args.apply and not kept:
                    products = desc.get(scope, {scope} if scope else set())
                    for budget_db in budget_dbs:
                        moved += update_budget_db(budget_db, old_code, new_code, products)

                    report_rows = conn.execute(
                        "SELECT report_acct_code FROM report_data_mapping WHERE data_acct_code = ?",
                        (old_code,),
                    ).fetchall()
                    added = 0
                    for rr in report_rows:
                        cur = conn.execute(
                            """
                            INSERT OR IGNORE INTO report_data_mapping(report_acct_code, data_acct_code)
                            VALUES (?, ?)
                            """,
                            (rr["report_acct_code"], new_code),
                        )
                        added += max(int(cur.rowcount or 0), 0)
                else:
                    added = 0

                result_rows.append({
                    "old_code": old_code,
                    "old_name": source["data_acct_name"],
                    "binding_code": b["binding_code"],
                    "metric_node_code": b["metric_node_code"],
                    "scope_code": scope,
                    "product_name": label,
                    "new_code": new_code,
                    "new_name": new_name,
                    "kept_original": "Y" if kept else "N",
                    "budget_rows_moved": moved,
                    "report_mappings_added": added,
                })

        old_codes = set(multi_codes)
        for r in conn.execute(
            """
            SELECT data_acct_code, data_acct_name, budget_formula, actual_formula
            FROM data_account
            WHERE COALESCE(budget_formula, '') <> '' OR COALESCE(actual_formula, '') <> ''
            """
        ):
            formula_text = f"{r['budget_formula'] or ''} {r['actual_formula'] or ''}"
            for old_code in sorted(old_codes):
                if re.search(rf"(?<![A-Z0-9_]){re.escape(old_code)}(?![A-Z0-9_])", formula_text):
                    formula_refs.append({
                        "formula_code": r["data_acct_code"],
                        "formula_name": r["data_acct_name"],
                        "referenced_old_code": old_code,
                        "formula_text": formula_text,
                    })

        if args.apply:
            conn.commit()
        audit_path = DATA_DIR / f"split_multi_metric_bindings_{'applied' if args.apply else 'preview'}_{run_ts}.xlsx"
        write_audit(result_rows, formula_refs, audit_path)
        print(audit_path)
        print(f"multi_codes={len(multi_codes)} rows={len(result_rows)} apply={args.apply}")
        return 0
    except Exception:
        if args.apply:
            conn.rollback()
        raise
    finally:
        conn.close()


if __name__ == "__main__":
    raise SystemExit(main())
