from __future__ import annotations

from io import BytesIO
import json
from pathlib import Path
import re
import sqlite3
from typing import Any, Awaitable, Callable

import aiosqlite
from fastapi import APIRouter, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from app.config import Settings
from app.db_paths import common_db_path
from app.schemas import (
    DataAccountCreate,
    DataAccountRow,
    DataAccountUpdate,
    ProductScopeMigrationFileItem,
    ProductScopeMigrationPreviewResponse,
)


def build_data_accounts_router(
    *,
    settings: Settings,
    load_budget_data_ref_counts: Callable[[], Awaitable[dict[str, int]]],
    row_to_account: Callable[[tuple[Any, ...]], DataAccountRow],
    enrich_account_usage_flags: Callable[[DataAccountRow], Awaitable[DataAccountRow]],
    load_data_account_scope_map: Callable[[aiosqlite.Connection], Awaitable[dict[str, bool]]],
    validate_formula_reference_scope: Callable[..., None],
    write_operation_log: Callable[..., Awaitable[None]],
    count_budget_data_refs: Callable[[str], Awaitable[int]],
    formulas_reference_code: Callable[[Any, Any, str], bool],
    get_account_row: Callable[[aiosqlite.Connection, str], Awaitable[dict[str, Any] | None]],
    normalize_cell: Callable[[Any], str],
    color_row: Callable[[Any, int, int, str], None],
    normalize_formula: Callable[[str | None], str],
    latest_version: Callable[[], Awaitable[tuple[int, str, str]]],
    try_latest_version_id: Callable[[], Awaitable[int | None]],
    recalculate_data_account_formula_all_products: Callable[..., Awaitable[None]],
    recalculate_data_account_rule_all_products: Callable[..., Awaitable[None]],
    set_data_account_need_calc: Callable[[str, int], Awaitable[None]],
    preview_insert_single_to_all_rows: Callable[[str, str], tuple[int, list[ProductScopeMigrationFileItem]]],
    preview_delete_all_to_single_rows: Callable[[str, str], tuple[int, list[ProductScopeMigrationFileItem]]],
    migrate_single_to_all_budget_data: Callable[[str, str], tuple[int, list[ProductScopeMigrationFileItem]]],
    migrate_all_to_single_budget_data: Callable[[str, str], tuple[int, list[ProductScopeMigrationFileItem]]],
) -> APIRouter:
    router = APIRouter()

    @router.get("/api/data-accounts", response_model=list[DataAccountRow])
    async def list_data_accounts():
        path = common_db_path()
        async with aiosqlite.connect(path) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            db.row_factory = aiosqlite.Row
            cur = await db.execute(
                """
                SELECT d.data_acct_code, d.data_acct_name, d.product_code,
                       d.product_codes,
                       d.budget_formula, d.actual_formula, d.need_calc, d.budget_rule_code,
                       d.budget_rule_config_json, d.value_type, d.remark,
                       p.product_name
                FROM data_account d
                LEFT JOIN product_type p ON d.product_code = p.product_code
                ORDER BY d.data_acct_code
                """
            )
            rows = await cur.fetchall()
            cur = await db.execute(
                """
                SELECT data_acct_code, COUNT(*)
                FROM report_data_mapping
                GROUP BY data_acct_code
                """
            )
            mapping_rows = await cur.fetchall()
        report_mapping_counts = {str(r[0]): int(r[1] or 0) for r in mapping_rows if r[0]}
        budget_ref_counts = await load_budget_data_ref_counts()
        accounts: list[DataAccountRow] = []
        for r in rows:
            account = row_to_account(tuple(r))
            budget_ref_count = budget_ref_counts.get(account.data_acct_code, 0)
            mapping_ref_count = report_mapping_counts.get(account.data_acct_code, 0)
            account.budget_data_ref_count = budget_ref_count
            account.report_mapping_ref_count = mapping_ref_count
            account.has_budget_data_records = budget_ref_count > 0
            accounts.append(account)
        return accounts

    @router.get("/api/data-accounts/export")
    async def export_data_accounts():
        path = common_db_path()
        async with aiosqlite.connect(path) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            cur = await db.execute(
                """
                SELECT d.data_acct_code, d.data_acct_name, d.budget_formula, d.actual_formula,
                       d.product_code, p.product_name, d.value_type, d.remark, d.need_calc,
                       d.budget_rule_code, d.budget_rule_config_json,
                       d.product_codes
                FROM data_account d
                LEFT JOIN product_type p ON d.product_code = p.product_code
                ORDER BY d.data_acct_code
                """
            )
            rows = await cur.fetchall()

        template_path = Path(__file__).resolve().parents[2] / "download_template" / "data_acct_temp.xlsx"
        if template_path.exists():
            wb = load_workbook(template_path)
        else:
            wb = Workbook()
        if "数据模版" in wb.sheetnames:
            ws = wb["数据模版"]
        else:
            ws = wb.active
            ws.title = "数据模版"

        required_headers = [
            "数据科目代码",
            "数据科目名称",
            "预算数计算公式",
            "实际数计算公式",
            "产品科目代码",
            "产品科目名称",
            "数值类型",
            "备注",
            "是否需重计算",
        ]
        header_to_col: dict[str, int] = {}
        max_col = ws.max_column or 1
        for col in range(1, max_col + 1):
            v = ws.cell(row=1, column=col).value
            if isinstance(v, str) and v.strip():
                header_to_col[v.strip()] = col
        next_col = max_col + 1
        for h in required_headers:
            if h not in header_to_col:
                ws.cell(row=1, column=next_col, value=h)
                header_to_col[h] = next_col
                next_col += 1

        if ws.max_row >= 2:
            ws.delete_rows(2, ws.max_row - 1)

        row_idx = 2
        for r in rows:
            ws.cell(row=row_idx, column=header_to_col["数据科目代码"], value=r[0])
            ws.cell(row=row_idx, column=header_to_col["数据科目名称"], value=r[1])
            ws.cell(row=row_idx, column=header_to_col["预算数计算公式"], value=r[2])
            ws.cell(row=row_idx, column=header_to_col["实际数计算公式"], value=r[3])
            # 产品科目：product_codes 三态：'all'=全部产品, ''=公司级, 'Z01,Z02'=指定产品
            _pcs_raw = r[9]  # product_codes 列（第10列，索引9）
            if _pcs_raw is None or str(_pcs_raw).upper() == 'ALL':
                # 'all' 或 NULL = 全部产品（兼容旧NULL）
                prod_code = "*"
                prod_name = "(全部产品)"
            elif (r[9] or "").strip() == "":
                # 空字符串 = 公司级科目，显示为空
                prod_code = ""
                prod_name = "(公司级)"
            else:
                # 具体产品代码
                prod_code = (_pcs_raw or "").strip()
                prod_name = ""
            ws.cell(row=row_idx, column=header_to_col["产品科目代码"], value=prod_code)
            ws.cell(row=row_idx, column=header_to_col["产品科目名称"], value=prod_name)
            ws.cell(row=row_idx, column=header_to_col["数值类型"], value=r[6])
            ws.cell(row=row_idx, column=header_to_col["备注"], value=r[7])
            ws.cell(row=row_idx, column=header_to_col["是否需重计算"], value=int(r[8] or 0))
            row_idx += 1

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": 'attachment; filename="data_account_export.xlsx"',
            },
        )

    @router.post("/api/data-accounts", response_model=DataAccountRow)
    async def create_data_account(body: DataAccountCreate):
        path = common_db_path()
        # 新逻辑：优先使用 product_codes（多产品列表）
        # product_codes 为 None 或 [] 都表示"适用所有产品科目"
        pc_list = body.product_codes if body.product_codes is not None else []
        if len(pc_list) == 0:
            # 空列表/None = 全部产品
            applies_all = True
            pc = None
            pc_str = None
        else:
            applies_all = False
            pc_str = ",".join(pc_list)
            pc = pc_list[0] if len(pc_list) == 1 else None
            # 验证所有产品代码存在
            async with aiosqlite.connect(path) as db:
                await db.execute("PRAGMA foreign_keys = ON")
                for p in pc_list:
                    cur = await db.execute(
                        "SELECT 1 FROM product_type WHERE product_code = ?", (p,)
                    )
                    if not await cur.fetchone():
                        raise HTTPException(status_code=400, detail=f"产品代码不存在：{p}")
        async with aiosqlite.connect(path) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            cur = await db.execute(
                "SELECT 1 FROM data_account WHERE data_acct_code = ?",
                (body.data_acct_code,),
            )
            if await cur.fetchone():
                raise HTTPException(status_code=409, detail="科目代码已存在")
            if not applies_all:
                # applies_all=False 说明 pc_list 有内容，已在上方通过产品存在性验证
                pass
            scope_by_code = await load_data_account_scope_map(db)
            validate_formula_reference_scope(
                formula=body.budget_formula,
                target_is_all=applies_all,
                scope_by_code=scope_by_code,
                formula_label="预算公式",
            )
            validate_formula_reference_scope(
                formula=body.actual_formula,
                target_is_all=applies_all,
                scope_by_code=scope_by_code,
                formula_label="实际公式",
            )
            await db.execute(
                """
                INSERT INTO data_account (
                  data_acct_code, data_acct_name, product_code,
                  product_codes,
                  budget_formula, actual_formula, budget_rule_code, budget_rule_config_json, value_type, remark
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    body.data_acct_code,
                    body.data_acct_name,
                    pc,
                    pc_str,
                    body.budget_formula,
                    body.actual_formula,
                    body.budget_rule_code,
                    body.budget_rule_config_json,
                    body.value_type,
                    body.remark,
                ),
            )
            await db.commit()
        ctx = f"业务年度 budget_{settings.budget_year}.db；当前版本见 session"
        await write_operation_log(
            action_type="INSERT",
            action_desc=f"新增数据科目 {body.data_acct_code}。{ctx}",
            target_table="data_account",
            affected_rows=1,
            after_data=body.model_dump(),
        )
        async with aiosqlite.connect(path) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            cur = await db.execute(
                """
                SELECT d.data_acct_code, d.data_acct_name, d.product_code,
                       d.product_codes,
                       d.budget_formula, d.actual_formula, d.need_calc, d.budget_rule_code,
                       d.budget_rule_config_json, d.value_type, d.remark,
                       p.product_name
                FROM data_account d
                LEFT JOIN product_type p ON d.product_code = p.product_code
                WHERE d.data_acct_code = ?
                """,
                (body.data_acct_code,),
            )
            r = await cur.fetchone()
        account = row_to_account(tuple(r))
        return await enrich_account_usage_flags(account)

    @router.delete("/api/data-accounts/{code}")
    async def delete_data_account(code: str):
        path = common_db_path()
        async with aiosqlite.connect(path) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            before = await get_account_row(db, code)
            if not before:
                raise HTTPException(status_code=404, detail="科目不存在")
            budget_ref_count = await count_budget_data_refs(code)
            cur = await db.execute(
                "SELECT COUNT(*) FROM report_data_mapping WHERE data_acct_code = ?",
                (code,),
            )
            n = (await cur.fetchone())[0]
            if budget_ref_count or n:
                reasons: list[str] = []
                if n:
                    reasons.append(f"已和报告科目建立映射（{n} 条）")
                if budget_ref_count:
                    reasons.append(f"已在预算数据库中有数据（{budget_ref_count} 条）")
                raise HTTPException(
                    status_code=409,
                    detail=f"该数据科目{ '，'.join(reasons) }，不能删除",
                )
            cur = await db.execute(
                """
                SELECT data_acct_code, budget_formula, actual_formula
                FROM data_account WHERE data_acct_code != ?
                """,
                (code,),
            )
            refs = await cur.fetchall()
            for ocode, bf, af in refs:
                if formulas_reference_code(bf, af, code):
                    raise HTTPException(
                        status_code=409,
                        detail=f"科目 {ocode} 的公式仍引用 {code}，无法删除",
                    )
            await db.execute(
                "DELETE FROM data_account WHERE data_acct_code = ?", (code,)
            )
            await db.commit()
        ctx = f"业务年度 budget_{settings.budget_year}.db；当前版本见 session"
        await write_operation_log(
            action_type="DELETE",
            action_desc=f"删除数据科目 {code}。{ctx}",
            target_table="data_account",
            affected_rows=1,
            before_data=before,
            after_data=None,
        )
        return {"ok": True}

    @router.post("/api/data-accounts/import-preview")
    async def preview_data_account_import(file: UploadFile = File(...)):
        content = await file.read()
        try:
            wb = load_workbook(filename=BytesIO(content), data_only=True)
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"无法读取Excel文件：{exc}") from exc

        if "数据模版" not in wb.sheetnames:
            raise HTTPException(status_code=400, detail='上传文件缺失“数据模版”工作表，不能上传数据。')

        ws = wb["数据模版"]
        headers = [normalize_cell(c.value) for c in ws[1]]
        if not any(headers):
            raise HTTPException(status_code=400, detail="数据模版工作表第一行字段头为空")

        preview_rows: list[dict[str, str]] = []
        total_rows = 0
        for ridx in range(2, ws.max_row + 1):
            row_values = [normalize_cell(ws.cell(ridx, c).value) for c in range(1, ws.max_column + 1)]
            if not any(row_values):
                continue
            total_rows += 1
            if len(preview_rows) < 20:
                preview_rows.append(
                    {headers[i]: row_values[i] for i in range(min(len(headers), len(row_values))) if headers[i]}
                )

        return {
            "columns": [h for h in headers if h],
            "preview_rows": preview_rows,
            "row_count": total_rows,
        }

    @router.post("/api/data-accounts/import-apply")
    async def apply_data_account_import(
        file: UploadFile = File(...),
        mappings_json: str = Form(...),
    ):
        try:
            mappings = json.loads(mappings_json)
        except json.JSONDecodeError as exc:
            raise HTTPException(status_code=400, detail="字段映射格式不合法") from exc
        if not isinstance(mappings, dict):
            raise HTTPException(status_code=400, detail="字段映射格式不合法")

        required_fields = ["code", "name", "valueType"]
        for f in required_fields:
            if not mappings.get(f):
                raise HTTPException(status_code=400, detail=f"字段映射缺失：{f}")

        content = await file.read()
        try:
            wb = load_workbook(filename=BytesIO(content))
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"无法读取Excel文件：{exc}") from exc
        if "数据模版" not in wb.sheetnames:
            raise HTTPException(status_code=400, detail='上传文件缺失“数据模版”工作表，不能上传数据。')
        ws = wb["数据模版"]
        headers = [normalize_cell(c.value) for c in ws[1]]
        header_col = {h: i + 1 for i, h in enumerate(headers) if h}

        for key, col_name in mappings.items():
            if col_name and col_name not in header_col:
                raise HTTPException(status_code=400, detail=f"映射列不存在：{col_name}")

        rows: list[dict[str, Any]] = []
        code_count: dict[str, int] = {}
        for ridx in range(2, ws.max_row + 1):
            raw_by_header = {
                h: normalize_cell(ws.cell(ridx, cidx).value) for h, cidx in header_col.items()
            }
            if not any(raw_by_header.values()):
                continue
            code = raw_by_header.get(mappings.get("code", ""), "").upper()
            rows.append(
                {
                    "row_idx": ridx,
                    "code": code,
                    "name": raw_by_header.get(mappings.get("name", ""), ""),
                    "budget_formula": raw_by_header.get(mappings.get("budgetFormula", ""), ""),
                    "actual_formula": raw_by_header.get(mappings.get("actualFormula", ""), ""),
                    "product_code": raw_by_header.get(mappings.get("product", ""), "").upper(),
                    "value_type": raw_by_header.get(mappings.get("valueType", ""), ""),
                    "remark": raw_by_header.get(mappings.get("remark", ""), ""),
                }
            )
            if code:
                code_count[code] = code_count.get(code, 0) + 1

        allowed_value_types = {"金额", "百分比", "户数"}
        data_code_re = re.compile(r"^[A-Z]\d{4}$")
        product_code_re = re.compile(r"^Z\w+$")  # 接受Z开头任意长度层级代码，如Z01/Z0101/Z020101

        path = common_db_path()
        success_count = 0
        overwrite_count = 0
        failed_count = 0
        pending_logs: list[dict[str, Any]] = []
        recalc_targets: list[tuple[str, str | None, int]] = []

        async with aiosqlite.connect(path) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            cur = await db.execute(
                "SELECT data_acct_code, budget_formula, actual_formula, product_codes FROM data_account"
            )
            existing_rows = await cur.fetchall()
            existing_codes = {str(r[0]) for r in existing_rows}
            existing_formulas = {
                str(r[0]): (r[1], r[2])
                for r in existing_rows
            }
            existing_scope_by_code = {
                str(r[0]): (r[3] or '').strip() == ''  # True=全部产品，False=指定产品 (product_codes为NULL或空字符串表示全部产品)
                for r in existing_rows
            }
            cur = await db.execute("SELECT product_code FROM product_type")
            existing_products = {str(r[0]) for r in await cur.fetchall()}
            staged_scope_by_code = dict(existing_scope_by_code)
            for staged_row in rows:
                staged_code = str(staged_row.get("code", "")).strip()
                if not staged_code:
                    continue
                staged_pc = str(staged_row.get("product_code", "")).strip().upper()
                if "-" in staged_pc:
                    staged_pc = staged_pc.split("-", 1)[0].strip().upper()
                staged_scope_by_code[staged_code] = not bool(staged_pc)

            for row in rows:
                errors: list[str] = []
                code = row["code"].strip()
                name = row["name"].strip()
                product_code_raw = row["product_code"].strip()
                value_type = row["value_type"].strip()

                if not code:
                    errors.append("数据科目代码不能为空")
                elif not data_code_re.match(code):
                    errors.append("数据科目代码格式错误（示例：A1001）")
                elif code_count.get(code, 0) > 1:
                    errors.append("上传文件内存在重复数据科目代码")

                if not name:
                    errors.append("数据科目名称不能为空")

                # 解析多产品代码（逗号分隔）
                product_codes_list: list[str] = []
                if product_code_raw:
                    for pc in product_code_raw.split(","):
                        pc = pc.strip().upper()
                        if "-" in pc:
                            pc = pc.split("-", 1)[0].strip().upper()
                        if pc:
                            product_codes_list.append(pc)
                
                # 验证每个产品代码
                for pc in product_codes_list:
                    if not product_code_re.match(pc):
                        errors.append(f"产品科目代码格式错误：{pc}")
                    elif pc not in existing_products:
                        errors.append(f"产品科目代码不存在：{pc}")

                if value_type not in allowed_value_types:
                    errors.append("数值类型必须为 金额/百分比/户数 且不能为空")

                # 多产品模式：只要列表不为空就属于特定产品，否则为全行
                target_is_all = len(product_codes_list) == 0
                try:
                    validate_formula_reference_scope(
                        formula=row["budget_formula"] or None,
                        target_is_all=target_is_all,
                        scope_by_code=staged_scope_by_code,
                        formula_label="预算公式",
                    )
                except HTTPException as exc:
                    errors.append(str(exc.detail))
                try:
                    validate_formula_reference_scope(
                        formula=row["actual_formula"] or None,
                        target_is_all=target_is_all,
                        scope_by_code=staged_scope_by_code,
                        formula_label="实际公式",
                    )
                except HTTPException as exc:
                    errors.append(str(exc.detail))

                if errors:
                    failed_count += 1
                    color_row(ws, row["row_idx"], ws.max_column, "FFFF0000")
                    continue

                if code in existing_codes:
                    prev_budget_formula, prev_actual_formula = existing_formulas.get(code, (None, None))
                    new_budget_formula = row["budget_formula"] or None
                    new_actual_formula = row["actual_formula"] or None
                    budget_formula_changed = (
                        normalize_formula(prev_budget_formula) != normalize_formula(new_budget_formula)
                    )
                    actual_formula_changed = (
                        normalize_formula(prev_actual_formula) != normalize_formula(new_actual_formula)
                    )
                    # 新逻辑：使用 product_codes 字段（逗号分隔）
                    product_codes_str = ",".join(product_codes_list) if product_codes_list else None
                    applies_all = len(product_codes_list) == 0
                    await db.execute(
                        """
                        UPDATE data_account
                        SET data_acct_name = ?, product_code = ?,
                            product_codes = ?,
                            budget_formula = ?, actual_formula = ?,
                            need_calc = CASE WHEN ? = 1 THEN 1 ELSE need_calc END,
                            value_type = ?, remark = ?
                        WHERE data_acct_code = ?
                        """,
                        (
                            name,
                            product_codes_list[0] if len(product_codes_list) == 1 else None,
                            product_codes_str,
                            new_budget_formula,
                            new_actual_formula,
                            int(budget_formula_changed or actual_formula_changed),
                            value_type,
                            row["remark"] or None,
                            code,
                        ),
                    )
                    if budget_formula_changed:
                        recalc_targets.append((code, new_budget_formula, 0))
                    if actual_formula_changed:
                        recalc_targets.append((code, new_actual_formula, 1))
                    existing_formulas[code] = (new_budget_formula, new_actual_formula)
                    overwrite_count += 1
                    success_count += 1
                    color_row(ws, row["row_idx"], ws.max_column, "FF0000FF")
                    pending_logs.append(
                        {
                            "action_type": "UPDATE",
                            "action_desc": f"Excel导入覆盖数据科目 {code}",
                            "target_table": "data_account",
                            "affected_rows": 1,
                            "after_data": {
                                "data_acct_code": code,
                                "data_acct_name": name,
                                "product_code": product_codes_list[0] if len(product_codes_list) == 1 else None,
                                "budget_formula": row["budget_formula"] or None,
                                "actual_formula": row["actual_formula"] or None,
                                "value_type": value_type,
                                "remark": row["remark"] or None,
                            },
                        }
                    )
                else:
                    # 新逻辑：使用 product_codes 字段（逗号分隔）
                    product_codes_str = ",".join(product_codes_list) if product_codes_list else None
                    applies_all = len(product_codes_list) == 0
                    await db.execute(
                        """
                        INSERT INTO data_account (
                          data_acct_code, data_acct_name, product_code,
                          product_codes,
                          budget_formula, actual_formula, value_type, remark
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            code,
                            name,
                            product_codes_list[0] if len(product_codes_list) == 1 else None,
                            product_codes_str,
                            row["budget_formula"] or None,
                            row["actual_formula"] or None,
                            value_type,
                            row["remark"] or None,
                        ),
                    )
                    existing_codes.add(code)
                    success_count += 1
                    pending_logs.append(
                        {
                            "action_type": "INSERT",
                            "action_desc": f"Excel导入新增数据科目 {code}",
                            "target_table": "data_account",
                            "affected_rows": 1,
                            "after_data": {
                                "data_acct_code": code,
                                "data_acct_name": name,
                                "product_code": product_codes_list[0] if len(product_codes_list) == 1 else None,
                                "budget_formula": row["budget_formula"] or None,
                                "actual_formula": row["actual_formula"] or None,
                                "value_type": value_type,
                                "remark": row["remark"] or None,
                            },
                        }
                    )

            await db.commit()

        if recalc_targets:
            version_id, _, _ = await latest_version()
            reset_codes: set[str] = set()
            for data_code, formula, budget_actual in recalc_targets:
                await recalculate_data_account_formula_all_products(
                    data_acct_code=data_code,
                    formula=formula,
                    version_id=version_id,
                    budget_actual=budget_actual,
                )
                reset_codes.add(data_code)
            for data_code in reset_codes:
                await set_data_account_need_calc(data_code, 0)

        for log in pending_logs:
            await write_operation_log(
                action_type=log["action_type"],
                action_desc=log["action_desc"],
                target_table=log["target_table"],
                affected_rows=log["affected_rows"],
                after_data=log["after_data"],
            )

        result_buf = BytesIO()
        wb.save(result_buf)
        result_buf.seek(0)
        total_rows = len(rows)
        headers = {
            "Content-Disposition": 'attachment; filename="data_account_import_result.xlsx"',
            "X-Import-Total": str(total_rows),
            "X-Import-Success": str(success_count),
            "X-Import-Overwrite": str(overwrite_count),
            "X-Import-Failed": str(failed_count),
        }
        return StreamingResponse(
            result_buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )

    @router.get(
        "/api/data-accounts/{code}/product-scope-migration-preview",
        response_model=ProductScopeMigrationPreviewResponse,
    )
    async def preview_product_scope_migration(
        code: str,
        target_all_products: bool = Query(..., description="目标是否为「适用所有产品」"),
        target_product_code: str | None = Query(None, description="目标单产品代码；target_all_products=false 时必填"),
    ):
        path = common_db_path()
        async with aiosqlite.connect(path) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            before = await get_account_row(db, code.strip().upper())
            if not before:
                raise HTTPException(status_code=404, detail="科目不存在")
        code_u = code.strip().upper()
        old_is_all = (before.get("product_codes") or "").strip() == ""
        old_pc = before.get("product_code")
        if not target_all_products:
            if not target_product_code or not str(target_product_code).strip():
                raise HTTPException(
                    status_code=400,
                    detail="选择单一产品时必须提供 target_product_code",
                )
        elif target_product_code and str(target_product_code).strip():
            raise HTTPException(
                status_code=400,
                detail="选择「适用所有产品」时不应再提供 target_product_code",
            )
        new_is_all = bool(target_all_products)
        new_pc: str | None = None if new_is_all else str(target_product_code).strip().upper()
        scope_changed = (new_is_all != old_is_all) or (str(new_pc or "") != str((old_pc or "") or ""))
        if not scope_changed:
            return ProductScopeMigrationPreviewResponse(
                data_acct_code=code_u,
                files=[],
                total_rows_to_insert=0,
                total_rows_to_delete=0,
                message="产品范围与目标一致，无需迁移。",
            )
        if (not old_is_all and new_is_all) and old_pc:
            tot, items = preview_insert_single_to_all_rows(code_u, str(old_pc))
            return ProductScopeMigrationPreviewResponse(
                data_acct_code=code_u,
                files=items,
                total_rows_to_insert=tot,
                total_rows_to_delete=0,
                message="将从单一产品扩展为「适用所有产品」：在 Data 目录下全部 budget_*.db 中补齐各产品明细行（新行默认值为 0）。",
            )
        if old_is_all and not new_is_all and new_pc:
            tot, items = preview_delete_all_to_single_rows(code_u, new_pc)
            return ProductScopeMigrationPreviewResponse(
                data_acct_code=code_u,
                files=items,
                total_rows_to_insert=0,
                total_rows_to_delete=tot,
                message="将从「适用所有产品」收敛为单一产品：在全部年度预算库中删除非保留产品的 budget_data 行（不可恢复）。",
            )
        raise HTTPException(
            status_code=400,
            detail="该预览仅用于「单产品 ↔ 适用所有产品」切换；其他产品绑定变更请使用常规编辑策略。",
        )

    @router.patch("/api/data-accounts/{code}", response_model=DataAccountRow)
    async def update_data_account(code: str, body: DataAccountUpdate):
        path = common_db_path()
        budget_formula_changed = False
        actual_formula_changed = False
        recalc_budget_formula: str | None = None
        recalc_actual_formula: str | None = None
        target_code = code
        migration_inserted_total: int | None = None
        migration_deleted_total: int | None = None
        migration_files_list: list[ProductScopeMigrationFileItem] | None = None
        async with aiosqlite.connect(path) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            before = await get_account_row(db, code)
            if not before:
                raise HTTPException(status_code=404, detail="科目不存在")
            # effective_is_all 已在旧字段块或 product_codes 块中设置
            updates: list[str] = []
            vals: list[Any] = []
            if body.data_acct_code is not None:
                next_code = body.data_acct_code.strip().upper()
                if next_code != code:
                    budget_ref_count = await count_budget_data_refs(code)
                    cur = await db.execute(
                        "SELECT COUNT(*) FROM report_data_mapping WHERE data_acct_code = ?",
                        (code,),
                    )
                    mapping_ref_count = int((await cur.fetchone())[0] or 0)
                    if budget_ref_count or mapping_ref_count:
                        reasons: list[str] = []
                        if mapping_ref_count:
                            reasons.append(f"已和报告科目建立映射（{mapping_ref_count} 条）")
                        if budget_ref_count:
                            reasons.append(f"已在预算数据库中有数据（{budget_ref_count} 条）")
                        raise HTTPException(
                            status_code=409,
                            detail=f"该数据科目{ '，'.join(reasons) }，不能修改科目代码",
                        )
                    cur = await db.execute(
                        "SELECT 1 FROM data_account WHERE data_acct_code = ?",
                        (next_code,),
                    )
                    if await cur.fetchone():
                        raise HTTPException(status_code=409, detail="科目代码已存在")
                    updates.append("data_acct_code = ?")
                    vals.append(next_code)
                    target_code = next_code
            if body.data_acct_name is not None:
                updates.append("data_acct_name = ?")
                vals.append(body.data_acct_name)
            # ── 三态 product_codes（方案B：已移除 product_codes三态）─────────────
            # 三态：None=所有产品, ''=公司级, 'Z01,Z02'=指定产品
            effective_is_all = True  # 默认值，product_codes 为空/None 时适用所有产品
            if "product_codes" in body.model_fields_set:
                pc_list = body.product_codes or []
                if len(pc_list) == 0:
                    # 空列表/None = 所有产品
                    pc_str = None
                    new_pc2 = None
                    effective_is_all = True
                else:
                    pc_str = ",".join(pc_list)
                    new_pc2 = pc_list[0] if len(pc_list) == 1 else None
                    effective_is_all = False
                    # 验证所有产品代码存在
                    for pc in pc_list:
                        cur = await db.execute(
                            "SELECT 1 FROM product_type WHERE product_code = ?", (pc,)
                        )
                        if not await cur.fetchone():
                            raise HTTPException(status_code=400, detail=f"产品代码不存在：{pc}")
                updates.append("product_codes = ?")
                vals.append(pc_str)
                updates.append("product_code = ?")
                vals.append(new_pc2)
            scope_by_code = await load_data_account_scope_map(db)
            scope_by_code[str(code).strip().upper()] = effective_is_all
            scope_by_code[str(target_code).strip().upper()] = effective_is_all
            if body.budget_formula is not None:
                validate_formula_reference_scope(
                    formula=body.budget_formula,
                    target_is_all=effective_is_all,
                    scope_by_code=scope_by_code,
                    formula_label="预算公式",
                )
                updates.append("budget_formula = ?")
                vals.append(body.budget_formula)
                budget_formula_changed = (
                    normalize_formula(body.budget_formula) != normalize_formula(before.get("budget_formula"))
                )
                recalc_budget_formula = body.budget_formula
                if normalize_formula(body.budget_formula):
                    updates.append("budget_rule_code = ?")
                    vals.append(None)
                    updates.append("budget_rule_config_json = ?")
                    vals.append(None)
            if body.actual_formula is not None:
                validate_formula_reference_scope(
                    formula=body.actual_formula,
                    target_is_all=effective_is_all,
                    scope_by_code=scope_by_code,
                    formula_label="实际公式",
                )
                updates.append("actual_formula = ?")
                vals.append(body.actual_formula)
                actual_formula_changed = (
                    normalize_formula(body.actual_formula) != normalize_formula(before.get("actual_formula"))
                )
                recalc_actual_formula = body.actual_formula
            if body.budget_rule_code is not None:
                updates.append("budget_rule_code = ?")
                vals.append(body.budget_rule_code or None)
                budget_formula_changed = True
                if str(body.budget_rule_code or "").strip():
                    updates.append("budget_formula = ?")
                    vals.append(None)
                    recalc_budget_formula = None
            if body.budget_rule_config_json is not None:
                updates.append("budget_rule_config_json = ?")
                vals.append(body.budget_rule_config_json or None)
                budget_formula_changed = True
            if budget_formula_changed or actual_formula_changed:
                updates.append("need_calc = 1")
            if body.value_type is not None:
                updates.append("value_type = ?")
                vals.append(body.value_type)
            if body.remark is not None:
                updates.append("remark = ?")
                vals.append(body.remark)
            if not updates:
                cur = await db.execute(
                    """
                    SELECT d.data_acct_code, d.data_acct_name, d.product_code,
                           d.product_codes,
                           d.budget_formula, d.actual_formula, d.need_calc, d.budget_rule_code,
                           d.budget_rule_config_json, d.value_type, d.remark,
                           p.product_name
                    FROM data_account d
                    LEFT JOIN product_type p ON d.product_code = p.product_code
                    WHERE d.data_acct_code = ?
                    """,
                    (code,),
                )
                r = await cur.fetchone()
                account = row_to_account(tuple(r))
                return await enrich_account_usage_flags(account)
            sql = f"UPDATE data_account SET {', '.join(updates)} WHERE data_acct_code = ?"
            vals.append(code)
            await db.execute(sql, vals)
            await db.commit()
            after = await get_account_row(db, target_code)
        ctx = f"业务年度 budget_{settings.budget_year}.db；当前版本见 session"
        await write_operation_log(
            action_type="UPDATE",
            action_desc=f"更新数据科目 {code}。{ctx}",
            target_table="data_account",
            affected_rows=1,
            before_data=before,
            after_data=after,
        )
        if budget_formula_changed or actual_formula_changed:
            version_id, _, _ = await latest_version()
            if budget_formula_changed:
                if normalize_formula(recalc_budget_formula):
                    await recalculate_data_account_formula_all_products(
                        data_acct_code=target_code,
                        formula=recalc_budget_formula,
                        version_id=version_id,
                        budget_actual=0,
                    )
                elif after and str(after.get("budget_rule_code") or "").strip():
                    await recalculate_data_account_rule_all_products(
                        data_acct_code=target_code,
                        rule_code=str(after.get("budget_rule_code") or "").strip(),
                        rule_config_json=after.get("budget_rule_config_json"),
                        version_id=version_id,
                        budget_actual=0,
                    )
            if actual_formula_changed:
                await recalculate_data_account_formula_all_products(
                    data_acct_code=target_code,
                    formula=recalc_actual_formula,
                    version_id=version_id,
                    budget_actual=1,
                )
            await set_data_account_need_calc(target_code, 0)
        elif migration_inserted_total is not None or migration_deleted_total is not None:
            version_id = await try_latest_version_id()
            if version_id is not None and after:
                bf = after.get("budget_formula")
                af = after.get("actual_formula")
                if normalize_formula(bf):
                    await recalculate_data_account_formula_all_products(
                        data_acct_code=target_code,
                        formula=bf,
                        version_id=version_id,
                        budget_actual=0,
                    )
                elif str(after.get("budget_rule_code") or "").strip():
                    await recalculate_data_account_rule_all_products(
                        data_acct_code=target_code,
                        rule_code=str(after.get("budget_rule_code") or "").strip(),
                        rule_config_json=after.get("budget_rule_config_json"),
                        version_id=version_id,
                        budget_actual=0,
                    )
                if normalize_formula(af):
                    await recalculate_data_account_formula_all_products(
                        data_acct_code=target_code,
                        formula=af,
                        version_id=version_id,
                        budget_actual=1,
                    )
                await set_data_account_need_calc(target_code, 0)
        async with aiosqlite.connect(path) as db:
            await db.execute("PRAGMA foreign_keys = ON")
            cur = await db.execute(
                """
                SELECT d.data_acct_code, d.data_acct_name, d.product_code,
                       d.product_codes, d.budget_formula, d.actual_formula, d.need_calc, d.budget_rule_code,
                       d.budget_rule_config_json, d.value_type, d.remark,
                       p.product_name
                FROM data_account d
                LEFT JOIN product_type p ON d.product_code = p.product_code
                WHERE d.data_acct_code = ?
                """,
                (target_code,),
            )
            r = await cur.fetchone()
        account = row_to_account(tuple(r))
        account = await enrich_account_usage_flags(account)
        if migration_inserted_total is not None or migration_deleted_total is not None:
            account = account.model_copy(
                update={
                    "migration_inserted_total": migration_inserted_total,
                    "migration_deleted_total": migration_deleted_total,
                    "migration_files": migration_files_list,
                }
            )
        return account

    return router
