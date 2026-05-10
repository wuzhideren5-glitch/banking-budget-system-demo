from __future__ import annotations

import re
from io import BytesIO
from pathlib import Path

from fastapi import APIRouter, File, HTTPException, Request, UploadFile
from openpyxl import load_workbook

from app.agent_graph import AgentGraphService
from app.agent_memory import ConversationMemoryStore
from app.schemas import (
    AgentChatRequest,
    AgentChatResponse,
    AgentFeedbackRequest,
    AgentFeedbackResponse,
    AgentFileParseResponse,
    AgentPivotSuggestion,
    AgentReplyOption,
)


def _compact_text(text: str) -> str:
    cleaned = re.sub(r"\s+", " ", text or "").strip()
    return cleaned


def _extract_key_points(text: str, limit: int = 5) -> list[str]:
    lines = [line.strip(" -\t") for line in text.splitlines() if line.strip()]
    if not lines:
        return []
    bullet_like = [x for x in lines if any(token in x for token in ("：", ":", "；", ";"))]
    chosen = bullet_like[:limit] if bullet_like else lines[:limit]
    return [_compact_text(x)[:120] for x in chosen if x]


def _summarize_text_for_agent(text: str, max_len: int = 220) -> str:
    compact = _compact_text(text)
    if not compact:
        return "文档可读内容较少，建议补充更清晰的业务描述后再分析。"
    if len(compact) <= max_len:
        return compact
    return f"{compact[:max_len]}..."


def _suggest_actions_from_text(text: str) -> list[str]:
    hints: list[str] = []
    raw = text or ""
    if any(k in raw for k in ["预算", "实际", "差异", "同比", "环比"]):
        hints.append("可继续让我按时间、部门、对比方式做预算执行差异分析。")
    if any(k in raw for k in ["部门", "条线", "分行", "业务对象"]):
        hints.append("可指定业务对象（如个人金融部/企业金融部）进行钻取。")
    if any(k in raw for k in ["季度", "月", "年度", "时间"]):
        hints.append("可补充明确时间范围（如 2026 年一季度/全年）提升分析准确性。")
    if not hints:
        hints.append("可告诉我你希望的时间范围、业务对象、对比方式和分析粒度，我来继续分析。")
    return hints[:3]


def _extract_text_from_upload(filename: str, content: bytes) -> tuple[str, list[str], str]:
    suffix = Path(filename or "").suffix.lower()
    warnings: list[str] = []

    if suffix in {".txt", ".md", ".csv", ".json"}:
        return content.decode("utf-8", errors="ignore"), warnings, suffix or "text"

    if suffix in {".html", ".htm"}:
        raw = content.decode("utf-8", errors="ignore")
        no_script = re.sub(r"<(script|style)\b[^>]*>.*?</\1>", " ", raw, flags=re.I | re.S)
        text = re.sub(r"<[^>]+>", " ", no_script)
        return text, warnings, suffix

    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        wb = load_workbook(filename=BytesIO(content), data_only=True, read_only=True)
        lines: list[str] = []
        for sheet in wb.worksheets[:3]:
            lines.append(f"工作表：{sheet.title}")
            for row in sheet.iter_rows(min_row=1, max_row=30, values_only=True):
                cells = [str(v).strip() for v in row if v is not None and str(v).strip()]
                if cells:
                    lines.append(" | ".join(cells))
        return "\n".join(lines), warnings, suffix

    if suffix == ".docx":
        try:
            from docx import Document
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"缺少 docx 解析依赖：{exc}") from exc
        doc = Document(BytesIO(content))
        lines = [p.text.strip() for p in doc.paragraphs if p.text and p.text.strip()]
        return "\n".join(lines), warnings, suffix

    if suffix == ".doc":
        warnings.append("检测到 .doc 老格式，建议转为 .docx 后上传以获得更完整解析。")
        return content.decode("utf-8", errors="ignore"), warnings, suffix

    if suffix == ".pdf":
        try:
            from pypdf import PdfReader
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"缺少 PDF 解析依赖：{exc}") from exc
        reader = PdfReader(BytesIO(content))
        pages = [page.extract_text() or "" for page in reader.pages]
        return "\n".join(pages), warnings, suffix

    if suffix in {".png", ".jpg", ".jpeg", ".webp", ".bmp", ".tiff"}:
        try:
            from PIL import Image
            import pytesseract
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"缺少图片 OCR 依赖：{exc}") from exc
        img = Image.open(BytesIO(content))
        text = pytesseract.image_to_string(img, lang="chi_sim+eng")
        return text, warnings, suffix

    warnings.append(f"暂未针对 {suffix or '未知类型'} 做专用解析，已按文本尝试提取。")
    return content.decode("utf-8", errors="ignore"), warnings, suffix or "unknown"


def build_agent_runtime_router(
    agent_service: AgentGraphService,
    memory_store: ConversationMemoryStore,
) -> APIRouter:
    router = APIRouter()

    def _prepend_user_greeting(reply: str, user_name: str) -> str:
        clean_reply = (reply or "").strip()
        clean_name = (user_name or "").strip()
        if not clean_reply or not clean_name:
            return clean_reply
        if clean_reply.startswith(f"{clean_name}，") or clean_reply.startswith(f"{clean_name},"):
            return clean_reply
        return f"{clean_name}，{clean_reply}"

    @router.post("/api/agent/chat", response_model=AgentChatResponse)
    async def agent_chat(req: AgentChatRequest, request: Request):
        query = req.message.strip()
        if not query:
            raise HTTPException(status_code=400, detail="message 不能为空")
        history: list[dict] = []
        for m in req.history:
            item: dict = {"role": m.role, "content": m.content}
            if m.dialogue_id is not None:
                item["dialogue_id"] = m.dialogue_id
            history.append(item)
        dialogue_state = {
            "last_dialogue_id": int(req.last_dialogue_id) if req.last_dialogue_id is not None else 0,
            "pending_query_spec": req.pending_query_spec if isinstance(req.pending_query_spec, dict) else None,
        }
        user_ctx = getattr(request.state, "current_user", None)
        result = agent_service.chat(
            query=query,
            history=history,
            top_k=req.top_k,
            dialogue_state=dialogue_state,
            trace_context={
                "channel": "web",
                "session_id": str((user_ctx or {}).get("session_id") or ""),
                "user_id": int((user_ctx or {}).get("user_id") or 0),
                "user_name": str((user_ctx or {}).get("user_name") or ""),
            },
        )
        user_name = str((user_ctx or {}).get("user_name", "")).strip()
        is_first_turn = int(req.last_dialogue_id or 0) <= 0
        is_simple_dialogue = bool(result.get("is_lightweight_social", False))
        reply_text = str(result.get("reply", ""))
        if user_name and (is_first_turn or is_simple_dialogue):
            reply_text = _prepend_user_greeting(reply_text, user_name)
        raw_opts = result.get("reply_options") or []
        reply_options: list[AgentReplyOption] = []
        if isinstance(raw_opts, list):
            for item in raw_opts:
                if not isinstance(item, dict):
                    continue
                oid = item.get("id")
                label = item.get("label")
                if isinstance(oid, str) and oid.strip() and isinstance(label, str) and label.strip():
                    reply_options.append(AgentReplyOption(id=oid.strip(), label=label.strip()))
        raw_pivot = result.get("pivot_suggestion")
        pivot_suggestion: AgentPivotSuggestion | None = None
        if isinstance(raw_pivot, dict):
            try:
                pivot_suggestion = AgentPivotSuggestion(**raw_pivot)
            except Exception:
                pivot_suggestion = None
        raw_pending = result.get("pending_query_spec")
        pending_out = raw_pending if isinstance(raw_pending, dict) else None
        return AgentChatResponse(
            reply=reply_text,
            intent_type=str(result.get("intent_type", "general")),
            next_action=str(result.get("next_action", "general_answer")),
            need_clarification=bool(result.get("need_clarification", False)),
            missing_slots=[str(x) for x in result.get("missing_slots", [])],
            clarification_options={
                str(k): [str(v) for v in vals]
                for k, vals in (result.get("clarification_options", {}) or {}).items()
            },
            assumptions=[str(x) for x in result.get("assumptions", [])],
            suggested_sql=result.get("suggested_sql"),
            kb_context=result.get("kb_context", {}),
            executed=bool(result.get("executed_result")),
            result_row_count=int((result.get("executed_result") or {}).get("row_count", 0)),
            result_preview=(result.get("executed_result") or {}).get("display_preview_rows", []),
            memory_id=result.get("memory_id"),
            reply_options=reply_options,
            open_pivot_table=bool(result.get("open_pivot_table", False)),
            pivot_suggestion=pivot_suggestion,
            dialogue_id=int(result.get("dialogue_id") or 1),
            pending_query_spec=pending_out,
        )

    @router.post("/api/agent/feedback", response_model=AgentFeedbackResponse)
    async def agent_feedback(req: AgentFeedbackRequest):
        memory_id = req.memory_id.strip()
        if not memory_id:
            raise HTTPException(status_code=400, detail="memory_id 不能为空")
        updated = memory_store.update_feedback(
            memory_id,
            satisfied=req.satisfied,
            comment=req.comment,
        )
        if not updated:
            raise HTTPException(status_code=404, detail="未找到对应 memory_id")
        return AgentFeedbackResponse(updated=True, memory_id=memory_id)

    @router.post("/api/agent/file/parse", response_model=AgentFileParseResponse)
    async def parse_agent_file(file: UploadFile = File(...)):
        filename = file.filename or "未命名文件"
        content = await file.read()
        if not content:
            raise HTTPException(status_code=400, detail="上传文件为空")
        try:
            extracted_text, warnings, file_type = _extract_text_from_upload(filename, content)
        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"文件解析失败：{exc}") from exc

        compact = _compact_text(extracted_text)
        if not compact:
            raise HTTPException(status_code=400, detail="未提取到有效文本内容")

        return AgentFileParseResponse(
            filename=filename,
            file_type=file_type,
            char_count=len(compact),
            summary=_summarize_text_for_agent(extracted_text),
            key_points=_extract_key_points(extracted_text),
            suggested_actions=_suggest_actions_from_text(extracted_text),
            warnings=warnings,
        )

    return router
