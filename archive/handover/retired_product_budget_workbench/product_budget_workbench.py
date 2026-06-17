from __future__ import annotations

import asyncio
import json
import logging
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from uuid import uuid4

import aiosqlite
from fastapi import APIRouter, HTTPException, Query
from pydantic import BaseModel, Field

from app.config import settings
from app.data_account_write import (
    ensure_metric_binding,
    ensure_product_metric_data_account,
    sync_budget_rule_to_data_account,
)
from app.db_paths import common_db_path
from app.deepseek_client import DeepseekClient


logger = logging.getLogger("uvicorn.error")

RULE_LABELS = {
    "formula": "公式计算",
    "driver": "动因测算",
    "allocation": "分摊规则",
    "manual": "人工录入",
}


def _metric_display_level(code: str | None, fallback: int = 1) -> int:
    text = str(code or "").strip()
    if re.fullmatch(r"\d+(?:\.\d+)*", text):
        return text.count(".") + 1
    return max(1, fallback - 1)


def _is_formal_metric_code(code: str | None) -> bool:
    return bool(re.fullmatch(r"\d{2}\.\d{2}\.\d{3}", str(code or "").strip()))


class WorkbenchComponentUpsert(BaseModel):
    product_code: str | None = None
    report_acct_code: str | None = None
    metric_node_code: str | None = None
    component_name: str | None = None
    template_id: str | None = None
    template_name: str | None = None
    rule_code: str | None = None
    source_type: str | None = None
    data_acct_code: str | None = None
    formula: str | None = None
    value_type: str | None = None
    status: str | None = None


class WorkbenchCreateComponent(BaseModel):
    product_code: str
    report_acct_code: str
    metric_node_code: str | None = None
    component_name: str = "新数据科目"


class WorkbenchTemplateCreate(BaseModel):
    template_name: str


class WorkbenchTemplateApply(BaseModel):
    product_code: str
    report_acct_code: str
    metric_node_code: str | None = None


class WorkbenchCreateDataAccount(BaseModel):
    data_acct_name: str | None = None
    metric_group_code: str | None = None
    metric_group_name: str | None = None


class WorkbenchCreateMetricDataAccount(BaseModel):
    product_code: str
    parent_metric_node_code: str
    metric_node_name: str
    component_name: str | None = None
    formula: str | None = None
    value_type: str = "金额"
    rule_code: str = "formula"
    source_type: str = "manual"
    node_type: str = "METRIC"


class WorkbenchConfirmMetricNode(BaseModel):
    suggested_name: str
    parent_code: str | None = "03"
    node_type: str = "METRIC"


class WorkbenchSuggestionRequest(BaseModel):
    product_code: str
    report_acct_code: str | None = None
    component_id: str | None = None


class WorkbenchSuggestionAdopt(BaseModel):
    suggestion: dict[str, Any] = Field(default_factory=dict)


class WorkbenchProductAction(BaseModel):
    product_code: str


class WorkbenchBulkProductAction(BaseModel):
    product_codes: list[str] = Field(default_factory=list)


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _new_id(prefix: str) -> str:
    return f"{prefix}-{uuid4().hex[:12]}"


def _safe_upper(value: str | None) -> str:
    return (value or "").strip().upper()


def _metric_node_compat_code(metric_node_code: str | None, fallback: str | None = None) -> str:
    return _safe_upper(metric_node_code) or _safe_upper(fallback) or "UNBOUND"


def _status_from_formula_and_binding(formula: str | None, data_acct_code: str | None) -> str:
    if not data_acct_code:
        return "warning"
    if formula and formula.strip():
        return "ready"
    return "draft"


def _synthetic_trial(seed: str, negative: bool = False) -> dict[str, float]:
    base = (sum(ord(ch) for ch in seed) % 700) + 180
    sign = -1 if negative else 1
    m01 = sign * float(base)
    m02 = sign * float(round(base * 1.05, 2))
    m03 = sign * float(round(base * 1.09, 2))
    return {
        "trial_m01": m01,
        "trial_m02": m02,
        "trial_m03": m03,
        "trial_annual": float(round((m01 + m02 + m03) * 4, 2)),
    }


def _loads_json_object(value: str | None) -> dict[str, Any] | None:
    if not value:
        return None
    try:
        parsed = json.loads(value)
    except Exception:
        return None
    return parsed if isinstance(parsed, dict) else None


def _extract_json_array(text: str | None) -> list[dict[str, Any]] | None:
    if not text:
        return None
    match = re.search(r"\[[\s\S]*\]", text)
    if not match:
        return None
    try:
        parsed = json.loads(match.group(0))
    except Exception:
        return None
    if not isinstance(parsed, list):
        return None
    return [item for item in parsed if isinstance(item, dict)]


def _clean_rule_code(value: Any) -> str:
    rule_code = str(value or "formula").strip()
    return rule_code if rule_code in RULE_LABELS else "formula"


def _clean_source_type(value: Any) -> str:
    source_type = str(value or "ai_suggestion").strip()
    return source_type or "ai_suggestion"


def _clean_value_type(value: Any) -> str:
    text = str(value or "金额").strip()
    return text if text in {"金额", "百分比", "户数"} else "金额"


def _clean_formula(value: Any) -> str:
    formula = str(value or "").strip()
    placeholders = ("选择数据科目", "待配置", "无公式", "N/A", "null", "None")
    if not formula or any(token in formula for token in placeholders):
        return ""
    return formula


def _formula_has_data_account_ref(formula: str) -> bool:
    return bool(re.search(r"\b[A-Z][0-9]{2,}\b", formula))


def _fallback_suggestion_for_component(component: aiosqlite.Row) -> dict[str, Any] | None:
    formula = _clean_formula(component["formula"])
    if not formula or not _formula_has_data_account_ref(formula):
        return None
    return {
        "component_id": component["component_id"],
        "title": "复用当前组件公式",
        "rule_code": _clean_rule_code(component["rule_code"]),
        "source_type": "ai_batch_validated",
        "formula": formula,
        "data_account_name": component["component_name"],
        "reason": "该组件已有可用公式，产品级 AI 配置优先保持现有口径，避免误改底层计算。",
    }


def _suggestion_by_component_id(items: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    for item in items:
        component_id = str(item.get("component_id") or "").strip()
        if component_id:
            result[component_id] = item
    return result


async def _table_has_fk_to(db: aiosqlite.Connection, table_name: str, target_table: str) -> bool:
    cur = await db.execute(f"PRAGMA foreign_key_list({table_name})")
    return any(str(row[2]) == target_table for row in await cur.fetchall())


async def _rebuild_runtime_table_without_report_fk(db: aiosqlite.Connection, table_name: str) -> None:
    if table_name == "product_budget_config_package":
        legacy_name = "product_budget_config_package__legacy_report_fk"
        create_sql = """
        CREATE TABLE product_budget_config_package (
          package_id INTEGER PRIMARY KEY AUTOINCREMENT,
          product_code TEXT NOT NULL REFERENCES product_type(product_code),
          status TEXT NOT NULL DEFAULT 'draft' CHECK (status IN ('draft', 'dispatched')),
          owner_user TEXT,
          active_report_acct_code TEXT,
          created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
          updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
          UNIQUE (product_code)
        )
        """
        columns = """
          package_id, product_code, status, owner_user,
          active_report_acct_code, created_at, updated_at
        """
        select_columns = columns
    elif table_name == "product_budget_component":
        legacy_name = "product_budget_component__legacy_report_fk"
        create_sql = """
        CREATE TABLE product_budget_component (
          component_id TEXT PRIMARY KEY NOT NULL,
          product_code TEXT NOT NULL REFERENCES product_type(product_code),
          report_acct_code TEXT NOT NULL,
          metric_node_code TEXT REFERENCES data_account_metric_node(node_code),
          component_name TEXT NOT NULL,
          template_id TEXT,
          template_name TEXT,
          template_source_component_id TEXT,
          rule_code TEXT NOT NULL DEFAULT 'formula',
          source_type TEXT NOT NULL DEFAULT 'manual',
          data_acct_code TEXT REFERENCES data_account(data_acct_code),
          formula TEXT,
          value_type TEXT NOT NULL DEFAULT '金额',
          status TEXT NOT NULL DEFAULT 'draft' CHECK (status IN ('draft', 'ready', 'warning', 'dispatched')),
          ai_reason TEXT,
          trial_m01 REAL,
          trial_m02 REAL,
          trial_m03 REAL,
          trial_annual REAL,
          sort_order INTEGER NOT NULL DEFAULT 0,
          created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
          updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        )
        """
        columns = """
          component_id, product_code, report_acct_code, metric_node_code,
          component_name, template_id, template_name, template_source_component_id,
          rule_code, source_type, data_acct_code, formula, value_type, status,
          ai_reason, trial_m01, trial_m02, trial_m03, trial_annual,
          sort_order, created_at, updated_at
        """
        select_columns = """
          component_id, product_code,
          COALESCE(NULLIF(report_acct_code, ''), NULLIF(metric_node_code, ''), component_id) AS report_acct_code,
          metric_node_code,
          component_name, template_id, template_name, template_source_component_id,
          rule_code, source_type, data_acct_code, formula, value_type, status,
          ai_reason, trial_m01, trial_m02, trial_m03, trial_annual,
          sort_order, created_at, updated_at
        """
    else:
        return

    await db.commit()
    await db.execute("PRAGMA foreign_keys = OFF")
    try:
        await db.executescript(
            f"""
            DROP TABLE IF EXISTS {legacy_name};
            ALTER TABLE {table_name} RENAME TO {legacy_name};
            {create_sql};
            INSERT INTO {table_name} ({columns})
            SELECT {select_columns}
            FROM {legacy_name};
            DROP TABLE {legacy_name};
            """
        )
        await db.commit()
    finally:
        await db.execute("PRAGMA foreign_keys = ON")


async def _ensure_runtime_tables(db: aiosqlite.Connection) -> None:
    await db.executescript(
        """
        CREATE TABLE IF NOT EXISTS product_budget_config_package (
          package_id INTEGER PRIMARY KEY AUTOINCREMENT,
          product_code TEXT NOT NULL REFERENCES product_type(product_code),
          status TEXT NOT NULL DEFAULT 'draft' CHECK (status IN ('draft', 'dispatched')),
          owner_user TEXT,
          active_report_acct_code TEXT,
          created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
          updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
          UNIQUE (product_code)
        );

        CREATE TABLE IF NOT EXISTS product_budget_component (
          component_id TEXT PRIMARY KEY NOT NULL,
          product_code TEXT NOT NULL REFERENCES product_type(product_code),
          report_acct_code TEXT NOT NULL,
          metric_node_code TEXT REFERENCES data_account_metric_node(node_code),
          component_name TEXT NOT NULL,
          template_id TEXT,
          template_name TEXT,
          template_source_component_id TEXT,
          rule_code TEXT NOT NULL DEFAULT 'formula',
          source_type TEXT NOT NULL DEFAULT 'manual',
          data_acct_code TEXT REFERENCES data_account(data_acct_code),
          formula TEXT,
          value_type TEXT NOT NULL DEFAULT '金额',
          status TEXT NOT NULL DEFAULT 'draft' CHECK (status IN ('draft', 'ready', 'warning', 'dispatched')),
          ai_reason TEXT,
          trial_m01 REAL,
          trial_m02 REAL,
          trial_m03 REAL,
          trial_annual REAL,
          sort_order INTEGER NOT NULL DEFAULT 0,
          created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
          updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        );

        CREATE INDEX IF NOT EXISTS idx_product_budget_component_product
        ON product_budget_component(product_code, report_acct_code);

        CREATE INDEX IF NOT EXISTS idx_product_budget_component_metric
        ON product_budget_component(product_code, metric_node_code);

        CREATE TABLE IF NOT EXISTS product_budget_component_template (
          template_id TEXT PRIMARY KEY NOT NULL,
          template_name TEXT NOT NULL,
          component_name TEXT NOT NULL,
          rule_code TEXT NOT NULL DEFAULT 'formula',
          source_type TEXT NOT NULL DEFAULT 'template',
          formula TEXT,
          value_type TEXT NOT NULL DEFAULT '金额',
          data_acct_code TEXT,
          source_component_id TEXT,
          created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
          updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        );
        """
    )
    cur = await db.execute("PRAGMA table_info(product_budget_component)")
    component_cols = {str(row[1]) for row in await cur.fetchall()}
    if "metric_node_code" not in component_cols:
        await db.execute("ALTER TABLE product_budget_component ADD COLUMN metric_node_code TEXT REFERENCES data_account_metric_node(node_code)")
    if "metric_binding_code" in component_cols:
        await _rebuild_runtime_table_without_report_fk(db, "product_budget_component")
    if await _table_has_fk_to(db, "product_budget_config_package", "report_account"):
        await _rebuild_runtime_table_without_report_fk(db, "product_budget_config_package")
    if await _table_has_fk_to(db, "product_budget_component", "report_account"):
        await _rebuild_runtime_table_without_report_fk(db, "product_budget_component")
    await db.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_product_budget_component_product
        ON product_budget_component(product_code, report_acct_code)
        """
    )
    await db.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_product_budget_component_metric
        ON product_budget_component(product_code, metric_node_code)
        """
    )
    await db.execute(
        """
        DELETE FROM product_budget_component
        WHERE COALESCE(metric_node_code, '') = ''
          AND COALESCE(data_acct_code, '') = ''
        """
    )
    await db.execute(
        """
        DELETE FROM product_budget_component
        WHERE rowid NOT IN (
          SELECT MIN(rowid)
          FROM product_budget_component
          WHERE COALESCE(metric_node_code, '') <> ''
            AND COALESCE(data_acct_code, '') <> ''
          GROUP BY product_code, metric_node_code, data_acct_code, component_name
        )
          AND COALESCE(metric_node_code, '') <> ''
          AND COALESCE(data_acct_code, '') <> ''
        """
    )
    await db.execute(
        """
        DELETE FROM product_budget_component
        WHERE rowid NOT IN (
          SELECT MIN(rowid)
          FROM product_budget_component
          WHERE COALESCE(metric_node_code, '') <> ''
            AND COALESCE(data_acct_code, '') = ''
          GROUP BY product_code, metric_node_code, component_name
        )
          AND COALESCE(metric_node_code, '') <> ''
          AND COALESCE(data_acct_code, '') = ''
        """
    )
    await db.execute(
        """
        WITH ranked AS (
          SELECT rowid AS rid,
                 ROW_NUMBER() OVER (
                   PARTITION BY product_code, metric_node_code
                   ORDER BY
                     CASE WHEN COALESCE(data_acct_code, '') <> '' THEN 0 ELSE 1 END,
                     sort_order,
                     created_at,
                     rowid
                 ) AS rn
          FROM product_budget_component
          WHERE metric_node_code GLOB '[0-9][0-9].[0-9][0-9].[0-9][0-9][0-9]'
        )
        DELETE FROM product_budget_component
        WHERE rowid IN (SELECT rid FROM ranked WHERE rn > 1)
        """
    )
    await db.execute(
        """
        CREATE UNIQUE INDEX IF NOT EXISTS idx_product_budget_component_metric_data_name_unique
        ON product_budget_component(product_code, metric_node_code, data_acct_code, component_name)
        WHERE COALESCE(metric_node_code, '') <> ''
          AND COALESCE(data_acct_code, '') <> ''
        """
    )
    await db.execute(
        """
        CREATE UNIQUE INDEX IF NOT EXISTS idx_product_budget_component_metric_unbound_name_unique
        ON product_budget_component(product_code, metric_node_code, component_name)
        WHERE COALESCE(metric_node_code, '') <> ''
          AND COALESCE(data_acct_code, '') = ''
        """
    )
    await db.execute(
        """
        CREATE UNIQUE INDEX IF NOT EXISTS idx_product_budget_component_product_metric_unique
        ON product_budget_component(product_code, metric_node_code)
        WHERE COALESCE(metric_node_code, '') <> ''
          AND metric_node_code GLOB '[0-9][0-9].[0-9][0-9].[0-9][0-9][0-9]'
        """
    )
    await db.commit()


async def _ensure_package(db: aiosqlite.Connection, product_code: str) -> None:
    await db.execute(
        """
        INSERT INTO product_budget_config_package(product_code, status, updated_at)
        VALUES (?, 'draft', ?)
        ON CONFLICT(product_code) DO UPDATE SET updated_at = excluded.updated_at
        """,
        (product_code, _now()),
    )


async def _load_product_parent_map(db: aiosqlite.Connection) -> dict[str, str | None]:
    cur = await db.execute("SELECT product_code, parent_code FROM product_type")
    rows = await cur.fetchall()
    return {
        str(row["product_code"]).strip().upper(): (
            str(row["parent_code"]).strip().upper() if row["parent_code"] else None
        )
        for row in rows
    }


def _product_scope_codes(product_code: str, parent_map: dict[str, str | None]) -> list[str]:
    codes: list[str] = []
    seen: set[str] = set()
    current = _safe_upper(product_code)
    while current and current not in seen:
        seen.add(current)
        codes.append(current)
        current = parent_map.get(current) or ""
    if "CORP" not in seen:
        codes.append("CORP")
    return codes


def _binding_applies_to_scope(binding: aiosqlite.Row, scope_codes: set[str]) -> bool:
    scope_type = str(binding["scope_type"] or "").strip().upper()
    scope_code = str(binding["scope_code"] or "").strip().upper()
    if scope_type == "CORP" or scope_code == "CORP":
        return True
    return scope_code in scope_codes


def _binding_scope_rank(binding: aiosqlite.Row, scope_codes: list[str]) -> int | None:
    scope_type = str(binding["scope_type"] or "").strip().upper()
    scope_code = str(binding["scope_code"] or "").strip().upper()
    rank_by_scope = {code: idx for idx, code in enumerate(scope_codes)}
    if scope_type == "CORP" or scope_code == "CORP":
        return rank_by_scope.get("CORP", len(scope_codes))
    return rank_by_scope.get(scope_code)


def _filter_most_specific_bindings(rows: list[aiosqlite.Row], scope_codes: list[str]) -> list[aiosqlite.Row]:
    min_rank_by_metric: dict[str, int] = {}
    ranked: list[tuple[aiosqlite.Row, int]] = []
    for row in rows:
        rank = _binding_scope_rank(row, scope_codes)
        if rank is None:
            continue
        ranked.append((row, rank))
        metric_code = str(row["metric_node_code"] or "").strip().upper()
        min_rank_by_metric[metric_code] = min(rank, min_rank_by_metric.get(metric_code, rank))
    return [
        row
        for row, rank in ranked
        if rank == min_rank_by_metric.get(str(row["metric_node_code"] or "").strip().upper())
    ]




async def _load_products(db: aiosqlite.Connection) -> list[dict[str, Any]]:
    await _ensure_product_initialization_from_excel(db)
    cur = await db.execute(
        """
        SELECT product_code, product_name
        FROM product_type
        ORDER BY product_code
        LIMIT 40
        """
    )
    rows = await cur.fetchall()
    parent_map = await _load_product_parent_map(db)
    cur = await db.execute(
        """
        SELECT b.data_acct_code, b.metric_node_code, b.scope_type, b.scope_code,
               d.budget_formula
        FROM data_account_metric_binding b
        JOIN data_account d ON d.data_acct_code = b.data_acct_code
        WHERE b.is_active = 1
        """
    )
    binding_rows = await cur.fetchall()
    cur = await db.execute(
        """
        SELECT product_code,
               COUNT(*) AS component_count,
               SUM(CASE WHEN COALESCE(data_acct_code, '') <> '' AND COALESCE(formula, '') <> '' THEN 1 ELSE 0 END) AS ready_count
        FROM product_budget_component
        WHERE metric_node_code GLOB '[0-9][0-9].[0-9][0-9].[0-9][0-9][0-9]'
        GROUP BY product_code
        """
    )
    formal_counts = {str(row["product_code"]): row for row in await cur.fetchall()}
    products: list[dict[str, Any]] = []
    for row in rows:
        formal_count = formal_counts.get(str(row["product_code"]))
        if formal_count:
            component_count = int(formal_count["component_count"] or 0)
            ready_count = int(formal_count["ready_count"] or 0)
            status = "ready" if component_count and ready_count == component_count else "draft"
            products.append(
                {
                    "product_code": row["product_code"],
                    "product_name": row["product_name"],
                    "owner": "当前用户",
                    "status": status,
                    "component_count": component_count,
                }
            )
            continue
        scope_codes = _product_scope_codes(str(row["product_code"]), parent_map)
        scope_code_set = set(scope_codes)
        visible_bindings = _filter_most_specific_bindings(
            [binding for binding in binding_rows if _binding_applies_to_scope(binding, scope_code_set)],
            scope_codes,
        )
        component_count = len({str(binding["data_acct_code"]) for binding in visible_bindings})
        ready_count = sum(1 for binding in visible_bindings if str(binding["budget_formula"] or "").strip())
        status = "ready" if component_count and ready_count == component_count else "draft"
        products.append(
            {
                "product_code": row["product_code"],
                "product_name": row["product_name"],
                "owner": "当前用户",
                "status": status,
                "component_count": component_count,
            }
        )
    return products


async def _next_metric_node_code(db: aiosqlite.Connection, parent_code: str) -> tuple[str, int, int]:
    cur = await db.execute(
        """
        SELECT level
        FROM data_account_metric_node
        WHERE node_code = ?
          AND is_active = 1
        """,
        (parent_code,),
    )
    parent = await cur.fetchone()
    if not parent:
        raise HTTPException(status_code=404, detail="指标树父节点不存在")
    try:
        parent_level = int(parent["level"])
    except (TypeError, KeyError):
        parent_level = int(parent[0])
    child_level = parent_level + 1
    prefix = f"{parent_code}."
    cur = await db.execute(
        """
        SELECT node_code, sort_order
        FROM data_account_metric_node
        WHERE parent_code = ?
          AND is_active = 1
        ORDER BY sort_order DESC, node_code DESC
        """,
        (parent_code,),
    )
    rows = await cur.fetchall()
    max_seq = 0
    max_sort = 0
    width = 2
    for row in rows:
        try:
            code = str(row["node_code"])
            sort_order = int(row["sort_order"] or 0)
        except (TypeError, KeyError):
            code = str(row[0])
            sort_order = int(row[1] or 0)
        max_sort = max(max_sort, sort_order)
        if code.startswith(prefix):
            tail = code[len(prefix):]
            if re.fullmatch(r"\d+", tail):
                max_seq = max(max_seq, int(tail))
                width = max(width, len(tail))
    return f"{prefix}{max_seq + 1:0{width}d}", child_level, max_sort + 10


async def _find_or_create_metric_node(
    db: aiosqlite.Connection,
    *,
    parent_code: str,
    node_name: str,
    node_type: str,
    remark: str,
) -> tuple[str, str, bool]:
    parent_code = _safe_upper(parent_code)
    node_name = node_name.strip()
    if not parent_code:
        raise HTTPException(status_code=400, detail="请选择新增指标的上级指标节点")
    if not node_name:
        raise HTTPException(status_code=400, detail="指标名称不能为空")
    cur = await db.execute(
        """
        SELECT node_code, node_name
        FROM data_account_metric_node
        WHERE parent_code = ?
          AND node_name = ?
          AND is_active = 1
        ORDER BY node_code
        LIMIT 1
        """,
        (parent_code, node_name),
    )
    existing = await cur.fetchone()
    if existing:
        return str(existing["node_code"]), str(existing["node_name"]), False

    code, level, sort_order = await _next_metric_node_code(db, parent_code)
    clean_node_type = str(node_type or "METRIC").strip().upper()
    if clean_node_type not in {"CATEGORY", "GROUP", "METRIC"}:
        clean_node_type = "METRIC"
    await db.execute(
        """
        INSERT INTO data_account_metric_node(
          node_code, node_name, parent_code, level, node_type,
          sort_order, is_active, remark, created_at, updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, 1, ?, ?, ?)
        """,
        (
            code,
            node_name,
            parent_code,
            level,
            clean_node_type,
            sort_order,
            remark,
            _now(),
            _now(),
        ),
    )
    return code, node_name, True


async def _sync_component_formula_to_data_account(db: aiosqlite.Connection, component_id: str) -> None:
    cur = await db.execute(
        """
        SELECT c.component_id, c.product_code, c.report_acct_code, c.metric_node_code,
               c.component_name, c.rule_code, c.source_type,
               c.data_acct_code, c.formula, c.value_type, n.node_name AS metric_node_name
        FROM product_budget_component c
        LEFT JOIN data_account_metric_node n ON n.node_code = c.metric_node_code
        WHERE c.component_id = ?
        """,
        (component_id,),
    )
    component = await cur.fetchone()
    if not component or not component["data_acct_code"]:
        return
    formula = str(component["formula"] or "").strip()
    rule_config = {
        "source": "product_budget_workbench",
        "component_id": component["component_id"],
        "product_code": component["product_code"],
        "metric_node_code": component["metric_node_code"],
        "data_acct_code": component["data_acct_code"],
        "component_name": component["component_name"],
        "rule_code": component["rule_code"],
        "source_type": component["source_type"],
        "formula": formula,
    }
    await sync_budget_rule_to_data_account(
        db,
        data_acct_code=component["data_acct_code"],
        data_acct_name=component["component_name"],
        metric_group_code=component["metric_node_code"],
        metric_group_name=component["metric_node_name"] or component["component_name"],
        budget_formula=formula or None,
        budget_rule_code=component["rule_code"],
        budget_rule_config_json=json.dumps(rule_config, ensure_ascii=False),
        value_type=component["value_type"] or "金额",
        need_calc=1 if formula else 0,
    )


def _workbench_init_excel_path() -> Path:
    return settings.business_inputs_dir / "metric_tree_scheme" / "统一业务指标树全貌_20260514_全局语义编号版.xlsx"


def _value_type_from_metric(metric_code: str, name: str, path: str) -> str:
    text = f"{metric_code} {name} {path}"
    if "户数" in text or "客户数" in text:
        return "户数"
    if metric_code.startswith("02.") or any(token in text for token in ("率", "占比", "比例", "FTP")):
        return "百分比"
    return "金额"


async def _ensure_product_initialization_from_excel(db: aiosqlite.Connection) -> None:
    excel_path = _workbench_init_excel_path()
    if not excel_path.exists():
        logger.warning("[product_workbench.init_import] missing excel=%s", excel_path)
        return

    try:
        from openpyxl import load_workbook
    except Exception:
        logger.exception("[product_workbench.init_import] openpyxl unavailable")
        return

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    if "产品工作台初始化_分产品" not in wb.sheetnames:
        logger.warning("[product_workbench.init_import] sheet missing excel=%s", excel_path)
        return
    ws = wb["产品工作台初始化_分产品"]
    headers = [str(cell or "").strip() for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    index = {name: idx for idx, name in enumerate(headers)}
    required = ["范围代码", "新系统指标编码", "标准指标名称", "Excel展示路径"]
    if any(name not in index for name in required):
        logger.warning("[product_workbench.init_import] required columns missing excel=%s", excel_path)
        return

    rows_by_metric: dict[tuple[str, str], dict[str, Any]] = {}
    for sort_order, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        product_code = _safe_upper(str(row[index["范围代码"]] or ""))
        metric_code = str(row[index["新系统指标编码"]] or "").strip()
        if not product_code or not _is_formal_metric_code(metric_code):
            continue
        name = str(row[index["标准指标名称"]] or "").strip() or "数据科目"
        display_path = str(row[index["Excel展示路径"]] or "").strip()
        old_code = str(row[index.get("旧系统指标编码", -1)] or "").strip() if "旧系统指标编码" in index else ""
        source_code = str(row[index.get("源Excel代码", -1)] or "").strip() if "源Excel代码" in index else ""
        key = (product_code, metric_code)
        bucket = rows_by_metric.setdefault(
            key,
            {
                "product_code": product_code,
                "metric_code": metric_code,
                "name": name,
                "display_path": display_path,
                "sort_order": sort_order,
                "sources": [],
            },
        )
        bucket["sort_order"] = min(int(bucket["sort_order"]), sort_order)
        bucket["sources"].append(
            {
                "old_metric_node_code": old_code,
                "source_excel_code": source_code,
                "display_path": display_path,
                "source_row": sort_order + 1,
            }
        )

    inserted = 0
    for row in rows_by_metric.values():
        product_code = row["product_code"]
        metric_code = row["metric_code"]
        component_name = str(row["name"] or "数据科目").strip() or "数据科目"
        report_code = f"M{product_code}{metric_code.replace('.', '')}"
        component_id = _new_id("pbc")
        ai_reason = json.dumps(
            {
                "source": "产品工作台初始化_分产品",
                "formal_metric_node_code": metric_code,
                "merged_source_count": len(row["sources"]),
                "sources": row["sources"],
            },
            ensure_ascii=False,
        )
        await db.execute(
            """
            INSERT INTO product_budget_component(
              component_id, product_code, report_acct_code, metric_node_code,
              component_name, rule_code, source_type, formula, value_type,
              status, ai_reason, sort_order
            ) VALUES (?, ?, ?, ?, ?, 'formula', 'product_initialization', NULL, ?, 'warning', ?, ?)
            ON CONFLICT(product_code, metric_node_code)
            WHERE COALESCE(metric_node_code, '') <> ''
              AND metric_node_code GLOB '[0-9][0-9].[0-9][0-9].[0-9][0-9][0-9]'
            DO UPDATE SET
              report_acct_code = excluded.report_acct_code,
              component_name = excluded.component_name,
              source_type = excluded.source_type,
              value_type = excluded.value_type,
              ai_reason = excluded.ai_reason,
              sort_order = excluded.sort_order,
              updated_at = CURRENT_TIMESTAMP
            """,
            (
                component_id,
                product_code,
                report_code,
                metric_code,
                component_name,
                _value_type_from_metric(metric_code, component_name, row["display_path"]),
                ai_reason,
                int(row["sort_order"]),
            ),
        )
        await db.execute(
            """
            UPDATE data_account
            SET data_acct_name = ?,
                metric_group_code = ?,
                metric_group_name = COALESCE((SELECT node_name FROM data_account_metric_node WHERE node_code = ?), ?),
                value_type = ?
            WHERE data_acct_code IN (
              SELECT data_acct_code
              FROM product_budget_component
              WHERE product_code = ?
                AND metric_node_code = ?
                AND COALESCE(data_acct_code, '') <> ''
            )
            """,
            (
                component_name,
                metric_code,
                metric_code,
                component_name,
                _value_type_from_metric(metric_code, component_name, row["display_path"]),
                product_code,
                metric_code,
            ),
        )
        inserted += 1
    logger.info("[product_workbench.init_import] ensured formal product initialization rows=%s excel=%s", inserted, excel_path)


async def _ensure_default_components(db: aiosqlite.Connection, product_code: str) -> None:
    await _ensure_product_initialization_from_excel(db)

    cur = await db.execute(
        """
        SELECT COUNT(*)
        FROM product_budget_component
        WHERE product_code = ?
          AND metric_node_code GLOB '[0-9][0-9].[0-9][0-9].[0-9][0-9][0-9]'
        """,
        (product_code,),
    )
    if int((await cur.fetchone())[0] or 0) > 0:
        return

    parent_map = await _load_product_parent_map(db)
    scope_codes = _product_scope_codes(product_code, parent_map)
    scope_placeholders = ",".join(["?"] * len(scope_codes))

    cur = await db.execute(
        f"""
        SELECT b.metric_node_code, b.data_acct_code, b.scope_type, b.scope_code,
               n.node_name, d.data_acct_code, d.data_acct_name, d.budget_formula,
               d.value_type, b.sort_order
        FROM data_account_metric_binding b
        JOIN data_account_metric_node n ON n.node_code = b.metric_node_code
        JOIN data_account d ON d.data_acct_code = b.data_acct_code
        WHERE b.is_active = 1
          AND (
            b.scope_code = 'CORP'
            OR UPPER(COALESCE(b.scope_code, '')) IN ({scope_placeholders})
          )
        ORDER BY COALESCE(n.sort_order, 999999), b.metric_node_code, b.sort_order, b.scope_code, b.data_acct_code
        """,
        (*scope_codes,),
    )
    rows = _filter_most_specific_bindings(await cur.fetchall(), scope_codes)
    visible_data_codes = {str(row["data_acct_code"]) for row in rows}
    if visible_data_codes:
        placeholders = ",".join(["?"] * len(visible_data_codes))
        await db.execute(
            f"""
            DELETE FROM product_budget_component
            WHERE product_code = ?
              AND COALESCE(metric_node_code, '') <> ''
              AND COALESCE(data_acct_code, '') <> ''
              AND COALESCE(data_acct_code, '') NOT IN ({placeholders})
            """,
            (product_code, *visible_data_codes),
        )

    for idx, row in enumerate(rows):
        report_code = _metric_node_compat_code(row["metric_node_code"])
        formula = row["budget_formula"]
        data_code = row["data_acct_code"]
        status = _status_from_formula_and_binding(formula, data_code)
        trial = _synthetic_trial(f"{product_code}-{row['metric_node_code']}-{data_code}", negative="成本" in str(row["node_name"]))
        cur_existing = await db.execute(
            """
            SELECT component_id
            FROM product_budget_component
            WHERE product_code = ?
              AND data_acct_code = ?
            LIMIT 1
            """,
            (product_code, data_code),
        )
        existing = await cur_existing.fetchone()
        if existing:
            await db.execute(
                """
                UPDATE product_budget_component
                SET report_acct_code = ?, metric_node_code = ?, component_name = ?,
                    template_name = '从现有数据科目引用', rule_code = 'formula',
                    source_type = 'existing_data_account', formula = ?, value_type = ?,
                    status = ?, trial_m01 = COALESCE(trial_m01, ?),
                    trial_m02 = COALESCE(trial_m02, ?), trial_m03 = COALESCE(trial_m03, ?),
                    trial_annual = COALESCE(trial_annual, ?), sort_order = ?, updated_at = ?
                WHERE component_id = ?
                """,
                (
                    report_code,
                    row["metric_node_code"],
                    str(row["data_acct_name"] or row["node_name"] or "数据科目"),
                    formula,
                    row["value_type"] or "金额",
                    status,
                    trial["trial_m01"],
                    trial["trial_m02"],
                    trial["trial_m03"],
                    trial["trial_annual"],
                    idx,
                    _now(),
                    existing["component_id"],
                ),
            )
            continue
        component_id = _new_id("pbc")
        await db.execute(
            """
            INSERT INTO product_budget_component(
              component_id, product_code, report_acct_code, metric_node_code, component_name,
              template_name, rule_code, source_type, data_acct_code, formula,
              value_type, status, trial_m01, trial_m02, trial_m03, trial_annual, sort_order
            ) VALUES (?, ?, ?, ?, ?, ?, 'formula', 'existing_data_account', ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                component_id,
                product_code,
                report_code,
                row["metric_node_code"],
                str(row["data_acct_name"] or row["node_name"] or "预算组件"),
                "从现有数据科目引用",
                data_code,
                formula,
                row["value_type"] or "金额",
                status,
                trial["trial_m01"],
                trial["trial_m02"],
                trial["trial_m03"],
                trial["trial_annual"],
                idx,
            ),
        )


async def _load_templates(db: aiosqlite.Connection) -> list[dict[str, Any]]:
    cur = await db.execute(
        """
        SELECT template_id, template_name, component_name, rule_code, source_type,
               formula, value_type, data_acct_code, source_component_id, updated_at
        FROM product_budget_component_template
        ORDER BY updated_at DESC
        LIMIT 50
        """
    )
    return [dict(row) for row in await cur.fetchall()]


async def _load_rows(db: aiosqlite.Connection, product_code: str) -> list[dict[str, Any]]:
    cur = await db.execute(
        """
        SELECT c.report_acct_code,
               COALESCE(n.node_name, c.component_name) AS report_acct_name,
               n.parent_code AS parent_code,
               COALESCE(n.level, 1) AS level,
               COALESCE(c.metric_node_code, c.report_acct_code) AS row_code,
               n.node_name AS row_name,
               n.sort_order AS metric_sort_order,
               c.component_id, c.component_name, c.template_id, c.template_name,
               c.rule_code, c.source_type, c.metric_node_code, n.node_name AS metric_node_name,
               c.data_acct_code, d.data_acct_name,
               d.budget_formula AS data_budget_formula, d.value_type AS data_value_type,
               b.scope_type AS binding_scope_type, b.scope_code AS binding_scope_code,
               b.is_active AS binding_is_active,
               c.formula, c.value_type, c.status, c.ai_reason,
               c.trial_m01, c.trial_m02, c.trial_m03, c.trial_annual,
               c.updated_at, c.sort_order
        FROM product_budget_component c
        LEFT JOIN data_account d ON d.data_acct_code = c.data_acct_code
        LEFT JOIN data_account_metric_node n ON n.node_code = c.metric_node_code
        LEFT JOIN data_account_metric_binding b ON b.data_acct_code = c.data_acct_code
        WHERE c.product_code = ?
          AND COALESCE(c.metric_node_code, '') <> ''
        ORDER BY
          CASE WHEN c.metric_node_code IS NULL THEN 1 ELSE 0 END,
          n.sort_order,
          c.metric_node_code,
          c.report_acct_code,
          c.sort_order,
          c.created_at
        """,
        (product_code,),
    )
    fetched_rows = await cur.fetchall()
    parent_map = await _load_product_parent_map(db)
    scope_codes = _product_scope_codes(product_code, parent_map)
    min_rank_by_metric: dict[str, int] = {}
    ranked_rows: list[tuple[aiosqlite.Row, int]] = []
    formal_rows = [row for row in fetched_rows if _is_formal_metric_code(row["metric_node_code"])]
    if formal_rows:
        for row in formal_rows:
            metric_code = str(row["metric_node_code"] or "").strip().upper()
            min_rank_by_metric[metric_code] = 0
            ranked_rows.append((row, 0))
    else:
        for row in fetched_rows:
            if int(row["binding_is_active"] or 0) != 1:
                continue
            binding_view = {
                "metric_node_code": row["metric_node_code"],
                "scope_type": row["binding_scope_type"],
                "scope_code": row["binding_scope_code"],
            }
            rank = _binding_scope_rank(binding_view, scope_codes)
            if rank is None:
                continue
            metric_code = str(row["metric_node_code"] or "").strip().upper()
            min_rank_by_metric[metric_code] = min(rank, min_rank_by_metric.get(metric_code, rank))
            ranked_rows.append((row, rank))
    rows_by_code: dict[str, dict[str, Any]] = {}
    for row, rank in ranked_rows:
        metric_code = str(row["metric_node_code"] or "").strip().upper()
        if rank != min_rank_by_metric.get(metric_code):
            continue
        code = str(row["row_code"])
        bucket = rows_by_code.setdefault(
            code,
            {
                "report_acct_code": code,
                "report_acct_name": row["row_name"] or row["report_acct_name"],
                "metric_node_code": row["metric_node_code"],
                "metric_node_name": row["metric_node_name"],
                "compat_report_acct_code": row["report_acct_code"],
                "parent_code": row["parent_code"],
                "level": _metric_display_level(row["metric_node_code"], int(row["level"] or 1)),
                "row_type": "metric" if row["metric_node_code"] else "compat",
                "component_count": 0,
                "_sort_order": int(row["metric_sort_order"] or 999999),
                "status": "draft",
                "trial_m01": 0.0,
                "trial_m02": 0.0,
                "trial_m03": 0.0,
                "trial_annual": 0.0,
                "components": [],
            },
        )
        trial = {
            "trial_m01": row["trial_m01"],
            "trial_m02": row["trial_m02"],
            "trial_m03": row["trial_m03"],
            "trial_annual": row["trial_annual"],
        }
        if any(v is None for v in trial.values()):
            trial = _synthetic_trial(f"{product_code}-{row['component_id']}", negative="成本" in str(row["report_acct_name"]))
        formula = row["data_budget_formula"] or row["formula"] or ""
        value_type = row["data_value_type"] or row["value_type"] or "金额"
        status = _status_from_formula_and_binding(formula, row["data_acct_code"])
        component = {
            "component_id": row["component_id"],
            "component_name": row["data_acct_name"] or row["component_name"],
            "template_id": row["template_id"],
            "template_name": row["template_name"],
            "rule_code": row["rule_code"],
            "rule_label": RULE_LABELS.get(str(row["rule_code"]), str(row["rule_code"])),
            "source_type": row["source_type"],
            "metric_node_code": row["metric_node_code"],
            "metric_node_name": row["metric_node_name"],
            "data_acct_code": row["data_acct_code"],
            "data_acct_name": row["data_acct_name"],
            "formula": formula,
            "value_type": value_type,
            "status": status,
            "ai_reason": row["ai_reason"],
            "updated_at": row["updated_at"],
            **trial,
        }
        bucket["components"].append(component)
        bucket["component_count"] += 1
        for key in ("trial_m01", "trial_m02", "trial_m03", "trial_annual"):
            bucket[key] += float(trial[key] or 0)
        statuses = {str(item["status"]) for item in bucket["components"]}
        bucket["status"] = "warning" if "warning" in statuses else "ready" if statuses <= {"ready", "dispatched"} else "draft"
    if rows_by_code:
        cur = await db.execute(
            """
            SELECT node_code, node_name, parent_code, level, sort_order
            FROM data_account_metric_node
            WHERE is_active = 1
            """
        )
        node_map = {str(node["node_code"]): node for node in await cur.fetchall()}
        status_rank = {"warning": 3, "draft": 2, "ready": 1, "dispatched": 1}
        for leaf in list(rows_by_code.values()):
            metric_code = leaf.get("metric_node_code")
            if not metric_code:
                continue
            parent_code = leaf.get("parent_code")
            while parent_code and parent_code != "00":
                node = node_map.get(str(parent_code))
                if not node:
                    break
                parent = rows_by_code.setdefault(
                    str(node["node_code"]),
                    {
                        "report_acct_code": node["node_code"],
                        "report_acct_name": node["node_name"],
                        "metric_node_code": node["node_code"],
                        "metric_node_name": node["node_name"],
                        "compat_report_acct_code": None,
                        "parent_code": None if node["parent_code"] == "00" else node["parent_code"],
                        "level": _metric_display_level(node["node_code"], int(node["level"] or 1)),
                        "row_type": "group",
                        "component_count": 0,
                        "_sort_order": int(node["sort_order"] or 0),
                        "status": "draft",
                        "trial_m01": 0.0,
                        "trial_m02": 0.0,
                        "trial_m03": 0.0,
                        "trial_annual": 0.0,
                        "components": [],
                    },
                )
                parent["component_count"] += int(leaf["component_count"] or len(leaf["components"]))
                for key in ("trial_m01", "trial_m02", "trial_m03", "trial_annual"):
                    parent[key] += float(leaf[key] or 0)
                if status_rank.get(str(leaf["status"]), 0) > status_rank.get(str(parent["status"]), 0):
                    parent["status"] = leaf["status"]
                parent_code = node["parent_code"]

    children_by_parent: dict[str | None, list[dict[str, Any]]] = {}
    for row in rows_by_code.values():
        parent_code = row.get("parent_code")
        parent_key = str(parent_code) if parent_code and str(parent_code) != "00" else None
        children_by_parent.setdefault(parent_key, []).append(row)

    def sort_key(item: dict[str, Any]) -> tuple[str]:
        return (str(item["report_acct_code"]),)

    for children in children_by_parent.values():
        children.sort(key=sort_key)

    rows: list[dict[str, Any]] = []
    visited: set[str] = set()

    def walk(items: list[dict[str, Any]]) -> None:
        for item in items:
            code = str(item["report_acct_code"])
            if code in visited:
                continue
            visited.add(code)
            rows.append(item)
            walk(children_by_parent.get(code, []))

    roots = [
        row
        for row in children_by_parent.get(None, [])
        if not row.get("parent_code") or str(row.get("parent_code")) not in rows_by_code
    ]
    if not roots:
        roots = children_by_parent.get(None, [])
    walk(sorted(roots, key=sort_key))
    remaining = [row for code, row in rows_by_code.items() if code not in visited]
    walk(sorted(remaining, key=sort_key))
    for row in rows:
        row.pop("_sort_order", None)
    return rows


def build_product_budget_workbench_router(deepseek_client: DeepseekClient | None = None) -> APIRouter:
    router = APIRouter(prefix="/api/product-budget-workbench", tags=["product-budget-workbench"])

    @router.get("/overview")
    async def overview(product_code: str | None = Query(default=None)):
        async with aiosqlite.connect(common_db_path()) as db:
            db.row_factory = aiosqlite.Row
            await db.execute("PRAGMA foreign_keys = ON")
            await _ensure_runtime_tables(db)
            products = await _load_products(db)
            if not products:
                return {"products": [], "selected_product_code": None, "rows": [], "templates": [], "summary": {}}
            selected = _safe_upper(product_code) or str(products[0]["product_code"])
            if selected not in {str(p["product_code"]) for p in products}:
                selected = str(products[0]["product_code"])
            await _ensure_package(db, selected)
            await _ensure_default_components(db, selected)
            await db.commit()
            products = await _load_products(db)
            rows = await _load_rows(db, selected)
            templates = await _load_templates(db)
        component_count = sum(len(row["components"]) for row in rows)
        ready_count = sum(1 for row in rows for c in row["components"] if c["status"] in {"ready", "dispatched"})
        warning_count = sum(1 for row in rows for c in row["components"] if c["status"] == "warning")
        metric_row_count = sum(1 for row in rows if row.get("row_type") in {"metric", "compat"})
        return {
            "products": products,
            "selected_product_code": selected,
            "rows": rows,
            "templates": templates,
            "rule_options": [{"code": code, "label": label} for code, label in RULE_LABELS.items()],
            "summary": {
                "report_row_count": metric_row_count,
                "component_count": component_count,
                "ready_count": ready_count,
                "warning_count": warning_count,
                "trial_annual": round(sum(float(row["trial_annual"] or 0) for row in rows), 2),
            },
        }

    @router.post("/components")
    async def create_component(body: WorkbenchCreateComponent):
        product_code = _safe_upper(body.product_code)
        report_code = _safe_upper(body.report_acct_code)
        if not product_code or not report_code:
            raise HTTPException(status_code=400, detail="产品和指标节点不能为空")
        component_id = _new_id("pbc")
        async with aiosqlite.connect(common_db_path()) as db:
            db.row_factory = aiosqlite.Row
            await db.execute("PRAGMA foreign_keys = ON")
            await _ensure_runtime_tables(db)
            await _ensure_package(db, product_code)
            metric_node_code = _safe_upper(body.metric_node_code) or None
            if metric_node_code:
                report_code = _metric_node_compat_code(metric_node_code, report_code)
            component_name = body.component_name.strip() or "新数据科目"
            if metric_node_code:
                cur = await db.execute(
                    """
                    SELECT component_id
                    FROM product_budget_component
                    WHERE product_code = ?
                      AND metric_node_code = ?
                    ORDER BY created_at, component_id
                    LIMIT 1
                    """,
                    (product_code, metric_node_code),
                )
                existing_metric = await cur.fetchone()
                if existing_metric:
                    return {"component_id": str(existing_metric["component_id"]), "created": False}
            cur = await db.execute(
                """
                SELECT component_id
                FROM product_budget_component
                WHERE product_code = ?
                  AND component_name = ?
                  AND COALESCE(data_acct_code, '') = ''
                  AND (
                    (? IS NOT NULL AND metric_node_code = ?)
                    OR (? IS NULL AND report_acct_code = ?)
                  )
                ORDER BY created_at, component_id
                LIMIT 1
                """,
                (
                    product_code,
                    component_name,
                    metric_node_code,
                    metric_node_code,
                    metric_node_code,
                    report_code,
                ),
            )
            existing = await cur.fetchone()
            if existing:
                return {"component_id": str(existing["component_id"]), "created": False}
            trial = _synthetic_trial(f"{product_code}-{report_code}-{component_name}")
            await db.execute(
                """
                INSERT INTO product_budget_component(
                  component_id, product_code, report_acct_code, metric_node_code, component_name,
                  rule_code, source_type, status, trial_m01, trial_m02, trial_m03, trial_annual
                ) VALUES (?, ?, ?, ?, ?, 'formula', 'manual', 'warning', ?, ?, ?, ?)
                """,
                (
                    component_id,
                    product_code,
                    report_code,
                    metric_node_code,
                    component_name,
                    trial["trial_m01"],
                    trial["trial_m02"],
                    trial["trial_m03"],
                    trial["trial_annual"],
                ),
            )
            await db.commit()
        return {"component_id": component_id, "created": True}

    @router.put("/components/{component_id}")
    async def update_component(component_id: str, body: WorkbenchComponentUpsert):
        allowed = {
            "component_name",
            "template_id",
            "template_name",
            "metric_node_code",
            "rule_code",
            "source_type",
            "data_acct_code",
            "formula",
            "value_type",
            "status",
        }
        updates: list[str] = []
        values: list[Any] = []
        for field in allowed:
            if field in body.model_fields_set:
                updates.append(f"{field} = ?")
                value = getattr(body, field)
                values.append(_safe_upper(value) if field in {"data_acct_code", "metric_node_code"} and value else value)
        if "status" not in body.model_fields_set and ("formula" in body.model_fields_set or "data_acct_code" in body.model_fields_set):
            formula = body.formula if "formula" in body.model_fields_set else None
            data_code = body.data_acct_code if "data_acct_code" in body.model_fields_set else None
            updates.append("status = COALESCE(?, status)")
            values.append(_status_from_formula_and_binding(formula, data_code))
        if not updates:
            return {"updated": False}
        updates.append("updated_at = ?")
        values.append(_now())
        values.append(component_id)
        async with aiosqlite.connect(common_db_path()) as db:
            db.row_factory = aiosqlite.Row
            await _ensure_runtime_tables(db)
            cur = await db.execute("SELECT 1 FROM product_budget_component WHERE component_id = ?", (component_id,))
            if not await cur.fetchone():
                raise HTTPException(status_code=404, detail="组件不存在")
            await db.execute(f"UPDATE product_budget_component SET {', '.join(updates)} WHERE component_id = ?", values)
            if {"formula", "rule_code", "value_type", "component_name"} & set(body.model_fields_set):
                await _sync_component_formula_to_data_account(db, component_id)
            await db.commit()
        return {"updated": True}

    @router.post("/metric-data-accounts")
    async def create_metric_data_account(body: WorkbenchCreateMetricDataAccount):
        product_code = _safe_upper(body.product_code)
        parent_code = _safe_upper(body.parent_metric_node_code)
        metric_name = body.metric_node_name.strip()
        if not product_code:
            raise HTTPException(status_code=400, detail="产品不能为空")
        if not parent_code:
            raise HTTPException(status_code=400, detail="请选择新增指标的上级指标节点")
        if not metric_name:
            raise HTTPException(status_code=400, detail="指标名称不能为空")

        formula = _clean_formula(body.formula)
        value_type = _clean_value_type(body.value_type)
        rule_code = _clean_rule_code(body.rule_code)
        source_type = _clean_source_type(body.source_type or "manual_metric_create")
        component_name = (body.component_name or metric_name).strip() or metric_name
        logger.info(
            "[product_workbench.metric_create] start product=%s parent=%s metric_name=%s value_type=%s has_formula=%s",
            product_code,
            parent_code,
            metric_name,
            value_type,
            bool(formula),
        )

        async with aiosqlite.connect(common_db_path()) as db:
            db.row_factory = aiosqlite.Row
            await db.execute("PRAGMA foreign_keys = ON")
            await _ensure_runtime_tables(db)
            await _ensure_package(db, product_code)
            cur = await db.execute("SELECT product_name FROM product_type WHERE product_code = ?", (product_code,))
            product = await cur.fetchone()
            if not product:
                raise HTTPException(status_code=404, detail="产品不存在")
            node_code, node_name, node_created = await _find_or_create_metric_node(
                db,
                parent_code=parent_code,
                node_name=metric_name,
                node_type=body.node_type,
                remark=f"由产品预算工作台新增：product_code={product_code}",
            )

            cur = await db.execute(
                """
                SELECT data_acct_code
                FROM data_account_metric_binding
                WHERE metric_node_code = ?
                  AND scope_code = ?
                  AND is_active = 1
                LIMIT 1
                """,
                (node_code, product_code),
            )
            binding = await cur.fetchone()
            data_created = False
            if binding:
                data_code = str(binding["data_acct_code"])
            else:
                data_created = True
                rule_config = {
                    "source": "product_budget_workbench_metric_create",
                    "product_code": product_code,
                    "metric_node_code": node_code,
                    "metric_node_name": node_name,
                    "parent_metric_node_code": parent_code,
                    "component_name": component_name,
                    "rule_code": rule_code,
                    "source_type": source_type,
                    "formula": formula,
                }
                created = await ensure_product_metric_data_account(
                    db,
                    product_code=product_code,
                    metric_node_code=node_code,
                    data_acct_name=component_name,
                    metric_group_name=node_name,
                    budget_formula=formula or None,
                    budget_rule_code=rule_code,
                    budget_rule_config=rule_config,
                    need_calc=1 if formula else 0,
                    value_type=value_type,
                    remark=f"由产品预算工作台自动生成：metric_node_code={node_code}; product_code={product_code}",
                    binding_remark=f"由产品预算工作台新增指标自动绑定：{component_name}",
                )
                data_code = created.data_acct_code

            cur = await db.execute(
                """
                SELECT component_id
                FROM product_budget_component
                WHERE product_code = ?
                  AND metric_node_code = ?
                  AND COALESCE(metric_node_code, '') <> ''
                LIMIT 1
                """,
                (product_code, node_code),
            )
            existing_component = await cur.fetchone()
            status = _status_from_formula_and_binding(formula, data_code)
            if existing_component:
                component_id = str(existing_component["component_id"])
                await db.execute(
                    """
                    UPDATE product_budget_component
                    SET report_acct_code = ?, component_name = ?, formula = ?, rule_code = ?,
                        source_type = ?, value_type = ?, status = ?, updated_at = ?
                    WHERE component_id = ?
                    """,
                    (
                        _metric_node_compat_code(node_code),
                        component_name,
                        formula,
                        rule_code,
                        source_type,
                        value_type,
                        status,
                        _now(),
                        component_id,
                    ),
                )
                component_created = False
            else:
                component_id = _new_id("pbc")
                trial = _synthetic_trial(f"{product_code}-{node_code}-{component_name}", negative="成本" in node_name)
                await db.execute(
                    """
                    INSERT INTO product_budget_component(
                      component_id, product_code, report_acct_code, metric_node_code,
                      component_name, rule_code, source_type,
                      data_acct_code, formula, value_type, status,
                      trial_m01, trial_m02, trial_m03, trial_annual
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        component_id,
                        product_code,
                        _metric_node_compat_code(node_code),
                        node_code,
                        component_name,
                        rule_code,
                        source_type,
                        data_code,
                        formula,
                        value_type,
                        status,
                        trial["trial_m01"],
                        trial["trial_m02"],
                        trial["trial_m03"],
                        trial["trial_annual"],
                    ),
                )
                component_created = True

            await _sync_component_formula_to_data_account(db, component_id)
            await db.commit()
        logger.info(
            "[product_workbench.metric_create] done product=%s parent=%s metric_node=%s data_account=%s component=%s node_created=%s data_created=%s component_created=%s",
            product_code,
            parent_code,
            node_code,
            data_code,
            component_id,
            node_created,
            data_created,
            component_created,
        )
        return {
            "component_id": component_id,
            "metric_node_code": node_code,
            "metric_node_name": node_name,
            "metric_node_created": node_created,
            "data_acct_code": data_code,
            "data_account_created": data_created,
            "component_created": component_created,
        }

    @router.post("/components/{component_id}/create-data-account")
    async def create_data_account_for_component(component_id: str, body: WorkbenchCreateDataAccount):
        async with aiosqlite.connect(common_db_path()) as db:
            db.row_factory = aiosqlite.Row
            await db.execute("PRAGMA foreign_keys = ON")
            await _ensure_runtime_tables(db)
            cur = await db.execute(
                """
                SELECT c.*, COALESCE(n.node_name, c.component_name) AS report_acct_name
                FROM product_budget_component c
                LEFT JOIN data_account_metric_node n ON n.node_code = c.metric_node_code
                WHERE c.component_id = ?
                """,
                (component_id,),
            )
            component = await cur.fetchone()
            if not component:
                raise HTTPException(status_code=404, detail="组件不存在")
            existing = component["data_acct_code"]
            if existing:
                await _sync_component_formula_to_data_account(db, component_id)
                await db.commit()
                return {"data_acct_code": existing, "created": False}
            if component["metric_node_code"]:
                cur = await db.execute(
                    """
                    SELECT data_acct_code
                    FROM data_account_metric_binding
                    WHERE metric_node_code = ?
                      AND scope_code = ?
                      AND is_active = 1
                    LIMIT 1
                    """,
                    (component["metric_node_code"], component["product_code"]),
                )
                binding = await cur.fetchone()
                if binding:
                    await db.execute(
                        """
                        UPDATE product_budget_component
                        SET data_acct_code = ?, status = ?, updated_at = ?
                        WHERE component_id = ?
                        """,
                        (
                            binding["data_acct_code"],
                            _status_from_formula_and_binding(component["formula"], binding["data_acct_code"]),
                            _now(),
                            component_id,
                        ),
                    )
                    await _sync_component_formula_to_data_account(db, component_id)
                    await db.commit()
                    return {"data_acct_code": binding["data_acct_code"], "created": False}
            rule_config = {
                "source": "product_budget_workbench",
                "component_id": component_id,
                "product_code": component["product_code"],
                "compat_report_acct_code": component["report_acct_code"],
                "metric_node_code": component["metric_node_code"],
                "component_name": component["component_name"],
                "rule_code": component["rule_code"],
                "source_type": component["source_type"],
            }
            metric_node_code = str(component["metric_node_code"] or "").strip()
            if not _is_formal_metric_code(metric_node_code):
                raise HTTPException(
                    status_code=409,
                    detail="产品预算配置台不再单独生成临时数据科目。请先在数据科目维护页选择正式指标节点和产品后生成唯一指标号码。",
                )
            created = await ensure_product_metric_data_account(
                db,
                product_code=component["product_code"],
                metric_node_code=metric_node_code,
                data_acct_name=body.data_acct_name or component["component_name"],
                metric_group_name=body.metric_group_name or component["report_acct_name"],
                budget_formula=component["formula"],
                budget_rule_code=component["rule_code"],
                budget_rule_config=rule_config,
                need_calc=1,
                value_type=component["value_type"] or "金额",
                remark=f"由产品预算工作台生成：metric_node_code={metric_node_code}; component={component['component_name']}",
                binding_remark=f"由产品预算工作台生成：{component['component_name']}",
            )
            data_code = created.data_acct_code
            await db.execute(
                """
                UPDATE product_budget_component
                SET data_acct_code = ?, status = ?, updated_at = ?
                WHERE component_id = ?
                """,
                (
                    data_code,
                    _status_from_formula_and_binding(component["formula"], data_code),
                    _now(),
                    component_id,
                ),
            )
            await _sync_component_formula_to_data_account(db, component_id)
            await db.commit()
        return {"data_acct_code": data_code, "created": True}

    @router.post("/components/{component_id}/confirm-metric-node")
    async def confirm_metric_node_for_component(component_id: str, body: WorkbenchConfirmMetricNode):
        suggested_name = body.suggested_name.strip()
        if not suggested_name:
            raise HTTPException(status_code=400, detail="请提供要新增的指标节点名称")
        parent_code = _safe_upper(body.parent_code) or "09.01"
        node_type = str(body.node_type or "METRIC").strip().upper() or "METRIC"
        async with aiosqlite.connect(common_db_path()) as db:
            db.row_factory = aiosqlite.Row
            await db.execute("PRAGMA foreign_keys = ON")
            await _ensure_runtime_tables(db)
            cur = await db.execute(
                "SELECT * FROM product_budget_component WHERE component_id = ?",
                (component_id,),
            )
            component = await cur.fetchone()
            if not component:
                raise HTTPException(status_code=404, detail="组件不存在")
            if component["metric_node_code"]:
                node_code = component["metric_node_code"]
                cur = await db.execute(
                    "SELECT node_name FROM data_account_metric_node WHERE node_code = ?",
                    (node_code,),
                )
                existing_node = await cur.fetchone()
                return {
                    "metric_node_code": node_code,
                    "metric_node_name": existing_node["node_name"] if existing_node else node_code,
                    "created": False,
                }

            cur = await db.execute(
                """
                SELECT node_code, node_name
                FROM data_account_metric_node
                WHERE node_name = ?
                  AND is_active = 1
                ORDER BY level DESC, node_code
                LIMIT 1
                """,
                (suggested_name,),
            )
            existing = await cur.fetchone()
            created = False
            if existing:
                node_code = existing["node_code"]
                metric_node_name = existing["node_name"]
            else:
                node_code, level, sort_order = await _next_metric_node_code(db, parent_code)
                metric_node_name = suggested_name
                await db.execute(
                    """
                    INSERT INTO data_account_metric_node(
                      node_code, node_name, parent_code, level, node_type,
                      sort_order, is_active, remark, created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, 1, ?, ?, ?)
                    """,
                    (
                        node_code,
                        metric_node_name,
                        parent_code,
                        level,
                        node_type,
                        sort_order,
                        f"由产品预算工作台确认新增：component_id={component_id}",
                        _now(),
                        _now(),
                    ),
                )
                created = True

            report_code = _metric_node_compat_code(node_code)
            await db.execute(
                """
                UPDATE product_budget_component
                SET metric_node_code = ?, report_acct_code = ?, status = ?, updated_at = ?
                WHERE component_id = ?
                """,
                (
                    node_code,
                    report_code,
                    _status_from_formula_and_binding(component["formula"], component["data_acct_code"]),
                    _now(),
                    component_id,
                ),
            )
            await db.commit()
        return {
            "metric_node_code": node_code,
            "metric_node_name": metric_node_name,
            "created": created,
        }

    @router.post("/components/{component_id}/save-template")
    async def save_template(component_id: str, body: WorkbenchTemplateCreate):
        template_id = _new_id("pbt")
        async with aiosqlite.connect(common_db_path()) as db:
            db.row_factory = aiosqlite.Row
            await _ensure_runtime_tables(db)
            cur = await db.execute("SELECT * FROM product_budget_component WHERE component_id = ?", (component_id,))
            component = await cur.fetchone()
            if not component:
                raise HTTPException(status_code=404, detail="组件不存在")
            await db.execute(
                """
                INSERT INTO product_budget_component_template(
                  template_id, template_name, component_name, rule_code, source_type,
                  formula, value_type, data_acct_code, source_component_id
                ) VALUES (?, ?, ?, ?, 'template', ?, ?, ?, ?)
                """,
                (
                    template_id,
                    body.template_name.strip() or f"{component['component_name']}模板",
                    component["component_name"],
                    component["rule_code"],
                    component["formula"],
                    component["value_type"],
                    component["data_acct_code"],
                    component_id,
                ),
            )
            await db.commit()
        return {"template_id": template_id}

    @router.post("/templates/{template_id}/apply")
    async def apply_template(template_id: str, body: WorkbenchTemplateApply):
        component_id = _new_id("pbc")
        async with aiosqlite.connect(common_db_path()) as db:
            db.row_factory = aiosqlite.Row
            await db.execute("PRAGMA foreign_keys = ON")
            await _ensure_runtime_tables(db)
            cur = await db.execute("SELECT * FROM product_budget_component_template WHERE template_id = ?", (template_id,))
            template = await cur.fetchone()
            if not template:
                raise HTTPException(status_code=404, detail="模板不存在")
            product_code = _safe_upper(body.product_code)
            metric_node_code = _safe_upper(body.metric_node_code) or None
            report_code = _safe_upper(body.report_acct_code)
            if metric_node_code:
                report_code = _metric_node_compat_code(metric_node_code, report_code)
            trial = _synthetic_trial(f"{product_code}-{report_code}-{template_id}")
            await db.execute(
                """
                INSERT INTO product_budget_component(
                  component_id, product_code, report_acct_code, metric_node_code, component_name,
                  template_id, template_name, template_source_component_id,
                  rule_code, source_type, data_acct_code, formula, value_type,
                  status, trial_m01, trial_m02, trial_m03, trial_annual
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'template_copy', ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    component_id,
                    product_code,
                    report_code,
                    metric_node_code,
                    template["component_name"],
                    template_id,
                    template["template_name"],
                    template["source_component_id"],
                    template["rule_code"],
                    template["data_acct_code"],
                    template["formula"],
                    template["value_type"],
                    _status_from_formula_and_binding(template["formula"], template["data_acct_code"]),
                    trial["trial_m01"],
                    trial["trial_m02"],
                    trial["trial_m03"],
                    trial["trial_annual"],
                ),
            )
            await db.commit()
        return {"component_id": component_id}

    @router.post("/ai-suggestions")
    async def ai_suggestions(body: WorkbenchSuggestionRequest):
        async with aiosqlite.connect(common_db_path()) as db:
            db.row_factory = aiosqlite.Row
            await _ensure_runtime_tables(db)
            cur = await db.execute("SELECT product_name FROM product_type WHERE product_code = ?", (_safe_upper(body.product_code),))
            product = await cur.fetchone()
            report_name = ""
            if body.report_acct_code:
                cur = await db.execute(
                    """
                    SELECT COALESCE(n.node_name, c.component_name) AS report_acct_name
                    FROM product_budget_component c
                    LEFT JOIN data_account_metric_node n ON n.node_code = c.metric_node_code
                    WHERE c.product_code = ?
                      AND (c.metric_node_code = ? OR c.report_acct_code = ?)
                    LIMIT 1
                    """,
                    (_safe_upper(body.product_code), _safe_upper(body.report_acct_code), _safe_upper(body.report_acct_code)),
                )
                report = await cur.fetchone()
                report_name = str(report["report_acct_name"] if report else "")
            component = None
            if body.component_id:
                cur = await db.execute("SELECT * FROM product_budget_component WHERE component_id = ?", (body.component_id,))
                component = await cur.fetchone()
        product_name = str(product["product_name"] if product else body.product_code)
        component_name = str(component["component_name"] if component else "当前科目")
        fallback = [
            {
                "title": "按现有数据科目公式下发",
                "rule_code": "formula",
                "source_type": "existing_data_account",
                "formula": str(component["formula"] if component and component["formula"] else "选择数据科目后复用其预算公式"),
                "data_account_name": component_name,
                "reason": f"{product_name} / {report_name or component_name} 可优先复用已有数据科目，减少重复维护。",
            },
            {
                "title": "按动因项生成组件",
                "rule_code": "driver",
                "source_type": "driver",
                "formula": "日均余额 * 收益率 * 计息天数 / 360",
                "data_account_name": f"{component_name}_动因测算",
                "reason": "适合贷款、存款和收益率类科目，公式参数清晰，便于后续试算。",
            },
        ]
        if not deepseek_client or not deepseek_client.is_enabled():
            return {"model": "heuristic", "suggestions": fallback}

        system_prompt = "你是银行预算系统配置助手。只返回 JSON 数组，不要 Markdown。"
        user_prompt = (
            f"产品：{body.product_code} {product_name}\n"
            f"指标节点：{body.report_acct_code or ''} {report_name}\n"
            f"组件：{component_name}\n"
            "请给出 2-3 条预算规则配置建议，每条包含 title, rule_code(formula/driver/allocation/manual), "
            "source_type, formula, data_account_name, reason。"
        )
        text = await asyncio.to_thread(
            deepseek_client.chat_completion,
            system_prompt,
            user_prompt,
            0.2,
            900,
        )
        suggestions = _extract_json_array(text) or fallback
        return {"model": deepseek_client.model, "suggestions": suggestions}

    @router.post("/components/{component_id}/adopt-ai-suggestion")
    async def adopt_ai_suggestion(component_id: str, body: WorkbenchSuggestionAdopt):
        suggestion = body.suggestion or {}
        formula = _clean_formula(suggestion.get("formula"))
        rule_code = _clean_rule_code(suggestion.get("rule_code"))
        source_type = _clean_source_type(suggestion.get("source_type"))
        reason = str(suggestion.get("reason") or suggestion.get("title") or "").strip()
        if not formula:
            raise HTTPException(status_code=400, detail="AI 建议没有可采纳公式")
        async with aiosqlite.connect(common_db_path()) as db:
            await _ensure_runtime_tables(db)
            cur = await db.execute("SELECT data_acct_code FROM product_budget_component WHERE component_id = ?", (component_id,))
            row = await cur.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="组件不存在")
            status = _status_from_formula_and_binding(formula, row[0])
            await db.execute(
                """
                UPDATE product_budget_component
                SET formula = ?, rule_code = ?, source_type = ?, ai_reason = ?,
                    status = ?, updated_at = ?
                WHERE component_id = ?
                """,
                (formula, rule_code, source_type, reason, status, _now(), component_id),
            )
            await db.commit()
        return {"updated": True}

    @router.post("/ai-configure-product")
    async def ai_configure_product(body: WorkbenchProductAction):
        product_code = _safe_upper(body.product_code)
        async with aiosqlite.connect(common_db_path()) as db:
            db.row_factory = aiosqlite.Row
            await _ensure_runtime_tables(db)
            cur = await db.execute("SELECT product_name FROM product_type WHERE product_code = ?", (product_code,))
            product = await cur.fetchone()
            if not product:
                raise HTTPException(status_code=404, detail="产品不存在")
            cur = await db.execute(
                """
                SELECT c.component_id, c.product_code, c.report_acct_code, c.metric_node_code,
                       COALESCE(n.node_name, c.component_name) AS report_acct_name,
                       c.component_name, c.rule_code, c.source_type, c.data_acct_code,
                       d.data_acct_name, c.formula, c.value_type, c.status
                FROM product_budget_component c
                LEFT JOIN data_account_metric_node n ON n.node_code = c.metric_node_code
                LEFT JOIN data_account d ON d.data_acct_code = c.data_acct_code
                WHERE c.product_code = ?
                ORDER BY c.metric_node_code, c.report_acct_code, c.sort_order, c.created_at
                """,
                (product_code,),
            )
            components = await cur.fetchall()

        if not components:
            return {
                "model": "none",
                "product_code": product_code,
                "applied_count": 0,
                "skipped_count": 0,
                "warnings": ["当前产品暂无可配置组件"],
                "suggestions": [],
            }

        ai_items: list[dict[str, Any]] = []
        model = "heuristic"
        if deepseek_client and deepseek_client.is_enabled():
            payload = [
                {
                    "component_id": c["component_id"],
                    "report_acct_code": c["report_acct_code"],
                    "metric_node_code": c["metric_node_code"],
                    "report_acct_name": c["report_acct_name"],
                    "component_name": c["component_name"],
                    "data_acct_code": c["data_acct_code"],
                    "data_acct_name": c["data_acct_name"],
                    "current_formula": c["formula"],
                    "value_type": c["value_type"],
                }
                for c in components
            ]
            system_prompt = (
                "你是银行预算系统产品级配置助手。只返回 JSON 数组，不要 Markdown。"
                "每个数组项必须包含 component_id, title, rule_code, source_type, formula, reason。"
                "只为有把握的组件给公式；不能确定时 formula 留空。"
            )
            user_prompt = (
                f"产品：{product_code} {product['product_name']}\n"
                "请基于以下组件一次性生成产品预算配置建议。必须保持 component_id 原样返回；"
                "公式优先复用 current_formula；没有可靠依据不要编造公式。\n"
                f"{json.dumps(payload, ensure_ascii=False)}"
            )
            text = await asyncio.to_thread(
                deepseek_client.chat_completion,
                system_prompt,
                user_prompt,
                0.1,
                2600,
            )
            ai_items = _extract_json_array(text) or []
            model = deepseek_client.model

        ai_by_component = _suggestion_by_component_id(ai_items)
        applied: list[dict[str, Any]] = []
        pending_data_accounts: list[dict[str, Any]] = []
        pending_metric_nodes: list[dict[str, Any]] = []
        warnings: list[str] = []
        now = _now()
        async with aiosqlite.connect(common_db_path()) as db:
            await _ensure_runtime_tables(db)
            for component in components:
                component_id = str(component["component_id"])
                metric_node_code = str(component["metric_node_code"] or "").strip()
                data_acct_code = str(component["data_acct_code"] or "").strip()
                if not metric_node_code:
                    pending_metric_nodes.append({
                        "component_id": component_id,
                        "product_code": product_code,
                        "compat_report_acct_code": component["report_acct_code"],
                        "suggested_name": component["report_acct_name"] or component["component_name"],
                        "component_name": component["component_name"],
                        "reason": "该配置行尚未锚定标准数据科目指标树节点，需要用户确认或新增指标节点后才能进入 AI 自动配置草稿。",
                    })
                    continue
                if not data_acct_code:
                    pending_data_accounts.append({
                        "component_id": component_id,
                        "product_code": product_code,
                        "metric_node_code": metric_node_code,
                        "metric_node_name": component["report_acct_name"],
                        "suggested_data_acct_name": component["component_name"],
                        "value_type": component["value_type"],
                        "reason": "该指标节点在当前产品下尚未绑定数据科目，需要用户确认创建数据科目后才能写入公式草稿。",
                    })
                    continue
                suggestion = ai_by_component.get(component_id) or _fallback_suggestion_for_component(component)
                if not suggestion:
                    warnings.append(f"{component['report_acct_name']} / {component['component_name']} 缺少可靠公式，已跳过")
                    continue
                formula = _clean_formula(suggestion.get("formula"))
                if not formula:
                    warnings.append(f"{component['report_acct_name']} / {component['component_name']} AI 公式为空，已跳过")
                    continue
                if not _formula_has_data_account_ref(formula):
                    warnings.append(f"{component['report_acct_name']} / {component['component_name']} AI 公式未引用数据科目编码，已跳过")
                    continue
                rule_code = _clean_rule_code(suggestion.get("rule_code"))
                source_type = _clean_source_type(suggestion.get("source_type") or "ai_batch")
                reason = str(suggestion.get("reason") or suggestion.get("title") or "产品级 AI 配置").strip()
                status = _status_from_formula_and_binding(formula, component["data_acct_code"])
                await db.execute(
                    """
                    UPDATE product_budget_component
                    SET formula = ?, rule_code = ?, source_type = ?, ai_reason = ?,
                        status = ?, updated_at = ?
                    WHERE component_id = ?
                    """,
                    (formula, rule_code, source_type, reason, status, now, component_id),
                )
                applied.append({
                    "component_id": component_id,
                    "product_code": product_code,
                    "metric_node_code": metric_node_code,
                    "report_acct_code": component["report_acct_code"],
                    "component_name": component["component_name"],
                    "title": suggestion.get("title") or "产品级 AI 配置",
                    "rule_code": rule_code,
                    "formula": formula,
                    "reason": reason,
                })
            await db.execute(
                """
                UPDATE product_budget_config_package
                SET status = 'draft', updated_at = ?
                WHERE product_code = ?
                """,
                (now, product_code),
            )
            await db.commit()

        return {
            "model": model,
            "product_code": product_code,
            "applied_count": len(applied),
            "skipped_count": len(warnings) + len(pending_data_accounts) + len(pending_metric_nodes),
            "warnings": warnings,
            "suggestions": applied,
            "configured_drafts": applied,
            "pending_data_accounts": pending_data_accounts,
            "pending_metric_nodes": pending_metric_nodes,
            "result_groups": {
                "configured_drafts": applied,
                "pending_data_accounts": pending_data_accounts,
                "pending_metric_nodes": pending_metric_nodes,
            },
        }

    @router.post("/ai-configure-products")
    async def ai_configure_products(body: WorkbenchBulkProductAction):
        product_codes: list[str] = []
        seen: set[str] = set()
        for raw in body.product_codes or []:
            code = _safe_upper(raw)
            if code and code not in seen:
                seen.add(code)
                product_codes.append(code)
        if not product_codes:
            raise HTTPException(status_code=400, detail="请显式选择 AI 批量配置的产品范围")
        results = []
        for product_code in product_codes:
            results.append(await ai_configure_product(WorkbenchProductAction(product_code=product_code)))
        return {
            "scope_type": "explicit_products",
            "product_codes": product_codes,
            "product_count": len(product_codes),
            "results": results,
        }

    @router.post("/trial")
    async def trial(body: WorkbenchProductAction):
        product_code = _safe_upper(body.product_code)
        async with aiosqlite.connect(common_db_path()) as db:
            db.row_factory = aiosqlite.Row
            await _ensure_runtime_tables(db)
            rows = await _load_rows(db, product_code)
        warnings = []
        for row in rows:
            for component in row["components"]:
                if not component["data_acct_code"]:
                    warnings.append(f"{row['report_acct_name']} / {component['component_name']} 未绑定数据科目")
                if not str(component["formula"] or "").strip():
                    warnings.append(f"{row['report_acct_name']} / {component['component_name']} 未配置公式")
        return {
            "product_code": product_code,
            "rows": rows,
            "warnings": warnings,
            "trial_annual": round(sum(float(row["trial_annual"] or 0) for row in rows), 2),
        }

    @router.post("/dispatch")
    async def dispatch(body: WorkbenchProductAction):
        product_code = _safe_upper(body.product_code)
        warnings: list[str] = []
        dispatched = 0
        async with aiosqlite.connect(common_db_path()) as db:
            db.row_factory = aiosqlite.Row
            await db.execute("PRAGMA foreign_keys = ON")
            await _ensure_runtime_tables(db)
            cur = await db.execute("SELECT * FROM product_budget_component WHERE product_code = ?", (product_code,))
            components = await cur.fetchall()
            for component in components:
                data_code = component["data_acct_code"]
                if not data_code:
                    warnings.append(f"{component['component_name']} 未绑定数据科目，未下发")
                    continue
                formula = str(component["formula"] or "").strip()
                if not formula:
                    warnings.append(f"{component['component_name']} 未配置公式，未下发")
                    continue
                rule_config = {
                    "source": "product_budget_workbench",
                    "component_id": component["component_id"],
                    "product_code": component["product_code"],
                    "report_acct_code": component["report_acct_code"],
                    "metric_node_code": component["metric_node_code"],
                    "data_acct_code": component["data_acct_code"],
                    "component_name": component["component_name"],
                    "template_id": component["template_id"],
                    "template_name": component["template_name"],
                    "rule_code": component["rule_code"],
                    "source_type": component["source_type"],
                    "formula": formula,
                }
                await sync_budget_rule_to_data_account(
                    db,
                    data_acct_code=data_code,
                    data_acct_name=component["component_name"],
                    metric_group_code=component["metric_node_code"],
                    metric_group_name=component["component_name"],
                    budget_formula=formula,
                    budget_rule_code=component["rule_code"],
                    budget_rule_config_json=json.dumps(rule_config, ensure_ascii=False),
                    value_type=component["value_type"] or "金额",
                    need_calc=1,
                )
                if component["metric_node_code"]:
                    await ensure_metric_binding(
                        db,
                        metric_node_code=component["metric_node_code"],
                        scope_code=component["product_code"],
                        data_acct_code=data_code,
                        remark=f"由产品预算工作台下发：{component['component_name']}",
                        conflict_mode="reuse",
                    )
                await db.execute(
                    """
                    UPDATE product_budget_component
                    SET status = 'dispatched', updated_at = ?
                    WHERE component_id = ?
                    """,
                    (_now(), component["component_id"]),
                )
                dispatched += 1
            await db.execute(
                """
                UPDATE product_budget_config_package
                SET status = 'dispatched', updated_at = ?
                WHERE product_code = ?
                """,
                (_now(), product_code),
            )
            await db.commit()
        return {"product_code": product_code, "dispatched_count": dispatched, "warnings": warnings}

    return router
